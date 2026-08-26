import os
import re
import math
import openpyxl
import mysql.connector
from datetime import datetime, timedelta, timezone

def get_now_mexico_city():
    try:
        from zoneinfo import ZoneInfo
        return datetime.now(ZoneInfo("America/Mexico_City"))
    except Exception:
        return datetime.now(timezone(timedelta(hours=-6)))

def conectar_db():
    try:
        return mysql.connector.connect(
            host=os.getenv("DB_HOST", "localhost"),
            user=os.getenv("DB_USER", "root"),
            password=os.getenv("DB_PASSWORD", "los4valtierra"),
            database=os.getenv("DB_NAME", "sgh_portal"),
            port=int(os.getenv("DB_PORT", 3306))
        )
    except Exception as e:
        print("Error al conectar a DB en ciclicos_service:", e)
        return None

def crear_tablas_ciclicos_if_not_exists():
    db = conectar_db()
    if not db:
        return False
    try:
        cursor = db.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS ciclicos_historial (
                id INT AUTO_INCREMENT PRIMARY KEY,
                codigo_tienda VARCHAR(50) NOT NULL,
                nombre_tienda VARCHAR(100),
                marca VARCHAR(100),
                fecha_conteo DATETIME NOT NULL,
                varianza_pct DECIMAL(5,2) DEFAULT 0.00,
                total_sap_pzas INT DEFAULT 0,
                total_escaneo_pzas INT DEFAULT 0,
                total_falta_escaneo INT DEFAULT 0,
                total_falta_sap INT DEFAULT 0,
                total_negativos INT DEFAULT 0,
                comentarios TEXT,
                usuario VARCHAR(100),
                fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS ciclicos_detalle_historial (
                id INT AUTO_INCREMENT PRIMARY KEY,
                ciclico_id INT NOT NULL,
                upc VARCHAR(50) NOT NULL,
                descripcion VARCHAR(255),
                cantidad_sap INT DEFAULT 0,
                cantidad_escaneo INT DEFAULT 0,
                diferencia INT DEFAULT 0,
                tipo_alerta VARCHAR(50),
                comentario_item VARCHAR(255),
                FOREIGN KEY (ciclico_id) REFERENCES ciclicos_historial(id) ON DELETE CASCADE
            )
        """)
        # Purga automática de historial antiguo (retención de 6 meses)
        cursor.execute("DELETE FROM ciclicos_historial WHERE fecha_conteo < DATE_SUB(NOW(), INTERVAL 6 MONTH)")
        db.commit()
        db.close()
        return True
    except Exception as e:
        print("Error creando tablas de cíclicos:", e)
        if db:
            db.close()
        return False

def limpiar_upc(upc_raw):
    """
    Limpia y normaliza UPCs en cualquier formato:
    - Textos con espacios o apóstrofe
    - Notación científica (ej. 7.25126E+11 -> 725126000000)
    - Flotantes de Excel (ej. 72512594904.0 -> 72512594904)
    """
    if upc_raw is None:
        return ""
    
    val_str = str(upc_raw).strip()
    if not val_str:
        return ""
    
    # Si viene en notación científica o flotante de Excel (ej: 7.25126e+11 o 72512594904.0)
    if 'e' in val_str.lower() or '.' in val_str:
        try:
            val_float = float(val_str)
            val_str = f"{val_float:.0f}"
        except Exception:
            pass
            
    # Eliminar cualquier caracter que no sea dígito
    cleaned = re.sub(r'\D', '', val_str)
    # Normalizar ceros a la izquierda (empareja 08053672879018 con 8053672879018)
    return cleaned.lstrip('0')

def procesar_conciliacion_ciclico(file_escaneo_path, file_sap_path):
    """
    Lee y compara el archivo de escaneo vs el archivo de inventario de SAP.
    Estructura SAP (StockSummary_Details):
      - Col C (idx 2): EAN/UPC
      - Col D (idx 3): Descripción
      - Col H (idx 7): Cantidad Total
      - Col I (idx 8): Cantidad no confirmada (En tránsito)
      - Col J (idx 9): Cantidad en tienda (Stock Teórico en tienda)
    Estructura Escaneo (Libro1):
      - Col A (idx 0): UPC / EAN
      - Col B (idx 1): Piezas (conteo físico)
    """
    if not os.path.exists(file_escaneo_path):
        raise FileNotFoundError(f"No se encontró el archivo de escaneo: {file_escaneo_path}")
    if not os.path.exists(file_sap_path):
        raise FileNotFoundError(f"No se encontró el archivo de SAP: {file_sap_path}")

    # 1. Leer Escaneo Físico (Soporte Multi-Formato Inteligente)
    wb_esc = openpyxl.load_workbook(file_escaneo_path, data_only=True)
    
    # Búsqueda de hoja 'Scan' si existe (para el formato especial Solara/Multi-hoja), de lo contrario usa la activa
    sheet_esc = wb_esc.active
    for s_name in wb_esc.sheetnames:
        if s_name.strip().lower() == "scan":
            sheet_esc = wb_esc[s_name]
            break
            
    rows_esc = list(sheet_esc.iter_rows(values_only=True))
    
    col_upc_esc = 0  # Col A por defecto
    col_pzs_esc = 1  # Col B por defecto
    col_art_esc = 1  # Col B por defecto para ART
    col_marca_esc = 2 # Col C por defecto para MARCA
    start_row_esc = 0
    
    # Auto-detectar encabezados en las primeras 25 filas buscando únicamente en las primeras 8 columnas (A a H)
    for idx_r, row in enumerate(rows_esc[:25]):
        if not row:
            continue
        first_cols_lower = [str(c).lower().strip() if c is not None else "" for c in row[:8]]
        if any(cell in ["upc", "código", "codigo", "ean", "barcode"] for cell in first_cols_lower):
            for idx_c, cell_str in enumerate(first_cols_lower):
                if cell_str in ["upc", "código", "codigo", "ean", "barcode"]:
                    col_upc_esc = idx_c
                elif cell_str in ["pzs", "piezas", "cant", "cantidad", "conteo", "pz"]:
                    col_pzs_esc = idx_c
                elif cell_str in ["art", "articulo", "artículo", "modelo"]:
                    col_art_esc = idx_c
                elif cell_str in ["marca", "brand", "linea", "línea"]:
                    col_marca_esc = idx_c
            start_row_esc = idx_r + 1
            break
            
    conteo_escaneo = {} # {upc_limpio: cantidad_piezas}
    info_escaneo_extra = {} # {upc_limpio: {"art": str, "marca": str}}
    
    for row in rows_esc[start_row_esc:]:
        if not row or len(row) <= col_upc_esc or row[col_upc_esc] is None:
            continue
        upc_clean = limpiar_upc(row[col_upc_esc])
        if len(upc_clean) < 5:
            continue
            
        cant_pzas = 1
        if col_pzs_esc < len(row) and row[col_pzs_esc] is not None:
            try:
                cant_pzas = int(float(str(row[col_pzs_esc]).strip()))
            except Exception:
                cant_pzas = 1
                
        conteo_escaneo[upc_clean] = conteo_escaneo.get(upc_clean, 0) + cant_pzas
        
        art_val = str(row[col_art_esc]).strip() if col_art_esc < len(row) and row[col_art_esc] is not None else ""
        marca_val = str(row[col_marca_esc]).strip() if col_marca_esc < len(row) and row[col_marca_esc] is not None else ""
        if upc_clean not in info_escaneo_extra:
            info_escaneo_extra[upc_clean] = {"art": art_val, "marca": marca_val}

    # 2. Leer Inventario SAP (StockSummary_Details)
    wb_sap = openpyxl.load_workbook(file_sap_path, data_only=True)
    sheet_sap = wb_sap.active
    
    rows_sap = list(sheet_sap.iter_rows(values_only=True))
    if not rows_sap:
        raise ValueError("El archivo de SAP está vacío.")

    # Auto-detectar índices de columnas (por posición por defecto o encabezados de SAP)
    col_upc_idx = 2    # Col C: EAN/UPC
    col_desc_idx = 3   # Col D: Descripción
    col_stock_idx = 9  # Col J: Cantidad en tienda
    col_transito_idx = 8 # Col I: Cantidad no confirmada (en tránsito)
    header_row_idx = 0

    for idx, row in enumerate(rows_sap[:10]):
        if not row:
            continue
        row_str_lower = [str(c).lower() if c is not None else "" for c in row]
        
        for c_idx, cell_text in enumerate(row_str_lower):
            if any(k in cell_text for k in ["ean/upc", "ean", "upc", "codigo", "código", "barcode"]):
                col_upc_idx = c_idx
            if any(k in cell_text for k in ["descripción", "descripcion", "texto", "modelo"]):
                col_desc_idx = c_idx
            if any(k in cell_text for k in ["cantidad en tienda", "tienda", "libre utilización"]):
                col_stock_idx = c_idx
            if any(k in cell_text for k in ["no confirmada", "tránsito", "transito"]):
                col_transito_idx = c_idx

        if "ean/upc" in " ".join(row_str_lower) or "descripción" in " ".join(row_str_lower):
            header_row_idx = idx
            break

    datos_sap = {} # {upc_limpio: {"desc": str, "stock": int, "transito": int}}
    
    data_rows = rows_sap[header_row_idx + 1:] if header_row_idx < len(rows_sap) else rows_sap
    
    for row in data_rows:
        if not row or len(row) <= col_upc_idx:
            continue
        upc_clean = limpiar_upc(row[col_upc_idx])
        if not upc_clean or len(upc_clean) < 5:
            continue
            
        desc = str(row[col_desc_idx]).strip() if col_desc_idx < len(row) and row[col_desc_idx] is not None else "Sin descripción"
        
        # Cantidad Teórica en Tienda
        stock_raw = row[col_stock_idx] if col_stock_idx < len(row) else 0
        try:
            stock_int = int(float(str(stock_raw).strip())) if stock_raw is not None else 0
        except Exception:
            stock_int = 0

        # Cantidad no confirmada / En tránsito
        transito_raw = row[col_transito_idx] if col_transito_idx < len(row) else 0
        try:
            transito_int = int(float(str(transito_raw).strip())) if transito_raw is not None else 0
        except Exception:
            transito_int = 0
            
        if upc_clean in datos_sap:
            datos_sap[upc_clean]["stock"] += stock_int
            datos_sap[upc_clean]["transito"] += transito_int
        else:
            datos_sap[upc_clean] = {"desc": desc, "stock": stock_int, "transito": transito_int}

    # 3. Conciliación y Clasificación de Alertas
    falta_en_escaneo = [] # En SAP pero no escaneado o escaneado < SAP
    falta_en_sap = []     # Escaneado pero no en SAP o escaneado > SAP
    stock_negativo = []   # Stock SAP < 0
    tabla_suma_dif = []   # Lista resumida para tabla Suma de DIF
    
    todos_los_upcs = set(conteo_escaneo.keys()).union(set(datos_sap.keys()))
    
    total_sap_pzas = sum(item["stock"] for item in datos_sap.values())
    total_escaneo_pzas = sum(conteo_escaneo.values())
    
    diferencias_absolutas = 0
    
    for upc in sorted(todos_los_upcs):
        cant_esc = conteo_escaneo.get(upc, 0)
        sap_info = datos_sap.get(upc, {"desc": "No registrado en reporte SAP", "stock": 0, "transito": 0})
        cant_sap = sap_info["stock"]
        cant_transito = sap_info["transito"]
        desc = sap_info["desc"]
        
        diff = cant_esc - cant_sap # Positivo = Sobrante, Negativo = Faltante
        diferencias_absolutas += abs(diff)

        # Construir info de ART y MARCA para la tabla Suma de DIF
        extra_info = info_escaneo_extra.get(upc, {})
        art_code = extra_info.get("art") or (desc.split()[0] if desc and desc != "Sin descripción" else "N/A")
        marca_name = extra_info.get("marca") or "Ray-Ban"
        
        if diff != 0:
            tabla_suma_dif.append({
                "upc": upc,
                "art": art_code,
                "marca": marca_name,
                "total_dif": diff
            })

        # Alerta: Stock Negativo en SAP
        if cant_sap < 0:
            stock_negativo.append({
                "upc": upc,
                "descripcion": desc,
                "cant_sap": cant_sap,
                "cant_escaneo": cant_esc,
                "diferencia": diff,
                "mensaje": f"Esta pieza la tienes en negativa en SAP ({cant_sap} pzas). Verifica por qué."
            })
            
        # Alerta: Falta en Escaneo (SAP indica que debe estar en tienda pero falta en escaneo)
        elif cant_sap > cant_esc:
            faltan = cant_sap - cant_esc
            msg_transito = f" (Nota: Hay {cant_transito} pza(s) no confirmadas/en tránsito)" if cant_transito > 0 else ""
            falta_en_escaneo.append({
                "upc": upc,
                "descripcion": desc,
                "cant_sap": cant_sap,
                "cant_escaneo": cant_esc,
                "faltantes": faltan,
                "cant_transito": cant_transito,
                "mensaje": f"Este código no está en tu escaneo y tu inventario SAP dice que debes tener {cant_sap} pza(s) en tienda (Faltan {faltan} pza(s)){msg_transito}."
            })
            
        # Alerta: Falta en SAP / Sobrante en Escaneo
        elif cant_esc > cant_sap:
            sobran = cant_esc - cant_sap
            falta_en_sap.append({
                "upc": upc,
                "descripcion": desc,
                "cant_sap": cant_sap,
                "cant_escaneo": cant_esc,
                "sobrantes": sobran,
                "mensaje": f"Este código lo escaneaste ({cant_esc} pza(s)) pero en tu inventario SAP no figura o tiene registrado menos ({cant_sap} pza(s))."
            })

    # Ordenar la tabla Suma de DIF: Primero Positivos (+), luego Negativos (-)
    tabla_suma_dif.sort(key=lambda x: (-x["total_dif"], x["upc"]))

    # Cálculo de varianza (%) basado en piezas del segmento/conteo o total escaneado
    # Si hay coincidencias de marca o escaneo, calcular sobre la base contada (ej. 7 / 427 = 1.64%)
    pzas_conteo_marca = sum(conteo_escaneo[u] for u in conteo_escaneo if datos_sap.get(u))
    divisor_varianza = pzas_conteo_marca if pzas_conteo_marca > 0 else (total_escaneo_pzas if total_escaneo_pzas > 0 else total_sap_pzas)
    
    if divisor_varianza > 0:
        varianza_pct = round((diferencias_absolutas / divisor_varianza) * 100, 2)
    else:
        varianza_pct = 0.0

    suma_dif_total = sum(item["total_dif"] for item in tabla_suma_dif)

    return {
        "summary": {
            "total_sap_pzas": total_sap_pzas,
            "total_escaneo_pzas": total_escaneo_pzas,
            "total_falta_escaneo": len(falta_en_escaneo),
            "total_falta_sap": len(falta_en_sap),
            "total_negativos": len(stock_negativo),
            "varianza_pct": varianza_pct,
            "suma_dif_total": suma_dif_total
        },
        "tabla_suma_dif": tabla_suma_dif,
        "falta_en_escaneo": falta_en_escaneo,
        "falta_en_sap": falta_en_sap,
        "stock_negativo": stock_negativo
    }

def guardar_ciclico_db(codigo_tienda, nombre_tienda, marca, resumen, falta_esc, falta_sap, negativos, comentarios="", usuario="tienda"):
    crear_tablas_ciclicos_if_not_exists()
    db = conectar_db()
    if not db:
        return False, "No se pudo conectar a la base de datos MySQL local."

    try:
        cursor = db.cursor()
        fecha_actual = get_now_mexico_city().strftime("%Y-%m-%d %H:%M:%S")

        cursor.execute("""
            INSERT INTO ciclicos_historial (
                codigo_tienda, nombre_tienda, marca, fecha_conteo, varianza_pct,
                total_sap_pzas, total_escaneo_pzas, total_falta_escaneo,
                total_falta_sap, total_negativos, comentarios, usuario
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            str(codigo_tienda),
            str(nombre_tienda),
            str(marca),
            fecha_actual,
            resumen.get("varianza_pct", 0.0),
            resumen.get("total_sap_pzas", 0),
            resumen.get("total_escaneo_pzas", 0),
            resumen.get("total_falta_escaneo", 0),
            resumen.get("total_falta_sap", 0),
            resumen.get("total_negativos", 0),
            str(comentarios),
            str(usuario)
        ))
        
        ciclico_id = cursor.lastrowid

        detalles_to_insert = []
        
        for item in falta_esc:
            detalles_to_insert.append((
                ciclico_id, item["upc"], item["descripcion"], item["cant_sap"],
                item["cant_escaneo"], -item["faltantes"], "FALTA_EN_ESCANEO", item["mensaje"]
            ))

        for item in falta_sap:
            detalles_to_insert.append((
                ciclico_id, item["upc"], item["descripcion"], item["cant_sap"],
                item["cant_escaneo"], item["sobrantes"], "FALTA_EN_SAP", item["mensaje"]
            ))

        for item in negativos:
            detalles_to_insert.append((
                ciclico_id, item["upc"], item["descripcion"], item["cant_sap"],
                item["cant_escaneo"], item["diferencia"], "STOCK_NEGATIVO", item["mensaje"]
            ))

        if detalles_to_insert:
            cursor.executemany("""
                INSERT INTO ciclicos_detalle_historial (
                    ciclico_id, upc, descripcion, cantidad_sap, cantidad_escaneo,
                    diferencia, tipo_alerta, comentario_item
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, detalles_to_insert)

        db.commit()
        db.close()
        return True, ciclico_id
    except Exception as e:
        print("Error al guardar cíclico en DB:", e)
        if db:
            db.close()
        return False, str(e)

def obtener_historial_ciclicos(codigo_tienda=None):
    db = conectar_db()
    if not db:
        return []
    try:
        cursor = db.cursor(dictionary=True)
        if codigo_tienda and str(codigo_tienda).strip() != "TODAS":
            cursor.execute("""
                SELECT * FROM ciclicos_historial
                WHERE codigo_tienda = %s OR nombre_tienda = %s
                ORDER BY fecha_conteo DESC LIMIT 100
            """, (str(codigo_tienda), str(codigo_tienda)))
        else:
            cursor.execute("""
                SELECT * FROM ciclicos_historial
                ORDER BY fecha_conteo DESC LIMIT 100
            """)
        rows = cursor.fetchall()
        db.close()
        return rows
    except Exception as e:
        print("Error obteniendo historial de cíclicos:", e)
        if db:
            db.close()
        return []

def obtener_detalle_ciclico(ciclico_id):
    db = conectar_db()
    if not db:
        return []
    try:
        cursor = db.cursor(dictionary=True)
        cursor.execute("""
            SELECT * FROM ciclicos_detalle_historial
            WHERE ciclico_id = %s
        """, (ciclico_id,))
        rows = cursor.fetchall()
        db.close()
        return rows
    except Exception as e:
        print("Error obteniendo detalle de cíclico:", e)
        if db:
            db.close()
        return []
