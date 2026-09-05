# -*- coding: utf-8 -*-
"""
MÓDULO: POLAR RB Y OO (Polarizados Ray-Ban y Oakley)
Métricas semanales y anuales de polarizado por tienda a partir de reportes Excel oficiales.
"""

import os
import re
import json
import openpyxl
from datetime import datetime, timezone, timedelta
import flet as ft

POLAR_UPLOADS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads", "polar")

# Catálogo maestro oficial de relaciones (Número de Tienda -> Nombre en Excel)
TIENDAS_MASTER_DATA = [
    (3010, 10, "APVR"), (3019, 3, "PARQUE TOR"), (3028, 4, "OASIS COY"), (3035, 10, "PVR GAL"),
    (3049, 11, "AGDL T1A"), (3051, 6, "CDJ"), (3067, 11, "AGDL T1B"), (3080, 11, "PP KIOSCO"),
    (3138, 8, "CLN FORUM 2"), (3150, 4, "ACA3"), (3151, 10, "AND2"), (3154, 6, "CHI1"),
    (3155, 6, "CHI2"), (3160, 10, "GG4"), (3162, 10, "GP3"), (3164, 16, "MER1"),
    (3167, 16, "MER4"), (3169, 5, "MTY1"), (3170, 5, "MTY2"), (3178, 5, "MTY4"),
    (3179, 6, "MTY5"), (3180, 6, "MTY6"), (3181, 5, "MTY7"), (3196, 8, "ACLN"),
    (3200, 9, "LEON OUTLET"), (3201, 8, "MAZ T1"), (3203, 8, "CABO1"), (3204, 8, "CABO2"),
    (3206, 8, "CABO3"), (3238, 9, "URUAPAN"), (3259, 10, "GG STZA"), (3260, 10, "AND3"),
    (3280, 11, "GDL FORUM"), (3282, 16, "TAB2"), (3287, 6, "MTY TA"), (3289, 6, "MTY TB"),
    (3290, 13, "CUN2"), (3295, 13, "CUN3"), (3296, 6, "MTY TC"), (3347, 7, "HMO T1"),
    (3348, 10, "PVR T1"), (3456, 9, "LEON T1"), (3464, 11, "COL1"), (3482, 16, "CD CAR 1"),
    (3487, 1, "QRO1"), (3488, 3, "ATIZAPAN 1"), (3491, 7, "HMO1"), (3497, 11, "GDL FORUM 2"),
    (3500, 2, "SAMARA"), (3502, 3, "INTERLOMAS 1"), (3504, 5, "TOR1"), (3506, 2, "222 A"),
    (3507, 3, "LINDAVISTA 2"), (3508, 1, "P DELTA 1"), (3511, 5, "MTY11"), (3516, 2, "LERMA OUTLET 2"),
    (3518, 9, "ZAC1"), (3519, 2, "STA FE 1"), (3523, 6, "MTY9"), (3526, 15, "COATZA 1"),
    (3542, 3, "GAL PACH"), (3556, 13, "CUN5"), (3557, 2, "METEPEC 1"), (3567, 6, "MTY TA2"),
    (3574, 2, "GAL TOL"), (3582, 5, "SLP2"), (3583, 3, "SATELITE 1"), (3584, 1, "SAN MIGUEL 1"),
    (3585, 2, "STA FE 2"), (3586, 3, "PUNTA NORTE 1"), (3600, 13, "CUN6"), (3617, 5, "SLP1"),
    (3625, 14, "PLAYA AME"), (3645, 3, "VALLEJO"), (3754, 16, "CAMPECHE"), (3770, 10, "GG5"),
    (3911, 9, "SALAMANCA"), (4018, 4, "ANGELOPOLIS"), (4040, 4, "PERISUR"), (4066, 4, "PUEBLA OUTLET"),
    (4222, 15, "ANDAMAR VER"), (4849, 15, "VER T1B"), (4862, 7, "MEXICALI T1"), (5006, 6, "MTY8"),
    (5078, 12, "CUN FBO"), (5079, 16, "TAB"), (5080, 7, "TIJ"), (5081, 14, "PLAYA CH"),
    (5107, 9, "CIBELES"), (5112, 9, "LEON MEDIANA"), (5203, 14, "PLAYA ZOLAR"), (5205, 14, "PLAYA 16"),
    (5207, 13, "CUN1"), (5208, 1, "AICM T1A"), (5215, 1, "AICM T1D"), (5238, 4, "ZIHUATANEJO T1"),
    (5256, 3, "PUNTA NORTE 2"), (5342, 14, "PLAYA CORAZON"), (5352, 8, "CABO T1"), (5378, 1, "AICM T1F"),
    (5379, 1, "AICM T2A"), (5385, 1, "AICM T2B"), (5399, 1, "AICM T2D"), (5610, 9, "MOR ALTOZANO"),
    (5632, 9, "LEON GDE"), (5638, 9, "AGS"), (7048, 8, "CABO4"), (7049, 2, "ANTARA"),
    (7051, 8, "LA PAZ T1"), (7058, 12, "CUN T2E"), (7060, 5, "MTY ESFERA"), (7061, 6, "MTY LA FE"),
    (7062, 7, "TIJ 2"), (8471, 13, "CUN OUTLET"), (8474, 2, "MASARYK"), (8491, 7, "ENSENADA"),
    (8604, 10, "SENTURA ZAMORA"), (8686, 5, "MTY FASHION DRIVE"), (8691, 4, "FORUM CUERNAVACA"),
    (8692, 10, "LA ISLA PVR"), (8784, 9, "MOR T1"), (8820, 7, "NOGALES"), (8822, 15, "OAXACA VALLE"),
    (8823, 15, "HUATULCO BAHIAS"), (8829, 1, "ANTEA"), (8838, 11, "FORUM TEPIC"), (8839, 3, "FORTUNA"),
    (8845, 7, "TIJ T1C"), (8851, 13, "PUERTO CANCUN"), (8881, 8, "LA PAZ CENTRO"), (8882, 11, "PLAZA PATRIA"),
    (8885, 1, "LA VICTORIA"), (8897, 6, "MTY OUTLET"), (8902, 5, "SALTILLO 1"), (8983, 14, "PUNTA LANGOSTA"),
    (8992, 12, "CUN T3C"), (9006, 4, "AVERANDA"), (9037, 11, "AGDL T2A"), (9054, 4, "MANACAR"),
    (9056, 6, "PASEO DURANGO"), (9063, 4, "PASEO ACOXPA"), (9095, 16, "AMBAR TUXTLA"), (9116, 4, "PASEO ZIHUATANEJO"),
    (9176, 10, "MALECON VALLARTA"), (9183, 12, "CUN T4B"), (9184, 12, "CUN T4A"), (9191, 16, "LA ISLA MERIDA"),
    (9253, 11, "AGDL T2B"), (9277, 3, "AMP INTERLOMAS"), (9279, 8, "LOS MOCHIS"), (9341, 11, "AGDL T2C"),
    (9342, 4, "SOLARA PUEBLA"), (9343, 6, "PASEO CHIHUAHUA"), (9344, 2, "ANTENAS"), (9345, 7, "ALAMEDA TIJ"),
    (9393, 9, "PLAZA CIBELES"), (9428, 4, "EXPLANADA PUEBLA"), (9506, 5, "SLP AERO"), (9507, 1, "PASEO QRO"),
    (9536, 5, "PUNTO VALLE KIOSCO"), (9537, 5, "PUNTO VALLE"), (9539, 2, "TOWN SQUARE METEPEC"), (9551, 16, "HARBOR MERIDA"),
    (9562, 10, "LANDMARK GDL"), (9563, 1, "QRO PREM OUTLET"), (9593, 11, "MIDTOWN JALISCO"), (9602, 15, "AMERICAS XALAPA"),
    (9603, 2, "ANTENAS 2"), (10540, 13, "LA ISLA CUN 2"), (10615, 1, "ENCUENTRO OCEANIA"), (10637, 16, "GAL MERIDA"),
    (10876, 9, "MOR AME KIOSKO"), (10955, 9, "ALTACIA LEON"), (12943, 12, "CUN T4C"), (12964, 3, "EXPLANADA PACHUCA"),
    (14499, 11, "GAL SANTA ANITA"), (15364, 9, "MOR AMERICAS"), (15536, 15, "ALTAMA TAMPICO"), (15610, 7, "PENINSULA TIJ"),
    (25163, 6, "MTY TA3"), (25164, 5, "SLP PARK"), (25237, 8, "LA CEIBA"), (25245, 14, "TULUM KAPEN HA"),
    (25858, 7, "LANDMARK TIJUANA"), (25859, 13, "OUTLET RIVERA MAYA"), (25977, 1, "AIFA"), (26153, 7, "TIJ T1B"),
    (26246, 2, "GAL METEPEC AMP"), (26369, 7, "ANDENES HMO"), (26382, 3, "PARQUE TEPEYAC"), (26503, 8, "ANIMA CABOS"),
    (26558, 16, "LA ISLA MER 2"), (27405, 5, "PLAZA SAN LUIS SOLARIS")
]

MAP_NUM_TO_POLAR_NAME = {str(num): name.upper() for num, _, name in TIENDAS_MASTER_DATA}
MAP_POLAR_NAME_TO_NUM = {name.upper(): str(num) for num, _, name in TIENDAS_MASTER_DATA}

def asegurar_directorios():
    os.makedirs(POLAR_UPLOADS_DIR, exist_ok=True)

def get_now_mexico_city():
    utc_now = datetime.now(timezone.utc)
    return utc_now - timedelta(hours=6)

def crear_tablas_polar_if_not_exists(db):
    if not db:
        return False
    try:
        cursor = db.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS polar_archivos (
                id INT AUTO_INCREMENT PRIMARY KEY,
                semana VARCHAR(50) NOT NULL,
                anio INT NOT NULL,
                nombre_archivo VARCHAR(255) NOT NULL,
                subido_por VARCHAR(100) DEFAULT '',
                fecha_subida DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS polar_metricas (
                id INT AUTO_INCREMENT PRIMARY KEY,
                archivo_id INT NOT NULL,
                semana VARCHAR(50) NOT NULL,
                anio INT NOT NULL,
                zona_hoja VARCHAR(100) DEFAULT '',
                tienda VARCHAR(150) NOT NULL,
                marca_periodo VARCHAR(50) NOT NULL,
                polar_qty INT DEFAULT 0,
                total_qty INT DEFAULT 0,
                pct_polar FLOAT DEFAULT 0.0,
                INDEX idx_tienda_sem (tienda, semana, anio),
                INDEX idx_marca (marca_periodo)
            )
        """)
        db.commit()
        cursor.close()
        return True
    except Exception as ex:
        print("Notice crear_tablas_polar_if_not_exists:", ex)
        return False

def parse_float_pct(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        if val <= 1.0 and val > 0.0:
            return round(val * 100.0, 1)
        return round(float(val), 1)
    s = str(val).replace("%", "").strip()
    try:
        f = float(s)
        if f <= 1.0 and f > 0.0:
            return round(f * 100.0, 1)
        return round(f, 1)
    except Exception:
        return 0.0

def parse_int_qty(val):
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).replace(",", "").strip()
    try:
        return int(float(s))
    except Exception:
        return 0

def procesar_excel_polar(file_path, subido_por, db_fn):
    """
    Lee todas las hojas del Excel de Polar (AICM, DF CENTRO, VALLE, DF NORTE, etc.)
    e indexa las 4 matrices: RB Wxx, RB 2026, OO Wxx, OO 2026.
    """
    if not os.path.exists(file_path):
        return False, "El archivo no existe."
        
    db = db_fn()
    if not db:
        return False, "No hay conexión con la base de datos."
        
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as ex:
        if db: db.close()
        return False, f"Error abriendo Excel: {str(ex)}"

    # Detectar semana y año desde el nombre del archivo o celdas
    fname = os.path.basename(file_path).upper()
    match_sem = re.search(r'W\s*(\d+)', fname)
    if not match_sem:
        match_sem = re.search(r'SEMANA\s*(\d+)', fname)
    
    num_semana_str = f"W{match_sem.group(1)}" if match_sem else "W34"
    anio_detectado = 2026
    match_anio = re.search(r'202\d', fname)
    if match_anio:
        anio_detectado = int(match_anio.group(0))
        
    cursor = db.cursor(dictionary=True)
    crear_tablas_polar_if_not_exists(db)
    
    # Limpiar previamente registros de la misma semana y año para evitar mezclas o datos obsoletos
    try:
        cursor.execute("DELETE FROM polar_metricas WHERE semana = %s AND anio = %s", (num_semana_str, anio_detectado))
        cursor.execute("DELETE FROM polar_archivos WHERE semana = %s AND anio = %s", (num_semana_str, anio_detectado))
    except Exception as ex_del:
        print("Notice borrando previo polar:", ex_del)
    
    # Crear registro en polar_archivos
    cursor.execute("""
        INSERT INTO polar_archivos (semana, anio, nombre_archivo, subido_por, fecha_subida)
        VALUES (%s, %s, %s, %s, %s)
    """, (num_semana_str, anio_detectado, os.path.basename(file_path), subido_por, get_now_mexico_city().strftime('%Y-%m-%d %H:%M:%S')))
    archivo_id = cursor.lastrowid
    
    registros_insertados = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_r = ws.max_row or 100
        max_c = ws.max_column or 20
        
        # Buscar bloques en la hoja
        # 1. RB SEMANA (RB Wxx)
        # 2. OO SEMANA (OO Wxx)
        # 3. RB AÑO (RB 2026)
        # 4. OO AÑO (OO 2026)
        
        # Escanear celdas para ubicar las posiciones de inicio
        blocks = [] # tuples: (marca_periodo, start_row, start_col)
        
        for r in range(1, min(max_r, 40)):
            for c in range(1, min(max_c, 15)):
                val = str(ws.cell(row=r, column=c).value or "").strip().upper()
                if not val:
                    continue
                if "RB" in val and ("W" in val or "SEM" in val):
                    # RB Semana
                    blocks.append(("RB_SEM", r, c))
                elif "RB" in val and ("202" in val or "AÑO" in val or "YTD" in val):
                    # RB Año
                    blocks.append(("RB_ANIO", r, c))
                elif ("OO" in val or "OAKLEY" in val) and ("W" in val or "SEM" in val):
                    # OO Semana
                    blocks.append(("OO_SEM", r, c))
                elif ("OO" in val or "OAKLEY" in val) and ("202" in val or "AÑO" in val or "YTD" in val):
                    # OO Año
                    blocks.append(("OO_ANIO", r, c))

        # Si no encontró por título exacto, usar heurística por columnas típicas (B=col 2, H=col 8)
        if not blocks:
            # Fallback a posiciones estándar
            blocks = [
                ("RB_SEM", 3, 2),
                ("OO_SEM", 3, 8),
                ("RB_ANIO", 22, 2),
                ("OO_ANIO", 22, 8)
            ]

        for marca_p, start_r, start_c in blocks:
            # La fila de encabezados de columna suele ser start_r + 1 (TIENDA, POLAR, Total general, %POLAR)
            header_row = start_r + 1
            data_start_row = header_row + 1
            
            for curr_r in range(data_start_row, data_start_row + 60):
                tienda_val = ws.cell(row=curr_r, column=start_c).value
                if not tienda_val:
                    continue
                tienda_str = str(tienda_val).strip()
                if not tienda_str:
                    continue
                    
                # Si llega al total o a otro bloque, detener
                if tienda_str.upper().startswith("TOTAL") or tienda_str.upper().startswith("RB") or tienda_str.upper().startswith("OO"):
                    # Si es TOTAL, podemos guardarlo con nombre de tienda 'TOTAL' si se desea o detener
                    if tienda_str.upper().startswith("TOTAL"):
                        polar_v = ws.cell(row=curr_r, column=start_c + 1).value
                        tot_v = ws.cell(row=curr_r, column=start_c + 2).value
                        pct_v = ws.cell(row=curr_r, column=start_c + 3).value
                        
                        polar_qty = parse_int_qty(polar_v)
                        tot_qty = parse_int_qty(tot_v)
                        pct_polar = parse_float_pct(pct_v) if pct_v is not None else ((polar_qty / tot_qty * 100.0) if tot_qty > 0 else 0.0)
                        
                        cursor.execute("""
                            INSERT INTO polar_metricas 
                                (archivo_id, semana, anio, zona_hoja, tienda, marca_periodo, polar_qty, total_qty, pct_polar)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (archivo_id, num_semana_str, anio_detectado, sheet_name, f"TOTAL - {sheet_name}", marca_p, polar_qty, tot_qty, pct_polar))
                        registros_insertados += 1
                    break
                    
                polar_v = ws.cell(row=curr_r, column=start_c + 1).value
                tot_v = ws.cell(row=curr_r, column=start_c + 2).value
                pct_v = ws.cell(row=curr_r, column=start_c + 3).value
                
                polar_qty = parse_int_qty(polar_v)
                tot_qty = parse_int_qty(tot_v)
                pct_polar = parse_float_pct(pct_v)
                if pct_v is None and tot_qty > 0:
                    pct_polar = round(polar_qty / tot_qty * 100.0, 1)

                cursor.execute("""
                    INSERT INTO polar_metricas 
                        (archivo_id, semana, anio, zona_hoja, tienda, marca_periodo, polar_qty, total_qty, pct_polar)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (archivo_id, num_semana_str, anio_detectado, sheet_name, tienda_str, marca_p, polar_qty, tot_qty, pct_polar))
                registros_insertados += 1

    db.commit()
    cursor.close()
    db.close()
    
    return True, f"✅ Archivo '{num_semana_str}' procesado con éxito. Se indexaron {registros_insertados} registros en {len(wb.sheetnames)} zonas."


def build_polar_view(page, user_info, conectar_db_fn, mostrar_snack_fn, tr_fn=None, get_zona_region_fn=None, seleccionar_archivo_async=None):
    """
    Construye la vista interactiva del Módulo Polar RB y OO.
    """
    asegurar_directorios()
    db_fn = conectar_db_fn
    mostrar_snack = mostrar_snack_fn or (lambda msg, col: print(msg))
    tr = tr_fn or (lambda es, en, *args: es)
    
    usuario_id = user_info.get("id") or 1
    rol_usuario = (user_info.get("rol") or "").strip()
    nombre_usuario = user_info.get("nombre") or "Usuario"
    tienda_usuario = (user_info.get("tienda") or "").strip()
    es_admin = (rol_usuario.lower() == "admin")
    
    # Comprobar tablas
    db_init = db_fn()
    if db_init:
        crear_tablas_polar_if_not_exists(db_init)
        db_init.close()
        
    # Obtener semanas disponibles y catálogo de tiendas
    semanas_disponibles = []
    tiendas_disponibles = []
    
    # Copia local de mapeos combinando catálogo maestro oficial y BD de usuarios
    num_a_polar = dict(MAP_NUM_TO_POLAR_NAME)
    polar_a_num = dict(MAP_POLAR_NAME_TO_NUM)
    
    db_load = db_fn()
    if db_load:
        try:
            cur = db_load.cursor(dictionary=True)
            cur.execute("SELECT DISTINCT semana, anio FROM polar_metricas ORDER BY anio DESC, semana DESC")
            for r in cur.fetchall():
                sem_key = f"{r['semana']} - {r['anio']}"
                if sem_key not in semanas_disponibles:
                    semanas_disponibles.append(sem_key)
                    
            cur.execute("SELECT DISTINCT tienda FROM polar_metricas WHERE tienda NOT LIKE 'TOTAL%' ORDER BY tienda ASC")
            for r in cur.fetchall():
                if r["tienda"] and r["tienda"] not in tiendas_disponibles:
                    tiendas_disponibles.append(r["tienda"])
                    
            # Enriquecer con usuarios registrados en BD
            cur.execute("SELECT Usuario, Nombre_Completo, Tienda FROM usuarios WHERE Rol != 'Admin'")
            for u_row in cur.fetchall():
                t_nombre = (u_row.get("Tienda") or "").strip().upper()
                u_user = (u_row.get("Usuario") or "").strip()
                u_full = (u_row.get("Nombre_Completo") or "").strip()
                
                nums = re.findall(r'\d+', u_user) + re.findall(r'\d+', u_full)
                for num in nums:
                    if len(num) >= 3 and num not in num_a_polar:
                        if t_nombre:
                            num_a_polar[num] = t_nombre
                            polar_a_num[t_nombre] = num
            cur.close()
        except Exception as ex:
            print("Notice cargando semanas/tiendas polar:", ex)
        db_load.close()

    # Función inteligente para resolver tienda por número (ej: 3502) o por nombre (ej: Interlomas)
    def resolver_tienda_por_query(q):
        if not q:
            return None
        q_clean = str(q).strip().upper()
        
        # 1. Búsqueda por número de tienda (ej. '3502' o 'SGH3502')
        nums = re.findall(r'\d+', q_clean)
        for n in nums:
            if n in num_a_polar:
                target = num_a_polar[n]
                for td in tiendas_disponibles:
                    if td == target:
                        return td
                for td in tiendas_disponibles:
                    if target in td or td in target:
                        return td

        # 2. Búsqueda por coincidencia exacta de nombre
        for td in tiendas_disponibles:
            if td == q_clean:
                return td

        # 3. Reglas especiales para evitar cruces (ej: AMPLIACION vs NORMAL)
        es_amp = ("AMP" in q_clean or "AMPLIACION" in q_clean or "AMPLIACIÓN" in q_clean)
        
        # 4. Búsqueda por palabras / fragmentos
        candidates = []
        for td in tiendas_disponibles:
            td_is_amp = ("AMP" in td or "AMPLIACION" in td)
            if es_amp and td_is_amp and any(w in td for w in q_clean.split() if len(w) > 3 and w not in ["AMP", "AMPLIACION"]):
                return td
            if not es_amp and td_is_amp:
                continue
            if q_clean in td or td in q_clean:
                candidates.append(td)
                
        if candidates:
            candidates.sort(key=lambda x: len(x))
            return candidates[0]
            
        return None

    # Estado activo
    semana_activa = [semanas_disponibles[0] if semanas_disponibles else "W34 - 2026"]
    
    # Resolver tienda inicial según usuario o catálogo
    def resolver_tienda_inicial():
        if not es_admin:
            # Probar usuario (ej. sgh3502), nombre completo (ej. Tienda Interlomas 3502) o tienda
            for cand in [user_info.get("usuario"), nombre_usuario, tienda_usuario]:
                if cand:
                    matched = resolver_tienda_por_query(cand)
                    if matched: return matched
            return tienda_usuario or "INTERLOMAS 1"
        if tiendas_disponibles:
            return tiendas_disponibles[0]
        return tienda_usuario or "INTERLOMAS 1"

    tienda_activa = [resolver_tienda_inicial()]
    vista_modo = ["tarjetas"] # "tarjetas" o "ranking"
    
    # Helper para opciones de dropdown con número de tienda
    def make_tienda_opt(t):
        n = polar_a_num.get(t, "")
        return ft.dropdown.Option(key=t, text=f"{t} (#{n})" if n else t)
    
    # -------------------------------------------------------------
    # CONTENEDORES Y FILTROS
    # -------------------------------------------------------------
    dd_semana = ft.Dropdown(
        label="Semana",
        options=[ft.dropdown.Option(s) for s in semanas_disponibles] if semanas_disponibles else [ft.dropdown.Option("W34 - 2026")],
        value=semana_activa[0],
        width=160,
        border_color="#00FFFF",
        color="white",
        dense=True
    )
    
    tf_buscar = ft.TextField(
        label="🔍 Buscar tienda / Nº",
        hint_text="Ej: 3502, 9277, Interlomas...",
        width=200,
        border_color="#00FFFF",
        color="white",
        dense=True,
        visible=es_admin
    )
    
    dd_tienda = ft.Dropdown(
        label="Tienda",
        options=[make_tienda_opt(t) for t in tiendas_disponibles],
        value=tienda_activa[0] if tienda_activa[0] in tiendas_disponibles else (tiendas_disponibles[0] if tiendas_disponibles else ""),
        width=250,
        border_color="#00FFFF",
        color="white",
        dense=True,
        visible=es_admin
    )
    
    num_asignado = polar_a_num.get(tienda_activa[0], "")
    txt_tienda_fija = ft.Container(
        content=ft.Row([
            ft.Icon(ft.Icons.STORE_ROUNDED, color="#00FFFF", size=18),
            ft.Text(f"Tienda Asignada: {tienda_activa[0]}" + (f" (#{num_asignado})" if num_asignado else ""), color="white", weight="bold", size=13)
        ], spacing=6),
        padding=ft.padding.Padding(12, 8, 12, 8),
        bgcolor="#141424",
        border_radius=8,
        border=ft.Border.all(1, "#2A2A3E"),
        visible=not es_admin
    )

    content_area = ft.Column(spacing=12)

    # -------------------------------------------------------------
    # RENDER DE TARJETAS DE MÉTRICAS (RB y OO)
    # -------------------------------------------------------------
    def build_metricas_panel():
        is_mobile = (page.width < 700) if (page and page.width) else False
        
        sem_raw, anio_raw = semana_activa[0].split(" - ") if " - " in semana_activa[0] else (semana_activa[0], "2026")
        
        # Buscar métricas de la tienda en MySQL
        db = db_fn()
        metricas = {}
        zona_tienda = ""
        
        if db:
            try:
                cur = db.cursor(dictionary=True)
                # Búsqueda flexible por tienda
                cur.execute("""
                    SELECT * FROM polar_metricas 
                    WHERE (tienda = %s OR tienda LIKE %s)
                      AND semana = %s AND anio = %s
                """, (tienda_activa[0], f"%{tienda_activa[0]}%", sem_raw.strip(), int(anio_raw.strip())))
                
                rows = cur.fetchall()
                for r in rows:
                    metricas[r["marca_periodo"]] = r
                    if r.get("zona_hoja"):
                        zona_tienda = r["zona_hoja"]
                        
                cur.close()
            except Exception as ex:
                print("Error consultando polar_metricas:", ex)
            db.close()

        # Extraer valores
        rb_sem = metricas.get("RB_SEM", {"polar_qty": 0, "total_qty": 0, "pct_polar": 0.0})
        rb_anio = metricas.get("RB_ANIO", {"polar_qty": 0, "total_qty": 0, "pct_polar": 0.0})
        oo_sem = metricas.get("OO_SEM", {"polar_qty": 0, "total_qty": 0, "pct_polar": 0.0})
        oo_anio = metricas.get("OO_ANIO", {"polar_qty": 0, "total_qty": 0, "pct_polar": 0.0})

        # Función para crear tarjeta de marca
        def make_marca_card(marca_title, marca_badge_color, marca_bg_header, sem_data, anio_data, border_color):
            pct_s = sem_data.get("pct_polar") or 0.0
            pct_a = anio_data.get("pct_polar") or 0.0
            
            # Condicional de meta (60.0%)
            cumple_sem = (pct_s >= 60.0)
            cumple_anio = (pct_a >= 60.0)
            
            badge_sem_bg = "#003311" if cumple_sem else "#330D00"
            badge_sem_fg = "#7CFC00" if cumple_sem else "#FF4500"
            badge_sem_txt = f"🎯 {pct_s}% (META 60%)" if cumple_sem else f"⚠️ {pct_s}% (META 60%)"
            
            badge_anio_bg = "#003311" if cumple_anio else "#330D00"
            badge_anio_fg = "#7CFC00" if cumple_anio else "#FF4500"
            badge_anio_txt = f"🎯 {pct_a}% (META 60%)" if cumple_anio else f"⚠️ {pct_a}% (META 60%)"

            return ft.Container(
                content=ft.Column([
                    # Header de la marca
                    ft.Container(
                        content=ft.Row([
                            ft.Text(marca_title, color="white", weight="bold", size=15),
                            ft.Container(expand=True),
                            ft.Container(
                                content=ft.Text("OBJETIVO MIN: 60%", color=marca_badge_color, size=10, weight="bold"),
                                bgcolor="#0A0A14",
                                padding=ft.padding.Padding(8, 3, 8, 3),
                                border_radius=6,
                                border=ft.Border.all(1, marca_badge_color)
                            )
                        ]),
                        bgcolor=marca_bg_header,
                        padding=ft.padding.Padding(12, 10, 12, 10),
                        border_radius=ft.BorderRadius(10, 10, 0, 0)
                    ),
                    
                    # Contenido de Semana y Año
                    ft.Container(
                        content=ft.Column([
                            # 1. SEMANA
                            ft.Text(f"📅 DESEMPEÑO SEMANAL ({sem_raw})", color="#00FFFF", size=12, weight="bold"),
                            ft.Row([
                                ft.Column([
                                    ft.Text("POLARIZADAS", color="#888888", size=10),
                                    ft.Text(f"🕶️ {sem_data.get('polar_qty', 0)} pzs", color="white", weight="bold", size=16)
                                ], expand=True),
                                ft.Column([
                                    ft.Text("TOTAL VENTAS", color="#888888", size=10),
                                    ft.Text(f"📦 {sem_data.get('total_qty', 0)} pzs", color="white", weight="bold", size=16)
                                ], expand=True),
                                ft.Column([
                                    ft.Text("% POLAR", color="#888888", size=10),
                                    ft.Container(
                                        content=ft.Text(f"{pct_s}%", color=badge_sem_fg, weight="bold", size=15),
                                        bgcolor=badge_sem_bg,
                                        padding=ft.padding.Padding(8, 3, 8, 3),
                                        border_radius=6,
                                        border=ft.Border.all(1, badge_sem_fg)
                                    )
                                ], expand=True)
                            ], spacing=10),
                            
                            ft.Divider(height=12, color="#222233"),
                            
                            # 2. AÑO (YTD)
                            ft.Text(f"📈 ACUMULADO DEL AÑO ({anio_raw})", color="#D8B4FE", size=12, weight="bold"),
                            ft.Row([
                                ft.Column([
                                    ft.Text("POLARIZADAS", color="#888888", size=10),
                                    ft.Text(f"🕶️ {anio_data.get('polar_qty', 0)} pzs", color="white", weight="bold", size=16)
                                ], expand=True),
                                ft.Column([
                                    ft.Text("TOTAL VENTAS", color="#888888", size=10),
                                    ft.Text(f"📦 {anio_data.get('total_qty', 0)} pzs", color="white", weight="bold", size=16)
                                ], expand=True),
                                ft.Column([
                                    ft.Text("% POLAR", color="#888888", size=10),
                                    ft.Container(
                                        content=ft.Text(f"{pct_a}%", color=badge_anio_fg, weight="bold", size=15),
                                        bgcolor=badge_anio_bg,
                                        padding=ft.padding.Padding(8, 3, 8, 3),
                                        border_radius=6,
                                        border=ft.Border.all(1, badge_anio_fg)
                                    )
                                ], expand=True)
                            ], spacing=10)
                        ], spacing=10),
                        padding=14
                    )
                ], spacing=0),
                bgcolor="#141424",
                border_radius=12,
                border=ft.Border.all(1.5, border_color),
                shadow=[ft.BoxShadow(color="#10000000", blur_radius=8, spread_radius=1)]
            )

        card_rb = make_marca_card("🔴 RAY-BAN (RB)", "#FF4500", "#330D00", rb_sem, rb_anio, "#FF4500")
        card_oo = make_marca_card("🔵 OAKLEY (OO)", "#00FFFF", "#002B36", oo_sem, oo_anio, "#00FFFF")

        if is_mobile:
            cards_layout = ft.Column([card_rb, card_oo], spacing=14)
        else:
            cards_layout = ft.Row([
                ft.Container(content=card_rb, expand=1),
                ft.Container(content=card_oo, expand=1)
            ], spacing=14, vertical_alignment=ft.CrossAxisAlignment.START)

        return ft.Column([
            cards_layout
        ], spacing=14)

    # -------------------------------------------------------------
    # RENDER DE TABLA GENERAL / RANKING DE ZONA
    # -------------------------------------------------------------
    def build_ranking_panel():
        is_mobile = (page.width < 750) if (page and page.width) else False
        sem_raw, anio_raw = semana_activa[0].split(" - ") if " - " in semana_activa[0] else (semana_activa[0], "2026")
        
        db = db_fn()
        filas_rb = []
        filas_oo = []
        
        if db:
            try:
                cur = db.cursor(dictionary=True)
                cur.execute("""
                    SELECT * FROM polar_metricas 
                    WHERE semana = %s AND anio = %s AND marca_periodo = 'RB_SEM'
                    ORDER BY pct_polar DESC, total_qty DESC
                """, (sem_raw.strip(), int(anio_raw.strip())))
                filas_rb = cur.fetchall()
                
                cur.execute("""
                    SELECT * FROM polar_metricas 
                    WHERE semana = %s AND anio = %s AND marca_periodo = 'OO_SEM'
                    ORDER BY pct_polar DESC, total_qty DESC
                """, (sem_raw.strip(), int(anio_raw.strip())))
                filas_oo = cur.fetchall()
                
                cur.close()
            except Exception as ex:
                print("Error ranking polar:", ex)
            db.close()

        def make_dt_rows(filas, marca_color):
            dt_rows = []
            for idx, r in enumerate(filas, start=1):
                t_nom = r.get("tienda") or ""
                is_tot = t_nom.upper().startswith("TOTAL")
                pct = r.get("pct_polar") or 0.0
                cumple = (pct >= 60.0)
                
                bg_c = "#003311" if cumple else "#330D00"
                fg_c = "#7CFC00" if cumple else "#FF4500"
                
                dt_rows.append(
                    ft.DataRow(
                        cells=[
                            ft.DataCell(ft.Text(f"{idx}" if not is_tot else "∑", color="#aaaaaa", size=11)),
                            ft.DataCell(ft.Text(t_nom, color="white" if not is_tot else marca_color, weight="bold" if is_tot else "normal", size=12)),
                            ft.DataCell(ft.Text(f"{r.get('polar_qty', 0)}", color="white", size=12)),
                            ft.DataCell(ft.Text(f"{r.get('total_qty', 0)}", color="white", size=12)),
                            ft.DataCell(
                                ft.Container(
                                    content=ft.Text(f"{pct}%", color=fg_c, weight="bold", size=11),
                                    bgcolor=bg_c,
                                    padding=ft.padding.Padding(6, 2, 6, 2),
                                    border_radius=4,
                                    border=ft.Border.all(1, fg_c)
                                )
                            ),
                        ],
                        color="#1A1A2E" if is_tot else (None if idx % 2 == 0 else "#10101C")
                    )
                )
            return dt_rows

        table_rb = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("#", color="#aaaaaa", size=11)),
                ft.DataColumn(ft.Text("TIENDA", color="#FF4500", weight="bold", size=12)),
                ft.DataColumn(ft.Text("POLAR", color="#aaaaaa", size=11)),
                ft.DataColumn(ft.Text("TOTAL", color="#aaaaaa", size=11)),
                ft.DataColumn(ft.Text("%POLAR", color="#FF4500", weight="bold", size=12)),
            ],
            rows=make_dt_rows(filas_rb, "#FF4500"),
            border=ft.Border.all(1, "#222233"),
            heading_row_color="#1E1E2E",
            heading_row_height=36,
            data_row_min_height=32,
            data_row_max_height=36
        )

        table_oo = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("#", color="#aaaaaa", size=11)),
                ft.DataColumn(ft.Text("TIENDA", color="#00FFFF", weight="bold", size=12)),
                ft.DataColumn(ft.Text("POLAR", color="#aaaaaa", size=11)),
                ft.DataColumn(ft.Text("TOTAL", color="#aaaaaa", size=11)),
                ft.DataColumn(ft.Text("%POLAR", color="#00FFFF", weight="bold", size=12)),
            ],
            rows=make_dt_rows(filas_oo, "#00FFFF"),
            border=ft.Border.all(1, "#222233"),
            heading_row_color="#1E1E2E",
            heading_row_height=36,
            data_row_min_height=32,
            data_row_max_height=36
        )

        panel_rb = ft.Column([
            ft.Text("🔴 RANKING RAY-BAN (RB)", color="#FF4500", weight="bold", size=14),
            ft.Container(content=ft.Row([table_rb], scroll=ft.ScrollMode.ADAPTIVE), bgcolor="#141424", border_radius=8, padding=6)
        ])
        
        panel_oo = ft.Column([
            ft.Text("🔵 RANKING OAKLEY (OO)", color="#00FFFF", weight="bold", size=14),
            ft.Container(content=ft.Row([table_oo], scroll=ft.ScrollMode.ADAPTIVE), bgcolor="#141424", border_radius=8, padding=6)
        ])

        if is_mobile:
            tables_layout = ft.Column([panel_rb, panel_oo], spacing=14)
        else:
            tables_layout = ft.Row([
                ft.Container(content=panel_rb, expand=1),
                ft.Container(content=panel_oo, expand=1)
            ], spacing=14, vertical_alignment=ft.CrossAxisAlignment.START)

        return ft.Column([
            ft.Text(f"📊 TABLA GENERAL DE DESEMPEÑO - {sem_raw} ({anio_raw})", color="#00FFFF", size=16, weight="bold"),
            ft.Text("Comparativa de polarizado Ray-Ban vs Oakley para todas las sucursales con meta al 60%.", color="#888888", size=12),
            ft.Divider(height=10, color="#333333"),
            tables_layout
        ], spacing=12)

    def recargar_vista():
        if vista_modo[0] == "tarjetas":
            btn_modo_tarjetas.bgcolor = "#003344"
            btn_modo_tarjetas.border = ft.Border.all(1.2, "#00FFFF")
            txt_modo_t.color = "#00FFFF"
            
            btn_modo_ranking.bgcolor = "#141424"
            btn_modo_ranking.border = ft.Border.all(1.2, "#2A2A3E")
            txt_modo_r.color = "#aaaaaa"
            
            content_area.controls = [build_metricas_panel()]
        else:
            btn_modo_tarjetas.bgcolor = "#141424"
            btn_modo_tarjetas.border = ft.Border.all(1.2, "#2A2A3E")
            txt_modo_t.color = "#aaaaaa"
            
            btn_modo_ranking.bgcolor = "#003344"
            btn_modo_ranking.border = ft.Border.all(1.2, "#00FFFF")
            txt_modo_r.color = "#00FFFF"
            
            content_area.controls = [build_ranking_panel()]
        try: page.update()
        except Exception: pass

    # Handlers
    def on_semana_change(e):
        if dd_semana.value:
            semana_activa[0] = dd_semana.value
            recargar_vista()

    dd_semana.on_change = on_semana_change

    def on_tienda_change(e):
        if dd_tienda.value:
            tienda_activa[0] = dd_tienda.value
            recargar_vista()

    dd_tienda.on_change = on_tienda_change

    def on_buscar_submit(e):
        if not es_admin:
            return
        if tf_buscar.value and tf_buscar.value.strip():
            matched = resolver_tienda_por_query(tf_buscar.value)
            if matched:
                tienda_activa[0] = matched
                dd_tienda.value = matched
                mostrar_snack(f"🔍 Tienda seleccionada: {matched}", "#7CFC00")
                recargar_vista()
            else:
                mostrar_snack(f"⚠️ No se encontró la tienda con: '{tf_buscar.value.strip()}'", "orange")

    tf_buscar.on_submit = on_buscar_submit
    tf_buscar.on_change = on_buscar_submit

    # -------------------------------------------------------------
    # SUBIDA DE EXCEL POLAR (ADMINISTRADORES)
    # -------------------------------------------------------------
    def subir_excel_click(e):
        if not seleccionar_archivo_async:
            mostrar_snack("Selector de archivo no disponible", "orange")
            return
            
        def on_excel_selected(path):
            if path and os.path.exists(path):
                dest_dir = POLAR_UPLOADS_DIR
                os.makedirs(dest_dir, exist_ok=True)
                fname = f"POLAR_{get_now_mexico_city().strftime('%Y%m%d_%H%M%S')}_{os.path.basename(path)}"
                dest_file = os.path.join(dest_dir, fname)
                try:
                    with open(path, "rb") as fi, open(dest_file, "wb") as fo:
                        fo.write(fi.read())
                        
                    ok, msg = procesar_excel_polar(dest_file, nombre_usuario, db_fn)
                    if ok:
                        mostrar_snack(msg, "#7CFC00")
                        
                        # Recargar semanas y tiendas disponibles
                        db_r = db_fn()
                        if db_r:
                            try:
                                cur_r = db_r.cursor(dictionary=True)
                                cur_r.execute("SELECT DISTINCT semana, anio FROM polar_metricas ORDER BY anio DESC, semana DESC")
                                semanas_disponibles.clear()
                                for r in cur_r.fetchall():
                                    semanas_disponibles.append(f"{r['semana']} - {r['anio']}")
                                
                                cur_r.execute("SELECT DISTINCT tienda FROM polar_metricas WHERE tienda NOT LIKE 'TOTAL%' ORDER BY tienda ASC")
                                tiendas_disponibles.clear()
                                for r in cur_r.fetchall():
                                    tiendas_disponibles.append(r["tienda"])
                                cur_r.close()
                            except Exception: pass
                            db_r.close()
                            
                        if semanas_disponibles:
                            dd_semana.options = [ft.dropdown.Option(s) for s in semanas_disponibles]
                            dd_semana.value = semanas_disponibles[0]
                            semana_activa[0] = semanas_disponibles[0]
                            
                        if tiendas_disponibles and es_admin:
                            dd_tienda.options = [make_tienda_opt(t) for t in tiendas_disponibles]
                            
                        recargar_vista()
                    else:
                        mostrar_snack(msg, "red")
                except Exception as ex_u:
                    print("Error subiendo excel polar:", ex_u)
                    mostrar_snack(f"Error procesando archivo: {str(ex_u)}", "red")

        seleccionar_archivo_async("Seleccionar Archivo Excel de Polar (POLAR Wxx.xlsx)", "file", on_excel_selected)

    btn_subir = ft.ElevatedButton(
        "📤 Subir Excel Polar",
        icon=ft.Icons.UPLOAD_FILE_ROUNDED,
        bgcolor="#7CFC00",
        color="#05070D",
        height=38,
        style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8), text_style=ft.TextStyle(weight="bold")),
        on_click=subir_excel_click,
        visible=es_admin
    )

    # Botones de alternar vista (Tarjetas / Ranking)
    txt_modo_t = ft.Text("Mi Tienda (RB & OO)", weight="bold", size=12, color="#00FFFF")
    txt_modo_r = ft.Text("Tabla Ranking Zona", weight="bold", size=12, color="#aaaaaa")

    btn_modo_tarjetas = ft.Container(
        content=ft.Row([ft.Text("🕶️", size=14), txt_modo_t], spacing=6, alignment=ft.MainAxisAlignment.CENTER),
        bgcolor="#003344",
        border=ft.Border.all(1.2, "#00FFFF"),
        border_radius=8,
        width=175,
        height=38,
        padding=ft.padding.Padding(10, 6, 10, 6),
        ink=True,
        on_click=lambda e: cambiar_modo("tarjetas")
    )
    
    btn_modo_ranking = ft.Container(
        content=ft.Row([ft.Text("📊", size=14), txt_modo_r], spacing=6, alignment=ft.MainAxisAlignment.CENTER),
        bgcolor="#141424",
        border=ft.Border.all(1.2, "#2A2A3E"),
        border_radius=8,
        width=175,
        height=38,
        padding=ft.padding.Padding(10, 6, 10, 6),
        ink=True,
        on_click=lambda e: cambiar_modo("ranking")
    )

    def cambiar_modo(nuevo_modo):
        vista_modo[0] = nuevo_modo
        recargar_vista()

    recargar_vista()

    # Layout Principal
    main_layout = ft.Column([
        # Fila 1: Encabezado y Botones de Modo / Carga
        ft.Row([
            ft.Row([
                ft.Text("🕶️", size=24),
                ft.Column([
                    ft.Text("MÓDULO POLAR RB Y OO", size=16, weight="bold", color="white"),
                    ft.Text("Métricas oficiales de Polarizado Ray-Ban y Oakley | Meta mínima: 60%", size=11, color="#00FFFF")
                ], spacing=1)
            ], vertical_alignment=ft.CrossAxisAlignment.CENTER),
            ft.Row([
                btn_modo_tarjetas,
                btn_modo_ranking,
                btn_subir
            ], spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER)
        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, vertical_alignment=ft.CrossAxisAlignment.CENTER),
        
        # Fila 2: Filtros (Semana, Buscador, Tienda)
        ft.Container(
            content=ft.Row([
                dd_semana,
                tf_buscar,
                ft.IconButton(
                    icon=ft.Icons.SEARCH_ROUNDED,
                    icon_color="#00FFFF",
                    tooltip="Buscar",
                    on_click=on_buscar_submit,
                    visible=es_admin
                ),
                dd_tienda if es_admin else txt_tienda_fija,
                ft.Container(expand=True),
                ft.IconButton(
                    icon=ft.Icons.REFRESH_ROUNDED,
                    icon_color="#aaaaaa",
                    tooltip="Refrescar",
                    on_click=lambda e: recargar_vista()
                )
            ], vertical_alignment=ft.CrossAxisAlignment.CENTER, spacing=8),
            padding=ft.padding.Padding(12, 8, 12, 8),
            bgcolor="#10101C",
            border_radius=10,
            border=ft.Border.all(1, "#222233")
        ),
        
        ft.Divider(height=6, color="#333333"),
        content_area,
        ft.Container(height=30)
    ], expand=True, spacing=10, scroll=ft.ScrollMode.AUTO)

    return main_layout
