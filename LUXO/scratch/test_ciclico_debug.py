import os
import openpyxl
import ciclicos_service as cs

file_escaneo = r"c:\Users\MOISES\Desktop\luxo7.7\LUXO\uploads\SOLARAQ32026(2).xlsx"
file_sap = r"c:\Users\MOISES\Desktop\luxo7.7\LUXO\uploads\StockSummary_Details(1).xlsx"

print("File Escaneo Exists:", os.path.exists(file_escaneo))
print("File SAP Exists:", os.path.exists(file_sap))

if os.path.exists(file_escaneo):
    wb = openpyxl.load_workbook(file_escaneo, data_only=True)
    print("Sheets in Escaneo:", wb.sheetnames)
    sheet = wb.active
    for sname in wb.sheetnames:
        if sname.strip().lower() == "scan":
            sheet = wb[sname]
            print("Found sheet:", sname)
            break
    
    rows = list(sheet.iter_rows(values_only=True))
    print("Total rows in sheet:", len(rows))
    for idx, r in enumerate(rows[:20]):
        print(f"Row {idx+1}: {r[:6]}")

    if os.path.exists(file_sap):
        res = cs.procesar_conciliacion_ciclico(file_escaneo, file_sap)
        print("Total SAP:", res["total_sap_pzas"])
        print("Total Escaneo:", res["total_escaneo_pzas"])
        print("Falta en Escaneo count:", len(res["falta_en_escaneo"]))
        print("Falta en SAP count:", len(res["falta_en_sap"]))
