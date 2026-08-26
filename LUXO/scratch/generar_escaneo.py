import os
import openpyxl
import re

input_file = r"C:\Users\MOISES\Downloads\inventario\StockSummary_Details(4).xlsx"
output_file = r"C:\Users\MOISES\Downloads\inventario\Escaneo_Fisico_Generado.xlsx"

if not os.path.exists(input_file):
    print("ERROR: No se encontró el archivo:", input_file)
    exit(1)

wb = openpyxl.load_workbook(input_file, data_only=True)
sheet = wb.active
if "StockSummary_Details" in wb.sheetnames:
    sheet = wb["StockSummary_Details"]

rows = list(sheet.iter_rows(values_only=True))
print("Total de filas en SAP:", len(rows))

# Detectar índices de Columna C (UPC) y Columna J (Cantidad en tienda)
col_upc_idx = 2    # Col C
col_stock_idx = 9  # Col J

header_row = 0
for idx, r in enumerate(rows[:10]):
    if not r: continue
    r_str = [str(c).lower() if c is not None else "" for c in r]
    for c_idx, cell_text in enumerate(r_str):
        if any(k in cell_text for k in ["ean/upc", "ean", "upc", "codigo"]):
            col_upc_idx = c_idx
        if any(k in cell_text for k in ["cantidad en tienda", "tienda", "libre utilización"]):
            col_stock_idx = c_idx
    if "ean/upc" in " ".join(r_str) or "descripción" in " ".join(r_str):
        header_row = idx
        break

print(f"Columna UPC: idx {col_upc_idx}, Columna Stock: idx {col_stock_idx}, Header row: {header_row}")

def limpiar_upc(val):
    if val is None: return ""
    v_str = str(val).strip()
    if 'e' in v_str.lower() or '.' in v_str:
        try:
            v_float = float(v_str)
            v_str = f"{v_float:.0f}"
        except Exception: pass
    cleaned = re.sub(r'\D', '', v_str)
    return cleaned.lstrip('0')

lista_escaneo = []

for r in rows[header_row + 1:]:
    if not r or len(r) <= col_upc_idx or r[col_upc_idx] is None:
        continue
    upc_clean = limpiar_upc(r[col_upc_idx])
    if not upc_clean or len(upc_clean) < 5:
        continue
        
    stock_raw = r[col_stock_idx] if col_stock_idx < len(r) else 0
    try:
        cant_pzas = int(float(str(stock_raw).strip())) if stock_raw is not None else 0
    except Exception:
        cant_pzas = 0
        
    if cant_pzas > 0:
        for _ in range(cant_pzas):
            lista_escaneo.append(upc_clean)

print("Total de piezas en la lista de escaneo generada:", len(lista_escaneo))

# Crear el nuevo libro de Excel de Escaneo Físico
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Scan"

ws_out.cell(row=1, column=1, value="UPC")
ws_out.cell(row=1, column=2, value="Piezas")

for idx, upc_val in enumerate(lista_escaneo, start=2):
    ws_out.cell(row=idx, column=1, value=upc_val)
    ws_out.cell(row=idx, column=2, value=1)

wb_out.save(output_file)
print("Archivo guardado exitosamente en:", output_file)

# Guardar una copia adicional en Downloads directamente
output_file_downloads = r"C:\Users\MOISES\Downloads\Escaneo_Fisico_Generado.xlsx"
wb_out.save(output_file_downloads)
print("Copia guardada exitosamente en:", output_file_downloads)
