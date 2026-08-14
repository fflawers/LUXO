import re

with open("enfoque_diario.py", "r") as f:
    code = f.read()

# Replace guardar_semana_historico
code = code.replace(
"""def guardar_semana_historico():
    key = f"S{global_meta['semana']}_{global_meta['num_tienda']}_{global_meta['tienda']}"
    import copy
    historico_semanal_state[key] = copy.deepcopy(store_state)
    guardar_estado_persistente()""",
"""def guardar_semana_historico(user_id):
    g_meta = user_states[user_id]["global_meta"]
    s_state = user_states[user_id]["store_state"]
    h_state = user_states[user_id]["historico_semanal_state"]
    key = f"S{g_meta['semana']}_{g_meta['num_tienda']}_{g_meta['tienda']}"
    import copy
    h_state[key] = copy.deepcopy(s_state)
    guardar_estado_persistente(user_id)"""
)

# Replace cargar_semana_historico
code = code.replace(
"""def cargar_semana_historico(num_semana):
    global_meta["semana"] = str(num_semana)
    key = f"S{num_semana}_{global_meta['num_tienda']}_{global_meta['tienda']}"
    if key in historico_semanal_state:
        import copy
        global store_state
        store_state.clear()
        store_state.update(copy.deepcopy(historico_semanal_state[key]))
    guardar_estado_persistente()""",
"""def cargar_semana_historico(user_id, num_semana):
    g_meta = user_states[user_id]["global_meta"]
    s_state = user_states[user_id]["store_state"]
    h_state = user_states[user_id]["historico_semanal_state"]
    g_meta["semana"] = str(num_semana)
    key = f"S{num_semana}_{g_meta['num_tienda']}_{g_meta['tienda']}"
    if key in h_state:
        import copy
        s_state.clear()
        s_state.update(copy.deepcopy(h_state[key]))
    guardar_estado_persistente(user_id)"""
)

# Replace sincronizar_colaboradores_db
code = code.replace(
"""def sincronizar_colaboradores_db(user_info=None, tienda_name=None):
    \"\"\"Consulta los colaboradores registrados en la base de datos de Configuración de Tienda y los auto-llena en Enfoque Diario 2026.\"\"\"
    db_names = []
    try:
        target_t = tienda_name or global_meta.get("tienda", "Vallejo")""",
"""def sincronizar_colaboradores_db(user_info=None, tienda_name=None, user_id=None):
    \"\"\"Consulta los colaboradores registrados en la base de datos de Configuración de Tienda y los auto-llena en Enfoque Diario 2026.\"\"\"
    if not user_id: return
    g_meta = user_states[user_id]["global_meta"]
    s_state = user_states[user_id]["store_state"]
    db_names = []
    try:
        target_t = tienda_name or g_meta.get("tienda", "Vallejo")"""
)

code = code.replace(
"""    if db_names:
        for d in DIAS:
            for i in range(8):
                if i < len(db_names):
                    store_state[d]["colaboradores"][i]["nombre"] = db_names[i]
                    if store_state[d]["colaboradores"][i]["horas"] <= 0:
                        store_state[d]["colaboradores"][i]["horas"] = 10.0 if i == 0 else 8.0
                else:
                    store_state[d]["colaboradores"][i]["nombre"] = ""
                    store_state[d]["colaboradores"][i]["horas"] = 0.0
        guardar_estado_persistente()""",
"""    if db_names:
        for d in DIAS:
            for i in range(8):
                if i < len(db_names):
                    s_state[d]["colaboradores"][i]["nombre"] = db_names[i]
                    if s_state[d]["colaboradores"][i]["horas"] <= 0:
                        s_state[d]["colaboradores"][i]["horas"] = 10.0 if i == 0 else 8.0
                else:
                    s_state[d]["colaboradores"][i]["nombre"] = ""
                    s_state[d]["colaboradores"][i]["horas"] = 0.0
        guardar_estado_persistente(user_id)"""
)

# Replace calcular_dia
code = code.replace(
"""def calcular_dia(d_name):
    data = store_state[d_name]""",
"""def calcular_dia(d_name, user_id):
    if user_id not in user_states: init_user_state(user_id)
    s_state = user_states[user_id]["store_state"]
    data = s_state[d_name]"""
)

# Replace generar_excel_enfoque
code = code.replace(
"""def generar_excel_enfoque(d_name, page=None):
    try:
        import openpyxl
        calc = calcular_dia(d_name)
        data = store_state[d_name]""",
"""def generar_excel_enfoque(d_name, user_id, page=None):
    try:
        import openpyxl
        calc = calcular_dia(d_name, user_id)
        g_meta = user_states[user_id]["global_meta"]
        s_state = user_states[user_id]["store_state"]
        data = s_state[d_name]"""
)
code = code.replace(
"""                    ws['I1'] = int(global_meta['semana']) if str(global_meta['semana']).isdigit() else global_meta['semana']
                    ws['K1'] = datetime.datetime.now()
                    ws['M1'] = global_meta['tienda']""",
"""                    ws['I1'] = int(g_meta['semana']) if str(g_meta['semana']).isdigit() else g_meta['semana']
                    ws['K1'] = datetime.datetime.now()
                    ws['M1'] = g_meta['tienda']"""
)
code = code.replace(
"""                    ws = wb[d]
                    d_data = store_state[d]""",
"""                    ws = wb[d]
                    d_data = s_state[d]"""
)
code = code.replace(
"""            ws['A2'] = f"DÍA: {d_name} | SEMANA: {global_meta['semana']} | TIENDA: {global_meta['tienda']}\"""",
"""            ws['A2'] = f"DÍA: {d_name} | SEMANA: {g_meta['semana']} | TIENDA: {g_meta['tienda']}\""""
)

# Replace generar_pdf_enfoque_file
code = code.replace(
"""def generar_pdf_enfoque_file(d_name):
    if not REPORTLAB_AVAILABLE:
        return None
        
    try:
        calc = calcular_dia(d_name)
        data = store_state[d_name]""",
"""def generar_pdf_enfoque_file(d_name, user_id):
    if not REPORTLAB_AVAILABLE:
        return None
        
    try:
        calc = calcular_dia(d_name, user_id)
        g_meta = user_states[user_id]["global_meta"]
        s_state = user_states[user_id]["store_state"]
        data = s_state[d_name]"""
)
code = code.replace(
"""        header_data = [
            [Paragraph(f"<b>DÍA:</b> {d_name}", header_style),
             Paragraph(f"<b>SEMANA:</b> {global_meta['semana']}", header_style),
             Paragraph(f"<b>TIENDA:</b> {global_meta['tienda']}", header_style)]
        ]""",
"""        header_data = [
            [Paragraph(f"<b>DÍA:</b> {d_name}", header_style),
             Paragraph(f"<b>SEMANA:</b> {g_meta['semana']}", header_style),
             Paragraph(f"<b>TIENDA:</b> {g_meta['tienda']}", header_style)]
        ]"""
)

with open("enfoque_diario.py", "w") as f:
    f.write(code)

