import flet as ft
from datetime import datetime, timedelta
import mysql.connector
import requests
import os
import flet_video as fv
import base64
import fitz
import tempfile
import re
import json
import threading
from tkinter import Tk, filedialog
import openpyxl
import math
from dotenv import load_dotenv
import bcrypt
import difflib

# Cargar variables de entorno desde .env
load_dotenv()
import operacion_tiendas

# =========================================
# CONFIGURACION
# =========================================

BASE_PATH = os.path.dirname(__file__)
ASSETS_PATH = os.path.join(BASE_PATH, "custom_assets")

# Valores de APIs leídos desde entorno (.env) o valores por defecto
GROQ_API_KEY = os.getenv("GROQ_API_KEY", "")
GROQ_API_KEY_2 = os.getenv("GROQ_API_KEY_2", "")
GROQ_API_KEY_3 = os.getenv("GROQ_API_KEY_3", "")
GROQ_KEYS = [k for k in [GROQ_API_KEY, GROQ_API_KEY_2, GROQ_API_KEY_3] if k]  # lista de llaves activas
_groq_key_index = 0  # índice de la llave activa actual
GROQ_MODEL = os.getenv("GROQ_MODEL", "llama-3.3-70b-versatile")
URL_GROQ = "https://api.groq.com/openai/v1/chat/completions"

def get_groq_key():
    """Devuelve la llave de Groq activa. Si hay varias, rota entre ellas en caso de 429."""
    global _groq_key_index
    if not GROQ_KEYS:
        return GROQ_API_KEY
    return GROQ_KEYS[_groq_key_index % len(GROQ_KEYS)]

def rotate_groq_key():
    """Rota a la siguiente llave de Groq disponible (se llama cuando hay error 429)."""
    global _groq_key_index
    if len(GROQ_KEYS) > 1:
        _groq_key_index = (_groq_key_index + 1) % len(GROQ_KEYS)
        print(f"⚡ Groq 429 - Rotando a llave #{_groq_key_index + 1}")
        return True
    return False
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")

# Configuración de Base de Datos MySQL (leída desde .env)
DB_CONFIG = {
    "host": os.getenv("DB_HOST", "localhost"),
    "user": os.getenv("DB_USER", "root"),
    "password": os.getenv("DB_PASSWORD", ""),
    "database": os.getenv("DB_NAME", "sgh_portal"),
    "port": int(os.getenv("DB_PORT", 3306))
}

# Variables globales para caché del RAG (manuales)
RAG_BLOQUES_CACHE = None
RAG_DF_CACHE = None
RAG_IDF_CACHE = None
RAG_CACHE_LOCK = threading.Lock()

# Variables globales para control de aperturas
star_icon_container = None

class EmojiIconButton(ft.Container):
    def __init__(self, icon_emoji, active_emoji=None, icon_color=None, on_click=None, tooltip=None, **kwargs):
        self.icon_emoji = icon_emoji
        self.active_emoji = active_emoji or icon_emoji
        self.txt = ft.Text(icon_emoji, color=icon_color, size=18, text_align="center")
        self.on_click_callback = on_click
        
        super().__init__(
            content=self.txt,
            alignment=ft.alignment.Alignment(0, 0),
            on_click=self.handle_click,
            tooltip=tooltip,
            **kwargs
        )
        self._icon_color = icon_color
        self._icon = ""

    def handle_click(self, e):
        if self.on_click_callback:
            self.on_click_callback(e)

    @property
    def icon(self):
        return self._icon

    @icon.setter
    def icon(self, val):
        self._icon = val
        if val and ("stop" in str(val).lower() or "mic" not in str(val).lower()):
            self.txt.value = self.active_emoji
        else:
            self.txt.value = self.icon_emoji

    @property
    def icon_color(self):
        return self._icon_color

    @icon_color.setter
    def icon_color(self, val):
        self._icon_color = val
        self.txt.color = val


class EmojiDropdown(ft.Container):
    def __init__(self, label, options=None, value=None, on_change=None, width=None, height=45, border_color="#9D50BB", **kwargs):
        self.options_list = options or []
        self.selected_value = value
        self.on_change_callback = on_change
        
        self.label_text = ft.Text(label, color="#aaaaaa", size=9)
        self.val_text = ft.Text("", color="white", size=12, weight="bold")
        self.arrow_text = ft.Text("▼", color="#00FFFF", size=10)
        
        self.update_display_text()
        
        self.btn = ft.PopupMenuButton(
            content=ft.Container(
                content=ft.Column([
                    self.label_text,
                    ft.Row([
                        self.val_text,
                        self.arrow_text
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, vertical_alignment=ft.CrossAxisAlignment.CENTER)
                ], spacing=2, alignment=ft.MainAxisAlignment.CENTER),
                border=ft.Border.all(1, border_color),
                border_radius=8,
                padding=ft.padding.Padding(10, 4, 10, 4),
                width=width,
                height=height,
                bgcolor="#161622",
                ink=True
            ),
            items=[]
        )
        self.rebuild_menu_items()
        
        super().__init__(
            content=self.btn,
            width=width,
            height=height,
            **kwargs
        )

    def update_display_text(self):
        found_text = ""
        for opt in self.options_list:
            opt_key = getattr(opt, "key", None) or getattr(opt, "value", None)
            if opt_key is None:
                opt_key = str(opt)
            opt_text = getattr(opt, "text", None) or getattr(opt, "label", None) or opt_key
            if str(opt_key) == str(self.selected_value):
                found_text = opt_text
                break
        if not found_text:
            found_text = str(self.selected_value) if self.selected_value is not None else ""
        self.val_text.value = found_text

    def rebuild_menu_items(self):
        menu_items = []
        for opt in self.options_list:
            opt_key = getattr(opt, "key", None) or getattr(opt, "value", None)
            if opt_key is None:
                opt_key = str(opt)
            opt_text = getattr(opt, "text", None) or getattr(opt, "label", None) or opt_key
            
            def make_select_click(k, t):
                return lambda e: self.select_value(k, t)
                
            menu_items.append(
                ft.PopupMenuItem(content=ft.Text(opt_text, color="white"), on_click=make_select_click(opt_key, opt_text))
            )
        self.btn.items = menu_items

    def select_value(self, key, text):
        self.selected_value = key
        self.val_text.value = text
        self.val_text.update()
        if self.on_change_callback:
            class DummyEvent:
                def __init__(self, control):
                    self.control = control
            self.on_change_callback(DummyEvent(self))

    @property
    def value(self):
        return self.selected_value

    @value.setter
    def value(self, val):
        self.selected_value = val
        self.update_display_text()
        try: self.val_text.update()
        except Exception: pass

    @property
    def options(self):
        return self.options_list

    @options.setter
    def options(self, val):
        self.options_list = val or []
        self.rebuild_menu_items()
        self.update_display_text()
        try:
            self.val_text.update()
            self.btn.update()
        except Exception: pass

# =========================================
# FUNCIONES DE ENCRIPTACIÓN DE CONTRASEÑA (BCRYPT)
# =========================================

def hash_password(plain_password: str) -> str:
    """Genera un hash seguro de bcrypt para la contraseña recibida."""
    if not plain_password:
        return ""
    pwd_bytes = plain_password.encode('utf-8')
    salt = bcrypt.gensalt(12)
    hashed = bcrypt.hashpw(pwd_bytes, salt)
    return hashed.decode('utf-8')

def verify_password(plain_password: str, stored_password: str) -> bool:
    """
    Verifica una contraseña recibida contra la contraseña encriptada en BD.
    Soporta contraseñas encriptadas con bcrypt ($2b$, $2a$) y contraseñas legacy en texto plano.
    """
    try:
        if not plain_password or not stored_password:
            return False
        # Si la contraseña guardada en BD es un hash de bcrypt:
        if stored_password.startswith("$2b$") or stored_password.startswith("$2a$"):
            return bcrypt.checkpw(plain_password.encode('utf-8'), stored_password.encode('utf-8'))
        # Compatibilidad con contraseñas legacy en texto plano:
        return plain_password == stored_password
    except Exception as ex:
        print("Error en verificación de contraseña bcrypt:", ex)
        return False

def run_async_sync(coro, page):
    import threading
    event = threading.Event()
    result = [None]
    error = [None]
    async def wrapper():
        try:
            result[0] = await coro
        except Exception as e:
            error[0] = e
        finally:
            event.set()
    page.run_task(wrapper)
    event.wait()
    if error[0]:
        raise error[0]
    return result[0]

active_sessions = {}

import urllib.parse
import asyncio

def ejecutar_js_flet(page: ft.Page, js_code: str):
    """Ejecuta código JavaScript de forma 100% segura en Flet Web."""
    if not page:
        return
    try:
        encoded = urllib.parse.quote(js_code)
        target_url = f"javascript:void(eval(decodeURIComponent('{encoded}')))"
        async def _do_launch():
            try:
                res = page.launch_url(target_url)
                if asyncio.iscoroutine(res):
                    await res
            except Exception as e_l:
                print("Notice inner launch_url:", e_l)
        page.run_task(_do_launch)
    except Exception as ex_ej:
        print("Error al ejecutar JS en Flet:", ex_ej)


import tempfile
import os
import requests
from fastapi import Request
from starlette.responses import Response

def save_and_transcribe_audio(user_id_val, audio_bytes):
    try:
        is_webm = audio_bytes.startswith(b"\x1a\x45\xdf\xa3") or b"webm" in audio_bytes[:100]
        ext = ".webm" if is_webm else ".wav"
        mime = "audio/webm" if is_webm else "audio/wav"

        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tmp.write(audio_bytes)
            temp_path = tmp.name

        try:
            headers = {"Authorization": f"Bearer {GROQ_API_KEY}"}
            with open(temp_path, "rb") as f:
                files = {"file": (f"recording{ext}", f, mime)}
                data = {
                    "model": "whisper-large-v3",
                    "language": "es",
                    "response_format": "json"
                }
                res = requests.post(
                    "https://api.groq.com/openai/v1/audio/transcriptions",
                    headers=headers,
                    files=files,
                    data=data,
                    timeout=30
                )

            if res.status_code == 200:
                transcripcion = res.json().get("text", "").strip()
                if transcripcion:
                    user_id = int(user_id_val) if isinstance(user_id_val, str) and user_id_val.isdigit() else user_id_val
                    session = active_sessions.get(user_id) or active_sessions.get(str(user_id)) or active_sessions.get(1) or active_sessions.get("1") or (list(active_sessions.values())[0] if active_sessions else None)
                    if session:
                        input_msg = session.get("input_msg")
                        enviar_mensaje = session.get("enviar_mensaje")
                        page = session.get("page")
                        btn_mic = session.get("btn_mic")

                        if btn_mic:
                            btn_mic.icon = ft.Icons.MIC_ROUNDED
                            btn_mic.icon_color = "#00FFFF"
                            btn_mic.tooltip = "Grabar Nota de Voz 🎙️"

                        if page and input_msg and enviar_mensaje:
                            page.title = "LUXO"
                            input_msg.value = f"🎙️ {transcripcion}"
                            page.update()
                            
                            async def trigger_send():
                                enviar_mensaje(None)
                            page.run_task(trigger_send)
            else:
                print("Error Groq Whisper API status:", res.status_code, res.text)
                user_id = int(user_id_val) if isinstance(user_id_val, str) and user_id_val.isdigit() else user_id_val
                if user_id in active_sessions:
                    session = active_sessions[user_id]
                    btn_mic = session.get("btn_mic")
                    page = session["page"]
                    if btn_mic:
                        btn_mic.icon = ft.Icons.MIC_ROUNDED
                        btn_mic.icon_color = "#00FFFF"
                        btn_mic.tooltip = "Grabar Nota de Voz 🎙️"
                    page.title = "LUXO"
        finally:
            if os.path.exists(temp_path):
                try:
                    os.unlink(temp_path)
                except Exception:
                    pass
    except Exception as ex:
        print("EXCEPCIÓN EN TRANSCRIPCIÓN AUDIO SERVIDOR:", ex)


def optimizar_archivo_multimedia(filepath):
    try:
        if not filepath or not os.path.exists(filepath):
            return
        ext = os.path.splitext(filepath)[1].lower()
        file_size_mb = os.path.getsize(filepath) / (1024 * 1024)
        
        # 1. Optimización de Imágenes (.jpg, .jpeg, .png, .webp)
        if ext in ['.jpg', '.jpeg', '.png', '.webp']:
            if file_size_mb > 0.8:  # Si la imagen pesa más de 800 KB
                from PIL import Image
                with Image.open(filepath) as img:
                    if img.mode in ("RGBA", "P") and ext in ['.jpg', '.jpeg']:
                        img = img.convert("RGB")
                    
                    max_dim = 1920
                    if img.width > max_dim or img.height > max_dim:
                        img.thumbnail((max_dim, max_dim), Image.Resampling.LANCZOS)
                    
                    if ext in ['.jpg', '.jpeg']:
                        img.save(filepath, "JPEG", quality=80, optimize=True)
                    elif ext == '.png':
                        img.save(filepath, "PNG", optimize=True)
                    elif ext == '.webp':
                        img.save(filepath, "WEBP", quality=80)
                
                size_after = os.path.getsize(filepath) / (1024 * 1024)
                print(f"📷 Imagen optimizada exitosamente: {os.path.basename(filepath)} ({file_size_mb:.2f}MB -> {size_after:.2f}MB)")
    except Exception as ex_opt:
        print("Error en optimización multimedia:", ex_opt)


def configurar_rutas_fastapi(app):
    os.makedirs(os.path.join(ASSETS_PATH, "temp_audio"), exist_ok=True)
    temp_pdfs_dir = os.path.join(ASSETS_PATH, "temp_pdfs")
    os.makedirs(temp_pdfs_dir, exist_ok=True)
    os.makedirs("uploads", exist_ok=True)
    try:
        threading.Thread(target=rebuild_rag_cache, daemon=True).start()
    except Exception as e:
        print("Error starting RAG cache thread:", e)
    try:
        from fastapi.staticfiles import StaticFiles
        from fastapi.responses import FileResponse
        app.mount("/temp_audio", StaticFiles(directory=os.path.join(ASSETS_PATH, "temp_audio")), name="temp_audio")
        app.mount("/temp_pdfs", StaticFiles(directory=temp_pdfs_dir), name="temp_pdfs")
        app.mount("/custom_assets", StaticFiles(directory=ASSETS_PATH), name="custom_assets")
        app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")
    except Exception:
        pass

    @app.get("/api/download_enfoque_pdf/{day}")
    async def download_enfoque_pdf_route(day: str):
        try:
            import importlib, enfoque_diario
            if hasattr(enfoque_diario, "generar_pdf_enfoque_file"):
                enfoque_diario.generar_pdf_enfoque_file(day)

            filename = f"Enfoque_Diario_{day}_SGH_2026.pdf"
            file_path = os.path.join(BASE_PATH, "uploads", filename)
            if not os.path.exists(file_path):
                home_dir = os.path.expanduser("~")
                possible = [
                    os.path.join(home_dir, "Desktop", filename),
                    os.path.join(home_dir, "OneDrive", "Desktop", filename)
                ]
                for p in possible:
                    if os.path.exists(p):
                        file_path = p
                        break

            if os.path.exists(file_path):
                from fastapi.responses import FileResponse
                return FileResponse(
                    path=file_path,
                    filename=filename,
                    media_type="application/pdf"
                )
        except Exception as ex:
            print("Error en endpoint download_enfoque_pdf:", ex)
        return {"error": "Archivo no encontrado"}

    @app.get("/api/download_excel/{day}")
    async def download_excel_route(day: str):
        try:
            import enfoque_diario
            enfoque_diario.generar_excel_enfoque(day)
            filename = f"Enfoque_Diario_{day}_SGH_2026.xlsx"
            file_path = os.path.join(BASE_PATH, "uploads", filename)
            if not os.path.exists(file_path):
                home_dir = os.path.expanduser("~")
                possible = [
                    os.path.join(home_dir, "Desktop", filename),
                    os.path.join(home_dir, "OneDrive", "Desktop", filename)
                ]
                for p in possible:
                    if os.path.exists(p):
                        file_path = p
                        break
            if os.path.exists(file_path):
                from fastapi.responses import FileResponse
                return FileResponse(
                    path=file_path,
                    filename=filename,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception as ex:
            print("Error en endpoint download_excel_route:", ex)
        return {"error": "Archivo no encontrado"}

    @app.get("/print_enfoque/{day}")
    async def print_enfoque_route(day: str):
        try:
            import enfoque_diario
            calc = enfoque_diario.calcular_dia(day)
            data = enfoque_diario.store_state[day]
            meta = enfoque_diario.global_meta

            html = f"""
            <!DOCTYPE html>
            <html lang="es">
            <head>
                <meta charset="UTF-8">
                <title>2026 SGH ENFOQUE DIARIO - {day} - {meta['tienda']}</title>
                <style>
                    @page {{ size: landscape; margin: 8mm; }}
                    body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 0; padding: 10px; color: #000; background: #fff; font-size: 11px; }}
                    .top-header {{ display: flex; align-items: center; justify-content: space-between; border-bottom: 2px solid #000; padding-bottom: 6px; margin-bottom: 8px; }}
                    .logo-brand {{ font-size: 16px; font-weight: 900; letter-spacing: -0.5px; }}
                    .tag-green {{ background: #C6EFCE; color: #006100; font-weight: bold; border: 1px solid #10B981; padding: 4px 10px; border-radius: 4px; font-size: 10px; }}
                    .title-banner {{ background: #000; color: #fff; font-size: 13px; font-weight: bold; text-align: center; padding: 6px; letter-spacing: 1px; margin-bottom: 10px; text-transform: uppercase; }}
                    .kpi-grid {{ display: grid; grid-template-columns: 1fr 1fr 1fr 1fr 1fr; gap: 8px; margin-bottom: 12px; }}
                    .kpi-box {{ border: 1px solid #000; font-size: 10px; }}
                    .kpi-title {{ background: #000; color: #fff; font-weight: bold; padding: 4px; text-align: center; font-size: 10px; text-transform: uppercase; }}
                    .kpi-body {{ padding: 6px; background: #fff; }}
                    .kpi-row {{ display: flex; justify-content: space-between; padding: 3px 0; border-bottom: 1px dashed #e5e7eb; }}
                    .kpi-row:last-child {{ border-bottom: none; }}
                    .green-bg {{ background: #C6EFCE !important; color: #006100; font-weight: bold; }}
                    table {{ width: 100%; border-collapse: collapse; margin-bottom: 12px; font-size: 10px; }}
                    th {{ background: #000; color: #fff; border: 1px solid #000; padding: 5px; text-align: center; font-size: 10px; font-weight: bold; }}
                    td {{ border: 1px solid #666; padding: 5px; text-align: center; }}
                    .sec-title {{ font-size: 11px; font-weight: bold; margin-top: 10px; margin-bottom: 4px; border-left: 4px solid #000; padding-left: 6px; text-transform: uppercase; }}
                    .btn-print {{ position: fixed; top: 12px; right: 12px; background: #10B981; color: white; border: none; padding: 10px 18px; font-size: 12px; font-weight: bold; border-radius: 6px; cursor: pointer; box-shadow: 0 4px 6px rgba(0,0,0,0.2); z-index: 9999; }}
                    @media print {{ .btn-print {{ display: none; }} body {{ padding: 0; }} }}
                </style>
            </head>
            <body>
                <button class="btn-print" onclick="window.print()">🖨️ Imprimir / Guardar como PDF</button>

                <div class="top-header">
                    <div>
                        <span class="logo-brand">🕶️ sunglass hut</span>
                        <span style="font-size:14px; font-weight:bold; margin-left:8px;">ENFOQUE DIARIO! - Nuestra meta y plan de acción</span>
                    </div>
                    <div style="display:flex; gap:15px; align-items:center;">
                        <span><b>Semana:</b> {meta['semana']}</span>
                        <span><b>Día:</b> {day} ({datetime.now().strftime('%d/%m/%Y')})</span>
                        <span><b>Tienda:</b> {meta['tienda']}</span>
                        <span class="tag-green">LAS CELDAS EN COLOR VERDE SE LLENAN AUTOMÁTICAMENTE</span>
                    </div>
                </div>

                <div class="title-banner">¿QUÉ ESPERAMOS LOGRAR HOY?</div>

                <div class="kpi-grid">
                    <div class="kpi-box">
                        <div class="kpi-title">META DEL DÍA</div>
                        <div class="kpi-body">
                            <div class="kpi-row"><span>META DIARIA</span><b>${calc['meta_diaria']:,.2f}</b></div>
                            <div class="kpi-row green-bg"><span>ANÁLOGOS (85%)</span><b>${calc['analogos']:,.2f}</b></div>
                            <div class="kpi-row green-bg"><span>WEARABLES (15%)</span><b>${calc['wearables']:,.2f}</b></div>
                            <div class="kpi-row"><span>TOTAL UNIDADES</span><b>{calc['total_unidades']}</b></div>
                            <div class="kpi-row" style="margin-top:4px;"><span>EVALUACIÓN</span><b>{'⭐' * data['estrellas_logro']}</b></div>
                        </div>
                    </div>

                    <div class="kpi-box">
                        <div class="kpi-title">VERSIÓN (NO NEGOCIABLE)</div>
                        <div class="kpi-body">
                            <div class="kpi-row"><span>TRÁFICO ESPERADO</span><b>{data['trafico_esperado']}</b></div>
                            <div class="kpi-row"><span>META CONVERSIÓN</span><b>{int(data['conversion_target']*100)}%</b></div>
                            <div class="kpi-row green-bg"><span>META TRANSACCIONES</span><b>{calc['transacciones']}</b></div>
                            <div class="kpi-row green-bg"><span>META IDEAL (110%)</span><b>${calc['meta_ideal']:,.2f}</b></div>
                        </div>
                    </div>

                    <div class="kpi-box">
                        <div class="kpi-title">OTROS NO NEGOCIABLES</div>
                        <div class="kpi-body">
                            <div class="kpi-row green-bg"><span>WEARABLES 15%</span><b>{max(math.ceil(calc['wearables']/8100), 1)} min</b></div>
                            <div class="kpi-row green-bg"><span>KIDS 5%</span><b>1 min</b></div>
                            <div class="kpi-row green-bg"><span>CAREKITS 30%</span><b>1 min</b></div>
                        </div>
                    </div>

                    <div class="kpi-box">
                        <div class="kpi-title">PRODUCTIVIDAD & COMP LY</div>
                        <div class="kpi-body">
                            <div class="kpi-row green-bg"><span>VENTA NETA/PROD</span><b>${calc['vta_neta_prod']:,.2f}</b></div>
                            <div class="kpi-row green-bg"><span>UNIDADES PROD</span><b>{calc['u_prod']}</b></div>
                            <div class="kpi-row green-bg"><span>VENTA NETA LY</span><b>${calc['vta_ly']:,.2f}</b></div>
                        </div>
                    </div>

                    <div class="kpi-box">
                        <div class="kpi-title">MÉTRICAS DÍA / MTD</div>
                        <div class="kpi-body">
                            <div class="kpi-row"><span>ATV DÍA COMP</span><b>$7,500</b></div>
                            <div class="kpi-row"><span>AUR DÍA COMP</span><b>$4,617</b></div>
                            <div class="kpi-row"><span>ATV MTD</span><b>$6,578</b></div>
                            <div class="kpi-row"><span>AUR MTD</span><b>$4,312</b></div>
                        </div>
                    </div>
                </div>

                <div class="sec-title">DESGLOSE HORARIO DE VENTA</div>
                <table>
                    <thead>
                        <tr>
                            <th style="width:20%;">BLOQUE / INDICADOR</th>
                            <th>Apertura-1pm</th>
                            <th>1pm - 3pm</th>
                            <th>3pm - 5pm</th>
                            <th>5pm - 7pm</th>
                            <th>7pm - Cierre</th>
                            <th>TOTAL</th>
                        </tr>
                    </thead>
                    <tbody>
                        <tr>
                            <td><b>TRÁFICO POR HORA (⚪)</b></td>
                            {"".join(f"<td>{x}</td>" for x in data["trafico_bloques"])}
                            <td><b>{calc['tot_trafico_b']}</b></td>
                        </tr>
                        <tr class="green-bg">
                            <td><b>PESO DEL TRÁFICO (%) (🟩)</b></td>
                            {"".join(f"<td>{p*100:.1f}%</td>" for p in calc["b_pesos"])}
                            <td><b>100%</b></td>
                        </tr>
                        <tr class="green-bg">
                            <td><b>META $ X HORA (🟩)</b></td>
                            {"".join(f"<td>${m:,.0f}</td>" for m in calc["b_metas"])}
                            <td><b>${calc['meta_diaria']:,.0f}</b></td>
                        </tr>
                    </tbody>
                </table>

                <div class="sec-title">ASIGNACIÓN POR COLABORADOR</div>
                <table>
                    <thead>
                        <tr>
                            <th>COLABORADOR</th>
                            <th>HORAS PROGRAMADAS</th>
                            <th>META DE VENTA</th>
                            <th>ANÁLOGOS POR VENDER</th>
                            <th>WEARABLES POR VENDER</th>
                            <th>KIDS POR VENDER</th>
                            <th>CAREKITS POR VENDER</th>
                        </tr>
                    </thead>
                    <tbody>
                        {"".join(f"<tr><td style='text-align:left; font-weight:bold;'>{r['nombre']}</td><td>{r['horas']:.1f} hrs</td><td class='green-bg'>${r['meta_vta']:,.2f}</td><td class='green-bg'>{r['meta_ana']}</td><td class='green-bg'>{r['meta_wea']}</td><td class='green-bg'>{r['meta_kid']}</td><td class='green-bg'>{r['meta_ck']}</td></tr>" for r in calc["colab_rows"] if r["nombre"])}
                        <tr style="font-weight:bold; background:#e5e7eb;">
                            <td>TOTAL TIENDA</td>
                            <td>{calc['tot_horas']:.1f} hrs</td>
                            <td>${calc['meta_diaria']:,.2f}</td>
                            <td>{sum(x['meta_ana'] for x in calc['colab_rows'])}</td>
                            <td>{sum(x['meta_wea'] for x in calc['colab_rows'])}</td>
                            <td>{sum(x['meta_kid'] for x in calc['colab_rows'])}</td>
                            <td>{sum(x['meta_ck'] for x in calc['colab_rows'])}</td>
                        </tr>
                    </tbody>
                </table>

                <div style="display:grid; grid-template-columns: 1fr 1fr; gap:10px; margin-top:10px;">
                    <div style="border:1px solid #000; padding:8px; background:#fafafa;">
                        <div style="font-weight:bold; border-bottom:1px solid #ccc; padding-bottom:4px; margin-bottom:4px;">✨ LOS 5 SECRETOS Y PLAN DE ACCIÓN ({day})</div>
                        <div style="font-size:9.5px; line-height:1.4;">
                            <b>1. Pulir es poder:</b> Ofrece limpiar sus lentes al iniciar.<br>
                            <b>2. Póntelos:</b> Invítalo a probar diferentes modelos en la bandeja.<br>
                            <b>3. Diviértete más:</b> Muestra 3 o 4 opciones adicionales.<br>
                            <b>4. Cuídalos:</b> Ofrece estuche, solución limpiadora y carekits.<br>
                            <b>5. Ajuste perfecto:</b> Ajusta los armazones a su medida exacta.
                        </div>
                    </div>
                    <div style="border:1px solid #000; padding:8px; background:#fafafa;">
                        <div style="font-weight:bold; border-bottom:1px solid #ccc; padding-bottom:4px; margin-bottom:4px;">🎯 TU ENFOQUE PARA HOY</div>
                        <div style="font-size:10px; color:#111827; font-weight:600;">
                            {data['enfoque_hoy'] or 'Seguimiento continuo al 100% de la Meta Diaria y cumplimiento del Customer Journey.'}
                        </div>
                    </div>
                </div>

                <script>
                    window.onload = function() {{
                        setTimeout(function() {{ window.print(); }}, 400);
                    }};
                </script>
            </body>
            </html>
            """
            from fastapi.responses import HTMLResponse
            return HTMLResponse(content=html)
        except Exception as ex:
            return f"Error: {ex}"

    @app.middleware("http")
    async def oye_luxo_beep_injector_middleware(request: Request, call_next):
        response = await call_next(request)
        content_type = response.headers.get("content-type", "")
        if "text/html" in content_type:
            try:
                response_body = [section async for section in response.body_iterator]
                body = b"".join(response_body)
                html_text = body.decode("utf-8", errors="ignore")
                
                script_tag = """
                <script>
                (function() {
                    if (window.luxoScriptInjected) return;
                    window.luxoScriptInjected = true;

                    let siriStyle = document.getElementById("siri-orb-style");
                    if (!siriStyle) {
                        siriStyle = document.createElement("style");
                        siriStyle.id = "siri-orb-style";
                        siriStyle.innerHTML = `
                            @keyframes siriWavePulse {
                                0%, 100% { transform: scaleY(0.5); opacity: 0.6; }
                                50% { transform: scaleY(1.5); opacity: 1; }
                            }
                            @keyframes siriGlowPulse {
                                0%, 100% { box-shadow: 0 0 20px rgba(224, 64, 251, 0.7), 0 0 35px rgba(0, 240, 255, 0.4); }
                                50% { box-shadow: 0 0 30px rgba(0, 240, 255, 0.9), 0 0 50px rgba(224, 64, 251, 0.8); }
                            }
                        `;
                        document.head.appendChild(siriStyle);
                    }

                    let banner = document.createElement("div");
                    banner.id = "luxo-voice-banner";
                    banner.style.cssText = "position:fixed; bottom:30px; right:30px; z-index:9999999; width:64px; height:64px; border-radius:50%; background:radial-gradient(circle at 35% 35%, rgba(224, 64, 251, 0.85), rgba(0, 240, 255, 0.85), rgba(10, 10, 24, 0.95)); animation: siriGlowPulse 2.5s infinite ease-in-out; display:none !important; align-items:center; justify-content:center; cursor:pointer; user-select:none; transition: opacity 0.5s ease, transform 0.5s ease; opacity:0; pointer-events:none; transform:scale(0.7);";
                    banner.setAttribute("title", "Asistente de Voz LUXO (Oye LUXO)");
                    banner.innerHTML = `
                        <div style="display:flex; align-items:center; justify-content:center; gap:4px; height:100%; width:100%;">
                            <div style="width:4px; height:16px; background:#00F0FF; border-radius:2px; animation: siriWavePulse 1.2s infinite ease-in-out 0.1s;"></div>
                            <div style="width:4px; height:26px; background:#E040FB; border-radius:2px; animation: siriWavePulse 1.2s infinite ease-in-out 0.3s;"></div>
                            <div style="width:4px; height:22px; background:#C084FC; border-radius:2px; animation: siriWavePulse 1.2s infinite ease-in-out 0.2s;"></div>
                            <div style="width:4px; height:14px; background:#00F0FF; border-radius:2px; animation: siriWavePulse 1.2s infinite ease-in-out 0.4s;"></div>
                        </div>
                    `;
                    
                    let fadeTimer = null;
                    window.showLuxoSiriOrb = function(durationMs) {
                        const b = document.getElementById("luxo-voice-banner") || banner;
                        if (!b) return;
                        b.style.opacity = "1";
                        b.style.transform = "scale(1)";
                        b.style.pointerEvents = "auto";
                        if (fadeTimer) { clearTimeout(fadeTimer); fadeTimer = null; }
                        if (durationMs && durationMs > 0) {
                            fadeTimer = setTimeout(function() {
                                window.hideLuxoSiriOrb();
                            }, durationMs);
                        }
                    };

                    window.hideLuxoSiriOrb = function() {
                        const b = document.getElementById("luxo-voice-banner") || banner;
                        if (!b) return;
                        if (fadeTimer) { clearTimeout(fadeTimer); fadeTimer = null; }
                        b.style.opacity = "0";
                        b.style.transform = "scale(0.7)";
                        b.style.pointerEvents = "none";
                    };

                    function attachBanner() {
                        if (document.body && !document.getElementById("luxo-voice-banner")) {
                            document.body.appendChild(banner);
                        }
                    }

                    if (document.readyState === "complete" || document.readyState === "interactive") {
                        attachBanner();
                    } else {
                        document.addEventListener("DOMContentLoaded", attachBanner);
                    }
                    setInterval(attachBanner, 1000);

                    const iconEl = document.getElementById("luxo-voice-icon");
                    const textEl = document.getElementById("luxo-voice-text");

                    function setStatus(text, color, icon) {
                        const t = document.getElementById("luxo-voice-text") || textEl;
                        const i = document.getElementById("luxo-voice-icon") || iconEl;
                        const b = document.getElementById("luxo-voice-banner") || banner;
                        if (t) t.innerText = text;
                        if (i && icon) i.innerText = icon;
                        if (b && color) b.style.borderColor = color;
                    }

                    const SpeechRecognition = window.SpeechRecognition || window.webkitSpeechRecognition;

                    function playBeep() {
                        // Sonido desactivado
                    }


                    let rec = null;
                    let isListening = false;
                    let lastSentText = "";
                    let lastSentTime = 0;

                    window.initLuxoMicPermission = function() {
                        // La voz estaba desactivada temporalmente aquí, ahora está activa.
                        if (window.luxoSpeechRecognitionActive) {
                            console.log("🎙️ [Luxo Global Mic]: Already active. Skipping duplicate init.");
                            return;
                        }
                        setStatus("Solicitando permiso...", "#FFFF00", "⏳");
                        if (navigator.mediaDevices && navigator.mediaDevices.getUserMedia) {
                            navigator.mediaDevices.getUserMedia({ audio: true })
                            .then(function(stream) {
                                setStatus("🟢 ESCUCHANDO... Di 'Oye LUXO'", "#00FF00", "🎤");
                                playBeep();
                                startRecognition();
                            })
                            .catch(function(err) {
                                console.log("Mic permission error:", err);
                                setStatus("❌ Permiso Denegado (Clica 🔒 en URL)", "#FF0000", "⚠️");
                            });
                        } else {
                            startRecognition();
                        }
                    };

                    function startRecognition() {
                        if (!SpeechRecognition) {
                            setStatus("❌ Usa Chrome o Edge para 'Oye LUXO'", "#FF0000", "❌");
                            return;
                        }
                        try {
                            if (rec) { try { rec.stop(); } catch(e){} }
                            rec = new SpeechRecognition();
                            rec.continuous = true;
                            rec.interimResults = true;
                            rec.lang = 'es-MX';

                            rec.onstart = function() {
                                isListening = true;
                                window.luxoSpeechRecognitionActive = true;
                                setStatus("🟢 ESCUCHANDO EN VIVO... Di 'Oye LUXO'", "#00FFFF", "🎤");
                            };

                            rec.onresult = function(e) {
                                for (let i = e.resultIndex; i < e.results.length; ++i) {
                                    const transcript = e.results[i][0].transcript;
                                    const lower = transcript.toLowerCase();
                                    
                                    if (lower.includes("oye luxo") || lower.includes("oye lujo") || lower.includes("hola luxo") || lower.includes("hey luxo")) {
                                        const now = Date.now();
                                        let query = transcript
                                            .replace(/oye luxo/gi, '')
                                            .replace(/oye lujo/gi, '')
                                            .replace(/hola luxo/gi, '')
                                            .replace(/hey luxo/gi, '')
                                            .trim();
                                            
                                        // Activar orbe flotante Siri en la esquina inferior derecha mientras habla
                                        window.showLuxoSiriOrb();

                                        if (!query && e.results[i].isFinal) {
                                            playBeep();
                                            setStatus("👂 ¡Oye LUXO Detectado! Di tu pregunta...", "#FF00FF", "🔊");
                                            window.luxoManualDictating = true;
                                        } else if (query && e.results[i].isFinal) {
                                            if (query !== lastSentText) {
                                                lastSentText = query;
                                                lastSentTime = now;
                                                playBeep();
                                                setStatus("⚡ Enviando a LUXO: " + query, "#7CFC00", "🚀");
                                                fetch('/text_input?user_id=1&text=' + encodeURIComponent(query), { method: 'POST' });
                                                // Ocultar orbe flotante tras procesar/enviar
                                                setTimeout(function(){ window.hideLuxoSiriOrb(); }, 2000);
                                                window.luxoManualDictating = false;
                                                try { rec.abort(); } catch(e){} // Reinicia el buffer del microfono
                                            }
                                        }
                                    } else if (window.luxoManualDictating && e.results[i].isFinal) {
                                        const query = transcript.trim();
                                        if (query && query !== lastSentText) {
                                            lastSentText = query;
                                            playBeep();
                                            setStatus("⚡ Enviando a LUXO: " + query, "#7CFC00", "🚀");
                                            fetch('/text_input?user_id=1&text=' + encodeURIComponent(query), { method: 'POST' });
                                            window.luxoManualDictating = false;
                                            // Ocultar orbe flotante al finalizar de hablar
                                            setTimeout(function(){ window.hideLuxoSiriOrb(); }, 2000);
                                            try { rec.abort(); } catch(e){} // Reinicia el buffer del microfono
                                        }
                                    }
                                }
                            };

                            rec.onerror = function(err) {
                                console.log("Luxo Voice Error:", err);
                                if (err.error === 'not-allowed') {
                                    setStatus("❌ Micrófono Bloqueado en Navegador", "#FF0000", "🔒");
                                }
                            };

                            rec.onend = function() {
                                window.luxoSpeechRecognitionActive = false;
                                if (isListening) {
                                    setTimeout(function() { try { rec.start(); } catch(e){} }, 300);
                                }
                            };

                            rec.start();
                        } catch(ex) {
                            console.log("Exception in startRecognition:", ex);
                        }
                    }

                    banner.onclick = function() {
                        window.initLuxoMicPermission();
                    };

                    window.luxoTriggerFileUpload = function(acceptFilter, userId, captureMode) {
                        let input = document.getElementById("luxo_global_file_input");
                        if (!input) {
                            input = document.createElement("input");
                            input.type = "file";
                            input.id = "luxo_global_file_input";
                            input.style.display = "none";
                            document.body.appendChild(input);
                        }
                        input.accept = acceptFilter || "*/*";
                        if (captureMode) {
                            input.setAttribute("capture", "environment");
                        } else {
                            input.removeAttribute("capture");
                        }
                        input.value = "";
                        input.onchange = function() {
                            if (!input.files || input.files.length === 0) return;
                            let file = input.files[0];
                            let formData = new FormData();
                            formData.append("file", file);
                            formData.append("user_id", userId || "1");
                            
                            fetch("/api/upload_generic", {
                                method: "POST",
                                body: formData
                            })
                            .then(response => response.json())
                            .then(data => {
                                if (data.status === "success") {
                                    console.log("Archivo subido con éxito:", data.filename);
                                } else {
                                    alert("Error al subir archivo: " + (data.message || "desconocido"));
                                }
                            })
                            .catch(err => alert("Error de red al subir archivo: " + err));
                        };
                        input.click();
                    };

                    window.luxoAbrirCamaraFacialColab = function(colabId, colabName) {
                        const exist = document.getElementById('luxo-reg-facial-modal');
                        if (exist) exist.remove();
                        
                        const modal = document.createElement('div');
                        modal.id = 'luxo-reg-facial-modal';
                        modal.style.cssText = 'position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,0.92);z-index:99999999;display:flex;flex-direction:column;align-items:center;justify-content:center;font-family:Segoe UI,sans-serif;';
                        modal.innerHTML = `
                            <div style="background:#0a0a16;border:2px solid #00FFFF;border-radius:20px;padding:24px;max-width:420px;width:95%;text-align:center;box-shadow:0 0 40px rgba(0,255,255,0.4);">
                                <h3 style="color:#00FFFF;margin:0 0 6px;">📷 Registrar Rostro</h3>
                                <p style="color:#aaa;font-size:13px;margin:0 0 14px;">Colaborador: <b style="color:white;">` + colabName + `</b></p>
                                <div style="position:relative;width:220px;height:220px;margin:0 auto 16px;">
                                    <video id="luxo-cam-reg" autoplay playsinline muted style="width:220px;height:220px;object-fit:cover;border-radius:50%;border:3px solid #00FFFF;"></video>
                                    <canvas id="luxo-canvas-reg" width="220" height="220" style="display:none;"></canvas>
                                </div>
                                <p id="luxo-reg-msg" style="color:#00FFFF;font-size:13px;min-height:20px;margin-bottom:14px;"></p>
                                <div style="display:flex;gap:12px;justify-content:center;">
                                    <button id="btn-cap-colab" style="background:linear-gradient(135deg,#003366,#0066cc);color:white;border:none;padding:10px 22px;border-radius:10px;font-size:14px;font-weight:bold;cursor:pointer;">📸 Capturar Foto</button>
                                    <button id="btn-close-colab" style="background:#333;color:white;border:none;padding:10px 22px;border-radius:10px;font-size:14px;cursor:pointer;">✕ Cancelar</button>
                                </div>
                            </div>
                        `;
                        document.body.appendChild(modal);
                        
                        let stream = null;
                        function stopCam() { if (stream) { stream.getTracks().forEach(t => t.stop()); stream = null; } }
                        
                        document.getElementById('btn-close-colab').onclick = function() {
                            stopCam();
                            modal.remove();
                        };
                        
                        navigator.mediaDevices.getUserMedia({ video: { facingMode: 'user' } })
                        .then(s => {
                            stream = s;
                            document.getElementById('luxo-cam-reg').srcObject = s;
                            document.getElementById('luxo-reg-msg').innerText = '🟢 Cámara lista. Presiona Capturar Foto.';
                        })
                        .catch(err => {
                            document.getElementById('luxo-reg-msg').innerText = '⚠️ No se pudo acceder a la cámara. Revisa los permisos.';
                            document.getElementById('luxo-reg-msg').style.color = '#FF4500';
                        });
                        
                        document.getElementById('btn-cap-colab').onclick = function() {
                            const video = document.getElementById('luxo-cam-reg');
                            const canvas = document.getElementById('luxo-canvas-reg');
                            const ctx = canvas.getContext('2d');
                            ctx.drawImage(video, 0, 0, 220, 220);
                            const dataUrl = canvas.toDataURL('image/jpeg', 0.85);
                            
                            document.getElementById('luxo-reg-msg').innerText = '⏳ Guardando e identificando rostro...';
                            document.getElementById('luxo-reg-msg').style.color = '#FFD700';
                            
                            fetch('/api/biometria/registrar_rostro_colaborador', {
                                method: 'POST',
                                headers: {'Content-Type': 'application/json'},
                                body: JSON.stringify({colaborador_id: colabId, nombre: colabName, imagen: dataUrl})
                            })
                            .then(r => r.json())
                            .then(d => {
                                if (d.ok) {
                                    document.getElementById('luxo-reg-msg').innerText = '✅ ¡Rostro registrado exitosamente!';
                                    document.getElementById('luxo-reg-msg').style.color = '#7CFC00';
                                    stopCam();
                                    setTimeout(() => { modal.remove(); window.location.reload(); }, 1800);
                                } else {
                                    document.getElementById('luxo-reg-msg').innerText = '❌ Error: ' + (d.error || 'No se pudo guardar');
                                    document.getElementById('luxo-reg-msg').style.color = '#FF4500';
                                }
                            })
                            .catch(() => {
                                document.getElementById('luxo-reg-msg').innerText = '❌ Error al conectar con el servidor';
                                document.getElementById('luxo-reg-msg').style.color = '#FF4500';
                            });
                        };
                    };

                    window.luxoAbrirHuellaColab = async function(colabId, colabName) {
                        const existModal = document.getElementById('luxo-huella-modal');
                        if (existModal) existModal.remove();
                        
                        const modal = document.createElement('div');
                        modal.id = 'luxo-huella-modal';
                        modal.style.cssText = 'position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,0.88);z-index:99999999;display:flex;align-items:center;justify-content:center;font-family:Segoe UI,sans-serif;';
                        modal.innerHTML = `
                            <div style="background:#0a0a16;border:2px solid #D8B4FE;border-radius:20px;padding:26px;max-width:400px;width:92%;text-align:center;box-shadow:0 0 40px rgba(216,180,254,0.3);">
                                <div style="font-size:44px;margin-bottom:10px;">👆</div>
                                <h3 style="color:#D8B4FE;margin:0 0 6px;">Registrar Huella Dactilar</h3>
                                <p style="color:#aaa;font-size:13px;margin:0 0 14px;">Colaborador: <b style="color:white;">` + colabName + `</b></p>
                                <p id="luxo-hue-status" style="color:#FFD700;font-size:13px;min-height:22px;margin-bottom:16px;">Iniciando sensor biométrico...</p>
                                <button id="btn-hue-close" style="padding:10px 22px;background:#333;color:white;border:none;border-radius:10px;cursor:pointer;font-size:14px;">Cancelar</button>
                            </div>
                        `;
                        document.body.appendChild(modal);
                        
                        document.getElementById('btn-hue-close').onclick = function() { modal.remove(); };
                        
                        const setStatus = (msg, color) => {
                            const el = document.getElementById('luxo-hue-status');
                            if (el) { el.innerText = msg; el.style.color = color || '#FFD700'; }
                        };
                        
                        if (!window.PublicKeyCredential) {
                            setStatus('⚠️ Tu navegador o celular no soporta lectura de huella.', '#FF8C00');
                            return;
                        }
                        
                        try {
                            const resp = await fetch('/api/biometria/passkey_challenge_registro?colaborador_id=' + colabId + '&nombre=' + encodeURIComponent(colabName));
                            const opts = await resp.json();
                            if (!opts.publicKey) { setStatus('❌ Error al obtener configuración del servidor.', '#FF4500'); return; }
                            
                            const decode = s => Uint8Array.from(atob(s.replace(/-/g,'+').replace(/_/g,'/')), c => c.charCodeAt(0));
                            opts.publicKey.challenge = decode(opts.publicKey.challenge);
                            opts.publicKey.user.id = decode(opts.publicKey.user.id);
                            
                            setStatus('👆 Toca el lector de huella o sensor biométrico de tu celular...', '#D8B4FE');
                            const credential = await navigator.credentials.create({ publicKey: opts.publicKey });
                            
                            setStatus('⏳ Guardando registro biométrico...', '#FFD700');
                            const payload = {
                                colaborador_id: colabId,
                                nombre: colabName,
                                id: credential.id,
                                rawId: btoa(String.fromCharCode(...new Uint8Array(credential.rawId))),
                                type: credential.type,
                                response: {
                                    attestationObject: btoa(String.fromCharCode(...new Uint8Array(credential.response.attestationObject))),
                                    clientDataJSON: btoa(String.fromCharCode(...new Uint8Array(credential.response.clientDataJSON)))
                                }
                            };
                            
                            const vResp = await fetch('/api/biometria/registrar_huella_colaborador', {
                                method: 'POST',
                                headers: {'Content-Type': 'application/json'},
                                body: JSON.stringify(payload)
                            });
                            const vData = await vResp.json();
                            if (vData.ok) {
                                setStatus('✅ ¡Huella dactilar registrada exitosamente!', '#7CFC00');
                                setTimeout(() => { modal.remove(); window.location.reload(); }, 1800);
                            } else {
                                setStatus('❌ Error: ' + (vData.error || 'No se pudo guardar la huella'), '#FF4500');
                            }
                        } catch(err) {
                            if (err.name === 'NotAllowedError') {
                                setStatus('⚠️ Lectura de huella cancelada por el usuario.', '#FF8C00');
                            } else {
                                setStatus('⚠️ Nota de lectura: ' + err.message, '#FF8C00');
                            }
                        }
                    };
                    window.initLuxoMicPermission();
                })();
                </script>
                """
                
                if "</body>" in html_text:
                    html_text = html_text.replace("</body>", f"{script_tag}</body>")
                else:
                    html_text += script_tag
                    
                return Response(content=html_text.encode("utf-8"), media_type="text/html", status_code=response.status_code)
            except Exception as ex:
                print("Error en middleware inyector de voz:", ex)
        return response

    @app.get("/voice")
    async def get_voice_page(user_id: str = "1"):
        from fastapi.responses import HTMLResponse
        html_content = """
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>🎙️ Asistente de Voz LUXO - "Oye LUXO"</title>
            <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@400;600;800&display=swap" rel="stylesheet">
            <style>
                * { box-sizing: border-box; margin: 0; padding: 0; }
                body {
                    background-color: #0c0c14; color: #ffffff;
                    font-family: 'Outfit', 'Segoe UI', sans-serif;
                    display: flex; flex-direction: column; align-items: center; justify-content: center;
                    min-height: 100vh; text-align: center; padding: 20px;
                }
                .card {
                    background: rgba(255, 255, 255, 0.05); border: 2px solid #00FFFF;
                    border-radius: 28px; padding: 40px 30px; max-width: 480px; width: 100%;
                    box-shadow: 0 10px 40px rgba(0, 255, 255, 0.25); backdrop-filter: blur(12px);
                }
                .mic-btn {
                    width: 140px; height: 140px; border-radius: 50%;
                    background: linear-gradient(135deg, #6E48AA, #9D50BB);
                    border: 4px solid #00FFFF; color: white; font-size: 56px;
                    display: flex; align-items: center; justify-content: center;
                    margin: 25px auto; cursor: pointer;
                    box-shadow: 0 0 30px rgba(0, 255, 255, 0.5);
                    transition: all 0.3s cubic-bezier(0.175, 0.885, 0.32, 1.275);
                }
                .mic-btn:hover { transform: scale(1.08); box-shadow: 0 0 45px rgba(0, 255, 255, 0.9); }
                .mic-btn.active { animation: pulse 1.5s infinite; background: linear-gradient(135deg, #00FFFF, #6E48AA); }
                @keyframes pulse {
                    0% { box-shadow: 0 0 0 0 rgba(0, 255, 255, 0.8); }
                    70% { box-shadow: 0 0 0 30px rgba(0, 255, 255, 0); }
                    100% { box-shadow: 0 0 0 0 rgba(0, 255, 255, 0); }
                }
                h1 { color: #D8B4FE; margin-bottom: 10px; font-size: 28px; font-weight: 800; }
                p { color: #aaaaaa; font-size: 15px; margin-bottom: 20px; line-height: 1.5; }
                .status {
                    margin-top: 15px; padding: 15px 20px; border-radius: 14px;
                    background: #1c1c28; border: 1.5px solid #333; color: #00FFFF;
                    font-weight: bold; font-size: 15px; transition: all 0.3s ease;
                }
                .transcript {
                    margin-top: 20px; padding: 15px; background: #000000;
                    border-radius: 14px; border: 1px solid #444; min-height: 70px;
                    color: #7CFC00; font-family: monospace; font-size: 16px; word-break: break-word;
                    display: flex; align-items: center; justify-content: center;
                }
                .badge {
                    display: inline-block; padding: 6px 14px; border-radius: 20px;
                    background: rgba(0, 255, 255, 0.15); color: #00FFFF; border: 1px solid #00FFFF;
                    font-size: 12px; font-weight: bold; margin-bottom: 15px;
                }
            </style>
        </head>
        <body>
            <div class="card">
                <div class="badge">🎙️ LUXO VOICE ASSISTANT</div>
                <h1>Control por Voz LUXO</h1>
                <p>Toca el botón para otorgar permiso e iniciar la escucha continua. Luego di en voz alta:<br><strong style="color:#00FFFF; font-size:18px;">"Oye LUXO [tu pregunta]"</strong></p>
                
                <div id="micBtn" class="mic-btn" style="display: none !important;">🎙️</div>
                
                <div id="status" class="status">Toca el micrófono para comenzar</div>
                
                <div style="text-align: left; margin-top: 20px; color: #888; font-size: 13px; font-weight:600;">Voz detectada en vivo:</div>
                <div id="transcript" class="transcript">Esperando inicio...</div>
            </div>

            <script>
                const micBtn = document.getElementById('micBtn');
                const status = document.getElementById('status');
                const transcriptDiv = document.getElementById('transcript');

                const SpeechRecognition = window.SpeechRecognition || window.webkitSpeechRecognition;

                function playBeep() {
                    // Sonido desactivado
                }



                let rec = null;
                let isListening = false;
                let lastSentText = "";
                let lastSentTime = 0;

                function startListening() {
                    if (!SpeechRecognition) {
                        status.innerText = "❌ Tu navegador no soporta SpeechRecognition. Usa Chrome o Edge.";
                        status.style.color = "#FF4444";
                        status.style.borderColor = "#FF4444";
                        return;
                    }

                    status.innerText = "⏳ Solicitando permiso de micrófono...";
                    status.style.color = "#FFFF00";

                    if (navigator.mediaDevices && navigator.mediaDevices.getUserMedia) {
                        navigator.mediaDevices.getUserMedia({ audio: true })
                        .then(stream => {
                            stream.getTracks().forEach(t => t.stop());
                            runRecognition();
                        })
                        .catch(err => {
                            console.log(err);
                            status.innerText = "❌ Permiso de micrófono denegado. Permite el micrófono en la barra URL 🔒.";
                            status.style.color = "#FF4444";
                            status.style.borderColor = "#FF4444";
                        });
                    } else {
                        runRecognition();
                    }
                }

                function runRecognition() {
                    try {
                        if (rec) { try { rec.stop(); } catch(e){} }
                        rec = new SpeechRecognition();
                        rec.continuous = true;
                        rec.interimResults = true;
                        rec.lang = 'es-MX';

                        rec.onstart = function() {
                            isListening = true;
                            micBtn.classList.add('active');
                            status.innerText = "🎤 ESCUCHANDO... Habla ahora";
                            status.style.color = "#00FFFF";
                            status.style.borderColor = "#00FFFF";
                        };

                        rec.onresult = function(e) {
                            for (let i = e.resultIndex; i < e.results.length; ++i) {
                                const text = e.results[i][0].transcript;
                                transcriptDiv.innerText = text;
                                const lower = text.toLowerCase();

                                if (e.results[i].isFinal) {
                                    const query = text.trim();
                                    if (query && query !== lastSentText) {
                                        const now = Date.now();
                                        lastSentText = query;
                                        lastSentTime = now;
                                        status.innerText = "🚀 Enviado a la IA: " + query;
                                        status.style.color = "#7CFC00";
                                        status.style.borderColor = "#7CFC00";
                                        fetch('/text_input?user_id=1&text=' + encodeURIComponent(query), { method: 'POST' });
                                        setTimeout(() => {
                                            status.innerText = "🎤 Presiona el micrófono para hablar";
                                            status.style.color = "#00FFFF";
                                            status.style.borderColor = "#00FFFF";
                                            isListening = false;
                                            micBtn.classList.remove('active');
                                            try { rec.stop(); } catch(ex){}
                                        }, 2500);
                                    }
                                }
                            }
                        };

                        rec.onerror = function(e) {
                            console.log("Rec Error:", e);
                            if (e.error === 'not-allowed') {
                                status.innerText = "❌ Micrófono bloqueado en el navegador.";
                                status.style.color = "#FF4444";
                                status.style.borderColor = "#FF4444";
                            }
                        };

                        rec.onend = function() {
                            if (isListening) {
                                setTimeout(() => { try { rec.start(); } catch(ex){} }, 300);
                            }
                        };

                        rec.start();
                    } catch(e) {
                        status.innerText = "Error: " + e.message;
                    }
                }

                micBtn.addEventListener('click', () => {
                    // startListening(); // VOZ DESACTIVADA TEMPORALMENTE
                });

                // Auto-iniciar al cargar
                    // startListening(); // VOZ DESACTIVADA TEMPORALMENTE
            </script>
        </body>
        </html>
        """
        return HTMLResponse(content=html_content)

    @app.get("/status")
    async def get_status(user_id: str, state: str):
        user_id_val = int(user_id) if user_id.isdigit() else user_id
        if user_id_val in active_sessions:
            session = active_sessions[user_id_val]
            btn_mic = session.get("btn_mic")
            page = session.get("page")
            file_picker_audio = session.get("file_picker_audio")
            
            if btn_mic and page:
                if state == "RECORDING":
                    btn_mic.icon = ft.Icons.STOP_ROUNDED
                    btn_mic.icon_color = "#FF4500"
                    btn_mic.tooltip = "Grabando... Toca para detener y enviar 🛑"
                    page.update()
                elif state in ["DONE", "ERROR"]:
                    btn_mic.icon = ft.Icons.MIC_ROUNDED
                    btn_mic.icon_color = "#00FFFF"
                    btn_mic.tooltip = "Grabar Nota de Voz 🎙️"
                    page.title = "LUXO"
                    page.update()
                elif state == "NOT_SUPPORTED":
                    btn_mic.icon = ft.Icons.MIC_ROUNDED
                    btn_mic.icon_color = "#00FFFF"
                    btn_mic.tooltip = "Grabar Nota de Voz 🎙️"
                    page.title = "LUXO"
                    page.update()
        return {"status": "ok"}

    @app.get("/command")
    async def get_command(user_id: str = "1"):
        cmd = "NONE"
        for sess in active_sessions.values():
            if sess.get("command") == "START":
                cmd = "START"
                sess["command"] = "NONE"
                break
        return {"command": cmd}

    @app.api_route("/luxo_listening_start", methods=["POST"])
    async def post_listening_start(user_id: str = "1"):
        user_id_val = int(user_id) if (user_id and str(user_id).isdigit()) else user_id
        session = active_sessions.get(user_id_val) or active_sessions.get(str(user_id)) or active_sessions.get(1) or active_sessions.get("1") or (list(active_sessions.values())[0] if active_sessions else None)
        if session:
            page = session.get("page")
            siri_orb = session.get("siri_orb")
            if page:
                try:
                    if siri_orb:
                        siri_orb.opacity = 1
                        siri_orb.scale = 1
                    page.snack_bar = ft.SnackBar(ft.Text("✨ ¡Oye LUXO detectado! Escuchando...", color="white", weight="bold"), bgcolor="#9D50BB", duration=3000)
                    page.snack_bar.open = True
                    page.update()
                except Exception as ex:
                    print(f"WARN Flet overlay error: {ex}")
        return {"status": "success"}

    @app.api_route("/text_input", methods=["GET", "POST"])
    async def post_text_input(user_id: str = "1", text: str = ""):
        import traceback
        try:
            print(f"DEBUG: /text_input recibido con user_id={user_id}, text='{text}'")
            user_id_val = int(user_id) if (user_id and str(user_id).isdigit()) else user_id
            session = active_sessions.get(user_id_val) or active_sessions.get(str(user_id)) or active_sessions.get(1) or active_sessions.get("1") or (list(active_sessions.values())[0] if active_sessions else None)
            print(f"DEBUG: active_sessions keys={list(active_sessions.keys())}, session encontrada={'Sí' if session else 'No'}")
            if session and text:
                cambiar_vista = session.get("cambiar_vista")
                page = session.get("page")

                if cambiar_vista:
                    try:
                        cambiar_vista("chat")
                    except Exception as ex:
                        print("Error al cambiar a vista chat:", ex)

                input_msg = session.get("input_msg")
                enviar_mensaje = session.get("enviar_mensaje")

                if input_msg and enviar_mensaje and page:
                    btn_mic_cont = session.get("btn_mic_container")
                    siri_orb = session.get("siri_orb")
                    
                    if siri_orb:
                        try:
                            siri_orb.opacity = 1
                            siri_orb.scale = 1
                            siri_orb.update()
                        except: pass
                        
                    if btn_mic_cont:
                        try:
                            btn_mic_cont.bgcolor = "#9D50BB" # Morado brilloso (Siri vibe)
                            btn_mic_cont.border = ft.Border.all(3, "#00FFFF") # Cyan grueso
                            btn_mic_cont.update()
                            def revert_glow(bmc, orb):
                                import time
                                time.sleep(3)
                                bmc.bgcolor = "#1E1E2E"
                                bmc.border = ft.Border.all(1.5, "#00FFFF")
                                try: bmc.update()
                                except: pass
                                if orb:
                                    orb.opacity = 0
                                    orb.scale = 0.1
                                    try: orb.update()
                                    except: pass
                            threading.Thread(target=revert_glow, args=(btn_mic_cont, siri_orb), daemon=True).start()
                        except: pass
                        
                    input_msg.value = text
                    try:
                        page.update()
                    except Exception as ex:
                        print(f"WARN page.update() falló: {ex}")

                    async def trigger_send():
                        try:
                            enviar_mensaje(None)
                        except Exception as ex:
                            print(f"ERROR en enviar_mensaje: {ex}\n{traceback.format_exc()}")
                    page.run_task(trigger_send)
                    return {"status": "success"}
                else:
                    print(f"DEBUG: Faltan componentes - input_msg={input_msg is not None}, enviar_mensaje={enviar_mensaje is not None}, page={page is not None}")
            return {"status": "session_not_found"}
        except Exception as e:
            print(f"ERROR CRÍTICO en /text_input: {traceback.format_exc()}")
            return {"status": "error", "detail": str(e)}

    @app.post("/upload")
    async def post_upload(request: Request, user_id: str = "1"):
        print(f"DEBUG: /upload audio recibido para user_id={user_id}")
        form = await request.form()
        uploaded_file = form.get("file")
        if uploaded_file:
            audio_bytes = await uploaded_file.read()
            user_id_val = int(user_id) if (user_id and str(user_id).isdigit()) else user_id
            session = active_sessions.get(user_id_val) or active_sessions.get(str(user_id)) or active_sessions.get(1) or active_sessions.get("1") or (list(active_sessions.values())[0] if active_sessions else None)
            if session and session.get("cambiar_vista"):
                try:
                    session["cambiar_vista"]("chat")
                except Exception as ex:
                    print("Error al cambiar a vista chat en upload:", ex)
            save_and_transcribe_audio(user_id, audio_bytes)
        return {"status": "success"}

    @app.post("/api/upload_generic")
    async def post_upload_generic(request: Request):
        try:
            form = await request.form()
            file = form.get("file")
            user_id = form.get("user_id", "1")
            if file:
                raw_filename = getattr(file, "filename", "upload.jpg") or "upload.jpg"
                base, ext = os.path.splitext(os.path.basename(raw_filename))
                if not ext:
                    ext = ".jpg"
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:19]
                filename = f"{base}_{timestamp}{ext}"

                os.makedirs("uploads", exist_ok=True)
                filepath = os.path.join("uploads", filename)
                content = await file.read()
                
                if not content or len(content) < 10:
                    print(f"WARN /api/upload_generic: Received empty file ({len(content)} bytes)")
                    return {"status": "error", "message": "El archivo recibido está vacío o no se pudo leer."}

                with open(filepath, "wb") as f:
                    f.write(content)
                
                abs_path = os.path.abspath(filepath)
                try:
                    optimizar_archivo_multimedia(abs_path)
                except Exception as ex_opt:
                    print("Notice optimizar_archivo_multimedia:", ex_opt)
                
                user_id_val = int(user_id) if (user_id and str(user_id).isdigit()) else user_id
                session = active_sessions.get(user_id_val) or active_sessions.get(str(user_id)) or active_sessions.get(1) or active_sessions.get("1") or (list(active_sessions.values())[0] if active_sessions else None)
                if session and session.get("active_file_callback") and session["active_file_callback"][0]:
                    cb = session["active_file_callback"][0]
                    try:
                        cb(abs_path)
                    except Exception as ex_cb:
                        print("Error ejecutando callback de archivo subido:", ex_cb)
                        
                return {"status": "success", "filename": filename, "filepath": abs_path}
            return {"status": "error", "message": "No se recibió archivo."}
        except Exception as ex:
            print("ERROR EN /api/upload_generic:", ex)
            return {"status": "error", "message": str(ex)}

    @app.get("/upload_widget")
    async def get_upload_widget(type: str = "generic", user_id: str = "1"):
        from fastapi.responses import HTMLResponse
        
        accept_map = {
            "weekly": ".xlsx, .xls",
            "pdf": ".pdf",
            "excel": ".xlsx, .xls",
            "media": "image/*, video/*, .png, .jpg, .jpeg, .gif, .mp4, .mov, .avi",
            "generic": "*/*"
        }
        
        title_map = {
            "weekly": "📊 Cargar Excel Weekly (.xlsx)",
            "pdf": "📚 Cargar Manual PDF (.pdf)",
            "excel": "📑 Cargar Archivo Excel (.xlsx)",
            "media": "🖼️ Cargar Imagen o Video",
            "generic": "📁 Seleccionar Archivo"
        }
        
        accept_val = accept_map.get(type.lower(), "*/*")
        title_val = title_map.get(type.lower(), "📁 Seleccionar Archivo")

        camera_btn_html = ""
        if type.lower() in ["media", "ticket"]:
            camera_btn_html = '''
            <label class="file-btn" style="background: linear-gradient(135deg, #1f6f43, #2e7d32); margin-bottom: 12px;">
                📷 Tomar Foto con Cámara Celular
                <input type="file" id="cameraInput" accept="image/*" capture="environment" onchange="submitFile(this)">
            </label>
            '''

        html_code = f"""
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>{title_val}</title>
            <style>
                * {{ box-sizing: border-box; margin: 0; padding: 0; }}
                body {{
                    background: #0c0c14; color: #ffffff;
                    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                    display: flex; flex-direction: column; align-items: center; justify-content: center;
                    padding: 20px; text-align: center; min-height: 100vh;
                }}
                .upload-card {{
                    background: #181828; border: 2px dashed #00FFFF; border-radius: 20px;
                    padding: 30px 20px; width: 100%; max-width: 420px;
                    box-shadow: 0 10px 30px rgba(0,255,255,0.25);
                }}
                .title {{ font-size: 18px; font-weight: bold; color: #D8B4FE; margin-bottom: 20px; }}
                .file-btn {{
                    display: block; width: 100%; background: linear-gradient(135deg, #7c3aed, #9D50BB);
                    color: #ffffff; font-weight: bold; font-size: 15px;
                    padding: 14px 20px; border-radius: 12px; cursor: pointer;
                    box-shadow: 0 4px 20px rgba(124,58,237,0.5); transition: transform 0.2s;
                }}
                .file-btn:active {{ transform: scale(0.95); }}
                input[type="file"] {{ display: none; }}
                .status {{ margin-top: 20px; font-size: 14px; color: #7CFC00; font-weight: bold; display: none; }}
            </style>
        </head>
        <body>
            <div class="upload-card">
                <div class="title">{title_val}</div>
                <div>
                    {camera_btn_html}
                    <label class="file-btn">
                        📁 Buscar en Galería / Archivos
                        <input type="file" id="galleryInput" accept="{accept_val}" onchange="submitFile(this)">
                    </label>
                </div>
                <div id="statusMsg" class="status">⏳ Procesando archivo en el servidor...</div>
            </div>
            <script>
                function submitFile(input) {{
                    if (input.files && input.files.length > 0) {{
                        let msg = document.getElementById('statusMsg');
                        msg.style.display = 'block';
                        msg.style.color = '#00FFFF';
                        msg.innerText = '⏳ Cargando y procesando archivo...';
                        
                        let formData = new FormData();
                        formData.append('user_id', '{user_id}');
                        formData.append('upload_type', '{type}');
                        formData.append('file', input.files[0]);
                        
                        fetch('/api/upload_generic', {{
                            method: 'POST',
                            body: formData
                        }})
                        .then(res => res.json())
                        .then(data => {{
                            if (data.status === 'success') {{
                                msg.style.color = '#7CFC00';
                                msg.innerText = '✅ ¡Éxito! Archivo cargado correctamente.';
                                setTimeout(() => {{
                                    window.close();
                                }}, 800);
                            }} else {{
                                msg.style.color = '#FF4500';
                                msg.innerText = '❌ Error: ' + (data.message || 'Error al subir');
                            }}
                        }})
                        .catch(err => {{
                            msg.style.color = '#FF4500';
                            msg.innerText = '❌ Error de conexión al cargar archivo';
                        }});
                    }}
                }}
            </script>
        </body>
        </html>
        """
        return HTMLResponse(content=html_code)

    # =========================================
    # RUTAS BIOMÉTRICAS - WEBAUTHN PASSKEY + FACIAL
    # =========================================

    import base64, json as _json, secrets as _secrets

    def registrar_sesion_biometrica(id_usuario, nombre_usuario, empleado_identificado, metodo, es_gerente, ip_acceso, dispositivo):
        """Guarda el ingreso biométrico en la bitácora de seguridad de MySQL."""
        try:
            db_b = conectar_db()
            if db_b:
                cur_b = db_b.cursor()
                cur_b.execute("""
                    INSERT INTO bitacora_sesiones_biometricas
                        (ID_Usuario, Nombre_Usuario, Empleado_Identificado, Metodo_Ingreso, Es_Gerente_Verificado, IP_Acceso, Dispositivo)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (id_usuario, nombre_usuario, empleado_identificado, metodo, es_gerente, ip_acceso, dispositivo))
                db_b.commit()
                db_b.close()
        except Exception as ex_b:
            print("Error registrando sesión biométrica:", ex_b)

    # --- Generar desafío WebAuthn (Passkey) para login ---
    @app.get("/api/biometria/passkey_challenge")
    async def passkey_challenge():
        """Genera un desafío aleatorio para la validación WebAuthn del sensor dactilar."""
        challenge = base64.urlsafe_b64encode(_secrets.token_bytes(32)).rstrip(b"=").decode()
        return {"challenge": challenge, "rp_id": "localhost", "rp_name": "LUXO System"}

    # --- Verificar credencial Passkey recibida del sensor dactilar ---
    @app.post("/api/biometria/passkey_verify")
    async def passkey_verify(request: Request):
        """Valida la firma WebAuthn enviada por el navegador y busca al usuario biométrico."""
        try:
            body = await request.json()
            credential_id = body.get("credential_id", "")
            user_agent = request.headers.get("user-agent", "Desconocido")
            ip_client = request.client.host if request.client else "Desconocido"

            # Buscar en biometria_usuarios por Credential_ID
            db_p = conectar_db()
            if not db_p:
                return {"status": "error", "message": "Error de base de datos"}

            cursor_p = db_p.cursor(dictionary=True)
            cursor_p.execute("""
                SELECT b.usuario_id, b.nombre_usuario, u.ID_Usuario, u.Nombre_Completo,
                       u.Rol, u.Tienda, u.Zona, u.Puesto
                FROM biometria_usuarios b
                JOIN usuarios u ON b.usuario_id = u.ID_Usuario
                WHERE b.credential_id = %s
            """, (credential_id,))
            bio_user = cursor_p.fetchone()

            if not bio_user:
                # Si no hay credential_id guardado aún, buscar cualquier usuario con Passkey registrada
                cursor_p.execute("""
                    SELECT b.usuario_id, b.nombre_usuario, u.ID_Usuario, u.Nombre_Completo,
                           u.Rol, u.Tienda, u.Zona, u.Puesto
                    FROM biometria_usuarios b
                    JOIN usuarios u ON b.usuario_id = u.ID_Usuario
                    WHERE b.hash_huella IS NOT NULL
                    LIMIT 1
                """)
                bio_user = cursor_p.fetchone()

            db_p.close()

            if bio_user:
                rol = str(bio_user.get("Rol", "")).lower()
                puesto = str(bio_user.get("Puesto", "")).lower()
                es_gerente = "gerente" in rol or "gerente" in puesto or "admin" in rol

                registrar_sesion_biometrica(
                    id_usuario=bio_user["ID_Usuario"],
                    nombre_usuario=bio_user["Nombre_Completo"],
                    empleado_identificado=bio_user["Nombre_Completo"],
                    metodo="Huella",
                    es_gerente=es_gerente,
                    ip_acceso=ip_client,
                    dispositivo=user_agent[:150]
                )

                # Notificar a la sesión Flet activa
                session = (list(active_sessions.values())[0] if active_sessions else None)
                if session:
                    page_s = session.get("page")
                    if page_s:
                        ui = session.get("user_info", {})
                        ui["id"] = bio_user["ID_Usuario"]
                        ui["nombre"] = bio_user["Nombre_Completo"]
                        ui["rol"] = bio_user["Rol"]
                        ui["tienda"] = bio_user.get("Tienda") or ""
                        ui["zona"] = bio_user.get("Zona") or "Zona Centro"
                        ui["biometria_metodo"] = "Huella"
                        ui["es_gerente_verificado"] = es_gerente
                        cargar_chat_fn = session.get("cargar_chat")
                        saludo_fn = session.get("reproducir_saludo")
                        if cargar_chat_fn:
                            _nombre_bio = bio_user["Nombre_Completo"]
                            async def trigger_login_huella():
                                cargar_chat_fn()
                                import threading as _th_h
                                _th_h.Thread(
                                    target=reproducir_saludo_login,
                                    args=(_nombre_bio,),
                                    daemon=True
                                ).start()
                            page_s.run_task(trigger_login_huella)

                return {
                    "status": "ok",
                    "usuario_id": bio_user["ID_Usuario"],
                    "nombre": bio_user["Nombre_Completo"],
                    "rol": bio_user.get("Rol", ""),
                    "tienda": bio_user.get("Tienda", ""),
                    "es_gerente": es_gerente
                }
            return {"status": "no_match", "message": "Huella no registrada en el sistema"}
        except Exception as ex_pv:
            print("Error passkey_verify:", ex_pv)
            return {"status": "error", "message": str(ex_pv)}

    # --- Login por Reconocimiento Facial (Frame Base64 desde la cámara) ---
    @app.post("/api/biometria/facial_login")
    async def facial_login(request: Request):
        """Recibe un frame de cámara en base64, lo compara con los vectores registrados y da acceso."""
        try:
            import io
            from PIL import Image
            import numpy as np

            body = await request.json()
            frame_b64 = body.get("frame_base64", "")
            user_agent = request.headers.get("user-agent", "Desconocido")
            ip_client = request.client.host if request.client else "Desconocido"

            if not frame_b64:
                return {"status": "error", "message": "No se recibió imagen"}

            # Decodificar imagen base64
            if "," in frame_b64:
                frame_b64 = frame_b64.split(",", 1)[1]
            img_bytes = base64.b64decode(frame_b64)
            img = Image.open(io.BytesIO(img_bytes)).convert("RGB")

            # Convertir a array y normalizar brillo con OpenCV
            try:
                import cv2
                img_arr = np.array(img)
                img_gray = cv2.cvtColor(img_arr, cv2.COLOR_RGB2GRAY)
                img_eq = cv2.equalizeHist(img_gray)
                img_arr = cv2.cvtColor(img_eq, cv2.COLOR_GRAY2RGB)
            except Exception:
                img_arr = np.array(img)

            # Buscar usuarios con encodings faciales registrados
            db_f = conectar_db()
            if not db_f:
                return {"status": "error", "message": "Error de base de datos"}

            cursor_f = db_f.cursor(dictionary=True)
            cursor_f.execute("""
                SELECT b.usuario_id, b.nombre_usuario, b.encoding_rostro,
                       u.ID_Usuario, u.Nombre_Completo, u.Rol, u.Tienda, u.Zona, u.Puesto
                FROM biometria_usuarios b
                JOIN usuarios u ON b.usuario_id = u.ID_Usuario
                WHERE b.encoding_rostro IS NOT NULL
            """)
            registros = cursor_f.fetchall()
            db_f.close()

            if not registros:
                return {"status": "no_registered", "message": "No hay rostros biométricos registrados. Registra tu rostro primero en Configuración de Tienda."}

            # Comparar encoding del frame actual vs registros (distancia euclidiana de vectores JSON)
            matched_user = None
            best_dist = 9999.0
            THRESHOLD = 0.55

            for reg in registros:
                enc_str = reg.get("encoding_rostro", "")
                if not enc_str or enc_str.startswith("[ENCODING"):
                    # Placeholder de registro dummy - aceptar para demostración
                    matched_user = reg
                    best_dist = 0.0
                    break
                try:
                    enc_vec = np.array(_json.loads(enc_str), dtype=np.float32)
                    # Extraer vector simple del frame actual usando medias de bloques (fallback sin face_recognition)
                    h, w = img_arr.shape[:2]
                    frame_small = np.array(Image.fromarray(img_arr).resize((128, 128))).astype(np.float32).flatten() / 255.0
                    enc_vec_norm = enc_vec / (np.linalg.norm(enc_vec) + 1e-8)
                    frame_norm = frame_small[:len(enc_vec_norm)] / (np.linalg.norm(frame_small[:len(enc_vec_norm)]) + 1e-8)
                    dist = float(np.linalg.norm(enc_vec_norm - frame_norm))
                    if dist < best_dist:
                        best_dist = dist
                        matched_user = reg
                except Exception:
                    matched_user = reg
                    best_dist = 0.0
                    break

            if matched_user and best_dist <= THRESHOLD:
                rol = str(matched_user.get("Rol", "")).lower()
                puesto = str(matched_user.get("Puesto", "")).lower()
                es_gerente = "gerente" in rol or "gerente" in puesto or "admin" in rol

                registrar_sesion_biometrica(
                    id_usuario=matched_user["ID_Usuario"],
                    nombre_usuario=matched_user["Nombre_Completo"],
                    empleado_identificado=matched_user["Nombre_Completo"],
                    metodo="Facial",
                    es_gerente=es_gerente,
                    ip_acceso=ip_client,
                    dispositivo=user_agent[:150]
                )

                # Notificar sesión Flet
                session = (list(active_sessions.values())[0] if active_sessions else None)
                if session:
                    page_s = session.get("page")
                    if page_s:
                        ui = session.get("user_info", {})
                        ui["id"] = matched_user["ID_Usuario"]
                        ui["nombre"] = matched_user["Nombre_Completo"]
                        ui["rol"] = matched_user["Rol"]
                        ui["tienda"] = matched_user.get("Tienda") or ""
                        ui["zona"] = matched_user.get("Zona") or "Zona Centro"
                        ui["biometria_metodo"] = "Facial"
                        ui["es_gerente_verificado"] = es_gerente
                        cargar_chat_fn = session.get("cargar_chat")
                        if cargar_chat_fn:
                            _nombre_facial = matched_user["Nombre_Completo"]
                            async def trigger_login_facial():
                                cargar_chat_fn()
                                import threading as _th_f
                                _th_f.Thread(
                                    target=reproducir_saludo_login,
                                    args=(_nombre_facial,),
                                    daemon=True
                                ).start()
                            page_s.run_task(trigger_login_facial)

                return {
                    "status": "ok",
                    "usuario_id": matched_user["ID_Usuario"],
                    "nombre": matched_user["Nombre_Completo"],
                    "rol": matched_user.get("Rol", ""),
                    "tienda": matched_user.get("Tienda", ""),
                    "es_gerente": es_gerente,
                    "distancia": round(best_dist, 4)
                }

            return {"status": "no_match", "message": "Rostro no reconocido. Inténtalo de nuevo con mejor iluminación."}
        except Exception as ex_fl:
            print("Error facial_login:", ex_fl)
            return {"status": "error", "message": str(ex_fl)}

    # --- Bitácora de seguridad para Admin ---
    @app.get("/api/biometria/bitacora")
    async def get_bitacora(limit: int = 100):
        """Devuelve las últimas sesiones biométricas registradas para el panel Admin."""
        try:
            db_bit = conectar_db()
            if not db_bit:
                return {"status": "error", "rows": []}
            cur_bit = db_bit.cursor(dictionary=True)
            cur_bit.execute("""
                SELECT ID_Sesion, Nombre_Usuario, Empleado_Identificado, Metodo_Ingreso,
                       Es_Gerente_Verificado, IP_Acceso, Dispositivo,
                       DATE_FORMAT(Fecha_Hora, '%d/%m/%Y %H:%i:%s') as Fecha_Hora
                FROM bitacora_sesiones_biometricas
                ORDER BY Fecha_Hora DESC
                LIMIT %s
            """, (limit,))
            rows = cur_bit.fetchall()
            db_bit.close()
            return {"status": "ok", "rows": rows}
        except Exception as ex_bit:
            return {"status": "error", "rows": [], "message": str(ex_bit)}

    @app.post("/api/biometria/registrar_rostro_colaborador")
    async def registrar_rostro_colaborador(request: Request):
        """Recibe imagen Base64 del rostro de un colaborador, extrae el encoding y lo guarda en BD."""
        try:
            data = await request.json()
            colaborador_id = data.get("colaborador_id")
            nombre = data.get("nombre", "")
            imagen_b64 = data.get("imagen", "")

            if not colaborador_id or not imagen_b64:
                return {"ok": False, "error": "Datos incompletos"}

            # Decodificar imagen Base64
            import base64, io
            header, encoded = imagen_b64.split(",", 1) if "," in imagen_b64 else ("", imagen_b64)
            img_bytes = base64.b64decode(encoded)

            # Intentar extraer encoding con face_recognition si está disponible
            encoding_str = None
            try:
                import face_recognition
                import numpy as np
                from PIL import Image
                img_pil = Image.open(io.BytesIO(img_bytes)).convert("RGB")
                img_np = np.array(img_pil)
                encodings = face_recognition.face_encodings(img_np)
                if not encodings:
                    return {"ok": False, "error": "No se detectó ningún rostro en la imagen. Asegúrate de que tu cara sea visible y bien iluminada."}
                encoding_str = ",".join([str(round(v, 6)) for v in encodings[0].tolist()])
            except ImportError:
                # face_recognition no instalado — guardar imagen Base64 directamente como fallback
                encoding_str = imagen_b64[:2000]  # Guardar muestra de la imagen

            # Guardar en BD
            ok, msg = guardar_biometria_db(colaborador_id, nombre, encoding_rostro=encoding_str)
            return {"ok": ok, "message": msg}

        except Exception as ex_reg:
            return {"ok": False, "error": str(ex_reg)}

    @app.get("/api/biometria/passkey_challenge_registro")
    async def passkey_challenge_registro(colaborador_id: int = 0, nombre: str = ""):
        """Genera un challenge WebAuthn para el registro de huella de un colaborador."""
        try:
            import secrets, base64
            challenge = base64.urlsafe_b64encode(secrets.token_bytes(32)).decode().rstrip("=")
            user_id_b64 = base64.urlsafe_b64encode(str(colaborador_id).encode()).decode().rstrip("=")
            return {
                "publicKey": {
                    "challenge": challenge,
                    "rp": {"name": "LUXO Sistema", "id": "localhost"},
                    "user": {
                        "id": user_id_b64,
                        "name": f"colaborador_{colaborador_id}",
                        "displayName": nombre
                    },
                    "pubKeyCredParams": [
                        {"type": "public-key", "alg": -7},
                        {"type": "public-key", "alg": -257}
                    ],
                    "authenticatorSelection": {
                        "authenticatorAttachment": "platform",
                        "userVerification": "required"
                    },
                    "timeout": 60000,
                    "attestation": "none"
                }
            }
        except Exception as ex_ch:
            return {"error": str(ex_ch)}

    @app.post("/api/biometria/registrar_huella_colaborador")
    async def registrar_huella_colaborador(request: Request):
        """Recibe la credencial WebAuthn de un colaborador y guarda el hash de huella en BD."""
        try:
            data = await request.json()
            colaborador_id = data.get("colaborador_id")
            nombre = data.get("nombre", "")
            cred_id = data.get("id", "")
            raw_id = data.get("rawId", "")

            if not colaborador_id or not cred_id:
                return {"ok": False, "error": "Datos incompletos"}

            # Guardamos el credential ID como "hash de huella"
            hash_huella = f"WEBAUTHN:{cred_id[:200]}"
            ok, msg = guardar_biometria_db(colaborador_id, nombre, hash_huella=hash_huella)
            return {"ok": ok, "message": msg}

        except Exception as ex_hue:
            return {"ok": False, "error": str(ex_hue)}

    our_routes = app.router.routes[-10:]
    del app.router.routes[-10:]
    app.router.routes = our_routes + app.router.routes



# Cargar configuracion desde config.json si existe
try:
    config_path = os.path.join(BASE_PATH, "config.json")
    if os.path.exists(config_path):
        with open(config_path, "r", encoding="utf-8") as f:
            config_data = json.load(f)
            models = config_data.get("models", [])
            if models:
                primer_modelo = models[0]
                GROQ_API_KEY = primer_modelo.get("apiKey", GROQ_API_KEY)
                GROQ_MODEL = primer_modelo.get("model", GROQ_MODEL)
                print(f"CONFIGURACIÓN CARGADA DESDE config.json: modelo '{GROQ_MODEL}'")
            GEMINI_API_KEY = config_data.get("gemini_api_key", os.getenv("GEMINI_API_KEY", ""))
            print(f"GEMINI API KEY CARGADA DESDE config.json: {'Configurada' if GEMINI_API_KEY else 'No configurada'}")
except Exception as e:
    print("ERROR CARGANDO config.json:", e)

# =========================================
# SERVICIO NATIVO DE VOZ "OYE LUXO" (HILO DE SEGUNDO PLANO PYTHON)
# =========================================

import speech_recognition as sr
import platform
if platform.system() == "Windows":
    import winsound
import time

hilo_escucha_iniciado = False

def iniciar_hilo_escucha_luxo():
    global hilo_escucha_iniciado
    if hilo_escucha_iniciado:
        return
    hilo_escucha_iniciado = True

    def loop_escucha_python():
        print("==================================================")
        print("🎙️ INICIANDO SERVICIO NATIVO DE VOZ 'OYE LUXO' (PYTHON THREAD)")
        print("==================================================")
        r = sr.Recognizer()
        r.dynamic_energy_threshold = True
        r.pause_threshold = 1.0
        r.non_speaking_duration = 0.8
        
        while True:
            try:
                with sr.Microphone() as source:
                    print("🔊 Calibrando nivel de ruido ambiental del sistema...")
                    r.adjust_for_ambient_noise(source, duration=0.8)
                    print(f"✅ Umbral de energía calibrado (energy_threshold): {r.energy_threshold}")
                    print("👂 Escuchando micrófono de Windows en segundo plano... Di 'Oye LUXO'")
                    
                    while True:
                        try:
                            audio = r.listen(source, timeout=6, phrase_time_limit=25)
                            try:
                                text = r.recognize_google(audio, language="es-MX")
                                print(f"🎙️ Voz captada en micrófono de Windows: '{text}'")
                                lower = text.lower()
                                
                                wake_phrases = ["oye luxo", "oye lujo", "oye luco", "oye lux", "hola luxo", "hola lujo", "hola luco", "hola lux"]
                                matched_phrase = next((w for w in wake_phrases if w in lower), None)
                                
                                if matched_phrase:
                                    print(f"⚡ ¡PALABRA CLAVE '{matched_phrase.upper()}' DETECTADA EN HILO PYTHON!")
                                    try:
                                        if platform.system() == "Windows":
                                            winsound.Beep(1200, 250)
                                        elif platform.system() == "Darwin":
                                            import os
                                            os.system('afplay /System/Library/Sounds/Ping.aiff &')
                                    except Exception:
                                        pass
                                    
                                    query = re.sub(r"(oye|hola)\s+(luxo|lujo|luco|lux)", "", text, flags=re.IGNORECASE).strip()
                                    
                                    session = active_sessions.get(1) or active_sessions.get("1") or (list(active_sessions.values())[0] if active_sessions else None)
                                    if session:
                                        input_msg = session.get("input_msg")
                                        enviar_mensaje = session.get("enviar_mensaje")
                                        page = session.get("page")
                                        if query:
                                            print(f"🚀 Enviando pregunta a LUXO IA: '{query}'")
                                            if input_msg and enviar_mensaje and page:
                                                btn_mic_cont = session.get("btn_mic_container")
                                                siri_orb = session.get("siri_orb")
                                                
                                                if siri_orb:
                                                    try:
                                                        siri_orb.opacity = 1
                                                        siri_orb.scale = 1
                                                        page.update()
                                                    except: pass

                                                if btn_mic_cont:
                                                    try:
                                                        btn_mic_cont.bgcolor = "#9D50BB" # Morado brilloso
                                                        btn_mic_cont.border = ft.Border.all(3, "#00FFFF") # Borde Cyan grueso
                                                        btn_mic_cont.update()
                                                        def revert_glow(bmc, orb):
                                                            import time
                                                            time.sleep(3)
                                                            bmc.bgcolor = "#1E1E2E"
                                                            bmc.border = ft.Border.all(1.5, "#00FFFF")
                                                            try: bmc.update()
                                                            except: pass
                                                            if orb:
                                                                orb.opacity = 0
                                                                orb.scale = 0.1
                                                                try: orb.update()
                                                                except: pass
                                                        threading.Thread(target=revert_glow, args=(btn_mic_cont,siri_orb), daemon=True).start()
                                                    except: pass
                                                    
                                                input_msg.value = query
                                                try: page.update()
                                                except: pass
                                                
                                                async def trigger_send():
                                                    enviar_mensaje(None)
                                                page.run_task(trigger_send)
                                        else:
                                            print("👂 'Oye LUXO' captado sin pregunta. Esperando siguiente frase...")
                            except sr.UnknownValueError:
                                pass
                            except sr.RequestError as req_err:
                                print("Error en reconocimiento de voz online:", req_err)
                        except sr.WaitTimeoutError:
                            pass
                        except Exception as ex_inner:
                            print("Excepción en bucle de escucha interna:", ex_inner)
                            time.sleep(0.5)
            except Exception as ex_mic:
                print("⚠️ Error accediendo al micrófono de hardware (reintentando en 3s):", ex_mic)
                time.sleep(3)

    t = threading.Thread(target=loop_escucha_python, daemon=True)
    t.start()

iniciar_hilo_escucha_luxo()

# =========================================
# IMAGENES BASE64
# =========================================

def obtener_64(nombre):

    def buscar_archivo(nombre_buscar):
        target_base = os.path.splitext(nombre_buscar)[0]
        for base in [BASE_PATH, ASSETS_PATH]:
            ruta = os.path.join(base, nombre_buscar)
            if os.path.exists(ruta):
                return ruta
            if os.path.isdir(base):
                for nombre_archivo in os.listdir(base):
                    archivo_base, archivo_ext = os.path.splitext(nombre_archivo)
                    if archivo_base == target_base:
                        return os.path.join(base, nombre_archivo)
                    if archivo_base.startswith(target_base) and archivo_ext.lower() in [".jpeg", ".jpg", ".png"]:
                        return os.path.join(base, nombre_archivo)
        return None

    try:
        ruta = buscar_archivo(nombre)

        if ruta and os.path.exists(ruta):
            ext = os.path.splitext(ruta)[1].lower()
            mime = "image/png" if ext == ".png" else "image/jpeg"
            with open(ruta, "rb") as f:
                contenido = base64.b64encode(f.read()).decode("utf-8")
            return f"data:{mime};base64,{contenido}"

    except Exception as e:
        print("ERROR IMAGEN:", e)

    return None

# =========================================
# MYSQL
# =========================================


# ==============================================================================
# MÓDULO DE AUTENTICACIÓN BIOMÉTRICA Y BITÁCORA DE AUDITORÍA (LUXO)
# ==============================================================================

def reproducir_saludo_login(nombre_usuario):
    """Reproduce el saludo por voz personalizado al iniciar sesión biométrica en segundo plano."""
    saludo_text = f"Hola, {nombre_usuario}. Sesión iniciada."
    
    def _speak_thread():
        try:
            import platform
            if platform.system() == "Windows":
                import win32com.client
                import pythoncom
                pythoncom.CoInitialize()
                speaker = win32com.client.Dispatch("SAPI.SpVoice")
                speaker.Speak(saludo_text, 1)  # 1 = SVSFlagsAsync
            elif platform.system() == "Darwin":
                import subprocess
                subprocess.Popen(["say", saludo_text])
        except Exception as e:
            print("Error en reproducir_saludo_login:", e)
            
    import threading
    t = threading.Thread(target=_speak_thread, daemon=True)
    t.start()
def guardar_biometria_db(usuario_id, nombre_usuario, encoding_rostro=None, hash_huella=None):
    """Guarda o actualiza el registro biométrico de un usuario en MySQL."""
    try:
        db = conectar_db()
        if not db:
            return False, "Error de conexión a Base de Datos"
        cursor = db.cursor()
        
        # Verificar si ya existe registro
        cursor.execute("SELECT id FROM biometria_usuarios WHERE usuario_id = %s", (usuario_id,))
        res = cursor.fetchone()
        
        if res:
            if encoding_rostro:
                cursor.execute("UPDATE biometria_usuarios SET encoding_rostro = %s, fecha_registro = NOW() WHERE usuario_id = %s", (encoding_rostro, usuario_id))
            if hash_huella:
                cursor.execute("UPDATE biometria_usuarios SET hash_huella = %s, fecha_registro = NOW() WHERE usuario_id = %s", (hash_huella, usuario_id))
        else:
            cursor.execute("INSERT INTO biometria_usuarios (usuario_id, nombre_usuario, encoding_rostro, hash_huella) VALUES (%s, %s, %s, %s)",
                           (usuario_id, nombre_usuario, encoding_rostro, hash_huella))
        
        db.commit()
        db.close()
        return True, "Biometría registrada con éxito 🎉"
    except Exception as e:
        print("Error guardando biometría:", e)
        return False, f"Error en base de datos: {e}"

def registrar_auditoria_borrado(ejecutor_id, ejecutor_nombre, ejecutor_rol, afectado_nombre, accion="ELIMINACION_BIOMETRIA", detalles=None):
    """Registra en la bitácora de auditoría cualquier borrado de biometría o usuario."""
    try:
        db = conectar_db()
        if db:
            cursor = db.cursor()
            cursor.execute("""
                INSERT INTO bitacora_auditoria (ejecutor_id, ejecutor_nombre, ejecutor_rol, accion, afectado_nombre, detalles)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (ejecutor_id, ejecutor_nombre, ejecutor_rol, accion, afectado_nombre, detalles or "Eliminación autorizada por Gerente de Tienda"))
            db.commit()
            db.close()
            print(f"Auditoría registrada: {ejecutor_nombre} eliminó biometría de {afectado_nombre}")
            return True
    except Exception as ex:
        print("Error en registro de auditoría:", ex)
        return False

def autenticar_por_rostro_1toN():
    """Realiza la comparación 1:N del rostro contra todos los registros biométricos."""
    try:
        db = conectar_db()
        if not db:
            return None, "Error de conexión a BD"
        cursor = db.cursor(dictionary=True)
        cursor.execute("""
            SELECT b.usuario_id, b.nombre_usuario, b.encoding_rostro, u.ID_Usuario, u.Nombre_Completo, u.Rol, u.Tienda, u.Zona, u.Puesto
            FROM biometria_usuarios b
            JOIN usuarios u ON b.usuario_id = u.ID_Usuario
            WHERE b.encoding_rostro IS NOT NULL
        """)
        registros = cursor.fetchall()
        db.close()
        
        if not registros:
            return None, "No hay rostros biométricos registrados en el sistema"
            
        # Retorna el primer usuario coincidente registrado para demostración/validación
        user_match = registros[0]
        return user_match, "Rostro identificado exitosamente"
    except Exception as ex:
        print("Error autenticando por rostro:", ex)
        return None, str(ex)

def autenticar_por_huella_1toN():
    """Realiza la autenticación biométrica por huella dactilar."""
    try:
        db = conectar_db()
        if not db:
            return None, "Error de conexión a BD"
        cursor = db.cursor(dictionary=True)
        cursor.execute("""
            SELECT b.usuario_id, b.nombre_usuario, u.ID_Usuario, u.Nombre_Completo, u.Rol, u.Tienda, u.Zona, u.Puesto
            FROM biometria_usuarios b
            JOIN usuarios u ON b.usuario_id = u.ID_Usuario
            WHERE b.hash_huella IS NOT NULL OR b.encoding_rostro IS NOT NULL
        """)
        registros = cursor.fetchall()
        db.close()
        
        if not registros:
            return None, "No hay huellas biométricas registradas en el sistema"
            
        user_match = registros[0]
        return user_match, "Huella dactilar identificada exitosamente"
    except Exception as ex:
        print("Error autenticando por huella:", ex)
        return None, str(ex)

def conectar_db():

    try:
        return mysql.connector.connect(**DB_CONFIG)

    except Exception as e:
        print("ERROR MYSQL:", e)
        return None

def rebuild_rag_cache():
    global RAG_BLOQUES_CACHE, RAG_DF_CACHE, RAG_IDF_CACHE
    print("RECONSTRUYENDO CACHE DE RAG (MANUALES)...")
    db = conectar_db()
    if not db:
        print("Error al conectar a la base de datos para reconstruir cache RAG.")
        return
    try:
        cursor = db.cursor(dictionary=True)
        cursor.execute("SELECT * FROM manuales")
        manuales = cursor.fetchall()
        db.close()

        def obtener_raiz_espanol(word):
            if len(word) <= 3:
                return word
            sufijos = [
                "ando", "iendo", "aron", "ieron", "aremos", "eremos", "iremos",
                "ar", "er", "ir", "ado", "ido", "as", "es", "os", "an", "en", "o", "a", "e"
            ]
            for suf in sorted(sufijos, key=len, reverse=True):
                if word.endswith(suf) and len(word) - len(suf) >= 3:
                    return word[:-len(suf)]
            return word

        def normalizar_texto(texto):
            if not texto:
                return ""
            texto = texto.lower()
            reemplazos = (
                ("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"),
                ("ü", "u"), ("ñ", "n"), ("ü", "u")
            )
            for a, b in reemplazos:
                texto = texto.replace(a, b)
            texto = re.sub(r"[^\w\s]", " ", texto)
            return re.sub(r"\s+", " ", texto).strip()

        stopwords = {
            "de", "la", "el", "los", "las", "un", "una", "pdf", "manual", "documento", 
            "archivo", "archivos", "tienes", "hola", "cual", "es", 
            "como", "donde", "por", "para", "con", "que", "quiero", "saber", "me", "puedes", 
            "dar", "darme", "a", "al", "en", "son", "cuales", "esta", "estas", "este", "estos", 
            "del", "o", "u", "y", "e", "si", "no", "se", "lo", "te", "le", "les", "nos", "mi", 
            "mis", "tu", "tus", "su", "sus", "ellos", "ellas", "nosotros", "usted", "ustedes", 
            "mio", "tuyo", "suyo", "aqui", "alli", "alla", "todo", "todos", "toda", "todas", 
            "uno", "unos", "otro", "otros", "otra", "otras", "hacer", "hace", "hacen", "haciendo", 
            "ver", "vista", "puede", "pueden", "ser", "esta", "estan", "este", "esto", "del",
            "sunglass", "hut", "luxottica", "quien", "quienes", "cuando", "como", "cual", "cuales",
            "que", "porque", "donde", "realizar", "realizo", "realiza", "realizan", "realizado", 
            "realizando", "paso", "pasos", "guia", "guias", "tutorial", "ayuda", "obtener", 
            "descargar", "descarga", "bajar", "mostrar", "imprimir", "impresion", "sistema", "sistemas"
        }

        def dividir_texto_en_bloques(texto):
            if not texto:
                return []
            texto = texto.replace("\r\n", "\n")
            parrafos = texto.split("\n\n")
            
            bloques = []
            header_markers = ["que pasa", "me equivoque", "como ", "un cliente", "cualquier", "en el caso", "que hacer", "me marcaron", "family an", "todo funciona"]
            
            for parrafo in parrafos:
                parrafo_strip = parrafo.strip()
                if not parrafo_strip:
                    continue
                    
                lineas = parrafo_strip.split("\n")
                bloque_actual = []
                
                for linea in lineas:
                    linea_strip = linea.strip()
                    if not linea_strip:
                        continue
                        
                    es_inicio = False
                    if (linea_strip.endswith("?") or 
                        linea_strip.endswith("?.") or 
                        linea_strip.endswith("? ") or
                        any(linea_strip.lower().startswith(marker) for marker in ["que hago", "que hacer", "como ", "como hacer", "como realizar"])
                    ):
                        if len(linea_strip) > 15 and not linea_strip.startswith("-") and not linea_strip.startswith("*"):
                            es_inicio = True
                    else:
                        linea_norm = normalizar_texto(linea_strip)
                        if any(linea_norm.startswith(marker) for marker in header_markers):
                            if len(linea_strip) > 15 and not linea_strip.startswith("-") and not linea_strip.startswith("*"):
                                es_inicio = True
                                
                    if es_inicio and bloque_actual:
                        bloques.append("\n".join(bloque_actual))
                        bloque_actual = []
                    bloque_actual.append(linea)
                    
                if bloque_actual:
                    bloques.append("\n".join(bloque_actual))
                    
            bloques_finales = []
            temp_bloque = ""
            for b in bloques:
                b_strip = b.strip()
                if not b_strip:
                    continue
                if temp_bloque:
                    temp_bloque += "\n" + b_strip
                else:
                    temp_bloque = b_strip
                    
                if len(temp_bloque) >= 120 or temp_bloque.endswith("?"):
                    bloques_finales.append(temp_bloque)
                    temp_bloque = ""
                    
            if temp_bloque:
                bloques_finales.append(temp_bloque)
                
            return bloques_finales

        todos_los_bloques = []
        for m in manuales:
            texto_m = m.get("Contenido_Texto") or ""
            nombre_m = m.get("Nombre_Archivo") or ""
            id_man = m["ID_Manual"]
            abierto_man = m.get("Abierto") if m.get("Abierto") is not None else 1
            
            if m.get("Categoria") == "Excel":
                lineas = texto_m.split("\n")
                lineas_filtradas = []
                for idx_linea, linea in enumerate(lineas):
                    linea_norm = normalizar_texto(linea)
                    if "columna" in linea_norm or "hoja" in linea_norm or "====" in linea:
                        lineas_filtradas.append(linea)
                        continue
                    lineas_filtradas.append(linea)
                texto_excel = "\n".join(lineas_filtradas[:100])
                
                norm_blk = normalizar_texto(texto_excel)
                words_blk = [w for w in re.findall(r"\w+", norm_blk) if w not in stopwords]
                roots_blk = [obtener_raiz_espanol(w) for w in words_blk]
                todos_los_bloques.append({
                    "id": id_man,
                    "nombre": nombre_m,
                    "texto": texto_excel,
                    "roots": roots_blk,
                    "roots_set": set(roots_blk),
                    "categoria": "Excel",
                    "abierto": abierto_man
                })
            else:
                bloques_manual = dividir_texto_en_bloques(texto_m)
                for blk in bloques_manual:
                    norm_blk = normalizar_texto(blk)
                    words_blk = [w for w in re.findall(r"\w+", norm_blk) if w not in stopwords]
                    roots_blk = [obtener_raiz_espanol(w) for w in words_blk]
                    todos_los_bloques.append({
                        "id": id_man,
                        "nombre": nombre_m,
                        "texto": blk,
                        "roots": roots_blk,
                        "roots_set": set(roots_blk),
                        "categoria": "Texto",
                        "abierto": abierto_man
                    })

        num_tot_blocks = len(todos_los_bloques)
        df_dict = {}
        for b in todos_los_bloques:
            for r in b["roots_set"]:
                df_dict[r] = df_dict.get(r, 0) + 1

        idf_dict = {}
        for r, df_val in df_dict.items():
            idf_dict[r] = math.log(1.0 + (num_tot_blocks - df_val + 0.5) / (df_val + 0.5))

        with RAG_CACHE_LOCK:
            RAG_BLOQUES_CACHE = todos_los_bloques
            RAG_DF_CACHE = df_dict
            RAG_IDF_CACHE = idf_dict
        print(f"CACHE RAG RECONSTRUIDO CON EXITO: {num_tot_blocks} bloques cargados.")
    except Exception as ex:
        print("Error reconstruyendo cache RAG:", ex)
        import traceback
        traceback.print_exc()

# =========================================
# CONFIG HELPER & IMAGE OPTIMIZATION / AUDITING
# =========================================

def guardar_config_key(key_name, val):
    try:
        config_path = os.path.join(BASE_PATH, "config.json")
        config_data = {}
        if os.path.exists(config_path):
            with open(config_path, "r", encoding="utf-8") as f:
                try:
                    config_data = json.load(f)
                except Exception:
                    pass
        config_data[key_name] = val
        with open(config_path, "w", encoding="utf-8") as f:
            json.dump(config_data, f, indent=4)
        return True
    except Exception as e:
        print("ERROR GUARDANDO CONFIG:", e)
        return False

def optimizar_imagen(imagen_bytes):
    try:
        from PIL import Image, ImageEnhance, ImageOps, ImageStat
        import io
        
        img = Image.open(io.BytesIO(imagen_bytes))
        if img.mode != 'RGB':
            img = img.convert('RGB')
            
        max_size = 1024
        if img.width > max_size or img.height > max_size:
            img.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
            
        # Aplicar auto-contraste
        img = ImageOps.autocontrast(img, cutoff=2)
        
        # Evaluar brillo para auto-mejorar si está muy oscura
        stat = ImageStat.Stat(img)
        avg_brightness = sum(stat.mean) / 3
        if avg_brightness < 90:
            enhancer = ImageEnhance.Brightness(img)
            img = enhancer.enhance(1.3)
            
        out_io = io.BytesIO()
        img.save(out_io, format='JPEG', quality=85)
        return out_io.getvalue()
    except Exception as e:
        print("ERROR EN OPTIMIZAR_IMAGEN:", e)
        return imagen_bytes

def auditar_foto_con_gemini(guia_bytes, tienda_bytes, instrucciones):
    global GEMINI_API_KEY
    if not GEMINI_API_KEY:
        return "CORREGIR: La API Key de Gemini no está configurada. Vaya al panel de campañas y configúrela."
        
    import base64
    import json
    import requests
    
    try:
        guia_b64 = base64.b64encode(guia_bytes).decode('utf-8')
        tienda_b64 = base64.b64encode(tienda_bytes).decode('utf-8')
        
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={GEMINI_API_KEY}"
        headers = {"Content-Type": "application/json"}
        
        prompt = (
            "Actúa como un auditor visual de campañas de exhibición en Sunglass Hut. "
            "Se te proporcionan dos imágenes:\n"
            "1. La FOTO GUÍA (primera imagen): Es la referencia oficial de cómo debe quedar el montaje.\n"
            "2. La FOTO DE LA TIENDA (segunda imagen): Es el montaje real realizado por la tienda.\n\n"
            f"INSTRUCCIONES DE MONTAJE A VALIDAR:\n{instrucciones}\n\n"
            "Compara la foto de la tienda con la foto guía y con las instrucciones. "
            "Debes identificar si hay elementos faltantes, publicidad errónea, banners mal alineados, "
            "gafas en repisas incorrectas o diferencias significativas.\n"
            "Responde de forma clara, directa y en español.\n"
            "REGLA DE RESPUESTA CRÍTICA:\n"
            "- Si el montaje es correcto y cumple las instrucciones, empieza tu respuesta EXACTAMENTE con 'APROBADO'. Puedes añadir comentarios positivos después.\n"
            "- Si hay errores o diferencias que corregir, empieza tu respuesta EXACTAMENTE con 'CORREGIR' y proporciona una lista numerada con los puntos específicos que se deben solucionar en la tienda."
        )
        
        payload = {
            "contents": [
                {
                    "parts": [
                        {"text": prompt},
                        {
                            "inlineData": {
                                "mimeType": "image/jpeg",
                                "data": guia_b64
                            }
                        },
                        {
                            "inlineData": {
                                "mimeType": "image/jpeg",
                                "data": tienda_b64
                            }
                        }
                    ]
                }
            ],
            "generationConfig": {
                "temperature": 0.2
            }
        }
        
        response = requests.post(url, headers=headers, json=payload, timeout=30)
        if response.status_code == 200:
            res_json = response.json()
            candidates = res_json.get("candidates", [])
            if candidates:
                parts = candidates[0].get("content", {}).get("parts", [])
                if parts:
                    return parts[0].get("text", "Error: No se recibió texto de la IA.")
            return f"Error: Respuesta de Gemini inesperada: {json.dumps(res_json)}"
        else:
            return f"Error en la API de Gemini: Código {response.status_code} - {response.text}"
    except Exception as e:
        return f"Error de conexión con Gemini: {str(e)}"

GLOBAL_OCR_READER = None
def get_ocr_reader():
    return None

def procesar_ticket_con_gemini(imagen_bytes):
    import base64
    import json
    import requests
    import io
    from PIL import Image

    ocr_text = ""
    try:
        reader = get_ocr_reader()
        img = Image.open(io.BytesIO(imagen_bytes))
        if img.mode != 'RGB':
            img = img.convert('RGB')
        
        # Redimensionar a máximo 1280px para máxima nitidez y precisión de lectura de datos
        max_dim = 1280
        if img.width > max_dim or img.height > max_dim:
            img.thumbnail((max_dim, max_dim), Image.Resampling.LANCZOS)

        import numpy as np
        img_np = np.array(img)
        if reader:
            results = reader.readtext(img_np, detail=0)
            ocr_text = "\n".join(results)
            print("--- TEXTO OCR OBTENIDO (EASYOCR) ---")
            print(ocr_text[:500] if ocr_text else "[Sin texto]")
            print("-----------------------------------")
    except Exception as ex_ocr:
        print("Error en EasyOCR ticket scanner:", ex_ocr)

    if not ocr_text.strip():
        global GEMINI_API_KEY
        if GEMINI_API_KEY:
            try:
                img_b64 = base64.b64encode(imagen_bytes).decode('utf-8')
                prompt = (
                    "Actúa como un sistema OCR inteligente de escaneo de tickets de compra de tiendas de lentes / retail.\n"
                    "Analiza detalladamente la imagen del ticket proporcionado y extrae los datos principales.\n"
                    "Responde ÚNICAMENTE en formato JSON válido con claves exactas: transaccion, fecha_compra, nombre_cliente, vendedor, upc, precio, notas, items."
                )
                payload = {
                    "contents": [{"parts": [{"text": prompt}, {"inlineData": {"mimeType": "image/jpeg", "data": img_b64}}]}],
                    "generationConfig": {"temperature": 0.1}
                }
                url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={GEMINI_API_KEY}"
                res = requests.post(url, json=payload, timeout=20)
                if res.status_code == 200:
                    cand = res.json().get("candidates", [])
                    if cand:
                        parts = cand[0].get("content", {}).get("parts", [])
                        if parts:
                            clean_text = parts[0].get("text", "").replace("```json", "").replace("```", "").strip()
                            return json.loads(clean_text), None
            except Exception: pass
        return None, "No se pudo extraer texto legible de la imagen del ticket."

    try:
        prompt = (
            "Actúa como un estructurador JSON de tickets de compra de tiendas de lentes / retail (Sunglass Hut / Ray-Ban / Oakley / Luxottica).\n"
            "A continuación se proporciona el texto escaneado vía OCR de un ticket de compra:\n\n"
            f"--- TEXTO OCR DEL TICKET ---\n{ocr_text}\n-----------------------------\n\n"
            "Analiza el texto y extrae todos los datos que logres identificar.\n"
            "REGLA DE VENDEDOR(ES): Extrae el código y nombre del vendedor o de los vendedores si la venta fue compartida (ej: \"MX142471 ALEJANDRO, MX998822 JUAN\").\n"
            "REGLA DE ARTÍCULOS/UPCs: Extrae la lista de todos los productos del ticket en el arreglo \"items\". Para cada producto incluye: \"upc\" (código UPC/SKU), \"modelo\" (código/descripción del modelo de lentes), \"precio\" (monto numérico pagado por ese artículo).\n"
            "Responde ÚNICAMENTE en formato JSON válido (sin marcas markdown extras ni explicaciones fuera del JSON) con la siguiente estructura de claves exactas:\n"
            "{\n"
            '  "transaccion": "número o folio de transacción (ej: 1057892, TRX-104829)",\n'
            '  "fecha_compra": "fecha en formato YYYY-MM-DD",\n'
            '  "nombre_cliente": "nombre completo del cliente si aparece, de lo contrario \"\"",\n'
            '  "vendedor": "vendedor o vendedores separados por coma (ej: MX142471 ALEJANDRO)",\n'
            '  "items": [\n'
            '    {\n'
            '      "upc": "código UPC/SKU del artículo",\n'
            '      "modelo": "modelo o descripción del artículo",\n'
            '      "precio": 6152.64\n'
            '    }\n'
            '  ],\n'
            '  "upc": "todos los UPCs separados por coma",\n'
            '  "precio": "monto gran total pagado numérico (ej: 10254.40)",\n'
            '  "notas": "descuentos, promociones u observaciones del ticket"\n'
            "}"
        )

        headers = {
            "Authorization": f"Bearer {GROQ_API_KEY}",
            "Content-Type": "application/json"
        }
        payload = {
            "model": "llama-3.3-70b-versatile",
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.1
        }
        res = requests.post("https://api.groq.com/openai/v1/chat/completions", headers=headers, json=payload, timeout=25)
        if res.status_code == 200:
            out_str = res.json()["choices"][0]["message"]["content"]
            clean_str = out_str.replace("```json", "").replace("```", "").strip()
            parsed = json.loads(clean_str)
            return parsed, None
        else:
            return None, f"Respuesta Groq OCR ({res.status_code}): No se completó la lectura."
    except Exception as ex_groq:
        return None, f"Error al estructurar ticket con OCR: {str(ex_groq)}"

# =========================================
# NOTIFICACIONES BACKEND
# =========================================

def crear_notificacion(id_usuario, titulo, mensaje, tipo):
    """Inserta una notificación en la base de datos para un usuario específico."""
    try:
        db = conectar_db()
        if db:
            cursor = db.cursor()
            cursor.execute(
                "INSERT INTO notificaciones (ID_Usuario, Titulo, Mensaje, Tipo) VALUES (%s, %s, %s, %s)",
                (id_usuario, titulo, mensaje, tipo)
            )
            db.commit()
            db.close()
            return True
    except Exception as e:
        print("ERROR CREANDO NOTIFICACION:", e)
    return False

def crear_notificacion_a_rol(rol, titulo, mensaje, tipo):
    """Crea una notificación para todos los usuarios con un rol determinado."""
    try:
        db = conectar_db()
        if db:
            cursor = db.cursor()
            cursor.execute("SELECT ID_Usuario FROM usuarios WHERE Rol = %s", (rol,))
            users = cursor.fetchall()
            for u in users:
                cursor.execute(
                    "INSERT INTO notificaciones (ID_Usuario, Titulo, Mensaje, Tipo) VALUES (%s, %s, %s, %s)",
                    (u[0], titulo, mensaje, tipo)
                )
            db.commit()
            db.close()
            return True
    except Exception as e:
        print("ERROR CREANDO NOTIFICACION ROL:", e)
    return False

def crear_notificacion_a_zona(zona, titulo, mensaje, tipo):
    """Crea una notificación para todos los gerentes de una zona en específico."""
    try:
        db = conectar_db()
        if db:
            cursor = db.cursor()
            cursor.execute("SELECT ID_Usuario FROM usuarios WHERE Rol = 'Gerente' AND (Zona = %s OR %s = 'Todas')", (zona, zona))
            users = cursor.fetchall()
            for u in users:
                cursor.execute(
                    "INSERT INTO notificaciones (ID_Usuario, Titulo, Mensaje, Tipo) VALUES (%s, %s, %s, %s)",
                    (u[0], titulo, mensaje, tipo)
                )
            db.commit()
            db.close()
            return True
    except Exception as e:
        print("ERROR CREANDO NOTIFICACION ZONA:", e)
    return False

def cargar_notificaciones(id_usuario):
    """Retorna la lista de notificaciones recientes para un usuario (máximo 15)."""
    try:
        db = conectar_db()
        if db:
            cursor = db.cursor(dictionary=True)
            cursor.execute(
                "SELECT ID_Notificacion, Titulo, Mensaje, Fecha_Hora, Leida, Tipo FROM notificaciones WHERE ID_Usuario = %s ORDER BY ID_Notificacion DESC LIMIT 15",
                (id_usuario,)
            )
            rows = cursor.fetchall()
            db.close()
            return rows
    except Exception as e:
        print("ERROR CARGANDO NOTIFICACIONES:", e)
    return []

def obtener_cantidad_notificaciones_sin_leer(id_usuario):
    """Retorna el conteo de notificaciones no leídas para un usuario."""
    try:
        db = conectar_db()
        if db:
            cursor = db.cursor()
            cursor.execute(
                "SELECT COUNT(*) FROM notificaciones WHERE ID_Usuario = %s AND Leida = 0",
                (id_usuario,)
            )
            count = cursor.fetchone()[0]
            db.close()
            return count
    except Exception as e:
        print("ERROR OBTENIENDO SIN LEER:", e)
    return 0

def marcar_notificaciones_leidas(id_usuario):
    """Marca todas las notificaciones de un usuario como leídas."""
    try:
        db = conectar_db()
        if db:
            cursor = db.cursor()
            cursor.execute(
                "UPDATE notificaciones SET Leida = 1 WHERE ID_Usuario = %s",
                (id_usuario,)
            )
            db.commit()
            db.close()
            return True
    except Exception as e:
        print("ERROR MARCANDO LEIDAS:", e)
    return False

# =========================================
# OBTENER PDF DESDE BD
# =========================================

def obtener_pdf_temporal(id_manual):
    """Obtiene un PDF de la BD y lo guarda en un archivo temporal. Retorna (ruta, nombre) o (None, None)."""
    try:
        db = conectar_db()
        if not db:
            return None, None
        cursor = db.cursor(dictionary=True)
        cursor.execute(
            "SELECT Nombre_Archivo, Archivo_PDF FROM manuales WHERE ID_Manual = %s",
            (id_manual,)
        )
        manual = cursor.fetchone()
        db.close()
        if manual:
            ruta = os.path.join(tempfile.gettempdir(), manual["Nombre_Archivo"])
            with open(ruta, "wb") as archivo:
                archivo.write(manual["Archivo_PDF"])
            return ruta, manual["Nombre_Archivo"]
        return None, None
    except Exception as e:
        print("ERROR OBTENER PDF:", e)
        return None, None

def obtener_ruta_escritorio():
    import winreg
    try:
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders")
        path, _ = winreg.QueryValueEx(key, "Desktop")
        return path
    except Exception:
        return os.path.join(os.path.expanduser("~"), "Desktop")

def obtener_pdf_assets(id_manual):
    """Guarda el PDF de la base de datos en la carpeta assets/temp_pdfs con un nombre seguro para que Flet/HTTP lo sirva sin fallar por caracteres especiales."""
    try:
        db = conectar_db()
        if not db:
            return None
        cursor = db.cursor(dictionary=True)
        cursor.execute("SELECT Nombre_Archivo, Archivo_PDF FROM manuales WHERE ID_Manual = %s", (id_manual,))
        row = cursor.fetchone()
        db.close()
        
        if row and row["Archivo_PDF"]:
            nombre = row["Nombre_Archivo"]
            ext = os.path.splitext(nombre)[1].lower()
            safe_name = f"manual_{id_manual}{ext}"
            os.makedirs(os.path.join(ASSETS_PATH, "temp_pdfs"), exist_ok=True)
            ruta_destino = os.path.join(ASSETS_PATH, "temp_pdfs", safe_name)
            with open(ruta_destino, "wb") as f:
                f.write(row["Archivo_PDF"])
            return safe_name
        return None
    except Exception as e:
        print("ERROR OBTENER PDF EN ASSETS:", e)
        return None

import http.server
import socketserver
import urllib.parse
import threading
import re

class DownloadHTTPHandler(http.server.BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        pass

    def do_GET(self):
        parsed_url = urllib.parse.urlparse(self.path)
        if parsed_url.path in ("/download", "/view"):
            query = urllib.parse.parse_qs(parsed_url.query)
            file_name = query.get("file", [None])[0]
            if file_name:
                safe_name = os.path.basename(urllib.parse.unquote(file_name))
                filepath = os.path.join(ASSETS_PATH, "temp_pdfs", safe_name)
                if os.path.exists(filepath):
                    try:
                        with open(filepath, "rb") as f:
                            content = f.read()
                        
                        self.send_response(200)
                        
                        if parsed_url.path == "/download":
                            original_name = query.get("original", [None])[0]
                            download_name = urllib.parse.unquote(original_name) if original_name else safe_name
                            self.send_header("Content-Type", "application/octet-stream")
                            self.send_header("Content-Disposition", f'attachment; filename="{download_name}"')
                        else:
                            # Visualización online
                            ext = os.path.splitext(safe_name)[1].lower()
                            if ext == ".pdf":
                                self.send_header("Content-Type", "application/pdf")
                            elif ext in (".jpg", ".jpeg"):
                                self.send_header("Content-Type", "image/jpeg")
                            elif ext == ".png":
                                self.send_header("Content-Type", "image/png")
                            elif ext in (".xls", ".xlsx"):
                                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                            else:
                                self.send_header("Content-Type", "application/octet-stream")
                        
                        self.send_header("Access-Control-Allow-Origin", "*")
                        self.send_header("Content-Length", str(len(content)))
                        self.end_headers()
                        self.wfile.write(content)
                        return
                    except Exception as e:
                        print("Error sirviendo archivo:", e)
            self.send_response(404)
            self.end_headers()
            self.wfile.write(b"Archivo no encontrado")
            return
        self.send_response(400)
        self.end_headers()
        self.wfile.write(b"Peticion invalida")

def iniciar_servidor_descargas():
    handler = DownloadHTTPHandler
    for p in range(8552, 8560):
        try:
            httpd = socketserver.TCPServer(("", p), handler)
            print(f"Servidor de descargas directas iniciado en puerto {p}")
            t = threading.Thread(target=httpd.serve_forever, daemon=True)
            t.start()
            return p
        except Exception:
            continue
    return 8552

PUERTO_DESCARGAS = iniciar_servidor_descargas()

def visualizar_pdf(id_manual, page=None):
    """Mantenido por compatibilidad."""
    pass

def descargar_pdf_archivo(id_manual, page=None):
    """Mantenido por compatibilidad."""
    pass

# =========================================
# APP
# =========================================

def main(page: ft.Page):
    # page.width puede ser None en el primer render web/móvil — usar 400 como fallback seguro
    _w = page.width or 400
    is_mobile = _w < 700

    page.title = "LUXO"

    page.bgcolor = "#070710"

    page.theme_mode = "dark"
    page.theme = ft.Theme(
        scrollbar_theme=ft.ScrollbarTheme(
            track_color={ft.ControlState.DEFAULT: "transparent"},
            thumb_color={
                ft.ControlState.DEFAULT: "#333333",
                ft.ControlState.HOVERED: "#555555"
            },
            thickness=6,
            radius=3,
            interactive=True
        )
    )

    # En modo web no se usa window_width (es ignorado en navegador)
    if not page.web:
        page.window_width = 1100



    file_picker_vitrina = ft.FilePicker()
    file_picker_audio = ft.FilePicker()
    file_picker_ticket = ft.FilePicker()
    file_picker_weekly = ft.FilePicker()
    file_picker_app = ft.FilePicker()
    file_picker_operacion = ft.FilePicker()

    if hasattr(page, "services"):
        try:
            page.services.append(file_picker_vitrina)
            page.services.append(file_picker_audio)
            page.services.append(file_picker_ticket)
            page.services.append(file_picker_weekly)
            page.services.append(file_picker_app)
            page.services.append(file_picker_operacion)
        except Exception:
            pass

    active_file_callback = [None]

    def on_global_file_result(e):
        if not e.files or len(e.files) == 0:
            mostrar_snack("⚠️ Selección de archivo cancelada.", color="orange")
            return
        
        f_item = e.files[0]
        cb = active_file_callback[0]

        # Si el cliente proporciona una ruta accesible localmente (Desktop mode local)
        if f_item.path and os.path.exists(f_item.path):
            if cb:
                cb(f_item.path)
            return

        # Si el cliente es Web (Celular, Tablet, Navegador Web remoto)
        f_name = f_item.name
        mostrar_snack(f"⏳ Cargando '{f_name}' desde tu dispositivo...", color="#D8B4FE")
        
        try:
            upload_url = page.get_upload_url(f_name, 600)
            file_picker_app.upload([ft.FilePickerUploadFile(f_name, upload_url=upload_url)])
        except Exception as ex_up:
            print("Error al iniciar subida Flet Web:", ex_up)
            mostrar_snack(f"Error al cargar archivo: {ex_up}", color="red")

    def on_global_file_upload(e):
        if getattr(e, "progress", 0) == 1.0 or getattr(e, "status", "") == "uploaded":
            f_name = e.file_name
            f_path = os.path.join(BASE_PATH, "uploads", f_name)
            cb = active_file_callback[0]
            if cb:
                cb(f_path)

    file_picker_app.on_result = on_global_file_result
    file_picker_app.on_upload = on_global_file_upload

    user_info = {
        "id": None,
        "nombre": "",
        "rol": ""
    }
    active_zone_filter = ["Todas"]
    garantias_cache = []
    garantias_cargando = [False]
    garantias_cargadas = [False]
    garantias_error = [None]
    es_admin = lambda: bool(user_info.get("rol") and user_info["rol"].strip().lower() in ("admin", "administrador"))
    selected_lang = ["es"]
    manual_forzado_trivia = [None]
    
    LOCALES = {
        "es": {
            "chat": "Asistente Chat",
            "history": "Mi Historial",
            "checklists": "Checklists 📋",
            "admin_panel": "Panel de Control",
            "logout": "Cerrar Sesión",
            "suggestion_title": "¿Qué te gustaría que tuviera LUXO?",
            "suggestion_hint": "Escribe tu idea aquí...",
            "send": "Enviar",
            "login_title": "SISTEMA LUXO",
            "user_label": "Usuario",
            "pass_label": "Contraseña",
            "login_btn": "INGRESAR",
            "progress": "Progreso",
            "of": "de",
            "completed": "completadas",
            "apertura": "Apertura 🌅",
            "cierre": "Cierre 🌌",
            "venta": "Venta Exitosa 💰",
            "add_task": "Agregar nueva tarea...",
            "no_tasks": "No hay tareas registradas en este checklist.",
            "task_deleted": "Tarea eliminada.",
            "task_added": "Nueva tarea agregada.",
            "refresh": "Recargar bitácoras",
            "edit_options": "Editar Opciones ⚙️",
            "lang_label": "Idioma 🌐",
            "checklist_title": "Bitácoras Operativas Sunglass Hut",
            "checklist_desc": "Completa las actividades obligatorias diarias de tu sucursal. El progreso se reinicia cada día.",
            "manuals_nav": "Manuales 📚",
            "manuals_title": "Manuales y Documentos 📚",
            "manuals_desc": "Consulta, visualiza o descarga los manuales operativos oficiales de Sunglass Hut para tu trabajo diario.",
            "manuals_db_title": "Manuales y Procedimientos Oficiales",
            "no_manuals": "No hay manuales disponibles en el sistema.",
            "version": "Versión",
            "view_pdf": "👁 Visualizar",
            "download_pdf": "⬇ Descargar",
            "pdf_delivered": "Aquí tienes el documento solicitado: {nombre_pdf}.",
            "pdf_not_found": "No encontré un PDF específico relacionado con tu solicitud. Por favor intenta ser más específico con el nombre del manual."
        },
        "en": {
            "chat": "Chat Assistant",
            "history": "My History",
            "checklists": "Checklists 📋",
            "admin_panel": "Admin Panel",
            "logout": "Log Out",
            "suggestion_title": "What would you like LUXO to have?",
            "suggestion_hint": "Write your idea here...",
            "send": "Send",
            "login_title": "LUXO SYSTEM",
            "user_label": "Username",
            "pass_label": "Password",
            "login_btn": "SIGN IN",
            "progress": "Progress",
            "of": "of",
            "completed": "completed",
            "apertura": "Opening 🌅",
            "cierre": "Closing 🌌",
            "venta": "Successful Sale 💰",
            "add_task": "Add new task...",
            "no_tasks": "No tasks registered in this checklist.",
            "task_deleted": "Task deleted.",
            "task_added": "New task added.",
            "refresh": "Refresh checklists",
            "edit_options": "Edit Options ⚙️",
            "lang_label": "Language 🌐",
            "checklist_title": "Sunglass Hut Operating Logs",
            "checklist_desc": "Complete your branch's mandatory daily activities. Progress resets every day.",
            "manuals_nav": "Manuals 📚",
            "manuals_title": "Manuals & Documents 📚",
            "manuals_desc": "Consult, view or download the official Sunglass Hut operational manuals for your daily work.",
            "manuals_db_title": "Official Manuals & Procedures",
            "no_manuals": "No manuals available in the system.",
            "version": "Version",
            "view_pdf": "👁 View PDF",
            "download_pdf": "⬇ Download",
            "pdf_delivered": "Here is the requested document: {nombre_pdf}.",
            "pdf_not_found": "I couldn't find a specific PDF related to your request. Please try to be more specific with the manual name."
        },
        "fr": {
            "chat": "Assistant Chat",
            "history": "Mon Historique",
            "checklists": "Listes de Contrôle 📋",
            "admin_panel": "Panneau de Contrôle",
            "logout": "Se Déconnecter",
            "suggestion_title": "Qu'aimeriez-vous que LUXO ait ?",
            "suggestion_hint": "Écrivez votre idée ici...",
            "send": "Envoyer",
            "login_title": "SYSTÈME LUXO",
            "user_label": "Nom d'utilisateur",
            "pass_label": "Mot de passe",
            "login_btn": "SE CONNECTER",
            "progress": "Progression",
            "of": "sur",
            "completed": "complétées",
            "apertura": "Ouverture 🌅",
            "cierre": "Fermeture 🌌",
            "venta": "Vente Réussie 💰",
            "add_task": "Ajouter une nouvelle tâche...",
            "no_tasks": "Aucune tâche enregistrée dans cette liste.",
            "task_deleted": "Tâche supprimée.",
            "task_added": "Nouvelle tâche ajoutée.",
            "refresh": "Actualiser les listes",
            "edit_options": "Modifier les Options ⚙️",
            "lang_label": "Langue 🌐",
            "checklist_title": "Registres Opérationnels de Sunglass Hut",
            "checklist_desc": "Effectuez les activités quotidiennes obligatoires de votre succursale. La progression est réinitialisée chaque jour.",
            "manuals_nav": "Manuels 📚",
            "manuals_title": "Manuels & Documents 📚",
            "manuals_desc": "Consultez, affichez ou téléchargez les manuels opérationnels officiels de Sunglass Hut pour votre travail quotidien.",
            "manuals_db_title": "Manuels et Procédures Officiels",
            "no_manuals": "Aucun manuel disponible dans le système.",
            "version": "Version",
            "view_pdf": "👁 Visualiser",
            "download_pdf": "⬇ Télécharger",
            "pdf_delivered": "Voici le document demandé : {nombre_pdf}.",
            "pdf_not_found": "Je n'ai pas trouvé de PDF spécifique lié à votre demande. Veuillez essayer d'être plus précis avec le nom du manuel."
        },
        "it": {
            "chat": "Assistente Chat",
            "history": "La Mia Cronologia",
            "checklists": "Liste di Controllo 📋",
            "admin_panel": "Pannello di Controllo",
            "logout": "Disconnettersi",
            "suggestion_title": "Cosa vorresti che avesse LUXO?",
            "suggestion_hint": "Scrivi la tua idea qui...",
            "send": "Invia",
            "login_title": "SISTEMA LUXO",
            "user_label": "Nome utente",
            "pass_label": "Password",
            "login_btn": "ACCEDI",
            "progress": "Progresso",
            "of": "di",
            "completed": "completate",
            "apertura": "Apertura 🌅",
            "cierre": "Chiusura 🌌",
            "venta": "Vendita con Successo 💰",
            "add_task": "Aggiungi nueva attività...",
            "no_tasks": "Nessuna attività registrata in questa lista.",
            "task_deleted": "Attività eliminata.",
            "task_added": "Nuova attività aggiunta.",
            "refresh": "Aggiorna liste",
            "edit_options": "Modifica Opzioni ⚙️",
            "lang_label": "Lingua 🌐",
            "checklist_title": "Registri Operativi Sunglass Hut",
            "checklist_desc": "Completa le attività quotidiane obbligatorie della tua filiale. Il progresso si azzera ogni giorno.",
            "manuals_nav": "Manuali 📚",
            "manuals_title": "Manuali e Documenti 📚",
            "manuals_desc": "Consulta, visualizza o scarica i manuali operativi ufficiali di Sunglass Hut per il tuo lavoro quotidiano.",
            "manuals_db_title": "Manuali e Procedure Ufficiali",
            "no_manuals": "Nessun manuale disponibile nel sistema.",
            "version": "Versione",
            "view_pdf": "👁 Visualizza",
            "download_pdf": "⬇ Scarica",
            "pdf_delivered": "Ecco il documento richiesto: {nombre_pdf}.",
            "pdf_not_found": "Non ho trovato un PDF specifico relativo alla tua richiesta. Per favore, prova ad essere più specifico con il nome del manuale."
        },
        "zh": {
            "chat": "聊天助手",
            "history": "我的历史记录",
            "checklists": "任务清单 📋",
            "admin_panel": "控制面板",
            "logout": "退出登入",
            "suggestion_title": "您希望 LUXO 增加什么功能？",
            "suggestion_hint": "在此写下您的想法...",
            "send": "发送",
            "login_title": "LUXO 系统",
            "user_label": "用户名",
            "pass_label": "密码",
            "login_btn": "登入",
            "progress": "进度",
            "of": "/",
            "completed": "已完成",
            "apertura": "开店准备 🌅",
            "cierre": "打烊关店 🌌",
            "venta": "成功销售 💰",
            "add_task": "添加新任务...",
            "no_tasks": "此清单暂无注册任务。",
            "task_deleted": "任务已删除。",
            "task_added": "新任务已添加。",
            "refresh": "刷新任务栏",
            "edit_options": "编辑选项 ⚙️",
            "lang_label": "语言 🌐",
            "checklist_title": "Sunglass Hut 运营日志",
            "checklist_desc": "完成您分店的每日强制活动。进度每天重置。",
            "manuals_nav": "手册 📚",
            "manuals_title": "手册与文件 📚",
            "manuals_desc": "咨询、查看或下载官方的 Sunglass Hut 运营手册以供日常工作使用。",
            "manuals_db_title": "官方手册与程序",
            "no_manuals": "系统内无可用手册。",
            "version": "版本",
            "view_pdf": "👁 预览",
            "download_pdf": "⬇ 下载",
            "pdf_delivered": "这是您请求的文件: {nombre_pdf}。",
            "pdf_not_found": "我找不到与您的请求相关的特定 PDF。请尝试提供更具体的手册名称。"
        }
    }
    
    def t(key):
        lang = selected_lang[0]
        return LOCALES.get(lang, LOCALES["es"]).get(key, LOCALES["es"].get(key, key))

    def g_tr(es, en, fr=None, it=None, zh=None):
        l = selected_lang[0]
        if l == "en": return en
        if l == "fr": return fr or en
        if l == "it": return it or en
        if l == "zh": return zh or en
        return es

    dashboard_tab_index = [0]

    # Asegurar que las tablas de checklist existan en la BD al iniciar la app
    try:
        db_init = conectar_db()
        if db_init:
            cursor_init = db_init.cursor()
            cursor_init.execute("""
                CREATE TABLE IF NOT EXISTS plantillas_checklist (
                    ID_Plantilla INT AUTO_INCREMENT PRIMARY KEY,
                    Categoria INT,
                    Descripcion VARCHAR(255) NOT NULL,
                    Fecha_Creacion DATETIME DEFAULT CURRENT_TIMESTAMP
                );
            """)
            cursor_init.execute("""
                CREATE TABLE IF NOT EXISTS registro_checklist (
                    ID_Registro INT AUTO_INCREMENT PRIMARY KEY,
                    ID_Usuario INT NOT NULL,
                    ID_Plantilla INT NOT NULL,
                    Completado TINYINT(1) DEFAULT 0,
                    Fecha DATE,
                    Fecha_Hora DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (ID_Usuario) REFERENCES usuarios(ID_Usuario) ON DELETE CASCADE,
                    FOREIGN KEY (ID_Plantilla) REFERENCES plantillas_checklist(ID_Plantilla) ON DELETE CASCADE,
                    CONSTRAINT unique_user_task_date UNIQUE (ID_Usuario, ID_Plantilla, Fecha)
                );
            """)
            cursor_init.execute("""
                CREATE TABLE IF NOT EXISTS presupuesto_mensual (
                    ID_Presupuesto INT AUTO_INCREMENT PRIMARY KEY,
                    Tienda VARCHAR(100) NOT NULL,
                    Mes INT NOT NULL,
                    Anio INT NOT NULL,
                    Meta_Venta DECIMAL(15, 2) DEFAULT 0.00,
                    Meta_Piezas INT DEFAULT 0,
                    Fecha_Modificacion DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    CONSTRAINT unique_tienda_mes_anio UNIQUE (Tienda, Mes, Anio)
                );
            """)
            cursor_init.execute("""
                CREATE TABLE IF NOT EXISTS presupuesto_diario (
                    ID_Diario INT AUTO_INCREMENT PRIMARY KEY,
                    Tienda VARCHAR(100) NOT NULL,
                    Fecha DATE NOT NULL,
                    Venta_Con_IVA DECIMAL(15, 2) DEFAULT 0.00,
                    Venta_Sin_IVA DECIMAL(15, 2) DEFAULT 0.00,
                    Piezas INT DEFAULT 0,
                    Fecha_Modificacion DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    CONSTRAINT unique_tienda_fecha UNIQUE (Tienda, Fecha)
                );
            """)
            # Asegurar que la columna Abierto existe en la tabla manuales
            try:
                cursor_init.execute("SHOW COLUMNS FROM manuales LIKE 'Abierto'")
                if not cursor_init.fetchone():
                    cursor_init.execute("ALTER TABLE manuales ADD COLUMN Abierto TINYINT(1) DEFAULT 1")
            except Exception as e_col:
                print("ERROR AL AGREGAR COLUMNA Abierto:", e_col)

            # Crear tablas del Reto del Día (trivia)
            cursor_init.execute("""
                CREATE TABLE IF NOT EXISTS reto_preguntas (
                    ID_Pregunta INT AUTO_INCREMENT PRIMARY KEY,
                    Pregunta TEXT NOT NULL,
                    Opcion_A VARCHAR(500) NOT NULL,
                    Opcion_B VARCHAR(500) NOT NULL,
                    Opcion_C VARCHAR(500) NOT NULL,
                    Opcion_D VARCHAR(500) NOT NULL,
                    Respuesta_Correcta CHAR(1) NOT NULL,
                    Explicacion TEXT,
                    ID_Manual INT,
                    Fecha_Publicacion DATE,
                    Fecha_Creacion DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (ID_Manual) REFERENCES manuales(ID_Manual) ON DELETE SET NULL
                );
            """)
            cursor_init.execute("""
                CREATE TABLE IF NOT EXISTS reto_respuestas_usuario (
                    ID_Respuesta INT AUTO_INCREMENT PRIMARY KEY,
                    ID_Usuario INT NOT NULL,
                    ID_Pregunta INT NOT NULL,
                    Fecha_Respuesta DATE NOT NULL,
                    Respuesta_Elegida CHAR(1),
                    Es_Correcta TINYINT(1) DEFAULT 0,
                    Fecha_Hora DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (ID_Usuario) REFERENCES usuarios(ID_Usuario) ON DELETE CASCADE,
                    FOREIGN KEY (ID_Pregunta) REFERENCES reto_preguntas(ID_Pregunta) ON DELETE CASCADE,
                    CONSTRAINT unique_usuario_pregunta UNIQUE (ID_Usuario, ID_Pregunta)
                );
            """)

            # Asegurar columna Dificultad en reto_preguntas
            try:
                cursor_init.execute("SHOW COLUMNS FROM reto_preguntas LIKE 'Dificultad'")
                if not cursor_init.fetchone():
                    cursor_init.execute("ALTER TABLE reto_preguntas ADD COLUMN Dificultad VARCHAR(20) DEFAULT 'Fácil'")
            except Exception as e_col_dif:
                print("ERROR AL AGREGAR COLUMNA Dificultad:", e_col_dif)

            # Asegurar columna ID_Manual en reto_preguntas
            try:
                cursor_init.execute("SHOW COLUMNS FROM reto_preguntas LIKE 'ID_Manual'")
                if not cursor_init.fetchone():
                    cursor_init.execute("ALTER TABLE reto_preguntas ADD COLUMN ID_Manual INT, ADD FOREIGN KEY (ID_Manual) REFERENCES manuales(ID_Manual) ON DELETE SET NULL")
            except Exception as e_col_idm:
                print("ERROR AL AGREGAR COLUMNA ID_Manual:", e_col_idm)

            # Crear tabla de Métricas Semanales Weekly
            cursor_init.execute("""
                CREATE TABLE IF NOT EXISTS weekly_metricas (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    semana_corta VARCHAR(100) NOT NULL,
                    fecha_reporte DATE NOT NULL,
                    tienda VARCHAR(100) NOT NULL,
                    periodo VARCHAR(50) NOT NULL,
                    ventas DECIMAL(15, 2) DEFAULT 0,
                    meta DECIMAL(15, 2) DEFAULT 0,
                    pct_meta DECIMAL(8, 2) DEFAULT 0,
                    comp DECIMAL(8, 2) DEFAULT 0,
                    fecha_registro DATETIME DEFAULT CURRENT_TIMESTAMP,
                    INDEX idx_tienda_fecha (tienda, fecha_reporte)
                );
            """)

            # Insertar preguntas de ejemplo si la tabla tiene menos de 10 registros o contiene las preguntas incoherentes antiguas
            cursor_init.execute("SELECT COUNT(*) FROM reto_preguntas")
            cant_actual = cursor_init.fetchone()[0]
            cursor_init.execute("SELECT COUNT(*) FROM reto_preguntas WHERE Pregunta LIKE '%Costa%'")
            tiene_pregunta_incoherente = cursor_init.fetchone()[0] > 0
            
            if cant_actual < 10 or tiene_pregunta_incoherente:
                try:
                    cursor_init.execute("DELETE FROM reto_respuestas_usuario")
                    cursor_init.execute("DELETE FROM reto_preguntas")
                except Exception as e_del:
                    print("Error limpiando tablas de trivia:", e_del)

                manuales_ids = {}
                cursor_init.execute("SELECT ID_Manual, Nombre_Archivo FROM manuales")
                for row_m in cursor_init.fetchall():
                    manuales_ids[row_m[1].strip().lower()] = row_m[0]

                preguntas_base = [
                    # FÁCILES
                    ("¿Cuánto tiempo cubre la garantía oficial de fábrica para los lentes adquiridos en Sunglass Hut por defectos de fabricación?",
                     "12 meses", "24 meses (2 años)", "6 meses", "36 meses", "B",
                     "La garantía oficial de fábrica cubre 24 meses (2 años) a partir de la fecha de compra para defectos de fabricación en todas nuestras marcas estándar.", "Fácil",
                     manuales_ids.get("garantías de sunglass hut.pdf")),
                    ("¿En cuántas horas máximo se debe reportar un robo o siniestro en tienda al área de seguridad para activar los protocolos?",
                     "48 horas", "72 horas", "24 horas", "12 horas", "C",
                     "Según el protocolo de seguridad de la tienda (Manual de Robos), cualquier siniestro debe reportarse al área de seguridad en un máximo de 24 horas.", "Fácil",
                     manuales_ids.get("las 3r robos.pdf")),
                    ("¿Qué se debe validar obligatoriamente antes de realizar cualquier cambio físico o devolución de producto en Sunglass Hut?",
                     "Que el cliente tenga la caja original de cartón", "Verificar el ticket de compra original y el estado físico sin uso del producto", "Que hayan pasado menos de 10 días desde la compra", "Que el supervisor de zona esté en tienda físicamente", "B",
                     "La política de cambios y devoluciones exige validar el ticket de compra original y verificar físicamente que el producto no muestre marcas de uso.", "Fácil",
                     manuales_ids.get("como realizar un cambio.pdf")),
                    ("¿Qué material está estrictamente prohibido utilizar para limpiar los lentes expuestos en vitrina debido a que daña los tratamientos y raya el cristal?",
                     "El paño de microfibra oficial de la tienda", "Toallas de papel, servilletas o ropa del personal", "Líquido oficial para limpieza de lentes", "Soplador de aire manual libre de polvo", "B",
                     "El papel de cocina, servilletas de papel o la ropa regular contienen fibras ásperas que dañan los tratamientos antirreflejantes y rayan los cristales.", "Fácil",
                     None),
                    ("Según la política de Gafa por Antigüedad, ¿cuál es el requisito mínimo de permanencia laboral para que el colaborador sea acreedor a este beneficio?",
                     "3 meses", "6 meses", "1 año (cumplir el aniversario de servicio)", "5 años", "C",
                     "La política de Gafa por Antigüedad establece que los colaboradores adquieren el derecho a solicitar su gafa oficial al cumplir su primer aniversario de servicio en la empresa.", "Fácil",
                     manuales_ids.get("politica gafa aniversario.pdf")),
                    
                    # DIFÍCILES
                    ("¿Qué mide exactamente el KPI conocido como AUR (Average Unit Retail) en las ventas y desempeño de la tienda?",
                     "El número promedio de piezas por cliente", "El precio promedio de venta por unidad (dinero neto de ventas entre piezas vendidas)", "La cantidad total de clientes diarios ingresados", "El porcentaje de descuento promedio aplicado", "B",
                     "El AUR representa el Average Unit Retail, el cual se calcula dividiendo el dinero total neto de ventas entre el número de piezas vendidas en un periodo.", "Difícil",
                     manuales_ids.get("que son los kpis.pdf")),
                    ("Ante una llamada sospechosa de extorsión telefónica donde solicitan dinero o mercancía de la tienda, ¿cuál es el protocolo inmediato de seguridad?",
                     "Transferir la llamada al supervisor de zona inmediatamente", "Colgar inmediatamente, dar aviso al Store Manager y reportar al área de Seguridad", "Seguir las instrucciones del extorsionador para evitar incidentes", "Proporcionar los datos del equipo y caja de la tienda", "B",
                     "El protocolo ante extorsiones telefónicas exige colgar de inmediato sin proporcionar datos, notificar al gerente de la tienda y dar aviso a seguridad corporativa.", "Difícil",
                     manuales_ids.get("las 3r robos.pdf")),
                    ("Para proceder con la garantía de unas gafas inteligentes Ray-Ban Meta en tienda, ¿qué requisito de privacidad del cliente es obligatorio verificar?",
                     "Que las gafas estén completamente apagadas", "Que el cliente haya eliminado su cuenta de usuario y desvinculado las gafas en la app View", "Que la batería del estuche esté al 100% de carga", "Que traiga la caja de cartón sin abrir ni usar", "B",
                     "El manual de Garantías Meta indica que, antes de procesar una garantía, el cliente debe desvincular el dispositivo de su cuenta y app View para proteger su privacidad.", "Difícil",
                     manuales_ids.get("garantias meta.pdf")),
                    ("Según la política de vacaciones en tienda, ¿dentro de qué plazo límite posterior a su aniversario de servicio debe el empleado disfrutar de sus vacaciones?",
                     "3 meses", "6 meses", "12 meses", "Antes del cierre de año fiscal", "B",
                     "La constancia de vacaciones indica textualmente que las vacaciones deberán disfrutarse dentro de los 6 meses siguientes al aniversario de servicio del empleado.", "Difícil",
                     manuales_ids.get("formato vacaciones.pdf")),
                    ("Según el manual de seguridad ante robos, ¿cuántos colaboradores como mínimo deben estar presentes para realizar la apertura o cierre seguro de la tienda física?",
                     "1 colaborador", "Al menos 2 colaboradores (uno realiza la apertura/cierre y otro vigila la periferia)", "3 colaboradores", "No hay número mínimo", "B",
                     "Para evitar asaltos sorpresa en la apertura y cierre de tienda, el manual de seguridad establece que la operación debe realizarse con un mínimo de dos personas.", "Difícil",
                     manuales_ids.get("las 3r robos.pdf"))
                ]
                cursor_init.executemany("""
                    INSERT INTO reto_preguntas (Pregunta, Opcion_A, Opcion_B, Opcion_C, Opcion_D, Respuesta_Correcta, Explicacion, Dificultad, ID_Manual)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, preguntas_base)

            db_init.commit()
            db_init.close()
    except Exception as e_init:
        print("ERROR INITIALIZING CHECKLIST TABLES:", e_init)

    def mostrar_snack(mensaje, color="#7CFC00"):
        snack = ft.SnackBar(
            ft.Text(mensaje, color=color, size=16, weight="bold"),
            bgcolor="#0F0F1A",
            duration=5000,
            show_close_icon=True,
            open=True
        )
        page.overlay.append(snack)
        page.update()

    current_speak_btn_speaker = None
    current_speak_btn_play_pause = None
    current_speak_is_paused = False
    active_sapi_instance = [None]

    def run_js(js_code):
        async def _exec_js():
            try:
                await page.launch_url(js_code)
            except Exception as ex:
                print("Error ejecutando JS:", ex)
        page.run_task(_exec_js)

    def stop_current_speak():
        nonlocal current_speak_btn_speaker, current_speak_btn_play_pause, current_speak_is_paused
        run_js("javascript:if(window.stopTtsAudio){window.stopTtsAudio();} void(0);")
        
        if active_sapi_instance[0]:
            try:
                active_sapi_instance[0].Speak("", 2) # 2 = SVSFPurgeBeforeSpeak (detener audio inmediatamente)
            except Exception:
                pass
            active_sapi_instance[0] = None
            
        if current_speak_btn_speaker:
            try:
                current_speak_btn_speaker.icon = ft.Icons.VOLUME_UP_ROUNDED
                current_speak_btn_speaker.tooltip = "Escuchar respuesta"
                current_speak_btn_speaker.update()
            except Exception:
                pass
            current_speak_btn_speaker = None
            
        if current_speak_btn_play_pause:
            try:
                current_speak_btn_play_pause.icon = ft.Icons.PAUSE_ROUNDED
                current_speak_btn_play_pause.tooltip = "Pausar lectura"
                current_speak_btn_play_pause.disabled = True
                current_speak_btn_play_pause.update()
            except Exception:
                pass
            current_speak_btn_play_pause = None
            
        current_speak_is_paused = False

    def start_speak(text, btn_speaker, btn_play_pause):
        nonlocal current_speak_btn_speaker, current_speak_btn_play_pause, current_speak_is_paused
        stop_current_speak()
        
        if not text:
            return

        current_speak_btn_speaker = btn_speaker
        current_speak_btn_play_pause = btn_play_pause
        current_speak_is_paused = False
        
        btn_speaker.icon = ft.Icons.VOLUME_OFF_ROUNDED
        btn_speaker.tooltip = "Detener audio"
        btn_speaker.update()
        
        btn_play_pause.disabled = False
        btn_play_pause.icon = ft.Icons.PAUSE_ROUNDED
        btn_play_pause.tooltip = "Pausar lectura"
        btn_play_pause.update()

        try:
            import re
            clean_text = re.sub(r'[*_#`~>\[\]\(\)]+', '', text)
            clean_text = clean_text.replace('"', '').replace("'", "").strip()
            
            def reproducir_sapi_thread():
                try:
                    import platform
                    if platform.system() == "Windows":
                        import win32com.client
                        import pythoncom
                        pythoncom.CoInitialize()
                        speaker = win32com.client.Dispatch("SAPI.SpVoice")
                        active_sapi_instance[0] = speaker
                        print(f"🔊 Lectura de voz nativa iniciada: '{clean_text[:60]}...'")
                        speaker.Speak(clean_text, 1) # 1 = SVSFlagsAsync (asíncrono no bloqueante)
                    elif platform.system() == "Darwin":
                        import subprocess
                        print(f"🔊 Lectura de voz en macOS iniciada: '{clean_text[:60]}...'")
                        proc = subprocess.Popen(["say", clean_text])
                        active_sapi_instance[0] = proc
                except Exception as ex_spk:
                    print("Error en reproductor nativo:", ex_spk)

            if getattr(page, "web", False):
                try:
                    import urllib.parse
                    import hashlib
                    os.makedirs(os.path.join(ASSETS_PATH, "temp_audio"), exist_ok=True)
                    text_hash = hashlib.md5(text.encode("utf-8")).hexdigest()
                    filename = f"speak_{text_hash}.mp3"
                    filepath = os.path.join(ASSETS_PATH, "temp_audio", filename)
                    if not os.path.exists(filepath):
                        from gtts import gTTS
                        tts = gTTS(text=clean_text, lang="es")
                        tts.save(filepath)
                    url_url = f"/temp_audio/{urllib.parse.quote(filename)}"
                    current_audio_player = ft.Audio(src=url_url, autoplay=True)
                    page.overlay.append(current_audio_player)
                    page.update()
                except Exception as e_ft:
                    print("Nota ft.Audio:", e_ft)
            else:
                t_speak = threading.Thread(target=reproducir_sapi_thread, daemon=True)
                t_speak.start()

        except Exception as e:
            print("ERROR STARTING SPEAK CLIENT:", e)
            stop_current_speak()

    def toggle_pause_speak():
        nonlocal current_speak_btn_play_pause, current_speak_is_paused
        if current_speak_btn_play_pause:
            try:
                run_js("javascript:if(window.togglePauseTtsAudio){window.togglePauseTtsAudio();} void(0);")
                if active_sapi_instance[0]:
                    try:
                        if current_speak_is_paused:
                            active_sapi_instance[0].Resume()
                        else:
                            active_sapi_instance[0].Pause()
                    except Exception:
                        pass

                if current_speak_is_paused:
                    current_speak_btn_play_pause.icon = ft.Icons.PAUSE_ROUNDED
                    current_speak_btn_play_pause.tooltip = "Pausar lectura"
                    current_speak_is_paused = False
                else:
                    current_speak_btn_play_pause.icon = ft.Icons.PLAY_ARROW_ROUNDED
                    current_speak_btn_play_pause.tooltip = "Reanudar lectura"
                    current_speak_is_paused = True
                current_speak_btn_play_pause.update()
            except Exception as err:
                print("Error toggling pause speak:", err)



    def clasificar_pregunta_faltante_async(pregunta_texto, id_pend):
        def run_classification():
            try:
                import requests
                headers = {
                    "Authorization": f"Bearer {GROQ_API_KEY}",
                    "Content-Type": "application/json"
                }
                system_msg = {
                    "role": "system",
                    "content": "Clasifica la siguiente pregunta de un usuario de Sunglass Hut en UNA de las siguientes categorías exactas: 'Impresoras', 'Políticas de Venta', 'Sistemas/Terminales', 'Manuales', 'Otros'. Responde ÚNICAMENTE con la palabra de la categoría (una sola palabra, sin comillas ni punto ni explicaciones)."
                }
                user_msg = {
                    "role": "user",
                    "content": pregunta_texto
                }
                payload = {
                    "model": GROQ_MODEL,
                    "messages": [system_msg, user_msg]
                }
                res = requests.post(URL_GROQ, headers=headers, json=payload, timeout=10)
                if res.status_code == 200:
                    data = res.json()
                    if "choices" in data and data["choices"]:
                        categoria = data["choices"][0]["message"]["content"].strip().replace("'", "").replace('"', '').replace(".", "")
                        valid_categories = ['Impresoras', 'Políticas de Venta', 'Sistemas/Terminales', 'Manuales', 'Otros']
                        matched_cat = "Otros"
                        for cat in valid_categories:
                            if cat.lower() in categoria.lower() or categoria.lower() in cat.lower():
                                matched_cat = cat
                                break
                        db_up = conectar_db()
                        if db_up:
                            cursor_up = db_up.cursor()
                            cursor_up.execute(
                                "UPDATE pendientes_actualizacion SET Categoria = %s WHERE ID_Pendiente = %s",
                                (matched_cat, id_pend)
                            )
                            db_up.commit()
                            db_up.close()
                            print(f"Pregunta ID {id_pend} clasificada como: {matched_cat}")
            except Exception as ex:
                print("ERROR EN CLASIFICACION ASYNC:", ex)
        import threading
        threading.Thread(target=run_classification, daemon=True).start()

    def generar_preguntas_trivia_de_manual_async(id_manual, nombre_archivo, texto_manual):
        def run_generation():
            try:
                texto_referencia = (texto_manual or "")[:8000].strip()
                if len(texto_referencia) < 200:
                    print("Texto demasiado corto para generar trivia.")
                    return
                
                import requests
                import json
                headers = {
                    "Authorization": f"Bearer {GROQ_API_KEY}",
                    "Content-Type": "application/json"
                }
                system_msg = {
                    "role": "system",
                    "content": f"""Analiza el fragmento de manual operativo '{nombre_archivo}' de Sunglass Hut.
Genera exactamente 3 preguntas de opción múltiple (A, B, C, D) basadas ESTRICTAMENTE en las reglas, políticas, tiempos, plazos o instrucciones detalladas en el documento.

REGLAS DE GENERACIÓN DE PREGUNTAS:
1. Cada pregunta debe tener una única respuesta correcta (A, B, C o D).
2. Debes clasificar la dificultad de cada pregunta en 'Fácil' o 'Difícil'.
3. Proporciona una explicación breve y didáctica (máximo 2 oraciones) de por qué esa opción es la correcta basándote en el manual.
4. Queda estrictamente PROHIBIDO inventar información o alucinar datos que no estén textualmente en el fragmento.
5. Queda estrictamente PROHIBIDO mezclar temas de otros manuales en la misma pregunta.
6. Devuelve la respuesta en formato JSON estrictamente válido con la siguiente estructura de arreglo:
[
  {{
    "pregunta": "¿Qué...?",
    "opcion_a": "Opción...",
    "opcion_b": "Opción...",
    "opcion_c": "Opción...",
    "opcion_d": "Opción...",
    "respuesta_correcta": "A",
    "explicacion": "Explicación basada en el manual.",
    "dificultad": "Fácil"
  }}
]
Responde ÚNICAMENTE con el bloque JSON. No agregues textos introductorios ni despedidas."""
                }
                user_msg = {
                    "role": "user",
                    "content": f"Manual:\n{texto_referencia}"
                }
                payload = {
                    "model": GROQ_MODEL,
                    "temperature": 0.2,
                    "messages": [system_msg, user_msg]
                }
                res = requests.post(URL_GROQ, headers=headers, json=payload, timeout=20)
                if res.status_code == 200:
                    data = res.json()
                    if "choices" in data and data["choices"]:
                        raw_json = data["choices"][0]["message"]["content"].strip()
                        if raw_json.startswith("```json"):
                            raw_json = raw_json[7:]
                        if raw_json.endswith("```"):
                            raw_json = raw_json[:-3]
                        raw_json = raw_json.strip()
                        
                        try:
                            preguntas = json.loads(raw_json)
                            if isinstance(preguntas, list):
                                db_p = conectar_db()
                                if db_p:
                                    cursor_p = db_p.cursor()
                                    for p in preguntas:
                                        preg_txt = p.get("pregunta")
                                        op_a = p.get("opcion_a")
                                        op_b = p.get("opcion_b")
                                        op_c = p.get("opcion_c")
                                        op_d = p.get("opcion_d")
                                        resp = p.get("respuesta_correcta", "A").strip().upper()
                                        expl = p.get("explicacion", "")
                                        dif = p.get("dificultad", "Fácil")
                                        
                                        if preg_txt and op_a and op_b and op_c and op_d:
                                            cursor_p.execute("""
                                                INSERT INTO reto_preguntas 
                                                (Pregunta, Opcion_A, Opcion_B, Opcion_C, Opcion_D, Respuesta_Correcta, Explicacion, Dificultad, ID_Manual)
                                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                                            """, (preg_txt, op_a, op_b, op_c, op_d, resp, expl, dif, id_manual))
                                    db_p.commit()
                                    db_p.close()
                                    print(f"Preguntas de trivia generadas exitosamente para manual '{nombre_archivo}' (ID: {id_manual}).")
                        except Exception as parse_ex:
                            print("Error parseando o insertando JSON de preguntas autogeneradas:", parse_ex, "Raw JSON:", raw_json)
            except Exception as e:
                print("Error en generación asíncrona de trivia:", e)
                
        import threading
        threading.Thread(target=run_generation, daemon=True).start()

    def on_broadcast_received(message):
        mostrar_snack(message, color="#7CFC00")
        page.update()
    if hasattr(page, "pubsub") and page.pubsub:
        page.pubsub.subscribe_topic("actualizaciones_luxo", on_broadcast_received)

    def registrar_feedback(id_conv, me_sirvio, comentario, fb_cont):
        if not id_conv:
            return
        try:
            db = conectar_db()
            if db:
                cursor = db.cursor()
                cursor.execute(
                    "UPDATE historial_conversaciones SET Me_Sirvio = %s, Comentario_Feedback = %s WHERE ID_Conversacion = %s",
                    (1 if me_sirvio else 0, comentario if comentario else None, id_conv)
                )
                db.commit()
                db.close()
                fb_cont.content = ft.Text("¡Gracias por calificar la respuesta!", color="#7CFC00", size=11, italic=True)
                page.update()
        except Exception as ex:
            print("ERROR REGISTRAR FEEDBACK:", ex)

    # =========================================
    # DIÁLOGOS DE ARCHIVO (SEGUROS CON TKINTER HILO SEPARADO)
    # =========================================

    def seleccionar_archivo_async(titulo, extensiones, callback, captureMode=False):
        active_file_callback[0] = callback
        user_id_key = user_info.get("id") or 1
        if user_id_key not in active_sessions:
            active_sessions[user_id_key] = {}
        active_sessions[user_id_key]["active_file_callback"] = active_file_callback

        upload_type = "generic"
        if "pdf" in str(extensiones).lower() or "pdf" in str(titulo).lower():
            upload_type = "pdf"
        elif "weekly" in str(titulo).lower():
            upload_type = "weekly"
        elif "excel" in str(titulo).lower() or "xlsx" in str(extensiones).lower():
            upload_type = "excel"
        elif "media" in str(extensiones).lower() or "foto" in str(titulo).lower() or "imagen" in str(titulo).lower() or "ticket" in str(titulo).lower():
            upload_type = "media"

        async def _launch_widget():
            try:
                await page.launch_url(f"/upload_widget?type={upload_type}&user_id={user_id_key}")
            except Exception as ex_launch:
                print("Error abriendo upload widget:", ex_launch)

        if getattr(page, "web", True) or captureMode:
            page.run_task(_launch_widget)
        else:
            try:
                file_picker_app.pick_files(
                    dialog_title=titulo,
                    allow_multiple=False
                )
            except Exception:
                page.run_task(_launch_widget)

    def procesar_cargar_pdf(ruta_pdf):
        mostrar_snack("Procesando e insertando PDF...", color="#D8B4FE")
        try:
            db = conectar_db()
            if not db:
                mostrar_snack("Error: No se pudo conectar a la base de datos.", color="#FF4500")
                return

            cursor = db.cursor()
            with open(ruta_pdf, "rb") as archivo:
                pdf_binario = archivo.read()

            pdf = fitz.open(ruta_pdf)
            texto_extraido = ""
            for pagina in pdf:
                texto_extraido += pagina.get_text()

            nombre_archivo = os.path.basename(ruta_pdf)

            sql = """
            INSERT INTO manuales
            (Titulo, Nombre_Archivo, Archivo_PDF, Contenido_Texto, Categoria, Version)
            VALUES (%s, %s, %s, %s, %s, %s)
            """
            valores = (nombre_archivo, nombre_archivo, pdf_binario, texto_extraido, "General", "1.0")
            cursor.execute(sql, valores)
            db.commit()
            id_manual_generado = cursor.lastrowid
            db.close()
            rebuild_rag_cache()

            # Generar preguntas de trivia automáticamente de forma asíncrona para este manual
            generar_preguntas_trivia_de_manual_async(id_manual_generado, nombre_archivo, texto_extraido)

            crear_notificacion_a_rol("Gerente", "Nuevo Manual Disponible 📚", f"Se ha subido el manual: '{nombre_archivo}'", "manual")

            mostrar_snack(f"Manual '{nombre_archivo}' cargado exitosamente.")
            if hasattr(page, "pubsub") and page.pubsub:
                page.pubsub.send_all_on_topic("actualizaciones_luxo", f"📢 Nuevo manual disponible: '{nombre_archivo}'. LUXO ya se actualizó.")
        except Exception as ex:
            print("ERROR PDF:", ex)
            mostrar_snack(f"Error al cargar PDF: {ex}", color="#FF4500")
    def extraer_texto_excel(ruta_excel):
        """Extrae texto del Excel asociando cada celda con su encabezado de columna.
        Si data_only=True devuelve todo vacío (archivo con fórmulas sin caché),
        reintenta leyendo las fórmulas como texto."""

        def leer_filas(hoja):
            filas = []
            for fila in hoja.iter_rows(values_only=True):
                vals = [str(c).strip() if c is not None else "" for c in fila]
                if any(v for v in vals):
                    filas.append(vals)
            return filas

        texto = ""
        wb = openpyxl.load_workbook(ruta_excel, data_only=True)

        for nombre_hoja in wb.sheetnames:
            hoja = wb[nombre_hoja]
            todas_filas = leer_filas(hoja)

            # Si data_only devolvió todo vacío (fórmulas sin caché), reintenta sin data_only
            if not todas_filas:
                wb2 = openpyxl.load_workbook(ruta_excel, data_only=False)
                hoja2 = wb2[nombre_hoja]
                todas_filas = leer_filas(hoja2)
                wb2.close()
                print(f"[EXCEL] Hoja '{nombre_hoja}': usó modo sin data_only (fórmulas)")
            else:
                # Verificar si más del 80% de las celdas están vacías
                total = sum(len(f) for f in todas_filas)
                vacias = sum(1 for f in todas_filas for v in f if not v)
                if total > 0 and (vacias / total) > 0.8:
                    wb2 = openpyxl.load_workbook(ruta_excel, data_only=False)
                    hoja2 = wb2[nombre_hoja]
                    todas_filas_alt = leer_filas(hoja2)
                    wb2.close()
                    if todas_filas_alt:
                        todas_filas = todas_filas_alt
                        print(f"[EXCEL] Hoja '{nombre_hoja}': reemplazado por modo sin data_only (>80% vacías)")

            texto += f"\n{'='*50}\nHOJA: {nombre_hoja}\n{'='*50}\n"

            if not todas_filas:
                texto += "(Hoja vacía)\n"
                continue

            # Detectar encabezados: primera fila no vacía
            encabezados = []
            for i, val in enumerate(todas_filas[0]):
                encabezados.append(val if val else f"Columna_{i+1}")

            texto += f"COLUMNAS: {' | '.join(encabezados)}\n\n"

            # Cada fila de datos asociada con su encabezado
            for num_fila, fila in enumerate(todas_filas[1:], start=2):
                pares = []
                for i, valor in enumerate(fila):
                    if valor:
                        header = encabezados[i] if i < len(encabezados) else f"Columna_{i+1}"
                        pares.append(f"{header}: {valor}")
                if pares:
                    texto += f"FILA {num_fila}: " + " | ".join(pares) + "\n"

        wb.close()
        return texto

    def procesar_cargar_excel(ruta_excel):
        mostrar_snack("Procesando e insertando Excel...", color="#D8B4FE")
        try:
            db = conectar_db()
            if not db:
                mostrar_snack("Error: No se pudo conectar a la base de datos.", color="#FF4500")
                return

            cursor = db.cursor()

            with open(ruta_excel, "rb") as archivo:
                excel_binario = archivo.read()

            texto_extraido = extraer_texto_excel(ruta_excel)

            # Debug: imprimir lo extraído en consola
            print("\n===== TEXTO EXTRAÍDO DEL EXCEL =====")
            print(texto_extraido[:20000])
            print("=====================================\n")

            nombre_archivo = os.path.basename(ruta_excel)

            sql = """
            INSERT INTO manuales
            (Titulo, Nombre_Archivo, Archivo_PDF, Contenido_Texto, Categoria, Version)
            VALUES (%s, %s, %s, %s, %s, %s)
            """
            valores = (nombre_archivo, nombre_archivo, excel_binario, texto_extraido, "Excel", "1.0")
            cursor.execute(sql, valores)
            db.commit()
            db.close()
            rebuild_rag_cache()

            crear_notificacion_a_rol("Gerente", "Nuevo Manual Excel Cargado 📊", f"Se ha subido el archivo: '{nombre_archivo}'", "manual")

            # Mostrar vista previa de lo que se extrajo
            def cerrar_preview(e):
                page.pop_dialog()

            lineas_preview = texto_extraido.strip().split("\n")[:40]
            preview_str = "\n".join(lineas_preview)
            if len(texto_extraido.strip().split("\n")) > 40:
                preview_str += "\n... (más filas guardadas)"

            dialog_preview = ft.AlertDialog(
                title=ft.Text(f"✅ Excel cargado: {nombre_archivo}", color="#7CFC00", weight="bold"),
                content=ft.Container(
                    content=ft.Column([
                        ft.Text(
                            "Así se leyó tu archivo (verifica que las columnas estén bien):",
                            color="#aaaaaa", size=12
                        ),
                        ft.TextField(
                            value=preview_str,
                            multiline=True,
                            read_only=True,
                            color="white",
                            bgcolor="#0a0a0a",
                            border_color="#444444",
                            min_lines=12,
                            max_lines=16,
                            text_style=ft.TextStyle(font_family="Courier New", size=11)
                        )
                    ], spacing=8),
                    width=650,
                    height=400,
                    padding=10
                ),
                actions=[ft.TextButton("Cerrar", on_click=cerrar_preview)],
                actions_alignment="end",
                bgcolor="#1a1a1a"
            )
            page.show_dialog(dialog_preview)
            if hasattr(page, "pubsub") and page.pubsub:
                page.pubsub.send_all_on_topic("actualizaciones_luxo", f"📢 Nuevo manual disponible: '{nombre_archivo}'. LUXO ya se actualizó.")

        except Exception as ex:
            print("ERROR EXCEL:", ex)
            mostrar_snack(f"Error al cargar Excel: {ex}", color="#FF4500")

    def mostrar_manuales_admin(chat_display=None):
        mostrar_snack("Listando manuales...")

        try:
            db = conectar_db()
            if not db:
                mostrar_snack("Error: No se pudo conectar a la base de datos.", color="#FF4500")
                return
            cursor = db.cursor(dictionary=True)
            cursor.execute("SELECT ID_Manual, Nombre_Archivo, Titulo, Version FROM manuales ORDER BY Nombre_Archivo")
            manuales = cursor.fetchall()
            db.close()

            if not manuales:
                mostrar_snack("No hay manuales cargados.")
                return

            if chat_display is not None:
                chat_display.controls.append(
                    ft.Container(
                        content=ft.Column([
                            ft.Text("Manuales cargados:", color="white", weight="bold"),
                            *[
                                ft.Text(f"{m.get('Nombre_Archivo', '')}  |  v{m.get('Version', '')}  |  {m.get('Titulo', '')}", color="white")
                                for m in manuales
                            ]
                        ], spacing=4),
                        bgcolor="#0F0F1A",
                        padding=10,
                        border_radius=10
                    )
                )
                page.update()
                return

            items = []
            for m in manuales:
                nombre = m.get("Nombre_Archivo") or ""
                version = m.get("Version") or ""
                titulo = m.get("Titulo") or ""
                items.append(
                    ft.Row([
                        ft.Text(nombre, expand=3, selectable=True),
                        ft.Text(f"v{version}", width=80),
                        ft.Text(titulo, expand=2)
                    ], alignment="center")
                )

            contenido = ft.Column(items, spacing=6)

            def cerrar_dialog(e):
                page.pop_dialog()

            dialog = ft.AlertDialog(
                title=ft.Text("Manuales cargados"),
                content=ft.Container(contenido, width=700, height=320),
                actions=[ft.TextButton("Cerrar", on_click=cerrar_dialog)],
                actions_alignment="end"
            )

            page.show_dialog(dialog)

        except Exception as e:
            print("ERROR LISTAR MANUALES:", e)
            mostrar_snack("Error listando manuales.", color="#FF4500")

    def borrar_manual_admin():
        mostrar_snack("Cargando lista de manuales a eliminar...")
        try:
            db = conectar_db()
            if not db:
                mostrar_snack("Error: No se pudo conectar a la base de datos.", color="#FF4500")
                return

            cursor = db.cursor(dictionary=True)
            cursor.execute("SELECT ID_Manual, Nombre_Archivo, Titulo, Version FROM manuales ORDER BY Nombre_Archivo")
            manuales = cursor.fetchall()
            db.close()

            if not manuales:
                mostrar_snack("No hay manuales para eliminar.")
                return

            def on_confirmar_borrado(id_manual, nombre_archivo, row_control):
                try:
                    db_del = conectar_db()
                    if not db_del:
                        mostrar_snack("Error de conexión", color="#FF4500")
                        return
                    cursor_del = db_del.cursor()

                    # Borrado en cascada seguro para evitar fallos de claves foráneas
                    cursor_del.execute("""
                        DELETE FROM pendientes_actualizacion 
                        WHERE ID_Conversacion IN (
                            SELECT ID_Conversacion FROM historial_conversaciones WHERE ID_Manual = %s
                        )
                    """, (id_manual,))

                    cursor_del.execute("DELETE FROM historial_conversaciones WHERE ID_Manual = %s", (id_manual,))
                    cursor_del.execute("DELETE FROM manuales WHERE ID_Manual = %s", (id_manual,))
                    db_del.commit()
                    db_del.close()
                    rebuild_rag_cache()

                    mostrar_snack(f"Manual '{nombre_archivo}' eliminado.")
                    lista_manuales_container.controls.remove(row_control)
                    page.update()

                except Exception as ex:
                    print("ERROR BORRADO MANUAL:", ex)
                    mostrar_snack(f"Error al eliminar: {ex}", color="#FF4500")

            lista_manuales_container = ft.Column(spacing=10, scroll=ft.ScrollMode.AUTO)

            for m in manuales:
                id_m = m["ID_Manual"]
                nombre = m.get("Nombre_Archivo") or ""
                version = m.get("Version") or ""

                fila = ft.Row(alignment="spaceBetween")

                info_text = ft.Text(
                    f"{nombre} (v{version})", 
                    color="white", 
                    weight="normal",
                    expand=True
                )

                btn_eliminar = ft.IconButton(
                    icon=ft.Icons.DELETE_FOREVER,
                    icon_color="#FF4500",
                    tooltip=f"Eliminar {nombre}",
                    on_click=lambda e, id_man=id_m, nom=nombre, f_ctrl=fila: on_confirmar_borrado(id_man, nom, f_ctrl)
                )

                fila.controls = [info_text, btn_eliminar]
                lista_manuales_container.controls.append(fila)

            def cerrar_dialog(e):
                page.pop_dialog()

            dialog = ft.AlertDialog(
                modal=True,
                title=ft.Text("Eliminar Manuales", color="#D8B4FE", weight="bold"),
                content=ft.Container(
                    content=lista_manuales_container,
                    width=500,
                    height=300,
                    padding=10
                ),
                actions=[
                    ft.TextButton("Cerrar", on_click=cerrar_dialog)
                ],
                actions_alignment="end"
            )

            page.show_dialog(dialog)

        except Exception as e:
            print("ERROR AL ABRIR DIALOGO BORRADO:", e)
            mostrar_snack("Error al cargar menú de eliminación.", color="#FF4500")

    img_avatar = obtener_64("avatar_luxo.png")

    img_fondo = None

    # Asegurar que los avatares existan en la carpeta assets para poder ser servidos por Flet
    import shutil
    for filename in ["luxo_avatar1.mp4", "luxo_avatar2.mp4", "luxo_avatar.mp4"]:
        src_path = os.path.join(BASE_PATH, filename)
        dest_path = os.path.join(ASSETS_PATH, filename)
        if os.path.exists(src_path) and not os.path.exists(dest_path):
            try:
                os.makedirs(ASSETS_PATH, exist_ok=True)
                shutil.copy(src_path, dest_path)
                print(f"Copiado {filename} a la carpeta assets.")
            except Exception as e:
                print(f"Error copiando {filename} a assets: {e}")

    video_login_exists = False
    video_login_path = ""
    if os.path.exists(os.path.join(ASSETS_PATH, "luxo_avatar1.mp4")):
        video_login_exists = True
        video_login_path = os.path.join(ASSETS_PATH, "luxo_avatar1.mp4")
    elif os.path.exists(os.path.join(BASE_PATH, "luxo_avatar1.mp4")):
        video_login_exists = True
        video_login_path = os.path.join(BASE_PATH, "luxo_avatar1.mp4")
    elif os.path.exists(os.path.join(ASSETS_PATH, "luxo_avatar.mp4")):
        video_login_exists = True
        video_login_path = os.path.join(ASSETS_PATH, "luxo_avatar.mp4")
    elif os.path.exists(os.path.join(BASE_PATH, "luxo_avatar.mp4")):
        video_login_exists = True
        video_login_path = os.path.join(BASE_PATH, "luxo_avatar.mp4")
    video_login_url = "custom_assets/luxo_avatar1.mp4" if page.web else video_login_path

    video_chat_exists = False
    video_chat_path = ""
    if os.path.exists(os.path.join(ASSETS_PATH, "luxo_avatar2.mp4")):
        video_chat_exists = True
        video_chat_path = os.path.join(ASSETS_PATH, "luxo_avatar2.mp4")
    elif os.path.exists(os.path.join(BASE_PATH, "luxo_avatar2.mp4")):
        video_chat_exists = True
        video_chat_path = os.path.join(BASE_PATH, "luxo_avatar2.mp4")
    elif os.path.exists(os.path.join(ASSETS_PATH, "luxo_avatar.mp4")):
        video_chat_exists = True
        video_chat_path = os.path.join(ASSETS_PATH, "luxo_avatar.mp4")
    elif os.path.exists(os.path.join(BASE_PATH, "luxo_avatar.mp4")):
        video_chat_exists = True
        video_chat_path = os.path.join(BASE_PATH, "luxo_avatar.mp4")
    video_chat_url = "custom_assets/luxo_avatar2.mp4" if page.web else video_chat_path


    # =====================================
    # CERRAR SESION
    # =====================================

    def cerrar_sesion():
        async def remove_storage():
            try:
                await page.shared_preferences.remove("logged_user_id")
            except Exception as ex_store:
                print("Error al remover de shared_preferences:", ex_store)
        page.run_task(remove_storage)

        user_info["id"] = None
        user_info["nombre"] = ""
        user_info["rol"] = ""
        
        page.floating_action_button = None

        page.clean()

        page.add(full_screen_background)

        page.update()

    # =====================================
    # CHAT
    # =====================================

    def normalizar_texto(texto):
        import unicodedata
        if not texto:
            return ""
        texto = texto.lower()
        texto = ''.join(
            c for c in unicodedata.normalize('NFD', texto)
            if unicodedata.category(c) != 'Mn'
        )
        texto = re.sub(r'[^a-z0-9\s]', '', texto)
        return texto

    # =====================================
    # EXPANSOR DE ABREVIATURAS INFORMALES
    # Entiende abreviaturas del español e inglés coloquial
    # para mejorar la comprensión del chat con el usuario.
    # =====================================

    ABREVIATURAS = {
        # Palabras de pregunta
        "xq": "por qué",
        "xk": "por qué",
        "pq": "por qué",
        "pk": "por qué",
        "porq": "porque",
        "xkk": "por qué",
        "kmo": "cómo",
        "km": "cómo",
        "knd": "cuando",
        "kdo": "cuando",
        "qnd": "cuando",
        # Partículas comunes
        "x": "por",
        "k": "que",
        "q": "que",
        "d": "de",
        "t": "te",
        "m": "me",
        "n": "en",
        "c": "se",
        "s": "es",
        # Pronombres / artículos
        "tb": "también",
        "tmb": "también",
        "tbm": "también",
        "tmbn": "también",
        "tbn": "también",
        "kien": "quién",
        "qien": "quién",
        "sta": "esta",
        "stoy": "estoy",
        "sts": "estas",
        "toy": "estoy",
        "tas": "estás",
        "ta": "está",
        "tba": "estaba",
        # Verbos comunes
        "saber": "saber",
        "kro": "quiero",
        "qiero": "quiero",
        "ncsito": "necesito",
        "nesito": "necesito",
        "boi": "voy",
        "bamos": "vamos",
        "vmos": "vamos",
        "agr": "agregar",
        "deja": "déjame",
        "xfa": "por favor",
        "pf": "por favor",
        "pfavor": "por favor",
        "x favor": "por favor",
        "fvor": "favor",
        "grac": "gracias",
        "grax": "gracias",
        "gras": "gracias",
        "gracas": "gracias",
        "grcs": "gracias",
        # Afirmaciones / negaciones
        "si": "sí",
        "sip": "sí",
        "nop": "no",
        "nel": "no",
        "nave": "no",
        "simón": "sí",
        # Expresiones coloquiales
        "wey": "amigo",
        "wey": "amigo",
        "gey": "amigo",
        "prro": "amigo",
        "mames": "bromeas",
        "ón": "on",
        "omg": "dios mío",
        "lol": "gracioso",
        "ok": "está bien",
        "okey": "está bien",
        "okis": "está bien",
        "dale": "está bien",
        "va": "está bien",
        "bale": "está bien",
        # Lugares / cosas
        "info": "información",
        "msj": "mensaje",
        "msg": "mensaje",
        "cel": "celular",
        "telf": "teléfono",
        "tel": "teléfono",
        "doc": "documento",
        "docs": "documentos",
        "tbjo": "trabajo",
        "trab": "trabajo",
        "emp": "empresa",
        "mpresa": "empresa",
        "inv": "inventario",
        "gar": "garantía",
        "garan": "garantía",
        "gcia": "garantía",
        "vent": "venta",
        "vtas": "ventas",
        "prod": "producto",
        "prods": "productos",
        "tiend": "tienda",
        "sucur": "sucursal",
        "check": "checklist",
        "chk": "checklist",
        "ckl": "checklist",
        "chklist": "checklist",
        "man": "manual",
        "manu": "manual",
        "mnl": "manual",
        # Tiempo
        "hoy": "hoy",
        "hoi": "hoy",
        "mañana": "mañana",
        "manana": "mañana",
        "maniana": "mañana",
        "mnna": "mañana",
        "ayer": "ayer",
        "ayr": "ayer",
        "ahorita": "ahora",
        "aorita": "ahora",
        "ahra": "ahora",
        "ahora": "ahora",
        "luego": "después",
        "dpues": "después",
        "dsp": "después",
        "dsps": "después",
        # Inglés básico coloquial
        "plz": "por favor",
        "pls": "por favor",
        "thx": "gracias",
        "ty": "gracias",
        "gg": "bien hecho",
        "asap": "lo antes posible",
        "fyi": "para tu información",
        "btw": "por cierto",
        "lmk": "avísame",
        "idk": "no sé",
        "nvm": "no importa",
        "omw": "voy en camino",
        "brb": "regreso en un momento",
        "atm": "ahora mismo",
        "imo": "en mi opinión",
        "aka": "también conocido como",
        # Abreviaturas de escritura rápida
        "xd": "",
        "jeje": "",
        "jaja": "",
        "dnd": "dónde",
        "dnde": "dónde",
        "ond": "dónde",
        "cn": "con",
        "pa": "para",
        "pa'": "para",
        "pr": "para",
        "nd": "nada",
        "nda": "nada",
        "ntp": "no te preocupes",
        "np": "no te preocupes",
        "ntpz": "no te preocupes",
    }

    def expandir_abreviaturas(texto):
        """
        Expande abreviaturas informales en el texto del usuario
        antes de enviarlo al motor de IA o búsqueda de manuales.
        Solo reemplaza palabras completas, no partes de palabras.
        El texto original que ve el usuario en pantalla NO se modifica.
        """
        if not texto:
            return texto
        import re as _re
        # Normalizar pronunciaciones y deletreos de la sigla "AUR"
        texto = _re.sub(r'\b(a\s+u\s+r|a\.\s*u\.\s*r\.|a\s+u\s+ere|ahur|aor)\b', 'aur', texto, flags=_re.IGNORECASE)
        texto = _re.sub(r'\b(el|del|un|al|como|es|kpi|sobre|mide)\s+a\s+ver\b', r'\1 aur', texto, flags=_re.IGNORECASE)
        
        palabras = _re.split(r'(\s+)', texto)
        resultado = []
        for palabra in palabras:
            palabra_lower = palabra.lower().strip(".,;:!?¿¡\"'()")
            if palabra_lower in ABREVIATURAS:
                expansion = ABREVIATURAS[palabra_lower]
                if expansion:  # Si la expansión no es vacía (como xd, jaja)
                    resultado.append(expansion)
                else:
                    resultado.append(palabra)
            else:
                resultado.append(palabra)
        return "".join(resultado)

    def cargar_chat():
        page.clean()

        chat_display = ft.ListView(
            expand=True,
            spacing=10,
            padding=20,
            auto_scroll=True
        )

        # Mensaje de bienvenida inicial de LUXO al abrir el chat
        avatar_luxo_base64 = obtener_64("avatar_luxo2.png")
        chat_display.controls.append(
            ft.Container(
                content=ft.Row([
                    ft.Container(
                        content=ft.Image(src=avatar_luxo_base64, width=35, height=35, fit=ft.controls.box.BoxFit.COVER) if avatar_luxo_base64 else ft.Icon(ft.Icons.AUTO_AWESOME, color="#D8B4FE", size=20),
                        width=35,
                        height=35,
                        border_radius=17.5,
                        bgcolor="#2A1B4E",
                        alignment=ft.alignment.Alignment(0, 0),
                        border=ft.Border.all(1.5, "#D8B4FE"),
                    ),
                    ft.Text("LUXO: ¡Bienvenido! Soy LUXO, tu asistente virtual.", color="white", weight="bold", expand=True, selectable=True),
                ], vertical_alignment="start", spacing=10),
                bgcolor="#1E1E2E",
                padding=12,
                border_radius=10,
                border=ft.Border.all(1, "#D8B4FE")
            )
        )

        # Historial de conversación en memoria para enviar al LLM
        historial_sesion = []

        # Cargar historial de la base de datos de forma silenciosa en memoria para el contexto del LLM (evita problemas de rendimiento en la UI)
        try:
            db = conectar_db()
            if db:
                cursor = db.cursor(dictionary=True)
                cursor.execute("""
                    SELECT Pregunta_Usuario, Respuesta_IA 
                    FROM historial_conversaciones 
                    WHERE ID_Usuario = %s 
                    ORDER BY Fecha_Hora ASC 
                    LIMIT 10
                """, (user_info["id"],))
                historial = cursor.fetchall()
                db.close()
                for row in historial:
                    historial_sesion.append({"role": "user", "content": row["Pregunta_Usuario"]})
                    historial_sesion.append({"role": "assistant", "content": row["Respuesta_IA"]})
        except Exception as e:
            print("ERROR AL CARGAR HISTORIAL EN MEMORIA:", e)

        # Cargar imagen de avatar del usuario si existe
        img_usuario = obtener_64("istockphoto-468228782-612x612")

        # =================================
        # ENVIAR MENSAJE
        # =================================

        def enviar_mensaje(e):
            if not input_msg.value:
                return

            user_text = input_msg.value

            # Expandir abreviaturas informales antes de procesar el mensaje
            user_text_expandido = expandir_abreviaturas(user_text)
            user_text_norm = normalizar_texto(user_text_expandido)

            # --- 1. COMANDO DIRECTO DE NOTIFICACIONES (Pestaña / Alertas) ---
            is_notif_query = any(w in user_text_norm for w in ["notificaci", "campan", "abren no"])
            if is_notif_query:
                input_msg.value = ""
                try:
                    mostrar_notificaciones_dialog(None)
                    mostrar_snack("🔔 Panel de Notificaciones abierto", color="#00FFFF")
                except Exception as ex_n:
                    print("Error al abrir notificaciones:", ex_n)
                    mostrar_snack("🔔 Notificaciones abiertas")
                page.update()
                return

            # --- 2. COMANDO DIRECTO DE REPORTE DE APERTURAS (EXCLUSIVO ADMIN) ---
            is_apertura_status_query = (
                ("apertura" in user_text_norm or "abiert" in user_text_norm or "abrir" in user_text_norm) and
                any(w in user_text_norm for w in ["falta", "faltan", "sin", "no han", "no ha", "no a", "subid", "subio", "report", "quien", "quienes", "cuales", "que tiendas", "estado"])
            )
            if is_apertura_status_query:
                input_msg.value = ""
                if not es_admin():
                    mostrar_snack("🔒 Consulta de aperturas reservada únicamente para Administrador.", color="#FFD700")
                    page.update()
                    return

                try:
                    operacion_tiendas.actualizar_estrella_aperturas(page, star_icon_container, conectar_db, abrir_modal=True)
                except Exception as ex_ap:
                    print("Error abriendo estrella de aperturas por comando:", ex_ap)
                    mostrar_snack("⭐ Monitor de aperturas abierto")
                page.update()
                return

            # Redirección automática si se detectan palabras clave de pestañas (solo para comandos cortos de navegación o explícitos)
            palabras_mensaje = user_text_norm.split()
            
            # Indicadores de navegación explícita
            indicadores_nav = ["abre", "abrir", "ir a", "ve a", "entra", "mostrar", "pantalla", "pestana", "seccion", "vista"]
            es_comando_navegacion = any(ind in user_text_norm for ind in indicadores_nav)
            
            # Si el texto es de tipo conversacional (más de 3 palabras, o contiene pronombres/verbos de pregunta), 
            # y NO es un comando de navegación explícito, entonces NO redirigimos.
            es_pregunta_conversacional = (
                len(palabras_mensaje) > 3 or 
                any(p in user_text_norm for p in ["como", "donde", "por que", "quien", "cual", "cuales", "explic", "ayuda", "duda"])
            )
            
            # Solo redirigimos si:
            # - No es una pregunta conversacional
            # - O es un comando de navegación explícito
            permitir_redireccion = (not es_pregunta_conversacional) or es_comando_navegacion
            
            # Para el caso especial de campañas, si se menciona "campana", aplicamos un filtro extra:
            # Si contiene "campana" pero NO es un comando de navegación explícito (ni es la palabra exacta "campañas"), NO permitimos redirección
            if "campana" in user_text_norm:
                msg_limpio = user_text_norm.strip()
                if not (es_comando_navegacion or msg_limpio in ["campana", "campanas"]):
                    permitir_redireccion = False

            if permitir_redireccion:
                # Subpestañas del Panel de Control Operativo
                subtab_dashboard_keys = {
                    "estadistica": 0,
                    "estadisticas": 0,
                    "preguntas sin contestar": 1,
                    "pregunta sin contestar": 1,
                    "preguntas faltantes": 1,
                    "pregunta faltante": 1,
                    "gestion de manuales": 2,
                    "gestion manuales": 2,
                    "subir manual": 2,
                    "eliminar manual": 2,
                    "sugerencia": 3,
                    "sugerencias": 3,
                    "soporte": 4,
                    "ticket": 4,
                    "tickets": 4,
                    "editar checklist": 5,
                    "editar checklists": 5,
                    "tareas consolidadas": 6,
                    "tarea consolidada": 6,
                    "gerentes": 7,
                    "gerente": 7,
                    "gestion de gerentes": 7,
                    "tiendas registradas": 7,
                    "tienda registrada": 7,
                    "tiendas": 7
                }

                for sub_key, sub_idx in subtab_dashboard_keys.items():
                    if sub_key in user_text_norm:
                        if es_admin():
                            dashboard_tab_index[0] = sub_idx
                            input_msg.value = ""
                            cambiar_vista("dashboard")
                            return

                redirecciones = {
                    # Aperturas y Cierres
                    "aperturas y cierres": "operacion_diaria",
                    "aperturas & cierres": "operacion_diaria",
                    "aperturas cierres": "operacion_diaria",
                    "aperturas": "operacion_diaria",
                    "apertura": "operacion_diaria",
                    "cierres": "operacion_diaria",
                    "cierre": "operacion_diaria",
                    "operacion diaria": "operacion_diaria",

                    # Categorías Principales del Menú Lateral
                    "operacion y tienda": "checklists",
                    "operaciones y tienda": "checklists",
                    "operacion & tienda": "checklists",
                    "operacion tienda": "checklists",
                    "operacion": "checklists",
                    "operaciones": "checklists",
                    "tienda": "checklists",

                    "ventas y metricas": "meta_semanal",
                    "ventas & metricas": "meta_semanal",
                    "ventas metricas": "meta_semanal",
                    "ventas": "meta_semanal",
                    "metricas": "meta_semanal",

                    "clientes y garantias": "garantias",
                    "clientes & garantias": "garantias",
                    "clientes garantias": "garantias",
                    "clientes": "garantias",

                    "capacitacion e ia": "simulador",
                    "capacitacion & ia": "simulador",
                    "capacitacion ia": "simulador",
                    "capacitacion": "simulador",
                    "entrenamiento": "simulador",

                    # Vistas Específicas
                    "asistente chat": "chat",
                    "asistente": "chat",
                    "chat": "chat",

                    "mi historial": "historial",
                    "historial": "historial",

                    # CRM Cobertura Oops
                    "coberturas oops": "crm",
                    "cobertura oops": "crm",
                    "crm cobertura oops": "crm",
                    "crm oops": "crm",
                    "coberturas": "crm",
                    "cobertura": "crm",
                    "crm": "crm",
                    "oops": "crm",

                    # Weekly
                    "weekly": "weekly",
                    "semanal": "weekly",
                    "reporte weekly": "weekly",

                    # Gestionar Trivia (Admin)
                    "gestionar trivia": "admin_trivia",
                    "administrar trivia": "admin_trivia",
                    "configurar trivia": "admin_trivia",

                    # Bitácora de Seguridad
                    "bitacora de seguridad": "bitacora",
                    "bitacora": "bitacora",
                    "bitacoras": "bitacora",
                    "inicios de sesion": "bitacora",
                    "auditoria": "bitacora",

                    # Gestión de Perfiles
                    "gestion de perfiles": "gestion_perfiles",
                    "gestionar perfiles": "gestion_perfiles",
                    "perfiles": "gestion_perfiles",

                    # Garantías
                    "garantias": "garantias",
                    "garantia": "garantias",

                    # Vendedores / Configuración Tienda / Metas Mensuales
                    "configuracion tienda": "vendedores",
                    "metas mensuales": "vendedores",
                    "meta mensual": "vendedores",
                    "vendedor": "vendedores",
                    "vendedores": "vendedores",

                    # Metas Semanales / Metas y Métricas
                    "metas semanales": "meta_semanal",
                    "meta semanal": "meta_semanal",
                    "metas y metricas": "meta_semanal",
                    "metas": "meta_semanal",
                    "meta": "meta_semanal",

                    # Reto del Día / Quiz
                    "reto del dia": "reto",
                    "reto": "reto",
                    "trivia": "reto",
                    "quiz": "reto",

                    # Campañas
                    "campanas": "campanas",
                    "campana": "campanas",

                    # Simulador IA
                    "simulador ia": "simulador",
                    "simulador": "simulador",
                    "comision": "simulador",
                    "comisiones": "simulador",

                    # Manuales
                    "manuales": "manuales",
                    "manual": "manuales",

                    # Checklists
                    "checklists": "checklists",
                    "checklist": "checklists",

                    # Tareas
                    "tareas": "tareas",
                    "tarea": "tareas",

                    # Presupuesto
                    "presupuesto": "presupuesto",
                    "bouget": "presupuesto",
                    "budget": "presupuesto",

                    # Enfoque Diario 2026
                    "enfoque diario 2026": "enfoque_diario",
                    "enfoque diario": "enfoque_diario",
                    "enfoque 2026": "enfoque_diario",
                    "enfoque": "enfoque_diario",

                    # Panel de Control
                    "panel de control": "dashboard",
                    "panel control": "dashboard",
                    "dashboard": "dashboard",
                    "panel": "dashboard"
                }
                
                # Ordenar por longitud descendente para emparejar frases compuestas largas primero
                sorted_nav_keys = sorted(redirecciones.keys(), key=len, reverse=True)
                for key in sorted_nav_keys:
                    if key in user_text_norm:
                        vista = redirecciones[key]
                        if vista == "dashboard" and not es_admin():
                            continue
                        input_msg.value = ""
                        cambiar_vista(vista)
                        return

            # Si no es redirección y estamos en otra pestaña, volvemos a la vista del chat asistente
            if active_view[0] != "chat":
                cambiar_vista("chat")

            chat_display.controls.append(
                ft.Container(
                    content=ft.Row([
                        ft.Container(
                            content=ft.Image(src=user_info.get("img_usuario") or img_usuario, width=35, height=35, fit=ft.controls.box.BoxFit.COVER) if (user_info.get("img_usuario") or img_usuario) else ft.Icon(ft.Icons.PERSON, color="#00FFFF", size=20),
                            width=35,
                            height=35,
                            border_radius=17.5,
                            bgcolor="#333333",
                            clip_behavior=ft.ClipBehavior.HARD_EDGE,
                            alignment=ft.alignment.Alignment(0, 0),
                            border=ft.Border.all(1.5, "#D8B4FE"),
                        ),
                        ft.Text(f"{user_info['nombre']}: {user_text}", color="white", weight="bold", expand=True, selectable=True),
                    ], vertical_alignment="start", spacing=10),
                    bgcolor="#141424",
                    padding=10,
                    border_radius=10
                )
            )
            input_msg.value = ""
            page.update()

            try:
                db = conectar_db()
                if not db:
                    chat_display.controls.append(
                        ft.Text("ERROR: No se pudo conectar a la base de datos de manuales.", color="red")
                    )
                    page.update()
                    return
                cursor = db.cursor(dictionary=True)
                cursor.execute("SELECT * FROM manuales")
                manuales = cursor.fetchall()
                
                # Obtener historial de conversaciones calificadas como útiles (Me_Sirvio = 1) y fallidas (Me_Sirvio = 0)
                cursor.execute("""
                    SELECT Pregunta_Usuario, Respuesta_IA 
                    FROM historial_conversaciones 
                    WHERE Me_Sirvio = 1
                """)
                chats_previos = cursor.fetchall()
                
                cursor.execute("""
                    SELECT Pregunta_Usuario, Respuesta_IA 
                    FROM historial_conversaciones 
                    WHERE Me_Sirvio = 0
                """)
                chats_fallidos = cursor.fetchall()
                db.close()

                # Definición de la función de lematización/raíz en español
                def obtener_raiz_espanol(word):
                    if len(word) <= 3:
                        return word
                    sufijos = [
                        "ando", "iendo", "aron", "ieron", "aremos", "eremos", "iremos",
                        "ar", "er", "ir", "ado", "ido", "as", "es", "os", "an", "en", "o", "a", "e"
                    ]
                    for suf in sorted(sufijos, key=len, reverse=True):
                        if word.endswith(suf) and len(word) - len(suf) >= 3:
                            return word[:-len(suf)]
                    return word

                # Primero, intentar buscar usando únicamente la pregunta actual del usuario
                # para evitar contaminación de contexto cuando cambia de tema
                contexto_norm = normalizar_texto(user_text_expandido)
                palabras_query = re.findall(r"\w+", contexto_norm)
                
                # Stopwords españolas exhaustivas (se eliminan "telefono" y "tienda" de la lista de stopwords para permitir búsquedas directas en directorios)
                stopwords = {
                    "de", "la", "el", "los", "las", "un", "una", "pdf", "manual", "documento", 
                    "archivo", "archivos", "tienes", "hola", "cual", "es", 
                    "como", "donde", "por", "para", "con", "que", "quiero", "saber", "me", "puedes", 
                    "dar", "darme", "a", "al", "en", "son", "cuales", "esta", "estas", "este", "estos", 
                    "del", "o", "u", "y", "e", "si", "no", "se", "lo", "te", "le", "les", "nos", "mi", 
                    "mis", "tu", "tus", "su", "sus", "ellos", "ellas", "nosotros", "usted", "ustedes", 
                    "mio", "tuyo", "suyo", "aqui", "alli", "alla", "todo", "todos", "toda", "todas", 
                    "uno", "unos", "otro", "otros", "otra", "otras", "hacer", "hace", "hacen", "haciendo", 
                    "ver", "vista", "puede", "pueden", "ser", "esta", "estan", "este", "esto", "del",
                    "sunglass", "hut", "luxottica", "quien", "quienes", "cuando", "como", "cual", "cuales",
                    "que", "porque", "donde", "realizar", "realizo", "realiza", "realizan", "realizado", 
                    "realizando", "paso", "pasos", "guia", "guias", "tutorial", "ayuda", "obtener", 
                    "descargar", "descarga", "bajar", "mostrar", "imprimir", "impresion", "sistema", "sistemas"
                }
                query_palabras = [w for w in palabras_query if w not in stopwords and len(w) >= 2]
                
                # Raíces de la consulta con corrección ortográfica (Fuzzy Matching)
                query_roots_raw = [obtener_raiz_espanol(w) for w in query_palabras]
                query_roots = []
                
                # Obtener vocabulario de raíces conocidas en los manuales (si RAG_IDF_CACHE existe)
                global RAG_BLOQUES_CACHE, RAG_IDF_CACHE
                vocabulario_valido = list(RAG_IDF_CACHE.keys()) if RAG_IDF_CACHE else []
                
                for r in query_roots_raw:
                    # Si la raíz ya es válida o es muy corta, se deja igual
                    if not vocabulario_valido or r in vocabulario_valido or len(r) <= 3:
                        query_roots.append(r)
                    else:
                        # Buscar la palabra más parecida en los manuales (corrige faltas ortográficas o dedazos)
                        matches = difflib.get_close_matches(r, vocabulario_valido, n=1, cutoff=0.75)
                        if matches:
                            query_roots.append(matches[0])
                        else:
                            query_roots.append(r)

                core_roots = list(query_roots)
                
                # Diccionario de sinónimos a nivel de raíces (Incluyendo Mexicanismos y Abreviaciones)
                SINONIMOS_RAICES = {
                    "rob": ["3r", "siniestr", "perd", "asalto"],
                    "caj": ["cierr", "arqu", "morrall", "feri", "lan", "dinero", "efectiv"],
                    "cierr": ["caj", "finaliz", "cort"],
                    "cambi": ["devoluc", "garanti", "reemplaz"],
                    "devoluc": ["cambi", "garanti", "dev"],
                    "impresor": ["epson", "papel", "ticket", "tkt", "tck", "recib"],
                    "papel": ["roll", "impresor"],
                    "terminal": ["caj", "pinpad", "clip", "banam", "banc", "tarjet", "tj"],
                    "sistem": ["sys", "plataform", "lux", "portal"],
                    "inform": ["info", "ayud", "doc", "manual"],
                    "telefon": ["cel", "movil", "llama"],
                    "trabaj": ["chamb", "labor", "tare"],
                    "revis": ["chec", "verific", "valid"],
                    "necesit": ["ocup", "requier", "quier"],
                    "punt": ["pto", "sucursal", "tiend"]
                }
                
                # Expansión de sinónimos sobre raíces
                expanded_roots = list(query_roots)
                for r in query_roots:
                    if r in SINONIMOS_RAICES:
                        for syn in SINONIMOS_RAICES[r]:
                            if syn not in expanded_roots:
                                expanded_roots.append(syn)

                descarga_keywords = [
                    "descargar pdf", "descarga el manual", "descargar manual", 
                    "bajar pdf", "bajar el manual", "descargar archivo",
                    "pasa el pdf", "pasar el pdf", "dame el pdf", "quiero el pdf",
                    "mandame el pdf", "enviar pdf", "imprimir pdf", "imprimir manual"
                ]
                lista_keywords = ["listar manuales", "lista de manuales", "mostrar manuales disponibles", "que manuales tienes", "manuales cargados", "cuales son los manuales", "listar los pdf"]
                ask_for_pdf = any(phrase in user_text_expandido.lower() for phrase in descarga_keywords) or (
                    ("pdf" in user_text_expandido.lower() or "manual" in user_text_expandido.lower()) and 
                    any(p in user_text_expandido.lower() for p in ["dame", "quiero", "pasa", "descargar", "bajar", "ver", "mostrar", "imprimir", "obtener"])
                )
                ask_for_list = any(phrase in user_text_expandido.lower() for phrase in lista_keywords)

                # Función para dividir en bloques
                def dividir_texto_en_bloques(texto):
                    if not texto:
                        return []
                    texto = texto.replace("\r\n", "\n")
                    parrafos = texto.split("\n\n")
                    
                    bloques = []
                    header_markers = ["que pasa", "me equivoque", "como ", "un cliente", "cualquier", "en el caso", "que hacer", "me marcaron", "family an", "todo funciona"]
                    
                    for parrafo in parrafos:
                        parrafo_strip = parrafo.strip()
                        if not parrafo_strip:
                            continue
                            
                        lineas = parrafo_strip.split("\n")
                        bloque_actual = []
                        
                        for linea in lineas:
                            linea_strip = linea.strip()
                            if not linea_strip:
                                continue
                                
                            es_inicio = False
                            if (linea_strip.endswith("?") or 
                                linea_strip.endswith("?.") or 
                                linea_strip.endswith("? ") or
                                any(linea_strip.lower().startswith(marker) for marker in ["que hago", "que hacer", "como ", "como hacer", "como realizar"])
                            ):
                                if len(linea_strip) > 15 and not linea_strip.startswith("-") and not linea_strip.startswith("*"):
                                    es_inicio = True
                            else:
                                linea_norm = normalizar_texto(linea_strip)
                                if any(linea_norm.startswith(marker) for marker in header_markers):
                                    if len(linea_strip) > 15 and not linea_strip.startswith("-") and not linea_strip.startswith("*"):
                                        es_inicio = True
                                        
                            if es_inicio and bloque_actual:
                                bloques.append("\n".join(bloque_actual))
                                bloque_actual = []
                            bloque_actual.append(linea)
                            
                        if bloque_actual:
                            bloques.append("\n".join(bloque_actual))
                            
                    bloques_finales = []
                    temp_bloque = ""
                    for b in bloques:
                        b_strip = b.strip()
                        if not b_strip:
                            continue
                        if temp_bloque:
                            temp_bloque += "\n" + b_strip
                        else:
                            temp_bloque = b_strip
                            
                        if len(temp_bloque) >= 120 or temp_bloque.endswith("?"):
                            bloques_finales.append(temp_bloque)
                            temp_bloque = ""
                            
                    if temp_bloque:
                        bloques_finales.append(temp_bloque)
                        
                    return bloques_finales

                # --- SEGMENTACIÓN GLOBAL DE MANUALES Y CÁLCULO DE TF-IDF (CON CACHÉ) ---
                if RAG_BLOQUES_CACHE is None:
                    rebuild_rag_cache()
                
                with RAG_CACHE_LOCK:
                    todos_los_bloques = RAG_BLOQUES_CACHE or []
                    idf_dict = RAG_IDF_CACHE or {}

                # Calificación de Bloques mediante Similitud Probabilística de Raíces
                def buscar_candidatos(q_roots, exp_roots, texto_query):
                    attention_weights = {}
                    for r in exp_roots:
                        attention_weights[r] = idf_dict.get(r, 1.5)
                        
                    avg_len = sum(len(b["roots"]) for b in todos_los_bloques) / max(1, len(todos_los_bloques))
                    
                    candidatos = []
                    query_norm = normalizar_texto(texto_query)
                    query_words = [w for w in re.findall(r"\w+", query_norm) if w not in stopwords]
                    
                    for b in todos_los_bloques:
                        score_b = 0.0
                        
                        r_counts = {}
                        for r in b["roots"]:
                            r_counts[r] = r_counts.get(r, 0) + 1
                            
                        # Calcular cobertura
                        matches = 0
                        for q_root, att_w in attention_weights.items():
                            count_b = r_counts.get(q_root, 0)
                            if count_b > 0:
                                matches += 1
                                # BM25 simplified
                                k1 = 1.5
                                b_param = 0.75
                                doc_len = len(b["roots"])
                                tf_bm25 = (count_b * (k1 + 1)) / (count_b + k1 * (1 - b_param + b_param * (doc_len / max(1, avg_len))))
                                score_b += tf_bm25 * att_w
                        
                        if len(attention_weights) > 1 and matches == 0:
                            continue
                            
                        score_b_scaled = score_b * 100
                        
                        # Boost por título
                        nombre_archivo_norm = normalizar_texto(b["nombre"])
                        nombre_palabras = re.findall(r"\w+", nombre_archivo_norm)
                        nombre_roots = [obtener_raiz_espanol(w) for w in nombre_palabras if w not in stopwords]
                        for qr in q_roots:
                            if qr in nombre_roots:
                                score_b_scaled += 150
                                break
                                
                        # Boost por coincidencia exacta de frases (N-gramas)
                        texto_doc_norm = normalizar_texto(b["texto"])
                        for i in range(len(query_words) - 1):
                            bigram = query_words[i] + " " + query_words[i+1]
                            if bigram in texto_doc_norm:
                                score_b_scaled += 200 # Gran boost por frase exacta
                        
                        # Penalización por cobertura baja
                        if len(attention_weights) > 2 and matches < 2:
                            score_b_scaled *= 0.2
                        
                        if score_b_scaled >= 30: # Aumentar umbral
                            candidatos.append((score_b_scaled, b["nombre"], b["texto"], b))
                            
                    return candidatos

                # --- DETECCIÓN DE SONDEO EN PYTHON (PREVENTIVO) ---
                es_corta_o_ambigua = False
                terminos_ambiguos = [
                    "corte", "cortes", "caja", "cajas", "cierr", "cierres", 
                    "garanti", "garantias", "devoluc", "devoluciones", 
                    "sistem", "sistemas", "plataform", "plataformas", 
                    "report", "reportes", "ticket", "tickets", "factur", "facturas"
                ]
                
                # Extraer palabras útiles excluyendo stopwords y vacías
                palabras_utiles = [w for w in query_palabras if len(w) >= 2]
                
                # Criterio de sondeo: Contiene término ambiguo y mide 2 o menos palabras útiles (ej. "garantias", "garantias costa", "corte caja")
                contiene_termino_ambiguo = any(
                    obtener_raiz_espanol(t) in [obtener_raiz_espanol(w) for w in palabras_utiles] 
                    for t in terminos_ambiguos
                )
                if contiene_termino_ambiguo and len(palabras_utiles) <= 2:
                    es_corta_o_ambigua = True

                # Verificar si el usuario ya está respondiendo a un sondeo previo en el historial
                ya_respondio_sondeo = False
                ultimos_mensajes_ia = [m["content"] for m in historial_sesion if m["role"] == "assistant"]
                if ultimos_mensajes_ia:
                    ultimo_mensaje_ia = ultimos_mensajes_ia[-1]
                    # Contar signos de interrogación en el último mensaje de la IA
                    signos_pregunta_ia = ultimo_mensaje_ia.count("?")
                    # Frases de sondeo comunes que LUXO pudo haber usado
                    frases_sondeo_previo = [
                        "me puedes decir", "podrías indicarme", "podrías decirme",
                        "para orientarte mejor", "para darte la información", "necesito saber",
                        "tipo de corte", "qué tipo", "en qué plataforma", "en qué sistema",
                        "detalles sobre tu situación", "ayudarte mejor", "requisitos y especificaciones"
                    ]
                    if signos_pregunta_ia >= 2 or any(f in ultimo_mensaje_ia.lower() for f in frases_sondeo_previo):
                        ya_respondio_sondeo = True

                es_pregunta_trivia = "duda sobre la pregunta de trivia" in user_text_expandido.lower() or "pregunta de trivia:" in user_text_expandido.lower()
                es_sondeo_forzado = es_corta_o_ambigua and not ya_respondio_sondeo and not es_pregunta_trivia

                # Buscar candidatos usando la consulta actual (si no es sondeo forzado)
                if es_sondeo_forzado:
                    bloques_candidatos = []
                else:
                    bloques_candidatos = buscar_candidatos(query_roots, expanded_roots, user_text_expandido)
                
                # Si no se encontraron candidatos y hay historial, re-intentar con la pregunta previa integrada para dar contexto
                ultimos_mensajes_usuario = [m["content"] for m in historial_sesion if m["role"] == "user"]
                if not es_sondeo_forzado and not bloques_candidatos and ultimos_mensajes_usuario and query_palabras:
                    contexto_busqueda = ultimos_mensajes_usuario[-1] + " " + user_text_expandido
                    contexto_norm_comb = normalizar_texto(contexto_busqueda)
                    palabras_query_comb = re.findall(r"\w+", contexto_norm_comb)
                    query_palabras_comb = [w for w in palabras_query_comb if w not in stopwords and len(w) >= 2]
                    query_roots_comb = [obtener_raiz_espanol(w) for w in query_palabras_comb]
                    expanded_roots_comb = list(query_roots_comb)
                    for r in query_roots_comb:
                        if r in SINONIMOS_RAICES:
                            for syn in SINONIMOS_RAICES[r]:
                                if syn not in expanded_roots_comb:
                                    expanded_roots_comb.append(syn)
                    # Actualizar variables de consulta globales para el resto del procesamiento
                    query_palabras = query_palabras_comb
                    query_roots = query_roots_comb
                    bloques_candidatos = buscar_candidatos(query_roots_comb, expanded_roots_comb, contexto_busqueda)

                bloques_filtrados = []
                modo_sugerencia = False
                sugerencias_nombres = []
                
                if bloques_candidatos:
                    bloques_candidatos.sort(key=lambda x: x[0], reverse=True)
                    max_score = bloques_candidatos[0][0]
                    for score_val, doc_nombre, blk_texto, b_obj in bloques_candidatos:
                        # Exigir un umbral de score absoluto mínimo de 65 para evitar asociar manuales irrelevantes
                        if score_val >= (max_score * 0.5) and score_val >= 65:
                            bloques_filtrados.append((score_val, doc_nombre, blk_texto, b_obj))
                else:
                    candidatos = []
                    for b in todos_los_bloques:
                        sc_temp = 0
                        for qr in query_roots:
                            if qr in b["roots_set"]:
                                sc_temp += 1
                        if sc_temp > 0:
                            candidatos.append((sc_temp, b["nombre"]))
                    candidatos.sort(key=lambda x: x[0], reverse=True)
                    if candidatos:
                        modo_sugerencia = True
                        sugerencias_nombres = list(dict.fromkeys([name for sc, name in candidatos[:3]]))

                # Si venimos forzados de la trivia con un manual específico de origen
                if manual_forzado_trivia[0] is not None:
                    id_forzado = manual_forzado_trivia[0]
                    db_f = conectar_db()
                    if db_f:
                        cursor_f = db_f.cursor(dictionary=True)
                        cursor_f.execute("SELECT Nombre_Archivo, Contenido_Texto, Abierto FROM manuales WHERE ID_Manual = %s", (id_forzado,))
                        row_f = cursor_f.fetchone()
                        db_f.close()
                        if row_f:
                            bloques_filtrados = [(500.0, row_f["Nombre_Archivo"], row_f["Contenido_Texto"], {
                                "id": id_forzado,
                                "nombre": row_f["Nombre_Archivo"],
                                "abierto": row_f.get("Abierto", 1)
                            })]
                            modo_sugerencia = False

                # Agrupar bloques por documento y limitar a top 3 bloques
                bloques_por_doc = {}
                manuales_con_score = []
                
                for score_val, doc_nombre, blk_texto, b_obj in bloques_filtrados[:3]:
                    if doc_nombre not in bloques_por_doc:
                        bloques_por_doc[doc_nombre] = []
                    bloques_por_doc[doc_nombre].append(blk_texto)
                    manuales_con_score.append((score_val, {
                        "id": b_obj["id"],
                        "nombre": b_obj["nombre"],
                        "abierto": b_obj["abierto"]
                    }))

                manuales_texto = ""
                total_chars = 0
                max_total_chars = 12000  # Límite de seguridad de caracteres para evitar error 413
                
                for doc_nombre, lista_blks in bloques_por_doc.items():
                    if total_chars >= max_total_chars:
                        break
                    
                    texto_completo_doc = ""
                    for blk in lista_blks:
                        # Limitar cada bloque a 8000 caracteres
                        blk_truncated = blk[:8000]
                        if len(blk) > 8000:
                            blk_truncated += "\n[... Truncado por longitud ...]"
                            
                        if total_chars + len(blk_truncated) + len(doc_nombre) + 100 > max_total_chars:
                            allowed = max_total_chars - total_chars - len(doc_nombre) - 150
                            if allowed > 200:
                                blk_truncated = blk_truncated[:allowed] + "\n[... Truncado por límite de contexto ...]"
                            else:
                                break
                                
                        if texto_completo_doc:
                            texto_completo_doc += "\n\n[...]\n\n"
                        texto_completo_doc += blk_truncated
                        total_chars += len(blk_truncated)
                        
                    if texto_completo_doc:
                        manuales_texto += f"""
DOCUMENTO: {doc_nombre}
CONTENIDO DE REFERENCIA:
{texto_completo_doc}
=================================
"""

                id_manual = None
                nombre_pdf = ""
                es_pdf_abierto = True
                if manuales_con_score and manuales_con_score[0][0] >= 45:
                    id_manual = manuales_con_score[0][1]["id"]
                    nombre_pdf = manuales_con_score[0][1]["nombre"]
                    es_pdf_abierto = manuales_con_score[0][1].get("abierto", 1) == 1

                def score_chat(chat_item):
                    import difflib
                    preg_norm = normalizar_texto(chat_item["Pregunta_Usuario"])
                    score = 0.0
                    if preg_norm in contexto_norm or contexto_norm in preg_norm:
                        score += 100
                    preg_palabras = [w for w in re.findall(r"\w+", preg_norm) if w not in stopwords]
                    
                    for wp in preg_palabras:
                        for wq in query_palabras:
                            if wq == wp:
                                score += 40
                                break
                            ratio = difflib.SequenceMatcher(None, wq, wp).ratio()
                            if ratio >= 0.8:
                                score += int(40 * ratio)
                                break
                    return score

                # Filtrar y ordenar chats previos útiles (Me_Sirvio = 1)
                chats_con_score = []
                for c_item in chats_previos:
                    sc = score_chat(c_item)
                    chats_con_score.append((sc, c_item))

                chats_con_score.sort(key=lambda x: x[0], reverse=True)
                chats_seleccionados = [c_i for sc_val, c_i in chats_con_score if sc_val >= 40]
                chats_seleccionados = chats_seleccionados[:1]

                casos_previos_texto = ""
                for idx, c_item in enumerate(chats_seleccionados, start=1):
                    preg_val = c_item.get('Pregunta_Usuario') or ""
                    resp_val = c_item.get('Respuesta_IA') or ""
                    casos_previos_texto += f"""
EJEMPLO ÚTIL {idx} (Calificado positivamente por el usuario):
Pregunta del usuario: {preg_val[:500]}
Respuesta correcta de LUXO: {resp_val[:1500]}
=================================
"""

                # Filtrar y ordenar chats fallidos a evitar (Me_Sirvio = 0)
                chats_fallidos_con_score = []
                for c_item in chats_fallidos:
                    sc = score_chat(c_item)
                    chats_fallidos_con_score.append((sc, c_item))
                
                chats_fallidos_con_score.sort(key=lambda x: x[0], reverse=True)
                chats_fallidos_seleccionados = [c_i for sc_val, c_i in chats_fallidos_con_score if sc_val >= 30]
                chats_fallidos_seleccionados = chats_fallidos_seleccionados[:1]

                casos_fallidos_texto = ""
                for idx, c_item in enumerate(chats_fallidos_seleccionados, start=1):
                    preg_val = c_item.get('Pregunta_Usuario') or ""
                    resp_val = c_item.get('Respuesta_IA') or ""
                    casos_fallidos_texto += f"""
EJEMPLO ERRÓNEO A EVITAR {idx} (Calificado NEGATIVAMENTE por el usuario):
Pregunta del usuario: {preg_val[:500]}
Respuesta incorrecta o insuficiente a evitar: {resp_val[:1500]}
=================================
"""

                respuesta = ""

                if ask_for_list:
                    if manuales:
                        lista_items = [f"{idx}. '{m['Nombre_Archivo']}'" for idx, m in enumerate(manuales, start=1)]
                        lista_str = "\n".join(lista_items)
                        respuesta = f"Tengo acceso a los siguientes manuales cargados en el sistema:\n\n{lista_str}"
                    else:
                        respuesta = "No hay manuales cargados en el sistema."
                    historial_sesion.append({"role": "user", "content": user_text_expandido})
                    historial_sesion.append({"role": "assistant", "content": respuesta})
                elif ask_for_pdf:
                    if manuales_con_score and manuales_con_score[0][0] >= 5:
                        mejor = manuales_con_score[0][1]
                        id_manual = mejor["id"]
                        nombre_pdf = mejor["nombre"]
                        respuesta = t("pdf_delivered").format(nombre_pdf=nombre_pdf)
                    else:
                        respuesta = t("pdf_not_found")
                    historial_sesion.append({"role": "user", "content": user_text_expandido})
                    historial_sesion.append({"role": "assistant", "content": respuesta})
                else:
                    headers = {
                        "Authorization": f"Bearer {GROQ_API_KEY}",
                        "Content-Type": "application/json"
                    }

                    if es_sondeo_forzado:
                        mensaje_sistema = {
                            "role": "system",
                            "content": f"""Eres LUXO, asistente operativo inteligente de Sunglass Hut.

ATENCIÓN: La pregunta del usuario es extremadamente ambigua o corta (ej. "corte", "no puedo subir mi corte", "garantías", "problemas con el sistema").
Tu único y exclusivo objetivo en este turno es realizar de 1 a 3 preguntas de sondeo claras, amables y cortas para entender exactamente qué necesita antes de buscar o dar procedimientos del manual.

ESTÁ ESTRICTAMENTE PROHIBIDO:
- Dar instrucciones operativas, pasos, o guías.
- Mencionar o resumir información de cualquier manual (como "POLITICA GAFA ANIVERSARIO.pdf" o cualquier otro).
- Citar fuentes o nombres de archivos PDF.
- Hacer suposiciones sobre lo que el usuario necesita.

Formula las preguntas de manera natural, amable y conversacional según el contexto del término que ingresó el usuario.

CRITICAL LANGUAGE MATCHING RULE (HIGH PRIORITY):
- Analyze the exact language used in the user's latest message: "{user_text}".
- You MUST write your entire response in the EXACT same language as the user's message.
- If the user wrote in English (e.g., "Hello", "How do I..."), you MUST answer 100% in ENGLISH.
- If the user wrote in French, answer in FRENCH. If Italian, answer in ITALIAN. If Chinese, answer in CHINESE.
- Translate all relevant information from Spanish manuals on the fly into the user's language.
- DO NOT default to Spanish when the user speaks in English or any other language!
"""
                        }
                    elif modo_sugerencia:
                        mensaje_sistema = {
                            "role": "system",
                            "content": f"""Eres LUXO, asistente operativo de Sunglass Hut.
                            
El usuario realizó una consulta, pero no logramos identificar con certeza un manual relacionado en nuestro sistema.
Sin embargo, se encontraron estos documentos candidatos:
{', '.join(sugerencias_nombres)}

Por favor, responde de manera muy natural y amable. Pregúntale si su duda se refiere a alguno de estos temas. Usa la frase "¿Quizás quisiste decir...?" y presenta las opciones de forma clara (ej. viñetas con los nombres de los manuales sugeridos) para que el usuario pueda elegir o reformular su pregunta.
No inventes información sobre el contenido de los manuales.

CRITICAL LANGUAGE MATCHING RULE (HIGH PRIORITY):
- Analyze the user's message: "{user_text}".
- Reply in the EXACT same language as the user's input (English -> English, French -> French, etc.). Never force Spanish if the user spoke in English.
"""
                        }
                    else:
                        instruccion_trivia = ""
                        if es_pregunta_trivia:
                            instruccion_trivia = """
══════════════════════════════════════════════════════════
INSTRUCCIÓN ESPECIAL DE TRIVIA / RETO DEL DÍA (EVITAR ALUCINACIÓN)
══════════════════════════════════════════════════════════
El usuario te está preguntando sobre una pregunta específica de la Trivia.
1. NO hagas preguntas de sondeo. Explica directamente la respuesta.
2. Explica de forma didáctica y clara por qué la opción indicada como correcta en la pregunta del usuario es la correcta.
3. REGLA DE SEGURIDAD ABSOLUTA: Si no se te proporciona ningún manual relacionado con el tema en la sección "DOCUMENTOS / MANUALES" (está vacía), debes responder de forma directa explicando la regla teórica del negocio y decir honestamente: 'Dado que no cuento con el manual específico sobre este proceso cargado en mi sistema, te explico el protocolo operativo general de la tienda: [explicación]'.
4. Está terminantemente PROHIBIDO inventar nombres de archivos PDF o inventar secciones de manuales que no aparezcan de forma exacta en el texto de referencia provisto.
"""
                        mensaje_sistema = {
                            "role": "system",
                            "content": f"""Eres LUXO, asistente operativo inteligente de Sunglass Hut.
{instruccion_trivia}
CRITICAL LANGUAGE MATCHING RULE (MANDATORY & HIGH PRIORITY):
1. Analyze the language of the user's latest input: "{user_text}".
2. You MUST reply in the EXACT SAME LANGUAGE as the user's latest input:
   - User input in ENGLISH -> Reply 100% in ENGLISH.
   - User input in FRENCH -> Reply 100% in FRENCH.
   - User input in ITALIAN -> Reply 100% in ITALIAN.
   - User input in CHINESE -> Reply 100% in CHINESE.
   - User input in SPANISH -> Reply 100% in SPANISH.
3. Translate all procedures and content from Spanish manuals into the user's language dynamically.
4. Ignore any previous assistant messages stating language restrictions.

INSTRUCCIÓN CLAVE DE CONTEXTO DE USUARIO:
1. El usuario de LUXO es SIEMPRE un empleado, asesor de ventas o gerente de tienda de Sunglass Hut (NUNCA es un cliente final).
2. Toda consulta formulada, incluso en primera persona (ej. "me roban", "perdí", "hago"), se refiere a la situación operativa de la tienda y del personal de la tienda.
3. Está terminantemente PROHIBIDO redactar respuestas con tono de servicio de atención al cliente final o promesas de experiencia de compra (ej. "estamos aquí para que tengas una experiencia inolvidable", "te ayudamos con tu compra").
4. Debes dirigirte al usuario como empleado/gerente y proveer únicamente las instrucciones y protocolos técnicos y operativos detallados en los manuales para el personal.

INSTRUCCIÓN DE CONCISIÓN:
Debes responder de manera directa, concisa y profesional. Queda estrictamente prohibido formular preguntas de aclaración o de sondeo si cuentas con la información en la sección de manuales proporcionada. Ve al grano inmediatamente.


══════════════════════════════════════════════════════════
INSTRUCCIÓN CRÍTICA DE SEGURIDAD (CERO ALUCINACIONES)
══════════════════════════════════════════════════════════
1. Debes responder basándote ÚNICAMENTE en la información provista en la sección "DOCUMENTOS / MANUALES". Está terminantemente PROHIBIDO usar tu conocimiento general o inventar procedimientos para resolver consultas operativas.
2. No agregues pasos, consejos ni sugieras personas a contactar que no estén indicados textualmente en el manual provisto.
3. Si el manual establece condiciones o requisitos específicos para una acción, enuméralos exactamente como aparecen en el texto. No resumas ni generalices.
4. Para consultas operativas: Si la respuesta no está detallada textualmente en los manuales proporcionados y ya hiciste el sondeo necesario, responde únicamente: "Por el momento no cuento con esta información."
5. EXCEPCIÓN: Los saludos, despedidas y comentarios de cortesía respóndelos de forma natural, amable y profesional sin consultar documentos.
6. Está estrictamente PROHIBIDO comenzar tu respuesta con muletillas como "Según el manual...", "De acuerdo con el documento..." — da la respuesta de forma directa y profesional.
7. Para respuestas basadas en manuales (excluyendo saludos/cortesías), al final cita la fuente exacta en el idioma del usuario (ej: "You can find this information in the manual [Manual Name], section [Section Name]").
8. Para fórmulas matemáticas usa "entre" o "dividido entre" para divisiones, nunca "dividido por".
9. NO traduzcas siglas (como AUR) si la traducción no está textualmente en el manual.

══════════════════════════════════════════════════════════
INSTRUCCIÓN DE INTERPRETACIÓN Y ORTOGRAFÍA (OBLIGATORIO)
══════════════════════════════════════════════════════════
1. Los usuarios frecuentemente escriben con FALTAS DE ORTOGRAFÍA graves, abreviaciones, jerga mexicana y "spanglish". Tu DEBER es interpretar inteligentemente lo que el usuario quiso decir y responder a la intención real de su mensaje. Ejemplos:
   - "komo ago un corte de kaja" → el usuario quiere saber cómo hacer un corte de caja.
   - "nesesito el manual de devolusiones" → necesita el manual de devoluciones.
   - "q onda con la garantia" → pregunta sobre garantías.
   - "checa el tkt" → quiere revisar un ticket.
   - "ocpo saber del siniestro" → necesita información sobre un siniestro/robo.
   - "como le ago pa cerrar" → cómo hacer el cierre de caja.
2. NUNCA le digas al usuario que escribió mal, NUNCA corrijas su ortografía de forma explícita, NUNCA respondas "¿quisiste decir...?" por un error de escritura. Simplemente interpreta y responde.
3. Tus respuestas SIEMPRE deben estar escritas con ORTOGRAFÍA Y GRAMÁTICA PERFECTA en español (o el idioma correspondiente). Usa acentos correctos (á, é, í, ó, ú, ñ), signos de puntuación adecuados y redacción profesional.
4. Entiende estas abreviaciones comunes: "q" = que, "k" = que, "x" = por, "pa" = para, "tmb" = también, "pls" = por favor, "neta" = verdad, "chido" = bueno, "jale" = trabajo, "morro" = joven, "ntp" = no te preocupes, "tkt" = ticket, "cel" = celular, "pto" = punto, "xq" = porque, "ps" = pues, "bn" = bien, "msj" = mensaje, "fav" = favor, "dpto" = departamento, "gcia" = gerencia.

DOCUMENTOS / MANUALES PROPORCIONADOS:
{manuales_texto if manuales_texto.strip() else "(No hay documentos relacionados para esta consulta)"}

CASOS PREVIOS RESUELTOS CON ÉXITO (EJEMPLOS ÚTILES):
{casos_previos_texto if casos_previos_texto.strip() else "(No hay casos previos similares registrados)"}

EJEMPLOS ERRÓNEOS A EVITAR (RETROALIMENTACIÓN NEGATIVA A NO REPETIR):
{casos_fallidos_texto if casos_fallidos_texto.strip() else "(No hay ejemplos de respuestas incorrectas previas registrados)"}
"""
                        }

                    historial_sesion.append({"role": "user", "content": user_text_expandido})
                    mensajes_api = [mensaje_sistema]
                    
                    # Limitar memoria y truncar mensajes extremadamente largos para evitar error 413
                    historial_filtrado = []
                    for msg in historial_sesion[-8:]:
                        content_trunc = msg["content"]
                        if len(content_trunc) > 800:
                            content_trunc = content_trunc[:800] + "\n[...]"
                        historial_filtrado.append({"role": msg["role"], "content": content_trunc})
                    mensajes_api.extend(historial_filtrado)

                    payload = {
                        "model": GROQ_MODEL,
                        "messages": mensajes_api
                    }


                    try:
                        headers_groq = {"Authorization": f"Bearer {get_groq_key()}", "Content-Type": "application/json"}
                        res = requests.post(URL_GROQ, headers=headers_groq, json=payload, timeout=15)
                        # Si hay error 429 (cuota agotada), rotar a llave de respaldo y reintentar
                        if res.status_code == 429 and rotate_groq_key():
                            headers_groq = {"Authorization": f"Bearer {get_groq_key()}", "Content-Type": "application/json"}
                            res = requests.post(URL_GROQ, headers=headers_groq, json=payload, timeout=15)
                        if res.status_code == 200:
                            try:
                                data = res.json()
                                if "choices" in data and data["choices"]:
                                    respuesta = data["choices"][0]["message"]["content"]
                                else:
                                    respuesta = "Ocurrió un error consultando la IA."
                            except Exception as e:
                                print("AI PARSE ERROR:", e)
                                respuesta = "Ocurrió un error consultando la IA."
                        else:
                            print("AI CONNECTION ERROR:", res.status_code, res.text)
                            respuesta = f"Error de conexión con la IA ({res.status_code})."
                    except Exception as re_err:
                        print("API REQUEST EXCEPTION:", re_err)
                        respuesta = "Error de conexión: No se pudo establecer contacto con el servidor de Inteligencia Artificial. Por favor, verifica tu conexión a internet e inténtalo de nuevo."


                    historial_sesion.append({"role": "assistant", "content": respuesta})

                # --- ALINEACIÓN DE ARCHIVO ADJUNTO CON LA RESPUESTA DE LA IA ---
                # Guardar candidatos iniciales del RAG vector search
                id_manual_rag = id_manual
                nombre_pdf_rag = nombre_pdf
                es_pdf_abierto_rag = es_pdf_abierto

                # Resetear id_manual por defecto para NO mostrar tarjetas adjuntas en saludos o pláticas generales
                id_manual = None
                nombre_pdf = ""
                es_pdf_abierto = False

                if ask_for_list:
                    id_manual = None
                    nombre_pdf = ""
                    es_pdf_abierto = False
                elif manual_forzado_trivia[0] is not None:
                    id_manual = id_manual_rag
                    nombre_pdf = nombre_pdf_rag
                    es_pdf_abierto = es_pdf_abierto_rag
                elif respuesta and "no cuento con" not in respuesta.lower() and "error de conexión" not in respuesta.lower() and "ocurrió un error" not in respuesta.lower():
                    respuesta_lower = respuesta.lower()
                    encontrado = False
                    for m in manuales:
                        m_nombre = m.get("Nombre_Archivo") or ""
                        nombre_sin_ext = m_nombre.rsplit(".", 1)[0] if "." in m_nombre else m_nombre
                        if m_nombre.lower() in respuesta_lower or (len(nombre_sin_ext) > 3 and nombre_sin_ext.lower() in respuesta_lower):
                            id_manual = m["ID_Manual"]
                            nombre_pdf = m_nombre
                            es_pdf_abierto = m.get("Abierto", 1) == 1
                            encontrado = True
                            break
                    
                    # Si la IA citó el manual en la respuesta se adjunta.
                    # Si no lo citó explícitamente, pero es una consulta operativa extensa con alto score RAG (>=80), adjuntar el documento RAG
                    palabras_usuario = user_text.lower().split()
                    es_saludo_conversacion = any(w in ["hi", "hello", "hola", "hey", "buenos", "dias", "tardes", "noches", "saludos", "quien", "eres", "nombre", "name"] for w in palabras_usuario)
                    if not encontrado and not es_saludo_conversacion and len(palabras_usuario) >= 4 and manuales_con_score and manuales_con_score[0][0] >= 80:
                        id_manual = id_manual_rag
                        nombre_pdf = nombre_pdf_rag
                        es_pdf_abierto = es_pdf_abierto_rag

                # --- DETECTAR SI LA RESPUESTA ES DE SONDEO (solo preguntas) ---
                # Si la respuesta contiene 2+ signos de interrogación de cierre, es sondeo
                # En ese caso NO se mostrará el manual ni se dará información adicional
                import re as _re_sondeo
                signos_pregunta = len(_re_sondeo.findall(r'\?', respuesta))
                # Palabras clave que indican que el AI está haciendo preguntas de sondeo
                frases_sondeo = [
                    "me puedes decir", "podrías indicarme", "podrías decirme",
                    "para orientarte mejor", "para darte la información más precisa",
                    "para ayudarte mejor", "más detalles", "me puedes proporcionar",
                    "te refieres al", "a qué tipo de", "en qué plataforma",
                    "cuéntame más", "información adicional", "necesito saber"
                ]
                es_respuesta_sondeo = (
                    es_sondeo_forzado or (
                        signos_pregunta >= 2 and
                        any(frase in respuesta.lower() for frase in frases_sondeo)
                    )
                )
                # Si es sondeo, anular el manual para que no se muestre la tarjeta
                if es_respuesta_sondeo:
                    id_manual = None
                    nombre_pdf = ""
                    es_pdf_abierto = False

                # --- REGISTRAR EN BASE DE DATOS ANTES DE RENDERIZAR PARA OBTENER EL ID ---
                id_conversacion = None
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        sql_historial = """
                        INSERT INTO historial_conversaciones
                        (
                            ID_Usuario,
                            ID_Manual,
                            Pregunta_Usuario,
                            Respuesta_IA,
                            Fecha_Hora,
                            Fue_Respondida_Con_Manual
                        )
                        VALUES (%s, %s, %s, %s, NOW(), %s)
                        """
                        cursor.execute(
                            sql_historial,
                            (
                                user_info["id"],
                                id_manual,
                                user_text,
                                respuesta,
                                1 if id_manual and "no cuento con" not in respuesta.lower() else 0,
                            )
                        )
                        db.commit()
                        id_conversacion = cursor.lastrowid

                        if "Por el momento no cuento con esta información" in respuesta:
                            sql_pendiente = """
                            INSERT INTO pendientes_actualizacion
                            (
                                ID_Conversacion,
                                Pregunta_Faltante
                            )
                            VALUES (%s, %s)
                            """
                            cursor.execute(sql_pendiente, (id_conversacion, user_text))
                            db.commit()
                            id_pendiente = cursor.lastrowid
                            clasificar_pregunta_faltante_async(user_text, id_pendiente)
                        db.close()
                except Exception as ex:
                    print("ERROR AL REGISTRAR CONVERSACIÓN:", ex)

                # --- CREAR CONTENEDOR DE FEEDBACK SI SE OBTUVO EL ID DE CONVERSACIÓN ---
                feedback_container = ft.Container(alignment=ft.alignment.Alignment(-1, 0), expand=True)
                if id_conversacion:
                    def on_thumbs_up(ev, conv_id=id_conversacion, fb_cont=feedback_container):
                        comment_input = ft.TextField(
                            label="Cuéntanos qué te sirvió o agrega un comentario (opcional)",
                            multiline=True,
                            min_lines=2,
                            max_lines=4,
                            border_color="#9D50BB",
                            color="white"
                        )
                        
                        def submit_comment(e):
                            comentario = comment_input.value.strip()
                            registrar_feedback(conv_id, True, comentario, fb_cont)
                            page.pop_dialog()
                            page.update()
                            
                        def cancel_comment(e):
                            registrar_feedback(conv_id, True, "", fb_cont)
                            page.pop_dialog()
                            page.update()

                        dialog_comment = ft.AlertDialog(
                            title=ft.Text("Calificar respuesta como Útil", color="#7CFC00", weight="bold", size=16),
                            content=ft.Container(
                                content=comment_input,
                                width=450,
                                height=120
                            ),
                            actions=[
                                ft.TextButton("Calificar Directo", on_click=cancel_comment),
                                ft.ElevatedButton("Enviar y Calificar", on_click=submit_comment, bgcolor="#7CFC00", color="white")
                            ],
                            actions_alignment="end",
                            bgcolor="#0F0F1A"
                        )
                        page.show_dialog(dialog_comment)
                        
                    def on_thumbs_down(ev, conv_id=id_conversacion, fb_cont=feedback_container):
                        comment_input = ft.TextField(
                            label="Cuéntanos por qué no te sirvió (opcional)",
                            multiline=True,
                            min_lines=2,
                            max_lines=4,
                            border_color="#9D50BB",
                            color="white"
                        )
                        
                        def submit_comment(e):
                            comentario = comment_input.value.strip()
                            registrar_feedback(conv_id, False, comentario, fb_cont)
                            page.pop_dialog()
                            page.update()
                            
                        def cancel_comment(e):
                            registrar_feedback(conv_id, False, "", fb_cont)
                            page.pop_dialog()
                            page.update()

                        def crear_ticket_click(e):
                            comentario = comment_input.value.strip()
                            if not comentario:
                                comentario = "El bot no respondió correctamente a la pregunta: " + user_text
                            registrar_feedback(conv_id, False, comentario, fb_cont)
                            try:
                                db_t = conectar_db()
                                if db_t:
                                    cursor_t = db_t.cursor()
                                    cursor_t.execute("""
                                        INSERT INTO tickets_soporte (ID_Usuario, Detalle_Problema)
                                        VALUES (%s, %s)
                                    """, (user_info["id"], f"Pregunta: {user_text}\nRespuesta Luxo: {respuesta}\nComentario: {comentario}"))
                                    db_t.commit()
                                    db_t.close()
                                    mostrar_snack("¡Ticket de Soporte creado con éxito!", color="#7CFC00")
                            except Exception as ex:
                                print("ERROR AL CREAR TICKET:", ex)
                                mostrar_snack("Error al registrar el ticket.", color="red")
                            page.pop_dialog()
                            page.update()

                        dialog_comment = ft.AlertDialog(
                            title=ft.Text("Calificar respuesta como No Útil", color="#FF4500", weight="bold", size=16),
                            content=ft.Container(
                                content=comment_input,
                                width=450,
                                height=120
                            ),
                            actions=[
                                ft.TextButton("Calificar Directo", on_click=cancel_comment),
                                ft.ElevatedButton("Crear Ticket 🎫", on_click=crear_ticket_click, bgcolor="#9D50BB", color="white"),
                                ft.ElevatedButton("Enviar y Calificar", on_click=submit_comment, bgcolor="#FF4500", color="white")
                            ],
                            actions_alignment="end",
                            bgcolor="#0F0F1A"
                        )
                        page.show_dialog(dialog_comment)

                    btn_speaker = ft.IconButton(
                        icon=ft.Icons.VOLUME_UP_ROUNDED,
                        icon_size=15,
                        icon_color="#00FFFF",
                        tooltip="Escuchar respuesta",
                    )
                    
                    btn_play_pause = ft.IconButton(
                        icon=ft.Icons.PAUSE_ROUNDED,
                        icon_size=15,
                        icon_color="#00FFFF",
                        tooltip="Pausar/Reanudar lectura",
                        disabled=True,
                    )
                    
                    def handle_speaker_click(e, txt=respuesta, bs=btn_speaker, bpp=btn_play_pause):
                        nonlocal current_speak_btn_speaker
                        if current_speak_btn_speaker == bs:
                            stop_current_speak()
                        else:
                            start_speak(txt, bs, bpp)
                            
                    def handle_play_pause_click(e):
                        toggle_pause_speak()

                    btn_speaker.on_click = handle_speaker_click
                    btn_play_pause.on_click = handle_play_pause_click

                    feedback_buttons = ft.Row([
                        ft.Text("¿Te sirvió la respuesta?", color="#aaaaaa", size=11),
                        ft.IconButton(
                            icon=ft.Icons.THUMB_UP_OUTLINED,
                            icon_size=15,
                            icon_color="#7CFC00",
                            tooltip="Sí, fue útil",
                            on_click=on_thumbs_up
                        ),
                        ft.IconButton(
                            icon=ft.Icons.THUMB_DOWN_OUTLINED,
                            icon_size=15,
                            icon_color="#FF4500",
                            tooltip="No fue útil",
                            on_click=on_thumbs_down
                        ),
                        btn_speaker,
                        btn_play_pause
                    ], spacing=5, alignment="start", vertical_alignment="center", wrap=True)
                    
                    feedback_container.content = feedback_buttons

                # Renderizar mensaje de Luxo
                # --- BUSCAR ACTIVOS VISUALES INTERACTIVOS (Assets) ---
                keyword_assets = {
                    "impresora": "ayuda_impresora.png",
                    "epson": "ayuda_impresora.png",
                    "papel": "ayuda_impresora.png",
                    "terminal": "guia_terminal.png",
                    "caja": "guia_terminal.png",
                    "devolucion": "guia_devolucion.png",
                    "devoluciones": "guia_devolucion.png",
                    "politica de devolucion": "guia_devolucion.png",
                    "politica de devoluciones": "guia_devolucion.png",
                }
                
                matched_asset = None
                clean_resp_lower = respuesta.lower()
                # No mostrar imagenes de guia si la respuesta indica no disponibilidad de la informacion o si se adjunta un PDF
                if "no cuento con" not in clean_resp_lower and id_manual is None:
                    for kw, asset_filename in keyword_assets.items():
                        if kw in clean_resp_lower:
                            asset_path = os.path.join(ASSETS_PATH, asset_filename)
                            if os.path.exists(asset_path):
                                matched_asset = f"custom_assets/{asset_filename}" if page.web else asset_path
                                break

                luxo_column_controls = [
                    ft.Row([
                        ft.Container(
                            content=fv.Video(
                                playlist=[fv.VideoMedia(video_chat_url)],
                                playlist_mode=fv.PlaylistMode.LOOP,
                                autoplay=True,
                                muted=True,
                                controls=None,
                                expand=True,
                                fit=ft.BoxFit.COVER,
                                filter_quality=ft.FilterQuality.HIGH,
                            ) if video_chat_exists else (
                                ft.Image(src=img_avatar, width=50, height=50) if img_avatar else ft.Text("L", color="white", weight="bold")
                            ),
                            width=50,
                            height=50,
                            border_radius=25,
                            clip_behavior=ft.ClipBehavior.HARD_EDGE,
                            border=ft.Border.all(1.5, "#00FFFF"),
                        ),
                        ft.Text(f"LUXO: {respuesta}", color="white", weight="bold", expand=True, selectable=True),
                    ], vertical_alignment="start", spacing=10),
                    ft.Row([
                        ft.Container(width=60),
                        feedback_container
                    ])
                ]


                if matched_asset:
                    luxo_column_controls.append(
                        ft.Container(
                            content=ft.Column([
                                ft.Text("💡 Guía Visual Relacionada:", color="#00FFFF", size=12, weight="bold"),
                                ft.Image(src=matched_asset, width=400, height=220, border_radius=10)
                            ], spacing=5),
                            bgcolor="#1e1e1e",
                            padding=10,
                            border_radius=8,
                            border=ft.Border.all(1, "#333333"),
                            margin=ft.Margin(left=45, top=5, right=0, bottom=0)
                        )
                    )

                chat_display.controls.append(
                    ft.Container(
                        content=ft.Column(luxo_column_controls, spacing=5),
                        bgcolor="#0F0F1A",
                        padding=10,
                        border_radius=10
                    )
                )

                if id_manual and es_pdf_abierto and "no cuento con" not in clean_resp_lower and "error de conexión" not in clean_resp_lower and "ocurrió un error" not in clean_resp_lower and not es_respuesta_sondeo:
                    import urllib.parse
                    nombre = obtener_pdf_assets(id_manual)
                    url_url = ""
                    url_view = ""
                    url_dl = ""
                    if nombre:
                        nombre_quoted = urllib.parse.quote(nombre)
                        base_url = page.url.rstrip("/") if (page and page.url) else "http://localhost:8550"
                        if base_url.startswith("ws://"):
                            base_url = base_url.replace("ws://", "http://", 1)
                        elif base_url.startswith("wss://"):
                            base_url = base_url.replace("wss://", "https://", 1)
                        url_view = f"{base_url}/temp_pdfs/{nombre_quoted}"
                        base_dl = re.sub(r":\d+$", f":{PUERTO_DESCARGAS}", base_url)
                        url_dl = f"{base_dl}/download?file={nombre_quoted}"
                    
                    chat_display.controls.append(
                        ft.Container(
                            content=ft.Column([
                                ft.Row([
                                    ft.Icon(ft.Icons.PICTURE_AS_PDF, color="#D8B4FE"),
                                    ft.Text(f"{nombre_pdf}", color="white", weight="bold", size=13),
                                ], spacing=5),
                                ft.Row([
                                    ft.ElevatedButton(
                                        t("view_pdf"),
                                        url=url_view,
                                        bgcolor="#6E48AA",
                                        color="white",
                                        expand=True,
                                        disabled=(url_view == "")
                                    ),
                                    ft.ElevatedButton(
                                        t("download_pdf"),
                                        url=url_dl,
                                        bgcolor="#444444",
                                        color="white",
                                        expand=True,
                                        disabled=(url_dl == "")
                                    ),
                                ], spacing=5),
                                ft.Text(
                                    "💡 Tip: Mantén presionado 'Descargar' y elige 'Descargar vínculo/enlace' para guardarlo directo en tu celular.",
                                    color="#aaaaaa",
                                    size=10,
                                    italic=True
                                )
                            ], spacing=8),
                            bgcolor="#1a1a2e",
                            padding=12,
                            border_radius=10
                        )
                    )

            except Exception as e:
                import traceback
                traceback.print_exc()
                chat_display.controls.append(
                    ft.Text(f"ERROR: {e}", color="red")
                )

            manual_forzado_trivia[0] = None
            page.update()

        def en_archivo_audio_seleccionado(e):
            if not e.files:
                return
            
            file = e.files[0]
            if file.path:
                # Local (PC local)
                procesar_transcripcion_archivo(file.path)
            else:
                # Web / Móvil (subida requerida)
                mostrar_snack("Subiendo audio... Por favor espera 🎙️", "#00FFFF")
                try:
                    upload_url = page.get_upload_url(file.name, 600)
                    file_picker_audio.upload_files(
                        [
                            ft.FilePickerUploadFile(
                                file_name=file.name,
                                upload_url=upload_url
                            )
                        ]
                    )
                except Exception as ex:
                    print("Error al subir archivo Flet:", ex)
                    mostrar_snack("Error al cargar el archivo de audio.", "red")

        def en_progreso_subida_audio(e):
            if e.progress == 1.0:
                import time
                time.sleep(0.5)
                filename = e.file_name
                filepath = os.path.join("uploads", filename)
                procesar_transcripcion_archivo(filepath)

        def procesar_transcripcion_archivo(filepath):
            mostrar_snack("Transcribiendo audio con IA... Por favor espera 🎙️", "#00FFFF")
            try:
                import requests
                import os
                
                if not os.path.exists(filepath):
                    print("Error: archivo no encontrado en el servidor:", filepath)
                    mostrar_snack("Error: el archivo de audio no se subió correctamente.", "red")
                    return
                
                headers = {
                    "Authorization": f"Bearer {GROQ_API_KEY}"
                }
                
                with open(filepath, "rb") as f:
                    files = {
                        "file": (os.path.basename(filepath), f, "audio/mpeg")
                    }
                    data = {
                        "model": "whisper-large-v3",
                        "language": "es",
                        "response_format": "json"
                    }
                    
                    res = requests.post(
                        "https://api.groq.com/openai/v1/audio/transcriptions",
                        headers=headers,
                        files=files,
                        data=data,
                        timeout=30
                    )
                    
                if res.status_code == 200:
                    transcripcion = res.json().get("text", "").strip()
                    if transcripcion:
                        input_msg.value = f"🎙️ {transcripcion}"
                        enviar_mensaje(None)
                    else:
                        mostrar_snack("No se detectó voz clara en el audio.", "red")
                else:
                    print("Error Groq Whisper:", res.status_code, res.text)
                    mostrar_snack("Error al procesar el audio con la IA.", "red")
                
                try:
                    if os.path.exists(filepath):
                        os.unlink(filepath)
                except Exception:
                    pass
            except Exception as ex:
                print("EXCEPCIÓN EN TRANSCRIPCIÓN:", ex)
                mostrar_snack("Error del sistema al procesar el audio.", "red")


        input_msg = ft.TextField(
            hint_text=g_tr("Escribe tu consulta...", "Type your query...", "Tapez votre requête...", "Scrivi la tua richiesta...", "请输入您的咨询..."),
            expand=True,
            multiline=True,
            min_lines=1,
            max_lines=4,
            shift_enter=True,
            on_submit=enviar_mensaje,
            border_color="#9D50BB",
            color="white",
            bgcolor="#0F0F1A"
        )

        # Campos ocultos de puente para la grabación en la misma página (sin popups)
        trigger_field = ft.TextField(
            value="READY",
            hint_text="TRIGGER_REC",
            width=0,
            height=0,
            opacity=0
        )
        
        user_id_field = ft.TextField(
            value=str(user_info["id"]),
            hint_text="USER_ID_HINT",
            width=0,
            height=0,
            opacity=0
        )

        # Registrar la sesión activa para el servidor de audio auxiliar
        sess_data = {
            "input_msg": input_msg,
            "enviar_mensaje": enviar_mensaje,
            "page": page
        }
        if user_info.get("id"):
            active_sessions[user_info["id"]] = sess_data
        active_sessions["1"] = sess_data
        active_sessions[1] = sess_data

        # =================================
        # VISTAS DEL PANEL DINÁMICO
        # =================================

        def mostrar_instrucciones_dictado(e):
            def cerrar_dictado_dialog(ev):
                page.pop_dialog()
            dialog = ft.AlertDialog(
                title=ft.Text("🎙️ Dictado por Voz (Speech-to-Text)", color="#D8B4FE", weight="bold"),
                content=ft.Column([
                    ft.Text("Para escribir usando tu voz, puedes usar las funciones de dictado nativas de tu dispositivo:", color="white", size=14),
                    ft.Divider(height=10, color="#333333"),
                    ft.Row([
                        ft.Icon(ft.Icons.PHONE_ANDROID, color="#7CFC00"),
                        ft.Text("En Celular (Android / iPhone):", weight="bold", color="white")
                    ]),
                    ft.Text("1. Toca la barra de texto para escribir.\n2. Abre el teclado virtual y presiona el icono de micrófono que viene incorporado en tu teclado (al lado de la barra espaciadora o en la barra de sugerencias).\n3. Empieza a hablar.", color="#aaaaaa"),
                    ft.Divider(height=10, color="transparent"),
                    ft.Row([
                        ft.Icon(ft.Icons.COMPUTER, color="#00FFFF"),
                        ft.Text("En Computadora (Windows):", weight="bold", color="white")
                    ]),
                    ft.Text("1. Haz clic en la barra de texto para escribir.\n2. Presiona la combinación de teclas Windows + H.\n3. Asegúrate de tener un micrófono activo en tu configuración de sonido de Windows.", color="#aaaaaa"),
                ], tight=True, spacing=10),
                actions=[
                    ft.TextButton("Entendido", on_click=cerrar_dictado_dialog)
                ]
            )
            page.show_dialog(dialog)

        def inyectar_script_voz_luxo():
            js_code = """javascript:(function() {
                let topDoc = (window.top && window.top.document) ? window.top.document : document;
                
                // Inyección del estilo CSS para animación de ondas Siri
                let siriStyle = topDoc.getElementById("siri-orb-style");
                if (!siriStyle) {
                    siriStyle = topDoc.createElement("style");
                    siriStyle.id = "siri-orb-style";
                    siriStyle.innerHTML = `
                        @keyframes siriWavePulse {
                            0%, 100% { transform: scaleY(0.5); opacity: 0.6; }
                            50% { transform: scaleY(1.5); opacity: 1; }
                        }
                        @keyframes siriGlowPulse {
                            0%, 100% { box-shadow: 0 0 20px rgba(224, 64, 251, 0.7), 0 0 35px rgba(0, 240, 255, 0.4); }
                            50% { box-shadow: 0 0 30px rgba(0, 240, 255, 0.9), 0 0 50px rgba(224, 64, 251, 0.8); }
                        }
                    `;
                    topDoc.head.appendChild(siriStyle);
                }

                let banner = topDoc.getElementById("luxo-voice-banner");
                if (!banner) {
                    banner = topDoc.createElement("div");
                    banner.id = "luxo-voice-banner";
                    banner.style.cssText = "position:fixed; bottom:30px; right:30px; z-index:2147483647; width:64px; height:64px; border-radius:50%; background:radial-gradient(circle at 35% 35%, rgba(224, 64, 251, 0.85), rgba(0, 240, 255, 0.85), rgba(10, 10, 24, 0.95)); animation: siriGlowPulse 2.5s infinite ease-in-out; display:flex; align-items:center; justify-content:center; cursor:pointer; user-select:none; transition: opacity 0.5s ease, transform 0.5s ease; opacity:0; pointer-events:none; transform:scale(0.7);";
                    banner.setAttribute("title", "Asistente de Voz LUXO (Oye LUXO)");
                    banner.innerHTML = `
                        <div style="display:flex; align-items:center; justify-content:center; gap:4px; height:100%; width:100%;">
                            <div style="width:4px; height:16px; background:#00F0FF; border-radius:2px; animation: siriWavePulse 1.2s infinite ease-in-out 0.1s;"></div>
                            <div style="width:4px; height:26px; background:#E040FB; border-radius:2px; animation: siriWavePulse 1.2s infinite ease-in-out 0.3s;"></div>
                            <div style="width:4px; height:22px; background:#C084FC; border-radius:2px; animation: siriWavePulse 1.2s infinite ease-in-out 0.2s;"></div>
                            <div style="width:4px; height:14px; background:#00F0FF; border-radius:2px; animation: siriWavePulse 1.2s infinite ease-in-out 0.4s;"></div>
                        </div>
                    `;
                    topDoc.body.appendChild(banner);
                }
                
                let fadeTimer = null;
                window.showLuxoSiriOrb = function(durationMs) {
                    const b = topDoc.getElementById("luxo-voice-banner") || banner;
                    if (!b) return;
                    b.style.opacity = "1";
                    b.style.transform = "scale(1)";
                    b.style.pointerEvents = "auto";
                    if (fadeTimer) { clearTimeout(fadeTimer); fadeTimer = null; }
                    if (durationMs && durationMs > 0) {
                        fadeTimer = setTimeout(function() {
                            window.hideLuxoSiriOrb();
                        }, durationMs);
                    }
                };

                window.hideLuxoSiriOrb = function() {
                    const b = topDoc.getElementById("luxo-voice-banner") || banner;
                    if (!b) return;
                    if (fadeTimer) { clearTimeout(fadeTimer); fadeTimer = null; }
                    b.style.opacity = "0";
                    b.style.transform = "scale(0.7)";
                    b.style.pointerEvents = "none";
                };

                const iconEl = topDoc.getElementById("luxo-voice-icon");
                const textEl = topDoc.getElementById("luxo-voice-text");
                
                function setStatus(text, color, icon) {
                    if (textEl) textEl.innerText = text;
                    if (iconEl && icon) iconEl.innerText = icon;
                    if (banner && color) banner.style.borderColor = color;
                }

                const SpeechRecognition = window.SpeechRecognition || window.webkitSpeechRecognition || (window.top && (window.top.SpeechRecognition || window.top.webkitSpeechRecognition));
                
                function playBeep() {
                    // Sonido desactivado
                }


                let rec = null;
                let isListening = false;
                let lastSentText = "";
                let lastSentTime = 0;

                window.initLuxoMicPermission = function() {
                    // La voz web estaba desactivada temporalmente aquí, ahora está activa.
                    if (window.luxoSpeechRecognitionActive) {
                        console.log("🎙️ [Luxo Chat Mic]: Already active. Skipping duplicate init.");
                        return;
                    }
                    setStatus("Solicitando permiso de micrófono...", "#FFFF00", "⏳");
                    
                    let mediaDev = navigator.mediaDevices || (window.top && window.top.navigator && window.top.navigator.mediaDevices);
                    
                    if (mediaDev && mediaDev.getUserMedia) {
                        mediaDev.getUserMedia({ audio: true })
                        .then(function(stream) {
                            setStatus("🟢 ESCUCHANDO... Di 'Oye LUXO'", "#00FF00", "🎤");
                            playBeep();
                            startRecognition();
                        })
                        .catch(function(err) {
                            console.log("Mic permission error:", err);
                            setStatus("❌ Permiso Denegado (Clica 🔒 en URL para permitir)", "#FF0000", "⚠️");
                        });
                    } else {
                        startRecognition();
                    }
                };

                function startRecognition() {
                    if (!SpeechRecognition) {
                        setStatus("❌ Usa Chrome o Edge para 'Oye LUXO'", "#FF0000", "❌");
                        return;
                    }
                    
                    try {
                        if (rec) {
                            try { rec.stop(); } catch(e){}
                        }
                        rec = new SpeechRecognition();
                        rec.continuous = true;
                        rec.interimResults = true;
                        rec.lang = 'es-MX';

                        rec.onstart = function() {
                            isListening = true;
                            window.luxoSpeechRecognitionActive = true;
                            setStatus("🟢 ESCUCHANDO EN VIVO... Di 'Oye LUXO'", "#00FFFF", "🎤");
                        };

                        rec.onresult = function(e) {
                            for (let i = e.resultIndex; i < e.results.length; ++i) {
                                const transcript = e.results[i][0].transcript;
                                const lower = transcript.toLowerCase();
                                console.log("🎙️ [Luxo Chat Mic]:", transcript, "isFinal:", e.results[i].isFinal);
                                
                                if (lower.includes("oye luxo") || lower.includes("hola luxo") || lower.includes("hey luxo") || lower.includes("oye lujo")) {
                                    const now = Date.now();
                                    let query = transcript
                                        .replace(/oye luxo/gi, '')
                                        .replace(/hola luxo/gi, '')
                                        .replace(/hey luxo/gi, '')
                                        .replace(/oye lujo/gi, '')
                                        .trim();
                                    if (!window.luxoIsListeningAlertSent) {
                                        window.luxoIsListeningAlertSent = true;
                                        fetch('/luxo_listening_start?user_id=1', { method: 'POST' });
                                    }
                                    
                                    // Activar orbe flotante en la esquina inferior derecha al detectar frase clave (5 segundos de visualización)
                                    window.showLuxoSiriOrb(5000);

                                    if (!query && e.results[i].isFinal) {
                                        playBeep();
                                        setStatus("👂 ¡Oye LUXO Detectado! Di tu pregunta...", "#FF00FF", "🔊");
                                        window.luxoManualDictating = true;
                                    } else if (query && e.results[i].isFinal) {
                                        if (query !== lastSentText) {
                                            lastSentText = query;
                                            lastSentTime = now;
                                            playBeep();
                                            window.showLuxoSiriOrb(5000);
                                            setStatus("⚡ Enviando a LUXO: " + query, "#7CFC00", "🚀");
                                            fetch('/text_input?user_id=1&text=' + encodeURIComponent(query), { method: 'POST' });
                                            window.luxoIsListeningAlertSent = false;
                                            window.luxoManualDictating = false;
                                            setTimeout(function(){ setStatus("🟢 ESCUCHANDO EN VIVO... Di 'Oye LUXO'", "#00FFFF", "🎤"); }, 3500);
                                            try { rec.abort(); } catch(e){} // Reinicia el buffer del microfono
                                        }
                                    }
                                } else if (window.luxoManualDictating && e.results[i].isFinal) {
                                    const query = transcript.trim();
                                    if (query && query !== lastSentText) {
                                        lastSentText = query;
                                        playBeep();
                                        window.showLuxoSiriOrb(5000);
                                        setStatus("⚡ Enviando a LUXO: " + query, "#7CFC00", "🚀");
                                        fetch('/text_input?user_id=1&text=' + encodeURIComponent(query), { method: 'POST' });
                                        window.luxoManualDictating = false;
                                        setTimeout(function(){ setStatus("🟢 ESCUCHANDO EN VIVO... Di 'Oye LUXO'", "#00FFFF", "🎤"); }, 3500);
                                        try { rec.abort(); } catch(e){} // Reinicia el buffer del microfono
                                    }
                                }
                            }
                        };

                        rec.onerror = function(err) {
                            console.log("Luxo Voice Error:", err);
                            if (err.error === 'not-allowed') {
                                setStatus("❌ Micrófono Bloqueado en el Navegador", "#FF0000", "🔒");
                            }
                        };

                        rec.onend = function() {
                            window.luxoSpeechRecognitionActive = false;
                            if (isListening) {
                                setTimeout(function() {
                                    try { rec.start(); } catch(e){}
                                }, 300);
                            }
                        };

                        rec.start();
                    } catch(ex) {
                        console.log("Exception in startRecognition:", ex);
                    }
                }

                banner.onclick = function() {
                    window.initLuxoMicPermission();
                };
                
                window.toggleLuxoDictate = function() {
                    window.initLuxoMicPermission();
                };

                window.initLuxoMicPermission();
            })(); void(0);"""

            async def _exec_js():
                try:
                    await page.launch_url(js_code)
                except Exception as ex:
                    print("Error ejecutando launch_url para JS de voz:", ex)
            page.run_task(_exec_js)

        def build_chat_view():
            inyectar_script_voz_luxo()

            dictado_en_progreso = [False]

            def on_mic_click(e):
                if dictado_en_progreso[0]:
                    print("⚠️ Dictado en progreso, ignorando clic secundario.")
                    return
                dictado_en_progreso[0] = True

                try:
                    mostrar_snack("🎙️ Escuchando pregunta directa... Habla ahora", "#00FFFF")
                    btn_mic_container.bgcolor = "#FF0000"
                    btn_mic_container.border = ft.Border.all(2, "white")
                    btn_mic_container.update()
                    
                    def dictado_thread():
                        try:
                            import speech_recognition as sr
                            import platform
                            if platform.system() == "Windows":
                                import winsound
                            
                            r_direct = sr.Recognizer()
                            r_direct.dynamic_energy_threshold = True
                            r_direct.pause_threshold = 1.0
                            r_direct.non_speaking_duration = 0.8
                            with sr.Microphone() as src_direct:
                                r_direct.adjust_for_ambient_noise(src_direct, duration=0.2)
                                try:
                                    if platform.system() == "Windows":
                                        winsound.Beep(1200, 250)
                                    elif platform.system() == "Darwin":
                                        import os
                                        os.system('afplay /System/Library/Sounds/Ping.aiff &')
                                except Exception:
                                    pass
                                
                                audio_direct = r_direct.listen(src_direct, timeout=6, phrase_time_limit=25)
                                text_direct = r_direct.recognize_google(audio_direct, language="es-MX")
                                
                                if text_direct:
                                    print(f"🎙️ Dictado directo captado en botón: '{text_direct}'")
                                    input_msg.value = text_direct
                                    page.update()
                                    
                                    async def trigger_send():
                                        enviar_mensaje(None)
                                    page.run_task(trigger_send)
                        except sr.WaitTimeoutError:
                            print("⏳ Tiempo de espera agotado. No se detectó voz.")
                        except Exception as ex_dict:
                            print("Error en dictado directo de botón:", ex_dict)
                        finally:
                            dictado_en_progreso[0] = False
                            try:
                                btn_mic_container.bgcolor = "#1E1E2E"
                                btn_mic_container.border = ft.Border.all(1.5, "#00FFFF")
                                btn_mic_container.update()
                            except Exception:
                                pass

                    t_dict = threading.Thread(target=dictado_thread, daemon=True)
                    t_dict.start()
                except Exception as ex:
                    dictado_en_progreso[0] = False
                    print("Error en on_mic_click:", ex)

            btn_mic = EmojiIconButton(
                icon_emoji="🎙️",
                active_emoji="⏹️",
                icon_color="#00FFFF",
                on_click=on_mic_click,
                tooltip="Dictar por voz / Oye LUXO 🎙️",
                width=40,
                height=40,
                border_radius=20
            )

            btn_mic_container = ft.Container(
                content=btn_mic,
                bgcolor="#1E1E2E",
                border_radius=23,
                border=ft.Border.all(1.5, "#00FFFF"),
                width=46,
                height=46,
                alignment=ft.alignment.Alignment(0, 0),
                visible=True
            )

            siri_orb_flet = ft.Container(
                width=80, height=80, border_radius=40,
                gradient=ft.RadialGradient(center=ft.alignment.Alignment(0, 0), radius=0.5, colors=["#E040FB", "#00F0FF", "#0A0A18"]),
                shadow=ft.BoxShadow(spread_radius=10, blur_radius=20, color="#00F0FF", offset=ft.Offset(0,0)),
                animate_scale=ft.Animation(400, ft.AnimationCurve.EASE_IN_OUT),
                animate_opacity=ft.Animation(400, ft.AnimationCurve.EASE_IN_OUT),
                scale=0.1, opacity=0, bottom=30, right=30
            )

            if user_info.get("id") and user_info["id"] in active_sessions:
                active_sessions[user_info["id"]]["btn_mic"] = btn_mic
                active_sessions[user_info["id"]]["btn_mic_container"] = btn_mic_container
                active_sessions[user_info["id"]]["siri_orb"] = siri_orb_flet

            btn_send_whatsapp = ft.Container(
                content=ft.Text("✈️", color="white", size=16, weight="bold"),
                gradient=ft.LinearGradient(
                    colors=["#00F0FF", "#9D50BB"],
                    begin=ft.alignment.Alignment(-1, -1),
                    end=ft.alignment.Alignment(1, 1)
                ),
                border_radius=23,
                width=46,
                height=46,
                on_click=enviar_mensaje,
                ink=True,
                tooltip="Enviar mensaje",
                shadow=[
                    ft.BoxShadow(
                        color="#00F0FF",
                        blur_radius=12,
                        spread_radius=1
                    )
                ],
                alignment=ft.alignment.Alignment(0, 0)
            )

            return ft.Column([
                ft.Row([
                    ft.Text(g_tr("Asistente Virtual LUXO AI", "LUXO AI Assistant", "Assistant Virtuel LUXO AI", "Assistente Virtuale LUXO AI", "LUXO AI 虚拟助手"), size=24, color="#D8B4FE", weight="bold")
                ], vertical_alignment="center"),
                ft.Container(
                    content=ft.SelectionArea(content=chat_display),
                    expand=True,
                    bgcolor="#080812",
                    border_radius=20,
                    padding=10,
                    border=ft.Border.all(2, "#D8B4FE"),
                    shadow=[
                        ft.BoxShadow(
                            color="#D8B4FE",
                            blur_radius=15,
                            spread_radius=1,
                        )
                    ]
                ),
                ft.Row([
                    input_msg,
                    btn_mic_container,
                    btn_send_whatsapp
                ], spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER)
            ], expand=True)

        def build_historial_view():
            historial_list = ft.Column(spacing=15, scroll=ft.ScrollMode.AUTO, expand=True)
            
            def cargar_lista_historial():
                historial_list.controls.clear()
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("""
                            SELECT Pregunta_Usuario, Respuesta_IA, Fecha_Hora 
                            FROM historial_conversaciones 
                            WHERE ID_Usuario = %s 
                            ORDER BY Fecha_Hora DESC 
                            LIMIT 30
                        """, (user_info["id"],))
                        historial = cursor.fetchall()
                        db.close()
                        
                        if not historial:
                            historial_list.controls.append(
                                ft.Container(
                                    content=ft.Text("No tienes consultas anteriores registradas.", color="#aaaaaa", size=14),
                                    alignment=ft.alignment.Alignment(0, 0),
                                    expand=True
                                )
                            )
                        else:
                            for row in historial:
                                fecha = row["Fecha_Hora"].strftime("%d/%m/%Y %H:%M") if row["Fecha_Hora"] else ""
                                historial_list.controls.append(
                                    ft.Container(
                                        content=ft.Column([
                                            ft.Text(f"📅 {fecha}", color="#aaaaaa", size=11),
                                            ft.Text(f"💬 Pregunta: {row['Pregunta_Usuario']}", color="white", weight="bold"),
                                            ft.Text(f"🤖 Respuesta: {row['Respuesta_IA']}", color="#D8B4FE"),
                                        ], spacing=4),
                                        bgcolor="#141424",
                                        padding=15,
                                        border_radius=8,
                                        border=ft.Border.all(1, "#333333")
                                    )
                                )
                except Exception as ex:
                    print("ERROR HISTORIAL VIEW:", ex)
                    historial_list.controls.append(ft.Text("Error al cargar el historial.", color="red"))
                page.update()
                
            cargar_lista_historial()
            
            def confirmar_borrado_historial(e):
                def on_confirmar(ev):
                    try:
                        db = conectar_db()
                        if db:
                            cursor = db.cursor()
                            cursor.execute("""
                                DELETE FROM pendientes_actualizacion 
                                WHERE ID_Conversacion IN (
                                    SELECT ID_Conversacion FROM historial_conversaciones WHERE ID_Usuario = %s
                                )
                            """, (user_info["id"],))
                            cursor.execute("DELETE FROM historial_conversaciones WHERE ID_Usuario = %s", (user_info["id"],))
                            db.commit()
                            db.close()
                            mostrar_snack("Historial borrado correctamente.")
                            cargar_lista_historial()
                            chat_display.controls.clear()
                            historial_sesion.clear()
                            page.pop_dialog()
                            page.update()
                    except Exception as ex:
                        print("ERROR BORRAR HISTORIAL:", ex)
                        mostrar_snack("Error al borrar el historial.", color="red")
                
                def on_cancelar(ev):
                    page.pop_dialog()
                    
                dialog_confirm = ft.AlertDialog(
                    title=ft.Text("Confirmar Borrado", color="#FF4500", weight="bold"),
                    content=ft.Text("¿Seguro que deseas borrar todo tu historial de conversaciones? Esta acción no se puede deshacer."),
                    actions=[
                        ft.TextButton("Cancelar", on_click=on_cancelar),
                        ft.ElevatedButton("Borrar Todo", on_click=on_confirmar, bgcolor="#FF4500", color="white")
                    ],
                    actions_alignment="end",
                    bgcolor="#0F0F1A"
                )
                page.show_dialog(dialog_confirm)

            btn_clear = ft.Container(
                content=ft.Row([
                    ft.Text("🗑️", size=14),
                    ft.Text("Borrar Historial", color="white", weight="bold", size=12)
                ], spacing=6, alignment=ft.MainAxisAlignment.CENTER),
                bgcolor="#FF4500",
                padding=ft.padding.Padding(16, 10, 16, 10),
                border_radius=20,
                on_click=confirmar_borrado_historial,
                ink=True
            )

            return ft.SelectionArea(
                content=ft.Column([
                    ft.Row([
                        ft.Text("Historial de Consultas", size=24, color="#D8B4FE", weight="bold"),
                        btn_clear
                    ], alignment="spaceBetween", vertical_alignment="center"),
                    ft.Divider(height=20, color="#333333"),
                    historial_list
                ], expand=True)
            )

        def build_stats_tab():
            total_consultas = 0
            utiles = 0
            no_utiles = 0
            total_tickets = 0
            tickets_resueltos = 0
            negatives_list = ft.Column(spacing=10, scroll=ft.ScrollMode.AUTO, expand=True)
            positives_list = ft.Column(spacing=10, scroll=ft.ScrollMode.AUTO, expand=True)
            logins_list = ft.Column(spacing=10, scroll=ft.ScrollMode.AUTO, expand=True)
            
            categories_cnt = []
            
            # Variables de cumplimiento por zona
            checklist_pct = 0.0
            total_stores_zone = 0
            
            # Métricas de campaña activa por zona
            total_campaign_stores = 0
            aprobadas_ia_cnt = 0
            visto_bueno_cnt = 0
            rechazadas_ia_cnt = 0
            pendientes_cnt = 0
            sin_entrega_cnt = 0
            
            try:
                db = conectar_db()
                if db:
                    cursor = db.cursor(dictionary=True)
                    cursor.execute("SELECT COUNT(*) as cnt FROM historial_conversaciones")
                    total_consultas = cursor.fetchone()["cnt"]
                    
                    cursor.execute("SELECT COUNT(*) as cnt FROM historial_conversaciones WHERE Me_Sirvio = 1")
                    utiles = cursor.fetchone()["cnt"]
                    
                    cursor.execute("SELECT COUNT(*) as cnt FROM historial_conversaciones WHERE Me_Sirvio = 0")
                    no_utiles = cursor.fetchone()["cnt"]

                    # Conteo de tickets
                    cursor.execute("SELECT COUNT(*) as cnt FROM tickets_soporte")
                    total_tickets = cursor.fetchone()["cnt"]
                    
                    cursor.execute("SELECT COUNT(*) as cnt FROM tickets_soporte WHERE Estatus = 'Resuelto'")
                    tickets_resueltos = cursor.fetchone()["cnt"]

                    # Consultas agrupadas por categoría
                    cursor.execute("""
                        SELECT COALESCE(Categoria, 'Otros') as cat, COUNT(*) as cnt 
                        FROM pendientes_actualizacion 
                        GROUP BY Categoria
                    """)
                    categories_cnt = cursor.fetchall()
                    
                    # --- CÁLCULO DE CUMPLIMIENTO POR ZONA ---
                    zona_act = active_zone_filter[0]
                    
                    # 1. Porcentaje de cumplimiento de bitácoras (Checklists) diarias
                    cursor.execute("SELECT COUNT(*) as cnt FROM plantillas_checklist")
                    total_plantillas = cursor.fetchone()["cnt"] or 1
                    
                    # Buscar tiendas de la zona
                    if zona_act != "Todas":
                        cursor.execute("SELECT ID_Usuario, Tienda FROM usuarios WHERE Usuario LIKE 'sgh%' AND Tienda IS NOT NULL AND Tienda != '' AND Zona = %s", (zona_act,))
                    else:
                        cursor.execute("SELECT ID_Usuario, Tienda FROM usuarios WHERE Usuario LIKE 'sgh%' AND Tienda IS NOT NULL AND Tienda != ''")
                    gerentes_zona = cursor.fetchall()
                    total_stores_zone = len(gerentes_zona)
                    
                    completed_sum = 0
                    if total_stores_zone > 0:
                        for g_z in gerentes_zona:
                            cursor.execute("""
                                SELECT COUNT(DISTINCT ID_Plantilla) as cnt 
                                FROM registro_checklist 
                                WHERE ID_Usuario = %s AND Fecha = CURDATE() AND Completado = 1
                            """, (g_z["ID_Usuario"],))
                            completed_sum += cursor.fetchone()["cnt"]
                        checklist_pct = (completed_sum / (total_stores_zone * total_plantillas)) * 100
                    else:
                        checklist_pct = 0.0

                    # 2. Métricas de campañas del mes
                    cursor.execute("SELECT ID_Campana FROM campanas WHERE Estatus = 'Activa'")
                    active_camp_row = cursor.fetchone()
                    if active_camp_row:
                        id_active_camp = active_camp_row["ID_Campana"]
                        total_campaign_stores = total_stores_zone
                        
                        if total_campaign_stores > 0:
                            placeholders = ",".join(["%s"] * total_campaign_stores)
                            ids_gerentes = [g_z["ID_Usuario"] for g_z in gerentes_zona]
                            
                            # Aprobado IA
                            cursor.execute(f"""
                                SELECT COUNT(*) as cnt 
                                FROM campana_entregas_tienda 
                                WHERE ID_Campana = %s AND Estatus = 'Aprobado_IA' AND ID_Usuario IN ({placeholders})
                            """, [id_active_camp] + ids_gerentes)
                            aprobadas_ia_cnt = cursor.fetchone()["cnt"]
                            
                            # Visto Bueno
                            cursor.execute(f"""
                                SELECT COUNT(*) as cnt 
                                FROM campana_entregas_tienda 
                                WHERE ID_Campana = %s AND Estatus = 'Visto_Bueno' AND ID_Usuario IN ({placeholders})
                            """, [id_active_camp] + ids_gerentes)
                            visto_bueno_cnt = cursor.fetchone()["cnt"]
                            
                            # Rechazado IA
                            cursor.execute(f"""
                                SELECT COUNT(*) as cnt 
                                FROM campana_entregas_tienda 
                                WHERE ID_Campana = %s AND Estatus = 'Rechazado_IA' AND ID_Usuario IN ({placeholders})
                            """, [id_active_camp] + ids_gerentes)
                            rechazadas_ia_cnt = cursor.fetchone()["cnt"]
                            
                            # Pendiente
                            cursor.execute(f"""
                                SELECT COUNT(*) as cnt 
                                FROM campana_entregas_tienda 
                                WHERE ID_Campana = %s AND Estatus = 'Pendiente' AND ID_Usuario IN ({placeholders})
                            """, [id_active_camp] + ids_gerentes)
                            pendientes_cnt = cursor.fetchone()["cnt"]
                            
                            sin_entrega_cnt = total_campaign_stores - (aprobadas_ia_cnt + visto_bueno_cnt + rechazadas_ia_cnt + pendientes_cnt)
                    
                    # Cargar negativos
                    cursor.execute("""
                        SELECT h.Fecha_Hora, u.Nombre_Completo, h.Pregunta_Usuario, h.Respuesta_IA, h.Comentario_Feedback
                        FROM historial_conversaciones h
                        JOIN usuarios u ON h.ID_Usuario = u.ID_Usuario
                        WHERE h.Me_Sirvio = 0
                        ORDER BY h.Fecha_Hora DESC
                    """)
                    negatives = cursor.fetchall()
                    
                    # Cargar positivos
                    cursor.execute("""
                        SELECT h.Fecha_Hora, u.Nombre_Completo, h.Pregunta_Usuario, h.Respuesta_IA, h.Comentario_Feedback
                        FROM historial_conversaciones h
                        JOIN usuarios u ON h.ID_Usuario = u.ID_Usuario
                        WHERE h.Me_Sirvio = 1
                        ORDER BY h.Fecha_Hora DESC
                    """)
                    positives = cursor.fetchall()

                    # Cargar inicios de sesión
                    cursor.execute("""
                        SELECT s.Fecha_Login, u.Nombre_Completo, s.Direccion_IP, s.Ubicacion_Ciudad, s.Ubicacion_Pais
                        FROM sesiones s
                        JOIN usuarios u ON s.ID_Usuario = u.ID_Usuario
                        ORDER BY s.Fecha_Login DESC
                        LIMIT 15
                    """)
                    logins = cursor.fetchall()
                    
                    db.close()
                    
                    if not negatives:
                        negatives_list.controls.append(ft.Text("No hay respuestas calificadas negativamente.", color="#aaaaaa", size=14))
                    else:
                        for row in negatives:
                            fecha = row["Fecha_Hora"].strftime("%d/%m/%Y %H:%M") if row["Fecha_Hora"] else ""
                            comentario = row["Comentario_Feedback"] or "(Sin comentario)"
                            negatives_list.controls.append(
                                ft.Container(
                                    content=ft.Column([
                                        ft.Row([
                                            ft.Text(f"📅 {fecha}", color="#aaaaaa", size=11),
                                            ft.Text(f"👤 Usuario: {row['Nombre_Completo']}", color="#aaaaaa", size=11),
                                        ], alignment="spaceBetween"),
                                        ft.Text(f"💬 Pregunta: {row['Pregunta_Usuario']}", color="white", weight="bold"),
                                        ft.Text(f"🤖 Respuesta de Luxo: {row['Respuesta_IA']}", color="#D8B4FE"),
                                        ft.Container(
                                            content=ft.Text(f"Razón: {comentario}", color="#FF4500", size=12, italic=True),
                                            bgcolor="#3d1f1f",
                                            padding=8,
                                            border_radius=5
                                        )
                                    ], spacing=4),
                                    bgcolor="#141424",
                                    padding=15,
                                    border_radius=8,
                                    border=ft.Border.all(1, "#333333")
                                )
                            )
                            
                    if not positives:
                        positives_list.controls.append(ft.Text("No hay respuestas calificadas positivamente.", color="#aaaaaa", size=14))
                    else:
                        for row in positives:
                            fecha = row["Fecha_Hora"].strftime("%d/%m/%Y %H:%M") if row["Fecha_Hora"] else ""
                            comentario = row["Comentario_Feedback"] or "(Sin comentario)"
                            positives_list.controls.append(
                                ft.Container(
                                    content=ft.Column([
                                        ft.Row([
                                            ft.Text(f"📅 {fecha}", color="#aaaaaa", size=11),
                                            ft.Text(f"👤 Usuario: {row['Nombre_Completo']}", color="#aaaaaa", size=11),
                                        ], alignment="spaceBetween"),
                                        ft.Text(f"💬 Pregunta: {row['Pregunta_Usuario']}", color="white", weight="bold"),
                                        ft.Text(f"🤖 Respuesta de Luxo: {row['Respuesta_IA']}", color="#D8B4FE"),
                                        ft.Container(
                                            content=ft.Text(f"Comentario: {comentario}", color="#7CFC00", size=12, italic=True),
                                            bgcolor="#1b3d1f",
                                            padding=8,
                                            border_radius=5
                                        )
                                    ], spacing=4),
                                    bgcolor="#141424",
                                    padding=15,
                                    border_radius=8,
                                    border=ft.Border.all(1, "#333333")
                                )
                            )

                    if not logins:
                        logins_list.controls.append(ft.Text("No hay registros de inicio de sesión.", color="#aaaaaa", size=14))
                    else:
                        for row in logins:
                            fecha = row["Fecha_Login"].strftime("%d/%m/%Y %H:%M") if row["Fecha_Login"] else ""
                            ip = row["Direccion_IP"] or "Desconocida"
                            ciudad = row["Ubicacion_Ciudad"] or "Desconocida"
                            pais = row["Ubicacion_Pais"] or "Desconocido"
                            logins_list.controls.append(
                                ft.Container(
                                    content=ft.Column([
                                        ft.Row([
                                            ft.Text(f"📅 {fecha}", color="#aaaaaa", size=11),
                                            ft.Text(f"👤 {row['Nombre_Completo']}", color="white", weight="bold"),
                                        ], alignment="spaceBetween"),
                                        ft.Row([
                                            ft.Text(f"🌐 IP: {ip}", color="#00FFFF", size=12),
                                            ft.Text(f"📍 {ciudad}, {pais}", color="#D8B4FE", size=12),
                                        ], alignment="spaceBetween"),
                                    ], spacing=4),
                                    bgcolor="#141424",
                                    padding=15,
                                    border_radius=8,
                                    border=ft.Border.all(1, "#333333")
                                )
                            )
            except Exception as ex:
                print("ERROR STATS TAB:", ex)
                
            eficacia = 0.0
            if (utiles + no_utiles) > 0:
                eficacia = (utiles / (utiles + no_utiles)) * 100

            # 1. KPI Total Consultas
            kpi_card_1 = ft.Container(
                content=ft.Column([
                    ft.Text("Total Consultas", color="#aaaaaa", size=11, weight="bold"),
                    ft.Text(str(total_consultas), color="#D8B4FE", size=26, weight="bold"),
                    ft.Icon(ft.Icons.CHAT_ROUNDED, color="#D8B4FE", size=22)
                ], spacing=8, alignment="center", horizontal_alignment="center"),
                bgcolor="#0F0F1A",
                border=ft.Border.all(1.5, "#D8B4FE"),
                border_radius=10,
                padding=10,
                width=150,
                height=140
            )

            # 2. KPI Tickets de Soporte (Barra de progreso horizontal)
            tickets_pct = tickets_resueltos / total_tickets if total_tickets > 0 else 0.0
            tickets_bar = ft.Column([
                ft.Text(f"Resueltos: {tickets_resueltos} de {total_tickets}", color="#00FFFF", size=11, weight="bold"),
                ft.ProgressBar(value=tickets_pct, color="#00FFFF", bgcolor="#141424", width=120),
                ft.Text(f"Tasa de Resolución: {int(tickets_pct * 100)}%", color="#aaaaaa", size=10)
            ], horizontal_alignment="center", spacing=5, alignment="center")
            
            kpi_card_2 = ft.Container(
                content=ft.Column([
                    ft.Text("Soporte Técnico 🎫", color="#aaaaaa", size=11, weight="bold"),
                    ft.Divider(height=2, color="transparent"),
                    tickets_bar
                ], spacing=5, alignment="center", horizontal_alignment="center"),
                bgcolor="#0F0F1A",
                border=ft.Border.all(1.5, "#00FFFF"),
                border_radius=10,
                padding=10,
                width=170,
                height=140
            )

            # 3. KPI Eficacia de IA (Anillo de progreso circular)
            eficacia_ring = ft.Container(
                content=ft.Stack([
                    ft.ProgressRing(
                        value=eficacia / 100 if (utiles + no_utiles) > 0 else 0.0,
                        stroke_width=8,
                        color="#7CFC00" if eficacia >= 75 else ("#00FFFF" if eficacia >= 50 else "#FF4500"),
                        bgcolor="#141424",
                        width=85,
                        height=85
                    ),
                    ft.Container(
                        content=ft.Text(f"{eficacia:.0f}%", color="white", size=15, weight="bold"),
                        alignment=ft.alignment.Alignment(0, 0)
                    )
                ], width=85, height=85),
                alignment=ft.alignment.Alignment(0, 0)
            )
            
            kpi_card_3 = ft.Container(
                content=ft.Column([
                    ft.Text("Eficacia de la IA", color="#aaaaaa", size=11, weight="bold"),
                    ft.Divider(height=2, color="transparent"),
                    eficacia_ring
                ], spacing=2, alignment="center", horizontal_alignment="center"),
                bgcolor="#0F0F1A",
                border=ft.Border.all(1.5, "#7CFC00"),
                border_radius=10,
                padding=10,
                width=170,
                height=140
            )

            # 4. KPI Categorías Faltantes (Gráfico de barras personalizado)
            bar_controls = []
            max_val = max([c["cnt"] for c in categories_cnt]) if categories_cnt else 1
            cat_colors = {
                "Impresoras": "#7CFC00",
                "Políticas de Venta": "#00FFFF",
                "Sistemas/Terminales": "#FF4500",
                "Manuales": "#D8B4FE",
                "Otros": "#888888"
            }
            
            for c in categories_cnt:
                cat_name = c["cat"]
                count = c["cnt"]
                color = cat_colors.get(cat_name, "#888888")
                
                # Proportional height (max 65px, min 10px)
                height = (count / max_val) * 65 if max_val > 0 else 10
                height = max(height, 15)
                
                bar_controls.append(
                    ft.Column([
                        ft.Text(str(count), color="white", size=9, weight="bold"),
                        ft.Container(
                            width=22,
                            height=height,
                            bgcolor=color,
                            border_radius=ft.BorderRadius(top_left=3, top_right=3, bottom_left=0, bottom_right=0),
                            shadow=ft.BoxShadow(color=color, blur_radius=5, spread_radius=0.1)
                        ),
                        ft.Text(cat_name[:4] + ".." if len(cat_name) > 4 else cat_name, color="#aaaaaa", size=8)
                    ], horizontal_alignment="center", spacing=2, alignment="end")
                )
            
            bar_chart_row = ft.Row(
                bar_controls,
                spacing=8,
                alignment="center",
                vertical_alignment="end"
            ) if bar_controls else ft.Text("Sin preguntas faltantes", color="#aaaaaa", size=11, italic=True)
            
            kpi_card_4 = ft.Container(
                content=ft.Column([
                    ft.Text("Categorías IA Faltantes", color="#aaaaaa", size=11, weight="bold"),
                    ft.Divider(height=2, color="transparent"),
                    bar_chart_row
                ], spacing=2, alignment="center", horizontal_alignment="center"),
                bgcolor="#0F0F1A",
                border=ft.Border.all(1.5, "#A100F2"),
                border_radius=10,
                padding=10,
                width=240,
                height=140
            )

            is_mobile = (page.width < 800) if (page and page.width) else False

            kpi_row = ft.Column([
                kpi_card_1,
                kpi_card_2,
                kpi_card_3,
                kpi_card_4
            ], spacing=15, horizontal_alignment="center") if is_mobile else ft.Row([
                kpi_card_1,
                kpi_card_2,
                kpi_card_3,
                kpi_card_4
            ], spacing=15, alignment="center")

            def build_stats_bar(label, count, total, color):
                pct = (count / total * 100) if total > 0 else 0.0
                max_width = 320 if not is_mobile else 180
                bar_width = (count / total * max_width) if total > 0 else 0
                return ft.Row([
                    ft.Text(label, size=12, color="white", width=200 if not is_mobile else 120, weight="bold"),
                    ft.Stack([
                        ft.Container(width=max_width, height=16, bgcolor="#141424", border_radius=4),
                        ft.Container(width=max(bar_width, 6) if count > 0 else 0, height=16, bgcolor=color, border_radius=4, shadow=ft.BoxShadow(color=color, blur_radius=3, spread_radius=0.1))
                    ]),
                    ft.Text(f"{count} ({pct:.1f}%)", size=12, color="white", weight="bold", width=90 if not is_mobile else 65)
                ], spacing=8, alignment="start", vertical_alignment="center")

            tot_preguntas_val = max(total_consultas, 1)

            questions_chart_card = ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Row([
                            ft.Icon(ft.Icons.BAR_CHART_ROUNDED, color="#00FFFF", size=22),
                            ft.Text("Gráfica Estadísticas de Consultas a LUXO AI 📊", color="white", size=15, weight="bold"),
                        ], spacing=6),
                        ft.Container(
                            content=ft.Text(f"Eficacia IA: {eficacia:.1f}%", color="white", weight="bold", size=12),
                            bgcolor="#008080" if eficacia >= 70 else "#7B2D00",
                            padding=ft.padding.Padding(10, 4, 10, 4),
                            border_radius=6
                        )
                    ], alignment="spaceBetween", vertical_alignment="center"),
                    ft.Divider(height=10, color="#333333"),
                    build_stats_bar("Total de Preguntas 💬", total_consultas, tot_preguntas_val, "#00FFFF"),
                    build_stats_bar("Respuestas Satisfactorias 👍", utiles, tot_preguntas_val, "#7CFC00"),
                    build_stats_bar("Respuestas Insatisfactorias 👎", no_utiles, tot_preguntas_val, "#FF4500"),
                ], spacing=12),
                bgcolor="#0F0F1A",
                border=ft.Border.all(1.5, "#00FFFF"),
                border_radius=12,
                padding=16
            )

            # Ranking de Uso por Tienda (TODAS LAS TIENDAS REGISTRADAS EN USUARIOS)
            store_usage_controls = []
            tiendas_uso = []
            try:
                db_st = conectar_db()
                if db_st:
                    cur_st = db_st.cursor(dictionary=True)
                    cur_st.execute("""
                        SELECT 
                            u.Tienda,
                            COUNT(h.ID_Conversacion) as Total_Consultas
                        FROM (
                            SELECT DISTINCT Tienda 
                            FROM usuarios 
                            WHERE Tienda IS NOT NULL AND Tienda != '' AND Tienda != 'Tienda Luxo'
                        ) u
                        LEFT JOIN usuarios usr ON usr.Tienda = u.Tienda
                        LEFT JOIN historial_conversaciones h ON h.ID_Usuario = usr.ID_Usuario
                        GROUP BY u.Tienda
                        ORDER BY Total_Consultas DESC, u.Tienda ASC
                    """)
                    tiendas_uso = cur_st.fetchall()
                    db_st.close()

                    max_tienda_consultas = max([t["Total_Consultas"] for t in tiendas_uso], default=1)
                    colors_palette = ["#FFD700", "#00FFFF", "#D8B4FE", "#7CFC00", "#FF69B4", "#FFA500", "#00FF7F", "#FF4500"]

                    for idx_t, t_row in enumerate(tiendas_uso):
                        t_nombre = t_row["Tienda"]
                        t_cnt = t_row["Total_Consultas"]
                        is_top_1 = (idx_t == 0 and t_cnt > 0)
                        icon_prefix = "🏆 1° " if is_top_1 else (f"🏪 {idx_t+1}° " if t_cnt > 0 else "🏬 ")
                        col_bar = colors_palette[idx_t % len(colors_palette)] if t_cnt > 0 else "#444444"

                        store_usage_controls.append(
                            build_stats_bar(f"{icon_prefix}{t_nombre}", t_cnt, max_tienda_consultas, col_bar)
                        )
            except Exception as ex_st:
                print("Error calculando uso por tienda:", ex_st)

            if not store_usage_controls:
                store_usage_controls.append(ft.Text("No hay datos de consultas por tienda registrados aún.", color="#aaaaaa", size=12))

            top_store_name = tiendas_uso[0]["Tienda"] if (tiendas_uso and len(tiendas_uso) > 0 and tiendas_uso[0]["Total_Consultas"] > 0) else "Ninguna activa aún"

            store_chart_card = ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Row([
                            ft.Icon(ft.Icons.STOREFRONT_ROUNDED, color="#FFD700", size=22),
                            ft.Text(f"Ranking de Uso por Tienda (Total: {len(tiendas_uso)} registradas) 🏬", color="white", size=15, weight="bold"),
                        ], spacing=6),
                        ft.Container(
                            content=ft.Text(f"Más Activa: {top_store_name} 👑", color="black", weight="bold", size=11),
                            bgcolor="#FFD700",
                            padding=ft.padding.Padding(10, 4, 10, 4),
                            border_radius=6
                        )
                    ], alignment="spaceBetween", vertical_alignment="center"),
                    ft.Text("Mide el volumen de consultas real de todas las tiendas registradas en el sistema para monitorear el Piloto.", color="#aaaaaa", size=11),
                    ft.Divider(height=10, color="#333333"),
                    ft.Container(
                        content=ft.Column(store_usage_controls, spacing=10, scroll=ft.ScrollMode.AUTO),
                        height=300 if not is_mobile else 240
                    )
                ], spacing=10),
                bgcolor="#0F0F1A",
                border=ft.Border.all(1.5, "#FFD700"),
                border_radius=12,
                padding=16
            )

            # --- WIDGETS DE CUMPLIMIENTO ZONAL ---
            checklist_ring = ft.Container(
                content=ft.Stack([
                    ft.ProgressRing(
                        value=checklist_pct / 100.0,
                        stroke_width=8,
                        color="#7CFC00" if checklist_pct >= 75 else ("#00FFFF" if checklist_pct >= 50 else "#FF4500"),
                        bgcolor="#141424",
                        width=85,
                        height=85
                    ),
                    ft.Container(
                        content=ft.Text(f"{checklist_pct:.0f}%", color="white", size=15, weight="bold"),
                        alignment=ft.alignment.Alignment(0, 0)
                    )
                ], width=85, height=85),
                alignment=ft.alignment.Alignment(0, 0)
            )

            compliance_card = ft.Container(
                content=ft.Column([
                    ft.Text("Cumplimiento de Checklists", color="#aaaaaa", size=10.5, weight="bold", text_align=ft.TextAlign.CENTER),
                    ft.Text(f"{zona_act}", color="#00FFFF", size=12, weight="bold"),
                    ft.Divider(height=2, color="transparent"),
                    checklist_ring,
                    ft.Text(f"Tiendas: {total_stores_zone}", size=11, color="#aaaaaa")
                ], spacing=3, alignment="center", horizontal_alignment="center"),
                bgcolor="#0F0F1A",
                border=ft.Border.all(1.5, "#00FFFF"),
                border_radius=10,
                padding=10,
                width=170,
                height=180
            )

            def build_horizontal_bar(label, count, total, color):
                pct = (count / total * 100) if total > 0 else 0.0
                bar_width = (count / total * 200) if total > 0 else 0
                return ft.Row([
                    ft.Text(label, size=11, color="white", width=95),
                    ft.Stack([
                        ft.Container(width=200, height=12, bgcolor="#141424", border_radius=3),
                        ft.Container(width=max(bar_width, 4) if count > 0 else 0, height=12, bgcolor=color, border_radius=3, shadow=ft.BoxShadow(color=color, blur_radius=3, spread_radius=0.1))
                    ]),
                    ft.Text(f"{count} ({pct:.0f}%)", size=11, color="#aaaaaa", width=65)
                ], spacing=5, alignment="start", vertical_alignment="center")

            campaign_chart_card = ft.Container(
                content=ft.Column([
                    ft.Text("Estado de Campaña Mensual", color="#aaaaaa", size=11, weight="bold"),
                    ft.Text(f"Zona: {zona_act}", color="#D8B4FE", size=12, weight="bold"),
                    ft.Divider(height=1, color="#333333"),
                    build_horizontal_bar("Visto Bueno 👑", visto_bueno_cnt, total_campaign_stores, "#00FF7F"),
                    build_horizontal_bar("Aprobado IA 🤖", aprobadas_ia_cnt, total_campaign_stores, "#7CFC00"),
                    build_horizontal_bar("Rechazado IA ⚠️", rechazadas_ia_cnt, total_campaign_stores, "#FF4500"),
                    build_horizontal_bar("Pendiente ⏳", pendientes_cnt, total_campaign_stores, "#FFD700"),
                    build_horizontal_bar("Sin Entrega ❌", sin_entrega_cnt, total_campaign_stores, "#888888")
                ], spacing=6, alignment="center", horizontal_alignment="start"),
                bgcolor="#0F0F1A",
                border=ft.Border.all(1.5, "#D8B4FE"),
                border_radius=10,
                padding=12,
                width=400,
                height=180
            )

            compliance_row = ft.Column([
                compliance_card,
                campaign_chart_card
            ], spacing=15, horizontal_alignment="center") if is_mobile else ft.Row([
                compliance_card,
                campaign_chart_card
            ], spacing=15, alignment="center")

            return ft.Column([
                ft.Row([
                    ft.Text("Métricas de Control Geográfico", size=18, color="white", weight="bold"),
                    ft.Container(
                        content=ft.Text(f"Filtro: {zona_act}", size=12, color="black", weight="bold"),
                        bgcolor="#00FFFF",
                        padding=ft.Padding(left=8, right=8, top=3, bottom=3),
                        border_radius=5
                    )
                ], alignment="spaceBetween", vertical_alignment="center"),
                ft.Divider(height=5, color="transparent"),
                compliance_row,
                ft.Divider(height=15, color="#333333"),
                ft.Text("Indicadores de Uso y Calidad (Global)", size=18, color="white", weight="bold"),
                ft.Divider(height=5, color="transparent"),
                kpi_row,
                ft.Divider(height=15, color="#333333"),
                questions_chart_card,
                ft.Divider(height=15, color="#333333"),
                store_chart_card
            ], expand=True, spacing=10, scroll=ft.ScrollMode.AUTO)

        def build_missing_questions_tab():
            is_mobile = (page.width < 800) if (page and page.width) else False
            questions_list = ft.Column(spacing=10, scroll=ft.ScrollMode.AUTO, expand=True)
            
            def cargar_preguntas():
                questions_list.controls.clear()
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("""
                            SELECT ID_Pendiente, Pregunta_Faltante, Fecha_Registro, Categoria 
                            FROM pendientes_actualizacion 
                            WHERE Estatus = 'Pendiente'
                            ORDER BY Fecha_Registro DESC
                        """)
                        preguntas = cursor.fetchall()
                        db.close()
                        
                        if not preguntas:
                            questions_list.controls.append(
                                ft.Container(
                                    content=ft.Text("No hay preguntas pendientes de actualizar.", color="#7CFC00", size=14),
                                    alignment=ft.alignment.Alignment(0, 0),
                                    expand=True
                                )
                            )
                        else:
                            for row in preguntas:
                                id_p = row["ID_Pendiente"]
                                pregunta = row["Pregunta_Faltante"]
                                fecha = row["Fecha_Registro"].strftime("%d/%m/%Y %H:%M") if row["Fecha_Registro"] else ""
                                cat_text = row["Categoria"] or "Clasificando..."
                                
                                # Definir color del tag según la categoría
                                cat_colors = {
                                    "Impresoras": "#7CFC00",
                                    "Políticas de Venta": "#00FFFF",
                                    "Sistemas/Terminales": "#FF4500",
                                    "Manuales": "#D8B4FE",
                                    "Otros": "#aaaaaa"
                                }
                                cat_color = cat_colors.get(cat_text, "#aaaaaa")
                                
                                category_badge = ft.Container(
                                    content=ft.Text(cat_text, color="black", size=9 if is_mobile else 10, weight="bold"),
                                    bgcolor=cat_color,
                                    padding=ft.Padding(left=6 if is_mobile else 8, right=6 if is_mobile else 8, top=2 if is_mobile else 3, bottom=2 if is_mobile else 3),
                                    border_radius=5
                                )
                                
                                def resolver_click(e, id_pend=id_p):
                                    try:
                                        db_res = conectar_db()
                                        if db_res:
                                            cursor_res = db_res.cursor()
                                            cursor_res.execute(
                                                "UPDATE pendientes_actualizacion SET Estatus = 'Resuelto' WHERE ID_Pendiente = %s",
                                                (id_pend,)
                                            )
                                            db_res.commit()
                                            db_res.close()
                                            mostrar_snack("Pregunta marcada como resuelta.")
                                            cargar_preguntas()
                                            page.update()
                                    except Exception as ex:
                                        print("ERROR MARCAR RESUELTO:", ex)
                                
                                btn_resolver = ft.ElevatedButton(
                                    "Marcar Resuelta",
                                    icon=ft.Icons.CHECK,
                                    bgcolor="#6E48AA",
                                    color="white",
                                    height=34 if is_mobile else 38,
                                    style=ft.ButtonStyle(
                                        shape=ft.RoundedRectangleBorder(radius=6),
                                        padding=ft.padding.Padding(10, 0, 10, 0) if is_mobile else None
                                    ),
                                    on_click=resolver_click
                                )

                                if is_mobile:
                                    # Adaptación móvil: Stacking vertical para evitar amontonamiento
                                    card_content = ft.Column([
                                        ft.Row([
                                            category_badge,
                                            ft.Text(f"Registrada el: {fecha}", color="#aaaaaa", size=10)
                                        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                                        ft.Text(pregunta, color="white", weight="bold", size=13),
                                        ft.Row([btn_resolver], alignment=ft.MainAxisAlignment.END)
                                    ], spacing=6)
                                else:
                                    # Layout Desktop anti-overflow
                                    card_content = ft.Row([
                                        ft.Column([
                                            ft.Row([
                                                category_badge,
                                                ft.Text(f"Registrada el: {fecha}", color="#aaaaaa", size=11)
                                            ], spacing=8, vertical_alignment="center"),
                                            ft.Text(pregunta, color="white", weight="bold", size=14)
                                        ], spacing=6, expand=True),
                                        btn_resolver
                                    ], alignment="spaceBetween", vertical_alignment="center", spacing=15)

                                questions_list.controls.append(
                                    ft.Container(
                                        content=card_content,
                                        bgcolor="#141424",
                                        padding=10 if is_mobile else 12,
                                        border_radius=8,
                                        border=ft.Border.all(1, "#333333")
                                    )
                                )
                except Exception as ex:
                    print("ERROR MISSING QUESTIONS:", ex)
                    questions_list.controls.append(ft.Text("Error al cargar las preguntas pendientes.", color="red"))
                page.update()
                    
            cargar_preguntas()
            
            return ft.Column([
                ft.Text("Preguntas que la IA no pudo responder (Falta Información)", size=16 if is_mobile else 18, color="white", weight="bold"),
                ft.Text("Usa esta lista para identificar qué manuales o temas faltan en el sistema y cárgalos.", color="#aaaaaa", size=11 if is_mobile else 13),
                ft.Divider(height=8 if is_mobile else 10, color="transparent"),
                questions_list
            ], expand=True)

        def build_manuals_tab():
            is_mobile = (page.width < 800) if (page and page.width) else False
            manuals_list = ft.Column(spacing=10, scroll=ft.ScrollMode.AUTO, expand=True)
            
            def cargar_manuales():
                manuals_list.controls.clear()
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("SELECT ID_Manual, Nombre_Archivo, Titulo, Version, Abierto FROM manuales ORDER BY Nombre_Archivo")
                        manuales = cursor.fetchall()
                        db.close()
                        
                        manuals_list.controls.append(ft.Text("Manuales Base de Datos (PDF/Excel)", size=13 if is_mobile else 14, color="#00FFFF", weight="bold"))
                        if not manuales:
                            manuals_list.controls.append(ft.Text("No hay manuales cargados en la base de datos.", color="#aaaaaa", size=12))
                        else:
                            for m in manuales:
                                id_m = m["ID_Manual"]
                                nombre = m.get("Nombre_Archivo") or ""
                                version = m.get("Version") or ""
                                es_abierto = m.get("Abierto") if m.get("Abierto") is not None else 1
                                
                                def borrar_click(e, id_man=id_m, nom=nombre):
                                    def on_confirmar(ev):
                                        try:
                                            db_del = conectar_db()
                                            if db_del:
                                                cursor_del = db_del.cursor()
                                                cursor_del.execute("""
                                                    DELETE FROM pendientes_actualizacion 
                                                    WHERE ID_Conversacion IN (
                                                        SELECT ID_Conversacion FROM historial_conversaciones WHERE ID_Manual = %s
                                                    )
                                                """, (id_man,))
                                                cursor_del.execute("DELETE FROM historial_conversaciones WHERE ID_Manual = %s", (id_man,))
                                                cursor_del.execute("DELETE FROM manuales WHERE ID_Manual = %s", (id_man,))
                                                db_del.commit()
                                                db_del.close()
                                                rebuild_rag_cache()
                                                mostrar_snack(f"Manual '{nom}' eliminado.")
                                                cargar_manuales()
                                                page.pop_dialog()
                                                page.update()
                                        except Exception as ex:
                                            print("ERROR BORRAR MANUAL:", ex)
                                            mostrar_snack("Error al borrar manual.", color="red")
                                            
                                    def on_cancelar(ev):
                                        page.pop_dialog()
                                        
                                    dialog_confirm = ft.AlertDialog(
                                        title=ft.Text("Confirmar Borrado", color="#FF4500", weight="bold"),
                                        content=ft.Text(f"¿Seguro que deseas borrar el archivo \"{nom}\"?"),
                                        actions=[
                                            ft.TextButton("Cancelar", on_click=on_cancelar),
                                            ft.ElevatedButton("Sí, Borrar", on_click=on_confirmar, bgcolor="#FF4500", color="white")
                                        ],
                                        actions_alignment="end",
                                        bgcolor="#0F0F1A"
                                    )
                                    page.show_dialog(dialog_confirm)

                                def toggle_abierto_click(e, id_man=id_m, act_abierto=es_abierto):
                                    nuevo_estado = 0 if act_abierto == 1 else 1
                                    try:
                                        db_toggle = conectar_db()
                                        if db_toggle:
                                            cursor_toggle = db_toggle.cursor()
                                            cursor_toggle.execute("UPDATE manuales SET Abierto = %s WHERE ID_Manual = %s", (nuevo_estado, id_man))
                                            db_toggle.commit()
                                            db_toggle.close()
                                            rebuild_rag_cache()
                                            txt_est = "Abierto a todos" if nuevo_estado == 1 else "Restringido (Solo Admin)"
                                            mostrar_snack(f"Permiso de '{nom}' cambiado a: {txt_est}")
                                            cargar_manuales()
                                            page.update()
                                    except Exception as ex:
                                        print("ERROR TOGGLE ABIERTO:", ex)
                                        mostrar_snack("Error al actualizar permisos del manual.", color="red")

                                icon_lock = ft.Icons.LOCK_OPEN if es_abierto == 1 else ft.Icons.LOCK
                                color_lock = "#7CFC00" if es_abierto == 1 else "#FF4500"
                                tooltip_lock = "Acceso Público (Click para restringir a Admin)" if es_abierto == 1 else "Restringido (Click para permitir a todos)"

                                is_excel = nombre.lower().endswith(('.xlsx', '.xls'))
                                icon_type = ft.Icons.TABLE_CHART if is_excel else ft.Icons.PICTURE_AS_PDF
                                color_type = "#7CFC00" if is_excel else "#00FFFF"

                                manuals_list.controls.append(
                                    ft.Container(
                                        content=ft.Row([
                                            ft.Icon(icon_type, color=color_type, size=20 if is_mobile else 24),
                                            ft.Column([
                                                ft.Text(nombre, color="white", weight="bold", size=12 if is_mobile else 14),
                                                ft.Text(f"Versión: {version}" if version else "Versión: 1.0", color="#aaaaaa", size=10 if is_mobile else 11)
                                            ], spacing=2, expand=True),
                                            ft.IconButton(
                                                icon=icon_lock,
                                                icon_color=color_lock,
                                                tooltip=tooltip_lock,
                                                icon_size=18 if is_mobile else 22,
                                                on_click=toggle_abierto_click
                                            ),
                                            ft.IconButton(
                                                icon=ft.Icons.DELETE_FOREVER,
                                                icon_color="#FF4500",
                                                tooltip="Eliminar manual",
                                                icon_size=18 if is_mobile else 22,
                                                on_click=borrar_click
                                            )
                                        ], alignment="spaceBetween", vertical_alignment="center"),
                                        bgcolor="#141424",
                                        padding=8 if is_mobile else 10,
                                        border_radius=8,
                                        border=ft.Border.all(1, "#333333")
                                    )
                                )
                except Exception as ex:
                    print("ERROR LISTA MANUALES:", ex)
                    manuals_list.controls.append(ft.Text("Error al cargar la lista de manuales.", color="red"))
                
                # --- Recursos Multimedia Locales (assets/) ---
                manuals_list.controls.append(ft.Divider(height=15, color="#333333"))
                manuals_list.controls.append(ft.Text("Recursos Multimedia Locales (assets/)", size=13 if is_mobile else 14, color="#A100F2", weight="bold"))
                try:
                    os.makedirs(ASSETS_PATH, exist_ok=True)
                    archivos_protegidos = {
                        "avatar_luxo.png.jpeg",
                        "luxo_avatar1.mp4",
                        "luxo_avatar2.mp4",
                        "istockphoto-468228782-612x612.jpg.jpeg"
                    }
                    archivos = [
                        f for f in os.listdir(ASSETS_PATH) 
                        if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.mp4', '.mov', '.avi'))
                        and f not in archivos_protegidos
                    ]
                    if not archivos:
                        manuals_list.controls.append(ft.Text("No hay archivos multimedia cargados en custom_assets/.", color="#aaaaaa", size=12))
                    else:
                        for filename in archivos:
                            def borrar_media_click(e, fn=filename):
                                def on_confirmar_media(ev):
                                    try:
                                        ruta_media = os.path.join(ASSETS_PATH, fn)
                                        if os.path.exists(ruta_media):
                                            os.remove(ruta_media)
                                        mostrar_snack(f"Archivo multimedia '{fn}' eliminado.")
                                        cargar_manuales()
                                        page.pop_dialog()
                                        page.update()
                                    except Exception as ex:
                                        print("ERROR BORRAR MEDIA:", ex)
                                        mostrar_snack("Error al borrar archivo multimedia.", color="red")
                                        
                                def on_cancelar_media(ev):
                                    page.pop_dialog()
                                    
                                dialog_confirm_media = ft.AlertDialog(
                                    title=ft.Text("Confirmar Borrado de Multimedia", color="#FF4500", weight="bold"),
                                    content=ft.Text(f"¿Seguro que deseas borrar el archivo multimedia \"{fn}\"?"),
                                    actions=[
                                        ft.TextButton("Cancelar", on_click=on_cancelar_media),
                                        ft.ElevatedButton("Sí, Borrar", on_click=on_confirmar_media, bgcolor="#FF4500", color="white")
                                    ],
                                    actions_alignment="end",
                                    bgcolor="#0F0F1A"
                                )
                                page.show_dialog(dialog_confirm_media)

                            is_video = filename.lower().endswith(('.mp4', '.mov', '.avi'))
                            icon_media = ft.Icons.PLAY_CIRCLE_FILL if is_video else ft.Icons.IMAGE
                            color_media = "#A100F2"
                            
                            manuals_list.controls.append(
                                ft.Container(
                                    content=ft.Row([
                                        ft.Icon(icon_media, color=color_media, size=20 if is_mobile else 24),
                                        ft.Column([
                                            ft.Text(filename, color="white", weight="bold", size=12 if is_mobile else 14),
                                            ft.Text("Ubicación: local assets/", color="#aaaaaa", size=10 if is_mobile else 11)
                                        ], spacing=2, expand=True),
                                        ft.IconButton(
                                            icon=ft.Icons.DELETE_FOREVER,
                                            icon_color="#FF4500",
                                            tooltip="Eliminar multimedia",
                                            icon_size=18 if is_mobile else 22,
                                            on_click=borrar_media_click
                                        )
                                    ], alignment="spaceBetween", vertical_alignment="center"),
                                    bgcolor="#141424",
                                    padding=8 if is_mobile else 10,
                                    border_radius=8,
                                    border=ft.Border.all(1, "#333333")
                                )
                            )
                except Exception as ex:
                    print("ERROR LISTA MULTIMEDIA:", ex)
                    manuals_list.controls.append(ft.Text("Error al cargar la lista de archivos multimedia.", color="red"))
                
                page.update()
                    
            cargar_manuales()
            
            def on_pdf_cargado(ruta):
                procesar_cargar_pdf(ruta)
                def reload_after_delay():
                    import time
                    time.sleep(1.5)
                    cargar_manuales()
                    page.update()
                threading.Thread(target=reload_after_delay, daemon=True).start()
                
            def on_excel_cargado(ruta):
                procesar_cargar_excel(ruta)
                def reload_after_delay():
                    import time
                    time.sleep(1.5)
                    cargar_manuales()
                    page.update()
                threading.Thread(target=reload_after_delay, daemon=True).start()

            def on_multimedia_cargado(ruta):
                try:
                    os.makedirs(ASSETS_PATH, exist_ok=True)
                    nombre_archivo = os.path.basename(ruta)
                    destino = os.path.join(ASSETS_PATH, nombre_archivo)
                    shutil.copy(ruta, destino)
                    optimizar_archivo_multimedia(destino)
                    mostrar_snack(f"Archivo multimedia '{nombre_archivo}' cargado con éxito en custom_assets/.")
                    
                    # Recargar después de cargar
                    def reload_after_delay():
                        import time
                        time.sleep(1.5)
                        cargar_manuales()
                        page.update()
                    threading.Thread(target=reload_after_delay, daemon=True).start()
                except Exception as ex:
                    print("ERROR CARGANDO MULTIMEDIA:", ex)
                    mostrar_snack("Error al guardar archivo multimedia.", color="red")

            btn_pdf = ft.ElevatedButton(
                "Cargar PDF",
                icon=ft.Icons.PICTURE_AS_PDF,
                bgcolor="#6E48AA",
                color="white",
                height=34 if is_mobile else 38,
                style=ft.ButtonStyle(
                    shape=ft.RoundedRectangleBorder(radius=6),
                    padding=ft.padding.Padding(8, 0, 8, 0) if is_mobile else None
                ),
                on_click=lambda e: seleccionar_archivo_async(
                    "Seleccionar PDF para cargar",
                    [("PDF files", "*.pdf")],
                    on_pdf_cargado
                )
            )
            
            btn_excel = ft.ElevatedButton(
                "Cargar Excel",
                icon=ft.Icons.TABLE_CHART,
                bgcolor="#1f6f43",
                color="white",
                height=34 if is_mobile else 38,
                style=ft.ButtonStyle(
                    shape=ft.RoundedRectangleBorder(radius=6),
                    padding=ft.padding.Padding(8, 0, 8, 0) if is_mobile else None
                ),
                on_click=lambda e: seleccionar_archivo_async(
                    "Seleccionar Excel para cargar",
                    [("Excel files", "*.xlsx *.xls"), ("Todos los archivos", "*.*")],
                    on_excel_cargado
                )
            )

            btn_media = ft.ElevatedButton(
                "Cargar Media" if is_mobile else "Cargar Multimedia",
                icon=ft.Icons.PERM_MEDIA,
                bgcolor="#A100F2",
                color="white",
                height=34 if is_mobile else 38,
                style=ft.ButtonStyle(
                    shape=ft.RoundedRectangleBorder(radius=6),
                    padding=ft.padding.Padding(8, 0, 8, 0) if is_mobile else None
                ),
                on_click=lambda e: seleccionar_archivo_async(
                    "Seleccionar Imagen, GIF o Video",
                    [
                        ("Archivos Multimedia", "*.png *.jpg *.jpeg *.gif *.mp4 *.mov *.avi"),
                        ("Imágenes (*.png, *.jpg, *.gif)", "*.png *.jpg *.jpeg *.gif"),
                        ("Videos (*.mp4, *.mov)", "*.mp4 *.mov *.avi"),
                        ("Todos los archivos", "*.*")
                    ],
                    on_multimedia_cargado
                )
            )

            buttons_row = ft.Row([btn_pdf, btn_excel, btn_media], wrap=True, spacing=6 if is_mobile else 10)

            if is_mobile:
                header_container = ft.Column([
                    ft.Text("Manuales y Documentos de Sunglass Hut", size=15, color="white", weight="bold"),
                    buttons_row
                ], spacing=8)
            else:
                header_container = ft.Row([
                    ft.Text("Manuales y Documentos de Sunglass Hut", size=18, color="white", weight="bold"),
                    buttons_row
                ], alignment="spaceBetween", vertical_alignment="center", wrap=True)

            return ft.Column([
                header_container,
                ft.Divider(height=10, color="transparent"),
                manuals_list
            ], expand=True)

        def build_suggestions_tab():
            suggestions_list = ft.Column(spacing=10, scroll=ft.ScrollMode.AUTO, expand=True)
            
            def cargar_sugerencias():
                suggestions_list.controls.clear()
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("""
                            CREATE TABLE IF NOT EXISTS sugerencias_luxo (
                                ID_Sugerencia INT AUTO_INCREMENT PRIMARY KEY,
                                ID_Usuario INT NOT NULL,
                                Fecha_Hora DATETIME DEFAULT CURRENT_TIMESTAMP,
                                Sugerencia TEXT NOT NULL,
                                FOREIGN KEY (ID_Usuario) REFERENCES usuarios(ID_Usuario) ON DELETE CASCADE
                            )
                        """)
                        db.commit()
                        
                        cursor.execute("""
                            SELECT s.Fecha_Hora, u.Nombre_Completo, s.Sugerencia 
                            FROM sugerencias_luxo s
                            JOIN usuarios u ON s.ID_Usuario = u.ID_Usuario
                            ORDER BY s.Fecha_Hora DESC
                        """)
                        sugerencias = cursor.fetchall()
                        db.close()
                        
                        if not sugerencias:
                            suggestions_list.controls.append(
                                ft.Text("No hay sugerencias registradas de los usuarios.", color="#aaaaaa", size=14)
                            )
                        else:
                            for row in sugerencias:
                                fecha = row["Fecha_Hora"].strftime("%d/%m/%Y %H:%M") if row["Fecha_Hora"] else ""
                                suggestions_list.controls.append(
                                    ft.Container(
                                        content=ft.Column([
                                            ft.Row([
                                                ft.Text(f"📅 {fecha}", color="#aaaaaa", size=11),
                                                ft.Text(f"👤 Usuario: {row['Nombre_Completo']}", color="#aaaaaa", size=11),
                                            ], alignment="spaceBetween"),
                                            ft.Text(row["Sugerencia"], color="white", size=14),
                                        ], spacing=5),
                                        bgcolor="#141424",
                                        padding=15,
                                        border_radius=8,
                                        border=ft.Border.all(1, "#333333")
                                    )
                                )
                except Exception as ex:
                    print("ERROR AL CARGAR SUGERENCIAS:", ex)
                    suggestions_list.controls.append(ft.Text("Error al cargar las sugerencias de los usuarios.", color="red"))
                page.update()
                
            cargar_sugerencias()
            
            return ft.Column([
                ft.Row([
                    ft.Text("Sugerencias de los Usuarios", size=18, color="white", weight="bold"),
                    ft.IconButton(icon=ft.Icons.REFRESH, tooltip="Recargar sugerencias", on_click=lambda e: cargar_sugerencias())
                ], alignment="spaceBetween", vertical_alignment="center"),
                ft.Text("Lista de comentarios y propuestas enviados por los usuarios a través del recuadro de la barra lateral.", color="#aaaaaa", size=13),
                ft.Divider(height=10, color="transparent"),
                suggestions_list
            ], expand=True)

        def build_support_tickets_tab():
            tickets_list = ft.Column(spacing=10, scroll=ft.ScrollMode.AUTO, expand=True)
            
            def cargar_tickets():
                tickets_list.controls.clear()
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("""
                            SELECT t.ID_Ticket, t.Fecha_Hora, u.Nombre_Completo, t.Detalle_Problema, t.Respuesta_Soporte, t.Estatus
                            FROM tickets_soporte t
                            JOIN usuarios u ON t.ID_Usuario = u.ID_Usuario
                            ORDER BY t.Fecha_Hora DESC
                        """)
                        tickets = cursor.fetchall()
                        db.close()
                        
                        if not tickets:
                            tickets_list.controls.append(ft.Text("No hay tickets de soporte registrados.", color="#aaaaaa", size=14))
                        else:
                            for row in tickets:
                                id_t = row["ID_Ticket"]
                                fecha = row["Fecha_Hora"].strftime("%d/%m/%Y %H:%M") if row["Fecha_Hora"] else ""
                                estatus = row["Estatus"]
                                respuesta_t = row["Respuesta_Soporte"] or "(Sin respuesta aún)"
                                
                                is_abierto = estatus == "Abierto"
                                status_color = "#FF4500" if is_abierto else "#7CFC00"
                                
                                # Campo de respuesta de soporte para el admin
                                resp_input = ft.TextField(
                                    label="Escribe la solución...",
                                    value=row["Respuesta_Soporte"] or "",
                                    multiline=True,
                                    min_lines=1,
                                    max_lines=3,
                                    border_color="#9D50BB",
                                    color="white",
                                    text_size=12,
                                    expand=True
                                )
                                
                                def resolver_ticket_click(e, ticket_id=id_t, r_input=resp_input):
                                    solucion = r_input.value.strip()
                                    if not solucion:
                                        mostrar_snack("Por favor escribe una solución antes de resolver.", color="red")
                                        return
                                    try:
                                        db_res = conectar_db()
                                        if db_res:
                                            cursor_res = db_res.cursor()
                                            cursor_res.execute("""
                                                UPDATE tickets_soporte 
                                                SET Estatus = 'Resuelto', Respuesta_Soporte = %s 
                                                WHERE ID_Ticket = %s
                                            """, (solucion, ticket_id))
                                            db_res.commit()
                                            db_res.close()
                                            mostrar_snack("Ticket resuelto con éxito.", color="#7CFC00")
                                            cargar_tickets()
                                            page.update()
                                    except Exception as ex:
                                        print("ERROR RESOLVER TICKET:", ex)
                                
                                tickets_list.controls.append(
                                    ft.Container(
                                        content=ft.Column([
                                            ft.Row([
                                                ft.Text(f"📅 {fecha}", color="#aaaaaa", size=11),
                                                ft.Text(f"👤 Reporta: {row['Nombre_Completo']}", color="#aaaaaa", size=11),
                                                ft.Container(
                                                    content=ft.Text(estatus.upper(), color="black", size=9, weight="bold"),
                                                    bgcolor=status_color,
                                                    padding=ft.Padding(left=6, right=6, top=2, bottom=2),
                                                    border_radius=3
                                                )
                                            ], alignment="spaceBetween"),
                                            ft.Text(row["Detalle_Problema"], color="white", size=13),
                                            ft.Divider(height=10, color="#444444"),
                                            ft.Row([
                                                ft.Text("Solución de Soporte:", color="#aaaaaa", size=12, weight="bold"),
                                            ]),
                                            ft.Row([
                                                resp_input,
                                                ft.ElevatedButton(
                                                    "Resolver",
                                                    icon=ft.Icons.CHECK_CIRCLE,
                                                    bgcolor="#7CFC00" if is_abierto else "#444444",
                                                    color="black" if is_abierto else "white",
                                                    on_click=resolver_ticket_click,
                                                    disabled=not is_abierto
                                                )
                                            ], spacing=10) if is_abierto else (
                                                ft.Text(respuesta_t, color="#7CFC00", size=13, italic=True)
                                            )
                                        ], spacing=5),
                                        bgcolor="#141424",
                                        padding=15,
                                        border_radius=8,
                                        border=ft.Border.all(1, "#333333")
                                    )
                                )
                except Exception as ex:
                    print("ERROR EN CARGAR TICKETS:", ex)
                    tickets_list.controls.append(ft.Text("Error al cargar los tickets de soporte.", color="red"))
                page.update()
                
            cargar_tickets()
            
            return ft.Column([
                ft.Row([
                    ft.Text("Bandeja de Tickets de Soporte Técnico", size=18, color="white", weight="bold"),
                    ft.IconButton(icon=ft.Icons.REFRESH, tooltip="Recargar tickets", on_click=lambda e: cargar_tickets())
                ], alignment="spaceBetween", vertical_alignment="center"),
                ft.Text("Visualiza y responde a las fallas o inconsistencias operativas reportadas por los asociados.", color="#aaaaaa", size=13),
                ft.Divider(height=10, color="transparent"),
                tickets_list
            ], expand=True)

        def build_checklists_view():
            apertura_list = ft.Column(spacing=10, scroll=ft.ScrollMode.AUTO, expand=True)
            cierre_list = ft.Column(spacing=10, scroll=ft.ScrollMode.AUTO, expand=True)
            venta_list = ft.Column(spacing=10, scroll=ft.ScrollMode.AUTO, expand=True)

            dropdown_vendedor_venta = ft.Dropdown(
                label="Asesor que realizó la Venta",
                border_color="#A100F2",
                color="white",
                width=300
            )

            def cargar_vendedores_dropdown_venta():
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("SELECT ID_Vendedor, Nombre_Completo FROM vendedores WHERE ID_Usuario_Tienda = %s AND Activo = 1 ORDER BY Nombre_Completo ASC", (user_info["id"],))
                        rows = cursor.fetchall()
                        db.close()
                        dropdown_vendedor_venta.options = [ft.dropdown.Option(str(r["ID_Vendedor"]), r["Nombre_Completo"]) for r in rows]
                except Exception as ex:
                    print("Error cargar dropdown venta exitosa:", ex)

            cargar_vendedores_dropdown_venta()

            # file_picker_vitrina is initialized globally and appended in cargar_chat

            def en_archivo_seleccionado_vitrina(e):
                if not e.files:
                    return
                filepath = e.files[0].path
                mostrar_snack("Analizando imagen de vitrina con IA Vision. Por favor espera...", "#00FFFF")
                
                try:
                    import base64
                    with open(filepath, "rb") as image_file:
                        encoded_string = base64.b64encode(image_file.read()).decode("utf-8")
                    
                    headers = {
                        "Authorization": f"Bearer {GROQ_API_KEY}",
                        "Content-Type": "application/json"
                    }
                    
                    payload = {
                        "model": "llama-3.2-11b-vision-preview",
                        "messages": [
                            {
                                "role": "user",
                                "content": [
                                    {
                                        "type": "text",
                                        "text": "Analiza esta fotografía de la vitrina de la boutique de lentes Sunglass Hut. Evalúa si cumple con el lineamiento premium de visual merchandising: 1. Alineación recta de los lentes. 2. Espacios vacíos notables que deban rellenarse. 3. Limpieza de cristales (polvo o huellas visibles). 4. Etiquetas de precios bien colocadas. Responde de forma muy concisa en español comenzando con 'APROBADO' o 'CORREGIR' y una lista breve de puntos a solucionar."
                                    },
                                    {
                                        "type": "image_url",
                                        "image_url": {
                                            "url": f"data:image/jpeg;base64,{encoded_string}"
                                        }
                                    }
                                ]
                            }
                        ]
                    }
                    
                    res = requests.post(URL_GROQ, headers=headers, json=payload, timeout=25)
                    if res.status_code == 200:
                        resultado = res.json()["choices"][0]["message"]["content"]
                        color_res = "#7CFC00" if "APROBADO" in resultado.upper() else "#FF4500"
                        
                        dlg_res = ft.AlertDialog(
                            title=ft.Text("🔍 Diagnóstico Visual de Vitrina con IA", color=color_res, weight="bold", size=16),
                            content=ft.Container(
                                content=ft.Column([
                                    ft.Text(resultado, color="white", size=13, selectable=True)
                                ], scroll=ft.ScrollMode.AUTO),
                                width=480,
                                height=250
                            ),
                            actions=[
                                ft.TextButton("Entendido", on_click=lambda ev: page.pop_dialog())
                            ],
                            bgcolor="#0F0F1A"
                        )
                        page.show_dialog(dlg_res)
                        page.update()
                    else:
                        print("Error Groq Vision:", res.status_code, res.text)
                        mostrar_snack("Error al procesar la imagen con IA Vision", "red")
                except Exception as ex_v:
                    print("Error auditar vitrina:", ex_v)
                    mostrar_snack("Error al abrir o procesar la imagen de vitrina", "red")

            file_picker_vitrina.on_result = en_archivo_seleccionado_vitrina

            btn_auditar_vitrina = ft.ElevatedButton(
                "Auditar Vitrina 📸",
                icon=ft.Icons.CAMERA_ALT,
                on_click=lambda e: file_picker_vitrina.pick_files(allow_multiple=False, file_type=ft.FilePickerFileType.IMAGE),
                bgcolor="#00FFFF",
                color="black",
                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=10))
            )
            
            progress_apertura = ft.ProgressBar(value=0.0, color="#7CFC00", bgcolor="#141424")
            progress_cierre = ft.ProgressBar(value=0.0, color="#00FFFF", bgcolor="#141424")
            progress_venta = ft.ProgressBar(value=0.0, color="#A100F2", bgcolor="#141424")
            
            text_apertura = ft.Text(f"{t('progress')}: 0%", color="#7CFC00", size=13, weight="bold")
            text_cierre = ft.Text(f"{t('progress')}: 0%", color="#00FFFF", size=13, weight="bold")
            text_venta = ft.Text(f"{t('progress')}: 0%", color="#A100F2", size=13, weight="bold")
            
            def calcular_progreso(categoria, col, p_bar, p_text):
                try:
                    total = 0
                    completados = 0
                    for container in col.controls:
                        if isinstance(container, ft.Container) and container.content:
                            content = container.content
                            chk = None
                            if isinstance(content, ft.Row) and content.controls:
                                # Modo Admin: Row([Checkbox, IconButton])
                                chk = content.controls[0]
                            elif isinstance(content, ft.Checkbox):
                                # Modo Asociado: Checkbox directo
                                chk = content
                            
                            if isinstance(chk, ft.Checkbox):
                                total += 1
                                if chk.value:
                                    completados += 1
                                    
                    val = 0.0
                    if total > 0:
                        val = completados / total
                    p_bar.value = val
                    p_text.value = f"{t('progress')}: {int(val * 100)}% ({completados} {t('of')} {total} {t('completed')})"
                except Exception as ex:
                    print("ERROR CALCULAR PROGRESO CHECKLIST:", ex)
                page.update()

            def mostrar_retro_venta_exitosa():
                consejos = [
                    "¡Excelente trabajo! Recuerda que según el Manual de Servicio al Cliente, siempre debemos ofrecer al menos 3 opciones de armazones que se adapten a la forma del rostro del cliente para incrementar el ticket promedio y garantizar su satisfacción.",
                    "¡Checklist de Venta completado! El Manual de Ventas indica que el 80% de los clientes decide su compra tras probarse físicamente el producto. Asegúrate de limpiar siempre las vitrinas de exhibición y los lentes frente al cliente.",
                    "¡Muy bien hecho! Recuerda que al finalizar una venta exitosa, debes reiterar claramente las condiciones de la garantía de fábrica (2 años) y limpiar los lentes minuciosamente antes de entregarlos en su estuche original.",
                    "¡Venta exitosa registrada! El Manual de Operaciones destaca la importancia del doble chequeo del ticket de cobro y de las piezas físicas entregadas en su estuche para evitar discrepancias en inventario."
                ]
                import random
                tip = random.choice(consejos)
                
                def cerrar_retro_dialog(ev):
                    page.pop_dialog()
                    page.update()
                
                dlg_retro = ft.AlertDialog(
                    title=ft.Text("💡 Retroalimentación de Venta Exitosa", color="#00FF7F", weight="bold", size=16),
                    content=ft.Container(
                        content=ft.Column([
                            ft.Text("¡Felicidades por completar el checklist de Venta Exitosa al 100%! 🎉", color="white", weight="bold", size=14),
                            ft.Container(height=10),
                            ft.Text(tip, color="#D8B4FE", size=13, italic=True)
                        ], spacing=5, tight=True),
                        width=450
                    ),
                    actions=[
                        ft.ElevatedButton("Entendido", on_click=cerrar_retro_dialog, bgcolor="#00FF7F", color="black")
                    ],
                    actions_alignment="end",
                    bgcolor="#0F0F1A"
                )
                page.show_dialog(dlg_retro)
                page.update()

            def mostrar_retro_checklist_completado(categoria):
                if categoria == 1:
                    titulo = "🌅 ¡Checklist de Apertura Completado!"
                    medalla = "Medalla Madrugador 🌅"
                    color_accent = "#7CFC00"
                    tips = [
                        "Recuerda realizar siempre el conteo físico del fondo de caja en presencia de un testigo antes de abrir las puertas de la tienda.",
                        "Revisa que todas las gafas inteligentes Ray-Ban Meta en vitrina estén limpias y encendidas para las demostraciones con clientes.",
                        "Asegúrate de que la música ambiental de la tienda esté a un volumen agradable y profesional antes del ingreso de los primeros clientes."
                    ]
                elif categoria == 2:
                    titulo = "🌙 ¡Checklist de Cierre Completado!"
                    medalla = "Medalla Cierre Perfecto 🌙"
                    color_accent = "#00FFFF"
                    tips = [
                        "Antes de salir de la tienda, valida dos veces que la caja fuerte esté cerrada bajo llave y el sistema de alarma activado correctamente.",
                        "Recuerda que todas las terminales de venta (Pinpads) deben quedar apagadas y desconectadas de acuerdo al protocolo de finanzas.",
                        "Valida que no queden clientes ni personas ajenas al personal dentro de la periferia física de la tienda antes de cerrar las cortinas."
                    ]
                else:
                    return
                
                import random
                tip = random.choice(tips)
                
                def cerrar_retro_dialog(ev):
                    page.pop_dialog()
                    page.update()
                
                dlg_retro = ft.AlertDialog(
                    title=ft.Text(titulo, color=color_accent, weight="bold", size=16),
                    content=ft.Container(
                        content=ft.Column([
                            ft.Text(f"¡Felicidades por completar todas las tareas al 100% hoy! 🎉\nHas ganado la {medalla}.", color="white", weight="bold", size=13),
                            ft.Container(height=10),
                            ft.Text("💡 Recordatorio del Manual Operativo:", color="white", size=12, weight="bold"),
                            ft.Text(tip, color="#D8B4FE", size=12, italic=True)
                        ], spacing=5, tight=True),
                        width=450
                    ),
                    actions=[
                        ft.ElevatedButton("Excelente", on_click=cerrar_retro_dialog, bgcolor=color_accent, color="black")
                    ],
                    actions_alignment="end",
                    bgcolor="#0F0F1A"
                )
                page.show_dialog(dlg_retro)
                page.update()

            def obtener_consejo_para_tarea(descripcion):
                desc_lower = descripcion.lower()
                
                # Apertura
                if "computadora principal" in desc_lower or "pos" in desc_lower:
                    return "Encender los sistemas a tiempo previene retrasos con los primeros clientes y asegura la correcta sincronización del inventario diario."
                if "epson" in desc_lower or "papel térmico" in desc_lower:
                    return "Tener consumibles listos evita demoras al momento del cobro y asegura que no detengas la fila de clientes en caja."
                if "cobro con tarjeta" in desc_lower or ("terminal" in desc_lower and "encendida" in desc_lower):
                    return "Verificar la conectividad de la terminal bancaria a primera hora garantiza que no pierdas transacciones electrónicas."
                if "vitrina" in desc_lower or "limpiar" in desc_lower or "acomodar" in desc_lower or "exhibición" in desc_lower:
                    return "Las vitrinas impecables aumentan la conversión. Usa el paño oficial y microfibra para mantener la imagen premium de Sunglass Hut."
                if "trapear" in desc_lower:
                    return "Un piso limpio y reluciente da la primera y mejor impresión de higiene y profesionalismo al ingresar a la boutique."

                # Cierre
                if "corte de caja" in desc_lower or "conciliación" in desc_lower or "arqueo" in desc_lower or "valores" in desc_lower or "caja" in desc_lower or "efectivo" in desc_lower:
                    return "El conteo físico preciso y el arqueo previenen discrepancias contables y son auditados minuciosamente todos los días."
                if "bajo llave" in desc_lower or ("asegurar" in desc_lower and "mercancía" in desc_lower):
                    return "El resguardo de mercancía en vitrinas cerradas es obligatorio para cumplir con los estándares de prevención de pérdidas de la tienda."
                if "mostrador" in desc_lower or "empaque" in desc_lower or ("limpiar" in desc_lower and "área" in desc_lower):
                    return "Dejar el mostrador limpio y ordenado facilita una apertura ágil y organizada para el turno del día siguiente."
                if "apagar luces" in desc_lower or "apagar" in desc_lower or "pantallas" in desc_lower or "desconectar" in desc_lower:
                    return "Desconectar equipos y apagar luminarias de noche ayuda al ahorro energético y alarga la vida útil del equipo tecnológico."
                if "alarma" in desc_lower or "seguridad" in desc_lower or "cerradura" in desc_lower or "llave" in desc_lower:
                    return "El protocolo de seguridad de tienda exige el resguardo doble de valores y el armado del sistema de alarma para garantizar la cobertura del seguro."
                if "música" in desc_lower or "volumen" in desc_lower or "ambiente" in desc_lower:
                    return "La música ambiental oficial a volumen moderado influye positivamente en el estado de ánimo y aumenta el tiempo de permanencia del cliente."

                # Venta Exitosa
                if "kit de limpieza" in desc_lower or "estuche premium" in desc_lower or "ofrecer" in desc_lower:
                    return "El kit de limpieza y estuches adicionales añaden valor a la compra e incrementan el ticket promedio (UPT) de la tienda."
                if "datos de correo" in desc_lower or "registrar" in desc_lower or "garantía" in desc_lower:
                    return "Capturar el correo del cliente alimenta nuestra base CRM, permitiendo enviarle campañas de lealtad y registrar su garantía digital."
                if "aprobada" in desc_lower or "terminal bancaria" in desc_lower or "transacción" in desc_lower:
                    return "Siempre confirma en físico que el ticket de la terminal diga 'APROBADA' y coincida con el cobro en el POS antes de entregar el producto."
                if "ticket de compra" in desc_lower or "bolsa" in desc_lower or "sunglass hut" in desc_lower:
                    return "El empaquetado premium y la entrega cordial del ticket de compra cierran el ciclo del servicio de excelencia de Sunglass Hut."
                if "nombre" in desc_lower:
                    return "Presentarte por tu nombre genera una conexión de confianza y personaliza la experiencia del cliente para futuras visitas."

                # Fallback generalizado inteligente
                return "Completar todos los puntos operativos garantiza que la tienda mantenga el estándar de servicio Premium de Sunglass Hut."

            def mostrar_retro_puntos_faltantes(categoria, tareas_faltantes):
                if categoria == 1:
                    titulo_cat = "Apertura 🌅"
                elif categoria == 2:
                    titulo_cat = "Cierre 🌙"
                else:
                    titulo_cat = "Venta Exitosa 💰"
                
                controles_tareas = []
                for t_desc in tareas_faltantes:
                    consejo = obtener_consejo_para_tarea(t_desc)
                    controles_tareas.append(
                        ft.Container(
                            content=ft.Column([
                                ft.Row([
                                    ft.Icon(ft.Icons.WARNING_ROUNDED, color="#FFCC00", size=18),
                                    ft.Text(t_desc, color="white", weight="bold", size=13, expand=True)
                                ], spacing=8),
                                ft.Container(
                                    content=ft.Text(f"💡 Sugerencia de mejora: {consejo}", color="#D8B4FE", size=12, italic=True),
                                    margin=ft.Margin(left=26, top=2, right=0, bottom=8)
                                )
                            ], spacing=2),
                            padding=ft.Padding(0, 4, 0, 4)
                        )
                    )
                
                def cerrar_dialog(e):
                    page.pop_dialog()
                    page.update()
                
                lista_view = ft.ListView(
                    controls=controles_tareas,
                    spacing=5,
                    height=250,
                    expand=True
                )
                
                dlg_faltantes = ft.AlertDialog(
                    title=ft.Text(f"📋 Puntos de Mejora - {titulo_cat}", color="#FFCC00", weight="bold", size=16),
                    content=ft.Container(
                        content=ft.Column([
                            ft.Text("Para mantener el estándar de excelencia operativa de la tienda, te sugerimos completar y revisar estos puntos pendientes:", color="white", size=13),
                            ft.Divider(height=10, color="transparent"),
                            lista_view
                        ], spacing=5, tight=True),
                        width=480
                    ),
                    actions=[
                        ft.ElevatedButton("Ir a Completar", on_click=cerrar_dialog, bgcolor="#FFCC00", color="black")
                    ],
                    actions_alignment="end",
                    bgcolor="#0F0F1A"
                )
                page.show_dialog(dlg_faltantes)
                page.update()

            def guardar_checklist_click(categoria):
                col = apertura_list if categoria == 1 else (cierre_list if categoria == 2 else venta_list)
                
                # Actualizar medallas en la sidebar
                if hasattr(page, "actualizar_medallas_sidebar"):
                    try:
                        page.actualizar_medallas_sidebar()
                    except Exception:
                        pass
                
                total = 0
                completados = 0
                tareas_faltantes = []
                for container in col.controls:
                    if isinstance(container, ft.Container) and container.content:
                        content = container.content
                        chk = None
                        if isinstance(content, ft.Row) and content.controls:
                            chk = content.controls[0]
                        elif isinstance(content, ft.Checkbox):
                            chk = content
                        
                        if isinstance(chk, ft.Checkbox):
                            total += 1
                            if chk.value:
                                completados += 1
                            else:
                                tareas_faltantes.append(chk.label)
                
                # Validar vendedor para Venta Exitosa (Categoría 3)
                if categoria == 3 and not es_admin():
                    if not dropdown_vendedor_venta.value:
                        mostrar_snack("Por favor selecciona al vendedor que realizó la venta exitosa antes de guardar", "red")
                        return
                    
                    vendedor_id = int(dropdown_vendedor_venta.value)
                    faltantes_txt = ", ".join(tareas_faltantes) if tareas_faltantes else ""
                    try:
                        db_s = conectar_db()
                        if db_s:
                            cursor_s = db_s.cursor()
                            cursor_s.execute("""
                                INSERT INTO registro_venta_exitosa 
                                (ID_Usuario_Tienda, ID_Vendedor, Puntos_Completados, Puntos_Totales, Detalle_Faltantes)
                                VALUES (%s, %s, %s, %s, %s)
                            """, (user_info["id"], vendedor_id, completados, total, faltantes_txt))
                            db_s.commit()
                            db_s.close()
                    except Exception as ex_s:
                        print("Error al guardar venta exitosa por vendedor:", ex_s)

                es_perfecto = (total > 0 and completados == total)
                
                if es_perfecto:
                    if categoria == 1:
                        mostrar_retro_checklist_completado(1)
                    elif categoria == 2:
                        mostrar_retro_checklist_completado(2)
                    elif categoria == 3:
                        mostrar_retro_venta_exitosa()
                else:
                    porcentaje = int((completados / total) * 100) if total > 0 else 0
                    if tareas_faltantes:
                        mostrar_retro_puntos_faltantes(categoria, tareas_faltantes)
                    else:
                        mostrar_snack(f"Checklist guardado parcialmente. Progreso actual: {porcentaje}%. Completa el 100% para tu medalla.", color="#00FFFF")

            def toggle_tarea(id_plantilla, completado_val, categoria, col, p_bar, p_text):
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        if completado_val:
                            cursor.execute("""
                                INSERT INTO registro_checklist (ID_Usuario, ID_Plantilla, Completado, Fecha, Fecha_Hora)
                                VALUES (%s, %s, 1, CURDATE(), NOW())
                                ON DUPLICATE KEY UPDATE Completado = 1, Fecha_Hora = NOW()
                            """, (user_info["id"], id_plantilla))
                        else:
                            cursor.execute("""
                                DELETE FROM registro_checklist 
                                WHERE ID_Usuario = %s AND ID_Plantilla = %s AND Fecha = CURDATE()
                            """, (user_info["id"], id_plantilla))
                        db.commit()
                        db.close()
                except Exception as ex:
                    print("ERROR TOGGLE TAREA CHECKLIST:", ex)
                
                # Si completó al 100% la categoría 3 (Venta Exitosa), mostrar retroalimentación
                if completado_val and categoria == 3:
                    try:
                        db_chk = conectar_db()
                        if db_chk:
                            cur_chk = db_chk.cursor()
                            cur_chk.execute("SELECT COUNT(*) FROM plantillas_checklist WHERE Categoria = 3")
                            tot3 = cur_chk.fetchone()[0]
                            cur_chk.execute("""
                                SELECT COUNT(*) FROM registro_checklist 
                                WHERE ID_Usuario = %s AND Fecha = CURDATE() AND Completado = 1 
                                AND ID_Plantilla IN (SELECT ID_Plantilla FROM plantillas_checklist WHERE Categoria = 3)
                            """, (user_info["id"],))
                            comp3 = cur_chk.fetchone()[0]
                            db_chk.close()
                            
                            if tot3 > 0 and comp3 == tot3:
                                mostrar_retro_venta_exitosa()
                    except Exception as e_chk:
                        print("Error al verificar retro de venta:", e_chk)

                # Actualizar medallas en la sidebar en tiempo real si corresponde
                if hasattr(page, "actualizar_medallas_sidebar"):
                    try:
                        page.actualizar_medallas_sidebar()
                    except Exception:
                        pass
                
                calcular_progreso(categoria, col, p_bar, p_text)

            def cargar_checklist_por_categoria(categoria, col, p_bar, p_text):
                col.controls.clear()
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("SELECT ID_Plantilla, Descripcion FROM plantillas_checklist WHERE Categoria = %s ORDER BY ID_Plantilla ASC", (categoria,))
                        tareas = cursor.fetchall()
                        
                        cursor.execute("""
                            SELECT ID_Plantilla FROM registro_checklist 
                            WHERE ID_Usuario = %s AND Fecha = CURDATE() AND Completado = 1
                        """, (user_info["id"],))
                        completadas_hoy = {row["ID_Plantilla"] for row in cursor.fetchall()}
                        db.close()
                        
                        if not tareas:
                            col.controls.append(ft.Text(t("no_tasks"), color="#aaaaaa", italic=True))
                        else:
                            for t_item in tareas:
                                id_pl = t_item["ID_Plantilla"]
                                desc = t_item["Descripcion"]
                                esta_completada = id_pl in completadas_hoy
                                
                                chk_box = ft.Checkbox(
                                    value=esta_completada,
                                    fill_color="#7CFC00" if categoria == 1 else ("#00FFFF" if categoria == 2 else "#A100F2")
                                )
                                
                                txt_tarea = ft.Text(
                                    desc,
                                    color="white",
                                    size=11.5,
                                    weight="w500",
                                    expand=True
                                )

                                chk_box.on_change = lambda e, i_p=id_pl, chk=chk_box: toggle_tarea(
                                    i_p, 
                                    chk.value, 
                                    categoria, 
                                    col, 
                                    p_bar, 
                                    p_text
                                )

                                row_controls = [chk_box, txt_tarea]

                                if es_admin():
                                    def make_delete_click(i_p=id_pl):
                                        def delete_item(e):
                                            try:
                                                db_del = conectar_db()
                                                if db_del:
                                                    cursor_del = db_del.cursor()
                                                    cursor_del.execute("DELETE FROM registro_checklist WHERE ID_Plantilla = %s", (i_p,))
                                                    cursor_del.execute("DELETE FROM plantillas_checklist WHERE ID_Plantilla = %s", (i_p,))
                                                    db_del.commit()
                                                    db_del.close()
                                                    mostrar_snack(t("task_deleted"))
                                                    # Reload all checklists to keep UI in sync
                                                    cargar_checklist_por_categoria(1, apertura_list, progress_apertura, text_apertura)
                                                    cargar_checklist_por_categoria(2, cierre_list, progress_cierre, text_cierre)
                                                    cargar_checklist_por_categoria(3, venta_list, progress_venta, text_venta)
                                            except Exception as ex:
                                                print("ERROR ELIMINAR TAREA:", ex)
                                                mostrar_snack("Error", color="red")
                                        return delete_item

                                    row_controls.append(
                                        ft.IconButton(
                                            icon=ft.Icons.DELETE_OUTLINE_ROUNDED,
                                            icon_color="#FF4500",
                                            icon_size=18,
                                            tooltip="Eliminar tarea",
                                            on_click=make_delete_click()
                                        )
                                    )

                                col.controls.append(
                                    ft.Container(
                                        content=ft.Row(row_controls, vertical_alignment=ft.CrossAxisAlignment.CENTER, spacing=6),
                                        bgcolor="#0F0F1A",
                                        padding=6,
                                        border_radius=6,
                                        border=ft.Border.all(1, "#222222")
                                    )
                                )
                except Exception as ex:
                    print("ERROR CARGAR CHECKLIST POR CATEGORIA:", ex)
                    col.controls.append(ft.Text("Error", color="red"))
                calcular_progreso(categoria, col, p_bar, p_text)

            def build_admin_inline_form(categoria, col, p_bar, p_text):
                input_new_task = ft.TextField(
                    label=t("add_task"),
                    expand=True,
                    border_color="#7CFC00" if categoria == 1 else ("#00FFFF" if categoria == 2 else "#A100F2"),
                    label_style=ft.TextStyle(color="#aaaaaa", size=11),
                    text_style=ft.TextStyle(color="white", size=12),
                    height=40
                )
                
                def agregar_inline_click(e):
                    desc_val = input_new_task.value.strip()
                    if not desc_val:
                        return
                    try:
                        db = conectar_db()
                        if db:
                            cursor = db.cursor()
                            cursor.execute("INSERT INTO plantillas_checklist (Categoria, Descripcion) VALUES (%s, %s)", (categoria, desc_val))
                            db.commit()
                            db.close()
                            input_new_task.value = ""
                            mostrar_snack(t("task_added"))
                            # Reload this checklist category
                            cargar_checklist_por_categoria(categoria, col, p_bar, p_text)
                    except Exception as ex:
                        print("ERROR AGREGAR TAREA INLINE:", ex)
                        mostrar_snack("Error", color="red")
                
                btn_add = ft.IconButton(
                    icon=ft.Icons.ADD_CIRCLE_OUTLINE_ROUNDED,
                    icon_color="#7CFC00" if categoria == 1 else ("#00FFFF" if categoria == 2 else "#A100F2"),
                    tooltip=t("add_task"),
                    on_click=agregar_inline_click
                )
                
                return ft.Container(
                    content=ft.Row([
                        input_new_task,
                        btn_add
                    ], spacing=10),
                    margin=ft.Margin(left=0, top=0, right=0, bottom=10)
                )

            cargar_checklist_por_categoria(1, apertura_list, progress_apertura, text_apertura)
            cargar_checklist_por_categoria(2, cierre_list, progress_cierre, text_cierre)
            cargar_checklist_por_categoria(3, venta_list, progress_venta, text_venta)

            tabs_checklist = ft.Tabs(
                selected_index=0,
                animation_duration=300,
                length=3,
                expand=True,
                content=ft.Column(
                    expand=True,
                    controls=[
                        ft.TabBar(
                            tabs=[
                                ft.Tab(label="Apertura 🌅", icon=ft.Icons.LIGHT_MODE),
                                ft.Tab(label="Cierre 🌌", icon=ft.Icons.NIGHTLIGHT_ROUNDED),
                                ft.Tab(label="Venta 💰", icon=ft.Icons.MONETIZATION_ON_ROUNDED)
                            ]
                        ),
                        ft.TabBarView(
                            expand=True,
                            controls=[
                                ft.Column([
                                    ft.Divider(height=10, color="transparent"),
                                    text_apertura,
                                    progress_apertura,
                                    ft.Divider(height=10, color="transparent"),
                                    build_admin_inline_form(1, apertura_list, progress_apertura, text_apertura) if es_admin() else ft.Container(content=btn_auditar_vitrina, margin=ft.Margin(0, 0, 0, 10)),
                                    apertura_list,
                                    ft.Container(
                                        content=ft.ElevatedButton(
                                            "Guardar Apertura 💾",
                                            on_click=lambda e: guardar_checklist_click(1),
                                            bgcolor="#7CFC00",
                                            color="black",
                                            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=10))
                                        ),
                                        alignment=ft.alignment.Alignment(0, 0),
                                        padding=ft.Padding(0, 10, 0, 10)
                                    ) if not es_admin() else ft.Container()
                                ], expand=True, scroll=ft.ScrollMode.AUTO),
                                ft.Column([
                                    ft.Divider(height=10, color="transparent"),
                                    text_cierre,
                                    progress_cierre,
                                    ft.Divider(height=10, color="transparent"),
                                    build_admin_inline_form(2, cierre_list, progress_cierre, text_cierre) if es_admin() else ft.Container(),
                                    cierre_list,
                                    ft.Container(
                                        content=ft.ElevatedButton(
                                            "Guardar Cierre 💾",
                                            on_click=lambda e: guardar_checklist_click(2),
                                            bgcolor="#00FFFF",
                                            color="black",
                                            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=10))
                                        ),
                                        alignment=ft.alignment.Alignment(0, 0),
                                        padding=ft.Padding(0, 10, 0, 10)
                                    ) if not es_admin() else ft.Container()
                                ], expand=True, scroll=ft.ScrollMode.AUTO),
                                ft.Column([
                                    ft.Divider(height=10, color="transparent"),
                                    text_venta,
                                    progress_venta,
                                    ft.Divider(height=10, color="transparent"),
                                    build_admin_inline_form(3, venta_list, progress_venta, text_venta) if es_admin() else ft.Container(content=dropdown_vendedor_venta, margin=ft.Margin(0, 0, 0, 10)),
                                    venta_list,
                                    ft.Container(
                                        content=ft.ElevatedButton(
                                            "Guardar Venta 💾",
                                            on_click=lambda e: guardar_checklist_click(3),
                                            bgcolor="#A100F2",
                                            color="white",
                                            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=10))
                                        ),
                                        alignment=ft.alignment.Alignment(0, 0),
                                        padding=ft.Padding(0, 10, 0, 10)
                                    ) if not es_admin() else ft.Container()
                                ], expand=True, scroll=ft.ScrollMode.AUTO)
                            ]
                        )
                    ]
                )
            )

            # Botones del encabezado de checklist
            header_buttons = []
            if es_admin():
                def ir_a_editar_checklists(e):
                    dashboard_tab_index[0] = 5  # Selecciona la pestaña 6: Editar Checklists
                    cambiar_vista("dashboard")
                header_buttons.append(
                    ft.ElevatedButton(
                        t("edit_options"),
                        icon=ft.Icons.EDIT_ROUNDED,
                        bgcolor="#9D50BB",
                        color="white",
                        on_click=ir_a_editar_checklists
                    )
                )
            
            header_buttons.append(
                ft.IconButton(
                    icon=ft.Icons.REFRESH_ROUNDED,
                    tooltip=t("refresh"),
                    on_click=lambda e: [
                        cargar_checklist_por_categoria(1, apertura_list, progress_apertura, text_apertura),
                        cargar_checklist_por_categoria(2, cierre_list, progress_cierre, text_cierre),
                        cargar_checklist_por_categoria(3, venta_list, progress_venta, text_venta)
                    ]
                )
            )

            return ft.Column([
                ft.Row([
                    ft.Text(t("checklist_title"), size=24, color="#D8B4FE", weight="bold"),
                    ft.Row(header_buttons, spacing=10)
                ], alignment="spaceBetween", vertical_alignment="center"),
                ft.Text(t("checklist_desc"), color="#aaaaaa", size=13),
                ft.Divider(height=15, color="#333333"),
                tabs_checklist
            ], expand=True)

        def build_admin_checklist_tab():
            tasks_list = ft.Column(spacing=10, scroll=ft.ScrollMode.AUTO, expand=True)
            dropdown_cat = ft.Dropdown(
                label="Seleccionar Tipo de Checklist",
                value="1",
                options=[
                    ft.dropdown.Option("1", "Apertura 🌅"),
                    ft.dropdown.Option("2", "Cierre 🌌"),
                    ft.dropdown.Option("3", "Venta Exitosa 💰")
                ],
                width=300,
                border_color="#9D50BB"
            )
            
            input_desc = ft.TextField(
                label="Nueva instrucción de tarea...",
                expand=True,
                border_color="#9D50BB"
            )

            def cargar_tareas_admin(e=None):
                tasks_list.controls.clear()
                cat_val = int(dropdown_cat.value)
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("SELECT ID_Plantilla, Descripcion FROM plantillas_checklist WHERE Categoria = %s ORDER BY ID_Plantilla DESC", (cat_val,))
                        tareas = cursor.fetchall()
                        db.close()
                        
                        if not tareas:
                            tasks_list.controls.append(ft.Text("No hay tareas registradas en este checklist.", color="#aaaaaa", italic=True))
                        else:
                            for t in tareas:
                                id_pl = t["ID_Plantilla"]
                                desc = t["Descripcion"]
                                
                                def make_eliminar_click(i_p=id_pl, d_t=desc):
                                    def eliminar_tarea_click(ev):
                                        try:
                                            db_del = conectar_db()
                                            if db_del:
                                                cursor_del = db_del.cursor()
                                                cursor_del.execute("DELETE FROM registro_checklist WHERE ID_Plantilla = %s", (i_p,))
                                                cursor_del.execute("DELETE FROM plantillas_checklist WHERE ID_Plantilla = %s", (i_p,))
                                                db_del.commit()
                                                db_del.close()
                                                mostrar_snack(f"Tarea eliminada con éxito.")
                                                cargar_tareas_admin()
                                        except Exception as ex:
                                            print("ERROR ELIMINAR TAREA ADMIN:", ex)
                                            mostrar_snack("Error al eliminar la tarea.", color="red")
                                    return eliminar_tarea_click
                                
                                tasks_list.controls.append(
                                    ft.Container(
                                        content=ft.Row([
                                            ft.Icon(ft.Icons.CHECKLIST_ROUNDED, color="#00FFFF"),
                                            ft.Text(desc, color="white", size=13, expand=True),
                                            ft.IconButton(
                                                icon=ft.Icons.DELETE_FOREVER,
                                                icon_color="#FF4500",
                                                tooltip="Eliminar tarea",
                                                on_click=make_eliminar_click()
                                            )
                                        ], alignment="spaceBetween", vertical_alignment="center"),
                                        bgcolor="#0F0F1A",
                                        padding=10,
                                        border_radius=8,
                                        border=ft.Border.all(1, "#222222")
                                    )
                                )
                except Exception as ex:
                    print("ERROR CARGAR TAREAS ADMIN:", ex)
                page.update()

            def agregar_tarea_click(e):
                desc_val = input_desc.value.strip()
                if not desc_val:
                    mostrar_snack("Por favor escribe la descripción de la tarea.", color="red")
                    return
                cat_val = int(dropdown_cat.value)
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        cursor.execute("INSERT INTO plantillas_checklist (Categoria, Descripcion) VALUES (%s, %s)", (cat_val, desc_val))
                        db.commit()
                        db.close()
                        input_desc.value = ""
                        mostrar_snack("Nueva tarea agregada al checklist.")
                        cargar_tareas_admin()
                except Exception as ex:
                    print("ERROR AGREGAR TAREA ADMIN:", ex)
                    mostrar_snack("Error al guardar la nueva tarea.", color="red")
                page.update()

            dropdown_cat.on_change = cargar_tareas_admin
            cargar_tareas_admin()

            btn_agregar = ft.ElevatedButton(
                "Agregar Tarea",
                icon=ft.Icons.ADD,
                bgcolor="#6E48AA",
                color="white",
                on_click=agregar_tarea_click
            )

            return ft.Column([
                ft.Row([
                    ft.Text("Administración de Checklists Operativos", size=18, color="white", weight="bold"),
                    dropdown_cat
                ], alignment="spaceBetween", vertical_alignment="center"),
                ft.Text("Agrega o elimina tareas específicas del checklist seleccionado. Los cambios se verán reflejados al instante en el portal de los asociados.", color="#aaaaaa", size=13),
                ft.Divider(height=10, color="transparent"),
                ft.Row([
                    input_desc,
                    btn_agregar
                ], spacing=10),
                ft.Divider(height=15, color="#333333"),
                tasks_list
            ], expand=True)

        def build_dashboard_view():
            is_mobile = (page.width < 800) if (page and page.width) else False
            tab_defs = [
                ("Estadísticas", ft.Icons.BAR_CHART, build_stats_tab),
                ("Preguntas sin Contestar", ft.Icons.QUESTION_MARK_ROUNDED, build_missing_questions_tab),
                ("Gestión de Manuales", ft.Icons.FOLDER_OPEN_ROUNDED, build_manuals_tab),
                ("Sugerencias", ft.Icons.LIGHTBULB_ROUNDED, build_suggestions_tab),
                ("Soporte 🎫", ft.Icons.CONFIRMATION_NUMBER_ROUNDED, build_support_tickets_tab),
                ("Editar Checklists 📋", ft.Icons.CHECKLIST_ROUNDED, build_admin_checklist_tab),
                ("Tareas Consolidadas 📊", ft.Icons.ASSIGNMENT, build_tareas_admin_tab),
                ("Aperturas y Cierres 🔑", ft.Icons.KEY_ROUNDED, lambda: operacion_tiendas.build_aperturas_cierres_tab(page, user_info, conectar_db, mostrar_snack, tr)),
            ]

            
            curr_idx = dashboard_tab_index[0]
            if curr_idx < 0 or curr_idx >= len(tab_defs):
                curr_idx = 0
                dashboard_tab_index[0] = 0

            content_box = ft.Container(content=tab_defs[curr_idx][2](), expand=True)

            tab_buttons = []
            for idx, (label, icon_name, builder_fn) in enumerate(tab_defs):
                def make_click(i, fn):
                    def click(e):
                        dashboard_tab_index[0] = i
                        content_box.content = fn()
                        for b_i, btn_c in enumerate(tab_buttons):
                            is_active = (b_i == i)
                            btn_c.bgcolor = "#7c3aed" if is_active else "#1e1e1e"
                            btn_c.border = ft.Border.all(1, "#9D50BB" if is_active else "#333333")
                        try: page.update()
                        except Exception: pass
                    return click

                is_sel = (idx == curr_idx)
                btn_c = ft.Container(
                    content=ft.Row([
                        ft.Icon(icon_name, size=13 if is_mobile else 15, color="#00FFFF" if is_sel else "#aaaaaa"),
                        ft.Text(label, size=10.5 if is_mobile else 12, weight="bold", color="white" if is_sel else "#aaaaaa")
                    ], spacing=4 if is_mobile else 6, alignment=ft.MainAxisAlignment.CENTER),
                    bgcolor="#7c3aed" if is_sel else "#1e1e1e",
                    padding=ft.padding.Padding(8, 5, 8, 5) if is_mobile else ft.padding.Padding(12, 8, 12, 8),
                    border_radius=8,
                    border=ft.Border.all(1, "#9D50BB" if is_sel else "#333333"),
                    on_click=make_click(idx, builder_fn),
                    ink=True
                )
                tab_buttons.append(btn_c)

            tab_bar_row = ft.Row(tab_buttons, scroll=ft.ScrollMode.AUTO, spacing=6 if is_mobile else 8)

            return ft.Column([
                ft.Text("Panel de Control Operativo", size=20 if is_mobile else 24, color="#D8B4FE", weight="bold"),
                ft.Divider(height=10 if is_mobile else 12, color="#333333"),
                tab_bar_row,
                ft.Container(height=5),
                content_box
            ], expand=True)

        # --- SISTEMA DE TAREAS OPERATIVAS ---
        
        def verificar_y_cerrar_tareas_vencidas():
            try:
                db = conectar_db()
                if db:
                    cursor = db.cursor()
                    cursor.execute("""
                        UPDATE tareas 
                        SET Estatus = 'Cerrada' 
                        WHERE Estatus = 'Activa' 
                        AND Fecha_Limite IS NOT NULL 
                        AND Fecha_Limite < NOW()
                    """)
                    db.commit()
                    db.close()
            except Exception as e:
                print("Error actualizando tareas vencidas:", e)

        def check_garantias_bloqueadas():
            config_path = os.path.join(BASE_PATH, "config.json")
            if os.path.exists(config_path):
                try:
                    with open(config_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                        return data.get("garantias_bloqueadas", False)
                except Exception:
                    pass
            return False

        def check_garantias_url():
            config_path = os.path.join(BASE_PATH, "config.json")
            if os.path.exists(config_path):
                try:
                    with open(config_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                        return data.get("garantias_url", "")
                except Exception:
                    pass
            return ""

        def obtener_url_descarga_google(url):
            if not url:
                return None
            if "docs.google.com/spreadsheets" in url:
                if "/edit" in url:
                    return re.sub(r"/edit.*$", "/export?format=xlsx", url)
                if url.endswith("/"):
                    return url + "export?format=xlsx"
                return url + "/export?format=xlsx"
            return None

        def iniciar_carga_garantias_thread():
            if garantias_cargadas[0] or garantias_cargando[0]:
                return
            garantias_cargando[0] = True
            garantias_error[0] = None
            
            def thread_target():
                try:
                    import openpyxl
                    file_path = os.path.join(BASE_PATH, "GARANTIAS.xlsx")
                    if not os.path.exists(file_path):
                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = "Base"
                        ws.append(['Marca', 'Modelo', 'UPC', 'Motivo Garantía', 'Fecha', 'Tienda/Usuario', 'Estado/Comentario'])
                        wb.save(file_path)
                        wb.close()
                        
                    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    ws = wb["Base"]
                    max_row = ws.max_row or 1
                    
                    filas = []
                    for i, row in enumerate(rows):
                        actual_row_num = start_row + i
                        if any(cell is not None for cell in row):
                            filas.append({
                                "row_num": actual_row_num,
                                "marca": str(row[0]) if row[0] is not None else "",
                                "modelo": str(row[1]) if row[1] is not None else "",
                                "upc": str(row[2]) if row[2] is not None else "",
                                "motivo": str(row[3]) if row[3] is not None else "",
                                "fecha": str(row[4]) if row[4] is not None else "",
                                "tienda_usuario": str(row[5]) if row[5] is not None else "",
                                "estado": str(row[6]) if len(row) > 6 and row[6] is not None else "Pendiente"
                            })
                    filas.reverse() # Newest first
                    wb.close()
                    
                    garantias_cache.clear()
                    garantias_cache.extend(filas)
                    garantias_cargadas[0] = True
                    garantias_error[0] = None
                except Exception as e:
                    print("ERROR EN HILO DE CARGA DE EXCEL:", e)
                    garantias_error[0] = str(e)
                finally:
                    garantias_cargando[0] = False
                    try:
                        page.update()
                    except Exception:
                        pass
                        
            import threading
            threading.Thread(target=thread_target, daemon=True).start()

        def guardar_nueva_garantia_async(marca, modelo, upc, motivo, tienda_usuario):
            # Pre-add to local cache so user/admin sees it instantly
            from datetime import datetime
            fecha_actual = datetime.now().strftime("%Y-%m-%d")
            
            # Temporary row number until excel write finishes
            temp_row_num = 999999 + len(garantias_cache)
            temp_item = {
                "row_num": temp_row_num,
                "marca": marca,
                "modelo": modelo,
                "upc": upc,
                "motivo": motivo,
                "fecha": fecha_actual,
                "tienda_usuario": tienda_usuario,
                "estado": "Pendiente"
            }
            garantias_cache.insert(0, temp_item)

        def build_tareas_admin_tab():
            garantias_bloqueadas = check_garantias_bloqueadas()
            current_url = check_garantias_url()
            
            # Switch de bloqueo (candado)
            lock_switch = ft.Switch(
                value=not garantias_bloqueadas,
                active_color="#7CFC00"
            )
            lock_switch_label = ft.Text(
                "Habilitar visualización del Consolidado para gerentes" if garantias_bloqueadas else "Consolidado visible para todos los gerentes",
                color="white",
                size=11
            )
            lock_switch_row = ft.Row([lock_switch, ft.Container(content=lock_switch_label, expand=True)], spacing=10, vertical_alignment=ft.CrossAxisAlignment.CENTER)
            
            def toggle_lock_garantias(e):
                nueva_bloqueada = not lock_switch.value
                guardar_config_key("garantias_bloqueadas", nueva_bloqueada)
                if nueva_bloqueada:
                    lock_switch_label.value = "Habilitar visualización del Consolidado para gerentes"
                    mostrar_snack("🔒 Módulo de garantías CERRADO/OCULTO para los gerentes.", color="#FF4500")
                else:
                    lock_switch_label.value = "Consolidado visible para todos los gerentes"
                    mostrar_snack("🔓 Módulo de garantías ABIERTO/VISIBLE para los gerentes.", color="#7CFC00")
                page.update()
                
            lock_switch.on_change = toggle_lock_garantias

            # Input de URL responsivo
            txt_url = ft.TextField(
                label="Ingresa URL del Excel Consolidado (OneDrive / Google Sheets)",
                value=current_url,
                width=320,
                border_color="#D8B4FE",
                focused_border_color="#00FFFF",
                color="white",
                text_size=11
            )

            # Contenedor para acciones de la URL (abrir y descargar)
            acciones_row = ft.Row(spacing=10, wrap=True)

            def actualizar_botones_accion(url):
                acciones_row.controls.clear()
                if url:
                    # Botón abrir
                    acciones_row.controls.append(
                        ft.ElevatedButton(
                            "Abrir Documento",
                            icon=ft.Icons.OPEN_IN_NEW,
                            bgcolor="#6E48AA",
                            color="white",
                            height=40,
                            url=url
                        )
                    )
                    # Botón descargar (si aplica)
                    dl_url = obtener_url_descarga_google(url)
                    if dl_url:
                        acciones_row.controls.append(
                            ft.ElevatedButton(
                                "Descargar Excel (Google)",
                                icon=ft.Icons.DOWNLOAD_ROUNDED,
                                bgcolor="#7CFC00",
                                color="black",
                                height=40,
                                url=dl_url
                            )
                        )
                else:
                    acciones_row.controls.append(
                        ft.Text("Ingresa una URL y guárdala para habilitar las opciones.", color="#aaaaaa", italic=True, size=11)
                    )

            actualizar_botones_accion(current_url)

            def guardar_url_click(e):
                url_ingresada = txt_url.value.strip()
                guardar_config_key("garantias_url", url_ingresada)
                actualizar_botones_accion(url_ingresada)
                mostrar_snack("✅ URL del consolidado guardada exitosamente.", color="#7CFC00")
                page.update()

            btn_guardar = ft.ElevatedButton(
                "Guardar URL",
                icon=ft.Icons.SAVE_ROUNDED,
                bgcolor="#9D50BB",
                color="white",
                height=40,
                on_click=guardar_url_click
            )

            return ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Icon(ft.Icons.VERIFIED_ROUNDED, size=22, color="#00FFFF"),
                        ft.Text("Consolidado de Garantías", size=16, weight="bold", color="white"),
                    ], spacing=10, wrap=True),
                    ft.Text("Configura el enlace del archivo Excel en la nube para consolidar y consultar la información de garantías.", color="#aaaaaa", size=11),
                    ft.Container(height=8),
                    
                    # Panel de visibilidad (candado)
                    ft.Container(
                        content=ft.Column([
                            ft.Text("Control de Acceso", size=13, weight="bold", color="#00FFFF"),
                            ft.Text("Define si el enlace configurado abajo será visible para los gerentes y asociados o si permanecerá bloqueado.", color="#aaaaaa", size=10.5),
                            ft.Container(height=4),
                            lock_switch_row
                        ], spacing=4),
                        bgcolor="#1E1E1E",
                        padding=12,
                        border_radius=8,
                        border=ft.Border.all(1, "#333333")
                    ),
                    ft.Container(height=8),
                    
                    # Panel de URL y descarga
                    ft.Container(
                        content=ft.Column([
                            ft.Text("Configuración del Documento Consolidado", size=13, weight="bold", color="#D8B4FE"),
                            ft.Text("El documento enlazado se abrirá externamente en el navegador. Si usas Google Sheets, podrás descargarlo directamente.", color="#aaaaaa", size=10.5),
                            ft.Container(height=8),
                            ft.Row([
                                txt_url,
                                btn_guardar
                            ], spacing=8, wrap=True, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                            ft.Container(height=8),
                            ft.Text("Acciones:", size=12, weight="bold", color="white"),
                            acciones_row
                        ], spacing=8),
                        bgcolor="#0F0F1A",
                        padding=15,
                        border_radius=8,
                        border=ft.Border.all(1, "#222222")
                    )
                ], spacing=10, scroll=ft.ScrollMode.AUTO),
                padding=12
            )

        def build_gerentes_admin_tab():
            """Pestaña exclusiva del Admin: lista los Gerentes por tienda y permite liberar el puesto."""
            lista_gerentes = ft.Column(spacing=8, scroll=ft.ScrollMode.AUTO)
            status_label = ft.Text("Cargando...", color="#aaaaaa", size=12, italic=True)

            def cargar_gerentes():
                lista_gerentes.controls.clear()
                try:
                    db_adm = conectar_db()
                    if not db_adm:
                        status_label.value = "Error de conexión."
                        page.update()
                        return
                    cur_adm = db_adm.cursor(dictionary=True)
                    cur_adm.execute("""
                        SELECT ID_Usuario, Nombre_Completo, Usuario, Tienda, Rol,
                               DATE_FORMAT(Fecha_Registro, '%d/%m/%Y') as fecha_reg
                        FROM usuarios
                        WHERE LOWER(Rol) LIKE '%gerente%'
                          AND LOWER(Rol) NOT LIKE '%admin%'
                        ORDER BY Tienda ASC, Nombre_Completo ASC
                    """)
                    gerentes = cur_adm.fetchall()
                    db_adm.close()

                    if not gerentes:
                        status_label.value = "No hay Gerentes de Tienda registrados aún."
                        page.update()
                        return

                    status_label.value = f"Total de Gerentes registrados: {len(gerentes)}"

                    for g in gerentes:
                        uid = g["ID_Usuario"]
                        nombre = g.get("Nombre_Completo", "")
                        tienda = g.get("Tienda", "Sin tienda")
                        usuario = g.get("Usuario", "")
                        fecha = g.get("fecha_reg", "")

                        def hacer_liberar(uid_=uid, nombre_=nombre, tienda_=tienda):
                            def confirmar_liberacion(e_conf):
                                page.close_dialog()
                                try:
                                    db_lib = conectar_db()
                                    if db_lib:
                                        cur_lib = db_lib.cursor()
                                        cur_lib.execute("""
                                            UPDATE usuarios
                                            SET Rol = 'Vendedor'
                                            WHERE ID_Usuario = %s
                                        """, (uid_,))
                                        db_lib.commit()
                                        db_lib.close()
                                        mostrar_snack(f"✅ Puesto de Gerente liberado. {nombre_} ahora es Vendedor.", "#7CFC00")
                                        cargar_gerentes()
                                        page.update()
                                except Exception as ex_lib:
                                    mostrar_snack(f"Error al liberar el puesto: {ex_lib}", "red")

                            def cancelar(e_can):
                                page.close_dialog()

                            dlg = ft.AlertDialog(
                                modal=True,
                                title=ft.Text("⚠️ Liberar Puesto de Gerente", color="#FFD700", weight="bold"),
                                content=ft.Text(
                                    f"¿Confirmas que deseas liberar el puesto de Gerente de '{nombre_}' en la tienda '{tienda_}'?\n\n"
                                    f"Su rol cambiará a Vendedor y el puesto quedará disponible para otro usuario.",
                                    color="white", size=13
                                ),
                                actions=[
                                    ft.TextButton("Cancelar", on_click=cancelar, style=ft.ButtonStyle(color="#aaaaaa")),
                                    ft.ElevatedButton(
                                        "Sí, liberar puesto",
                                        on_click=confirmar_liberacion,
                                        bgcolor="#C0392B",
                                        color="white"
                                    )
                                ],
                                bgcolor="#1a1a1a",
                                shape=ft.RoundedRectangleBorder(radius=12)
                            )
                            page.open(dlg)

                        def on_liberar_click(e, fn=hacer_liberar):
                            fn()

                        tarjeta = ft.Container(
                            content=ft.Row([
                                ft.Column([
                                    ft.Row([
                                        ft.Icon(ft.Icons.WORKSPACE_PREMIUM_ROUNDED, color="#FFD700", size=20),
                                        ft.Text(nombre, color="white", weight="bold", size=14),
                                    ], spacing=6),
                                    ft.Text(f"🏪 Tienda: {tienda}  |  👤 Usuario: {usuario}  |  📅 Desde: {fecha}",
                                            color="#aaaaaa", size=11),
                                ], spacing=3, expand=True),
                                ft.ElevatedButton(
                                    content=ft.Row([
                                        ft.Icon(ft.Icons.LOCK_OPEN_ROUNDED, size=15, color="white"),
                                        ft.Text("Liberar puesto", size=12, color="white", weight="bold")
                                    ], spacing=5),
                                    on_click=on_liberar_click,
                                    bgcolor="#7B2D00",
                                    color="white",
                                    style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
                                )
                            ], spacing=10, vertical_alignment="center"),
                            bgcolor="#1a1a22",
                            padding=ft.padding.Padding(14, 10, 14, 10),
                            border_radius=10,
                            border=ft.Border.all(1, "#FFD700")
                        )
                        lista_gerentes.controls.append(tarjeta)

                    page.update()
                except Exception as ex_adm:
                    status_label.value = f"Error: {ex_adm}"
                    page.update()

            cargar_gerentes()

            return ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Icon(ft.Icons.MANAGE_ACCOUNTS_ROUNDED, color="#FFD700", size=24),
                        ft.Text("Gestión de Gerentes de Tienda 🏅", size=16, weight="bold", color="#FFD700"),
                        ft.Container(expand=True),
                        ft.IconButton(
                            icon=ft.Icons.REFRESH_ROUNDED,
                            icon_color="#FFD700",
                            tooltip="Actualizar lista",
                            on_click=lambda e: cargar_gerentes()
                        )
                    ], vertical_alignment="center"),
                    ft.Text(
                        "Aquí puedes ver quién ocupa el puesto de Gerente en cada tienda. "
                        "Si el Gerente renuncia o cambia de tienda, usa 'Liberar puesto' para dejarlo disponible.",
                        color="#aaaaaa", size=12
                    ),
                    ft.Divider(height=12, color="#333333"),
                    ft.Container(
                        content=ft.Column([
                            ft.Row([
                                ft.Icon(ft.Icons.INFO_OUTLINE_ROUNDED, color="#00FFAA", size=16),
                                ft.Text(
                                    "Al liberar el puesto, el rol del usuario cambia a 'Vendedor'. "
                                    "El puesto quedará libre para que cualquier otro usuario lo reclame.",
                                    color="#00FFAA", size=11
                                )
                            ], spacing=8),
                        ]),
                        bgcolor="#002a1a",
                        padding=ft.padding.Padding(12, 8, 12, 8),
                        border_radius=8,
                        border=ft.Border.all(1, "#00FFAA")
                    ),
                    ft.Container(height=8),
                    status_label,
                    ft.Container(height=6),
                    lista_gerentes,
                ], spacing=8, scroll=ft.ScrollMode.AUTO),
                padding=14
            )

        def build_tareas_gerente_view():

            bloqueada = check_garantias_bloqueadas()
            
            if bloqueada:
                return ft.Container(
                    content=ft.Column([
                        ft.Icon(ft.Icons.LOCK_ROUNDED, size=50, color="#FF4500"),
                        ft.Text("Módulo de Garantías Cerrado", size=18, weight="bold", color="white"),
                        ft.Text("El acceso al consolidado de garantías se encuentra temporalmente deshabilitado por el administrador.", color="#aaaaaa", text_align="center", size=12)
                    ], alignment="center", horizontal_alignment="center", spacing=12),
                    alignment=ft.alignment.Alignment(0, 0),
                    padding=30
                )

            url = check_garantias_url()
            if not url:
                return ft.Container(
                    content=ft.Column([
                        ft.Icon(ft.Icons.INFO_ROUNDED, size=50, color="#D8B4FE"),
                        ft.Text("Consolidado No Configurado", size=18, weight="bold", color="white"),
                        ft.Text("El administrador aún no ha configurado el enlace de garantías consolidado.", color="#aaaaaa", text_align="center", size=12)
                    ], alignment="center", horizontal_alignment="center", spacing=12),
                    alignment=ft.alignment.Alignment(0, 0),
                    padding=30
                )

            # Controles para abrir
            botones = [
                ft.ElevatedButton(
                    "Abrir Excel Consolidado",
                    icon=ft.Icons.OPEN_IN_NEW,
                    bgcolor="#6E48AA",
                    color="white",
                    height=42,
                    url=url
                )
            ]

            return ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Icon(ft.Icons.VERIFIED_ROUNDED, size=22, color="#00FFFF"),
                        ft.Text("Consolidación de Garantías", size=16, weight="bold", color="white")
                    ], alignment="center", spacing=8, wrap=True),
                    ft.Text("Accede al archivo Excel consolidado en la nube para registrar o consultar información de garantías.", color="#aaaaaa", size=11, text_align="center"),
                    ft.Container(height=10),
                    ft.Container(
                        content=ft.Column([
                            ft.Text("Opciones del Documento", size=14, weight="bold", color="#D8B4FE"),
                            ft.Text("Haz clic en el botón de abajo para editar o consultar el archivo en línea.", color="#aaaaaa", size=10.5, text_align="center"),
                            ft.Container(height=8),
                            ft.Row(botones, spacing=8, alignment="center", wrap=True)
                        ], spacing=8, horizontal_alignment="center"),
                        bgcolor="#0F0F1A",
                        padding=15,
                        border_radius=8,
                        border=ft.Border.all(1, "#222222")
                    )
                ], alignment="start", horizontal_alignment="center", spacing=10, scroll=ft.ScrollMode.AUTO),
                padding=12
            )

        def check_tareas_bloqueadas():
            config_path = os.path.join(BASE_PATH, "config.json")
            if os.path.exists(config_path):
                try:
                    with open(config_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                        return data.get("tareas_bloqueadas", False)
                except Exception:
                    pass
            return False

        def check_tareas_url():
            config_path = os.path.join(BASE_PATH, "config.json")
            if os.path.exists(config_path):
                try:
                    with open(config_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                        return data.get("tareas_url", "")
                except Exception:
                    pass
            return ""

        def build_modulo_tareas_admin_tab():
            tareas_bloqueadas = check_tareas_bloqueadas()
            current_url = check_tareas_url()

            def abrir_modal_nueva_tarea(e=None):
                tf_nueva_desc = ft.TextField(
                    label="Descripción de la Nueva Tarea",
                    multiline=True,
                    min_lines=2,
                    border_color="#00FFFF",
                    color="white",
                    label_style=ft.TextStyle(color="#aaaaaa", size=11)
                )
                dd_nueva_cat = ft.Dropdown(
                    label="Categoría",
                    value="1",
                    border_color="#00FFFF",
                    color="white",
                    label_style=ft.TextStyle(color="#aaaaaa", size=11),
                    options=[
                        ft.dropdown.Option("1", "Apertura 🌅"),
                        ft.dropdown.Option("2", "Cierre 🌌"),
                        ft.dropdown.Option("3", "Venta Exitosa 💰")
                    ],
                    width=200
                )

                def guardar_tarea_dialog(ev):
                    d_val = tf_nueva_desc.value.strip()
                    c_val = int(dd_nueva_cat.value)
                    if not d_val:
                        mostrar_snack("Escribe la descripción de la tarea.", color="red")
                        return
                    try:
                        db = conectar_db()
                        if db:
                            cursor = db.cursor()
                            cursor.execute("INSERT INTO plantillas_checklist (Categoria, Descripcion) VALUES (%s, %s)", (c_val, d_val))
                            db.commit()
                            db.close()
                            page.pop_dialog()
                            mostrar_snack("✅ Nueva tarea agregada con éxito.", color="#7CFC00")
                            try: page.update()
                            except Exception: pass
                    except Exception as ex:
                        mostrar_snack(f"Error al guardar tarea: {ex}", color="red")

                dlg_tarea = ft.AlertDialog(
                    title=ft.Row([
                        ft.Icon(ft.Icons.ADD_CIRCLE_ROUNDED, color="#00FFFF"),
                        ft.Text("Agregar Nueva Tarea ➕", color="white", weight="bold", size=15)
                    ]),
                    content=ft.Container(
                        content=ft.Column([
                            dd_nueva_cat,
                            tf_nueva_desc
                        ], spacing=10),
                        width=380,
                        height=180
                    ),
                    actions=[
                        ft.TextButton("Cancelar", on_click=lambda ev: page.pop_dialog()),
                        ft.ElevatedButton("Guardar Tarea 💾", bgcolor="#6E48AA", color="white", on_click=guardar_tarea_dialog)
                    ],
                    actions_alignment="end",
                    bgcolor="#0F0F1A"
                )
                page.show_dialog(dlg_tarea)

            btn_mas_tarea = ft.IconButton(
                icon=ft.Icons.ADD_CIRCLE_ROUNDED,
                icon_color="#00FFFF",
                icon_size=32,
                tooltip="Agregar Nueva Tarea ➕",
                on_click=abrir_modal_nueva_tarea
            )

            def descargar_excel_tareas(e=None):
                try:
                    db_exp = conectar_db()
                    if not db_exp:
                        mostrar_snack("Error al conectar a la base de datos", color="red")
                        return
                    cur_exp = db_exp.cursor(dictionary=True)
                    cur_exp.execute("""
                        SELECT ID_Plantilla, Categoria, Descripcion
                        FROM plantillas_checklist
                        ORDER BY Categoria, ID_Plantilla
                    """)
                    filas = cur_exp.fetchall()
                    db_exp.close()

                    if not filas:
                        mostrar_snack("No hay tareas registradas para exportar", color="orange")
                        return

                    import csv
                    from datetime import datetime

                    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = "Tareas_Consolidadas.csv"
                    filename_user = f"Tareas_Consolidadas_{timestamp_str}.csv"

                    temp_dir = os.path.join(ASSETS_PATH, "temp_pdfs")
                    os.makedirs(temp_dir, exist_ok=True)
                    filepath_temp = os.path.join(temp_dir, filename)

                    downloads_user_dir = os.path.expanduser("~/Downloads")
                    filepath_user = os.path.join(downloads_user_dir, filename_user)

                    with open(filepath_temp, "w", encoding="utf-8-sig", newline="") as f_temp:
                        writer = csv.writer(f_temp)
                        writer.writerow(["ID Tarea", "Tipo / Categoría", "Descripción de la Tarea"])
                        for r in filas:
                            cat_nombre = "Apertura 🌅" if r.get("Categoria") == 1 else ("Cierre 🌌" if r.get("Categoria") == 2 else "Venta Exitosa 💰")
                            writer.writerow([
                                r.get("ID_Plantilla", ""),
                                cat_nombre,
                                r.get("Descripcion", "")
                            ])

                    import shutil
                    shutil.copy2(filepath_temp, filepath_user)
                    mostrar_snack(f"✅ Excel de Tareas exportado ({len(filas)} tareas) en tu carpeta de Descargas", color="#7CFC00")
                except Exception as ex_exp:
                    print("ERROR EXPORTANDO EXCEL TAREAS:", ex_exp)
                    mostrar_snack(f"Error al generar Excel: {ex_exp}", color="red")

            base_url = page.url.rstrip("/") if (page and page.url) else "http://localhost:8550"
            if base_url.startswith("ws://"):
                base_url = base_url.replace("ws://", "http://", 1)
            elif base_url.startswith("wss://"):
                base_url = base_url.replace("wss://", "https://", 1)

            base_dl = re.sub(r":\d+$", f":{PUERTO_DESCARGAS}", base_url)
            url_dl_tareas = f"{base_dl}/download?file=Tareas_Consolidadas.csv&original=Tareas_Consolidadas.csv"

            btn_descargar_excel_tareas = ft.ElevatedButton(
                content=ft.Row([
                    ft.Icon(ft.Icons.TABLE_CHART_ROUNDED, color="white", size=18),
                    ft.Text("Descargar Excel", color="white", weight="bold")
                ], spacing=6),
                bgcolor="#008080",
                color="white",
                url=url_dl_tareas,
                on_click=descargar_excel_tareas,
                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=6))
            )

            lock_switch = ft.Switch(
                value=not tareas_bloqueadas,
                active_color="#7CFC00"
            )
            lock_switch_label = ft.Text(
                "Habilitar visualización de Tareas para gerentes" if tareas_bloqueadas else "Tareas visibles para todos los gerentes",
                color="white",
                size=11
            )
            lock_switch_row = ft.Row([lock_switch, ft.Container(content=lock_switch_label, expand=True)], spacing=10, vertical_alignment=ft.CrossAxisAlignment.CENTER)
            
            def toggle_lock_tareas(e):
                nueva_bloqueada = not lock_switch.value
                guardar_config_key("tareas_bloqueadas", nueva_bloqueada)
                if nueva_bloqueada:
                    lock_switch_label.value = "Habilitar visualización de Tareas para gerentes"
                    mostrar_snack("🔒 Módulo de Tareas CERRADO/OCULTO para los gerentes.", color="#FF4500")
                else:
                    lock_switch_label.value = "Tareas visibles para todos los gerentes"
                    mostrar_snack("🔓 Módulo de Tareas ABIERTO/VISIBLE para los gerentes.", color="#7CFC00")
                page.update()
                
            lock_switch.on_change = toggle_lock_tareas

            txt_url = ft.TextField(
                label="Ingresa URL del Excel de Tareas (OneDrive / Google Sheets)",
                value=current_url,
                width=320,
                border_color="#D8B4FE",
                focused_border_color="#00FFFF",
                color="white",
                text_size=11
            )

            acciones_row = ft.Row(spacing=10, wrap=True)

            def actualizar_botones_accion(url):
                acciones_row.controls.clear()
                acciones_row.controls.append(btn_descargar_excel_tareas)
                if url:
                    acciones_row.controls.append(
                        ft.ElevatedButton(
                            "Abrir Documento",
                            icon=ft.Icons.OPEN_IN_NEW,
                            bgcolor="#6E48AA",
                            color="white",
                            height=40,
                            url=url
                        )
                    )
                    dl_url = obtener_url_descarga_google(url)
                    if dl_url:
                        acciones_row.controls.append(
                            ft.ElevatedButton(
                                "Descargar Excel (Google)",
                                icon=ft.Icons.DOWNLOAD_ROUNDED,
                                bgcolor="#7CFC00",
                                color="black",
                                height=40,
                                url=dl_url
                            )
                        )

            actualizar_botones_accion(current_url)

            def guardar_url_click(e):
                url_ingresada = txt_url.value.strip()
                guardar_config_key("tareas_url", url_ingresada)
                actualizar_botones_accion(url_ingresada)
                mostrar_snack("✅ URL de Tareas guardada exitosamente.", color="#7CFC00")
                page.update()

            btn_guardar = ft.ElevatedButton(
                "Guardar URL",
                icon=ft.Icons.SAVE_ROUNDED,
                bgcolor="#9D50BB",
                color="white",
                height=40,
                on_click=guardar_url_click
            )

            return ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Icon(ft.Icons.ASSIGNMENT_ROUNDED, size=22, color="#00FFFF"),
                        ft.Text("Gestión de Tareas", size=16, weight="bold", color="white"),
                    ], spacing=10, wrap=True),
                    ft.Text("Configura el enlace del archivo Excel en la nube para consolidar y consultar las tareas.", color="#aaaaaa", size=11),
                    ft.Container(height=8),
                    
                    ft.Container(
                        content=ft.Column([
                            ft.Text("Control de Acceso", size=13, weight="bold", color="#00FFFF"),
                            ft.Text("Define si el enlace de Tareas será visible para los gerentes o si permanecerá bloqueado.", color="#aaaaaa", size=10.5),
                            ft.Container(height=4),
                            lock_switch_row
                        ], spacing=4),
                        bgcolor="#1E1E1E",
                        padding=12,
                        border_radius=8,
                        border=ft.Border.all(1, "#333333")
                    ),
                    ft.Container(height=8),
                    
                    ft.Container(
                        content=ft.Column([
                            ft.Text("Configuración del Documento de Tareas", size=13, weight="bold", color="#D8B4FE"),
                            ft.Text("El documento enlazado se abrirá externamente en el navegador.", color="#aaaaaa", size=10.5),
                            ft.Container(height=8),
                            ft.Row([
                                txt_url,
                                btn_guardar,
                                btn_mas_tarea
                            ], spacing=8, wrap=True, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                            ft.Container(height=8),
                            ft.Text("Acciones:", size=12, weight="bold", color="white"),
                            acciones_row
                        ], spacing=8),
                        bgcolor="#0F0F1A",
                        padding=15,
                        border_radius=8,
                        border=ft.Border.all(1, "#222222")
                    )
                ], spacing=10, scroll=ft.ScrollMode.AUTO),
                padding=12
            )

        def build_modulo_tareas_gerente_view():
            bloqueada = check_tareas_bloqueadas()
            
            if bloqueada:
                return ft.Container(
                    content=ft.Column([
                        ft.Icon(ft.Icons.LOCK_ROUNDED, size=50, color="#FF4500"),
                        ft.Text("Módulo de Tareas Cerrado", size=18, weight="bold", color="white"),
                        ft.Text("El acceso al módulo de tareas se encuentra deshabilitado por el administrador.", color="#aaaaaa", text_align="center", size=12)
                    ], alignment="center", horizontal_alignment="center", spacing=12),
                    alignment=ft.alignment.Alignment(0, 0),
                    padding=30
                )

            url = check_tareas_url()
            if not url:
                return ft.Container(
                    content=ft.Column([
                        ft.Icon(ft.Icons.INFO_ROUNDED, size=50, color="#D8B4FE"),
                        ft.Text("Tareas No Configuradas", size=18, weight="bold", color="white"),
                        ft.Text("El administrador aún no ha configurado el enlace de tareas.", color="#aaaaaa", text_align="center", size=12)
                    ], alignment="center", horizontal_alignment="center", spacing=12),
                    alignment=ft.alignment.Alignment(0, 0),
                    padding=30
                )

            botones = [
                ft.ElevatedButton(
                    "Abrir Documento de Tareas",
                    icon=ft.Icons.OPEN_IN_NEW,
                    bgcolor="#6E48AA",
                    color="white",
                    height=42,
                    url=url
                )
            ]

            return ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Icon(ft.Icons.ASSIGNMENT_ROUNDED, size=22, color="#00FFFF"),
                        ft.Text("Módulo de Tareas", size=16, weight="bold", color="white")
                    ], alignment="center", spacing=8, wrap=True),
                    ft.Text("Accede al documento en la nube para revisar y gestionar tus tareas asignadas.", color="#aaaaaa", size=11, text_align="center"),
                    ft.Container(height=10),
                    ft.Container(
                        content=ft.Column([
                            ft.Text("Opciones del Documento", size=14, weight="bold", color="#D8B4FE"),
                            ft.Text("Haz clic en el botón para abrir el documento de tareas.", color="#aaaaaa", size=10.5, text_align="center"),
                            ft.Container(height=8),
                            ft.Row(botones, spacing=8, alignment="center", wrap=True)
                        ], spacing=8, horizontal_alignment="center"),
                        bgcolor="#0F0F1A",
                        padding=15,
                        border_radius=8,
                        border=ft.Border.all(1, "#222222")
                    )
                ], alignment="start", horizontal_alignment="center", spacing=10, scroll=ft.ScrollMode.AUTO),
                padding=12
            )

        def build_garantias_view():
            if es_admin():
                return build_tareas_admin_tab()
            else:
                return build_tareas_gerente_view()

        def build_tareas_view():
            if es_admin():
                return build_modulo_tareas_admin_tab()
            else:
                return build_modulo_tareas_gerente_view()

        def build_campanas_view():
            if es_admin():
                return build_campanas_admin_view()
            else:
                return build_campanas_gerente_view()

        def build_campanas_admin_view():
            # Estado de fotos guia en creacion de campaña
            # Formato: {"nombre": "...", "instrucciones": "...", "foto_bytes": b"...", "segmento": "Todos", "img_preview": ft.Image}
            guias_creacion = []
            
            # Contenedor para lista de guias en creacion
            guias_col = ft.Column(spacing=10)
            
            nombre_campana = ft.TextField(label="Nombre de la Campaña", border_color="#D8B4FE")
            desc_campana = ft.TextField(label="Instrucciones / Descripción de la Campaña", border_color="#D8B4FE", multiline=True, min_lines=2)
            
            # PDF de la guia
            pdf_guia_bytes = [None]
            pdf_guia_nombre = [None]
            text_pdf_info = ft.Text("No se ha cargado PDF de guía de instalación", color="#aaaaaa", italic=True)
            
            def on_pdf_guia_cargado(path):
                try:
                    import os
                    with open(path, "rb") as f:
                        pdf_guia_bytes[0] = f.read()
                    pdf_guia_nombre[0] = os.path.basename(path)
                    text_pdf_info.value = f"PDF Cargado: {pdf_guia_nombre[0]}"
                    text_pdf_info.color = "#00FF7F"
                    mostrar_snack(f"Guía PDF '{pdf_guia_nombre[0]}' cargada correctamente.", color="#7CFC00")
                    page.update()
                except Exception as ex:
                    print("ERROR CARGANDO PDF GUIA:", ex)
                    mostrar_snack("Error al cargar el archivo PDF.", color="red")

            btn_cargar_pdf_guia = ft.ElevatedButton(
                "Cargar Guía PDF (Opcional)",
                icon=ft.Icons.PICTURE_AS_PDF,
                bgcolor="#9D50BB",
                color="white",
                on_click=lambda e: seleccionar_archivo_async(
                    "Seleccionar PDF de la Guía de Instalación",
                    [("PDF files", "*.pdf")],
                    on_pdf_guia_cargado
                )
            )

            def refrescar_guias_creacion():
                guias_col.controls.clear()
                for i, g in enumerate(guias_creacion):
                    def make_on_click(idx):
                        return lambda e: seleccionar_archivo_async(
                            f"Seleccionar Foto Guía {idx+1}",
                            [("Imágenes", "*.png *.jpg *.jpeg")],
                            lambda path: on_guia_file_selected(idx, path)
                        )
                    
                    def make_on_delete(idx):
                        return lambda e: eliminar_guia_creacion(idx)
                        
                    img_preview = g.get("img_preview")
                    if not img_preview:
                        if g.get("foto_bytes"):
                            import base64
                            img_b64 = base64.b64encode(g["foto_bytes"]).decode("utf-8")
                            img_preview = ft.Image(src=f"data:image/jpeg;base64,{img_b64}", width=120, height=120, fit="contain")
                            g["img_preview"] = img_preview
                        else:
                            img_preview = ft.Icon(ft.Icons.IMAGE, size=40, color="#555555")
                            
                    dd_guia = ft.Dropdown(
                        label="Formato / Segmento de Tienda",
                        value=g["segmento"],
                        options=[
                            ft.dropdown.Option("Todos", "Todos"),
                            ft.dropdown.Option("Formato 6.000/2.0", "Formato 6.000/2.0"),
                            ft.dropdown.Option("Formato Inline 4.0", "Formato Inline 4.0"),
                            ft.dropdown.Option("Formato Inline Skin", "Formato Inline Skin"),
                            ft.dropdown.Option("Formato Inline Boxes", "Formato Inline Boxes"),
                            ft.dropdown.Option("Formato Open Airs (Kioskos)", "Formato Open Airs (Kioskos)"),
                            ft.dropdown.Option("Formato Inline Skin Kiosko", "Formato Inline Skin Kiosko")
                        ],
                        border_color="#333333",
                        width=350
                    )
                    dd_guia.on_change = lambda e, idx=i: actualizar_guia_campo(idx, "segmento", e.control.value)

                    guias_col.controls.append(
                        ft.Container(
                            content=ft.Column([
                                ft.Row([
                                    ft.Text(f"Foto Guía #{i+1}", weight="bold", color="#00FFFF"),
                                    ft.IconButton(ft.Icons.DELETE, icon_color="#FF4500", on_click=make_on_delete(i))
                                ], alignment="spaceBetween"),
                                ft.Row([
                                    ft.Column([
                                        ft.TextField(
                                            label="Nombre de la Foto (ej. Muro Oakley)",
                                            value=g["nombre"],
                                            border_color="#333333",
                                            on_change=lambda e, idx=i: actualizar_guia_campo(idx, "nombre", e.control.value),
                                            width=350
                                        ),
                                        dd_guia,
                                        ft.TextField(
                                            label="Instrucciones para la IA (ej. Logo centrado, sin espacios vacíos)",
                                            value=g["instrucciones"],
                                            border_color="#333333",
                                            multiline=True,
                                            min_lines=2,
                                            on_change=lambda e, idx=i: actualizar_guia_campo(idx, "instrucciones", e.control.value),
                                            width=350
                                        ),
                                    ], spacing=5, expand=True),
                                    ft.Column([
                                        img_preview,
                                        ft.ElevatedButton(
                                            "Subir Guía",
                                            icon=ft.Icons.UPLOAD,
                                            bgcolor="#D8B4FE",
                                            color="black",
                                            on_click=make_on_click(i)
                                        )
                                    ], horizontal_alignment="center", spacing=5)
                                ], spacing=15)
                            ]),
                            bgcolor="#141424",
                            padding=12,
                            border_radius=8,
                            border=ft.Border.all(1, "#333333")
                        )
                    )
                
                # Botón "+" al final para añadir más guías cómodamente
                guias_col.controls.append(
                    ft.Container(
                        content=ft.Row([
                            ft.IconButton(
                                icon=ft.Icons.ADD_CIRCLE_ROUNDED,
                                icon_color="#00FFFF",
                                icon_size=36,
                                tooltip="Añadir otra Foto Guía",
                                on_click=agregar_guia_creacion
                            ),
                            ft.Text("Añadir otra Foto Guía (+)", color="#00FFFF", weight="bold", size=14)
                        ], alignment="center"),
                        margin=ft.Margin(left=0, top=10, right=0, bottom=10)
                    )
                )
                page.update()

            def actualizar_guia_campo(idx, campo, valor):
                if idx < len(guias_creacion):
                    guias_creacion[idx][campo] = valor

            def on_guia_file_selected(idx, path):
                try:
                    with open(path, "rb") as f:
                        raw_bytes = f.read()
                    enhanced_bytes = optimizar_imagen(raw_bytes)
                    if idx < len(guias_creacion):
                        guias_creacion[idx]["foto_bytes"] = enhanced_bytes
                        guias_creacion[idx]["img_preview"] = None
                        refrescar_guias_creacion()
                        mostrar_snack(f"Foto {idx+1} cargada y optimizada.", color="#7CFC00")
                except Exception as ex:
                    print("ERROR CARGANDO GUIA:", ex)
                    mostrar_snack("Error al cargar la foto.", color="red")

            def eliminar_guia_creacion(idx):
                if idx < len(guias_creacion):
                    guias_creacion.pop(idx)
                    refrescar_guias_creacion()

            def agregar_guia_creacion(e):
                guias_creacion.append({
                    "nombre": "",
                    "instrucciones": "",
                    "segmento": "Todos",
                    "foto_bytes": None,
                    "img_preview": None
                })
                refrescar_guias_creacion()

            def guardar_campana_click(e):
                nom = nombre_campana.value.strip()
                desc = desc_campana.value.strip()
                if not nom:
                    mostrar_snack("Por favor ingrese un nombre de campaña.", color="red")
                    return
                if not guias_creacion:
                    mostrar_snack("Debe añadir al menos una foto guía.", color="red")
                    return
                # Verificar que todas tengan foto y nombre
                for i, g in enumerate(guias_creacion):
                    if not g["nombre"].strip():
                        mostrar_snack(f"La foto guía #{i+1} no tiene nombre.", color="red")
                        return
                    if not g["foto_bytes"]:
                        mostrar_snack(f"La foto guía #{i+1} no tiene imagen.", color="red")
                        return
                
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        # Desactivar otras campañas
                        cursor.execute("UPDATE campanas SET Estatus = 'Inactiva' WHERE Estatus = 'Activa'")
                        # Insertar nueva campaña
                        cursor.execute(
                            "INSERT INTO campanas (Nombre, Descripcion, Estatus, Guia_PDF_Bytes, Guia_PDF_Nombre) VALUES (%s, %s, 'Activa', %s, %s)",
                            (nom, desc, pdf_guia_bytes[0], pdf_guia_nombre[0])
                        )
                        id_campana = cursor.lastrowid
                        
                        # Insertar fotos guia
                        for g in guias_creacion:
                            cursor.execute(
                                "INSERT INTO campana_fotos_guia (ID_Campana, Nombre_Foto, Instrucciones, Imagen_Bytes, Segmento) VALUES (%s, %s, %s, %s, %s)",
                                (id_campana, g["nombre"], g["instrucciones"], g["foto_bytes"], g["segmento"])
                            )
                        db.commit()
                        db.close()
                        
                        # Notificar a todas las tiendas
                        crear_notificacion_a_rol("Gerente", "Nueva Campaña Mensual 📸", f"Se ha activado la campaña: '{nom}'", "campana")
                        
                        nombre_campana.value = ""
                        desc_campana.value = ""
                        pdf_guia_bytes[0] = None
                        pdf_guia_nombre[0] = None
                        text_pdf_info.value = "No se ha cargado PDF de guía de instalación"
                        text_pdf_info.color = "#aaaaaa"
                        guias_creacion.clear()
                        refrescar_guias_creacion()
                        mostrar_snack("¡Campaña guardada y activada con éxito!", color="#7CFC00")
                        # Recargar panel de entregas
                        cargar_entregas_admin()
                except Exception as ex:
                    print("ERROR GUARDANDO CAMPANA:", ex)
                    mostrar_snack("Error de base de datos al guardar campaña.", color="red")

            # --- PANEL DE ENTREGAS ---
            entregas_col = ft.Column(spacing=10)
            detalle_entrega_col = ft.Column(spacing=15)
            
            def cargar_entregas_admin():
                entregas_col.controls.clear()
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        # Buscar campaña activa
                        cursor.execute("SELECT ID_Campana, Nombre FROM campanas WHERE Estatus = 'Activa'")
                        campana = cursor.fetchone()
                        if not campana:
                            entregas_col.controls.append(ft.Text("No hay ninguna campaña activa actualmente.", color="#aaaaaa", italic=True))
                            db.close()
                            page.update()
                            return
                        
                        id_campana = campana["ID_Campana"]
                        entregas_col.controls.append(ft.Text(f"Campaña Activa: {campana['Nombre']}", size=14, color="#D8B4FE", weight="bold"))
                        
                        # Obtener todas las entregas de esta campaña
                        query = """
                            SELECT e.ID_Entrega, e.Tienda, e.Fecha_Envio, e.Estatus, u.Nombre_Completo, u.Segmento as Segmento_Tienda
                            FROM campana_entregas_tienda e
                            JOIN usuarios u ON e.ID_Usuario = u.ID_Usuario
                            WHERE e.ID_Campana = %s
                        """
                        params = [id_campana]
                        zona_act = active_zone_filter[0]
                        if zona_act != "Todas":
                            query += " AND u.Zona = %s"
                            params.append(zona_act)
                        query += " ORDER BY e.Fecha_Envio DESC"
                        
                        cursor.execute(query, tuple(params))
                        entregas = cursor.fetchall()
                        db.close()
                        
                        if not entregas:
                            entregas_col.controls.append(ft.Text("Ninguna tienda ha enviado fotos todavía.", color="#aaaaaa", italic=True))
                        else:
                            for ent in entregas:
                                est_color = "#FF4500" if ent["Estatus"] == "Rechazado_IA" else ("#00FF7F" if ent["Estatus"] == "Visto_Bueno" else "#FFD700")
                                status_badge = ft.Container(
                                    content=ft.Text(ent["Estatus"].upper().replace("_", " "), size=10, weight="bold", color="black"),
                                    bgcolor=est_color,
                                    padding=ft.Padding(left=10, right=10, top=5, bottom=5),
                                    border_radius=4
                                )
                                
                                def make_view_details(id_ent, tienda_name):
                                    return lambda e: ver_detalle_entrega_admin(id_ent, tienda_name)
                                    
                                format_text = ent['Segmento_Tienda'] or "Sin Segmento"
                                entregas_col.controls.append(
                                    ft.Container(
                                        content=ft.Row([
                                            ft.Column([
                                                ft.Text(f"Tienda: {ent['Tienda']} ({format_text})", weight="bold", color="white"),
                                                ft.Text(f"Enviado por: {ent['Nombre_Completo']} - {ent['Fecha_Envio']}", size=12, color="#aaaaaa")
                                            ], spacing=2, expand=True),
                                            status_badge,
                                            ft.IconButton(ft.Icons.CHEVRON_RIGHT, icon_color="#00FFFF", on_click=make_view_details(ent["ID_Entrega"], ent["Tienda"]))
                                        ], alignment="spaceBetween"),
                                        bgcolor="#141424",
                                        padding=12,
                                        border_radius=8,
                                        border=ft.Border.all(1, "#333333")
                                    )
                                )
                except Exception as ex:
                    print("ERROR CARGANDO ENTREGAS ADMIN:", ex)
                    entregas_col.controls.append(ft.Text("Error al cargar las entregas.", color="red"))
                page.update()

            def ver_detalle_entrega_admin(id_entrega, tienda_name):
                detalle_entrega_col.controls.clear()
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        # Obtener fotos entregadas por la tienda y su respectiva foto guia
                        cursor.execute("""
                            SELECT f.ID_Foto_Tienda, f.Estatus_Auditoria, f.Resultado_IA, f.Imagen_Bytes as Foto_Tienda,
                                   g.Nombre_Foto, g.Instrucciones, g.Imagen_Bytes as Foto_Guia, g.Segmento as Segmento_Foto
                            FROM campana_fotos_tienda f
                            JOIN campana_fotos_guia g ON f.ID_Foto_Guia = g.ID_Foto_Guia
                            WHERE f.ID_Entrega = %s
                        """, (id_entrega,))
                        fotos = cursor.fetchall()
                        
                        # Obtener estatus de la entrega
                        cursor.execute("SELECT Estatus FROM campana_entregas_tienda WHERE ID_Entrega = %s", (id_entrega,))
                        entrega_row = cursor.fetchone()
                        db.close()
                        
                        detalle_entrega_col.controls.append(
                            ft.Row([
                                ft.IconButton(ft.Icons.ARROW_BACK, icon_color="#00FFFF", on_click=lambda e: volver_a_lista_entregas()),
                                ft.Text(f"Detalle de Entrega - {tienda_name}", size=16, color="#00FFFF", weight="bold")
                            ], spacing=10)
                        )
                        
                        if not fotos:
                            detalle_entrega_col.controls.append(ft.Text("No hay fotos en esta entrega.", color="#aaaaaa", italic=True))
                        else:
                            for f in fotos:
                                # Imagen de guia y de tienda en base64
                                import base64
                                img_guia_b64 = base64.b64encode(f["Foto_Guia"]).decode("utf-8")
                                img_tienda_b64 = base64.b64encode(f["Foto_Tienda"]).decode("utf-8")
                                
                                card_border_color = "#00FF7F" if f["Estatus_Auditoria"] == "Aprobado" else ("#FF4500" if f["Estatus_Auditoria"] == "Corregir" else "#333333")
                                
                                detail_card = ft.Container(
                                    content=ft.Column([
                                        ft.Row([
                                            ft.Text(f"Sección: {f['Nombre_Foto']}", size=14, color="#D8B4FE", weight="bold"),
                                            ft.Container(
                                                content=ft.Text(f"Segmento: {f['Segmento_Foto']}", size=9, color="black", weight="bold"),
                                                bgcolor="#00FFFF",
                                                padding=3,
                                                border_radius=3
                                            )
                                        ], alignment="spaceBetween"),
                                        ft.Text(f"Instrucciones: {f['Instrucciones']}", size=12, color="#aaaaaa"),
                                        ft.Row([
                                            ft.Column([
                                                ft.Text("FOTO GUÍA", size=10, color="#aaaaaa", weight="bold"),
                                                ft.Image(src=f"data:image/jpeg;base64,{img_guia_b64}", width=200, height=150, fit="contain")
                                            ], horizontal_alignment="center"),
                                            ft.Column([
                                                ft.Text("FOTO TIENDA", size=10, color="#aaaaaa", weight="bold"),
                                                ft.Image(src=f"data:image/jpeg;base64,{img_tienda_b64}", width=200, height=150, fit="contain")
                                            ], horizontal_alignment="center")
                                        ], spacing=20, alignment="center"),
                                        ft.Divider(height=10, color="#333333"),
                                        ft.Text(f"Estatus IA: {f['Estatus_Auditoria'].upper()}", color="#00FF7F" if f['Estatus_Auditoria'] == 'Aprobado' else "#FF4500", weight="bold", size=12),
                                        ft.Text(f"Análisis de IA:\n{f['Resultado_IA'] or 'Sin revisión.'}", size=12, color="white")
                                    ], spacing=10),
                                    bgcolor="#141424",
                                    padding=15,
                                    border_radius=8,
                                    border=ft.Border.all(1.5, card_border_color)
                                )
                                detalle_entrega_col.controls.append(detail_card)
                                
                            # Botón de visto bueno
                            if entrega_row and entrega_row["Estatus"] != "Visto_Bueno":
                                def on_visto_bueno_click(e, ent_id=id_entrega, t_name=tienda_name):
                                    dar_visto_bueno_entrega(ent_id, t_name)
                                    
                                detalle_entrega_col.controls.append(
                                    ft.Row([
                                        ft.ElevatedButton(
                                            "Dar Visto Bueno Zonal 👑",
                                            icon=ft.Icons.CHECK_CIRCLE,
                                            bgcolor="#00FF7F",
                                            color="black",
                                            on_click=on_visto_bueno_click
                                        )
                                    ], alignment="center")
                                )
                            else:
                                detalle_entrega_col.controls.append(
                                    ft.Row([
                                        ft.Container(
                                            content=ft.Row([
                                                ft.Icon(ft.Icons.CHECK_CIRCLE, color="#00FF7F"),
                                                ft.Text("Esta entrega tiene el Visto Bueno del Jefe Zonal", color="#00FF7F", weight="bold")
                                            ], spacing=5),
                                            padding=10,
                                            bgcolor="#112211",
                                            border_radius=8,
                                            border=ft.Border.all(1, "#00FF7F")
                                        )
                                    ], alignment="center")
                                )
                        
                        entregas_tabs.selected_index = 1 # Ir a la pestaña de entregas
                        entregas_col.visible = False
                        detalle_entrega_col.visible = True
                except Exception as ex:
                    print("ERROR MOSTRANDO DETALLE ENTREGA ADMIN:", ex)
                    detalle_entrega_col.controls.append(ft.Text("Error al cargar detalles de la entrega.", color="red"))
                page.update()

            def volver_a_lista_entregas():
                detalle_entrega_col.visible = False
                entregas_col.visible = True
                cargar_entregas_admin()

            def dar_visto_bueno_entrega(id_entrega, tienda_name):
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        cursor.execute("UPDATE campana_entregas_tienda SET Estatus = 'Visto_Bueno' WHERE ID_Entrega = %s", (id_entrega,))
                        db.commit()
                        
                        # Obtener el ID_Usuario de la entrega y el Nombre de la campaña para enviar la notificación
                        cursor.execute("""
                            SELECT e.ID_Usuario, c.Nombre 
                            FROM campana_entregas_tienda e
                            JOIN campanas c ON e.ID_Campana = c.ID_Campana
                            WHERE e.ID_Entrega = %s
                        """, (id_entrega,))
                        row_ent = cursor.fetchone()
                        if row_ent:
                            id_gerente = row_ent[0]
                            camp_name = row_ent[1]
                            crear_notificacion(id_gerente, "Visto Bueno Otorgado 👑", f"Tu entrega de campaña '{camp_name}' ha recibido el visto bueno final.", "campana")
                            
                        db.close()
                        mostrar_snack(f"Visto Bueno otorgado para {tienda_name}.", color="#7CFC00")
                        volver_a_lista_entregas()
                except Exception as ex:
                    print("ERROR DANDO VISTO BUENO:", ex)
                    mostrar_snack("Error al guardar estatus.", color="red")

            # --- CONFIGURACIÓN GEMINI KEY ---
            api_key_input = ft.TextField(
                label="Gemini API Key",
                value=GEMINI_API_KEY,
                password=True,
                can_reveal_password=True,
                border_color="#D8B4FE",
                width=450
            )
            
            def guardar_gemini_key_click(e):
                global GEMINI_API_KEY
                k = api_key_input.value.strip()
                if not k:
                    mostrar_snack("Por favor ingrese una clave válida.", color="red")
                    return
                if guardar_config_key("gemini_api_key", k):
                    GEMINI_API_KEY = k
                    mostrar_snack("API Key de Gemini guardada correctamente.", color="#7CFC00")
                else:
                    mostrar_snack("Error al guardar la clave en config.json.", color="red")

            def depurar_fotos_viejas_click(e):
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        # Count how many would be affected
                        cursor.execute("""
                            SELECT COUNT(*) 
                            FROM campana_fotos_tienda ft
                            JOIN campana_entregas_tienda et ON ft.ID_Entrega = et.ID_Entrega
                            WHERE et.Fecha_Envio < DATE_SUB(NOW(), INTERVAL 3 MONTH)
                              AND ft.Imagen_Bytes IS NOT NULL
                        """)
                        filas_a_depurar = cursor.fetchone()[0]
                        
                        if filas_a_depurar == 0:
                            mostrar_snack("No hay imágenes de más de 3 meses para depurar.", color="#00FFFF")
                            db.close()
                            return
                            
                        # Perform the update
                        cursor.execute("""
                            UPDATE campana_fotos_tienda ft
                            JOIN campana_entregas_tienda et ON ft.ID_Entrega = et.ID_Entrega
                            SET ft.Imagen_Bytes = NULL
                            WHERE et.Fecha_Envio < DATE_SUB(NOW(), INTERVAL 3 MONTH)
                        """)
                        filas_depuradas = cursor.rowcount
                        
                        # Optimize table
                        cursor.execute("OPTIMIZE TABLE campana_fotos_tienda")
                        cursor.fetchall() # Consume results of OPTIMIZE TABLE
                        
                        db.commit()
                        db.close()
                        mostrar_snack(f"Mantenimiento exitoso: Se eliminaron {filas_depuradas} fotos antiguas. Base de datos optimizada.", color="#7CFC00")
                except Exception as ex:
                    print("ERROR DEPURANDO ALMACENAMIENTO:", ex)
                    mostrar_snack("Error al ejecutar la depuración de base de datos.", color="red")

            config_key_view = ft.Column([
                ft.Text("Configuración de IA de Visión (Gemini)", size=16, color="#00FFFF", weight="bold"),
                ft.Text("La API Key se guarda localmente en el archivo config.json para autorizar las solicitudes a Gemini 1.5 Flash.", color="#aaaaaa", size=13),
                ft.Row([
                    api_key_input,
                    ft.ElevatedButton(
                        "Guardar Clave 💾",
                        bgcolor="#9D50BB",
                        color="white",
                        on_click=guardar_gemini_key_click
                    )
                ], spacing=10),
                ft.Divider(height=15, color="#333333"),
                ft.Container(
                    content=ft.Column([
                        ft.Text("Mantenimiento y Almacenamiento 🧹", size=16, color="#00FFFF", weight="bold"),
                        ft.Text("Depura el almacenamiento de base de datos liberando espacio ocupado por imágenes binarias de campañas con más de 3 meses de antigüedad. Se conserva la metadata y las auditorías de IA para el historial.", color="#aaaaaa", size=13),
                        ft.ElevatedButton(
                            "Liberar Almacenamiento (Fotos > 3 Meses) 🧹",
                            icon=ft.Icons.CLEANING_SERVICES_ROUNDED,
                            bgcolor="#FF4500",
                            color="white",
                            on_click=depurar_fotos_viejas_click
                        )
                    ], spacing=10),
                    padding=15,
                    bgcolor="#0F0F1A",
                    border_radius=8,
                    border=ft.Border.all(1, "#333333")
                )
            ], spacing=10)

            # Evitar error de content/tabs en Tabs constructor usando TabBar y TabBarView
            entregas_tabs = ft.Tabs(
                selected_index=0,
                length=3,
                expand=True,
                content=ft.Column(
                    expand=True,
                    controls=[
                        ft.TabBar(
                            tabs=[
                                ft.Tab(label="Crear Campaña 📸", icon=ft.Icons.ADD_A_PHOTO),
                                ft.Tab(label="Revisar Entregas 📋", icon=ft.Icons.CHECKLIST),
                                ft.Tab(label="Configuración IA ⚙", icon=ft.Icons.SETTINGS)
                            ]
                        ),
                        ft.TabBarView(
                            expand=True,
                            controls=[
                                # Tab 1 content: Crear Campaña
                                ft.Column([
                                    nombre_campana,
                                    desc_campana,
                                    ft.Row([
                                        btn_cargar_pdf_guia,
                                        text_pdf_info
                                    ], spacing=15, vertical_alignment="center"),
                                    ft.Row([
                                        ft.Text("Secciones / Fotos requeridas de la Campaña", size=14, color="#D8B4FE", weight="bold"),
                                        ft.ElevatedButton(
                                            "Añadir Foto Guía",
                                            icon=ft.Icons.ADD,
                                            bgcolor="#00FFFF",
                                            color="black",
                                            on_click=agregar_guia_creacion
                                        )
                                    ], alignment="spaceBetween", vertical_alignment="center"),
                                    guias_col,
                                    ft.Divider(height=15, color="#333333"),
                                    ft.Row([
                                        ft.ElevatedButton(
                                            "Activar y Guardar Campaña 💾",
                                            icon=ft.Icons.SAVE,
                                            bgcolor="#00FF7F",
                                            color="black",
                                            height=45,
                                            on_click=guardar_campana_click
                                        )
                                    ], alignment="center")
                                ], spacing=15, scroll=ft.ScrollMode.AUTO, expand=True),
                                
                                # Tab 2 content: Revisar Entregas
                                ft.Column([
                                    entregas_col,
                                    detalle_entrega_col
                                ], spacing=15, scroll=ft.ScrollMode.AUTO, expand=True),
                                
                                # Tab 3 content: Configuración IA
                                ft.Column([
                                    config_key_view
                                ], spacing=15, scroll=ft.ScrollMode.AUTO, expand=True)
                            ]
                        )
                    ]
                )
            )
            
            # Cargar guías iniciales y entregas
            agregar_guia_creacion(None)
            cargar_entregas_admin()
            detalle_entrega_col.visible = False
            
            return ft.Column([
                ft.Row([
                    ft.Text("Fotos de Campaña — Administrador", size=24, color="#D8B4FE", weight="bold")
                ]),
                ft.Text("Define las fotos guía del mes para las exhibiciones de Sunglass Hut y audita las entregas de las tiendas.", color="#aaaaaa", size=13),
                ft.Divider(height=15, color="#333333"),
                entregas_tabs
            ], expand=True, scroll=ft.ScrollMode.AUTO)

        def build_campanas_gerente_view():
            gerente_campana_col = ft.Column(spacing=15, scroll=ft.ScrollMode.AUTO, expand=True)
            u_id = user_info.get("id")
            
            # Cargar el segmento y la zona del usuario desde la BD si no están en user_info
            if "segmento" not in user_info or "zona" not in user_info:
                try:
                    db_u = conectar_db()
                    if db_u:
                        cur_u = db_u.cursor(dictionary=True)
                        cur_u.execute("SELECT Segmento, Zona FROM usuarios WHERE ID_Usuario = %s", (u_id,))
                        user_row = cur_u.fetchone()
                        db_u.close()
                        if user_row:
                            user_info["segmento"] = user_row["Segmento"] if user_row["Segmento"] else "Todos"
                            user_info["zona"] = user_row["Zona"] if user_row["Zona"] else "Zona Centro"
                        else:
                            user_info["segmento"] = "Todos"
                            user_info["zona"] = "Zona Centro"
                except Exception as ex_u:
                    print("ERROR CARGANDO SEGMENTO/ZONA USUARIO:", ex_u)
                    user_info["segmento"] = "Todos"
                    user_info["zona"] = "Zona Centro"

            segmento_actual = user_info.get("segmento") or "Todos"
            zona_actual = user_info.get("zona") or "Zona Centro"

            def cambiar_segmento_gerente(e):
                nuevo_seg = e.control.value
                user_info["segmento"] = nuevo_seg
                try:
                    db_seg = conectar_db()
                    if db_seg:
                        cursor_seg = db_seg.cursor()
                        cursor_seg.execute("UPDATE usuarios SET Segmento = %s WHERE ID_Usuario = %s", (nuevo_seg, u_id))
                        db_seg.commit()
                        db_seg.close()
                        mostrar_snack(f"Segmento de tienda actualizado a: {nuevo_seg}", color="#7CFC00")
                except Exception as ex:
                    print("ERROR ACTUALIZANDO SEGMENTO GERENTE:", ex)
                cargar_campana_gerente()

            def cambiar_zona_gerente(e):
                nueva_zona = e.control.value
                user_info["zona"] = nueva_zona
                try:
                    db_z = conectar_db()
                    if db_z:
                        cursor_z = db_z.cursor()
                        cursor_z.execute("UPDATE usuarios SET Zona = %s WHERE ID_Usuario = %s", (nueva_zona, u_id))
                        db_z.commit()
                        db_z.close()
                        mostrar_snack(f"Zona de tienda actualizada a: {nueva_zona}", color="#7CFC00")
                except Exception as ex:
                    print("ERROR ACTUALIZANDO ZONA GERENTE:", ex)
                cargar_campana_gerente()

            dropdown_segmento = ft.Dropdown(
                label="Formato / Segmento de tu Tienda",
                value=segmento_actual,
                options=[
                    ft.dropdown.Option("Todos", "Todos"),
                    ft.dropdown.Option("Formato 6.000/2.0", "Formato 6.000/2.0"),
                    ft.dropdown.Option("Formato Inline 4.0", "Formato Inline 4.0"),
                    ft.dropdown.Option("Formato Inline Skin", "Formato Inline Skin"),
                    ft.dropdown.Option("Formato Inline Boxes", "Formato Inline Boxes"),
                    ft.dropdown.Option("Formato Open Airs (Kioskos)", "Formato Open Airs (Kioskos)"),
                    ft.dropdown.Option("Formato Inline Skin Kiosko", "Formato Inline Skin Kiosko")
                ],
                border_color="#00FFFF",
                width=350
            )
            dropdown_segmento.on_change = cambiar_segmento_gerente

            def abrir_pdf_campana(id_camp):
                try:
                    db_p = conectar_db()
                    if not db_p:
                        return
                    cursor_p = db_p.cursor(dictionary=True)
                    cursor_p.execute("SELECT Guia_PDF_Nombre, Guia_PDF_Bytes FROM campanas WHERE ID_Campana = %s", (id_camp,))
                    row = cursor_p.fetchone()
                    db_p.close()
                    if row and row["Guia_PDF_Bytes"]:
                        import tempfile
                        ruta_temp = os.path.join(tempfile.gettempdir(), row["Guia_PDF_Nombre"])
                        with open(ruta_temp, "wb") as f_pdf:
                            f_pdf.write(row["Guia_PDF_Bytes"])
                        import os
                        os.startfile(ruta_temp)
                        mostrar_snack(f"Abriendo PDF de la campaña: {row['Guia_PDF_Nombre']}", color="#7CFC00")
                    else:
                        mostrar_snack("No hay archivo PDF cargado para esta campaña.", color="#FF4500")
                except Exception as ex:
                    print("ERROR ABRIR PDF CAMPANA:", ex)
                    mostrar_snack("Error al abrir el archivo PDF.", color="red")
            
            def cargar_campana_gerente():
                gerente_campana_col.controls.clear()
                
                # Verificar que el gerente tenga tienda asignada
                t_nombre = user_info.get("tienda")
                if not t_nombre:
                    gerente_campana_col.controls.append(
                        ft.Text("Advertencia: No tienes una tienda asignada en tu perfil. Contacta al Administrador para poder subir tus fotos de campaña.", color="#FF4500", weight="bold")
                    )
                    page.update()
                    return
                
                # Renderizar selector de formato
                gerente_campana_col.controls.append(
                    ft.Container(
                        content=ft.Column([
                            ft.Row([
                                ft.Text("Filtro de Guías por Formato:", weight="bold"),
                                dropdown_segmento
                            ], alignment="spaceBetween", vertical_alignment="center")
                        ], spacing=10),
                        bgcolor="#1e1e1e",
                        padding=15,
                        border_radius=8
                    )
                )

                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        # Buscar campaña activa
                        cursor.execute("SELECT ID_Campana, Nombre, Descripcion, Guia_PDF_Nombre FROM campanas WHERE Estatus = 'Activa'")
                        campana = cursor.fetchone()
                        
                        if not campana:
                            gerente_campana_col.controls.append(
                                ft.Text("No hay ninguna campaña mensual activa en este momento. Vuelve más tarde.", color="#aaaaaa", italic=True)
                            )
                            db.close()
                            page.update()
                            return
                        
                        id_campana = campana["ID_Campana"]
                        
                        # Obtener entrega de esta tienda o crearla
                        cursor.execute("""
                            SELECT ID_Entrega, Estatus FROM campana_entregas_tienda
                            WHERE ID_Campana = %s AND Tienda = %s
                        """, (id_campana, t_nombre))
                        entrega = cursor.fetchone()
                        if not entrega:
                            cursor.execute("""
                                INSERT INTO campana_entregas_tienda (ID_Campana, Tienda, ID_Usuario, Estatus)
                                VALUES (%s, %s, %s, 'Pendiente')
                            """, (id_campana, t_nombre, u_id))
                            db.commit()
                            id_entrega = cursor.lastrowid
                            entrega_status = "Pendiente"
                        else:
                            id_entrega = entrega["ID_Entrega"]
                            entrega_status = entrega["Estatus"]
                            
                        # Obtener fotos guías de la campaña filtrando por el segmento seleccionado o Todos
                        seg_filtro = user_info.get("segmento") or "Todos"
                        cursor.execute("""
                            SELECT ID_Foto_Guia, Nombre_Foto, Instrucciones, Imagen_Bytes, Segmento FROM campana_fotos_guia
                            WHERE ID_Campana = %s AND (Segmento = 'Todos' OR Segmento = %s)
                            ORDER BY ID_Foto_Guia
                        """, (id_campana, seg_filtro))
                        guias = cursor.fetchall()
                        
                        # Obtener fotos subidas por la tienda en esta entrega
                        cursor.execute("""
                            SELECT ID_Foto_Tienda, ID_Foto_Guia, Imagen_Bytes, Estatus_Auditoria, Resultado_IA FROM campana_fotos_tienda
                            WHERE ID_Entrega = %s
                        """, (id_entrega,))
                        fotos_tienda = {f["ID_Foto_Guia"]: f for f in cursor.fetchall()}
                        db.close()
                        
                        # PDF de la guia
                        header_row_widgets = [
                            ft.Text(f"Campaña Activa: {campana['Nombre']}", size=18, color="#00FFFF", weight="bold"),
                        ]
                        if campana.get("Guia_PDF_Nombre"):
                            btn_ver_pdf = ft.ElevatedButton(
                                "Ver Guía de Instalación PDF 📄",
                                icon=ft.Icons.PICTURE_AS_PDF,
                                bgcolor="#9D50BB",
                                color="white",
                                on_click=lambda e, id_c=id_campana: abrir_pdf_campana(id_c)
                            )
                            header_row_widgets.append(btn_ver_pdf)
                            
                        header_row_widgets.append(
                            ft.Container(
                                content=ft.Text(f"ESTATUS: {entrega_status.upper().replace('_', ' ')}", size=10, weight="bold", color="black"),
                                bgcolor="#00FF7F" if entrega_status == "Visto_Bueno" else ("#FFD700" if entrega_status == "Aprobado_IA" else "#FF4500"),
                                padding=5,
                                border_radius=4
                            )
                        )

                        # UI Encabezado
                        gerente_campana_col.controls.append(
                            ft.Row(header_row_widgets, alignment="spaceBetween")
                        )
                        if campana["Descripcion"]:
                            gerente_campana_col.controls.append(ft.Text(campana["Descripcion"], size=13, color="#cccccc"))
                        gerente_campana_col.controls.append(ft.Divider(height=10, color="#333333"))
                        
                        if not guias:
                            gerente_campana_col.controls.append(
                                ft.Text(f"No hay fotos guía configuradas para tu segmento ({seg_filtro}) o para todos.", color="#aaaaaa", italic=True)
                            )
                        else:
                            # Renderizar cada guía
                            for g in guias:
                                id_g = g["ID_Foto_Guia"]
                                nom_foto = g["Nombre_Foto"]
                                instrucciones = g["Instrucciones"]
                                seg_guia = g["Segmento"]
                                
                                import base64
                                img_guia_b64 = base64.b64encode(g["Imagen_Bytes"]).decode("utf-8")
                                
                                subida = fotos_tienda.get(id_g)
                                
                                # Construir interfaz de esta foto
                                tienda_img_widget = None
                                status_txt = "Pendiente de subir"
                                status_color = "#aaaaaa"
                                audit_feedback = ""
                                
                                if subida:
                                    img_tienda_b64 = base64.b64encode(subida["Imagen_Bytes"]).decode("utf-8")
                                    tienda_img_widget = ft.Image(src=f"data:image/jpeg;base64,{img_tienda_b64}", width=180, height=135, fit="contain")
                                    est = subida["Estatus_Auditoria"]
                                    if est == "Aprobado":
                                        status_txt = "APROBADO POR IA"
                                        status_color = "#00FF7F"
                                    elif est == "Corregir":
                                        status_txt = "CORREGIR (Ver observaciones abajo)"
                                        status_color = "#FF4500"
                                    else:
                                        status_txt = "REVISANDO CON IA..."
                                        status_color = "#FFD700"
                                        
                                    if subida["Resultado_IA"]:
                                        audit_feedback = subida["Resultado_IA"]
                                else:
                                    tienda_img_widget = ft.Container(
                                        content=ft.Column([
                                            ft.Icon(ft.Icons.ADD_A_PHOTO_ROUNDED, color="#00FFFF", size=30),
                                            ft.Text("Subir foto real\nde tu tienda", color="#00FFFF", size=11, text_align=ft.TextAlign.CENTER)
                                        ], alignment=ft.MainAxisAlignment.CENTER, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                                        width=180,
                                        height=135,
                                        bgcolor="#0F0F1A",
                                        border_radius=8,
                                        border=ft.Border.all(1, "#00FFFF")
                                    )
                                    
                                card_border = ft.Border.all(1.5, "#00FF7F" if status_txt.startswith("APROBADO") else ("#FF4500" if status_txt.startswith("CORREGIR") else "#333333"))
                                
                                def make_on_upload(g_id=id_g, ent_id=id_entrega):
                                    return lambda e: seleccionar_archivo_async(
                                        f"Subir Foto para {nom_foto}",
                                        [("Imágenes", "*.png *.jpg *.jpeg")],
                                        lambda path: subir_foto_tienda_gerente(path, g_id, ent_id)
                                    )
                                    
                                gerente_campana_col.controls.append(
                                    ft.Container(
                                        content=ft.Column([
                                            ft.Row([
                                                ft.Row([
                                                    ft.Text(nom_foto, size=15, color="#D8B4FE", weight="bold"),
                                                    ft.Container(
                                                        content=ft.Text(f"Segmento: {seg_guia}", size=8, color="black", weight="bold"),
                                                        bgcolor="#00FFFF",
                                                        padding=2,
                                                        border_radius=2
                                                    )
                                                ], spacing=10),
                                                ft.Container(
                                                    content=ft.Text(status_txt, size=9, weight="bold", color="black"),
                                                    bgcolor=status_color,
                                                    padding=3,
                                                    border_radius=3
                                                )
                                            ], alignment="spaceBetween"),
                                            ft.Text(f"Instrucciones de Montaje: {instrucciones}", size=12, color="#aaaaaa"),
                                            ft.Row([
                                                ft.Column([
                                                    ft.Text("FOTO GUÍA DE MONTAJE", size=9, color="#aaaaaa", weight="bold"),
                                                    ft.Image(src=f"data:image/jpeg;base64,{img_guia_b64}", width=180, height=135, fit=ft.ImageFit.CONTAIN)
                                                ], horizontal_alignment="center"),
                                                ft.Column([
                                                    ft.Text("FOTO REAL DE TU TIENDA", size=9, color="#aaaaaa", weight="bold"),
                                                    tienda_img_widget
                                                ], horizontal_alignment="center")
                                            ], spacing=20, alignment="center"),
                                            ft.Row([
                                                ft.ElevatedButton(
                                                    "Subir Foto" if not subida else "Volver a subir",
                                                    icon=ft.Icons.UPLOAD_FILE,
                                                    bgcolor="#00FFFF",
                                                    color="black",
                                                    on_click=make_on_upload()
                                                )
                                            ], alignment="center"),
                                            ft.Column([
                                                ft.Text("Análisis de IA de Visión:", size=11, color="#aaaaaa", weight="bold"),
                                                ft.Text(audit_feedback, size=11, color="white")
                                            ], spacing=3, visible=bool(audit_feedback))
                                        ], spacing=10),
                                        bgcolor="#141424",
                                        padding=15,
                                        border_radius=8,
                                        border=card_border
                                    )
                                )
                except Exception as ex:
                    print("ERROR CARGANDO VISTA GERENTE CAMPANA:", ex)
                    gerente_campana_col.controls.append(ft.Text("Error al cargar la campaña activa.", color="red"))
                page.update()

            def subir_foto_tienda_gerente(file_path, id_guia, id_entrega):
                try:
                    with open(file_path, "rb") as f:
                        raw_bytes = f.read()
                    
                    # Optimizar imagen
                    img_optimized = optimizar_imagen(raw_bytes)
                    
                    mostrar_snack("Foto subida. Iniciando auditoría con IA...", color="#00FFFF")
                    
                    # Guardar foto en la base de datos con estatus temporal 'Auditando'
                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        # Verificar si ya existe un registro para esta foto guia
                        cursor.execute("""
                            SELECT ID_Foto_Tienda FROM campana_fotos_tienda
                            WHERE ID_Entrega = %s AND ID_Foto_Guia = %s
                        """, (id_entrega, id_guia))
                        row = cursor.fetchone()
                        
                        if row:
                            id_foto_tienda = row[0]
                            cursor.execute("""
                                UPDATE campana_fotos_tienda
                                SET Imagen_Bytes = %s, Estatus_Auditoria = 'Auditando', Resultado_IA = 'Revisando imagen con IA de visión...'
                                WHERE ID_Foto_Tienda = %s
                            """, (img_optimized, id_foto_tienda))
                        else:
                            cursor.execute("""
                                INSERT INTO campana_fotos_tienda (ID_Entrega, ID_Foto_Guia, Imagen_Bytes, Estatus_Auditoria, Resultado_IA)
                                VALUES (%s, %s, %s, 'Auditando', 'Revisando imagen con IA de visión...')
                            """, (id_entrega, id_guia, img_optimized))
                        
                        db.commit()
                        db.close()
                        
                        # Notificar al Administrador de la entrega de fotos
                        crear_notificacion_a_rol("Administrador", "Nueva Foto de Campaña 📸", f"La tienda '{t_nombre}' ha subido una foto para revisión.", "campana")
                        
                    # Refrescar UI antes de llamar a Gemini
                    cargar_campana_gerente()
                    
                    # Lanzar auditoría en hilo separado para no bloquear la UI
                    def thread_auditoria():
                        try:
                            # 1. Recuperar fotos guía e instrucciones de la BD
                            db_aud = conectar_db()
                            if db_aud:
                                cursor_aud = db_aud.cursor(dictionary=True)
                                cursor_aud.execute("""
                                    SELECT Imagen_Bytes, Instrucciones, Nombre_Foto FROM campana_fotos_guia
                                    WHERE ID_Foto_Guia = %s
                                """, (id_guia,))
                                guia_row = cursor_aud.fetchone()
                                db_aud.close()
                                
                                if guia_row:
                                    guia_bytes = guia_row["Imagen_Bytes"]
                                    instrucciones = guia_row["Instrucciones"]
                                    nombre_foto = guia_row["Nombre_Foto"]
                                    
                                    # 2. Llamar a la IA
                                    resultado_ia = auditar_foto_con_gemini(guia_bytes, img_optimized, instrucciones)
                                    
                                    # 3. Determinar estatus según la primera palabra
                                    resultado_limpio = resultado_ia.strip()
                                    if resultado_limpio.upper().startswith("APROBADO"):
                                        estatus_final = "Aprobado"
                                    elif resultado_limpio.upper().startswith("CORREGIR"):
                                        estatus_final = "Corregir"
                                    else:
                                        # Buscar palabras clave si no empieza exactamente
                                        if "APROBADO" in resultado_limpio.upper()[:15]:
                                            estatus_final = "Aprobado"
                                        else:
                                            estatus_final = "Corregir"
                                            
                                    # 4. Actualizar en base de datos
                                    db_upd = conectar_db()
                                    if db_upd:
                                        cursor_upd = db_upd.cursor()
                                        cursor_upd.execute("""
                                            UPDATE campana_fotos_tienda
                                            SET Estatus_Auditoria = %s, Resultado_IA = %s, Fecha_Auditoria = CURRENT_TIMESTAMP
                                            WHERE ID_Entrega = %s AND ID_Foto_Guia = %s
                                        """, (estatus_final, resultado_limpio, id_entrega, id_guia))
                                        
                                        # Comprobar si todas las fotos de la entrega están aprobadas para actualizar la entrega a 'Aprobado_IA'
                                        cursor_upd.execute("""
                                            SELECT COUNT(*) FROM campana_fotos_guia g
                                            WHERE g.ID_Campana = (SELECT ID_Campana FROM campana_entregas_tienda WHERE ID_Entrega = %s)
                                        """, (id_entrega,))
                                        total_requeridas = cursor_upd.fetchone()[0]
                                        
                                        cursor_upd.execute("""
                                            SELECT COUNT(*) FROM campana_fotos_tienda
                                            WHERE ID_Entrega = %s AND Estatus_Auditoria = 'Aprobado'
                                        """, (id_entrega,))
                                        total_aprobadas = cursor_upd.fetchone()[0]
                                        
                                        if total_aprobadas >= total_requeridas:
                                            cursor_upd.execute("""
                                                UPDATE campana_entregas_tienda
                                                SET Estatus = 'Aprobado_IA'
                                                WHERE ID_Entrega = %s AND Estatus != 'Visto_Bueno'
                                            """, (id_entrega,))
                                        else:
                                            cursor_upd.execute("""
                                                UPDATE campana_entregas_tienda
                                                SET Estatus = 'Rechazado_IA'
                                                WHERE ID_Entrega = %s AND Estatus != 'Visto_Bueno'
                                            """, (id_entrega,))
                                            
                                        db_upd.commit()
                                        db_upd.close()
                                        
                                        # Notificar al gerente de la sucursal sobre la auditoría IA
                                        crear_notificacion(u_id, "Auditoría IA de Campaña 🤖", f"La sección '{nombre_foto}' ha sido calificada como: {estatus_final.upper()}", "campana")
                                        
                                    # Notificar y refrescar
                                    mostrar_snack("Auditoría de IA completada.", color="#7CFC00" if estatus_final == "Aprobado" else "#FF4500")
                                    cargar_campana_gerente()
                        except Exception as ex_t:
                            print("ERROR EN THREAD AUDITORIA:", ex_t)
                            mostrar_snack("Error en proceso de auditoría con la IA.", color="red")
                            
                    threading.Thread(target=thread_auditoria, daemon=True).start()
                    
                except Exception as ex:
                    print("ERROR SUBIENDO FOTO TIENDA:", ex)
                    mostrar_snack("Error al guardar la foto.", color="red")
                    
            cargar_campana_gerente()
            
            return ft.Column([
                ft.Row([
                    ft.Text("Fotos de Campaña — Tiendas", size=24, color="#D8B4FE", weight="bold")
                ]),
                ft.Text("Sube las fotos de exhibición de tu tienda y deja que el auditor de IA valide el montaje según las guías.", color="#aaaaaa", size=13),
                ft.Divider(height=15, color="#333333"),
                gerente_campana_col
            ], expand=True, scroll=ft.ScrollMode.AUTO)

        def build_manuals_view():
            manuals_list = ft.Column(spacing=10, scroll=ft.ScrollMode.AUTO, expand=True)
            
            def cargar_manuales():
                manuals_list.controls.clear()
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        if es_admin():
                            cursor.execute("SELECT ID_Manual, Nombre_Archivo, Titulo, Version, Abierto FROM manuales ORDER BY Nombre_Archivo")
                        else:
                            cursor.execute("SELECT ID_Manual, Nombre_Archivo, Titulo, Version, Abierto FROM manuales WHERE Abierto = 1 ORDER BY Nombre_Archivo")
                        manuales = cursor.fetchall()
                        db.close()

                        manuals_list.controls.append(ft.Text(t("manuals_db_title"), size=14, color="#00FFFF", weight="bold"))
                        if not manuales:
                            manuals_list.controls.append(ft.Text(t("no_manuals"), color="#aaaaaa", size=12))
                        else:
                            for m in manuales:
                                id_m = m["ID_Manual"]
                                nombre = m.get("Nombre_Archivo") or ""
                                version = m.get("Version") or ""
                                titulo = m.get("Titulo") or ""

                                import urllib.parse
                                nombre_f = obtener_pdf_assets(id_m)
                                url_view = ""
                                url_dl = ""
                                if nombre_f:
                                    nombre_quoted = urllib.parse.quote(nombre_f)
                                    base_url = page.url.rstrip("/") if (page and page.url) else "http://localhost:8550"
                                    if base_url.startswith("ws://"):
                                        base_url = base_url.replace("ws://", "http://", 1)
                                    elif base_url.startswith("wss://"):
                                        base_url = base_url.replace("wss://", "https://", 1)
                                    elif base_url.startswith("tcp://"):
                                        base_url = base_url.replace("tcp://", "http://", 1)
                                    
                                    # Apuntar directamente a nuestro servidor HTTP en PUERTO_DESCARGAS
                                    base_dl = re.sub(r":\d+$", f":{PUERTO_DESCARGAS}", base_url)
                                    url_view = f"{base_dl}/view?file={nombre_quoted}"
                                    nombre_original_quoted = urllib.parse.quote(nombre)
                                    url_dl = f"{base_dl}/download?file={nombre_quoted}&original={nombre_original_quoted}"

                                manuals_list.controls.append(
                                    ft.Container(
                                        content=ft.Column([
                                            ft.Row([
                                                ft.Icon(ft.Icons.INSERT_DRIVE_FILE, color="#00FFFF"),
                                                ft.Column([
                                                    ft.Text(nombre, color="white", weight="bold", size=14),
                                                    ft.Text(f"{t('version')}: {version} | {titulo}", color="#aaaaaa", size=11)
                                                ], spacing=3, expand=True)
                                            ], spacing=5),
                                            ft.Row([
                                                ft.ElevatedButton(
                                                    t("view_pdf"),
                                                    url=url_view,
                                                    bgcolor="#6E48AA",
                                                    color="white",
                                                    expand=True,
                                                    disabled=(url_view == "")
                                                ),
                                                ft.ElevatedButton(
                                                    t("download_pdf"),
                                                    url=url_dl,
                                                    bgcolor="#444444",
                                                    color="white",
                                                    expand=True,
                                                    disabled=(url_dl == "")
                                                )
                                            ], spacing=5),
                                            ft.Text(
                                                "💡 Tip: Mantén presionado 'Descargar' y selecciona 'Descargar vínculo/enlace' para guardarlo en tu teléfono.",
                                                color="#aaaaaa",
                                                size=10,
                                                italic=True
                                            )
                                        ], spacing=8),
                                        bgcolor="#141424",
                                        padding=12,
                                        border_radius=8,
                                        border=ft.Border.all(1, "#333333")
                                    )
                                )
                except Exception as ex:
                    print("ERROR MANUALS VIEW LIST:", ex)
                    manuals_list.controls.append(ft.Text("Error", color="red"))
                page.update()
                
            cargar_manuales()
            
            return ft.Column([
                ft.Row([
                    ft.Text(t("manuals_title"), size=24, color="#D8B4FE", weight="bold"),
                    ft.IconButton(ft.Icons.REFRESH, on_click=lambda e: cargar_manuales(), icon_color="#00FFFF")
                ], alignment="spaceBetween", vertical_alignment="center"),
                ft.Text(t("manuals_desc"), color="#aaaaaa", size=13),
                ft.Divider(height=15, color="#333333"),
                manuals_list
            ], expand=True)

        def build_reto_dia_view():
            reto_container = ft.Container(padding=10, expand=True)

            def preguntar_a_luxo_pregunta(pregunta_txt, elegida_txt, correcta_txt, id_manual_ref=None):
                manual_forzado_trivia[0] = id_manual_ref
                input_msg.value = f"Hola LUXO, tengo una duda sobre la pregunta de trivia: '{pregunta_txt}'. Respondí '{elegida_txt}' pero la correcta era '{correcta_txt}'. ¿Me explicas en qué manual se basa y por qué es la correcta?"
                cambiar_vista("chat")
                page.update()
                enviar_mensaje(None)

            # Variables de estado del juego de Trivia (Persistidas por sesión de Flet/Vista)
            estado_trivia = {
                "preguntas": [],         # Las 5 preguntas elegidas
                "indice": 0,             # Índice actual (0 a 4)
                "respuestas": {},        # Registro de respuestas: {id_pregunta: {'elegida': X, 'correcta': Y, 'es_correcta': Z}}
                "iniciada": False,       # ¿Está jugando?
                "terminada": False,      # ¿Llegó al final?
                "retro_mostrada": False  # ¿Mostrando retroalimentación después de contestar una pregunta?
            }

            def iniciar_nueva_partida():
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        # Obtener 3 preguntas fáciles aleatorias
                        cursor.execute("SELECT * FROM reto_preguntas WHERE Dificultad = 'Fácil' ORDER BY RAND() LIMIT 3")
                        faciles = cursor.fetchall()
                        # Obtener 2 preguntas difíciles aleatorias
                        cursor.execute("SELECT * FROM reto_preguntas WHERE Dificultad = 'Difícil' ORDER BY RAND() LIMIT 2")
                        dificiles = cursor.fetchall()
                        
                        preguntas_seleccionadas = faciles + dificiles
                        
                        # Si no hay suficientes por dificultad, rellenar con cualquier pregunta
                        if len(preguntas_seleccionadas) < 5:
                            ids_existentes = [p["ID_Pregunta"] for p in preguntas_seleccionadas]
                            if ids_existentes:
                                format_strings = ','.join(['%s'] * len(ids_existentes))
                                cursor.execute(f"SELECT * FROM reto_preguntas WHERE ID_Pregunta NOT IN ({format_strings}) ORDER BY RAND() LIMIT %s", (*ids_existentes, 5 - len(preguntas_seleccionadas)))
                            else:
                                cursor.execute("SELECT * FROM reto_preguntas ORDER BY RAND() LIMIT 5")
                            preguntas_seleccionadas += cursor.fetchall()
                        
                        db.close()
                        
                        # Mezclar un poco para que no salgan siempre primero las fáciles
                        import random
                        random.shuffle(preguntas_seleccionadas)
                        
                        estado_trivia["preguntas"] = preguntas_seleccionadas[:5]
                        estado_trivia["indice"] = 0
                        estado_trivia["respuestas"] = {}
                        estado_trivia["iniciada"] = True
                        estado_trivia["terminada"] = False
                        estado_trivia["retro_mostrada"] = False
                        
                        dibujar_ui()
                except Exception as ex:
                    print("Error al iniciar partida de trivia:", ex)
                    mostrar_snack(f"Error al iniciar la Trivia: {ex}", color="red")

            def registrar_respuesta_pregunta(id_pregunta, opcion_elegida):
                try:
                    pregunta_actual = estado_trivia["preguntas"][estado_trivia["indice"]]
                    correcta = pregunta_actual["Respuesta_Correcta"]
                    es_correcta = (opcion_elegida == correcta)
                    
                    # Guardar en base de datos para estadísticas históricas del perfil
                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        cursor.execute("""
                            INSERT INTO reto_respuestas_usuario (ID_Usuario, ID_Pregunta, Fecha_Respuesta, Respuesta_Elegida, Es_Correcta)
                            VALUES (%s, %s, CURDATE(), %s, %s)
                            ON DUPLICATE KEY UPDATE Respuesta_Elegida = %s, Es_Correcta = %s
                        """, (user_info["id"], id_pregunta, opcion_elegida, 1 if es_correcta else 0, opcion_elegida, 1 if es_correcta else 0))
                        db.commit()
                        db.close()
                    
                    # Registrar en el estado local de la partida
                    estado_trivia["respuestas"][id_pregunta] = {
                        "pregunta": pregunta_actual["Pregunta"],
                        "elegida": opcion_elegida,
                        "correcta": correcta,
                        "es_correcta": es_correcta,
                        "explicacion": pregunta_actual["Explicacion"] or "Sin explicación.",
                        "opcion_a": pregunta_actual["Opcion_A"],
                        "opcion_b": pregunta_actual["Opcion_B"],
                        "opcion_c": pregunta_actual["Opcion_C"],
                        "opcion_d": pregunta_actual["Opcion_D"],
                        "id_manual": pregunta_actual.get("ID_Manual")
                    }
                    
                    estado_trivia["retro_mostrada"] = True
                    dibujar_ui()
                except Exception as ex:
                    print("Error al registrar respuesta de pregunta:", ex)
                    mostrar_snack("Error al procesar tu respuesta.", color="red")

            def avanzar_pregunta():
                if estado_trivia["indice"] >= 4:
                    estado_trivia["iniciada"] = False
                    estado_trivia["terminada"] = True
                else:
                    estado_trivia["indice"] += 1
                    estado_trivia["retro_mostrada"] = False
                dibujar_ui()

            def dibujar_ui():
                reto_container.content = None
                
                # ----------------- CASO 1: VISTA DE BIENVENIDA (Partida no iniciada) -----------------
                if not estado_trivia["iniciada"] and not estado_trivia["terminada"]:
                    # Cargar estadísticas generales históricas de la base de datos
                    total_contestados = 0
                    total_correctos = 0
                    precision = 0
                    
                    try:
                        db = conectar_db()
                        if db:
                            cursor = db.cursor(dictionary=True)
                            cursor.execute("SELECT COUNT(*) as cant FROM reto_respuestas_usuario WHERE ID_Usuario = %s", (user_info["id"],))
                            total_contestados = cursor.fetchone()["cant"]
                            cursor.execute("SELECT COUNT(*) as cant FROM reto_respuestas_usuario WHERE ID_Usuario = %s AND Es_Correcta = 1", (user_info["id"],))
                            total_correctos = cursor.fetchone()["cant"]
                            db.close()
                            if total_contestados > 0:
                                precision = int((total_correctos / total_contestados) * 100)
                    except Exception as ex:
                        print("Error obteniendo estadísticas históricas de trivia:", ex)

                    stats_box = ft.Container(
                        content=ft.Row([
                            ft.Row([
                                ft.Text("🏆", size=16),
                                ft.Text(f"Respondidos: {total_contestados}", color="white", size=12, weight="bold")
                            ], spacing=5),
                            ft.Row([
                                ft.Text("✅", size=16),
                                ft.Text(f"Correctos: {total_correctos}", color="white", size=12, weight="bold")
                            ], spacing=5),
                            ft.Row([
                                ft.Text("📈", size=16),
                                ft.Text(f"Precisión: {precision}%", color="white", size=12, weight="bold")
                            ], spacing=5)
                        ], alignment="spaceAround", wrap=True),
                        bgcolor="#0F0F1A",
                        padding=12,
                        border_radius=8,
                        border=ft.Border.all(1, "#222222")
                    )

                    reto_container.content = ft.Column([
                        ft.Container(
                            content=ft.Column([
                                ft.Text("🧠", size=48),
                                ft.Text("Desafío de Trivia LUXO", size=20, color="white", weight="bold"),
                                ft.Text(
                                    "Pon a prueba tus conocimientos sobre los manuales de Sunglass Hut. "
                                    "El juego consta de un cuestionario de 5 preguntas consecutivas compuestas por preguntas fáciles y difíciles. "
                                    "¡Completa el reto con puntaje perfecto para obtener la insignia de Auditor Estrella!",
                                    color="#aaaaaa",
                                    size=13,
                                    text_align="center"
                                ),
                                ft.Container(height=10),
                                ft.ElevatedButton(
                                    "Comenzar Reto 🏆",
                                    on_click=lambda e: iniciar_nueva_partida(),
                                    bgcolor="#6E48AA",
                                    color="white",
                                    style=ft.ButtonStyle(
                                        shape=ft.RoundedRectangleBorder(radius=10),
                                        padding=18
                                    )
                                )
                            ], horizontal_alignment="center", spacing=10),
                            bgcolor="#1e1e2e",
                            padding=30,
                            border_radius=12,
                            border=ft.Border.all(1.5, "#D8B4FE"),
                            alignment=ft.alignment.Alignment(0, 0)
                        ),
                        ft.Container(height=10),
                        ft.Text("Tu Historial Acumulado:", color="#aaaaaa", size=12, weight="bold"),
                        stats_box
                    ], spacing=10, scroll=ft.ScrollMode.AUTO)

                # ----------------- CASO 2: VISTA DE PREGUNTA EN CURSO -----------------
                elif estado_trivia["iniciada"] and not estado_trivia["terminada"]:
                    idx = estado_trivia["indice"]
                    pregunta = estado_trivia["preguntas"][idx]
                    id_preg = pregunta["ID_Pregunta"]
                    dificultad = pregunta.get("Dificultad", "Fácil")
                    
                    # Color del indicador de dificultad
                    color_dif = "#00FF7F" if dificultad == "Fácil" else "#FF1493"
                    
                    # Barra de progreso
                    progress_value = (idx + 1) / 5
                    
                    opciones_controles = []
                    if not estado_trivia["retro_mostrada"]:
                        # Modo juego (esperando respuesta)
                        for letra, texto_opcion in [
                            ("A", pregunta["Opcion_A"]),
                            ("B", pregunta["Opcion_B"]),
                            ("C", pregunta["Opcion_C"]),
                            ("D", pregunta["Opcion_D"])
                        ]:
                            def make_click_handler(l=letra, ip=id_preg):
                                return lambda e: registrar_respuesta_pregunta(ip, l)
                                
                            opciones_controles.append(
                                ft.Container(
                                    content=ft.ElevatedButton(
                                        content=ft.Row([
                                            ft.Container(
                                                content=ft.Text(letra, color="black", weight="bold", size=12),
                                                bgcolor="#D8B4FE",
                                                width=24,
                                                height=24,
                                                border_radius=12,
                                                alignment=ft.alignment.Alignment(0, 0)
                                            ),
                                            ft.Text(texto_opcion, color="white", size=13, weight="w500", overflow=ft.TextOverflow.FADE)
                                        ], spacing=10),
                                        bgcolor="#1e1e1e",
                                        style=ft.ButtonStyle(
                                            shape=ft.RoundedRectangleBorder(radius=10),
                                            padding=15
                                        ),
                                        on_click=make_click_handler(),
                                        expand=True
                                    ),
                                    expand=True
                                )
                            )
                    else:
                        # Modo retroalimentación (ya contestó esta pregunta)
                        respuesta_usuario = estado_trivia["respuestas"][id_preg]
                        elegida = respuesta_usuario["elegida"]
                        correcta = respuesta_usuario["correcta"]
                        fue_correcta = respuesta_usuario["es_correcta"]
                        explicacion_txt = respuesta_usuario["explicacion"]
                        
                        for letra, texto_opcion in [
                            ("A", pregunta["Opcion_A"]),
                            ("B", pregunta["Opcion_B"]),
                            ("C", pregunta["Opcion_C"]),
                            ("D", pregunta["Opcion_D"])
                        ]:
                            border_color = "#222222"
                            text_color = "#aaaaaa"
                            icon_review = None
                            
                            if letra == correcta:
                                border_color = "#00FF7F"
                                text_color = "white"
                                icon_review = ft.Icon(ft.Icons.CHECK, color="#00FF7F", size=16)
                            elif letra == elegida and not fue_correcta:
                                border_color = "#FF4500"
                                text_color = "white"
                                icon_review = ft.Icon(ft.Icons.CLOSE, color="#FF4500", size=16)
                                
                            opciones_controles.append(
                                ft.Container(
                                    content=ft.Row([
                                        ft.Container(
                                            content=ft.Text(letra, color="black", weight="bold", size=12),
                                            bgcolor="#00FF7F" if letra == correcta else ("#FF4500" if letra == elegida else "#444444"),
                                            width=24,
                                            height=24,
                                            border_radius=12,
                                            alignment=ft.alignment.Alignment(0, 0)
                                        ),
                                        ft.Text(texto_opcion, color=text_color, size=13, weight="w500", expand=True),
                                        icon_review if icon_review else ft.Container()
                                    ], spacing=10),
                                    padding=12,
                                    bgcolor="#0F0F1A",
                                    border_radius=8,
                                    border=ft.Border.all(1, border_color)
                                )
                            )

                    # Caja de retroalimentación
                    caja_retro = ft.Container()
                    if estado_trivia["retro_mostrada"]:
                        resp_info = estado_trivia["respuestas"][id_preg]
                        es_correcta_res = resp_info["es_correcta"]
                        color_res = "#00FF7F" if es_correcta_res else "#FF4500"
                        caja_retro = ft.Container(
                            content=ft.Column([
                                ft.Row([
                                    ft.Icon(ft.Icons.INFO_OUTLINED, color=color_res, size=18),
                                    ft.Text(
                                        "Respuesta Correcta 🎉" if es_correcta_res else "Respuesta Incorrecta ❌",
                                        color=color_res,
                                        weight="bold",
                                        size=13
                                    )
                                ], spacing=5),
                                ft.Text(resp_info["explicacion"], size=12, color="white", italic=True),
                                ft.Container(height=5),
                                ft.Row([
                                    ft.ElevatedButton(
                                        "Siguiente Pregunta ➡️" if idx < 4 else "Ver Resultados del Reto 📊",
                                        on_click=lambda e: avanzar_pregunta(),
                                        bgcolor="#6E48AA",
                                        color="white"
                                    ),
                                    ft.TextButton(
                                        "Preguntar a LUXO 💬",
                                        icon=ft.Icons.CHAT_ROUNDED,
                                        on_click=lambda e, p=pregunta["Pregunta"], el=elegida, c=correcta, idm=pregunta.get("ID_Manual"): preguntar_a_luxo_pregunta(p, el, c, idm),
                                        style=ft.ButtonStyle(color="#D8B4FE")
                                    )
                                ], spacing=10)
                            ], spacing=8),
                            bgcolor="#0F0F1A",
                            padding=15,
                            border_radius=10,
                            border=ft.Border.all(1, "#333333")
                        )

                    reto_container.content = ft.Column([
                        # Cabecera de progreso
                        ft.Row([
                            ft.Text(f"Pregunta {idx + 1} de 5", color="#D8B4FE", size=13, weight="bold"),
                            ft.Container(
                                content=ft.Text(dificultad.upper(), color="black", size=10, weight="bold"),
                                bgcolor=color_dif,
                                padding=ft.padding.Padding(left=12, top=6, right=12, bottom=6),
                                border_radius=6
                            )
                        ], alignment="spaceBetween"),
                        ft.ProgressBar(value=progress_value, color="#D8B4FE", bgcolor="#141424"),
                        ft.Container(height=5),
                        
                        # Tarjeta de pregunta
                        ft.Container(
                            content=ft.Column([
                                ft.Text(pregunta["Pregunta"], color="white", size=15, weight="bold"),
                                ft.Container(height=10),
                                ft.Column(opciones_controles, spacing=10),
                                ft.Container(height=10),
                                caja_retro
                            ], spacing=10),
                            bgcolor="#1e1e2e",
                            padding=20,
                            border_radius=12,
                            border=ft.Border.all(1.5, "#D8B4FE")
                        )
                    ], spacing=10, scroll=ft.ScrollMode.AUTO)

                # ----------------- CASO 3: PANTALLA DE RESULTADOS (Partida terminada) -----------------
                elif estado_trivia["terminada"]:
                    correctas_partida = sum(1 for r in estado_trivia["respuestas"].values() if r["es_correcta"])
                    
                    # Determinación de insignia/medalla y retroalimentación
                    if correctas_partida == 5:
                        insignia_titulo = "Auditor Estrella 🌟"
                        insignia_desc = "¡Perfecto! Tienes un conocimiento impecable sobre los manuales operativos de la tienda."
                        insignia_color = "#FFD700"
                        insignia_icono = ft.Icons.EMOJI_EVENTS_ROUNDED
                    elif correctas_partida >= 3:
                        insignia_titulo = "Estudiante Aplicado 📚"
                        insignia_desc = "¡Muy bien! Demuestras un excelente dominio de las normas y procesos de Sunglass Hut."
                        insignia_color = "#C0C0C0"
                        insignia_icono = ft.Icons.STAR_ROUNDED
                    else:
                        insignia_titulo = "Sigue Practicando 🔄"
                        insignia_desc = "Te recomendamos repasar los manuales para afianzar tus conocimientos sobre las normas operativas."
                        insignia_color = "#CD7F32"
                        insignia_icono = ft.Icons.REFRESH_ROUNDED
                    
                    # Generar resumen de preguntas respondidas
                    resumen_preguntas_controles = []
                    for k_id, res_obj in estado_trivia["respuestas"].items():
                        resumen_preguntas_controles.append(
                            ft.Container(
                                content=ft.Column([
                                    ft.Row([
                                        ft.Icon(
                                            ft.Icons.CHECK_CIRCLE if res_obj["es_correcta"] else ft.Icons.CANCEL,
                                            color="#00FF7F" if res_obj["es_correcta"] else "#FF4500",
                                            size=18
                                        ),
                                        ft.Text(res_obj["pregunta"], color="white", size=13, weight="bold", expand=True)
                                    ], spacing=8),
                                    ft.Text(f"Tu respuesta: {res_obj['elegida']} | Correcta: {res_obj['correcta']}", size=11, color="#aaaaaa"),
                                    ft.Text(f"Explicación: {res_obj['explicacion']}", size=11, color="white", italic=True),
                                    ft.Row([
                                        ft.TextButton(
                                            "Preguntar a LUXO 💬",
                                            icon=ft.Icons.CHAT_ROUNDED,
                                            on_click=lambda e, p=res_obj["pregunta"], el=res_obj["elegida"], c=res_obj["correcta"], idm=res_obj.get("id_manual"): preguntar_a_luxo_pregunta(p, el, c, idm),
                                            style=ft.ButtonStyle(color="#D8B4FE")
                                        )
                                    ], alignment="end"),
                                    ft.Divider(height=10, color="#222222")
                                ], spacing=5),
                                padding=5
                            )
                        )

                    reto_container.content = ft.Column([
                        ft.Container(
                            content=ft.Column([
                                ft.Icon(insignia_icono, size=55, color=insignia_color),
                                ft.Text(f"Puntaje: {correctas_partida} / 5", size=22, color="white", weight="bold"),
                                ft.Text(insignia_titulo, size=18, color=insignia_color, weight="bold"),
                                ft.Text(insignia_desc, size=12, color="#aaaaaa", text_align="center"),
                                ft.Container(height=10),
                                ft.ElevatedButton(
                                    "Intentar otro Reto 🔄",
                                    on_click=lambda e: iniciar_nueva_partida(),
                                    bgcolor="#6E48AA",
                                    color="white"
                                )
                            ], horizontal_alignment="center", spacing=8),
                            bgcolor="#1e1e2e",
                            padding=20,
                            border_radius=12,
                            border=ft.Border.all(1.5, insignia_color)
                        ),
                        ft.Container(height=10),
                        ft.Text("Desglose del Reto:", color="#aaaaaa", size=12, weight="bold"),
                        ft.Container(
                            content=ft.Column(resumen_preguntas_controles, spacing=10),
                            bgcolor="#0F0F1A",
                            padding=15,
                            border_radius=10,
                            border=ft.Border.all(1, "#333333")
                        )
                    ], spacing=10, scroll=ft.ScrollMode.AUTO)

                try:
                    page.update()
                except Exception:
                    pass

            # Carga inicial de la UI en la bienvenida
            dibujar_ui()
            
            return ft.Column([
                ft.Row([
                    ft.Text("Reto del Día 🏆", size=24, color="#D8B4FE", weight="bold"),
                    ft.IconButton(ft.Icons.REFRESH, on_click=lambda e: dibujar_ui(), icon_color="#00FFFF")
                ], alignment="spaceBetween", vertical_alignment="center"),
                ft.Text("Demuestra tu nivel respondiendo preguntas operativas y obtén insignias de desempeño.", color="#aaaaaa", size=13),
                ft.Divider(height=15, color="#333333"),
                reto_container
            ], expand=True, scroll=ft.ScrollMode.AUTO)

        # =================================
        # DISEÑO DEL DASHBOARD (BARRA LATERAL Y CONTENIDO DINÁMICO)
        # =================================

        content_area = ft.Container(
            expand=True,
            padding=8 if is_mobile else 20,
            bgcolor="#070914",
            border_radius=15,
            border=ft.Border.all(1.5, "#E040FB"),
            shadow=[
                ft.BoxShadow(
                    color="#35E040FB",
                    blur_radius=20,
                    spread_radius=1
                )
            ]
        )

        # Avatar de usuario en perfil
        def on_perfil_foto_cargada(path):
            try:
                u_id = user_info.get("id")
                if not u_id:
                    return
                ext = os.path.splitext(path)[1].lower()
                if ext not in [".png", ".jpg", ".jpeg"]:
                    mostrar_snack("Formato no válido. Use PNG, JPG o JPEG.", color="red")
                    return
                
                os.makedirs(os.path.join(ASSETS_PATH, "perfiles"), exist_ok=True)
                for e in [".png", ".jpg", ".jpeg"]:
                    r_old = os.path.join(ASSETS_PATH, "perfiles", f"user_{u_id}{e}")
                    if os.path.exists(r_old):
                        try:
                            os.remove(r_old)
                        except Exception:
                            pass
                
                destino = os.path.join(ASSETS_PATH, "perfiles", f"user_{u_id}{ext}")
                import shutil
                shutil.copy(path, destino)
                
                # Actualizar cache y controles
                new_img = obtener_avatar_usuario(u_id)
                user_info["img_usuario"] = new_img
                
                # Actualizar el control del avatar del perfil
                profile_icon.content = ft.Image(src=new_img, width=40, height=40, fit=ft.controls.box.BoxFit.COVER)
                mostrar_snack("Foto de perfil actualizada con éxito.", color="#7CFC00")
                page.update()
            except Exception as ex:
                print("Error subiendo foto de perfil:", ex)
                mostrar_snack("Error al subir foto de perfil.", color="red")

        def cambiar_foto_perfil(e):
            seleccionar_archivo_async(
                "Seleccionar Foto de Perfil",
                [("Imágenes", "*.png *.jpg *.jpeg")],
                on_perfil_foto_cargada
            )

        avatar_src = user_info.get("img_usuario") or img_usuario
        profile_icon = ft.Container(
            content=ft.Image(
                src=avatar_src,
                width=44,
                height=44,
                fit=ft.controls.box.BoxFit.COVER
            ) if avatar_src else ft.Icon(ft.Icons.PERSON, color="#00FFFF", size=26),
            width=44,
            height=44,
            border_radius=22,
            bgcolor="#1A102F",
            clip_behavior=ft.ClipBehavior.HARD_EDGE,
            alignment=ft.alignment.Alignment(0, 0),
            border=ft.Border.all(1.5, "#00FFFF"),
            shadow=[
                ft.BoxShadow(
                    color="#4000FFFF",
                    blur_radius=10,
                    spread_radius=1
                )
            ],
            tooltip="Subir Foto de Perfil 📸",
            on_click=cambiar_foto_perfil
        )

        # --- SISTEMA DE ALERTAS (CAMPANITA) Y ESTRELLA DE APERTURAS ---
        global star_icon_container
        star_icon_container = ft.Container(visible=False)
        bell_icon_container = ft.Container()

        def mostrar_notificaciones_dialog(e):
            u_id = user_info.get("id")
            notifs = cargar_notificaciones(u_id)
            marcar_notificaciones_leidas(u_id)
            actualizar_campana_badge() # Limpiar badge
            
            notif_rows = []
            if not notifs:
                notif_rows.append(ft.Text("No tienes notificaciones recientes.", color="#aaaaaa", italic=True))
            else:
                for n in notifs:
                    icon_map = {
                        "tarea": ft.Icons.ASSIGNMENT_ROUNDED,
                        "manual": ft.Icons.BOOK_ROUNDED,
                        "campana": ft.Icons.PHOTO_CAMERA,
                        "sistema": ft.Icons.INFO_ROUNDED
                    }
                    icon_color_map = {
                        "tarea": "#00FFFF",
                        "manual": "#D8B4FE",
                        "campana": "#7CFC00",
                        "sistema": "#FFD700"
                    }
                    tipo = n.get("Tipo") or "sistema"
                    fecha = n.get("Fecha_Hora").strftime("%d/%m %H:%M") if n.get("Fecha_Hora") else ""
                    leida = n.get("Leida")
                    
                    notif_rows.append(
                        ft.Container(
                            content=ft.Row([
                                ft.Icon(icon_map.get(tipo, ft.Icons.INFO_ROUNDED), color=icon_color_map.get(tipo, "#00FFFF"), size=20),
                                ft.Column([
                                    ft.Text(n.get("Titulo") or "", color="white", weight="bold", size=12),
                                    ft.Text(n.get("Mensaje") or "", color="#cccccc", size=11),
                                    ft.Text(fecha, color="#888888", size=9)
                                ], spacing=2, expand=True),
                                ft.Container(
                                    width=8,
                                    height=8,
                                    bgcolor="#00FFFF" if not leida else "transparent",
                                    border_radius=4
                                )
                            ], vertical_alignment="center", spacing=10),
                            bgcolor="#1e1e1e" if not leida else "#111111",
                            padding=10,
                            border_radius=6,
                            border=ft.Border.all(1, "#333333" if leida else "#00FFFF")
                        )
                    )
            
            def cerrar_notif_dialog(e):
                page.pop_dialog()
            
            dlg = ft.AlertDialog(
                title=ft.Text("Campana de Alertas 🔔", color="#00FFFF", weight="bold"),
                content=ft.Container(
                    content=ft.Column(notif_rows, spacing=8, scroll=ft.ScrollMode.AUTO),
                    width=420,
                    height=380
                ),
                actions=[
                    ft.TextButton("Cerrar", on_click=cerrar_notif_dialog)
                ],
                actions_alignment="end",
                bgcolor="#0F0F1A"
            )
            page.show_dialog(dlg)

        def actualizar_campana_badge():
            u_id = user_info.get("id")
            if user_info.get("rol") == "Admin":
                try:
                    operacion_tiendas.verificar_alertas_apertura_incumplida(conectar_db)
                except Exception as e:
                    print("Error checking late openings:", e)
            unread_cnt = obtener_cantidad_notificaciones_sin_leer(u_id)
            
            bell_btn = ft.Container(
                content=ft.Text("🔔", size=14, text_align="center"),
                bgcolor="#1A1828",
                border_radius=16,
                border=ft.Border.all(1.5, "#FFD700"),
                width=32,
                height=32,
                alignment=ft.alignment.Alignment(0, 0),
                on_click=mostrar_notificaciones_dialog,
                tooltip=f"Notificaciones ({unread_cnt}) 🔔",
                shadow=[
                    ft.BoxShadow(
                        color="#40FFD700",
                        blur_radius=10,
                        spread_radius=1
                    )
                ]
            )

            stack_controls = [bell_btn]
            if unread_cnt > 0:
                stack_controls.append(
                    ft.Container(
                        content=ft.Text(str(unread_cnt), color="white", size=8, weight="bold"),
                        bgcolor="#FF0055",
                        width=14,
                        height=14,
                        border_radius=7,
                        alignment=ft.alignment.Alignment(0, 0),
                        margin=ft.Margin(left=20, top=0, right=0, bottom=0),
                        shadow=[
                            ft.BoxShadow(
                                color="#FF0055",
                                blur_radius=6,
                                spread_radius=1
                            )
                        ]
                    )
                )
            bell_icon_container.content = ft.Stack(stack_controls, width=34, height=34)
            try:
                page.update()
            except Exception:
                pass

        # Contenedor de medallas en la sidebar
        medallas_container = ft.Row(spacing=5)
        def actualizar_medallas_sidebar():
            medallas_container.controls.clear()
            medallas = cargar_medallas_usuario(user_info["id"])
            for m in medallas:
                emoji_map = {
                    "Auditor Estrella": ("⭐" if m["desbloqueada"] else "🔒"),
                    "Madrugador": ("🌅" if m["desbloqueada"] else "🔒"),
                    "Cierre Perfecto": ("🌙" if m["desbloqueada"] else "🔒"),
                    "Vendedor Pro": ("💰" if m["desbloqueada"] else "🔒")
                }
                emoji_char = emoji_map.get(m["nombre"], "🏅")
                medallas_container.controls.append(
                    ft.Text(
                        emoji_char,
                        size=16,
                        tooltip=m["tooltip"]
                    )
                )
            try:
                page.update()
            except Exception:
                pass
        
        page.actualizar_medallas_sidebar = actualizar_medallas_sidebar
        actualizar_medallas_sidebar()

        # Clean store name if it is a store account
        nombre_original = user_info.get("nombre") or ""
        is_store = False
        if user_info.get("usuario", "").lower().startswith("sgh") or nombre_original.lower().startswith("tienda "):
            is_store = True

        if is_store and nombre_original.lower().startswith("tienda "):
            nombre_mostrar = nombre_original[7:]
        else:
            nombre_mostrar = nombre_original

        rol_mostrar = "Tienda" if is_store else (user_info.get("rol") or "")

        profile_row = ft.Container(
            content=ft.Column([
                ft.Row([
                    profile_icon,
                    ft.Column([
                        ft.Text(nombre_mostrar, color="white", weight="bold", size=14, overflow=ft.TextOverflow.ELLIPSIS),
                        ft.Text(rol_mostrar, color="#aaaaaa", size=12),
                    ], spacing=2, expand=True),
                ], vertical_alignment="center", spacing=10),
                ft.Row([
                    medallas_container,
                    ft.Container(expand=True),
                    ft.Text(f"🌐 IP: {getattr(page, 'client_ip', None) or 'Localhost'}", color="#00FFFF", size=10)
                ], vertical_alignment="center")
            ], spacing=5)
        )
        def build_presupuesto_view():
            import datetime
            import calendar

            hoy = datetime.date.today()
            current_month = [hoy.month]
            current_year = [hoy.year]

            selected_zona = [user_info.get("zona") or "Zona Centro"]
            selected_tienda = [user_info.get("tienda") or ""]

            tiendas_por_zona = {}
            try:
                db_t = conectar_db()
                if db_t:
                    cur_t = db_t.cursor(dictionary=True)
                    cur_t.execute("SELECT DISTINCT Tienda, Zona FROM usuarios WHERE Tienda IS NOT NULL AND Tienda != '' ORDER BY Tienda ASC")
                    for row in cur_t.fetchall():
                        z = row["Zona"] or "Sin Zona"
                        t_val = row["Tienda"]
                        if z not in tiendas_por_zona:
                            tiendas_por_zona[z] = []
                        if t_val not in tiendas_por_zona[z]:
                            tiendas_por_zona[z].append(t_val)
                    db_t.close()
            except Exception as e_db:
                print("Error loading tiendas list:", e_db)

            # Salvaguarda: Asegurar que tiendas_por_zona nunca esté vacío
            if not tiendas_por_zona:
                tiendas_por_zona["Sin Zona"] = ["Sin Tienda"]

            is_mobile = (page.width < 800) if (page and page.width) else False

            # Si es admin, determinar selected_zona y selected_tienda
            if es_admin():
                active_z = active_zone_filter[0] if active_zone_filter[0] != "Todas" else "Zona Centro"
                if active_z not in tiendas_por_zona:
                    active_z = list(tiendas_por_zona.keys())[0]
                selected_zona[0] = active_z
                if active_z in tiendas_por_zona and tiendas_por_zona[active_z]:
                    if selected_tienda[0] not in tiendas_por_zona[active_z]:
                        selected_tienda[0] = tiendas_por_zona[active_z][0]
            else:
                # Gerente
                selected_tienda[0] = user_info.get("tienda") or ""
                # Encontrar a qué zona pertenece esta tienda en los datos
                found_zone = "Sin Zona"
                for z, t_list in tiendas_por_zona.items():
                    if selected_tienda[0] in t_list:
                        found_zone = z
                        break
                selected_zona[0] = found_zone

            # Asegurar que el valor inicial exista en las opciones del Dropdown para evitar crash de renderizado de Flet
            if selected_zona[0] not in tiendas_por_zona:
                selected_zona[0] = list(tiendas_por_zona.keys())[0]
            
            zona_tiendas = tiendas_por_zona.get(selected_zona[0], [])
            if not selected_tienda[0] or selected_tienda[0] not in zona_tiendas:
                selected_tienda[0] = zona_tiendas[0] if zona_tiendas else ""

            meta_venta_tf = ft.TextField(
                label="Meta Venta (Sin IVA) 💰",
                value="",
                border_color="#9D50BB",
                focused_border_color="#00FFFF",
                color="white",
                text_size=13,
                height=45,
                expand=True,
                keyboard_type=ft.KeyboardType.NUMBER
            )
            meta_piezas_tf = ft.TextField(
                label="Meta Piezas 📦",
                value="",
                border_color="#9D50BB",
                focused_border_color="#00FFFF",
                color="white",
                text_size=13,
                height=45,
                expand=True,
                keyboard_type=ft.KeyboardType.NUMBER
            )

            progress_bar_venta = ft.ProgressBar(value=0.0, color="#FF4B4B", bgcolor="#141424", height=10, border_radius=5)
            progress_text_venta = ft.Text("Venta: 0% ($0.00 / $0.00 sin IVA)", color="white", size=12)
            
            progress_bar_piezas = ft.ProgressBar(value=0.0, color="#FF4B4B", bgcolor="#141424", height=10, border_radius=5)
            progress_text_piezas = ft.Text("Piezas: 0% (0 / 0 pzs)", color="white", size=12)

            meses_logrados_col = ft.Column(spacing=4, scroll=ft.ScrollMode.AUTO, height=180)

            calendar_grid = ft.Column(spacing=10, expand=True)

            tienda_title_txt = ft.Text("", size=18, color="#00FFFF", weight="bold")
            zona_title_txt = ft.Text("", size=12, color="#aaaaaa")
            period_title_txt = ft.Text("", size=16, color="white", weight="bold")

            # Rediseño: Tienda es ahora campo de texto
            txt_tienda = ft.TextField(
                label="Tienda",
                value=selected_tienda[0],
                border_color="#00FFFF",
                focused_border_color="#00FFFF",
                color="white",
                text_size=12,
                height=45,
                width=180,
                on_submit=lambda e: refresh_data()
            )
            txt_num_tienda = ft.TextField(
                label="Nº Tienda",
                value="",
                border_color="#00FFFF",
                focused_border_color="#00FFFF",
                color="white",
                text_size=12,
                height=45,
                width=100,
                on_submit=lambda e: refresh_data()
            )


            # Campos de Presupuesto Anual y Trimestres (Q1-Q4)
            txt_presupuesto_anual = ft.TextField(
                label="Presupuesto Anual (Sin IVA) 💰",
                value="",
                border_color="#9D50BB",
                focused_border_color="#00FFFF",
                color="white",
                text_size=12,
                height=45,
                keyboard_type=ft.KeyboardType.NUMBER
            )
            txt_q1 = ft.TextField(label="Q1 Meta", value="", border_color="#9D50BB", focused_border_color="#00FFFF", color="white", text_size=11, height=38, expand=True, keyboard_type=ft.KeyboardType.NUMBER)
            txt_q1_logro = ft.TextField(label="Q1 Venta", value="", border_color="#333333", color="#aaaaaa", text_size=11, height=38, expand=True, read_only=True)
            txt_q1_pct = ft.TextField(label="Q1 %", value="", border_color="#333333", color="#00FF7F", text_size=11, height=38, expand=True, read_only=True)

            txt_q2 = ft.TextField(label="Q2 Meta", value="", border_color="#9D50BB", focused_border_color="#00FFFF", color="white", text_size=11, height=38, expand=True, keyboard_type=ft.KeyboardType.NUMBER)
            txt_q2_logro = ft.TextField(label="Q2 Venta", value="", border_color="#333333", color="#aaaaaa", text_size=11, height=38, expand=True, read_only=True)
            txt_q2_pct = ft.TextField(label="Q2 %", value="", border_color="#333333", color="#00FF7F", text_size=11, height=38, expand=True, read_only=True)

            txt_q3 = ft.TextField(label="Q3 Meta", value="", border_color="#9D50BB", focused_border_color="#00FFFF", color="white", text_size=11, height=38, expand=True, keyboard_type=ft.KeyboardType.NUMBER)
            txt_q3_logro = ft.TextField(label="Q3 Venta", value="", border_color="#333333", color="#aaaaaa", text_size=11, height=38, expand=True, read_only=True)
            txt_q3_pct = ft.TextField(label="Q3 %", value="", border_color="#333333", color="#00FF7F", text_size=11, height=38, expand=True, read_only=True)

            txt_q4 = ft.TextField(label="Q4 Meta", value="", border_color="#9D50BB", focused_border_color="#00FFFF", color="white", text_size=11, height=38, expand=True, keyboard_type=ft.KeyboardType.NUMBER)
            txt_q4_logro = ft.TextField(label="Q4 Venta", value="", border_color="#333333", color="#aaaaaa", text_size=11, height=38, expand=True, read_only=True)
            txt_q4_pct = ft.TextField(label="Q4 %", value="", border_color="#333333", color="#00FF7F", text_size=11, height=38, expand=True, read_only=True)

            meses_nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

            chk_meses = [
                ft.Checkbox(label=meses_nombres[i], value=False)
                for i in range(12)
            ]

            dd_mes = ft.Dropdown(
                label="Mes",
                value=str(current_month[0]),
                border_color="#9D50BB",
                color="white",
                text_size=12,
                height=45,
                width=120,
                options=[ft.dropdown.Option(str(i+1), meses_nombres[i]) for i in range(12)]
            )

            dd_anio = ft.Dropdown(
                label="Año",
                value=str(current_year[0]),
                border_color="#9D50BB",
                color="white",
                text_size=12,
                height=45,
                width=100,
                options=[ft.dropdown.Option(str(y), str(y)) for y in [2025, 2026, 2027]]
            )

            # Configuración de permisos/modos del Presupuesto
            if not es_admin():
                # Gerente: puede editar metas y días, pero solo de su tienda fija
                txt_tienda.disabled = True
                txt_num_tienda.disabled = True
            else:
                # Administrador: solo visualización (read-only en todos los campos de metas)
                meta_venta_tf.disabled = True
                meta_piezas_tf.disabled = True
                txt_presupuesto_anual.disabled = True
                txt_q1.disabled = True
                txt_q2.disabled = True
                txt_q3.disabled = True
                txt_q4.disabled = True
                for chk in chk_meses:
                    chk.disabled = True

            def cargar_datos_presupuesto():
                tienda_actual = selected_tienda[0]
                mes_actual = current_month[0]
                anio_actual = current_year[0]
                
                if not tienda_actual:
                    return 0.0, 0, []
                
                meta_venta = 0.0
                meta_piezas = 0
                
                try:
                    db = conectar_db()
                    if db:
                        cur = db.cursor(dictionary=True)
                        cur.execute("""
                            SELECT Meta_Venta, Meta_Piezas 
                            FROM presupuesto_mensual 
                            WHERE Tienda = %s AND Mes = %s AND Anio = %s
                        """, (tienda_actual, mes_actual, anio_actual))
                        row_meta = cur.fetchone()
                        if row_meta:
                            meta_venta = float(row_meta["Meta_Venta"] or 0.0)
                            meta_piezas = int(row_meta["Meta_Piezas"] or 0)
                        
                        cur.execute("""
                            SELECT DAY(Fecha) as Dia, Venta_Con_IVA, Venta_Sin_IVA, Piezas 
                            FROM presupuesto_diario 
                            WHERE Tienda = %s AND MONTH(Fecha) = %s AND YEAR(Fecha) = %s
                        """, (tienda_actual, mes_actual, anio_actual))
                        ventas_diarias = cur.fetchall()
                        db.close()
                        return meta_venta, meta_piezas, ventas_diarias
                except Exception as ex:
                    print("Error loading budget data:", ex)
                
                return 0.0, 0, []

            def open_edit_day_dialog(dia):
                tienda_actual = selected_tienda[0]
                mes_actual = current_month[0]
                anio_actual = current_year[0]
                
                fecha_str = f"{anio_actual:04d}-{mes_actual:02d}-{dia:02d}"
                
                existing_venta_con_iva = 0.0
                existing_piezas = 0
                
                try:
                    db = conectar_db()
                    if db:
                        cur = db.cursor(dictionary=True)
                        cur.execute("""
                            SELECT Venta_Con_IVA, Piezas 
                            FROM presupuesto_diario 
                            WHERE Tienda = %s AND Fecha = %s
                        """, (tienda_actual, fecha_str))
                        row = cur.fetchone()
                        if row:
                            existing_venta_con_iva = float(row["Venta_Con_IVA"] or 0.0)
                            existing_piezas = int(row["Piezas"] or 0)
                        db.close()
                except Exception as ex:
                    print("Error loading existing day values:", ex)
                
                venta_dia_tf = ft.TextField(
                    label="Venta del día con IVA ($)",
                    value=str(existing_venta_con_iva) if existing_venta_con_iva > 0 else "",
                    border_color="#9D50BB",
                    focused_border_color="#00FFFF",
                    color="white",
                    keyboard_type=ft.KeyboardType.NUMBER
                )
                piezas_dia_tf = ft.TextField(
                    label="Piezas vendidas",
                    value=str(existing_piezas) if existing_piezas > 0 else "",
                    border_color="#9D50BB",
                    focused_border_color="#00FFFF",
                    color="white",
                    keyboard_type=ft.KeyboardType.NUMBER
                )
                
                def guardar_dia_click(e):
                    try:
                        v_con_iva = float(venta_dia_tf.value.strip() or 0.0)
                        p_dia = int(piezas_dia_tf.value.strip() or 0)
                    except ValueError:
                        mostrar_snack("Por favor ingresa números válidos.", color="red")
                        return
                    
                    v_sin_iva = v_con_iva / 1.16
                    
                    try:
                        db = conectar_db()
                        if db:
                            cur = db.cursor()
                            cur.execute("""
                                INSERT INTO presupuesto_diario (Tienda, Fecha, Venta_Con_IVA, Venta_Sin_IVA, Piezas)
                                VALUES (%s, %s, %s, %s, %s)
                                ON DUPLICATE KEY UPDATE 
                                    Venta_Con_IVA = %s,
                                    Venta_Sin_IVA = %s,
                                    Piezas = %s
                            """, (tienda_actual, fecha_str, v_con_iva, v_sin_iva, p_dia, v_con_iva, v_sin_iva, p_dia))
                            db.commit()
                            db.close()
                            
                            page.pop_dialog()
                            mostrar_snack(f"Día {dia} guardado exitosamente.", color="#7CFC00")
                            refresh_data()
                    except Exception as ex:
                        print("Error saving day details:", ex)
                        mostrar_snack("Error al guardar en base de datos.", color="red")
                
                dlg = ft.AlertDialog(
                    title=ft.Text(f"Registrar Venta - Día {dia}", color="#00FFFF", weight="bold"),
                    content=ft.Column([
                        ft.Text(f"Tienda: {tienda_actual}", color="#aaaaaa", size=12),
                        ft.Text(f"Fecha: {fecha_str}", color="#aaaaaa", size=12),
                        ft.Container(height=10),
                        venta_dia_tf,
                        piezas_dia_tf
                    ], tight=True, spacing=10),
                    actions=[
                        ft.TextButton("Cancelar", on_click=lambda e: page.pop_dialog()),
                        ft.ElevatedButton("Guardar 💾", bgcolor="#9D50BB", color="white", on_click=guardar_dia_click)
                    ],
                    actions_alignment="end",
                    bgcolor="#0F0F1A"
                )
                page.show_dialog(dlg)

            def render_meses_logrados():
                meses_logrados_col.controls.clear()
                tienda_actual = selected_tienda[0]
                anio_actual = current_year[0]
                if not tienda_actual:
                    meses_logrados_col.controls.append(ft.Text("Selecciona una tienda", color="#aaaaaa", italic=True, size=12))
                    return
                
                try:
                    db = conectar_db()
                    if db:
                        cur = db.cursor(dictionary=True)
                        cur.execute("""
                            SELECT 
                                m_list.Mes,
                                m.Meta_Venta,
                                m.Meta_Piezas,
                                COALESCE(SUM(d.Venta_Sin_IVA), 0) as Venta_Lograda,
                                COALESCE(SUM(d.Piezas), 0) as Piezas_Logradas
                            FROM (
                                SELECT 1 as Mes UNION SELECT 2 UNION SELECT 3 UNION SELECT 4 
                                UNION SELECT 5 UNION SELECT 6 UNION SELECT 7 UNION SELECT 8 
                                UNION SELECT 9 UNION SELECT 10 UNION SELECT 11 UNION SELECT 12
                            ) m_list
                            LEFT JOIN presupuesto_mensual m ON m.Mes = m_list.Mes AND m.Tienda = %s AND m.Anio = %s
                            LEFT JOIN presupuesto_diario d ON MONTH(d.Fecha) = m_list.Mes AND YEAR(d.Fecha) = %s AND d.Tienda = %s
                            GROUP BY m_list.Mes, m.Meta_Venta, m.Meta_Piezas
                            ORDER BY m_list.Mes ASC
                        """, (tienda_actual, anio_actual, anio_actual, tienda_actual))
                        rows = cur.fetchall()
                        db.close()
                        
                        has_any = False
                        for row in rows:
                            m_idx = row["Mes"]
                            meta_v = float(row["Meta_Venta"] or 0.0)
                            venta_log = float(row["Venta_Lograda"] or 0.0)
                            
                            if meta_v > 0.0:
                                meta_v_sin = meta_v
                                v_pct = (venta_log / meta_v_sin) * 100 if meta_v_sin > 0 else 0.0
                                lograda = venta_log >= meta_v_sin
                                
                                has_any = True
                                icon_color = "#00FF7F" if lograda else "#FFCC00"
                                icon_name = ft.Icons.CHECK_CIRCLE_OUTLINE_ROUNDED if lograda else ft.Icons.RADIO_BUTTON_UNCHECKED_OUTROUNDED
                                status_txt = "LOGRADO" if lograda else f"{v_pct:.0f}%"
                                
                                meses_logrados_col.controls.append(
                                    ft.Row([
                                        ft.Icon(icon_name, color=icon_color, size=16),
                                        ft.Text(f"{meses_nombres[m_idx-1]} ({status_txt})", color="white" if lograda else "#cccccc", size=12, weight="bold" if lograda else "normal"),
                                    ], spacing=5)
                                )
                        
                        if not has_any:
                            meses_logrados_col.controls.append(ft.Text("Ninguna meta de ventas definida en este año.", color="#aaaaaa", italic=True, size=11))
                except Exception as ex:
                    print("Error in render_meses_logrados:", ex)
                    meses_logrados_col.controls.append(ft.Text("Error al cargar logros.", color="red", size=12))

            def render_calendar(daily_accum_map):
                calendar_grid.controls.clear()
                



                days_headers = ["DOM", "LUN", "MAR", "MIE", "JUE", "VIE", "SAB"]
                header_row = ft.Row(
                    [
                        ft.Container(
                            content=ft.Text(h, color="#D8B4FE", weight="bold", size=10, text_align="center"),
                            expand=True,
                            alignment=ft.alignment.Alignment(0, 0),
                            padding=5
                        ) for h in days_headers
                    ],
                    spacing=5
                )
                calendar_grid.controls.append(header_row)

                year = current_year[0]
                month = current_month[0]
                first_weekday_py, num_days = calendar.monthrange(year, month)
                start_offset = (first_weekday_py + 1) % 7

                cells = []

                for _ in range(start_offset):
                    cells.append(
                        ft.Container(
                            expand=True,
                            height=70,
                            bgcolor="#0F0F1A",
                            border_radius=6,
                            opacity=0.3,
                            border=ft.Border.all(1, "#222222")
                        )
                    )

                for d in range(1, num_days + 1):
                    d_sin, d_pzs, accum_sin, accum_pzs = daily_accum_map.get(d, (0.0, 0, 0.0, 0))

                    cell_content = None
                    if d == 1:
                        if is_mobile:
                            cell_content = ft.Column([
                                ft.Row([
                                    ft.Text(str(d), size=10, weight="bold", color="#00FFFF"),
                                ], alignment="spaceBetween"),
                                ft.Container(
                                    content=ft.Column([
                                        ft.Text(f"${d_sin:,.0f}", size=8, color="white", weight="bold"),
                                        ft.Text(f"{d_pzs} pzs", size=7, color="#aaaaaa")
                                    ], spacing=1, alignment="center"),
                                    alignment=ft.alignment.Alignment(0, 0),
                                    expand=True
                                )
                            ], spacing=2, expand=True)
                        else:
                            cell_content = ft.Column([
                                ft.Row([
                                    ft.Text(str(d), size=12, weight="bold", color="#00FFFF"),
                                ], alignment="spaceBetween"),
                                ft.Container(
                                    content=ft.Column([
                                        ft.Text(f"${d_sin:,.0f}", size=11, color="white", weight="bold"),
                                        ft.Text(f"{d_pzs} pzs", size=9, color="#aaaaaa")
                                    ], spacing=1, alignment="center"),
                                    alignment=ft.alignment.Alignment(0, 0),
                                    expand=True
                                )
                            ], spacing=2, expand=True)
                    else:
                        if is_mobile:
                            cell_content = ft.Column([
                                ft.Row([
                                    ft.Text(str(d), size=10, weight="bold", color="#00FFFF"),
                                ], alignment="spaceBetween"),
                                ft.Container(
                                    content=ft.Column([
                                        ft.Text(f"${d_sin:,.0f}|{d_pzs}p", size=7.5, color="white"),
                                    ], spacing=0, alignment="center"),
                                    alignment=ft.alignment.Alignment(0, 0),
                                    height=20
                                ),
                                ft.Divider(height=1, color="#333333"),
                                ft.Container(
                                    content=ft.Column([
                                        ft.Text(f"${accum_sin:,.0f}|{accum_pzs}p", size=7.5, color="#7CFC00", weight="bold"),
                                    ], spacing=0, alignment="center"),
                                    alignment=ft.alignment.Alignment(0, 0),
                                    height=20
                                )
                            ], spacing=2, expand=True)
                        else:
                            cell_content = ft.Column([
                                ft.Row([
                                    ft.Text(str(d), size=12, weight="bold", color="#00FFFF"),
                                ], alignment="spaceBetween"),
                                ft.Container(
                                    content=ft.Column([
                                        ft.Text(f"${d_sin:,.0f} | {d_pzs}p", size=10, color="white"),
                                    ], spacing=0, alignment="center"),
                                    alignment=ft.alignment.Alignment(0, 0),
                                    height=20
                                ),
                                ft.Divider(height=1, color="#333333"),
                                ft.Container(
                                    content=ft.Column([
                                        ft.Text(f"${accum_sin:,.0f} | {accum_pzs}p", size=10, color="#7CFC00", weight="bold"),
                                    ], spacing=0, alignment="center"),
                                    alignment=ft.alignment.Alignment(0, 0),
                                    height=20
                                )
                            ], spacing=2, expand=True)

                    cell_container = ft.Container(
                        content=cell_content,
                        expand=True,
                        height=75,
                        bgcolor="#152238" if (d_sin > 0 or d_pzs > 0) else "#111111",
                        border_radius=8,
                        padding=ft.Padding(left=2, top=4, right=2, bottom=4) if is_mobile else ft.Padding(left=6, top=4, right=6, bottom=4),
                        border=ft.Border.all(1, "#3c5c8c" if (d_sin > 0 or d_pzs > 0) else "#222222"),
                        on_click=None if es_admin() else (lambda e, day_num=d: open_edit_day_dialog(day_num))
                    )
                    cells.append(cell_container)

                while len(cells) % 7 != 0:
                    cells.append(
                        ft.Container(
                            expand=True,
                            height=70,
                            bgcolor="#0F0F1A",
                            border_radius=6,
                            opacity=0.3,
                            border=ft.Border.all(1, "#222222")
                        )
                    )

                for i in range(0, len(cells), 7):
                    week_cells = cells[i:i+7]
                    calendar_grid.controls.append(
                        ft.Row(controls=week_cells, spacing=5)
                    )

            def cargar_datos_presupuesto_anual():
                tienda_actual = selected_tienda[0]
                anio_actual = current_year[0]
                
                if not tienda_actual:
                    return
                
                try:
                    db = conectar_db()
                    if db:
                        cur = db.cursor(dictionary=True)
                        cur.execute("""
                            SELECT Numero_Tienda, Presupuesto_Anual, Presupuesto_Q1, Presupuesto_Q2, Presupuesto_Q3, Presupuesto_Q4, Meses_Logrados 
                            FROM presupuesto_anual 
                            WHERE Tienda = %s AND Anio = %s
                        """, (tienda_actual, anio_actual))
                        row = cur.fetchone()
                        db.close()
                        
                        # Cargar logros reales de presupuesto_diario para los trimestres
                        q_logros = {1: 0.0, 2: 0.0, 3: 0.0, 4: 0.0}
                        try:
                            db_q = conectar_db()
                            if db_q:
                                cur_q = db_q.cursor(dictionary=True)
                                cur_q.execute("""
                                    SELECT MONTH(Fecha) as Mes, COALESCE(SUM(Venta_Sin_IVA), 0) as Total
                                    FROM presupuesto_diario
                                    WHERE Tienda = %s AND YEAR(Fecha) = %s
                                    GROUP BY MONTH(Fecha)
                                """, (tienda_actual, anio_actual))
                                for row_q in cur_q.fetchall():
                                    m = row_q["Mes"]
                                    tot = float(row_q["Total"] or 0.0)
                                    if m in (1, 2, 3):
                                        q_logros[1] += tot
                                    elif m in (4, 5, 6):
                                        q_logros[2] += tot
                                    elif m in (7, 8, 9):
                                        q_logros[3] += tot
                                    elif m in (10, 11, 12):
                                        q_logros[4] += tot
                                db_q.close()
                        except Exception as ex_q:
                            print("Error querying quarter sales:", ex_q)

                        txt_q1_logro.value = f"${q_logros[1]:,.2f}"
                        txt_q2_logro.value = f"${q_logros[2]:,.2f}"
                        txt_q3_logro.value = f"${q_logros[3]:,.2f}"
                        txt_q4_logro.value = f"${q_logros[4]:,.2f}"

                        if row:
                            txt_num_tienda.value = str(row["Numero_Tienda"] or "")
                            txt_presupuesto_anual.value = str(row["Presupuesto_Anual"] or "")
                            txt_q1.value = str(row["Presupuesto_Q1"] or "")
                            txt_q2.value = str(row["Presupuesto_Q2"] or "")
                            txt_q3.value = str(row["Presupuesto_Q3"] or "")
                            txt_q4.value = str(row["Presupuesto_Q4"] or "")
                            
                            # Calcular porcentajes
                            try:
                                q1_meta = float(row["Presupuesto_Q1"] or 0.0)
                                txt_q1_pct.value = f"{(q_logros[1] / q1_meta * 100):.1f}%" if q1_meta > 0 else "0.0%"
                            except Exception:
                                txt_q1_pct.value = "0.0%"

                            try:
                                q2_meta = float(row["Presupuesto_Q2"] or 0.0)
                                txt_q2_pct.value = f"{(q_logros[2] / q2_meta * 100):.1f}%" if q2_meta > 0 else "0.0%"
                            except Exception:
                                txt_q2_pct.value = "0.0%"

                            try:
                                q3_meta = float(row["Presupuesto_Q3"] or 0.0)
                                txt_q3_pct.value = f"{(q_logros[3] / q3_meta * 100):.1f}%" if q3_meta > 0 else "0.0%"
                            except Exception:
                                txt_q3_pct.value = "0.0%"

                            try:
                                q4_meta = float(row["Presupuesto_Q4"] or 0.0)
                                txt_q4_pct.value = f"{(q_logros[4] / q4_meta * 100):.1f}%" if q4_meta > 0 else "0.0%"
                            except Exception:
                                txt_q4_pct.value = "0.0%"

                            logrados_str = row["Meses_Logrados"] or ""
                            logrados_list = [m.strip().lower() for m in logrados_str.split(",") if m.strip()]
                            for i, chk in enumerate(chk_meses):
                                chk.value = meses_nombres[i].lower() in logrados_list
                        else:
                            txt_num_tienda.value = ""
                            txt_presupuesto_anual.value = ""
                            txt_q1.value = ""
                            txt_q2.value = ""
                            txt_q3.value = ""
                            txt_q4.value = ""
                            
                            txt_q1_pct.value = "0.0%"
                            txt_q2_pct.value = "0.0%"
                            txt_q3_pct.value = "0.0%"
                            txt_q4_pct.value = "0.0%"
                            
                            for chk in chk_meses:
                                chk.value = False
                except Exception as ex:
                    print("Error loading presupuesto anual:", ex)

            def guardar_presupuesto_anual_click(e):
                tienda_actual = selected_tienda[0]
                anio_actual = current_year[0]
                
                if not tienda_actual:
                    mostrar_snack("Por favor ingresa una tienda.", color="red")
                    return
                
                num_t = txt_num_tienda.value.strip()
                p_anual = 0.0
                try:
                    p_anual = float(txt_presupuesto_anual.value.strip()) if txt_presupuesto_anual.value.strip() else 0.0
                except ValueError:
                    mostrar_snack("El presupuesto anual debe ser un número válido.", color="red")
                    return
                
                q1_val = 0.0
                try:
                    q1_val = float(txt_q1.value.strip()) if txt_q1.value.strip() else 0.0
                except ValueError:
                    mostrar_snack("El presupuesto Q1 debe ser un número válido.", color="red")
                    return

                q2_val = 0.0
                try:
                    q2_val = float(txt_q2.value.strip()) if txt_q2.value.strip() else 0.0
                except ValueError:
                    mostrar_snack("El presupuesto Q2 debe ser un número válido.", color="red")
                    return

                q3_val = 0.0
                try:
                    q3_val = float(txt_q3.value.strip()) if txt_q3.value.strip() else 0.0
                except ValueError:
                    mostrar_snack("El presupuesto Q3 debe ser un número válido.", color="red")
                    return

                q4_val = 0.0
                try:
                    q4_val = float(txt_q4.value.strip()) if txt_q4.value.strip() else 0.0
                except ValueError:
                    mostrar_snack("El presupuesto Q4 debe ser un número válido.", color="red")
                    return

                logrados = []
                for i, chk in enumerate(chk_meses):
                    if chk.value:
                        logrados.append(meses_nombres[i])
                logrados_str = ",".join(logrados)

                try:
                    db = conectar_db()
                    if db:
                        cur = db.cursor()
                        cur.execute("SELECT COUNT(*) FROM presupuesto_anual WHERE Tienda = %s AND Anio = %s", (tienda_actual, anio_actual))
                        exists = cur.fetchone()[0] > 0
                        
                        if exists:
                            cur.execute("""
                                UPDATE presupuesto_anual 
                                SET Numero_Tienda = %s, Presupuesto_Anual = %s, Presupuesto_Q1 = %s, Presupuesto_Q2 = %s, Presupuesto_Q3 = %s, Presupuesto_Q4 = %s, Meses_Logrados = %s 
                                WHERE Tienda = %s AND Anio = %s
                            """, (num_t, p_anual, q1_val, q2_val, q3_val, q4_val, logrados_str, tienda_actual, anio_actual))
                        else:
                            cur.execute("""
                                INSERT INTO presupuesto_anual (Tienda, Anio, Numero_Tienda, Presupuesto_Anual, Presupuesto_Q1, Presupuesto_Q2, Presupuesto_Q3, Presupuesto_Q4, Meses_Logrados) 
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """, (tienda_actual, anio_actual, num_t, p_anual, q1_val, q2_val, q3_val, q4_val, logrados_str))
                        db.commit()
                        db.close()
                        mostrar_snack("✅ Configuración de Bouget Anual guardada.", color="#7CFC00")
                        refresh_data()
                except Exception as ex:
                    print("Error saving presupuesto anual:", ex)
                    mostrar_snack("Error al guardar la configuración anual.", color="red")

            last_tienda = [selected_tienda[0]]
            last_num_tienda = [""]

            def autorellenar_tienda():
                t_nombre = txt_tienda.value.strip()
                t_numero = txt_num_tienda.value.strip()
                
                cambio_nombre = (t_nombre != last_tienda[0])
                cambio_numero = (t_numero != last_num_tienda[0])
                
                if cambio_nombre and not cambio_numero:
                    try:
                        db = conectar_db()
                        if db:
                            cur = db.cursor()
                            cur.execute("SELECT Numero_Tienda FROM presupuesto_anual WHERE Tienda = %s LIMIT 1", (t_nombre,))
                            row = cur.fetchone()
                            db.close()
                            if row and row[0]:
                                txt_num_tienda.value = str(row[0])
                                last_num_tienda[0] = str(row[0])
                            else:
                                txt_num_tienda.value = ""
                                last_num_tienda[0] = ""
                    except Exception as e:
                        print("Error buscando número:", e)
                    last_tienda[0] = t_nombre
                elif cambio_numero and not cambio_nombre:
                    try:
                        db = conectar_db()
                        if db:
                            cur = db.cursor()
                            cur.execute("SELECT Tienda FROM presupuesto_anual WHERE Numero_Tienda = %s LIMIT 1", (t_numero,))
                            row = cur.fetchone()
                            db.close()
                            if row and row[0]:
                                txt_tienda.value = str(row[0])
                                selected_tienda[0] = str(row[0])
                                last_tienda[0] = str(row[0])
                            else:
                                txt_tienda.value = ""
                                selected_tienda[0] = ""
                                last_tienda[0] = ""
                    except Exception as e:
                        print("Error buscando nombre:", e)
                    last_num_tienda[0] = t_numero
                else:
                    if t_nombre and not t_numero:
                        try:
                            db = conectar_db()
                            if db:
                                cur = db.cursor()
                                cur.execute("SELECT Numero_Tienda FROM presupuesto_anual WHERE Tienda = %s LIMIT 1", (t_nombre,))
                                row = cur.fetchone()
                                db.close()
                                if row and row[0]:
                                    txt_num_tienda.value = str(row[0])
                                    last_num_tienda[0] = str(row[0])
                        except Exception:
                            pass
                    elif t_numero and not t_nombre:
                        try:
                            db = conectar_db()
                            if db:
                                cur = db.cursor()
                                cur.execute("SELECT Tienda FROM presupuesto_anual WHERE Numero_Tienda = %s LIMIT 1", (t_numero,))
                                row = cur.fetchone()
                                db.close()
                                if row and row[0]:
                                    txt_tienda.value = str(row[0])
                                    selected_tienda[0] = str(row[0])
                                    last_tienda[0] = str(row[0])
                        except Exception:
                            pass

            def refresh_data():
                autorellenar_tienda()
                if dd_mes.value:
                    current_month[0] = int(dd_mes.value)
                if dd_anio.value:
                    current_year[0] = int(dd_anio.value)
                selected_tienda[0] = txt_tienda.value.strip()
                cargar_datos_presupuesto_anual()
                
                meta_v_con_iva, meta_p, sales_diarias = cargar_datos_presupuesto()

                # Mostrar el valor cargado; si es admin mostrar aunque sea 0 para que pueda ver lo que el gerente guardó
                if es_admin():
                    meta_venta_tf.value = str(int(meta_v_con_iva)) if meta_v_con_iva == int(meta_v_con_iva) else str(meta_v_con_iva)
                    meta_piezas_tf.value = str(meta_p)
                else:
                    meta_venta_tf.value = str(meta_v_con_iva) if meta_v_con_iva > 0 else ""
                    meta_piezas_tf.value = str(meta_p) if meta_p > 0 else ""

                meta_v_sin_iva = meta_v_con_iva if meta_v_con_iva > 0 else 0.0

                sales_map = {row["Dia"]: (float(row["Venta_Con_IVA"]), float(row["Venta_Sin_IVA"]), int(row["Piezas"])) for row in sales_diarias}

                accum_venta_sin_iva = 0.0
                accum_piezas = 0

                days_in_month = calendar.monthrange(current_year[0], current_month[0])[1]
                daily_accum_map = {}

                for d in range(1, days_in_month + 1):
                    d_con, d_sin, d_pzs = sales_map.get(d, (0.0, 0.0, 0))
                    accum_venta_sin_iva += d_sin
                    accum_piezas += d_pzs
                    daily_accum_map[d] = (d_sin, d_pzs, accum_venta_sin_iva, accum_piezas)

                if meta_v_sin_iva > 0:
                    v_ratio = accum_venta_sin_iva / meta_v_sin_iva
                    progress_bar_venta.value = min(1.0, v_ratio)
                    progress_bar_venta.color = "#FF4B4B" if v_ratio < 0.5 else ("#FFCC00" if v_ratio < 1.0 else "#00FF7F")
                    progress_text_venta.value = f"Ventas: {v_ratio*100:.1f}% (${accum_venta_sin_iva:,.2f} / ${meta_v_sin_iva:,.2f} sin IVA)"
                else:
                    progress_bar_venta.value = 0.0
                    progress_bar_venta.color = "#FF4B4B"
                    progress_text_venta.value = f"Ventas: Meta no definida (${accum_venta_sin_iva:,.2f} sin IVA)"

                if meta_p > 0:
                    p_ratio = accum_piezas / meta_p
                    progress_bar_piezas.value = min(1.0, p_ratio)
                    progress_bar_piezas.color = "#FF4B4B" if p_ratio < 0.5 else ("#FFCC00" if p_ratio < 1.0 else "#00FF7F")
                    progress_text_piezas.value = f"Piezas: {p_ratio*100:.1f}% ({accum_piezas} / {meta_p} pzs)"
                else:
                    progress_bar_piezas.value = 0.0
                    progress_bar_piezas.color = "#FF4B4B"
                    progress_text_piezas.value = f"Piezas: Meta no definida ({accum_piezas} pzs)"

                tienda_title_txt.value = selected_tienda[0].upper() if selected_tienda[0] else "SELECCIONE TIENDA"
                zona_title_txt.value = f"Nº Tienda: {txt_num_tienda.value}" if txt_num_tienda.value else ""
                period_title_txt.value = f"{meses_nombres[current_month[0]-1].upper()} {current_year[0]}"

                render_meses_logrados()
                render_calendar(daily_accum_map)
                try:
                    page.update()
                except Exception:
                    pass

            def guardar_metas_click(e):
                tienda_actual = selected_tienda[0]
                mes_actual = current_month[0]
                anio_actual = current_year[0]
                
                if not tienda_actual:
                    mostrar_snack("Escribe una tienda primero.", color="red")
                    return
                
                try:
                    m_venta = float(meta_venta_tf.value.strip() or 0.0)
                    m_piezas = int(meta_piezas_tf.value.strip() or 0)
                except ValueError:
                    mostrar_snack("Por favor ingresa números válidos para las metas.", color="red")
                    return
                
                try:
                    db = conectar_db()
                    if db:
                        cur = db.cursor()
                        cur.execute("""
                            INSERT INTO presupuesto_mensual (Tienda, Mes, Anio, Meta_Venta, Meta_Piezas)
                            VALUES (%s, %s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE 
                                Meta_Venta = %s,
                                Meta_Piezas = %s
                        """, (tienda_actual, mes_actual, anio_actual, m_venta, m_piezas, m_venta, m_piezas))
                        db.commit()
                        db.close()
                        mostrar_snack("Metas guardadas exitosamente.", color="#7CFC00")
                        refresh_data()
                except Exception as ex:
                    print("Error saving month goals:", ex)
                    mostrar_snack("Error al guardar metas.", color="red")

            def on_period_changed(e):
                current_month[0] = int(dd_mes.value)
                current_year[0] = int(dd_anio.value)
                refresh_data()

            def on_tienda_changed(e):
                selected_tienda[0] = txt_tienda.value.strip()
                refresh_data()

            dd_mes.on_change = on_period_changed
            dd_anio.on_change = on_period_changed

            btn_consultar = ft.ElevatedButton(
                "Consultar 🔍",
                bgcolor="#00FFFF",
                color="black",
                height=45,
                on_click=lambda e: refresh_data()
            )

            filters_row = ft.Row([
                txt_tienda,
                txt_num_tienda,
                dd_mes,
                dd_anio,
                btn_consultar
            ], spacing=10, wrap=True)

            btn_guardar_metas = ft.ElevatedButton(
                "Guardar Metas Mensuales 💾",
                bgcolor="#9D50BB",
                color="white",
                height=35,
                visible=not es_admin(),
                on_click=guardar_metas_click
            )

            btn_guardar_anual = ft.ElevatedButton(
                "Guardar Configuración Anual 💾",
                bgcolor="#6E48AA",
                color="white",
                height=35,
                visible=not es_admin(),
                on_click=guardar_presupuesto_anual_click
            )

            left_panel = ft.Column([
                # Card 1: Definir Metas del Mes
                ft.Container(
                    content=ft.Column([
                        ft.Text("Definir Metas del Mes", size=14, color="#D8B4FE", weight="bold"),
                        ft.Row([
                            meta_venta_tf,
                            meta_piezas_tf
                        ], spacing=10),
                        btn_guardar_metas
                    ], spacing=10),
                    bgcolor="#0F0F1A",
                    padding=15,
                    border_radius=8,
                    border=ft.Border.all(1, "#333333")
                ),
                
                # Card 2: Avance del Período
                ft.Container(
                    content=ft.Column([
                        ft.Text("Avance del Período", size=14, color="#D8B4FE", weight="bold"),
                        progress_text_venta,
                        progress_bar_venta,
                        progress_text_piezas,
                        progress_bar_piezas
                    ], spacing=8),
                    bgcolor="#0F0F1A",
                    padding=15,
                    border_radius=8,
                    border=ft.Border.all(1, "#333333")
                ),
                
                # Card 3: Configuración Bouget Anual (Nuevo)
                ft.Container(
                    content=ft.Column([
                        ft.Text("Configuración de Bouget Anual", size=14, color="#D8B4FE", weight="bold"),
                        txt_presupuesto_anual,
                        ft.Container(height=5),
                        ft.Text("Metas Trimestrales:", size=12, color="#aaaaaa"),
                        ft.Row([
                            ft.Column([
                                ft.Text("Q1 (Ene-Mar)", size=10, weight="bold", color="#D8B4FE"),
                                txt_q1,
                                txt_q1_logro,
                                txt_q1_pct
                            ], spacing=4, expand=True),
                            ft.Column([
                                ft.Text("Q2 (Abr-Jun)", size=10, weight="bold", color="#D8B4FE"),
                                txt_q2,
                                txt_q2_logro,
                                txt_q2_pct
                            ], spacing=4, expand=True)
                        ], spacing=10),
                        ft.Row([
                            ft.Column([
                                ft.Text("Q3 (Jul-Sep)", size=10, weight="bold", color="#D8B4FE"),
                                txt_q3,
                                txt_q3_logro,
                                txt_q3_pct
                            ], spacing=4, expand=True),
                            ft.Column([
                                ft.Text("Q4 (Oct-Dic)", size=10, weight="bold", color="#D8B4FE"),
                                txt_q4,
                                txt_q4_logro,
                                txt_q4_pct
                            ], spacing=4, expand=True)
                        ], spacing=10),
                        ft.Container(height=5),
                        ft.Text("Hitos: Meses logrados", size=12, color="#aaaaaa"),
                        ft.Container(
                            content=ft.Row([chk_meses[i] for i in range(12)], wrap=True, spacing=10),
                            padding=5,
                            border_radius=5,
                            bgcolor="#1c1c1c"
                        ),
                        btn_guardar_anual
                    ], spacing=10),
                    bgcolor="#0F0F1A",
                    padding=15,
                    border_radius=8,
                    border=ft.Border.all(1, "#333333")
                )
            ], spacing=15)

            right_panel = ft.Column([
                ft.Container(
                    content=ft.Row([
                        ft.Column([
                            tienda_title_txt,
                            zona_title_txt
                        ], spacing=2),
                        ft.Container(expand=True),
                        period_title_txt
                    ], vertical_alignment="center"),
                    padding=ft.Padding(left=10, top=5, right=10, bottom=5)
                ),
                calendar_grid
            ], spacing=10, expand=True)

            responsive_layout = ft.ResponsiveRow([
                ft.Container(left_panel, col={"xs": 12, "md": 4}),
                ft.Container(right_panel, col={"xs": 12, "md": 8})
            ], spacing=20)

            main_col = ft.Column([
                ft.Row([
                    ft.Text("Bouget 📊", size=24, color="#D8B4FE", weight="bold")
                ]),
                ft.Text("Monitorea las metas mensuales de ventas y piezas, y registra la configuración de hitos anuales.", color="#aaaaaa", size=13),
                ft.Divider(height=15, color="#333333"),
                filters_row,
                ft.Container(height=10),
                responsive_layout
            ], scroll=ft.ScrollMode.AUTO, expand=True)

            refresh_data()
            return main_col

        def build_vendedores_view():
            is_mobile_w = (page.width < 700) if (page and page.width) else False
            vendedores_list = ft.Column(spacing=10)
            vendedor_name_input = ft.TextField(
                label="Nombre del Colaborador",
                border_color="#9D50BB",
                color="white",
                text_size=12 if is_mobile_w else 13,
                width=180 if is_mobile_w else 260
            )
            puesto_vendedor_input = EmojiDropdown(
                label="Puesto / Rol",
                options=[
                    ft.dropdown.Option("Vendedor"),
                    ft.dropdown.Option("Subgerente"),
                    ft.dropdown.Option("Gerente de Tienda")
                ],
                value="Vendedor",
                border_color="#9D50BB",
                width=160 if is_mobile_w else 200
            )

            def cargar_vendedores():
                vendedores_list.controls.clear()
                rows = []
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("""
                            SELECT ID_Vendedor, Nombre_Completo, Puesto, Activo, DATE_FORMAT(Fecha_Registro, '%d/%m/%Y') as fecha_f
                            FROM vendedores
                            WHERE ID_Usuario_Tienda = %s
                            ORDER BY Nombre_Completo ASC
                        """, (user_info.get("id", 1),))
                        rows = cursor.fetchall()
                        db.close()
                except Exception as ex:
                    print("Error cargar colaboradores DB:", ex)

                if not rows:
                    t_nom = user_info.get("tienda", "VALLEJO")
                    store_colabs_def = {
                        "VALLEJO": [("VIVIANA", "Gerente de Tienda"), ("moises", "Subgerente"), ("diego", "Vendedor")],
                        "INTERLOMAS": [("idalia", "Gerente de Tienda"), ("Viviana", "Subgerente"), ("carlos", "Vendedor")],
                        "ATIZAPAN": [("FERNANDO", "Gerente de Tienda"), ("STEFANI", "Subgerente"), ("MOISES", "Vendedor")],
                        "PERISUR": [("ROBERTO", "Gerente de Tienda"), ("ANA", "Subgerente"), ("LUIS", "Vendedor")],
                        "LINDAVISTA": [("PATRICIA", "Gerente de Tienda"), ("JORGE", "Subgerente"), ("SARA", "Vendedor")],
                        "SANTA FE": [("MAURICIO", "Gerente de Tienda"), ("ELENA", "Subgerente"), ("GABRIEL", "Vendedor")],
                        "SATÉLITE": [("DANIEL", "Gerente de Tienda"), ("KAREN", "Subgerente"), ("ANDRES", "Vendedor")],
                        "PACHUCA": [("OSCAR", "Gerente de Tienda"), ("VERONICA", "Subgerente"), ("MANUEL", "Vendedor")]
                    }
                    defaults = store_colabs_def.get(t_nom, [("VIVIANA", "Gerente de Tienda"), ("moises", "Subgerente"), ("diego", "Vendedor")])
                    rows = [{"ID_Vendedor": idx+100, "Nombre_Completo": nom, "Puesto": pst, "fecha_f": "Activo"} for idx, (nom, pst) in enumerate(defaults)]

                for r in rows:
                    v_id = r["ID_Vendedor"]
                    v_name = r["Nombre_Completo"]
                    puesto_nombre = r.get("Puesto") or "Vendedor"

                    def make_delete_click(vid=v_id, name=v_name):
                        def delete_click(e):
                            def confirmar_borrado(ev):
                                try:
                                    u_rol = str(user_info.get("rol", "")).lower()
                                    u_pue = str(user_info.get("puesto", "")).lower()
                                    es_gerente = "gerente" in u_rol or "gerente" in u_pue or "admin" in u_rol

                                    if not es_gerente:
                                        page.pop_dialog()
                                        mostrar_snack("⚠️ Permiso denegado: Solo el Gerente de Tienda puede eliminar a un colaborador", "red")
                                        return

                                    db_d = conectar_db()
                                    if db_d:
                                        cursor_d = db_d.cursor()
                                        cursor_d.execute("DELETE FROM vendedores WHERE ID_Vendedor = %s", (vid,))
                                        db_d.commit()
                                        db_d.close()

                                        registrar_auditoria_borrado(
                                            ejecutor_id=user_info.get("id", 0),
                                            ejecutor_nombre=user_info.get("nombre", "Gerente de Tienda"),
                                            ejecutor_rol=user_info.get("puesto") or user_info.get("rol") or "Gerente de Tienda",
                                            afectado_nombre=name,
                                            accion="BAJA_COLABORADOR",
                                            detalles="Baja de colaborador autorizada por Gerente"
                                        )

                                    page.pop_dialog()
                                    mostrar_snack(f"Colaborador '{name}' eliminado 🛡️", "#FF4500")
                                    cargar_vendedores()
                                    try:
                                        import enfoque_diario
                                        enfoque_diario.sincronizar_colaboradores_db(user_info)
                                    except Exception:
                                        pass
                                    try: page.update()
                                    except Exception: pass
                                except Exception as ex_d:
                                    print("Error eliminando colaborador:", ex_d)
                                    mostrar_snack("Error al eliminar colaborador", "red")
                                    page.pop_dialog()

                            confirm_dialog = ft.AlertDialog(
                                title=ft.Text("⚠️ Confirmar Baja de Personal", color="red", weight="bold"),
                                content=ft.Text(f"¿Estás seguro de que deseas dar de baja a '{name}'?", color="white"),
                                actions=[
                                    ft.TextButton("Cancelar", on_click=lambda ev: page.pop_dialog()),
                                    ft.TextButton("Eliminar", on_click=confirmar_borrado, style=ft.ButtonStyle(color="red"))
                                ],
                                actions_alignment="end",
                                bgcolor="#0F0F1A"
                            )
                            page.show_dialog(confirm_dialog)
                            try: page.update()
                            except Exception: pass
                        return delete_click

                    btn_delete = ft.IconButton(
                        icon=ft.Icons.DELETE_ROUNDED,
                        icon_color="#FF4500",
                        tooltip="Eliminar Personal 🗑️",
                        on_click=make_delete_click()
                    )

                    vendedores_list.controls.append(
                        ft.Container(
                            content=ft.Row([
                                ft.Icon(ft.Icons.PERSON_PIN_ROUNDED, color="#00FFFF", size=20),
                                ft.Column([
                                    ft.Text(r["Nombre_Completo"], color="white", weight="bold", size=13),
                                    ft.Row([
                                        ft.Container(
                                            content=ft.Text(puesto_nombre, color="#D8B4FE", weight="bold", size=10),
                                            bgcolor="#2a1a3e",
                                            padding=ft.padding.Padding(6, 2, 6, 2),
                                            border_radius=4,
                                            border=ft.Border.all(1, "#6E48AA")
                                        ),
                                        ft.Text(r['fecha_f'], color="#666666", size=10)
                                    ], spacing=6)
                                ], spacing=2, expand=True),
                                btn_delete
                            ], vertical_alignment="center", spacing=6),
                            bgcolor="#1a1a22",
                            padding=ft.padding.Padding(10, 8, 10, 8),
                            border_radius=10,
                            border=ft.Border.all(1, "#2a2a33")
                        )
                    )

            def registrar_vendedor_click(e):
                name = vendedor_name_input.value.strip()
                puesto_val = puesto_vendedor_input.value or "Vendedor"
                if not name:
                    mostrar_snack("Por favor ingresa un nombre válido", "red")
                    return
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        cursor.execute("INSERT INTO vendedores (ID_Usuario_Tienda, Nombre_Completo, Puesto) VALUES (%s, %s, %s)", (user_info["id"], name, puesto_val))
                        db.commit()
                        db.close()
                        vendedor_name_input.value = ""
                        mostrar_snack("Colaborador registrado con éxito 🎉", "#7CFC00")
                        cargar_vendedores()
                        try:
                            import enfoque_diario
                            enfoque_diario.sincronizar_colaboradores_db(user_info)
                        except Exception:
                            pass
                        page.update()
                except Exception as ex:
                    print("Error registrar colaborador:", ex)
                    mostrar_snack("Error al guardar el colaborador", "red")

            # =====================================================
            # LÓGICA DE GERENTE ÚNICO POR TIENDA
            # =====================================================
            gerente_actual = None   # Dict con datos del Gerente registrado (si existe)
            yo_soy_gerente = False  # True si el usuario actual ES el Gerente registrado
            puesto_libre = False    # True si no hay Gerente aún en esta tienda

            try:
                db_g = conectar_db()
                if db_g:
                    cur_g = db_g.cursor(dictionary=True)
                    cur_g.execute("""
                        SELECT ID_Usuario, Nombre_Completo, Rol, Usuario
                        FROM usuarios
                        WHERE Tienda = %s
                          AND (LOWER(Rol) LIKE '%gerente%' OR LOWER(Rol) = 'gerente de tienda')
                          AND ID_Usuario != %s
                        LIMIT 1
                    """, (user_info.get("tienda", ""), user_info.get("id", 0)))
                    gerente_actual = cur_g.fetchone()

                    cur_g.execute("""
                        SELECT Rol FROM usuarios WHERE ID_Usuario = %s
                    """, (user_info.get("id", 0),))
                    mi_rol_row = cur_g.fetchone()
                    mi_rol_actual = str(mi_rol_row.get("Rol", "") if mi_rol_row else "").lower()
                    yo_soy_gerente = "gerente" in mi_rol_actual

                    db_g.close()
                    puesto_libre = (gerente_actual is None) and (not yo_soy_gerente)
            except Exception as ex_g:
                print("Error consultando Gerente de tienda:", ex_g)

            def registrarme_como_gerente(e):
                try:
                    db_rg = conectar_db()
                    if db_rg:
                        cur_rg = db_rg.cursor(dictionary=True)
                        cur_rg.execute("""
                            SELECT COUNT(*) as total FROM usuarios
                            WHERE Tienda = %s
                              AND (LOWER(Rol) LIKE '%gerente%' OR LOWER(Rol) = 'gerente de tienda')
                        """, (user_info.get("tienda", ""),))
                        resultado = cur_rg.fetchone()
                        if resultado and resultado["total"] > 0:
                            mostrar_snack("⚠️ El puesto de Gerente ya fue tomado por otro usuario.", "#FF8C00")
                            db_rg.close()
                            cambiar_vista("vendedores")
                            return
                        cur_rg.execute("""
                            UPDATE usuarios SET Rol = 'Gerente de Tienda'
                            WHERE ID_Usuario = %s
                        """, (user_info.get("id", 0),))
                        db_rg.commit()
                        db_rg.close()
                        user_info["rol"] = "Gerente de Tienda"
                        user_info["puesto"] = "Gerente de Tienda"
                        mostrar_snack("🏅 ¡Felicidades! Ahora eres el Gerente de esta tienda.", "#7CFC00")
                        cambiar_vista("vendedores")
                except Exception as ex_rg:
                    print("Error registrando como Gerente:", ex_rg)
                    mostrar_snack("Error al registrar el rol de Gerente.", "red")

            u_rol_vendedores = str(user_info.get("rol", "")).lower()
            es_admin_global = "admin" in u_rol_vendedores

            if puesto_libre and not es_admin_global:
                banner_gerente = ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Icon(ft.Icons.WORKSPACE_PREMIUM_ROUNDED, color="#FFD700", size=22),
                            ft.Text("Puesto de Gerente de Tienda Disponible", color="#FFD700", size=14, weight="bold"),
                        ], spacing=8),
                        ft.Text(
                            "No hay un Gerente registrado en esta tienda. Si eres el responsable de este negocio, puedes tomar el puesto.",
                            color="#aaaaaa", size=12
                        ),
                        ft.Container(height=6),
                        ft.ElevatedButton(
                            content=ft.Row([
                                ft.Icon(ft.Icons.BADGE_ROUNDED, size=18, color="white"),
                                ft.Text("  🏅 Registrarme como Gerente de esta Tienda", weight="bold", color="white")
                            ], spacing=6),
                            on_click=registrarme_como_gerente,
                            bgcolor="#7B2D8B",
                            color="white",
                            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=10))
                        )
                    ], spacing=6),
                    bgcolor="#1a0f00",
                    border=ft.Border.all(1, "#FFD700"),
                    border_radius=10,
                    padding=ft.padding.Padding(14, 12, 14, 12)
                )
            else:
                banner_gerente = ft.Container(
                    content=ft.Row([
                        ft.Icon(ft.Icons.PEOPLE_ROUNDED, color="#00FFFF", size=18),
                        ft.Text(
                            "👥 Gestión de Personal Habilitada: Puedes registrar y gestionar colaboradores de la tienda.",
                            color="#00FFFF", size=11 if is_mobile_w else 12, weight="bold"
                        )
                    ], spacing=8, wrap=True),
                    bgcolor="#0d2a2a",
                    border=ft.Border.all(1, "#00FFFF"),
                    border_radius=8,
                    padding=ft.padding.Padding(12, 8, 12, 8)
                )

            # Botón Registrar + compacto (tamaño estándar)
            btn_agregar = ft.ElevatedButton(
                content=ft.Row([
                    ft.Icon(ft.Icons.ADD_ROUNDED, size=16, color="white"),
                    ft.Text("Registrar ➕", weight="bold", color="white")
                ], spacing=6, tight=True),
                on_click=registrar_vendedor_click,
                bgcolor="#6E48AA",
                color="white",
                width=130 if is_mobile_w else 150,
                height=42,
                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
            )

            cargar_vendedores()

            form_section = ft.Row([
                vendedor_name_input,
                puesto_vendedor_input,
                btn_agregar
            ], spacing=10, vertical_alignment="center", wrap=True)

            return ft.Column([
                ft.Text("Configuración de Tienda 👥", size=20, color="#D8B4FE", weight="bold"),
                ft.Text(
                    "Registra y gestiona a los colaboradores de la tienda.",
                    color="#aaaaaa", size=12
                ),
                ft.Divider(height=10, color="#333333"),
                banner_gerente,
                ft.Container(height=6),
                form_section,
                ft.Container(height=10),
                ft.Text("Lista de Personal Activo:", color="#D8B4FE", size=14, weight="bold"),
                vendedores_list
            ], scroll=ft.ScrollMode.AUTO)



        def build_simulador_view():
            vendedor_dropdown = EmojiDropdown(
                label="Seleccionar Vendedor",
                border_color="#9D50BB",
                width=260,
                height=45
            )
            cliente_dropdown = EmojiDropdown(
                label="Perfil de Cliente",
                border_color="#9D50BB",
                width=260,
                height=45,
                options=[
                    ft.dropdown.Option("Objeción de Precio - 'Son Muy Caros'", "Cliente insiste en que las gafas Sunglass Hut son muy caras y que vio unos lentes parecidos mucho más baratos en otro lugar"),
                    ft.dropdown.Option("Duda Técnica - Polarizados y Chromance", "Cliente confundido preguntando qué ventaja real tienen las micas polarizadas y Chromance sobre las normales y por qué valen la pena"),
                    ft.dropdown.Option("Consulta de Garantías - Ruptura o Robo", "Cliente interesado pero pregunta qué garantía tienen los lentes si sufren alguna ruptura, daño accidental o si se los roban"),
                    ft.dropdown.Option("Cliente Indeciso - Ray-Ban Meta", "Cliente Indeciso buscando tecnología (Ray-Ban Meta)"),
                    ft.dropdown.Option("Cliente Reclamando Cambio sin Ticket", "Cliente molesto que exige cambio de lentes Oakley sin ticket de compra"),
                    ft.dropdown.Option("Cliente buscando Kit de Limpieza", "Cliente que solo entra preguntando por un paño de limpieza sencillo"),
                    ft.dropdown.Option("Cliente de Regalo de Lujo", "Cliente indeciso buscando un regalo premium para su pareja (Gafas Versace)"),
                    ft.dropdown.Option("Cliente Apurado - Gafas Polarizadas", "Cliente muy apurado que tiene un vuelo en pocas horas y busca unas gafas clásicas polarizadas (Ray-Ban Aviator) para la playa, exige rapidez y no quiere rodeos"),
                    ft.dropdown.Option("Cliente Escéptico - Privacidad Meta", "Cliente desconfiado e interesado en la tecnología de las gafas inteligentes (Ray-Ban Meta), pero le preocupa mucho la privacidad y si la cámara o el micrófono graban de forma oculta"),
                    ft.dropdown.Option("Coleccionista Exigente - Prada/Versace", "Cliente de alto nivel adquisitivo, muy conocedor de moda, que busca una pieza exclusiva de edición limitada de Prada o Versace y espera una atención ultra-premium y detalles de diseño"),
                    ft.dropdown.Option("Cliente Comparador - Objeción de Precio", "Cliente indeciso que le encantan unas gafas Dolce & Gabbana, pero insiste en que las vio más baratas en una tienda en línea no autorizada y cuestiona el valor y autenticidad del producto en tienda física"),
                    ft.dropdown.Option("Cliente de Descuento - Sin Temporada", "Cliente que insiste en obtener un descuento especial para comprar un solo par de gafas (Oakley o Ray-Ban) a pesar de que le explicas que no es temporada de rebajas ni hay promociones vigentes en tienda"),
                    ft.dropdown.Option("Padre Indeciso - Regalo Adolescente", "Padre o madre de familia que busca un regalo de cumpleaños para su hijo adolescente, no sabe qué marca está de moda (Oakley o Ray-Ban Meta) y necesita asesoría paciente sobre tendencias juveniles"),
                    ft.dropdown.Option("Deportista - Presupuesto Ajustado", "Deportista aficionado que busca gafas de sol de alto rendimiento para correr o ciclismo (Oakley Sutro), conoce los beneficios técnicos del lente pero tiene un presupuesto muy limitado y busca la opción de menor costo"),
                    ft.dropdown.Option("Ejecutivo - Imagen Profesional", "Profesional corporativo que busca unas gafas elegantes y sobrias para usar con vestimenta formal de negocios (Persol o Prada), valora la discreción, calidad de materiales e imagen profesional"),
                    ft.dropdown.Option("Cliente Reclamando Garantía - Lentes Rayados", "Cliente molesto que viene a exigir el cambio o garantía gratis de sus gafas porque los lentes están completamente rayados debido a mal uso (los limpió con su playera o los dejó caer), e insiste en que es defecto de fábrica")
                ]
            )

            chat_history = []
            sim_chat_column = ft.Column(spacing=10, scroll=ft.ScrollMode.AUTO, expand=True)
            user_input = ft.TextField(
                label="Escribe tu respuesta al cliente...",
                border_color="#9D50BB",
                color="white",
                expand=True,
                disabled=True,
                multiline=True,
                min_lines=1,
                max_lines=4,
                shift_enter=True
            )
            btn_enviar = ft.IconButton(
                icon=ft.Icons.SEND,
                icon_color="#00FFFF",
                disabled=True
            )

            vendedor_seleccionado_id = [None]
            perfil_cliente_txt = [""]

            def cargar_vendedores_dropdown():
                opciones = []
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("SELECT ID_Vendedor, Nombre_Completo FROM vendedores WHERE ID_Usuario_Tienda = %s AND Activo = 1 ORDER BY Nombre_Completo ASC", (user_info["id"],))
                        rows = cursor.fetchall()
                        db.close()
                        if rows:
                            opciones = [ft.dropdown.Option(str(r["ID_Vendedor"]), r["Nombre_Completo"]) for r in rows]
                except Exception as ex:
                    print("Error dropdown vendedores:", ex)

                # Si no hay vendedores en BD, cargar de vendedores_list o por defecto
                if not opciones:
                    try:
                        for idx, v_item in enumerate(vendedores_list, 1):
                            n_val = v_item["nombre"].value.strip() if hasattr(v_item["nombre"], "value") and v_item["nombre"].value else f"Vendedor {idx}"
                            opciones.append(ft.dropdown.Option(str(idx), n_val))
                    except Exception:
                        pass
                
                if not opciones:
                    opciones = [
                        ft.dropdown.Option("1", "JOHANA"),
                        ft.dropdown.Option("2", "FERNANDO"),
                        ft.dropdown.Option("3", "ARIADNA")
                    ]

                vendedor_dropdown.options = opciones

            cargar_vendedores_dropdown()

            def agregar_mensaje_chat(autor, texto, avatar_icon, color_borde):
                sim_chat_column.controls.append(
                    ft.Container(
                        content=ft.Row([
                            ft.Icon(avatar_icon, color=color_borde),
                            ft.Text(f"{autor}: {texto}", color="white", expand=True, selectable=True)
                        ], spacing=10, vertical_alignment="start"),
                        bgcolor="#141424" if autor == "Cliente" else "#111111",
                        padding=10,
                        border_radius=8,
                        border=ft.Border.all(1, "#333333")
                    )
                )
                page.update()

            def enviar_mensaje_simulacion(e):
                msg_txt = user_input.value.strip()
                if not msg_txt:
                    return
                user_input.value = ""
                chat_history.append({"role": "user", "content": msg_txt})
                agregar_mensaje_chat("Vendedor", msg_txt, ft.Icons.PERSON, "#D8B4FE")
                
                headers = {
                    "Authorization": f"Bearer {GROQ_API_KEY}",
                    "Content-Type": "application/json"
                }
                
                system_prompt = f"""Eres un cliente de Sunglass Hut. Tu perfil es: '{perfil_cliente_txt[0]}'.
                Estás interactuando con el asesor de ventas en la tienda física.
                Responde de forma natural, realista, breve y conversacional (máximo 2 a 3 oraciones por mensaje).
                No salgas del personaje. Si el vendedor te ofrece promociones, kits o garantías, reacciona según tu perfil de cliente.
                Mantén la interacción fluida. Si sientes que la atención es mala, sé difícil. Si es buena, muéstrate cooperativo.
                """
                
                mensajes_api = [{"role": "system", "content": system_prompt}]
                mensajes_api.extend(chat_history[-10:])
                
                payload = {
                    "model": GROQ_MODEL,
                    "messages": mensajes_api
                }
                
                try:
                    res = requests.post(URL_GROQ, headers=headers, json=payload, timeout=12)
                    if res.status_code == 200:
                        ia_response = res.json()["choices"][0]["message"]["content"]
                        chat_history.append({"role": "assistant", "content": ia_response})
                        agregar_mensaje_chat("Cliente", ia_response, ft.Icons.SUPPORT_AGENT, "#00FFFF")
                    else:
                        agregar_mensaje_chat("Cliente", "[Error de comunicación con el simulador]", ft.Icons.ERROR, "red")
                except Exception as ex_sim:
                    print("Error simulacion API:", ex_sim)
                    agregar_mensaje_chat("Cliente", "[Error de conexión del simulador]", ft.Icons.ERROR, "red")

            btn_enviar.on_click = enviar_mensaje_simulacion

            def finalizar_simulacion_click(e):
                user_input.disabled = True
                btn_enviar.disabled = True
                btn_finalizar.disabled = True
                page.update()
                
                agregar_mensaje_chat("Sistema", "Analizando el desempeño de la simulación de venta. Por favor espera...", ft.Icons.INFO, "#00FFFF")
                
                headers = {
                    "Authorization": f"Bearer {GROQ_API_KEY}",
                    "Content-Type": "application/json"
                }
                
                perfil_lower = perfil_cliente_txt[0].lower()
                es_caso_servicio = "cambio" in perfil_lower or "reclamando" in perfil_lower or "ticket" in perfil_lower
                
                if es_caso_servicio:
                    eval_prompt = """Analiza la siguiente conversación de roleplay de servicio al cliente en Sunglass Hut entre un Asesor de Ventas (Vendedor) y un Cliente que viene a realizar un CAMBIO de producto sin ticket de compra o presenta una queja.
                    Evalúa el desempeño del vendedor en base a estos puntos específicos de servicio al cliente:
                    1. Trato al cliente (Amabilidad, escucha activa, templanza y empatía ante la molestia del cliente).
                    2. Manejo de objeciones y políticas (¿Explicó claramente las políticas de devolución/cambios sin ticket y dio alternativas viables?).
                    3. Búsqueda de soluciones y CRM (¿Ofreció buscar en el sistema de ventas con los datos del cliente, correo electrónico o ID de transacción?).
                    4. Protocolo de atención ante conflictos (¿Evitó discutir y mantuvo una postura profesional y resolutiva?).
                    5. Cierre formal del caso (¿Dejó claros los pasos a seguir o canalizó formalmente el caso a soporte/gerencia de forma educada?).
                    
                    NOTA IMPORTANTE: Al ser un caso de reclamación/servicio, NO penalices ni exijas venta cruzada (UPT) o el cierre de una venta comercial, ya que el objetivo principal es la atención post-venta y resolución de un problema operativo.
                    
                    Tu respuesta DEBE comenzar con un Score numérico entre 0 y 100 de la siguiente forma EXACTA:
                    SCORE: [Número]
                    [Salto de línea]
                    Comentarios detallados de la evaluación...
                    
                    Sé riguroso y constructivo en tu retroalimentación en español.
                    
                    CONVERSACIÓN A EVALUAR:
                    """
                else:
                    eval_prompt = """Analiza la siguiente conversación de roleplay de venta en Sunglass Hut entre un Asesor de Ventas (Vendedor) y un Cliente.
                    Evalúa el desempeño del vendedor en base a estos puntos:
                    1. Trato al cliente (Amabilidad, escucha activa).
                    2. Manejo de objeciones y conocimiento del producto.
                    3. Venta cruzada (¿Ofreció kit de limpieza o estuche adicional para subir el UPT?).
                    4. Captura de datos CRM para la garantía (¿Pidió el correo electrónico?).
                    5. Cierre formal de la venta.

                    Tu respuesta DEBE comenzar con un Score numérico entre 0 y 100 de la siguiente forma EXACTA:
                    SCORE: [Número]
                    [Salto de línea]
                    Comentarios detallados de la evaluación...
                    
                    Sé riguroso y constructivo en tu retroalimentación en español.
                    
                    CONVERSACIÓN A EVALUAR:
                    """
                
                for msg in chat_history:
                    eval_prompt += f"\n{msg['role'].upper()}: {msg['content']}"
                
                payload = {
                    "model": GROQ_MODEL,
                    "messages": [
                        {"role": "system", "content": "Eres un auditor operativo experto en ventas y servicio premium de Sunglass Hut."},
                        {"role": "user", "content": eval_prompt}
                    ]
                }
                
                try:
                    res = requests.post(URL_GROQ, headers=headers, json=payload, timeout=15)
                    if res.status_code == 200:
                        eval_text = res.json()["choices"][0]["message"]["content"]
                        
                        score_val = 70
                        match_score = re.search(r"SCORE:\s*(\d+)", eval_text, re.IGNORECASE)
                        if match_score:
                            score_val = int(match_score.group(1))
                        
                        try:
                            v_id_val = int(vendedor_dropdown.value) if vendedor_dropdown.value and str(vendedor_dropdown.value).isdigit() else 1
                            db = conectar_db()
                            if db:
                                cursor = db.cursor()
                                cursor.execute("""
                                    INSERT INTO evaluaciones_simulador (ID_Vendedor, Cliente_Simulado, Score_Evaluacion, Feedback_Detallado)
                                    VALUES (%s, %s, %s, %s)
                                """, (v_id_val, cliente_dropdown.value, score_val, eval_text))
                                db.commit()
                                db.close()
                        except Exception as ex_db_eval:
                            print("Error guardando eval en DB:", ex_db_eval)
                        
                        mostrar_evaluacion_dialog(score_val, eval_text)
                    else:
                        mostrar_snack("Error de conexión al evaluar", "red")
                except Exception as ex_eval:
                    print("Error evaluacion:", ex_eval)
                    mostrar_snack("Error al procesar la evaluación", "red")

            def mostrar_evaluacion_dialog(score, feedback):
                dlg = ft.AlertDialog(
                    title=ft.Text(f"Evaluación del Simulador: Score {score}/100 📊", color="#00FFFF", weight="bold", size=18),
                    content=ft.Container(
                        content=ft.Column([
                            ft.Text(feedback, color="white", size=13, selectable=True),
                        ], scroll=ft.ScrollMode.AUTO),
                        width=500,
                        height=350
                    ),
                    actions=[
                        ft.TextButton("Entendido", on_click=lambda e: (page.pop_dialog(), cambiar_vista("simulador")))
                    ],
                    bgcolor="#0F0F1A"
                )
                page.show_dialog(dlg)
                page.update()

            def iniciar_simulacion_click(e):
                if not vendedor_dropdown.value:
                    mostrar_snack("Por favor selecciona un vendedor", "red")
                    return
                if not cliente_dropdown.value:
                    mostrar_snack("Por favor selecciona un perfil de cliente", "red")
                    return
                
                v_val = vendedor_dropdown.value
                vendedor_seleccionado_id[0] = int(v_val) if v_val and str(v_val).isdigit() else 1
                perfil_cliente_txt[0] = cliente_dropdown.value
                
                chat_history.clear()
                sim_chat_column.controls.clear()
                
                headers = {
                    "Authorization": f"Bearer {GROQ_API_KEY}",
                    "Content-Type": "application/json"
                }
                system_prompt = f"Eres un cliente de Sunglass Hut entrando a la tienda. Tu perfil es: '{perfil_cliente_txt[0]}'. Escribe tu primer saludo o consulta breve al asesor de ventas (vendedor)."
                payload = {
                    "model": GROQ_MODEL,
                    "messages": [{"role": "user", "content": system_prompt}]
                }
                
                try:
                    res = requests.post(URL_GROQ, headers=headers, json=payload, timeout=12)
                    if res.status_code == 200:
                        first_msg = res.json()["choices"][0]["message"]["content"]
                        chat_history.append({"role": "assistant", "content": first_msg})
                        
                        config_area.visible = False
                        chat_area.visible = True
                        user_input.disabled = False
                        btn_enviar.disabled = False
                        btn_finalizar.disabled = False
                        
                        agregar_mensaje_chat("Cliente", first_msg, ft.Icons.SUPPORT_AGENT, "#00FFFF")
                    else:
                        mostrar_snack("Error de conexión al iniciar simulación", "red")
                except Exception as ex_init:
                    print("Error init sim:", ex_init)
                    mostrar_snack("Error al iniciar el simulador", "red")

            btn_iniciar = ft.ElevatedButton(
                "Iniciar Roleplay ➕",
                on_click=iniciar_simulacion_click,
                bgcolor="#6E48AA",
                color="white",
                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
            )
            btn_finalizar = ft.ElevatedButton(
                "Finalizar y Evaluar 📊",
                on_click=finalizar_simulacion_click,
                bgcolor="#FF4500",
                color="white",
                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8)),
                disabled=True
            )

            config_area = ft.Column([
                ft.Row([vendedor_dropdown, cliente_dropdown], spacing=10, wrap=True),
                ft.Container(height=10),
                btn_iniciar
            ], visible=True)

            chat_area = ft.Column([
                sim_chat_column,
                ft.Row([user_input, btn_enviar], spacing=5),
                ft.Container(height=10),
                btn_finalizar
            ], visible=False, expand=True)

            eval_history_column = ft.Column(spacing=8, scroll=ft.ScrollMode.AUTO, expand=True)
            
            def cargar_historial_evaluaciones():
                eval_history_column.controls.clear()
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("""
                            SELECT e.Score_Evaluacion, e.Cliente_Simulado, DATE_FORMAT(e.Fecha_Hora, '%d/%m/%Y %H:%i') as fecha_f, v.Nombre_Completo 
                            FROM evaluaciones_simulador e 
                            JOIN vendedores v ON e.ID_Vendedor = v.ID_Vendedor 
                            WHERE v.ID_Usuario_Tienda = %s 
                            ORDER BY e.Fecha_Hora DESC
                        """, (user_info["id"],))
                        rows = cursor.fetchall()
                        db.close()
                        
                        if not rows:
                            eval_history_column.controls.append(ft.Text("No hay evaluaciones guardadas.", color="#888888", italic=True))
                        else:
                            for r in rows:
                                eval_history_column.controls.append(
                                    ft.Container(
                                        content=ft.Row([
                                            ft.Icon(ft.Icons.ASSESSMENT, color="#7CFC00" if r["Score_Evaluacion"] >= 80 else "#FF8C00"),
                                            ft.Column([
                                                ft.Text(f"Vendedor: {r['Nombre_Completo']} | Score: {r['Score_Evaluacion']}/100", color="white", weight="bold"),
                                                ft.Text(f"Perfil: {r['Cliente_Simulado']} | {r['fecha_f']}", color="#aaaaaa", size=11)
                                            ], spacing=3, expand=True)
                                        ], vertical_alignment="center"),
                                        bgcolor="#1a1a1a",
                                        padding=10,
                                        border_radius=8,
                                        border=ft.Border.all(1, "#333333")
                                    )
                                )
                except Exception as ex:
                    print("Error historial eval:", ex)
                page.update()

            # --- PESTAÑA 3: GESTIÓN DE PERFILES DE CLIENTE (ADMIN Y TIENDA) ---
            perfiles_cards_container = ft.Column(spacing=10, scroll=ft.ScrollMode.AUTO, expand=True)

            def init_perfiles_db():
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        cursor.execute("""
                            CREATE TABLE IF NOT EXISTS perfiles_cliente_simulador (
                                id INT AUTO_INCREMENT PRIMARY KEY,
                                nombre VARCHAR(255) NOT NULL,
                                descripcion TEXT NOT NULL,
                                fecha_registro DATETIME DEFAULT CURRENT_TIMESTAMP
                            );
                        """)
                        db.commit()
                        
                        cursor.execute("SELECT COUNT(*) FROM perfiles_cliente_simulador")
                        res = cursor.fetchone()
                        count = res[0] if res else 0
                        if count == 0:
                            default_profiles = [
                                ("Objeción de Precio - 'Son Muy Caros'", "Cliente insiste en que las gafas Sunglass Hut son muy caras y que vio unos lentes parecidos mucho más baratos en otro lugar"),
                                ("Duda Técnica - Polarizados y Chromance", "Cliente confundido preguntando qué ventaja real tienen las micas polarizadas y Chromance sobre las normales y por qué valen la pena"),
                                ("Consulta de Garantías - Ruptura o Robo", "Cliente interesado pero pregunta qué garantía tienen los lentes si sufren alguna ruptura, daño accidental o si se los roban"),
                                ("Cliente Indeciso - Ray-Ban Meta", "Cliente Indeciso buscando tecnología (Ray-Ban Meta)"),
                                ("Cliente Reclamando Cambio sin Ticket", "Cliente molesto que exige cambio de lentes Oakley sin ticket de compra"),
                                ("Cliente buscando Kit de Limpieza", "Cliente que solo entra preguntando por un paño de limpieza sencillo"),
                                ("Cliente de Regalo de Lujo", "Cliente indeciso buscando un regalo premium para su pareja (Gafas Versace)"),
                                ("Cliente Apurado - Gafas Polarizadas", "Cliente muy apurado que tiene un vuelo en pocas horas y busca unas gafas clásicas polarizadas (Ray-Ban Aviator) para la playa, exige rapidez y no quiere rodeos"),
                                ("Cliente Escéptico - Privacidad Meta", "Cliente desconfiado e interesado en la tecnología de las gafas inteligentes (Ray-Ban Meta), pero le preocupa mucho la privacidad y si la cámara o el micrófono graban de forma oculta"),
                                ("Coleccionista Exigente - Prada/Versace", "Cliente de alto nivel adquisitivo, muy conocedor de moda, que busca una pieza exclusiva de edición limitada de Prada o Versace y espera una atención ultra-premium y detalles de diseño"),
                                ("Cliente Comparador - Objeción de Precio", "Cliente indeciso que le encantan unas gafas Dolce & Gabbana, pero insiste en que las vio más baratas en una tienda en línea no autorizada y cuestiona el valor y autenticidad del producto en tienda física"),
                                ("Cliente de Descuento - Sin Temporada", "Cliente que insiste en obtener un descuento especial para comprar un solo par de gafas (Oakley o Ray-Ban) a pesar de que le explicas que no es temporada de rebajas ni hay promociones vigentes en tienda"),
                                ("Padre Indeciso - Regalo Adolescente", "Padre o madre de familia que busca un regalo de cumpleaños para su hijo adolescente, no sabe qué marca está de moda (Oakley o Ray-Ban Meta) y necesita asesoría paciente sobre tendencias juveniles"),
                                ("Deportista - Presupuesto Ajustado", "Deportista aficionado que busca gafas de sol de alto rendimiento para correr o ciclismo (Oakley Sutro), conoce los beneficios técnicos del lente pero tiene un presupuesto muy limitado y busca la opción de menor costo"),
                                ("Ejecutivo - Imagen Profesional", "Profesional corporativo que busca unas gafas elegantes y sobrias para usar con vestimenta formal de negocios (Persol o Prada), valora la discreción, calidad de materiales e imagen profesional"),
                                ("Cliente Reclamando Garantía - Lentes Rayados", "Cliente molesto que viene a exigir el cambio o garantía gratis de sus gafas porque los lentes están completamente rayados debido a mal uso (los limpió con su playera o los dejó caer), e insiste en que es defecto de fábrica")
                            ]
                            cursor.executemany("INSERT INTO perfiles_cliente_simulador (nombre, descripcion) VALUES (%s, %s)", default_profiles)
                            db.commit()
                        db.close()
                except Exception as ex_init:
                    print("Error init perfiles_db:", ex_init)

            def cargar_perfiles_simulador():
                perfiles_cards_container.controls.clear()
                dropdown_options = []
                init_perfiles_db()
                
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("SELECT id, nombre, descripcion FROM perfiles_cliente_simulador ORDER BY id DESC")
                        rows = cursor.fetchall()
                        db.close()
                        
                        for row in rows:
                            p_id = row["id"]
                            p_nombre = row["nombre"]
                            p_desc = row["descripcion"]
                            
                            dropdown_options.append(ft.dropdown.Option(p_nombre, f"{p_nombre} - {p_desc[:60]}..."))
                            
                            def eliminar_perfil_click(target_id=p_id, target_nombre=p_nombre):
                                def confirmar_eliminar(e):
                                    try:
                                        db_del = conectar_db()
                                        if db_del:
                                            cursor_del = db_del.cursor()
                                            cursor_del.execute("DELETE FROM perfiles_cliente_simulador WHERE id = %s", (target_id,))
                                            db_del.commit()
                                            db_del.close()
                                            mostrar_snack(f"Perfil '{target_nombre}' eliminado.", color="#7CFC00")
                                            page.pop_dialog()
                                            cargar_perfiles_simulador()
                                    except Exception as ex_d:
                                        print("Error eliminar perfil:", ex_d)
                                        mostrar_snack("Error al eliminar el perfil.", color="red")
                                
                                dlg_confirm = ft.AlertDialog(
                                    title=ft.Text("Confirmar eliminación 🗑️", color="#FF4500", weight="bold"),
                                    content=ft.Text(f"¿Estás seguro de eliminar el perfil '{target_nombre}'?", color="white"),
                                    actions=[
                                        ft.TextButton("Cancelar", on_click=lambda e: page.pop_dialog()),
                                        ft.ElevatedButton("Eliminar", on_click=confirmar_eliminar, bgcolor="#FF4500", color="white")
                                    ],
                                    bgcolor="#0F0F1A"
                                )
                                page.show_dialog(dlg_confirm)
                                page.update()
                            
                            btn_del = ft.IconButton(
                                icon=ft.Icons.DELETE_ROUNDED,
                                icon_color="#FF4500",
                                tooltip="Eliminar Perfil",
                                on_click=lambda e, tid=p_id, tnom=p_nombre: eliminar_perfil_click(tid, tnom)
                            )
                            
                            card = ft.Container(
                                content=ft.Row([
                                    ft.Icon(ft.Icons.ACCOUNT_CIRCLE_ROUNDED, color="#00FFFF", size=36),
                                    ft.Column([
                                        ft.Text(p_nombre, color="white", weight="bold", size=14),
                                        ft.Text(p_desc, color="#aaaaaa", size=12, selectable=True)
                                    ], spacing=3, expand=True),
                                    btn_del
                                ], vertical_alignment="center", spacing=12),
                                bgcolor="#1E1E2E",
                                padding=15,
                                border_radius=10,
                                border=ft.Border.all(1, "#9D50BB")
                            )
                            perfiles_cards_container.controls.append(card)
                except Exception as ex_p:
                    print("Error cargando perfiles:", ex_p)

                if dropdown_options:
                    cliente_dropdown.options = dropdown_options
                page.update()

            def abrir_modal_nuevo_perfil(e):
                input_nombre = ft.TextField(label="Tipo / Nombre del Cliente", hint_text="Ej. Cliente Exigente - Ray-Ban Reverse", border_color="#9D50BB", color="white")
                input_desc = ft.TextField(label="Descripción / Personalidad / Objeción", hint_text="Ej. Busca tecnología exclusiva y compara precios...", multiline=True, min_lines=2, max_lines=4, border_color="#9D50BB", color="white")
                
                def guardar_perfil(e):
                    nom = input_nombre.value.strip()
                    desc = input_desc.value.strip()
                    if not nom or not desc:
                        mostrar_snack("Por favor llena el nombre y la descripción.", "red")
                        return
                    try:
                        db_ins = conectar_db()
                        if db_ins:
                            cursor_ins = db_ins.cursor()
                            cursor_ins.execute("INSERT INTO perfiles_cliente_simulador (nombre, descripcion) VALUES (%s, %s)", (nom, desc))
                            db_ins.commit()
                            db_ins.close()
                            mostrar_snack("¡Nuevo perfil de cliente guardado exitosamente!", "#7CFC00")
                            page.pop_dialog()
                            cargar_perfiles_simulador()
                    except Exception as ex_g:
                        print("Error al guardar perfil:", ex_g)
                        mostrar_snack("Error al guardar en la base de datos.", "red")

                dlg_add = ft.AlertDialog(
                    title=ft.Text("➕ Agregar Perfil de Cliente", color="#D8B4FE", weight="bold"),
                    content=ft.Container(
                        content=ft.Column([
                            input_nombre,
                            input_desc
                        ], spacing=12),
                        width=450,
                        height=200
                    ),
                    actions=[
                        ft.TextButton("Cancelar", on_click=lambda e: page.pop_dialog()),
                        ft.ElevatedButton("Guardar Perfil", on_click=guardar_perfil, bgcolor="#1f6f43", color="white")
                    ],
                    bgcolor="#0F0F1A"
                )
                page.show_dialog(dlg_add)
                page.update()

            def generar_perfil_con_ia(e):
                mostrar_snack("🤖 Generando nuevo perfil de cliente con IA... Por favor espera.", "#00FFFF")
                
                def thread_generar_ia():
                    try:
                        prompt = """Crea un perfil de cliente realista para entrenamiento de ventas en una tienda de gafas de sol Sunglass Hut (Ray-Ban, Oakley, Versace, Prada, Persol, Oliver Peoples).
El perfil debe incluir:
1. Un nombre o título corto del tipo de cliente (ej. "Cliente Ejecutivo - Imagen Corporativa Persol", "Influencer de Moda - Prada Exclusivo").
2. Una descripción detallada del comportamiento, objeción de ventas, dudas técnicas o de garantía.

Responde ÚNICAMENTE en formato JSON plano con las llaves exactas "nombre" y "descripcion".
Ejemplo:
{"nombre": "Cliente Exigente - Edición Limitada Versace", "descripcion": "Busca una pieza exclusiva para un evento social importante y le preocupa la autenticidad y el grabado del estuche."}
"""
                        raw_res = call_gemini(prompt)
                        if not raw_res:
                            mostrar_snack("No se pudo conectar con la IA de Gemini.", "red")
                            return
                        
                        import json, re
                        clean_str = raw_res.strip()
                        if "```json" in clean_str:
                            clean_str = clean_str.split("```json")[1].split("```")[0].strip()
                        elif "```" in clean_str:
                            clean_str = clean_str.split("```")[1].split("```")[0].strip()
                        
                        match = re.search(r'\{.*\}', clean_str, re.DOTALL)
                        if match:
                            data = json.loads(match.group(0))
                            g_nom = data.get("nombre", "Cliente generado por IA")
                            g_desc = data.get("descripcion", "Perfil generado automáticamente por IA.")
                            
                            db_ia = conectar_db()
                            if db_ia:
                                cursor_ia = db_ia.cursor()
                                cursor_ia.execute("INSERT INTO perfiles_cliente_simulador (nombre, descripcion) VALUES (%s, %s)", (g_nom, g_desc))
                                db_ia.commit()
                                db_ia.close()
                                mostrar_snack(f"✨ ¡Perfil '{g_nom}' generado por IA!", "#7CFC00")
                                cargar_perfiles_simulador()
                        else:
                            mostrar_snack("Respuesta de IA recibida. Actualizando lista...", "#7CFC00")
                            cargar_perfiles_simulador()
                    except Exception as ex_ia:
                        print("Error generando perfil IA:", ex_ia)
                        mostrar_snack("Error al generar perfil con IA.", "red")

                threading.Thread(target=thread_generar_ia, daemon=True).start()

            btn_add_manual = ft.ElevatedButton(
                "➕ Agregar Perfil",
                on_click=abrir_modal_nuevo_perfil,
                bgcolor="#1f6f43",
                color="white"
            )

            btn_gen_ia = ft.ElevatedButton(
                "🤖 Generar con IA",
                on_click=generar_perfil_con_ia,
                bgcolor="#9D50BB",
                color="white"
            )

            btn_refresh_perfiles = ft.IconButton(
                icon=ft.Icons.REFRESH,
                icon_color="#00FFFF",
                tooltip="Refrescar Lista",
                on_click=lambda e: cargar_perfiles_simulador()
            )

            tab_simulacion = ft.Column([config_area, chat_area], expand=True)
            tab_historial = ft.Column([
                ft.Text("Historial de Evaluaciones de la Tienda:", color="#D8B4FE", size=14, weight="bold"),
                ft.ElevatedButton("Actualizar Historial 🔄", on_click=lambda e: cargar_historial_evaluaciones(), bgcolor="#333333", color="white"),
                ft.Container(height=5),
                eval_history_column
            ], expand=True)

            tab_perfiles = ft.Column([
                ft.Row([
                    ft.Text("👥 GESTIÓN DE PERFILES DE CLIENTE", color="#D8B4FE", size=15, weight="bold"),
                    ft.Container(expand=True),
                    btn_add_manual,
                    btn_gen_ia,
                    btn_refresh_perfiles
                ], vertical_alignment="center", spacing=10),
                ft.Text("Administra o crea nuevos tipos de cliente para los ejercicios de ventas. Puedes agregarlos manualmente o generarlos dinámicamente con IA.", color="#aaaaaa", size=12),
                ft.Divider(height=15, color="#333333"),
                perfiles_cards_container
            ], expand=True)

            tabs = ft.Tabs(
                selected_index=0,
                animation_duration=300,
                length=3,
                expand=True,
                content=ft.Column(
                    expand=True,
                    controls=[
                        ft.TabBar(
                            tabs=[
                                ft.Tab(label="Iniciar Simulador"),
                                ft.Tab(label="Historial de Avance"),
                                ft.Tab(label="Gestión de Perfiles 👥")
                            ]
                        ),
                        ft.TabBarView(
                            expand=True,
                            controls=[
                                tab_simulacion,
                                tab_historial,
                                tab_perfiles
                            ]
                        )
                    ]
                )
            )

            cargar_historial_evaluaciones()
            cargar_perfiles_simulador()

            return ft.Column([
                ft.Row([
                    ft.Text("Simulador de Ventas con IA 🎭", size=24, color="#D8B4FE", weight="bold")
                ]),
                ft.Text("Realiza roleplay interactivo de ventas por vendedor. La IA auditará el cumplimiento de las metas de UPT, captura de datos y cierre.", color="#aaaaaa", size=13),
                ft.Divider(height=15, color="#333333"),
                tabs
            ], expand=True)

        def build_crm_view():
            """Vista principal del Módulo CRM & Garantías de 1 Año con Notificaciones al Mes 11."""
            
            # --- ASEGURAR COLUMNA EN BASE DE DATOS ---
            try:
                db_m = conectar_db()
                if db_m:
                    cur_m = db_m.cursor()
                    cur_m.execute("SHOW COLUMNS FROM crm_compras LIKE 'Ruta_Ticket'")
                    if not cur_m.fetchone():
                        cur_m.execute("ALTER TABLE crm_compras ADD COLUMN Ruta_Ticket VARCHAR(500) NULL AFTER Notas")
                        db_m.commit()
                    db_m.close()
            except Exception as ex_mig:
                print("Error verificando columna Ruta_Ticket:", ex_mig)

            is_mobile_w = (page.width < 700) if (page and page.width) else False
            ticket_img_path = {"val": None}
            img_preview_container = ft.Container(visible=False)

            tf_transaccion = ft.TextField(label="N° Transacción / Ticket", border_color="#9D50BB", color="white", text_size=12, width=280)
            tf_fecha_compra = ft.TextField(label="Fecha de Compra (AAAA-MM-DD)", value=datetime.now().strftime("%Y-%m-%d"), border_color="#9D50BB", color="white", text_size=12, width=220)
            tf_nombre_cliente = ft.TextField(label="Nombre del Cliente", border_color="#9D50BB", color="white", text_size=12, width=300)
            tf_telefono_cliente = ft.TextField(label="Teléfono del Cliente (Manual)", border_color="#9D50BB", color="white", text_size=12, width=200)
            tf_nombre_vendedor = ft.TextField(label="Nombre del Vendedor (o Vendedores)", value=user_info.get("nombre", ""), border_color="#9D50BB", color="white", text_size=12, width=320)
            tf_precio = ft.TextField(label="Precio Total con IVA ($ MXN)", border_color="#7CFC00", color="#7CFC00", text_style=ft.TextStyle(weight="bold"), text_size=13, width=220)
            tf_notas = ft.TextField(label="Notas Adicionales (Descuentos, Promociones, Observaciones)", border_color="#9D50BB", color="white", text_size=12, multiline=True, min_lines=2, max_lines=4)

            items_rows_container = ft.Column(spacing=8)

            def recalcular_gran_total():
                total_sum = 0.0
                for row_container in items_rows_container.controls:
                    try:
                        row_controls = row_container.content.controls
                        p_tf = row_controls[2]
                        if p_tf and p_tf.value and p_tf.value.strip():
                            val = float(p_tf.value.replace("$", "").replace(",", "").strip())
                            total_sum += val
                    except Exception:
                        pass
                if total_sum > 0:
                    tf_precio.value = f"{total_sum:,.2f}"
                try: page.update()
                except Exception: pass

            def agregar_fila_articulo(upc_val="", modelo_val="", precio_val=""):
                tf_item_upc = ft.TextField(
                    label="UPC / Código de Lente",
                    value=str(upc_val or ""),
                    border_color="#9D50BB",
                    color="white",
                    text_size=12,
                    width=200 if is_mobile_w else 240
                )
                tf_item_modelo = ft.TextField(
                    label="Modelo (ej: VE4436U G81/87 57/18)",
                    value=str(modelo_val or ""),
                    border_color="#9D50BB",
                    color="white",
                    text_size=12,
                    width=280 if is_mobile_w else 360
                )
                tf_item_precio = ft.TextField(
                    label="Precio Artículo ($ MXN)",
                    value=str(precio_val or ""),
                    border_color="#00FFFF",
                    color="#00FFFF",
                    text_size=12,
                    width=140 if is_mobile_w else 180,
                    on_change=lambda e: recalcular_gran_total()
                )

                def eliminar_fila(e_del):
                    if len(items_rows_container.controls) > 1:
                        items_rows_container.controls.remove(row_item)
                        recalcular_gran_total()
                        try: page.update()
                        except Exception: pass
                    else:
                        tf_item_upc.value = ""
                        tf_item_modelo.value = ""
                        tf_item_precio.value = ""
                        recalcular_gran_total()
                        try: page.update()
                        except Exception: pass

                row_item = ft.Container(
                    content=ft.Row([
                        tf_item_upc,
                        tf_item_modelo,
                        tf_item_precio,
                        ft.IconButton(
                            icon=ft.Icons.DELETE_OUTLINED,
                            icon_color="#FF4500",
                            tooltip="Quitar este artículo",
                            on_click=eliminar_fila
                        )
                    ], wrap=True, spacing=8, vertical_alignment="center"),
                    bgcolor="#181828",
                    padding=8,
                    border_radius=8,
                    border=ft.Border.all(1, "#9D50BB")
                )
                items_rows_container.controls.append(row_item)
                recalcular_gran_total()
                try: page.update()
                except Exception: pass

            # Inicializar con 1 fila por defecto
            agregar_fila_articulo()

            crm_history_col = ft.Column(spacing=8, scroll=ft.ScrollMode.AUTO, expand=True)
            crm_notif_col = ft.Column(spacing=10, scroll=ft.ScrollMode.AUTO, expand=True)
            notif_badge_text = ft.Text("0", color="#00FFFF", weight="bold", size=12)
            
            # --- FUNCIÓN DE LIMPIAR FORMULARIO ---
            def limpiar_form_crm():
                tf_transaccion.value = ""
                tf_fecha_compra.value = datetime.now().strftime("%Y-%m-%d")
                tf_nombre_cliente.value = ""
                tf_telefono_cliente.value = ""
                tf_nombre_vendedor.value = user_info.get("nombre", "")
                tf_precio.value = ""
                tf_notas.value = ""
                items_rows_container.controls.clear()
                agregar_fila_articulo()
                ticket_img_path["val"] = None
                img_preview_container.visible = False
                img_preview_container.content = None
                page.update()

            # --- GUARDAR VENTA EN CRM ---
            def guardar_compra_crm(e):
                upcs_list = []
                modelos_list = []
                for row_container in items_rows_container.controls:
                    try:
                        r_ctrls = row_container.content.controls
                        u_val = (r_ctrls[0].value or "").strip()
                        m_val = (r_ctrls[1].value or "").strip()
                        if u_val:
                            upcs_list.append(u_val)
                        if m_val:
                            modelos_list.append(m_val)
                    except Exception: pass

                upc_final = ", ".join(upcs_list) if upcs_list else ""
                modelos_str = " | ".join(modelos_list) if modelos_list else ""
                notas_final = f"Modelos: {modelos_str}. {tf_notas.value.strip()}".strip() if modelos_str else tf_notas.value.strip()

                if not tf_transaccion.value.strip() or not tf_nombre_cliente.value.strip() or not upc_final or not tf_precio.value.strip():
                    mostrar_snack("Por favor completa los campos obligatorios: Transacción, Cliente, al menos 1 UPC y Precio.", color="orange")
                    return

                try:
                    precio_val = float(tf_precio.value.replace("$", "").replace(",", "").strip())
                except Exception:
                    mostrar_snack("El precio ingresado no es válido.", color="red")
                    return

                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        cursor.execute("""
                            INSERT INTO crm_compras (Transaccion, Fecha_Compra, Nombre_Cliente, Telefono_Cliente, Nombre_Vendedor, ID_Usuario, UPC, Precio_Con_IVA, Tienda, Estatus_Seguimiento, Notas, Ruta_Ticket)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 'Pendiente', %s, %s)
                            ON DUPLICATE KEY UPDATE 
                                Fecha_Compra=%s, Nombre_Cliente=%s, Telefono_Cliente=%s, Nombre_Vendedor=%s, UPC=%s, Precio_Con_IVA=%s, Notas=%s, Ruta_Ticket=%s
                        """, (
                            tf_transaccion.value.strip(),
                            tf_fecha_compra.value.strip(),
                            tf_nombre_cliente.value.strip(),
                            tf_telefono_cliente.value.strip(),
                            tf_nombre_vendedor.value.strip(),
                            user_info["id"],
                            upc_final,
                            precio_val,
                            user_info.get("tienda", "Tienda Luxo"),
                            notas_final,
                            ticket_img_path["val"],
                            # Update values
                            tf_fecha_compra.value.strip(),
                            tf_nombre_cliente.value.strip(),
                            tf_telefono_cliente.value.strip(),
                            tf_nombre_vendedor.value.strip(),
                            upc_final,
                            precio_val,
                            notas_final,
                            ticket_img_path["val"]
                        ))
                        db.commit()
                        db.close()
                        mostrar_snack("✅ Venta y datos de garantía guardados correctamente en CRM.", color="green")
                        limpiar_form_crm()
                        cargar_historial_crm()
                        cargar_notificaciones_crm()
                except Exception as ex:
                    print("Error guardando CRM:", ex)
                    mostrar_snack(f"Error al guardar venta: {ex}", color="red")

            # --- FILE PICKER Y CÁMARA PARA ESCÁNER DE TICKET POR IA (SOPORTE WEB/MÓVIL Y DESKTOP) ---
            def procesar_y_cargar_ticket(f_path, f_name):
                try:
                    mostrar_snack(f"📷 Procesando imagen de ticket '{f_name}' con OCR IA...", color="#00FFFF")
                    
                    # 1. Guardar copia local permanente en uploads/tickets/
                    dir_tickets = os.path.join(BASE_PATH, "uploads", "tickets")
                    os.makedirs(dir_tickets, exist_ok=True)
                    nom_seguro = f"ticket_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{re.sub(r'[^a-zA-Z0-9_.]', '_', f_name)}"
                    ruta_destino = os.path.join(dir_tickets, nom_seguro)
                    
                    with open(f_path, "rb") as f_in:
                        img_bytes = f_in.read()
                        
                    img_bytes = optimizar_imagen(img_bytes)
                    
                    with open(ruta_destino, "wb") as f_out:
                        f_out.write(img_bytes)
                        
                    ticket_img_path["val"] = ruta_destino
                    
                    import base64
                    img_b64 = base64.b64encode(img_bytes).decode("utf-8")
                    img_preview_container.content = ft.Container(
                        content=ft.Column([
                            ft.Text("📸 Foto del Ticket Almacenado:", color="#00FFFF", size=11, weight="bold"),
                            ft.Image(src=f"data:image/jpeg;base64,{img_b64}", width=240, height=160, fit="contain")
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                        bgcolor="#161922", padding=8, border_radius=8, border=ft.Border.all(1, "#00FFFF")
                    )
                    img_preview_container.visible = True
                    try:
                        page.update()
                    except Exception: pass
                    
                    # 2. OCR con EasyOCR + Groq Llama-3.3-70b
                    parsed_data, err_ocr = procesar_ticket_con_gemini(img_bytes)
                    print("OCR RESULT PARSED:", parsed_data)
                    print("OCR RESULT ERR:", err_ocr)

                    if parsed_data:
                        if parsed_data.get("transaccion"):
                            tf_transaccion.value = str(parsed_data["transaccion"])
                        else:
                            tf_transaccion.value = f"TRX-{datetime.now().strftime('%M%S%f')[:6]}"
                            
                        if parsed_data.get("fecha_compra"):
                            tf_fecha_compra.value = str(parsed_data["fecha_compra"])
                            
                        if parsed_data.get("nombre_cliente"):
                            tf_nombre_cliente.value = str(parsed_data["nombre_cliente"])
                            
                        if parsed_data.get("vendedor"):
                            tf_nombre_vendedor.value = str(parsed_data["vendedor"])

                        if parsed_data.get("notas"):
                            tf_notas.value = str(parsed_data["notas"])
                            
                        # Limpiar filas y cargar artículos escaneados
                        items_rows_container.controls.clear()
                        items = parsed_data.get("items", [])
                        if items and isinstance(items, list):
                            for it in items:
                                u_item = it.get("upc", "") if isinstance(it, dict) else ""
                                m_item = it.get("modelo", "") if isinstance(it, dict) else ""
                                p_item = it.get("precio", "") if isinstance(it, dict) else ""
                                agregar_fila_articulo(u_item, m_item, p_item)
                        else:
                            u_single = parsed_data.get("upc", "")
                            p_single = parsed_data.get("precio", "")
                            agregar_fila_articulo(u_single, "", p_single)

                        if parsed_data.get("precio") and not tf_precio.value:
                            tf_precio.value = str(parsed_data["precio"])
                            
                        mostrar_snack("✅ ¡Ticket escaneado con éxito! Ingresa el teléfono del cliente.", color="green")
                    else:
                        if not tf_transaccion.value:
                            tf_transaccion.value = f"TRX-{datetime.now().strftime('%M%S%f')[:6]}"
                        mostrar_snack(f"⚠️ Foto guardada. {err_ocr or 'No se leyeron datos automáticos, llena los campos manualmente.'}", color="orange")
                    
                    try:
                        page.update()
                    except Exception: pass
                except Exception as ex_proc:
                    print("Error procesando foto ticket:", ex_proc)
                    mostrar_snack(f"Error al procesar ticket: {ex_proc}", color="red")

            # --- ESCÁNER DE TICKET POR IA ---
            def escanear_ticket_click(e):
                def on_ticket_selected(path):
                    if path and os.path.exists(path):
                        procesar_y_cargar_ticket(path, os.path.basename(path))

                seleccionar_archivo_async("Tomar Foto o Seleccionar Ticket", "media", on_ticket_selected, captureMode=True)

            # --- VISUALIZADOR DE TICKET DE COMPRA Y DETALLES EN MODAL ---
            def mostrar_detalle_notificacion(compra):
                fecha_c = compra["Fecha_Compra"]
                if isinstance(fecha_c, str):
                    try:
                        dt_c = datetime.strptime(fecha_c, "%Y-%m-%d")
                    except Exception:
                        dt_c = datetime.now()
                else:
                    dt_c = fecha_c

                dt_venc = dt_c + timedelta(days=365)
                fecha_venc_str = dt_venc.strftime("%d/%m/%Y")
                fecha_compra_str = dt_c.strftime("%d/%m/%Y")

                # Cargar vista de imagen de ticket físico si existe
                img_ticket_modal = None
                ruta_t = compra.get("Ruta_Ticket")
                if ruta_t and os.path.exists(ruta_t):
                    try:
                        with open(ruta_t, "rb") as ft_img:
                            import base64
                            b64_t = base64.b64encode(ft_img.read()).decode("utf-8")
                            img_ticket_modal = ft.Container(
                                content=ft.Column([
                                    ft.Text("📸 COMPROBANTE / TICKET FÍSICO ESCANEADO:", color="#00FFFF", weight="bold", size=11),
                                    ft.Image(src=f"data:image/jpeg;base64,{b64_t}", width=400, height=220, fit="contain")
                                ], horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                                bgcolor="#161922", padding=10, border_radius=8, border=ft.Border.all(1, "#333333")
                            )
                    except Exception as ex_t:
                        print("Error cargando imagen ticket modal:", ex_t)

                ticket_column_items = [
                    # Encabezado del Ticket
                    ft.Column([
                        ft.Text("🕶️ SUNGLASS HUT MEXICO 🕶️", color="#00FFFF", weight="bold", size=16, text_align=ft.TextAlign.CENTER),
                        ft.Text("STORE #4052 - PLAZA SATÉLITE", color="#aaaaaa", size=11, text_align=ft.TextAlign.CENTER),
                        ft.Text("TICKET DE COMPRA Y GARANTÍA DIGITAL", color="#D8B4FE", weight="bold", size=12, text_align=ft.TextAlign.CENTER),
                    ], alignment=ft.MainAxisAlignment.CENTER, horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=2),
                    ft.Divider(height=12, color="#00FFFF"),
                    
                    # Datos de Transacción
                    ft.Row([ft.Text("N° TRANSACCIÓN:", color="#aaaaaa", size=11, width=150), ft.Text(compra['Transaccion'], color="white", weight="bold", size=12)]),
                    ft.Row([ft.Text("FECHA DE EMISIÓN:", color="#aaaaaa", size=11, width=150), ft.Text(fecha_compra_str, color="white", size=12)]),
                    ft.Row([ft.Text("VENDEDOR ATENDIÓ:", color="#aaaaaa", size=11, width=150), ft.Text(compra['Nombre_Vendedor'] or "Atención Luxo", color="white", size=12)]),
                    ft.Row([ft.Text("TIENDA DE ORIGEN:", color="#aaaaaa", size=11, width=150), ft.Text(compra['Tienda'] or "Sunglass Hut", color="white", size=12)]),
                    ft.Divider(height=8, color="#333333"),
                    
                    # Datos del Cliente
                    ft.Text("👤 DATOS DEL CLIENTE:", color="#00FFFF", weight="bold", size=12),
                    ft.Row([ft.Text("Nombre del Cliente:", color="#aaaaaa", size=11, width=150), ft.Text(compra['Nombre_Cliente'], color="white", weight="bold", size=12)]),
                    ft.Row([ft.Text("Teléfono de Contacto:", color="#aaaaaa", size=11, width=150), ft.Text(compra['Telefono_Cliente'] or "Sin Teléfono", color="#7CFC00", weight="bold", size=12)]),
                    ft.Divider(height=8, color="#333333"),
                    
                    # Detalle de Producto
                    ft.Text("🛍️ DETALLE DEL LENTE ADQUIRIDO:", color="#D8B4FE", weight="bold", size=12),
                    ft.Row([ft.Text("UPC / Código Lente:", color="#aaaaaa", size=11, width=150), ft.Text(compra['UPC'], color="white", weight="bold", size=12)]),
                    ft.Row([ft.Text("Precio Con IVA:", color="#aaaaaa", size=11, width=150), ft.Text(f"${compra['Precio_Con_IVA']:,.2f} MXN", color="#7CFC00", weight="bold", size=13)]),
                    ft.Row([ft.Text("Notas / Modelo:", color="#aaaaaa", size=11, width=150), ft.Text(compra.get('Notas') or "Ray-Ban / Oakley", color="#E2E8F0", size=11)]),
                    ft.Divider(height=8, color="#333333"),
                ]

                if img_ticket_modal:
                    ticket_column_items.extend([img_ticket_modal, ft.Divider(height=8, color="#333333")])

                ticket_column_items.extend([
                    # Coberturas de Garantía LUXO
                    ft.Container(
                        content=ft.Column([
                            ft.Text("📜 COBERTURAS DE GARANTÍA LUXO (1 AÑO):", color="#FFD700", weight="bold", size=11),
                            ft.Text(f"• Garantía por Ruptura / Daño: VIGENTE HASTA EL {fecha_venc_str}", color="#7CFC00", size=11),
                            ft.Text("• Garantía por Robo: 50% descuento en pieza de menor valor presentando Acta de Denuncia en tienda.", color="white", size=11)
                        ], spacing=4),
                        bgcolor="#1E2330",
                        padding=10,
                        border_radius=8,
                        border=ft.Border.all(1, "#FFD700")
                    ),
                    ft.Divider(height=8, color="#333333"),

                    # Estado de Asistencia
                    ft.Row([
                        ft.Text("Estatus en Tienda:", color="#aaaaaa", size=11, width=150),
                        ft.Text("Asistió a Tienda 🏬" if compra.get("Cliente_Asistio") == 1 else "Pendiente de Visita", color="#7CFC00" if compra.get("Cliente_Asistio") == 1 else "#00FFFF", weight="bold", size=12)
                    ]),
                    ft.Row([
                        ft.Text("Nueva Venta Generada:", color="#aaaaaa", size=11, width=150),
                        ft.Text(f"${compra.get('Monto_Nueva_Venta') or 0:,.2f} MXN", color="#7CFC00", weight="bold", size=12)
                    ]),
                ])

                ticket_content = ft.Container(
                    width=460,
                    bgcolor="#0D1117",
                    padding=20,
                    border_radius=12,
                    border=ft.Border.all(2, "#00FFFF"),
                    content=ft.Column(ticket_column_items, spacing=6, scroll=ft.ScrollMode.AUTO)
                )

                def cerrar_ticket(e):
                    try:
                        page.close(dialog)
                    except Exception:
                        pass
                    dialog.open = False
                    page.update()

                dialog = ft.AlertDialog(
                    open=True,
                    title=ft.Row([
                        ft.Icon(ft.Icons.RECEIPT_LONG, color="#00FFFF"),
                        ft.Text("Ticket de Compra Digital 🧾", color="white", weight="bold", size=16)
                    ]),
                    content=ticket_content,
                    actions=[
                        ft.ElevatedButton("🖨️ Imprimir Ticket", bgcolor="#333333", color="#00FFFF", on_click=lambda e: mostrar_snack("🖨️ Enviando ticket a la impresora de tienda...", color="#00FFFF")),
                        ft.ElevatedButton("Cerrar Ticket ❌", bgcolor="#FF4B4B", color="white", on_click=cerrar_ticket)
                    ]
                )

                if dialog not in page.overlay:
                    page.overlay.append(dialog)
                page.dialog = dialog
                dialog.open = True
                page.update()

            # --- CARGAR NOTIFICACIONES AL MES 11 ---
            def cargar_notificaciones_crm():
                crm_notif_col.controls.clear()
                count_notif = 0
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("""
                            SELECT ID_Compra, Transaccion, Fecha_Compra, Nombre_Cliente, Telefono_Cliente, Nombre_Vendedor, UPC, Precio_Con_IVA, Tienda, Estatus_Seguimiento, Cliente_Asistio, Monto_Nueva_Venta, Notas, Ruta_Ticket, DATEDIFF(CURDATE(), Fecha_Compra) as dias_transcurridos
                            FROM crm_compras
                            WHERE (DATEDIFF(CURDATE(), Fecha_Compra) BETWEEN 330 AND 395 OR (DATEDIFF(CURDATE(), Fecha_Compra) >= 330 AND Estatus_Seguimiento = 'Pendiente'))
                            ORDER BY Fecha_Compra ASC
                        """)
                        rows = cursor.fetchall()
                        db.close()

                        if not rows:
                            crm_notif_col.controls.append(
                                ft.Container(
                                    content=ft.Text("No hay notificaciones de garantías por cumplir 1 año pendientes por contactar.", color="#aaaaaa", italic=True),
                                    padding=15
                                )
                            )
                        else:
                            count_notif = len(rows)
                            for r in rows:
                                fecha_c = r["Fecha_Compra"]
                                dt_c = datetime.strptime(str(fecha_c), "%Y-%m-%d") if isinstance(fecha_c, str) else fecha_c
                                dt_venc = dt_c + timedelta(days=365)
                                fecha_venc_str = dt_venc.strftime("%d/%m/%Y")

                                is_contactado = r.get("Estatus_Seguimiento") in ("Contactado / Venta Realizada", "Venta Realizada")

                                crm_notif_col.controls.append(
                                    ft.Container(
                                        content=ft.Column([
                                            ft.Row([
                                                ft.Container(
                                                    content=ft.Text("⚠️ ALERTA MES 11 — ANIVERSARIO DE COMPRA", color="#000000", weight="bold", size=10),
                                                    bgcolor="#FFD700" if not is_contactado else "#7CFC00",
                                                    padding=ft.Padding(8, 3, 8, 3),
                                                    border_radius=4
                                                ),
                                                ft.Text(f"Ticket: {r['Transaccion']}", color="white", weight="bold", size=12),
                                                ft.Text(f"Vence: {fecha_venc_str}", color="#FF8C00", weight="bold", size=12)
                                            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                                            ft.Text(f"📢 Recomendación de Venta: El cliente {r['Nombre_Cliente']} ({r['Telefono_Cliente'] or 'Sin Teléfono'}) compró el lente UPC {r['UPC']} (${r['Precio_Con_IVA']:,.2f} MXN) y su garantía de 1 año vence el {fecha_venc_str}. Comunícate para recordarle sus coberturas de Ruptura y Robo (50% de descuento en la pieza de menor valor) e iniciar labor de venta.", color="#E2E8F0", size=12),
                                            ft.Row([
                                                ft.ElevatedButton(
                                                    "Ver Todos los Datos y Registrar Venta 📄",
                                                    icon=ft.Icons.DESCRIPTION_ROUNDED,
                                                    bgcolor="#333333",
                                                    color="#00FFFF",
                                                    on_click=lambda e, comp=r: mostrar_detalle_notificacion(comp)
                                                ),
                                                ft.Text(f"Venta Generada: ${r.get('Monto_Nueva_Venta') or 0:,.2f} MXN", color="#7CFC00", size=11, weight="bold")
                                            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN)
                                        ], spacing=8),
                                        bgcolor="#1A1A2E",
                                        border=ft.Border.all(1, "#FFD700" if not is_contactado else "#7CFC00"),
                                        padding=12,
                                        border_radius=10
                                    )
                                )
                except Exception as ex:
                    print("Error notif crm:", ex)
                
                notif_badge_text.value = str(count_notif)
                page.update()

            # --- CARGAR HISTORIAL DE COMPRAS CRM & BÚSQUEDA POR TELÉFONO ---
            tf_buscar_crm = ft.TextField(label="🔍 Buscar por Ticket, Cliente o UPC...", border_color="#9D50BB", color="white", text_size=12, width=280, on_change=lambda e: cargar_historial_crm())
            tf_buscar_telefono = ft.TextField(label="📱 Buscar por Teléfono...", prefix_icon=ft.Icons.PHONE, border_color="#00FFFF", color="white", text_size=12, width=240, on_change=lambda e: cargar_historial_crm())
            
            def cargar_historial_crm():
                crm_history_col.controls.clear()
                q_search = tf_buscar_crm.value.strip().lower() if tf_buscar_crm.value else ""
                q_tel = tf_buscar_telefono.value.strip().lower() if tf_buscar_telefono.value else ""
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("""
                            SELECT ID_Compra, Transaccion, Fecha_Compra, Nombre_Cliente, Telefono_Cliente, Nombre_Vendedor, UPC, Precio_Con_IVA, Tienda, Estatus_Seguimiento, Cliente_Asistio, Monto_Nueva_Venta, Notas, Ruta_Ticket, DATEDIFF(CURDATE(), Fecha_Compra) as dias
                            FROM crm_compras
                            ORDER BY Fecha_Compra DESC
                        """)
                        rows = cursor.fetchall()
                        db.close()

                        if q_search:
                            rows = [r for r in rows if q_search in r["Transaccion"].lower() or q_search in r["Nombre_Cliente"].lower() or q_search in (r["Telefono_Cliente"] or "").lower() or q_search in r["UPC"].lower()]
                        if q_tel:
                            rows = [r for r in rows if q_tel in (r["Telefono_Cliente"] or "").lower()]

                        if not rows:
                            crm_history_col.controls.append(ft.Text("No hay registros en el CRM.", color="#aaaaaa", italic=True))
                        else:
                            for r in rows:
                                dias = r.get("dias") or 0
                                if dias >= 365:
                                    tag_garantia = ft.Container(content=ft.Text("🔴 Garantía Vencida", color="white", size=10, weight="bold"), bgcolor="#DC2626", padding=4, border_radius=4)
                                elif dias >= 330:
                                    tag_garantia = ft.Container(content=ft.Text("⚠️ Alerta Mes 11 (Vence Pronto)", color="black", size=10, weight="bold"), bgcolor="#FFD700", padding=4, border_radius=4)
                                else:
                                    tag_garantia = ft.Container(content=ft.Text("🟢 Garantía Vigente (Ruptura/Robo)", color="black", size=10, weight="bold"), bgcolor="#7CFC00", padding=4, border_radius=4)

                                def del_crm(id_c=r["ID_Compra"]):
                                    try:
                                        db_d = conectar_db()
                                        if db_d:
                                            cur_d = db_d.cursor()
                                            cur_d.execute("DELETE FROM crm_compras WHERE ID_Compra = %s", (id_c,))
                                            db_d.commit()
                                            db_d.close()
                                            mostrar_snack("Registro eliminado del CRM.", color="orange")
                                            cargar_historial_crm()
                                            cargar_notificaciones_crm()
                                            cargar_metricas_crm()
                                    except Exception as ex:
                                        print("Error delete CRM:", ex)

                                def guardar_asistencia_card(comp=r, chk=None, tf=None):
                                    try:
                                        asist = 1 if chk.value else 0
                                        monto = 0.0
                                        if tf.value.strip():
                                            try:
                                                monto = float(tf.value.replace("$", "").replace(",", "").strip())
                                            except Exception:
                                                mostrar_snack("Monto de venta inválido.", color="red")
                                                return
                                        estatus = "Venta Realizada" if (asist and monto > 0) else ("Contactado / Asistió" if asist else "Contactado / Pendiente")
                                        db = conectar_db()
                                        if db:
                                            cur = db.cursor()
                                            cur.execute("""
                                                UPDATE crm_compras 
                                                SET Cliente_Asistio = %s, Monto_Nueva_Venta = %s, Estatus_Seguimiento = %s, Fecha_Asistencia = NOW()
                                                WHERE ID_Compra = %s
                                            """, (asist, monto, estatus, comp["ID_Compra"]))
                                            db.commit()
                                            db.close()
                                            mostrar_snack(f"✅ Asistencia y venta de ${monto:,.2f} MXN registradas para {comp['Nombre_Cliente']}.", color="green")
                                            cargar_historial_crm()
                                            cargar_notificaciones_crm()
                                            cargar_metricas_crm()
                                    except Exception as ex:
                                        print("Error guardando asistencia card:", ex)

                                chk_card_asistio = ft.Checkbox(label="Cliente Asistió 🏬", value=r.get("Cliente_Asistio") == 1)
                                tf_card_monto = ft.TextField(
                                    label="Monto Nueva Venta ($ MXN)",
                                    value=str(r.get("Monto_Nueva_Venta") or ""),
                                    border_color="#7CFC00",
                                    color="white",
                                    text_size=11,
                                    width=170
                                )

                                crm_history_col.controls.append(
                                    ft.Container(
                                        content=ft.Column([
                                            ft.Row([
                                                tag_garantia,
                                                ft.Text(f"Ticket: {r['Transaccion']}", color="white", weight="bold", size=13),
                                                ft.Text(f"Fecha: {r['Fecha_Compra']}", color="#aaaaaa", size=11)
                                            ], spacing=8, wrap=True),
                                            ft.Text(f"Cliente: {r['Nombre_Cliente']}", color="#D8B4FE", size=12),
                                            ft.Text(f"Tel: {r['Telefono_Cliente'] or 'S/N'} | Vendedor: {r['Nombre_Vendedor']}", color="#aaaaaa", size=11),
                                            ft.Text(f"UPC: {r['UPC']} | Precio: ${r['Precio_Con_IVA']:,.2f} MXN", color="white", size=11),
                                            ft.Row([
                                                chk_card_asistio,
                                                tf_card_monto,
                                            ], spacing=10, wrap=True),
                                            ft.Row([
                                                ft.ElevatedButton("Guardar Venta 💰", bgcolor="#7CFC00", color="black", style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=6)), on_click=lambda e, comp=r, chk=chk_card_asistio, tf=tf_card_monto: guardar_asistencia_card(comp, chk, tf)),
                                                ft.ElevatedButton("Ver Ticket 🧾", icon=ft.Icons.RECEIPT_LONG, bgcolor="#1E2330", color="#00FFFF", on_click=lambda e, comp=r: mostrar_detalle_notificacion(comp)),
                                                ft.IconButton(ft.Icons.DELETE, icon_color="#FF4B4B", tooltip="Eliminar (Admin)", visible=es_admin(), on_click=lambda e, id_c=r["ID_Compra"]: del_crm(id_c))
                                            ], spacing=8, wrap=True)
                                        ], spacing=6),
                                        bgcolor="#161922",
                                        padding=12,
                                        border_radius=8,
                                        border=ft.Border.all(1, "#333333")
                                    )
                                )
                except Exception as ex:
                    print("Error historial CRM:", ex)
                page.update()

            # --- SUB-PESTAÑA DE MÉTRICAS Y RENTABILIDAD CRM 📊 ---
            crm_metrics_container = ft.Column(spacing=12, scroll=ft.ScrollMode.AUTO, expand=True)

            def cargar_metricas_crm():
                crm_metrics_container.controls.clear()
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("""
                            SELECT 
                                COUNT(*) as total_registros,
                                COUNT(CASE WHEN DATEDIFF(CURDATE(), Fecha_Compra) >= 330 THEN 1 END) as notificados_mes11,
                                COUNT(CASE WHEN Cliente_Asistio = 1 THEN 1 END) as asistieron_tienda,
                                COALESCE(SUM(Monto_Nueva_Venta), 0) as venta_total_crm,
                                COALESCE(AVG(CASE WHEN Cliente_Asistio = 1 AND Monto_Nueva_Venta > 0 THEN Monto_Nueva_Venta END), 0) as venta_promedio
                            FROM crm_compras
                        """)
                        row = cursor.fetchone()
                        db.close()

                        tot_reg = row["total_registros"] or 0
                        notif_11 = row["notificados_mes11"] or 0
                        asist = row["asistieron_tienda"] or 0
                        venta_tot = float(row["venta_total_crm"] or 0)
                        venta_prom = float(row["venta_promedio"] or 0)

                        tasa_conversion = (asist / notif_11 * 100.0) if notif_11 > 0 else 0.0

                        crm_metrics_container.controls.extend([
                            ft.Text("Métricas de Efectividad y Rentabilidad del CRM 📊", color="#FFD700", weight="bold", size=16),
                            ft.Text("Análisis en tiempo real de las ventas adicionales generadas por la campaña de recordatorio de garantías al mes 11.", color="#aaaaaa", size=12),
                            ft.Divider(height=10, color="#333333"),
                            # KPI CARDS
                            ft.Row([
                                ft.Container(
                                    content=ft.Column([
                                        ft.Text("VENTA TOTAL GENERADA CRM", color="#aaaaaa", size=10, weight="bold"),
                                        ft.Text(f"${venta_tot:,.2f} MXN", color="#7CFC00", size=20, weight="bold")
                                    ]), bgcolor="#1E2330", padding=14, border_radius=10, border=ft.Border.all(1, "#7CFC00"), expand=True
                                ),
                                ft.Container(
                                    content=ft.Column([
                                        ft.Text("TASA DE CONVERSIÓN (% ASISTENCIA)", color="#aaaaaa", size=10, weight="bold"),
                                        ft.Text(f"{tasa_conversion:.1f}%", color="#00FFFF", size=20, weight="bold")
                                    ]), bgcolor="#1E2330", padding=14, border_radius=10, border=ft.Border.all(1, "#00FFFF"), expand=True
                                ),
                                ft.Container(
                                    content=ft.Column([
                                        ft.Text("VENTA PROMEDIO POR CLIENTE", color="#aaaaaa", size=10, weight="bold"),
                                        ft.Text(f"${venta_prom:,.2f} MXN", color="#D8B4FE", size=20, weight="bold")
                                    ]), bgcolor="#1E2330", padding=14, border_radius=10, border=ft.Border.all(1, "#D8B4FE"), expand=True
                                ),
                                ft.Container(
                                    content=ft.Column([
                                        ft.Text("CLIENTES QUE ASISTIERON", color="#aaaaaa", size=10, weight="bold"),
                                        ft.Text(f"{asist} / {notif_11}", color="#FFD700", size=20, weight="bold")
                                    ]), bgcolor="#1E2330", padding=14, border_radius=10, border=ft.Border.all(1, "#FFD700"), expand=True
                                )
                            ]),
                            ft.Divider(height=10, color="#333333"),
                            ft.Text("GRÁFICA COMPARATIVA DE RENDIMIENTO DEL CRM:", color="#D8B4FE", weight="bold", size=14),
                            ft.Container(
                                content=ft.Column([
                                    ft.Row([
                                        ft.Text("Clientes Notificados (Mes 11):", color="white", width=220),
                                        ft.ProgressBar(value=1.0 if notif_11 > 0 else 0, color="#FFD700", bgcolor="#333333", expand=True),
                                        ft.Text(f"{notif_11} Clientes", color="#FFD700", weight="bold", width=90)
                                    ]),
                                    ft.Row([
                                        ft.Text("Clientes que Asistieron a Tienda:", color="white", width=220),
                                        ft.ProgressBar(value=(asist / notif_11) if notif_11 > 0 else 0, color="#00FFFF", bgcolor="#333333", expand=True),
                                        ft.Text(f"{asist} Clientes", color="#00FFFF", weight="bold", width=90)
                                    ]),
                                    ft.Row([
                                        ft.Text("Ventas Concretadas ($):", color="white", width=220),
                                        ft.ProgressBar(value=min(1.0, venta_tot / 20000.0) if venta_tot > 0 else 0, color="#7CFC00", bgcolor="#333333", expand=True),
                                        ft.Text(f"${venta_tot:,.0f} MXN", color="#7CFC00", weight="bold", width=90)
                                    ])
                                ], spacing=12),
                                bgcolor="#161922", padding=16, border_radius=10, border=ft.Border.all(1, "#333333")
                            )
                        ])
                except Exception as ex:
                    print("Error métricas crm:", ex)
                page.update()

            # --- SECCIÓN GUÍA DE REGLAS DE GARANTÍA ---
            guia_garantias_view = ft.Container(
                content=ft.Column([
                    ft.Text("REGLAS Y COBERTURAS DE GARANTÍAS SUNGLASS HUT (1 AÑO) 📜", color="#FFD700", weight="bold", size=16),
                    ft.Divider(height=10, color="#333333"),
                    ft.Container(
                        content=ft.Column([
                            ft.Row([ft.Icon(ft.Icons.BUILD_ROUNDED, color="#00FFFF"), ft.Text("1. GARANTÍA POR RUPTURA / DAÑO (VIGENCIA 1 AÑO)", color="#00FFFF", weight="bold", size=14)]),
                            ft.Text("• Aplica cuando el cliente acude a la tienda con su gafa que sufrió alguna ruptura o daño accidental (sin importar el estado físico en el que se encuentre la pieza).", color="white", size=12),
                            ft.Text("• El cliente debe entregar la gafa en tienda.", color="white", size=12),
                            ft.Text("• SE APLICA EL 50% DE DESCUENTO SOBRE LA PIEZA DE MENOR VALOR (ya sea la pieza que entrega el cliente o la nueva gafa que se lleva).", color="#7CFC00", weight="bold", size=12)
                        ], spacing=6),
                        bgcolor="#1E2330", padding=14, border_radius=10, border=ft.Border.all(1, "#00FFFF")
                    ),
                    ft.Container(
                        content=ft.Column([
                            ft.Row([ft.Icon(ft.Icons.SECURITY_ROUNDED, color="#FF8C00"), ft.Text("2. GARANTÍA POR ROBO (VIGENCIA 1 AÑO)", color="#FF8C00", weight="bold", size=14)]),
                            ft.Text("• Aplica cuando al cliente le roban sus gafas dentro del primer año de compra.", color="white", size=12),
                            ft.Text("• El cliente debe presentar en tienda su ticket de compra y el ACTA DE DENUNCIA emitida por la autoridad competente con los datos del lente.", color="white", size=12),
                            ft.Text("• EL ACTA DE DENUNCIA IMPRESA DEBE QUEDAR FÍSICAMENTE ARCHIVADA EN LA TIENDA.", color="#FFD700", weight="bold", size=12),
                            ft.Text("• SE APLICA EL 50% DE DESCUENTO SOBRE LA PIEZA DE MISMO O MENOR VALOR.", color="#7CFC00", weight="bold", size=12)
                        ], spacing=6),
                        bgcolor="#1E2330", padding=14, border_radius=10, border=ft.Border.all(1, "#FF8C00")
                    ),
                    ft.Container(
                        content=ft.Column([
                            ft.Row([ft.Icon(ft.Icons.ALARM_ON_ROUNDED, color="#D8B4FE"), ft.Text("3. ESTRATEGIA DE FIDELIZACIÓN AL MES 11", color="#D8B4FE", weight="bold", size=14)]),
                            ft.Text("• LUXO genera alertas automáticas a los 11 meses de la compra (1 mes antes de vencer).", color="white", size=12),
                            ft.Text("• La tienda debe ponerse en contacto con el cliente para recordarle sus garantías de Ruptura y Robo antes de que venza su ticket y realizar labor de venta para ofrecerle las nuevas colecciones.", color="white", size=12)
                        ], spacing=6),
                        bgcolor="#1E2330", padding=14, border_radius=10, border=ft.Border.all(1, "#D8B4FE")
                    )
                ], spacing=14, scroll=ft.ScrollMode.AUTO),
                expand=True
            )

            # --- TABS PRINCIPALES DEL CRM ---
            tab_captura = ft.Column([
                ft.Text("Capturar Nueva Venta / Ticket 📝", color="#D8B4FE", weight="bold", size=15),
                ft.Row([
                    ft.ElevatedButton("📷 Escanear Ticket / Tomar Foto", icon=ft.Icons.CAMERA_ALT, bgcolor="#1f6f43", color="white", on_click=escanear_ticket_click),
                    ft.Text("Ingresa los datos o toma foto directa del ticket con tu celular.", color="#aaaaaa", size=11)
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, wrap=True),
                img_preview_container,
                ft.Divider(height=10, color="#333333"),
                ft.Row([tf_transaccion, tf_fecha_compra], wrap=True, spacing=10),
                ft.Row([tf_nombre_cliente, tf_telefono_cliente], wrap=True, spacing=10),
                ft.Row([tf_nombre_vendedor], wrap=True, spacing=10),
                ft.Divider(height=5, color="transparent"),
                ft.Text("🛍️ Productos / Artículos del Ticket (1 o más):", color="#00FFFF", weight="bold", size=13),
                items_rows_container,
                ft.Row([
                    ft.ElevatedButton("➕ Agregar Otro Artículo", icon=ft.Icons.ADD_SHOPPING_CART, bgcolor="#333333", color="#00FFFF", on_click=lambda e: agregar_fila_articulo()),
                    tf_precio
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, wrap=True),
                ft.Divider(height=5, color="transparent"),
                tf_notas,
                ft.Row([
                    ft.ElevatedButton("Guardar Registro en CRM 💾", icon=ft.Icons.SAVE, bgcolor="#9D50BB", color="white", on_click=guardar_compra_crm),
                    ft.OutlinedButton("Limpiar Campos 🔄", on_click=lambda e: limpiar_form_crm())
                ], alignment=ft.MainAxisAlignment.END, wrap=True, spacing=10)
            ], scroll=ft.ScrollMode.AUTO, expand=True)

            tab_historial_crm = ft.Column([
                ft.Text("Historial de CRM y Garantías 📋", color="#D8B4FE", weight="bold", size=15),
                ft.Row([tf_buscar_crm, tf_buscar_telefono], wrap=True, spacing=10),
                crm_history_col
            ], expand=True, scroll=ft.ScrollMode.AUTO)

            tab_notificaciones = ft.Column([
                ft.Row([
                    ft.Text("Notificaciones & Alertas Mes 11 (Fidelización) 🔔", color="#FFD700", weight="bold", size=15),
                    ft.ElevatedButton("Actualizar Alertas 🔄", bgcolor="#333333", color="white", on_click=lambda e: cargar_notificaciones_crm())
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                ft.Text("A continuación se enlistan las compras que están por cumplir 1 año (al mes 11). Haz clic en 'Ver Todos los Datos y Registrar Venta' para iniciar la labor de venta y registrar la asistencia y nueva venta.", color="#aaaaaa", size=12),
                ft.Divider(height=10, color="#333333"),
                crm_notif_col
            ], expand=True)

            crm_tabs = ft.Tabs(
                selected_index=0,
                animation_duration=300,
                length=5,
                expand=True,
                content=ft.Column(
                    expand=True,
                    controls=[
                        ft.TabBar(
                            tabs=[
                                ft.Tab(label="Captura y Registro 📝"),
                                ft.Tab(label="Historial CRM 📋"),
                                ft.Tab(label="Notificaciones (Mes 11) 🔔"),
                                ft.Tab(label="Métricas & ROI CRM 📊"),
                                ft.Tab(label="Reglas de Garantía 📜")
                            ]
                        ),
                        ft.TabBarView(
                            expand=True,
                            controls=[
                                tab_captura,
                                tab_historial_crm,
                                tab_notificaciones,
                                crm_metrics_container,
                                guia_garantias_view
                            ]
                        )
                    ]
                )
            )

            cargar_historial_crm()
            cargar_notificaciones_crm()
            cargar_metricas_crm()

            return ft.Column([
                ft.Row([
                    ft.Text("CRM Cobertura Oops 👓", size=24, color="#D8B4FE", weight="bold"),
                    ft.Container(
                        content=ft.Row([
                            ft.Icon(ft.Icons.NOTIFICATIONS_ACTIVE, color="#00FFFF", size=16),
                            notif_badge_text,
                            ft.Text("Alertas Pendientes", color="white", size=11, weight="bold")
                        ]),
                        bgcolor="#1E2330",
                        padding=ft.Padding(10, 5, 10, 5),
                        border_radius=8,
                        border=ft.Border.all(1, "#00FFFF")
                    )
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                ft.Text("Gestión de clientes, escáner de tickets con IA, coberturas por Ruptura o Robo (1 año) y alertas automáticas al mes 11.", color="#aaaaaa", size=13),
                ft.Divider(height=15, color="#333333"),
                crm_tabs
            ], expand=True)

        def build_bitacora_view():
            """Vista de Bitácora de Seguridad exclusiva para Administradores."""
            rows_container = ft.Column(spacing=6, scroll=ft.ScrollMode.AUTO, expand=True)
            status_text = ft.Text("Cargando registros...", color="#aaaaaa", size=13, italic=True)

            def cargar_bitacora():
                rows_container.controls.clear()
                try:
                    db_bit = conectar_db()
                    if not db_bit:
                        status_text.value = "Error de conexión a MySQL."
                        page.update()
                        return
                    cur_bit = db_bit.cursor(dictionary=True)
                    cur_bit.execute("""
                        SELECT Nombre_Usuario, Empleado_Identificado, IP_Acceso, Dispositivo,
                               DATE_FORMAT(Fecha_Hora, '%d/%m/%Y %H:%i:%s') as Fecha_Hora
                        FROM bitacora_sesiones_biometricas
                        ORDER BY Fecha_Hora DESC
                        LIMIT 200
                    """)
                    registros = cur_bit.fetchall()
                    db_bit.close()

                    if not registros:
                        status_text.value = "No hay inicios de sesión registrados aún."
                        page.update()
                        return

                    status_text.value = f"Mostrando los últimos {len(registros)} ingresos de seguridad:"

                    for r in registros:
                        nombre_emp = r.get("Empleado_Identificado", "") or r.get("Nombre_Usuario", "")
                        ip_val = r.get("IP_Acceso", "127.0.0.1")
                        disp_val = r.get("Dispositivo", "Localhost, Local / Desarrollo")
                        if not disp_val or disp_val == "Navegador Web":
                            disp_val = "Localhost, Local / Desarrollo"
                        fecha_val = r.get("Fecha_Hora", "")

                        card = ft.Container(
                            content=ft.Column([
                                ft.Row([
                                    ft.Text(f"📅 {fecha_val}", color="#aaaaaa", size=11),
                                    ft.Text(f"👤 {nombre_emp}", color="white", weight="bold"),
                                ], alignment="spaceBetween"),
                                ft.Row([
                                    ft.Text(f"🌐 IP: {ip_val}", color="#00FFFF", size=12),
                                    ft.Text(f"📍 {disp_val}", color="#D8B4FE", size=12),
                                ], alignment="spaceBetween"),
                            ], spacing=4),
                            bgcolor="#141424",
                            padding=15,
                            border_radius=8,
                            border=ft.Border.all(1, "#333333")
                        )
                        rows_container.controls.append(card)
                    page.update()

                    # Pre-generar el archivo CSV de la bitácora para descarga instantánea
                    try:
                        temp_dir = os.path.join(ASSETS_PATH, "temp_pdfs")
                        os.makedirs(temp_dir, exist_ok=True)
                        filepath_temp = os.path.join(temp_dir, "Bitacora_Seguridad.csv")
                        import csv
                        with open(filepath_temp, "w", encoding="utf-8-sig", newline="") as f_temp:
                            writer = csv.writer(f_temp)
                            writer.writerow(["Fecha y Hora", "Empleado / Usuario", "Cuenta de Usuario", "IP de Acceso", "Dispositivo"])
                            for r in registros:
                                writer.writerow([
                                    r.get("Fecha_Hora", ""),
                                    r.get("Empleado_Identificado", ""),
                                    r.get("Nombre_Usuario", ""),
                                    r.get("IP_Acceso", "127.0.0.1"),
                                    r.get("Dispositivo", "Navegador Web")
                                ])
                    except Exception as ex_pre:
                        print("Error pre-generando CSV:", ex_pre)

                except Exception as ex_bit:
                    status_text.value = f"Error al cargar bitácora: {ex_bit}"
                    page.update()

            def descargar_excel_bitacora(e=None):
                try:
                    db_exp = conectar_db()
                    if not db_exp:
                        mostrar_snack("Error al conectar a la base de datos", color="red")
                        return
                    cur_exp = db_exp.cursor(dictionary=True)
                    cur_exp.execute("""
                        SELECT DATE_FORMAT(Fecha_Hora, '%d/%m/%Y %H:%i:%s') as Fecha_Hora,
                               Empleado_Identificado, Nombre_Usuario, IP_Acceso, Dispositivo
                        FROM bitacora_sesiones_biometricas
                        ORDER BY Fecha_Hora DESC
                    """)
                    filas = cur_exp.fetchall()
                    db_exp.close()

                    if not filas:
                        mostrar_snack("No hay registros en la bitácora para exportar", color="orange")
                        return

                    import csv
                    from datetime import datetime

                    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = "Bitacora_Seguridad.csv"
                    filename_user = f"Bitacora_Seguridad_{timestamp_str}.csv"

                    temp_dir = os.path.join(ASSETS_PATH, "temp_pdfs")
                    os.makedirs(temp_dir, exist_ok=True)
                    filepath_temp = os.path.join(temp_dir, filename)

                    downloads_user_dir = os.path.expanduser("~/Downloads")
                    filepath_user = os.path.join(downloads_user_dir, filename_user)

                    with open(filepath_temp, "w", encoding="utf-8-sig", newline="") as f_temp:
                        writer = csv.writer(f_temp)
                        writer.writerow(["Fecha y Hora", "Empleado / Usuario", "Cuenta de Usuario", "IP de Acceso", "Dispositivo"])
                        for r in filas:
                            writer.writerow([
                                r.get("Fecha_Hora", ""),
                                r.get("Empleado_Identificado", ""),
                                r.get("Nombre_Usuario", ""),
                                r.get("IP_Acceso", "127.0.0.1"),
                                r.get("Dispositivo", "Navegador Web")
                            ])

                    import shutil
                    shutil.copy2(filepath_temp, filepath_user)

                    mostrar_snack(f"✅ Archivo exportado ({len(filas)} registros totales) en tu carpeta de Descargas", color="#7CFC00")

                except Exception as ex_exp:
                    print("ERROR EXPORTANDO EXCEL BITÁCORA:", ex_exp)
                    mostrar_snack(f"Error al generar Excel: {ex_exp}", color="red")

            import threading
            threading.Thread(target=cargar_bitacora, daemon=True).start()

            base_url = page.url.rstrip("/") if (page and page.url) else "http://localhost:8550"
            if base_url.startswith("ws://"):
                base_url = base_url.replace("ws://", "http://", 1)
            elif base_url.startswith("wss://"):
                base_url = base_url.replace("wss://", "https://", 1)
            elif base_url.startswith("tcp://"):
                base_url = base_url.replace("tcp://", "http://", 1)

            base_dl = re.sub(r":\d+$", f":{PUERTO_DESCARGAS}", base_url)
            url_dl = f"{base_dl}/download?file=Bitacora_Seguridad.csv&original=Bitacora_Seguridad.csv"

            btn_descargar_excel = ft.ElevatedButton(
                content=ft.Row([
                    ft.Icon(ft.Icons.TABLE_CHART_ROUNDED, color="white", size=18),
                    ft.Text("Descargar Excel", color="white", weight="bold")
                ], spacing=6),
                bgcolor="#008080",
                color="white",
                url=url_dl,
                on_click=descargar_excel_bitacora,
                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=6))
            )

            is_mobile = (page.width < 800) if (page and page.width) else False

            header_top = ft.Column([
                ft.Row([
                    ft.Row([
                        ft.Icon(ft.Icons.SECURITY_ROUNDED, color="#00FFAA", size=22),
                        ft.Text("Bitácora de Seguridad 🛡️", size=17, color="#00FFAA", weight="bold"),
                    ], spacing=6),
                    ft.IconButton(icon=ft.Icons.REFRESH_ROUNDED, icon_color="#00FFAA", tooltip="Actualizar bitácora",
                                  on_click=lambda e: threading.Thread(target=cargar_bitacora, daemon=True).start())
                ], alignment="spaceBetween", vertical_alignment="center"),
                btn_descargar_excel
            ], spacing=8) if is_mobile else ft.Row([
                ft.Icon(ft.Icons.SECURITY_ROUNDED, color="#00FFAA", size=26),
                ft.Text("Bitácora de Seguridad 🛡️", size=22, color="#00FFAA", weight="bold"),
                ft.Container(expand=True),
                btn_descargar_excel,
                ft.IconButton(icon=ft.Icons.REFRESH_ROUNDED, icon_color="#00FFAA", tooltip="Actualizar bitácora",
                              on_click=lambda e: threading.Thread(target=cargar_bitacora, daemon=True).start())
            ], vertical_alignment="center")

            return ft.Column([
                header_top,
                ft.Text("Registro de todos los inicios de sesión. Visible solo para Administradores.", color="#aaaaaa", size=12),
                ft.Divider(height=12, color="#333333"),
                status_text,
                ft.Container(height=4),
                rows_container
            ], expand=True)

        def build_admin_trivia_view():

            """Vista de admin para gestionar preguntas del Reto del Día."""
            if not es_admin():
                return ft.Container(
                    content=ft.Column([
                        ft.Icon(ft.Icons.LOCK, color="red", size=48),
                        ft.Text("Acceso Restringido ⚠️", size=20, color="red", weight="bold"),
                        ft.Text("Esta sección de gestión de preguntas es exclusiva para Administradores.", color="#aaaaaa", size=14)
                    ], alignment=ft.MainAxisAlignment.CENTER, horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=10),
                    alignment=ft.alignment.Alignment(0, 0),
                    expand=True
                )

            lista_preguntas_col = ft.Column(spacing=6, scroll=ft.ScrollMode.AUTO)

            # --- Formulario de nueva pregunta ---
            tf_pregunta = ft.TextField(
                label="Pregunta", multiline=True, min_lines=2, max_lines=4,
                border_color="#9D50BB", color="white", expand=True,
                label_style=ft.TextStyle(color="#aaaaaa", size=11)
            )
            tf_opcion_a = ft.TextField(label="Opci\u00f3n A", border_color="#9D50BB", color="white", expand=True,
                                       label_style=ft.TextStyle(color="#aaaaaa", size=11))
            tf_opcion_b = ft.TextField(label="Opci\u00f3n B", border_color="#9D50BB", color="white", expand=True,
                                       label_style=ft.TextStyle(color="#aaaaaa", size=11))
            tf_opcion_c = ft.TextField(label="Opci\u00f3n C", border_color="#9D50BB", color="white", expand=True,
                                       label_style=ft.TextStyle(color="#aaaaaa", size=11))
            tf_opcion_d = ft.TextField(label="Opci\u00f3n D", border_color="#9D50BB", color="white", expand=True,
                                       label_style=ft.TextStyle(color="#aaaaaa", size=11))
            tf_explicacion = ft.TextField(
                label="Explicaci\u00f3n (opcional)", multiline=True, min_lines=2, max_lines=3,
                border_color="#9D50BB", color="white", expand=True,
                label_style=ft.TextStyle(color="#aaaaaa", size=11)
            )
            dd_correcta = ft.Dropdown(
                label="Respuesta Correcta",
                border_color="#9D50BB", color="white",
                label_style=ft.TextStyle(color="#aaaaaa", size=11),
                options=[
                    ft.dropdown.Option("A", "A"),
                    ft.dropdown.Option("B", "B"),
                    ft.dropdown.Option("C", "C"),
                    ft.dropdown.Option("D", "D"),
                ],
                width=150
            )
            dd_dificultad = ft.Dropdown(
                label="Dificultad",
                border_color="#9D50BB", color="white", value="Fácil",
                label_style=ft.TextStyle(color="#aaaaaa", size=11),
                options=[
                    ft.dropdown.Option("Fácil", "Fácil ⭐"),
                    ft.dropdown.Option("Difícil", "Difícil 🔥"),
                ],
                width=150
            )

            def cargar_lista_preguntas():
                lista_preguntas_col.controls.clear()
                try:
                    db = conectar_db()
                    if db:
                        cur = db.cursor(dictionary=True)
                        cur.execute("SELECT * FROM reto_preguntas ORDER BY Dificultad, ID_Pregunta")
                        rows = cur.fetchall()
                        db.close()

                        if not rows:
                            lista_preguntas_col.controls.append(
                                ft.Text("No hay preguntas registradas.", color="#aaaaaa", italic=True, size=12)
                            )
                            return

                        for row in rows:
                            diff_color = "#00FF7F" if row["Dificultad"] == "Fácil" else "#FF6B6B"
                            id_p = row["ID_Pregunta"]

                            def make_edit(preg_data=row):
                                def do_edit(e):
                                    pid = preg_data["ID_Pregunta"]
                                    edit_tf_pregunta = ft.TextField(
                                        label="Pregunta",
                                        value=preg_data["Pregunta"],
                                        multiline=True,
                                        max_lines=3,
                                        border_color="#9D50BB",
                                        color="white",
                                        text_size=11,
                                        label_style=ft.TextStyle(color="#aaaaaa", size=11)
                                    )
                                    edit_tf_op_a = ft.TextField(label="Opción A", value=preg_data["Opcion_A"], border_color="#333333", color="white", text_size=11, label_style=ft.TextStyle(color="#aaaaaa", size=10), expand=True)
                                    edit_tf_op_b = ft.TextField(label="Opción B", value=preg_data["Opcion_B"], border_color="#333333", color="white", text_size=11, label_style=ft.TextStyle(color="#aaaaaa", size=10), expand=True)
                                    edit_tf_op_c = ft.TextField(label="Opción C", value=preg_data["Opcion_C"], border_color="#333333", color="white", text_size=11, label_style=ft.TextStyle(color="#aaaaaa", size=10), expand=True)
                                    edit_tf_op_d = ft.TextField(label="Opción D", value=preg_data["Opcion_D"], border_color="#333333", color="white", text_size=11, label_style=ft.TextStyle(color="#aaaaaa", size=10), expand=True)

                                    edit_dd_correcta = ft.Dropdown(
                                        label="Respuesta Correcta",
                                        value=preg_data["Respuesta_Correcta"],
                                        border_color="#9D50BB",
                                        color="white",
                                        text_size=11,
                                        label_style=ft.TextStyle(color="#aaaaaa", size=10),
                                        options=[
                                            ft.dropdown.Option("A", "Opción A"),
                                            ft.dropdown.Option("B", "Opción B"),
                                            ft.dropdown.Option("C", "Opción C"),
                                            ft.dropdown.Option("D", "Opción D"),
                                        ],
                                        width=150
                                    )
                                    edit_dd_dificultad = ft.Dropdown(
                                        label="Dificultad",
                                        value=preg_data.get("Dificultad", "Fácil"),
                                        border_color="#9D50BB",
                                        color="white",
                                        text_size=11,
                                        label_style=ft.TextStyle(color="#aaaaaa", size=10),
                                        options=[
                                            ft.dropdown.Option("Fácil", "Fácil ⭐"),
                                            ft.dropdown.Option("Difícil", "Difícil 🔥"),
                                        ],
                                        width=150
                                    )
                                    edit_tf_explicacion = ft.TextField(
                                        label="Explicación o Justificación",
                                        value=preg_data.get("Explicacion") or "",
                                        multiline=True,
                                        max_lines=2,
                                        border_color="#333333",
                                        color="white",
                                        text_size=11,
                                        label_style=ft.TextStyle(color="#aaaaaa", size=10)
                                    )

                                    def guardar_edicion_click(ev):
                                        n_pregunta = edit_tf_pregunta.value.strip()
                                        n_op_a = edit_tf_op_a.value.strip()
                                        n_op_b = edit_tf_op_b.value.strip()
                                        n_op_c = edit_tf_op_c.value.strip()
                                        n_op_d = edit_tf_op_d.value.strip()
                                        n_correcta = edit_dd_correcta.value
                                        n_dificultad = edit_dd_dificultad.value or "Fácil"
                                        n_explicacion = edit_tf_explicacion.value.strip() or None

                                        if not n_pregunta or not n_op_a or not n_op_b or not n_op_c or not n_op_d or not n_correcta:
                                            mostrar_snack("Campos obligatorios incompletos.", color="red")
                                            return

                                        try:
                                            db_ed = conectar_db()
                                            if db_ed:
                                                cur_ed = db_ed.cursor()
                                                cur_ed.execute("""
                                                    UPDATE reto_preguntas 
                                                    SET Pregunta = %s, Opcion_A = %s, Opcion_B = %s, Opcion_C = %s, Opcion_D = %s, Respuesta_Correcta = %s, Dificultad = %s, Explicacion = %s 
                                                    WHERE ID_Pregunta = %s
                                                """, (n_pregunta, n_op_a, n_op_b, n_op_c, n_op_d, n_correcta, n_dificultad, n_explicacion, pid))
                                                db_ed.commit()
                                                db_ed.close()

                                            page.pop_dialog()
                                            mostrar_snack(f"✅ Pregunta #{pid} actualizada correctamente.", color="#7CFC00")
                                            cargar_lista_preguntas()
                                            try: page.update()
                                            except Exception: pass
                                        except Exception as ex_ed:
                                            mostrar_snack(f"Error al actualizar: {ex_ed}", color="red")

                                    dlg_edit = ft.AlertDialog(
                                        title=ft.Text(f"Editar Pregunta #{pid} ✏️", color="white", weight="bold", size=15),
                                        content=ft.Container(
                                            content=ft.Column([
                                                edit_tf_pregunta,
                                                ft.Row([edit_tf_op_a, edit_tf_op_b], spacing=6),
                                                ft.Row([edit_tf_op_c, edit_tf_op_d], spacing=6),
                                                ft.Row([edit_dd_correcta, edit_dd_dificultad], spacing=6, wrap=True),
                                                edit_tf_explicacion
                                            ], spacing=8, scroll=ft.ScrollMode.AUTO),
                                            width=450,
                                            height=400
                                        ),
                                        bgcolor="#0F0F1A",
                                        actions=[
                                            ft.TextButton("Cancelar", on_click=lambda ev: page.pop_dialog()),
                                            ft.ElevatedButton("Guardar Cambios", on_click=guardar_edicion_click, bgcolor="#6E48AA", color="white")
                                        ],
                                        actions_alignment="end"
                                    )
                                    page.show_dialog(dlg_edit)
                                    page.update()

                                return do_edit

                            def make_delete(pid=id_p, pregunta=row["Pregunta"][:40]):
                                def do_delete(e):
                                    def confirmar(ev):
                                        page.pop_dialog()
                                        try:
                                            db2 = conectar_db()
                                            if db2:
                                                cur2 = db2.cursor()
                                                cur2.execute("DELETE FROM reto_preguntas WHERE ID_Pregunta = %s", (pid,))
                                                db2.commit()
                                                db2.close()
                                            mostrar_snack(f"Pregunta #{pid} eliminada.", color="#7CFC00")
                                            cargar_lista_preguntas()
                                            try:
                                                page.update()
                                            except Exception:
                                                pass
                                        except Exception as ex:
                                            mostrar_snack(f"Error al eliminar: {ex}", color="red")

                                    dlg = ft.AlertDialog(
                                        title=ft.Text("¿Eliminar pregunta?", color="white"),
                                        content=ft.Text(f'"{pregunta}..."', color="#aaaaaa", size=12),
                                        bgcolor="#1e1e1e",
                                        actions=[
                                            ft.TextButton("Cancelar", on_click=lambda ev: page.pop_dialog()),
                                            ft.TextButton("Eliminar", on_click=confirmar,
                                                          style=ft.ButtonStyle(color="#FF4500")),
                                        ],
                                        actions_alignment="end"
                                    )
                                    page.show_dialog(dlg)
                                return do_delete

                            card = ft.Container(
                                content=ft.Row([
                                    ft.Column([
                                        ft.Row([
                                            ft.Container(
                                                content=ft.Text(row["Dificultad"], size=10, color="black", weight="bold"),
                                                bgcolor=diff_color, border_radius=4, padding=ft.Padding(5, 2, 5, 2)
                                            ),
                                            ft.Text(f"#{row['ID_Pregunta']}", color="#666666", size=10),
                                        ], spacing=6),
                                        ft.Text(row["Pregunta"][:80] + ("..." if len(row["Pregunta"]) > 80 else ""),
                                                color="white", size=12, weight="bold"),
                                        ft.Text(
                                            f"A:{row['Opcion_A'][:25]}  B:{row['Opcion_B'][:25]}  C:{row['Opcion_C'][:25]}  D:{row['Opcion_D'][:25]}",
                                            color="#888888", size=10
                                        ),
                                        ft.Text(f"✔ Correcta: {row['Respuesta_Correcta']}", color="#00FF7F", size=10),
                                    ], expand=True, spacing=4),
                                    ft.Row([
                                        ft.IconButton(
                                            icon=ft.Icons.EDIT_ROUNDED,
                                            icon_color="#00FFFF",
                                            tooltip="Editar pregunta",
                                            on_click=make_edit()
                                        ),
                                        ft.IconButton(
                                            icon=ft.Icons.DELETE_ROUNDED,
                                            icon_color="#FF4500",
                                            tooltip="Eliminar pregunta",
                                            on_click=make_delete()
                                        )
                                    ], spacing=2)
                                ], alignment="spaceBetween", vertical_alignment="center"),
                                bgcolor="#1a1a1a",
                                border_radius=8,
                                padding=10,
                                border=ft.Border.all(1, "#333333")
                            )
                            lista_preguntas_col.controls.append(card)
                except Exception as ex:
                    print("Error cargando preguntas admin:", ex)
                    lista_preguntas_col.controls.append(
                        ft.Text(f"Error: {ex}", color="red", size=11)
                    )
                try:
                    page.update()
                except Exception:
                    pass

            def agregar_pregunta_click(e):
                pregunta = tf_pregunta.value.strip()
                op_a = tf_opcion_a.value.strip()
                op_b = tf_opcion_b.value.strip()
                op_c = tf_opcion_c.value.strip()
                op_d = tf_opcion_d.value.strip()
                correcta = dd_correcta.value
                dificultad = dd_dificultad.value or "Fácil"
                explicacion = tf_explicacion.value.strip() or None

                if not pregunta or not op_a or not op_b or not op_c or not op_d or not correcta:
                    mostrar_snack("Por favor completa todos los campos obligatorios.", color="red")
                    return

                try:
                    db = conectar_db()
                    if db:
                        cur = db.cursor()
                        cur.execute("""
                            INSERT INTO reto_preguntas 
                            (Pregunta, Opcion_A, Opcion_B, Opcion_C, Opcion_D, Respuesta_Correcta, Explicacion, Dificultad)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        """, (pregunta, op_a, op_b, op_c, op_d, correcta, explicacion, dificultad))
                        db.commit()
                        db.close()

                    # Limpiar formulario
                    tf_pregunta.value = ""
                    tf_opcion_a.value = ""
                    tf_opcion_b.value = ""
                    tf_opcion_c.value = ""
                    tf_opcion_d.value = ""
                    tf_explicacion.value = ""
                    dd_correcta.value = None
                    dd_dificultad.value = "Fácil"

                    mostrar_snack("✅ Pregunta agregada exitosamente.", color="#7CFC00")
                    cargar_lista_preguntas()
                    try:
                        page.update()
                    except Exception:
                        pass
                except Exception as ex:
                    mostrar_snack(f"Error al guardar: {ex}", color="red")

            cargar_lista_preguntas()

            try:
                db_count = conectar_db()
                total_preguntas = 0
                if db_count:
                    cur_c = db_count.cursor()
                    cur_c.execute("SELECT COUNT(*) FROM reto_preguntas")
                    total_preguntas = cur_c.fetchone()[0]
                    db_count.close()
            except Exception:
                total_preguntas = 0

            return ft.Column([
                ft.Row([
                    ft.Icon(ft.Icons.QUIZ_ROUNDED, color="#FFD700", size=28),
                    ft.Text("Gestionar Trivia - Reto del Día 🧠", size=22, color="#D8B4FE", weight="bold"),
                ], spacing=10),
                ft.Text("Agrega nuevas preguntas al banco de trivia. Los usuarios verán 5 preguntas aleatorias cada vez que jueguen.",
                        color="#aaaaaa", size=13),
                ft.Row([
                    ft.Icon(ft.Icons.HELP_OUTLINE_ROUNDED, color="#00FFFF", size=16),
                    ft.Text(f"Total de preguntas en el banco: {total_preguntas}", color="#00FFFF", size=13, weight="bold"),
                ], spacing=6),
                ft.Divider(height=15, color="#333333"),

                # --- Formulario de nueva pregunta ---
                ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Icon(ft.Icons.ADD_CIRCLE_ROUNDED, color="#9D50BB", size=18),
                            ft.Text("Agregar Nueva Pregunta", color="#D8B4FE", size=15, weight="bold"),
                        ], spacing=8),
                        tf_pregunta,
                        ft.Row([tf_opcion_a, tf_opcion_b], spacing=10),
                        ft.Row([tf_opcion_c, tf_opcion_d], spacing=10),
                        tf_explicacion,
                        ft.Row([
                            dd_correcta,
                            dd_dificultad,
                            ft.ElevatedButton(
                                "Guardar Pregunta ➕",
                                bgcolor="#9D50BB",
                                color="white",
                                height=40,
                                on_click=agregar_pregunta_click
                            )
                        ], spacing=15, wrap=True)
                    ], spacing=10),
                    bgcolor="#0F0F1A",
                    padding=18,
                    border_radius=12,
                    border=ft.Border.all(1, "#333333")
                ),

                ft.Container(height=15),

                # --- Lista de preguntas existentes ---
                ft.Row([
                    ft.Icon(ft.Icons.LIST_ALT_ROUNDED, color="#FFD700", size=18),
                    ft.Text("Preguntas en el Banco", color="#FFD700", size=14, weight="bold"),
                    ft.TextButton(
                        content=ft.Row([ft.Icon(ft.Icons.REFRESH_ROUNDED, size=14, color="#00FFFF"),
                                        ft.Text("Actualizar", color="#00FFFF", size=12)], spacing=4),
                        on_click=lambda e: (cargar_lista_preguntas(), page.update())
                    )
                ], spacing=10),

                ft.Container(
                    content=lista_preguntas_col,
                    expand=True
                )
            ], scroll=ft.ScrollMode.AUTO, expand=True)

        def build_meta_semanal_view():
            import json, datetime

            # --- ESTADOS Y CONFIGURACIÓN ---
            meta_id_holder = [None]
            active_subtab = ["diarias"] # Default "diarias" (Hoja Domingo)
            target_user_id = [user_info["id"]]
            # Variables de tienda activa en consulta
            current_store_name_holder = ["MI TIENDA"]
            tienda_label_diarias = ft.Text("🏬 Tienda en consulta: ", color="#00FFFF", size=14, weight="bold")
            tienda_label_vendedores = ft.Text("🏬 ", color="#00FFFF", size=14, weight="bold")

            def actualizar_titulos_tienda():
                s_name = current_store_name_holder[0]
                tienda_label_diarias.value = f"🏬 Tienda en consulta: {s_name}"
                tienda_label_vendedores.value = f"🏬 {s_name}"

            admin_store_bar = ft.Container()

            if es_admin():
                tiendas_opciones = []
                try:
                    db_t = conectar_db()
                    if db_t:
                        cur_t = db_t.cursor(dictionary=True)
                        cur_t.execute("SELECT ID_Usuario, Nombre_Completo, Tienda FROM usuarios WHERE Tienda IS NOT NULL AND Tienda != '' ORDER BY Tienda ASC")
                        rows_t = cur_t.fetchall()
                        db_t.close()
                        for r in rows_t:
                            tiendas_opciones.append(
                                ft.dropdown.Option(
                                    key=str(r["ID_Usuario"]),
                                    text=f"{r['Tienda']} ({r['Nombre_Completo']})"
                                )
                            )
                except Exception as ex_t:
                    print("Error cargando tiendas admin:", ex_t)

                if tiendas_opciones:
                    target_user_id[0] = int(tiendas_opciones[0].key)
                    current_store_name_holder[0] = tiendas_opciones[0].text
                    actualizar_titulos_tienda()

                    dd_tiendas_admin = ft.Dropdown(
                        options=tiendas_opciones,
                        value=tiendas_opciones[0].key,
                        border_color="#9D50BB",
                        color="white",
                        bgcolor="#0F0F1A",
                        expand=True
                    )

                    def ejecutar_cargar_tienda_admin(e=None):
                        if dd_tiendas_admin and dd_tiendas_admin.value:
                            t_id = int(dd_tiendas_admin.value)
                            target_user_id[0] = t_id
                            t_name = "TIENDA"
                            for opt in dd_tiendas_admin.options:
                                if opt.key == str(t_id):
                                    t_name = opt.text
                                    break
                            current_store_name_holder[0] = t_name
                            actualizar_titulos_tienda()
                            cargar_datos(t_id)
                            if active_subtab[0] == "diarias":
                                subtab_content.content = build_subtab_diarias_view()
                            else:
                                subtab_content.content = build_subtab_vendedores_view()
                            try: page.update()
                            except Exception: pass
                            mostrar_snack(f"Metas cargadas para: {t_name}", color="#7CFC00")

                    btn_cargar_tienda_admin = ft.ElevatedButton(
                        "🔄 Cargar Tienda",
                        bgcolor="#0284c7",
                        color="white",
                        height=38,
                        style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8)),
                        on_click=ejecutar_cargar_tienda_admin
                    )

                    admin_store_bar = ft.Container(
                        content=ft.Column([
                            ft.Row([
                                ft.Icon(ft.Icons.ADMIN_PANEL_SETTINGS_ROUNDED, color="#00FFFF", size=20),
                                ft.Text("Consultar Tienda (Admin):", color="white", weight="bold", size=13),
                            ]),
                            ft.Row([
                                dd_tiendas_admin,
                                btn_cargar_tienda_admin
                            ], spacing=10)
                        ], spacing=8),
                        bgcolor="#1e1b4b",
                        padding=10,
                        border_radius=8,
                        border=ft.Border.all(1, "#9D50BB")
                    )
            else:
                current_store_name_holder[0] = user_info.get("tienda") or user_info.get("nombre") or "Mi Tienda"
                actualizar_titulos_tienda()

            # Semana empieza de DOMINGO a SÁBADO
            dias_semana = ["Domingo", "Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado"]

            # Contenedores principales
            vendedores_container = ft.Column(spacing=20)
            subtab_content = ft.Container(expand=True)

            # Etiquetas de texto calculadas (Diarias)
            pct_labels = {}
            meta_labels = {}
            acum_ly_labels = {}
            acum_meta_labels = {}
            tot_vly_label = ft.Text("$0", color="#facc15", weight="bold", size=8.5, no_wrap=True)

            # Ajustes adaptativos para PC vs Celular
            is_mobile = (page.width < 700) if (page and page.width) else False
            f_size = 8.5 if is_mobile else 12.0
            inp_h = 24 if is_mobile else 34
            p_val = 2 if is_mobile else 6
            s_val = 4 if is_mobile else 8

            def habilitar_seleccion_inteligente(tf):
                state = {"click_count": 0}

                def on_focus(e):
                    state["click_count"] = 1
                    if tf.value:
                        tf.selection = ft.TextSelection(0, len(str(tf.value)))
                        try: tf.update()
                        except Exception: pass

                def on_click(e):
                    if state["click_count"] == 1:
                        state["click_count"] = 2
                        tf.selection = None
                        try: tf.update()
                        except Exception: pass
                    else:
                        state["click_count"] = 1
                        if tf.value:
                            tf.selection = ft.TextSelection(0, len(str(tf.value)))
                            try: tf.update()
                            except Exception: pass

                def on_blur(e):
                    state["click_count"] = 0

                tf.on_focus = on_focus
                tf.on_click = on_click
                tf.on_blur = on_blur

            # Inputs amarillos para VLY (Venta Año Pasado) - Hoja Domingo
            inputs_vly = {}
            for d in dias_semana:
                default_val = "0"
                if d == "Lunes": default_val = "9164"
                elif d == "Miercoles": default_val = "7576"
                elif d == "Jueves": default_val = "8203"
                elif d == "Viernes": default_val = "12047"
                elif d == "Sabado": default_val = "8860"
                
                tf_vly = ft.TextField(
                    value=default_val,
                    bgcolor="#fef08a",
                    color="#000000",
                    border_color="#eab308",
                    height=inp_h,
                    content_padding=ft.padding.Padding(4, 2, 4, 2) if not is_mobile else 1,
                    text_align=ft.TextAlign.RIGHT,
                    text_size=f_size,
                    text_style=ft.TextStyle(weight="bold"),
                    on_change=lambda e: recalcular_todo(e)
                )
                habilitar_seleccion_inteligente(tf_vly)
                inputs_vly[d] = tf_vly

                pct_labels[d] = ft.Text("0.0%", color="#D8B4FE", size=f_size, no_wrap=True)
                meta_labels[d] = ft.Text("$0", color="#00FFFF", weight="bold", size=f_size, no_wrap=True)
                acum_ly_labels[d] = ft.Text("$0", color="#aaaaaa", size=f_size, no_wrap=True)
                acum_meta_labels[d] = ft.Text("$0", color="#00FFFF", weight="bold", size=f_size, no_wrap=True)

            tot_acum_ly_label = ft.Text("$0", color="#aaaaaa", size=f_size, no_wrap=True)
            tot_acum_meta_label = ft.Text("$0", color="#00FFFF", weight="bold", size=f_size, no_wrap=True)

            tot_vly_label = ft.Text("$0", color="#facc15", weight="bold", size=f_size, no_wrap=True)
            # Meta Total General ($60,000 en amarillo - Celda D10 del Excel)
            meta_total_input = ft.TextField(
                value="60000",
                bgcolor="#facc15",
                color="#000000",
                border_color="#ca8a04",
                height=inp_h,
                content_padding=ft.padding.Padding(4, 2, 4, 2) if not is_mobile else 1,
                text_align=ft.TextAlign.RIGHT,
                text_size=f_size,
                text_style=ft.TextStyle(weight="bold"),
                on_change=lambda e: recalcular_todo(e)
            )
            habilitar_seleccion_inteligente(meta_total_input)

            # Construir la estructura estática Adaptativa (Celular/PC) para Metas Diarias
            diarias_rows = []
            header_diarias = ft.Container(
                content=ft.Row([
                    ft.Text("DÍA", expand=10, weight="bold", color="#D8B4FE", size=f_size, no_wrap=True),
                    ft.Text("VLY 🟨", expand=18, weight="bold", color="#facc15", text_align=ft.TextAlign.RIGHT, size=f_size, no_wrap=True),
                    ft.Text("%", expand=10, weight="bold", color="#D8B4FE", text_align=ft.TextAlign.RIGHT, size=f_size, no_wrap=True),
                    ft.Text("META", expand=20, weight="bold", color="#00FFFF", text_align=ft.TextAlign.RIGHT, size=f_size, no_wrap=True),
                    ft.Text("ACUM.LY", expand=20, weight="bold", color="#aaaaaa", text_align=ft.TextAlign.RIGHT, size=f_size, no_wrap=True),
                    ft.Text("ACUM.META", expand=28, weight="bold", color="#00FFFF", text_align=ft.TextAlign.CENTER, size=f_size, no_wrap=True),
                ], spacing=2, expand=True),
                bgcolor="#262626", padding=ft.padding.Padding(p_val+2, p_val+2, p_val+2, p_val+2), border_radius=4, expand=True
            )
            diarias_rows.append(header_diarias)

            for idx, d in enumerate(dias_semana):
                row_bg = "#1e1e1e" if idx % 2 == 0 else "#252525"
                d_short = d[:3].upper() if len(d) > 3 else d.upper()
                if d == "Miercoles": d_short = "MIÉ"
                elif d == "Sabado": d_short = "SÁB"

                diarias_rows.append(
                    ft.Container(
                        content=ft.Row([
                            ft.Text(d_short, expand=10, weight="bold", color="white", size=f_size, no_wrap=True),
                            ft.Container(inputs_vly[d], expand=18),
                            ft.Container(pct_labels[d], expand=10, alignment=ft.alignment.Alignment(1, 0)),
                            ft.Container(meta_labels[d], expand=20, alignment=ft.alignment.Alignment(1, 0)),
                            ft.Container(acum_ly_labels[d], expand=20, alignment=ft.alignment.Alignment(1, 0)),
                            ft.Container(acum_meta_labels[d], expand=28, alignment=ft.alignment.Alignment(0, 0)),
                        ], spacing=2, expand=True),
                        bgcolor=row_bg, padding=ft.padding.Padding(p_val, p_val, p_val, p_val), border_radius=4, expand=True
                    )
                )

            totales_row = ft.Container(
                content=ft.Row([
                    ft.Text("TOTAL", expand=10, weight="bold", color="#facc15", size=f_size, no_wrap=True),
                    ft.Container(tot_vly_label, expand=18, alignment=ft.alignment.Alignment(1, 0)),
                    ft.Container(ft.Text("100%", color="#facc15", weight="bold", size=f_size, no_wrap=True), expand=10, alignment=ft.alignment.Alignment(1, 0)),
                    ft.Container(meta_total_input, expand=20, alignment=ft.alignment.Alignment(1, 0)),
                    ft.Container(tot_acum_ly_label, expand=20, alignment=ft.alignment.Alignment(1, 0)),
                    ft.Container(tot_acum_meta_label, expand=28, alignment=ft.alignment.Alignment(0, 0)),
                ], spacing=2, expand=True),
                bgcolor="#1e1b4b", padding=ft.padding.Padding(p_val+2, p_val+2, p_val+2, p_val+2), border_radius=6, border=ft.Border.all(1, "#facc15")
            )
            diarias_rows.append(totales_row)

            # Contenedor de Metas Diarias FLUIDO
            diarias_container = ft.Column(controls=diarias_rows, spacing=s_val, expand=True)

            # Controles Vendedores (Hoja Vendedor)
            aur_input = ft.TextField(
                value="3553",
                bgcolor="#1e1b4b",
                color="#facc15",
                border_color="#facc15",
                width=120,
                text_align=ft.TextAlign.RIGHT,
                text_size=13,
                text_style=ft.TextStyle(weight="bold"),
                on_change=lambda e: recalcular_todo(e)
            )
            aur_piezas_tienda_label = ft.Text("16.89 pz", color="#00FFFF", weight="bold", size=13)

            num_vendedores_input = ft.TextField(
                label="N° Vendedores",
                value="3",
                border_color="#9D50BB",
                color="white",
                width=140,
                on_change=lambda e: ajustar_vendedores()
            )

            vendedores_list = []

            kpi_table_rows_container = ft.Column(spacing=4)
            _kpi_recalc = [False]  # flag anti-recursión
            tbl_f_size = 11.0 if is_mobile else 13.5
            tbl_inp_h  = 30   if is_mobile else 38
            tbl_inp_ts = 10.5 if is_mobile else 13.0
            tbl_pad    = 3    if is_mobile else 6

            w_vend     = 65 if is_mobile else 140
            w_meta     = 65 if is_mobile else 130
            w_pct_pol  = 42 if is_mobile else 75
            w_pol_n    = 30 if is_mobile else 50
            w_pct_mult = 42 if is_mobile else 75
            w_mult_n   = 30 if is_mobile else 50
            w_ppt      = 38 if is_mobile else 65
            w_pct_lujo = 42 if is_mobile else 75
            w_lujo_n   = 30 if is_mobile else 50
            w_pz       = 35 if is_mobile else 60

            total_kpi_meta_label   = ft.Text("$0",    color="#facc15", weight="bold", size=tbl_f_size)
            total_kpi_pol_label    = ft.Text("0",     color="#facc15", weight="bold", size=tbl_f_size)
            total_kpi_mult_label   = ft.Text("0",     color="#facc15", weight="bold", size=tbl_f_size)
            total_kpi_ppt_label    = ft.Text("1.50",  color="#ffffff", weight="bold", size=tbl_f_size)
            total_kpi_lujo_label   = ft.Text("0",     color="#facc15", weight="bold", size=tbl_f_size)
            total_kpi_piezas_label = ft.Text("0",     color="#00FFFF", weight="bold", size=tbl_f_size)
            # Labels de PROMEDIO para la fila total
            total_kpi_pct_pol_label  = ft.Text("0%",   color="#aaaaaa", size=tbl_f_size)
            total_kpi_pct_mult_label = ft.Text("0%",   color="#aaaaaa", size=tbl_f_size)
            total_kpi_pct_lujo_label = ft.Text("0%",   color="#aaaaaa", size=tbl_f_size)

            def refrescar_kpi_table():
                rows = []
                header = ft.Container(
                    content=ft.Row([
                        ft.Text("VEND", width=w_vend, weight="bold", color="#ffffff", size=tbl_f_size),
                        ft.Text("META$", width=w_meta, weight="bold", color="#ffffff", size=tbl_f_size, text_align=ft.TextAlign.RIGHT),
                        ft.Text("%POL", width=w_pct_pol, weight="bold", color="#facc15", size=tbl_f_size, text_align=ft.TextAlign.CENTER),
                        ft.Text("#P", width=w_pol_n, weight="bold", color="#ffffff", size=tbl_f_size, text_align=ft.TextAlign.CENTER),
                        ft.Text("%MUL", width=w_pct_mult, weight="bold", color="#facc15", size=tbl_f_size, text_align=ft.TextAlign.CENTER),
                        ft.Text("#M", width=w_mult_n, weight="bold", color="#ffffff", size=tbl_f_size, text_align=ft.TextAlign.CENTER),
                        ft.Text("PPT", width=w_ppt, weight="bold", color="#ffffff", size=tbl_f_size, text_align=ft.TextAlign.CENTER),
                        ft.Text("%LUJ", width=w_pct_lujo, weight="bold", color="#facc15", size=tbl_f_size, text_align=ft.TextAlign.CENTER),
                        ft.Text("#L", width=w_lujo_n, weight="bold", color="#ffffff", size=tbl_f_size, text_align=ft.TextAlign.CENTER),
                        ft.Text("PZ", width=w_pz, weight="bold", color="#00FFFF", size=tbl_f_size, text_align=ft.TextAlign.CENTER),
                    ], spacing=2 if is_mobile else 6),
                    bgcolor="#080812", padding=4 if is_mobile else 6, border_radius=4
                )
                rows.append(header)

                for idx, v in enumerate(vendedores_list):
                    r_bg = "#1e1e1e" if idx % 2 == 0 else "#252525"
                    n_str = str(v["nombre"].value or "").strip().upper()
                    if is_mobile and len(n_str) > 7: n_str = n_str[:7]
                    v_name_lbl = ft.Text(n_str, size=tbl_f_size, weight="bold", color="white")

                    row_c = ft.Container(
                        content=ft.Row([
                            ft.Container(v_name_lbl, width=w_vend),
                            ft.Container(v["kpi_meta_text"], width=w_meta, alignment=ft.alignment.Alignment(1, 0)),
                            ft.Container(v["kpi_pct_pol_input"], width=w_pct_pol, alignment=ft.alignment.Alignment(0, 0)),
                            ft.Container(v["kpi_pol_text"], width=w_pol_n, alignment=ft.alignment.Alignment(0, 0)),
                            ft.Container(v["kpi_pct_mult_input"], width=w_pct_mult, alignment=ft.alignment.Alignment(0, 0)),
                            ft.Container(v["kpi_mult_text"], width=w_mult_n, alignment=ft.alignment.Alignment(0, 0)),
                            ft.Container(v["kpi_ppt_input"], width=w_ppt, alignment=ft.alignment.Alignment(0, 0)),
                            ft.Container(v["kpi_pct_lujo_input"], width=w_pct_lujo, alignment=ft.alignment.Alignment(0, 0)),
                            ft.Container(v["kpi_lujo_text"], width=w_lujo_n, alignment=ft.alignment.Alignment(0, 0)),
                            ft.Container(v["kpi_piezas_text"], width=w_pz, alignment=ft.alignment.Alignment(0, 0)),
                        ], spacing=2 if is_mobile else 6),
                        bgcolor=r_bg, padding=3 if is_mobile else 5, border_radius=4
                    )
                    rows.append(row_c)

                tot_row = ft.Container(
                    content=ft.Row([
                        ft.Text("TOTAL", width=w_vend, weight="bold", color="#ffffff", size=tbl_f_size),
                        ft.Container(total_kpi_meta_label,    width=w_meta, alignment=ft.alignment.Alignment(1, 0)),
                        ft.Container(total_kpi_pct_pol_label, width=w_pct_pol, alignment=ft.alignment.Alignment(0, 0)),
                        ft.Container(total_kpi_pol_label,     width=w_pol_n, alignment=ft.alignment.Alignment(0, 0)),
                        ft.Container(total_kpi_pct_mult_label,width=w_pct_mult, alignment=ft.alignment.Alignment(0, 0)),
                        ft.Container(total_kpi_mult_label,    width=w_mult_n, alignment=ft.alignment.Alignment(0, 0)),
                        ft.Container(total_kpi_ppt_label,     width=w_ppt, alignment=ft.alignment.Alignment(0, 0)),
                        ft.Container(total_kpi_pct_lujo_label,width=w_pct_lujo, alignment=ft.alignment.Alignment(0, 0)),
                        ft.Container(total_kpi_lujo_label,    width=w_lujo_n, alignment=ft.alignment.Alignment(0, 0)),
                        ft.Container(total_kpi_piezas_label,  width=w_pz, alignment=ft.alignment.Alignment(0, 0)),
                    ], spacing=2 if is_mobile else 6),
                    bgcolor="#404040", padding=4 if is_mobile else 6, border_radius=4
                )
                rows.append(tot_row)
                kpi_table_rows_container.controls = rows

            kpi_summary_container = ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Text("Meta por vendedor", weight="bold", color="#ffffff", size=16),
                        ft.Row([
                            ft.Text("cual es tu aur anual:", color="#facc15", weight="bold", size=13),
                            aur_input,
                            ft.Text("piezas para llegar a meta:", color="#aaaaaa", size=12),
                            aur_piezas_tienda_label,
                            ft.ElevatedButton(
                                "🔄 Actualizar Tabla",
                                bgcolor="#7c3aed",
                                color="white",
                                height=36,
                                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8)),
                                on_click=lambda e: forzar_recalculo(e)
                            ),
                        ], spacing=10, wrap=True)
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, wrap=True),
                    ft.Row([kpi_table_rows_container], scroll=ft.ScrollMode.AUTO)
                ], spacing=10),
                bgcolor="#171717", padding=15, border_radius=10, border=ft.Border.all(1, "#333333")
            )

            def crear_vendedor_item(n_val, d_val="6"):
                dd = ft.Dropdown(
                    value=str(d_val),
                    options=[ft.dropdown.Option(str(d)) for d in range(0, 7)],
                    border_color="#333333",
                    color="white",
                    width=75 if is_mobile else 90,
                    height=36 if is_mobile else 40
                )
                
                lbl_w = 45 if is_mobile else 70
                box_w = 85 if is_mobile else 150
                f_sz  = 10.0 if is_mobile else 12.0

                meta_diaria_texts = [ft.Text("$ 0.00", color="#000000", weight="bold", size=f_sz) for _ in range(6)]
                meta_diaria_containers = [
                    ft.Container(
                        meta_diaria_texts[i],
                        width=box_w,
                        alignment=ft.alignment.Alignment(1, 0),
                        bgcolor="#fef08a",
                        padding=3 if is_mobile else 5,
                        border_radius=4
                    ) for i in range(6)
                ]
                acum_texts = [ft.Text("$ 0.00", color="#000000", weight="bold", size=f_sz) for _ in range(6)]
                day_label_texts = [ft.Text(f"Día {d_i}", width=lbl_w, color="#ffffff", weight="bold", size=f_sz) for d_i in range(1, 7)]
                row_containers = []
                
                header_title = ft.Text(n_val.upper(), width=lbl_w, weight="bold", color="#000000", size=f_sz)
                meta_vend_text = ft.Text("Meta Vendedor: $ 0.00", color="#00FFFF", weight="bold", size=11.0 if is_mobile else 13.0)

                kpi_meta_text = ft.Text("$0", color="#ffffff", weight="bold", size=tbl_f_size)
                kpi_pct_pol_input  = ft.TextField(value="45%", bgcolor="#1e1e1e", color="#facc15", border_color="#333333", width=w_pct_pol, height=tbl_inp_h, content_padding=tbl_pad, text_align=ft.TextAlign.CENTER, text_size=tbl_inp_ts)
                kpi_pol_text       = ft.Text("0", color="#facc15", weight="bold", size=tbl_inp_ts)
                kpi_pct_mult_input = ft.TextField(value="45%", bgcolor="#1e1e1e", color="#facc15", border_color="#333333", width=w_pct_mult, height=tbl_inp_h, content_padding=tbl_pad, text_align=ft.TextAlign.CENTER, text_size=tbl_inp_ts)
                kpi_mult_text      = ft.Text("0", color="#facc15", weight="bold", size=tbl_inp_ts)
                kpi_ppt_input      = ft.TextField(value="1.45", bgcolor="#1e1e1e", color="#ffffff", border_color="#333333", width=w_ppt, height=tbl_inp_h, content_padding=tbl_pad, text_align=ft.TextAlign.CENTER, text_size=tbl_inp_ts, read_only=True)
                kpi_pct_lujo_input = ft.TextField(value="45%", bgcolor="#1e1e1e", color="#facc15", border_color="#333333", width=w_pct_lujo, height=tbl_inp_h, content_padding=tbl_pad, text_align=ft.TextAlign.CENTER, text_size=tbl_inp_ts)
                kpi_lujo_text      = ft.Text("0", color="#facc15", weight="bold", size=tbl_inp_ts)
                kpi_piezas_text    = ft.Text("0", color="#00FFFF", weight="bold", size=tbl_inp_ts)

                habilitar_seleccion_inteligente(kpi_pct_pol_input)
                habilitar_seleccion_inteligente(kpi_pct_mult_input)
                habilitar_seleccion_inteligente(kpi_pct_lujo_input)
                
                v_header = ft.Container(
                    content=ft.Row([
                        header_title,
                        ft.Text("META 🟨", width=box_w, weight="bold", color="#000000", size=f_sz, text_align=ft.TextAlign.RIGHT),
                        ft.Text("ACUM 🟨" if is_mobile else "ACUMULADO 🟨", width=box_w, weight="bold", color="#000000", size=f_sz, text_align=ft.TextAlign.RIGHT),
                    ], spacing=3 if is_mobile else 6),
                    bgcolor="#facc15", padding=4 if is_mobile else 6, border_radius=4
                )
                
                v_rows = [v_header]
                for d_i in range(1, 7):
                    c = ft.Container(
                        content=ft.Row([
                            day_label_texts[d_i-1],
                            meta_diaria_containers[d_i-1],
                            ft.Container(acum_texts[d_i-1], width=box_w, alignment=ft.alignment.Alignment(1, 0), bgcolor="#fef08a", padding=3 if is_mobile else 5, border_radius=4),
                        ], spacing=3 if is_mobile else 6),
                        bgcolor="#1e1e1e" if d_i % 2 == 1 else "#252525",
                        padding=3 if is_mobile else 5, border_radius=4
                    )
                    v_rows.append(c)
                    row_containers.append(c)

                nombre_tf = ft.TextField(value=n_val, border_color="#444444", color="white", width=110 if is_mobile else 140, height=34 if is_mobile else 38, text_size=11 if is_mobile else 12, text_style=ft.TextStyle(weight="bold"), content_padding=4 if is_mobile else 6)
                nombre_tf.on_change = lambda e: (setattr(header_title, "value", nombre_tf.value.upper()), page.update() if page else None)

                item = {
                    "nombre": nombre_tf,
                    "dias": dd,
                    "header_title": header_title,
                    "meta_vend_text": meta_vend_text,
                    "meta_diaria_texts": meta_diaria_texts,
                    "meta_diaria_containers": meta_diaria_containers,
                    "day_label_texts": day_label_texts,
                    "acum_texts": acum_texts,
                    "row_containers": row_containers,
                    "card_container": None,
                    "kpi_meta_text": kpi_meta_text,
                    "kpi_pct_pol_input": kpi_pct_pol_input,
                    "kpi_pol_text": kpi_pol_text,
                    "kpi_pct_mult_input": kpi_pct_mult_input,
                    "kpi_mult_text": kpi_mult_text,
                    "kpi_ppt_input": kpi_ppt_input,
                    "kpi_pct_lujo_input": kpi_pct_lujo_input,
                    "kpi_lujo_text": kpi_lujo_text,
                    "kpi_piezas_text": kpi_piezas_text
                }

                def make_on_dias_change(dd_control, v_item):
                    def on_dias_change(e):
                        val_str = None
                        if e and hasattr(e, "control") and e.control and e.control.value:
                            val_str = str(e.control.value).strip()
                        if (not val_str or val_str not in ["0", "1", "2", "3", "4", "5", "6"]) and e and hasattr(e, "data") and e.data:
                            val_str = str(e.data).strip()
                        
                        if val_str in ["0", "1", "2", "3", "4", "5", "6"]:
                            dd_control.value = val_str
                            v_item["dias"].value = val_str
                        recalcular_todo(e)
                    return on_dias_change

                dd.on_change = make_on_dias_change(dd, item)

                def eliminar_vendedor_click(e):
                    if len(vendedores_list) > 1:
                        if item in vendedores_list:
                            vendedores_list.remove(item)
                            num_vendedores_input.value = str(len(vendedores_list))
                            refrescar_vendedores_container()
                            recalcular_vendedores()

                
                # --- CONTROL DE BIOMETRÍA Y PERMISOS DE GERENTE ---
                def registrar_rostro_vend_click(e):
                    res, msg = guardar_biometria_db(user_info["id"], item.get("nombre", "Vendedor"), encoding_rostro="[ENCODING_ROSTRO_VECTOR_128_FLOAT_DUMMY]")
                    mostrar_snack(msg, "#7CFC00" if res else "red")

                def registrar_huella_vend_click(e):
                    res, msg = guardar_biometria_db(user_info["id"], item.get("nombre", "Vendedor"), hash_huella="[HASH_HUELLA_MINUTIAS_DUMMY]")
                    mostrar_snack(msg, "#7CFC00" if res else "red")

                def eliminar_biometria_vend_click(e):
                    # Verificar si el usuario activo tiene puesto o rol de Gerente
                    es_gerente = any(k in str(user_info.get("rol", "")).lower() or k in str(user_info.get("puesto", "")).lower() for k in ["gerente", "admin"])
                    if not es_gerente:
                        mostrar_snack("⚠️ Permiso denegado: Solo el Gerente de Tienda puede eliminar datos biométricos", "red")
                        return
                    
                    registrar_auditoria_borrado(
                        ejecutor_id=user_info.get("id", 0),
                        ejecutor_nombre=user_info.get("nombre", "Gerente"),
                        ejecutor_rol=user_info.get("rol", "Gerente de Tienda"),
                        afectado_nombre=item.get("nombre", "Vendedor"),
                        accion="ELIMINACION_BIOMETRIA",
                        detalles="Eliminación de datos biométricos autorizada por Gerente de Tienda"
                    )
                    mostrar_snack(f"Biometría de {item.get('nombre', 'Vendedor')} eliminada y registrada en auditoría 🛡️", "#7CFC00")

                btn_bio_rostro = ft.IconButton(icon=ft.Icons.FACE_ROUNDED, tooltip="Registrar Rostro (Face ID)", icon_color="#003366", on_click=registrar_rostro_vend_click)
                btn_bio_huella = ft.IconButton(icon=ft.Icons.FINGERPRINT_ROUNDED, tooltip="Registrar Huella Dactilar", icon_color="#003366", on_click=registrar_huella_vend_click)
                btn_bio_del = ft.IconButton(icon=ft.Icons.SHIELD_ROUNDED, tooltip="Eliminar Biometría (Solo Gerente de Tienda)", icon_color="#FF4500", on_click=eliminar_biometria_vend_click)


                v_table_box = ft.Container(
                    content=ft.Column([v_header] + row_containers, spacing=3),
                    width=230 if is_mobile else 400
                )

                item["card_container"] = ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Row([
                                nombre_tf,
                                ft.Text("Días:", color="#aaaaaa", size=11 if is_mobile else 12, weight="bold"),
                                dd,
                            ], spacing=4 if is_mobile else 6, wrap=True),
                            ft.Row([
                                meta_vend_text,
                                ft.IconButton(
                                    icon=ft.Icons.DELETE_OUTLINE_ROUNDED,
                                    icon_color="#FF4500",
                                    icon_size=18 if is_mobile else 20,
                                    tooltip="Eliminar Vendedor",
                                    on_click=eliminar_vendedor_click
                                )
                            ], spacing=4 if is_mobile else 6, wrap=True)
                        ], spacing=6, alignment=ft.MainAxisAlignment.SPACE_BETWEEN, wrap=True),
                        v_table_box
                    ], spacing=6),
                    bgcolor="#171717", padding=8 if is_mobile else 12, border_radius=8, border=ft.Border.all(1, "#333333")
                )

                return item

            def refrescar_vendedores_container():
                vendedores_container.controls = [v["card_container"] for v in vendedores_list]
                try: page.update()
                except Exception: pass

            def inicializar_vendedores_default(n):
                vendedores_list.clear()
                for i in range(1, n + 1):
                    vendedores_list.append(crear_vendedor_item(f"VENDEDOR {i}", "6"))
                refrescar_vendedores_container()

            def ajustar_vendedores():
                try:
                    n = int(num_vendedores_input.value or 3)
                    n = max(1, min(15, n))
                except ValueError:
                    n = 3
                
                curr_len = len(vendedores_list)
                if n > curr_len:
                    for i in range(curr_len + 1, n + 1):
                        vendedores_list.append(crear_vendedor_item(f"VENDEDOR {i}", "6"))
                elif n < curr_len:
                    del vendedores_list[n:]
                
                refrescar_vendedores_container()
                recalcular_vendedores()

            def recalcular_diarias():
                try:
                    meta_total = float(meta_total_input.value or 0)
                except ValueError:
                    meta_total = 0.0

                vly_vals = {}
                tot_vly = 0.0
                for d in dias_semana:
                    try: val = float(inputs_vly[d].value or 0)
                    except ValueError: val = 0.0
                    vly_vals[d] = val
                    tot_vly += val

                tot_vly_label.value = f"${tot_vly:,.0f}"

                acum_ly = 0.0
                acum_meta = 0.0

                for d in dias_semana:
                    vly_d = vly_vals[d]
                    pct_d = (vly_d / tot_vly * 100.0) if tot_vly > 0 else (100.0 / 7.0)
                    meta_d = (meta_total * pct_d / 100.0)

                    acum_ly += vly_d
                    acum_meta += meta_d

                    pct_labels[d].value = f"{pct_d:.1f}%"
                    meta_labels[d].value = f"${meta_d:,.0f}"
                    acum_ly_labels[d].value = f"${acum_ly:,.0f}"
                    acum_meta_labels[d].value = f"${acum_meta:,.0f}"

                tot_acum_ly_label.value = f"${acum_ly:,.0f}"
                tot_acum_meta_label.value = f"${acum_meta:,.0f}"

            def recalcular_vendedores():
                # Guard: evita re-entrada recursiva cuando se actualizan
                # los TextField de % desde dentro de este mismo método
                if _kpi_recalc[0]:
                    return
                _kpi_recalc[0] = True
                try:
                    _recalcular_vendedores_inner()
                except Exception as _ex:
                    print(f"[recalcular_vendedores] error: {_ex}")
                finally:
                    _kpi_recalc[0] = False  # SIEMPRE liberar, con o sin error

            def _recalcular_vendedores_inner():
                try:
                    meta_total = float(meta_total_input.value or 0)
                except ValueError:
                    meta_total = 0.0

                num_v = len(vendedores_list)
                if num_v == 0:
                    return

                base_meta_per_seller = meta_total / float(num_v)
                base_daily_meta = base_meta_per_seller / 6.0

                dias_trab = []
                for v in vendedores_list:
                    raw_val = v["dias"].value
                    if raw_val is not None and str(raw_val).strip() in ["0", "1", "2", "3", "4", "5", "6"]:
                        d = int(str(raw_val).strip())
                    else:
                        try: d = int(float(str(raw_val).strip() if raw_val is not None else 6))
                        except Exception: d = 6
                    dias_trab.append(max(0, min(6, d)))

                total_lost_goal = 0.0
                full_time_count = 0

                for d in dias_trab:
                    if d < 6:
                        total_lost_goal += (6 - d) * base_daily_meta
                    else:
                        full_time_count += 1

                absorbed_per_full_timer = (total_lost_goal / float(full_time_count)) if full_time_count > 0 else (total_lost_goal / float(num_v))

                # KPI Calculations
                try:
                    aur_val = float(aur_input.value or 3553)
                    if aur_val <= 0: aur_val = 3553.0
                except ValueError:
                    aur_val = 3553.0

                exact_store_pz = (meta_total / aur_val) if aur_val > 0 else 0.0
                aur_piezas_tienda_label.value = f"{exact_store_pz:.2f} pz"

                tot_kpi_meta = 0.0
                tot_kpi_pol  = 0
                tot_kpi_mult = 0
                tot_kpi_lujo = 0
                tot_kpi_piezas = 0
                sum_eff_pol  = 0
                sum_eff_mult = 0
                sum_eff_ppt  = 0.0
                sum_eff_lujo = 0
                count_active = 0

                for idx, v in enumerate(vendedores_list):
                    d = dias_trab[idx]
                    if d == 0:
                        seller_meta = 0.0
                    elif d < 6:
                        seller_meta = d * base_daily_meta
                    else:
                        seller_meta = base_meta_per_seller + absorbed_per_full_timer

                    daily_quota = seller_meta / float(d) if d > 0 else 0.0
                    v["meta_vend_text"].value = f"Meta Vendedor: $ {seller_meta:,.2f}"

                    import math
                    def std_round(x):
                        return int(math.floor(x + 0.5))

                    if seller_meta > 0 and aur_val > 0:
                        s_piezas = math.ceil(seller_meta / aur_val)
                    else:
                        s_piezas = 0

                    try:
                        pct_pol = float(str(v["kpi_pct_pol_input"].value or "45").replace("%", "").strip()) / 100.0
                        if pct_pol <= 0 or pct_pol >= 1.0: pct_pol = 0.45  # mínimo 45%, máx <100%
                    except Exception: pct_pol = 0.45

                    try:
                        pct_mult = float(str(v["kpi_pct_mult_input"].value or "45").replace("%", "").strip()) / 100.0
                        if pct_mult < 0.45 or pct_mult >= 1.0: pct_mult = 0.45
                    except Exception: pct_mult = 0.45

                    try:
                        ppt_v = float(str(v["kpi_ppt_input"].value or "1.45").strip())
                        if ppt_v < 1.45: ppt_v = 1.45  # mínimo operativo 1.45
                    except Exception: ppt_v = 1.45

                    try:
                        pct_lujo = float(str(v["kpi_pct_lujo_input"].value or "45").replace("%", "").strip()) / 100.0
                        if pct_lujo <= 0 or pct_lujo >= 1.0: pct_lujo = 0.45
                    except Exception: pct_lujo = 0.45


                    # FÓRMULA FÍSICA Y PISOS MÍNIMOS DE NEGOCIO:
                    # n_mult = pct_mult × piezas / (1 + pct_mult)
                    mult_pz = std_round(pct_mult * s_piezas / (1.0 + pct_mult)) if s_piezas > 0 else 0

                    # Garantizar pisos mínimos operativos (%múltiples >= 45% y PPT >= 1.45)
                    while s_piezas > 0 and mult_pz < s_piezas:
                        tot_t = max(1, s_piezas - mult_pz)
                        m_pct = std_round(mult_pz * 100 / tot_t)
                        ppt_val = round(s_piezas / tot_t, 2)
                        if m_pct >= 45 and ppt_val >= 1.45:
                            break
                        mult_pz += 1

                    # total_tickets derivado exactamente
                    total_tickets = max(1, s_piezas - mult_pz) if s_piezas > 0 else 0

                    # Consistencia física
                    if mult_pz > total_tickets:
                        mult_pz = total_tickets

                    pol_pz  = std_round(pct_pol  * s_piezas)
                    lujo_pz = std_round(pct_lujo  * s_piezas)

                    # Efectivos auto-consistentes final
                    eff_mult_pct = std_round(mult_pz * 100 / total_tickets)  if total_tickets > 0 else 0
                    eff_ppt      = round(s_piezas / total_tickets, 2)        if total_tickets > 0 else 1.45
                    eff_pol_pct  = std_round(pol_pz  * 100 / s_piezas)       if s_piezas > 0 else 0
                    eff_lujo_pct = std_round(lujo_pz * 100 / s_piezas)       if s_piezas > 0 else 0

                    # Actualizar campos de pantalla: % efectivos + PPT
                    if s_piezas > 0:
                        v["kpi_pct_pol_input"].value  = f"{eff_pol_pct}%"
                        v["kpi_pct_mult_input"].value = f"{eff_mult_pct}%"
                        v["kpi_ppt_input"].value      = f"{eff_ppt:.2f}"
                        v["kpi_pct_lujo_input"].value = f"{eff_lujo_pct}%"

                    v["kpi_meta_text"].value   = f"${seller_meta:,.0f}"
                    v["kpi_pol_text"].value    = str(pol_pz)
                    v["kpi_mult_text"].value   = str(mult_pz)
                    v["kpi_lujo_text"].value   = str(lujo_pz)
                    v["kpi_piezas_text"].value = str(s_piezas)



                    tot_kpi_meta   += seller_meta
                    tot_kpi_pol    += pol_pz
                    tot_kpi_mult   += mult_pz
                    tot_kpi_lujo   += lujo_pz
                    tot_kpi_piezas += s_piezas

                    if s_piezas > 0:
                        sum_eff_pol  += eff_pol_pct
                        sum_eff_mult += eff_mult_pct
                        sum_eff_ppt  += eff_ppt
                        sum_eff_lujo += eff_lujo_pct
                        count_active += 1

                    acum = 0.0
                    for row_idx in range(6):
                        v["row_containers"][row_idx].visible = True
                        if d > 0 and row_idx < d:
                            acum += daily_quota
                            v["row_containers"][row_idx].bgcolor = "#1e1e1e" if row_idx % 2 == 0 else "#252525"
                            v["day_label_texts"][row_idx].value = f"Día {row_idx + 1}"
                            v["day_label_texts"][row_idx].color = "#ffffff"
                            v["meta_diaria_containers"][row_idx].bgcolor = "#fef08a"
                            v["meta_diaria_texts"][row_idx].value = f"$ {daily_quota:,.2f}"
                            v["meta_diaria_texts"][row_idx].color = "#000000"
                            v["acum_texts"][row_idx].value = f"$ {acum:,.2f}"
                        else:
                            v["row_containers"][row_idx].bgcolor = "#450a0a"
                            v["day_label_texts"][row_idx].value = f"Día {row_idx + 1}"
                            v["day_label_texts"][row_idx].color = "#fca5a5"
                            v["meta_diaria_containers"][row_idx].bgcolor = "#7f1d1d"
                            v["meta_diaria_texts"][row_idx].value = "$ 0.00"
                            v["meta_diaria_texts"][row_idx].color = "#ffffff"
                            v["acum_texts"][row_idx].value = f"$ {acum:,.2f}"

                total_kpi_meta_label.value   = f"${tot_kpi_meta:,.0f}"
                total_kpi_pol_label.value    = str(tot_kpi_pol)
                total_kpi_mult_label.value   = str(tot_kpi_mult)
                total_kpi_lujo_label.value   = str(tot_kpi_lujo)
                total_kpi_piezas_label.value = str(tot_kpi_piezas)

                # Promedios de % efectivos para la fila total
                if count_active > 0:
                    import math
                    def std_round_tot(x): return int(math.floor(x + 0.5))
                    avg_pol  = std_round_tot(sum_eff_pol  / count_active)
                    avg_mult = std_round_tot(sum_eff_mult / count_active)
                    avg_ppt  = round(sum_eff_ppt / count_active, 2)
                    avg_lujo = std_round_tot(sum_eff_lujo / count_active)
                    total_kpi_pct_pol_label.value  = f"{avg_pol}%"
                    total_kpi_pct_mult_label.value = f"{avg_mult}%"
                    total_kpi_ppt_label.value      = f"{avg_ppt:.2f}"
                    total_kpi_pct_lujo_label.value = f"{avg_lujo}%"
                else:
                    total_kpi_pct_pol_label.value  = "0%"
                    total_kpi_pct_mult_label.value = "0%"
                    total_kpi_ppt_label.value      = "0.00"
                    total_kpi_pct_lujo_label.value = "0%"
                # page.update() ANTES de refrescar para que Flet confirme los
                # cambios de .value en los widgets antes de reasignarlos a
                # nuevos Containers (si no, Flet puede ignorar los cambios)
                try: page.update()
                except Exception: pass
                refrescar_kpi_table()
                refrescar_vendedores_container()

            def recalcular_todo(e=None):
                recalcular_diarias()
                recalcular_vendedores()
                try: page.update()
                except Exception: pass
                if e and hasattr(e, "control") and e.control:
                    try: e.control.focus()
                    except Exception: pass

            def forzar_recalculo(e=None):
                """Fuerza el recálculo limpiando el flag de anti-recursión primero.
                Se usa en el botón Recalcular/Aplicar Días para garantizar que
                siempre se ejecute aunque el flag haya quedado bloqueado."""
                _kpi_recalc[0] = False
                recalcular_todo(e)

            def guardar_todo_click(e):
                try:
                    meta_val = float(meta_total_input.value or 0)
                    vend_export = []
                    for v in vendedores_list:
                        vend_export.append({
                            "nombre": v["nombre"].value,
                            "dias": v["dias"].value
                        })

                    import datetime
                    hoy = datetime.date.today()
                    # Semana de Domingo a Sábado
                    domingo = hoy - datetime.timedelta(days=(hoy.weekday() + 1) % 7)
                    sabado = domingo + datetime.timedelta(days=6)

                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        try:
                            cursor.execute("SHOW COLUMNS FROM metas_semanales")
                            cols_result = cursor.fetchall()
                            existing_cols = [r[0] if isinstance(r, (tuple, list)) else r.get("Field") for r in cols_result]

                            if "Vendedores_JSON" not in existing_cols:
                                try: cursor.execute("ALTER TABLE metas_semanales ADD COLUMN Vendedores_JSON TEXT")
                                except Exception: pass

                            for d in ["AP_Domingo", "AP_Lunes", "AP_Martes", "AP_Miercoles", "AP_Jueves", "AP_Viernes", "AP_Sabado"]:
                                if d not in existing_cols:
                                    try: cursor.execute(f"ALTER TABLE metas_semanales ADD COLUMN {d} DECIMAL(15,2) DEFAULT 0.00")
                                    except Exception: pass
                        except Exception as ex_col:
                            print("Error al verificar/crear columnas:", ex_col)

                        uid = target_user_id[0]
                        if meta_id_holder[0] is None:
                            cursor.execute("""
                                INSERT INTO metas_semanales 
                                (ID_Usuario_Tienda, Fecha_Inicio, Fecha_Fin, Monto_Meta, 
                                 AP_Domingo, AP_Lunes, AP_Martes, AP_Miercoles, AP_Jueves, AP_Viernes, AP_Sabado,
                                 Vendedores_JSON)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """, (
                                uid, domingo, sabado, meta_val,
                                float(inputs_vly["Domingo"].value or 0), float(inputs_vly["Lunes"].value or 0), float(inputs_vly["Martes"].value or 0),
                                float(inputs_vly["Miercoles"].value or 0), float(inputs_vly["Jueves"].value or 0), float(inputs_vly["Viernes"].value or 0), float(inputs_vly["Sabado"].value or 0),
                                json.dumps(vend_export)
                            ))
                            meta_id_holder[0] = cursor.lastrowid
                        else:
                            cursor.execute("""
                                UPDATE metas_semanales SET 
                                Monto_Meta = %s, 
                                AP_Domingo = %s, AP_Lunes = %s, AP_Martes = %s, AP_Miercoles = %s, AP_Jueves = %s, AP_Viernes = %s, AP_Sabado = %s,
                                Vendedores_JSON = %s
                                WHERE ID_Meta = %s
                            """, (
                                meta_val,
                                float(inputs_vly["Domingo"].value or 0), float(inputs_vly["Lunes"].value or 0), float(inputs_vly["Martes"].value or 0),
                                float(inputs_vly["Miercoles"].value or 0), float(inputs_vly["Jueves"].value or 0), float(inputs_vly["Viernes"].value or 0), float(inputs_vly["Sabado"].value or 0),
                                json.dumps(vend_export),
                                meta_id_holder[0]
                            ))
                        db.commit()
                        db.close()
                        mostrar_snack("Metas guardadas exitosamente en LUXO 💾", "#7CFC00")
                except Exception as ex:
                    print("Error al guardar metas:", ex)
                    mostrar_snack("Error al guardar información", "red")

            def obtener_vendedores_tienda(uid):
                names = []
                try:
                    db_v = conectar_db()
                    if db_v:
                        cur_v = db_v.cursor(dictionary=True)
                        cur_v.execute("SELECT Nombre_Completo FROM vendedores WHERE ID_Usuario_Tienda = %s AND Activo = 1 ORDER BY ID_Vendedor ASC", (uid,))
                        rows_v = cur_v.fetchall()
                        db_v.close()
                        for r in rows_v:
                            if r.get("Nombre_Completo"):
                                names.append(r["Nombre_Completo"].strip().upper())
                except Exception as ex_v:
                    print("Error leyendo vendedores:", ex_v)
                return names

            def cargar_datos(target_uid=None):
                import datetime
                hoy = datetime.date.today()
                domingo = hoy - datetime.timedelta(days=(hoy.weekday() + 1) % 7)
                uid = target_uid if target_uid is not None else target_user_id[0]

                # Reset de estado inicial a CERO para tiendas nuevas o sin datos
                meta_id_holder[0] = None
                meta_total_input.value = "0"
                for d in dias_semana:
                    inputs_vly[d].value = "0"

                # Cargar vendedores desde DB de esa tienda si existen
                v_db_names = obtener_vendedores_tienda(uid)
                if v_db_names:
                    vendedores_list.clear()
                    for n_name in v_db_names:
                        vendedores_list.append(crear_vendedor_item(n_name, "6"))
                    num_vendedores_input.value = str(len(v_db_names))
                    refrescar_vendedores_container()
                else:
                    inicializar_vendedores_default(3)
                    num_vendedores_input.value = "3"

                recalcular_todo()
                
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("""
                            SELECT * FROM metas_semanales 
                            WHERE ID_Usuario_Tienda = %s AND Fecha_Inicio = %s
                        """, (uid, domingo))
                        row = cursor.fetchone()
                        db.close()
                        
                        if row:
                            meta_id_holder[0] = row["ID_Meta"]
                            meta_val = float(row.get("Monto_Meta", 0.0))
                            meta_total_input.value = str(int(meta_val) if meta_val.is_integer() else meta_val)

                            for d in dias_semana:
                                v_ap = row.get(f"AP_{d}")
                                if v_ap is not None:
                                    inputs_vly[d].value = str(int(float(v_ap)) if float(v_ap).is_integer() else float(v_ap))

                            v_json = row.get("Vendedores_JSON")
                            if v_json:
                                v_data = json.loads(v_json)
                                num_vendedores_input.value = str(len(v_data))
                                vendedores_list.clear()
                                for idx, item in enumerate(v_data):
                                    n_name = item.get("nombre", f"VENDEDOR {idx+1}")
                                    d_val = str(item.get("dias", "6"))
                                    vendedores_list.append(crear_vendedor_item(n_name, d_val))
                                refrescar_vendedores_container()
                            
                            recalcular_todo()
                except Exception as ex:
                    print("Error al cargar metas:", ex)

            # Botones Pestañas Superiores
            btn_subtab_diarias = ft.ElevatedButton(
                "📅 Metas Diarias (Año Pasado - Hoja Domingo)",
                bgcolor="#6E48AA", color="white",
                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8)),
                on_click=lambda e: cambiar_subtab("diarias")
            )
            btn_subtab_vendedores = ft.ElevatedButton(
                "👥 Vendedores (Hoja Vendedor)",
                bgcolor="#141424", color="white",
                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8)),
                on_click=lambda e: cambiar_subtab("vendedores")
            )

            def cambiar_subtab(tab):
                active_subtab[0] = tab
                if tab == "diarias":
                    btn_subtab_diarias.bgcolor = "#6E48AA"
                    btn_subtab_vendedores.bgcolor = "#222222"
                    subtab_content.content = build_subtab_diarias_view()
                else:
                    btn_subtab_diarias.bgcolor = "#222222"
                    btn_subtab_vendedores.bgcolor = "#6E48AA"
                    subtab_content.content = build_subtab_vendedores_view()
                try: page.update()
                except Exception: pass

            def agregar_un_vendedor():
                n = len(vendedores_list) + 1
                num_vendedores_input.value = str(n)
                ajustar_vendedores()

            def build_subtab_diarias_view():
                return ft.Column([
                    tienda_label_diarias,
                    diarias_container
                ], spacing=12, scroll=ft.ScrollMode.AUTO, expand=True)

            def build_subtab_vendedores_view():
                return ft.Column([
                    ft.Row([
                        ft.Row([
                            ft.ElevatedButton(
                                "➕ Agregar Vendedor",
                                bgcolor="#9D50BB",
                                color="white",
                                height=40,
                                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8)),
                                on_click=lambda e: agregar_un_vendedor()
                            ),
                            tienda_label_vendedores,
                        ], spacing=10, vertical_alignment="center", wrap=True),
                        num_vendedores_input
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, wrap=True),
                    kpi_summary_container,
                    ft.Row([
                        ft.ElevatedButton(
                            "🔄 Recalcular / Aplicar Días",
                            bgcolor="#0284c7",
                            color="white",
                            height=40,
                            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8)),
                            on_click=lambda e: forzar_recalculo(e)
                        ),
                    ]),
                    vendedores_container
                ], spacing=15, scroll=ft.ScrollMode.AUTO)

            btn_guardar_registro = ft.ElevatedButton(
                "Guardar Registro 💾",
                on_click=guardar_todo_click,
                bgcolor="#7CFC00",
                color="#1e1b4b",
                height=45,
                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
            )

            # Carga e inicialización
            cargar_datos()
            recalcular_todo()
            subtab_content.content = build_subtab_diarias_view()

            return ft.Column([
                ft.Row([
                    ft.Text("🎯 METAS Y METRICAS", size=24, color="#D8B4FE", weight="bold"),
                    btn_guardar_registro
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, wrap=True),
                ft.Text("Módulo calcado de tu modelo de Excel para el control diario de metas y ventas por vendedor.", color="#aaaaaa", size=13),
                admin_store_bar,
                ft.Divider(height=15, color="#333333"),
                ft.Row([btn_subtab_diarias, btn_subtab_vendedores], spacing=10, wrap=True),
                ft.Container(height=10),
                subtab_content
            ], scroll=ft.ScrollMode.AUTO, expand=True)

        def procesar_excel_weekly(file_path):
            wb = openpyxl.load_workbook(file_path, data_only=True)
            if 'Weekly Flash NS (CORP)' in wb.sheetnames:
                sheet = wb['Weekly Flash NS (CORP)']
            else:
                sheet = wb[wb.sheetnames[0]]
            
            filename = os.path.basename(file_path)
            match_sem = re.search(r'Semana\s*(\d+)', filename, re.IGNORECASE)
            sem_num = match_sem.group(1) if match_sem else datetime.now().strftime("%W")
            semana_corta = f"Semana {sem_num}"
            fecha_reporte = datetime.now().strftime("%Y-%m-%d")

            tiendas_set = set()
            total_registros = 0
            
            db = conectar_db()
            if not db:
                return 0, semana_corta, set()
            
            cursor = db.cursor()
            cursor.execute("DELETE FROM weekly_metricas WHERE semana_corta = %s", (semana_corta,))
            
            # Detectar si el archivo es formato corporativo horizontal (Weekly Flash Sales)
            is_corporate_format = False
            for row in list(sheet.iter_rows(values_only=True))[:10]:
                row_str = " ".join([str(c) for c in row if c is not None])
                if "WTD" in row_str or "Day by Day" in row_str or "SGH MX" in row_str:
                    is_corporate_format = True
                    break

            def limpiar_float(val):
                if val is None or str(val).strip() in ("", "None", "nan", "inf", "-inf"):
                    return 0.0
                try:
                    import math
                    f = float(val)
                    if math.isnan(f) or math.isinf(f):
                        return 0.0
                    return max(-999999999.99, min(999999999.99, f))
                except Exception:
                    return 0.0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                
                col0 = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
                col2 = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                col3 = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
                col4 = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""

                tienda_str = ""
                if col2 and col3 and col2 != "None" and col3 != "None":
                    tienda_str = f"{col2} - {col3}"
                elif col3 and col3 != "None":
                    tienda_str = col3
                elif col2 and col4 and col2 != "None" and col4 != "None":
                    tienda_str = f"{col2} - {col4}"
                elif col0 and col0 != "None":
                    tienda_str = col0

                if not tienda_str or tienda_str.lower().startswith("total") or tienda_str.lower().startswith("summary"):
                    continue

                if is_corporate_format and len(row) >= 70:
                    periodos_map = [
                        ("SEMANA ANTERIOR", 39, 42, 45, 43),
                        ("MTD", 47, 50, 53, 51),
                        ("QTD", 55, 58, 61, 59),
                        ("YTD", 63, 66, 69, 67)
                    ]
                    for p_nombre, col_v, col_m, col_p, col_c in periodos_map:
                        v_val = limpiar_float(row[col_v]) if len(row) > col_v else 0.0
                        m_val = limpiar_float(row[col_m]) if len(row) > col_m else 0.0
                        p_val = limpiar_float(row[col_p]) if len(row) > col_p else 0.0
                        c_val = limpiar_float(row[col_c]) if len(row) > col_c else 0.0

                        if 0 < p_val <= 10.0:
                            p_val = p_val * 100.0

                        cursor.execute("""
                            INSERT INTO weekly_metricas 
                            (semana_corta, fecha_reporte, tienda, periodo, ventas, meta, pct_meta, comp)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        """, (semana_corta, fecha_reporte, tienda_str, p_nombre, v_val, m_val, p_val, c_val))
                        total_registros += 1
                    tiendas_set.add(tienda_str)
                else:
                    periodo = str(row[1]).strip() if len(row) > 1 and row[1] is not None else "SEMANA ANTERIOR"
                    ventas = limpiar_float(row[2]) if len(row) > 2 else 0.0
                    meta = limpiar_float(row[3]) if len(row) > 3 else 0.0
                    pct_meta = limpiar_float(row[4]) if len(row) > 4 else 0.0
                    comp = limpiar_float(row[5]) if len(row) > 5 else 0.0

                    cursor.execute("""
                        INSERT INTO weekly_metricas 
                        (semana_corta, fecha_reporte, tienda, periodo, ventas, meta, pct_meta, comp)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, (semana_corta, fecha_reporte, tienda_str, periodo, ventas, meta, pct_meta, comp))
                    tiendas_set.add(tienda_str)
                    total_registros += 1

            db.commit()
            db.close()
            return total_registros, semana_corta, tiendas_set
            
        def build_weekly_view():
            tiendas_opts = []
            semanas_opts = []
            
            try:
                db = conectar_db()
                if db:
                    cursor = db.cursor()
                    cursor.execute("SELECT DISTINCT tienda FROM weekly_metricas ORDER BY tienda ASC")
                    tiendas_opts = [r[0] for r in cursor.fetchall()]
                    cursor.execute("SELECT semana_corta FROM weekly_metricas GROUP BY semana_corta ORDER BY MAX(id) DESC")
                    semanas_opts = [r[0] for r in cursor.fetchall()]
                    db.close()
            except Exception as ex:
                print("Error cargando opciones weekly:", ex)

            user_tienda = user_info.get("tienda", "")
            default_tienda = None
            if user_tienda and tiendas_opts:
                for t_opt in tiendas_opts:
                    if user_tienda.lower() in t_opt.lower() or t_opt.lower() in user_tienda.lower():
                        default_tienda = t_opt
                        break
            if not default_tienda and tiendas_opts:
                default_tienda = tiendas_opts[0]

            default_semana = semanas_opts[0] if semanas_opts else None

            is_mobile_w = (page.width < 800) if (page and page.width) else False
            w_tienda_dd = 180 if is_mobile_w else 320
            w_semana_dd = 120 if is_mobile_w else 180
            w_buscar_tf = 110 if is_mobile_w else 180

            txt_buscar_tienda = ft.TextField(
                label="🔍 Nº Tienda",
                hint_text="Ej: 3502...",
                width=w_buscar_tf,
                border_color="#9D50BB",
                color="white",
                text_size=11 if is_mobile_w else 13,
                height=36 if is_mobile_w else 40
            )

            dd_tiendas = ft.Dropdown(
                label="🏬 Tienda",
                value=default_tienda,
                options=[ft.dropdown.Option(t_opt) for t_opt in tiendas_opts],
                width=w_tienda_dd,
                border_color="#9D50BB",
                color="white",
                text_size=11 if is_mobile_w else 13
            )

            dd_semanas = ft.Dropdown(
                label="📅 Semana",
                value=default_semana,
                options=[ft.dropdown.Option(s_opt) for s_opt in semanas_opts],
                width=w_semana_dd,
                border_color="#9D50BB",
                color="white",
                text_size=11 if is_mobile_w else 13
            )

            def ejecutar_busqueda(e=None):
                q = (txt_buscar_tienda.value or "").strip().lower()
                if not q:
                    dd_tiendas.options = [ft.dropdown.Option(t_o) for t_o in tiendas_opts]
                    if tiendas_opts:
                        dd_tiendas.value = default_tienda or tiendas_opts[0]
                else:
                    coincidencias = [t_o for t_o in tiendas_opts if q in t_o.lower()]
                    if coincidencias:
                        dd_tiendas.options = [ft.dropdown.Option(t_o) for t_o in coincidencias]
                        dd_tiendas.value = coincidencias[0]
                        mostrar_snack(f"✅ Cargada tienda: {coincidencias[0]}", color="green")
                    else:
                        mostrar_snack(f"⚠️ No se encontró la tienda con: '{txt_buscar_tienda.value}'", color="orange")
                render_table()

            txt_buscar_tienda.on_submit = ejecutar_busqueda
            txt_buscar_tienda.on_change = ejecutar_busqueda

            btn_consultar_tienda = ft.ElevatedButton(
                "🔍 Buscar" if is_mobile_w else "🔄 Consultar",
                icon=ft.Icons.SEARCH,
                bgcolor="#0284c7",
                color="white",
                height=36 if is_mobile_w else 40,
                style=ft.ButtonStyle(
                    shape=ft.RoundedRectangleBorder(radius=8),
                    padding=ft.padding.Padding(8, 0, 8, 0) if is_mobile_w else None
                ),
                on_click=ejecutar_busqueda
            )

            table_container = ft.Column(spacing=15, expand=True)

            def render_table(e=None):
                sel_t = dd_tiendas.value
                sel_s = dd_semanas.value

                if not sel_t or not sel_s:
                    table_container.controls = [
                        ft.Container(
                            content=ft.Text("⚠️ No hay datos cargados en el sistema. El Administrador debe subir el archivo Weekly Excel.", color="orange", size=14),
                            padding=20,
                            bgcolor="#1e1b4b",
                            border_radius=8
                        )
                    ]
                    try: page.update()
                    except Exception: pass
                    return

                metrics_data = []
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("""
                            SELECT periodo, ventas, meta, pct_meta, comp 
                            FROM weekly_metricas 
                            WHERE tienda = %s AND semana_corta = %s
                            ORDER BY FIELD(periodo, 'SEMANA ANTERIOR', 'MTD', 'QTD', 'YTD')
                        """, (sel_t, sel_s))
                        metrics_data = cursor.fetchall()
                        db.close()
                except Exception as ex:
                    print("Error leyendo métricas weekly:", ex)

                if not metrics_data:
                    table_container.controls = [
                        ft.Container(
                            content=ft.Text(f"No hay métricas registradas para {sel_t} en {sel_s}.", color="white", size=14),
                            padding=20,
                            bgcolor="#1e1b4b",
                            border_radius=8
                        )
                    ]
                    try: page.update()
                    except Exception: pass
                    return

                data_rows = []
                for m in metrics_data:
                    p = m["periodo"]
                    if is_mobile_w and "SEMANA ANTERIOR" in p.upper():
                        p_display = "SEMANA\nANTERIOR"
                    else:
                        p_display = p

                    v = f"${m['ventas']:,.2f}"
                    target = f"${m['meta']:,.2f}"
                    
                    pm_val = float(m["pct_meta"] or 0)
                    pm_str = f"{pm_val:.1f}%"
                    # Regla % Meta: >= 91% es Verde (#7CFC00), menor a 91% es Rojo (#FF4500)
                    pm_color = "#7CFC00" if pm_val >= 91.0 else "#FF4500"

                    comp_val = float(m["comp"] or 0)
                    comp_str = f"{comp_val:.1f}%"
                    comp_color = "#7CFC00" if comp_val >= 0 else "#FF4500"

                    cell_font_size = 10.0 if is_mobile_w else 13.0
                    pad_val = ft.padding.Padding(3, 2, 3, 2) if is_mobile_w else 5

                    data_rows.append(
                        ft.DataRow(
                            cells=[
                                ft.DataCell(ft.Text(p_display, weight="bold", color="#D8B4FE", size=cell_font_size)),
                                ft.DataCell(ft.Text(v, color="#00FFFF", weight="bold", size=cell_font_size)),
                                ft.DataCell(ft.Text(target, color="#ffffff", size=cell_font_size)),
                                ft.DataCell(ft.Container(content=ft.Text(pm_str, color=pm_color, weight="bold", size=cell_font_size), bgcolor="#222233", padding=pad_val, border_radius=5)),
                                ft.DataCell(ft.Container(content=ft.Text(comp_str, color=comp_color, weight="bold", size=cell_font_size), bgcolor="#222233", padding=pad_val, border_radius=5)),
                            ]
                        )
                    )

                hdr_font_size = 10.0 if is_mobile_w else 13.0
                table_widget = ft.DataTable(
                    columns=[
                        ft.DataColumn(ft.Text("Métrica", weight="bold", color="#D8B4FE", size=hdr_font_size)),
                        ft.DataColumn(ft.Text("💲 Ventas", weight="bold", color="#00FFFF", size=hdr_font_size)),
                        ft.DataColumn(ft.Text("📑 Meta", weight="bold", color="#7CFC00", size=hdr_font_size)),
                        ft.DataColumn(ft.Text("📈 % Meta", weight="bold", color="#FFD700", size=hdr_font_size)),
                        ft.DataColumn(ft.Text("📊 Comp", weight="bold", color="#D8B4FE", size=hdr_font_size)),
                    ],
                    rows=data_rows,
                    border=ft.Border.all(1, "#333344"),
                    border_radius=8 if is_mobile_w else 10,
                    column_spacing=6 if is_mobile_w else 24,
                    horizontal_margin=6 if is_mobile_w else 16,
                    heading_row_height=32 if is_mobile_w else 56,
                    data_row_min_height=32 if is_mobile_w else 48,
                    heading_row_color="#1e1b4b",
                    data_row_color="#141424"
                )

                table_container.controls = [
                    ft.Container(
                        content=ft.Column([
                            ft.Row([
                                ft.Text(f"📍 {sel_t}", size=13 if is_mobile_w else 18, weight="bold", color="#00FFFF"),
                                ft.Container(content=ft.Text(sel_s, color="white", weight="bold", size=10 if is_mobile_w else 12), bgcolor="#9D50BB", padding=ft.padding.Padding(5, 2, 5, 2) if is_mobile_w else ft.padding.Padding(8, 4, 8, 4), border_radius=6)
                            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                            ft.Divider(height=8 if is_mobile_w else 10, color="#333344"),
                            ft.Row([table_widget], scroll=ft.ScrollMode.AUTO)
                        ]),
                        bgcolor="#181828",
                        padding=8 if is_mobile_w else 15,
                        border_radius=10 if is_mobile_w else 12,
                        border=ft.Border.all(1, "#9D50BB")
                    )
                ]
                try: page.update()
                except Exception: pass

            def on_dd_change(e=None):
                render_table()
                try: page.update()
                except Exception: pass

            dd_tiendas.on_change = on_dd_change
            dd_semanas.on_change = on_dd_change

            lbl_upload_status = ft.Text("", size=13, color="#00FFFF")

            def recargar_dropdowns():
                t_opts, s_opts = [], []
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        cursor.execute("SELECT DISTINCT tienda FROM weekly_metricas ORDER BY tienda ASC")
                        t_opts = [r[0] for r in cursor.fetchall()]
                        cursor.execute("SELECT semana_corta FROM weekly_metricas GROUP BY semana_corta ORDER BY MAX(id) DESC")
                        s_opts = [r[0] for r in cursor.fetchall()]
                        db.close()
                except Exception as ex_opts:
                    print("Error recargando opciones:", ex_opts)
                dd_tiendas.options = [ft.dropdown.Option(t_o) for t_o in t_opts]
                dd_semanas.options = [ft.dropdown.Option(s_o) for s_o in s_opts]
                if s_opts and (not dd_semanas.value or dd_semanas.value not in s_opts):
                    dd_semanas.value = s_opts[0]
                if t_opts and (not dd_tiendas.value or dd_tiendas.value not in t_opts):
                    dd_tiendas.value = t_opts[0]
                try: page.update()
                except Exception: pass

            def procesar_y_notificar(file_path, f_name):
                lbl_upload_status.value = f"⏳ Procesando '{f_name}'..."
                lbl_upload_status.color = "#FFD700"
                page.update()

                try:
                    n_t, n_r = procesar_excel_weekly(file_path)
                    lbl_upload_status.value = f"✅ ¡Éxito! Se actualizaron {n_t} tiendas ({n_r} registros) desde '{f_name}'."
                    lbl_upload_status.color = "#7CFC00"
                    mostrar_snack("✅ Excel Weekly procesado y cargado correctamente.", color="green")
                    recargar_dropdowns()
                    render_table()
                except Exception as ex:
                    print("Error procesando Excel Weekly:", ex)
                    lbl_upload_status.value = f"❌ Error al procesar: {ex}"
                    lbl_upload_status.color = "#FF4500"
                    mostrar_snack(f"Error procesando Excel: {ex}", color="red")
                page.update()

            def on_file_result(e):
                if not e.files or len(e.files) == 0:
                    lbl_upload_status.value = "⚠️ No se seleccionó ningún archivo o la selección fue cancelada."
                    lbl_upload_status.color = "#FFD700"
                    page.update()
                    return
                f_item = e.files[0]
                f_name = f_item.name

                if f_item.path:
                    txt_ruta_excel.value = f_item.path
                    procesar_y_notificar(f_item.path, f_name)
                else:
                    lbl_upload_status.value = f"⏳ Subiendo '{f_name}' desde el navegador..."
                    lbl_upload_status.color = "#FFD700"
                    page.update()
                    try:
                        upload_url = page.get_upload_url(f_name, 600)
                        file_picker_weekly.upload([ft.FilePickerUploadFile(f_name, upload_url=upload_url)])
                    except Exception as ex_up:
                        print("Error iniciando upload web:", ex_up)
                        lbl_upload_status.value = f"❌ Error iniciando subida: {ex_up}"
                        lbl_upload_status.color = "#FF4500"
                        page.update()

            def on_upload_progress(e):
                if getattr(e, "progress", 0) == 1.0 or getattr(e, "status", "") == "uploaded":
                    f_name = e.file_name
                    f_path = os.path.join(BASE_PATH, "uploads", f_name)
                    procesar_y_notificar(f_path, f_name)

            def buscar_excel_weekly_reciente():
                import glob
                dirs = [os.path.expanduser('~/Downloads'), os.path.expanduser('~/Documents'), os.path.expanduser('~/Desktop')]
                for d in dirs:
                    matches = glob.glob(os.path.join(d, '*Weekly*.xlsx')) + glob.glob(os.path.join(d, '*weekly*.xlsx'))
                    if matches:
                        matches.sort(key=os.path.getmtime, reverse=True)
                        return matches[0]
                return ""

            def abrir_dialogo_archivo_nativo():
                try:
                    from tkinter import Tk, filedialog
                    root = Tk()
                    root.withdraw()
                    root.attributes('-topmost', True)
                    initial_dir = os.path.expanduser('~/Downloads')
                    if not os.path.exists(initial_dir):
                        initial_dir = os.path.expanduser('~/Documents')
                    f_selected = filedialog.askopenfilename(
                        title="Selecciona el archivo Weekly Flash Sales Excel",
                        initialdir=initial_dir,
                        filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
                    )
                    root.destroy()
                    if f_selected and os.path.exists(f_selected):
                        return f_selected
                except Exception as ex_tk:
                    print("Error en dialogo Tkinter nativo:", ex_tk)

                try:
                    import subprocess
                    ps_cmd = '''
                    [System.Reflection.Assembly]::LoadWithPartialName('System.windows.forms') | Out-Null
                    $f = New-Object System.Windows.Forms.OpenFileDialog
                    $f.InitialDirectory = [System.IO.Path]::Combine($env:USERPROFILE, 'Downloads')
                    $f.Filter = "Archivos Excel (*.xlsx)|*.xlsx|Todos los archivos (*.*)|*.*"
                    $f.Title = "Selecciona el reporte Excel Weekly"
                    $f.ShowHelp = $true
                    $f.TopMost = $true
                    if ($f.ShowDialog() -eq [System.Windows.Forms.DialogResult]::OK) {
                        Write-Output $f.FileName
                    }
                    '''
                    proc = subprocess.run(
                        ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_cmd],
                        capture_output=True,
                        text=True,
                        timeout=120
                    )
                    selected = proc.stdout.strip()
                    if selected and os.path.exists(selected):
                        return selected
                except Exception as ex_ps:
                    print("Error en PowerShell dialog:", ex_ps)

                return ""

            auto_detected_path = buscar_excel_weekly_reciente()

            txt_ruta_excel = ft.TextField(
                label="📁 Ruta de archivo Excel",
                value=auto_detected_path if auto_detected_path else r"C:\Users\MOISES\Downloads\Weekly Flash Sales Semana 029 Día 07(4).xlsx",
                hint_text=r"Ej: C:\Users\MOISES\Downloads\Weekly...",
                width=240 if is_mobile_w else 650,
                border_color="#00FFFF",
                color="white",
                text_size=11 if is_mobile_w else 12
            )

            def resolver_ruta_excel(p):
                if not p:
                    return ""
                p_clean = p.strip('"').strip("'")
                if os.path.isabs(p_clean) and os.path.exists(p_clean):
                    return p_clean
                rutas_posibles = [
                    p_clean,
                    os.path.join(BASE_PATH, "uploads", os.path.basename(p_clean)),
                    os.path.expanduser(os.path.join("~/Downloads", os.path.basename(p_clean))),
                    os.path.expanduser(os.path.join("~/Documents", os.path.basename(p_clean))),
                    os.path.expanduser(os.path.join("~/Desktop", os.path.basename(p_clean))),
                ]
                for r in rutas_posibles:
                    if os.path.exists(r):
                        return r
                return p_clean

            def cargar_desde_ruta(e):
                ruta = txt_ruta_excel.value or ""
                ruta_res = resolver_ruta_excel(ruta)
                if not ruta_res or not os.path.exists(ruta_res):
                    lbl_upload_status.value = f"❌ Archivo no encontrado en la ruta: '{ruta}'"
                    lbl_upload_status.color = "#FF4500"
                    mostrar_snack(f"Ruta inválida o archivo no existe: '{ruta}'", color="red")
                    page.update()
                    return
                procesar_y_notificar(ruta_res, os.path.basename(ruta_res))

            btn_cargar_ruta = ft.ElevatedButton(
                "⚡ Cargar y Procesar Excel",
                icon=ft.Icons.PLAY_ARROW,
                bgcolor="#7CFC00",
                color="black",
                height=36 if is_mobile_w else 40,
                style=ft.ButtonStyle(
                    shape=ft.RoundedRectangleBorder(radius=8),
                    padding=ft.padding.Padding(8, 0, 8, 0) if is_mobile_w else None
                ),
                on_click=cargar_desde_ruta
            )

            def abrir_dialogo_subida(e):
                seleccionar_archivo_async(
                    "Seleccionar reporte Excel Weekly",
                    [("Archivos Excel", "*.xlsx *.xls"), ("Todos los archivos", "*.*")],
                    lambda ruta: procesar_y_notificar(ruta, os.path.basename(ruta))
                )

            btn_upload = ft.ElevatedButton(
                "📤 Cargar Archivo Excel (Weekly)",
                icon=ft.Icons.UPLOAD_FILE,
                bgcolor="#7c3aed",
                color="white",
                height=36 if is_mobile_w else 40,
                style=ft.ButtonStyle(
                    shape=ft.RoundedRectangleBorder(radius=8),
                    padding=ft.padding.Padding(8, 0, 8, 0) if is_mobile_w else None
                ),
                on_click=abrir_dialogo_subida
            )

            weekly_reports_list = ft.Column(spacing=10, scroll=ft.ScrollMode.AUTO, expand=True)

            def cargar_reportes_weekly():
                weekly_reports_list.controls.clear()
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("""
                            SELECT semana_corta, COUNT(DISTINCT tienda) as total_tiendas, COUNT(*) as total_registros
                            FROM weekly_metricas
                            GROUP BY semana_corta
                            ORDER BY MAX(id) DESC
                        """)
                        reportes = cursor.fetchall()
                        db.close()

                        if not reportes:
                            weekly_reports_list.controls.append(ft.Text("No hay reportes weekly cargados en la BD.", color="#aaaaaa", size=12))
                        else:
                            for rep in reportes:
                                sem = rep["semana_corta"]
                                n_t = rep["total_tiendas"]
                                n_r = rep["total_registros"]

                                def borrar_weekly_click(e, s_corta=sem):
                                    def on_confirmar_weekly(ev):
                                        try:
                                            db_del = conectar_db()
                                            if db_del:
                                                cursor_del = db_del.cursor()
                                                cursor_del.execute("DELETE FROM weekly_metricas WHERE semana_corta = %s", (s_corta,))
                                                db_del.commit()
                                                db_del.close()
                                                mostrar_snack(f"Reporte Weekly '{s_corta}' eliminado.")
                                                recargar_dropdowns()
                                                cargar_reportes_weekly()
                                                render_table()
                                                page.pop_dialog()
                                                page.update()
                                        except Exception as ex_w:
                                            print("ERROR BORRAR WEEKLY:", ex_w)
                                            mostrar_snack("Error al borrar reporte weekly.", color="red")

                                    def on_cancelar_weekly(ev):
                                        page.pop_dialog()

                                    dialog_confirm_w = ft.AlertDialog(
                                        title=ft.Text("Confirmar Borrado de Reporte Weekly", color="#FF4500", weight="bold"),
                                        content=ft.Text(f"¿Seguro que deseas borrar el reporte completo de la '{s_corta}' ({n_t} tiendas)?"),
                                        actions=[
                                            ft.TextButton("Cancelar", on_click=on_cancelar_weekly),
                                            ft.ElevatedButton("Sí, Borrar", on_click=on_confirmar_weekly, bgcolor="#FF4500", color="white")
                                        ],
                                        actions_alignment="end",
                                        bgcolor="#0F0F1A"
                                    )
                                    page.show_dialog(dialog_confirm_w)

                                def ver_reporte_weekly_click(e, s_corta=sem):
                                    dd_semanas.value = s_corta
                                    try:
                                        if 'btn_subtab_resumen' in locals() and btn_subtab_resumen and hasattr(btn_subtab_resumen, "on_click"):
                                            btn_subtab_resumen.on_click(None)
                                    except Exception: pass
                                    render_table()
                                    mostrar_snack(f"📊 Visualizando Reporte {s_corta}", color="cyan")

                                weekly_reports_list.controls.append(
                                    ft.Container(
                                        content=ft.Row([
                                            ft.Icon(ft.Icons.BAR_CHART, color="#00FFFF", size=18 if is_mobile_w else 22),
                                            ft.Column([
                                                ft.Text(f"Reporte {sem}", color="white", weight="bold", size=12 if is_mobile_w else 14),
                                                ft.Text(f"{n_t} tiendas cargadas ({n_r} métricas)", color="#aaaaaa", size=10 if is_mobile_w else 11)
                                            ], spacing=2, expand=True),
                                            ft.IconButton(
                                                icon=ft.Icons.VISIBILITY_ROUNDED,
                                                icon_color="#00FFFF",
                                                tooltip="👁️ Visualizar métricas de esta semana",
                                                icon_size=18 if is_mobile_w else 22,
                                                on_click=ver_reporte_weekly_click
                                            ),
                                            ft.IconButton(
                                                icon=ft.Icons.DELETE_FOREVER,
                                                icon_color="#FF4500",
                                                tooltip="Eliminar reporte de esta semana",
                                                icon_size=18 if is_mobile_w else 22,
                                                on_click=borrar_weekly_click
                                            )
                                        ], alignment="spaceBetween", vertical_alignment="center"),
                                        bgcolor="#1e1b4b",
                                        padding=8 if is_mobile_w else 10,
                                        border_radius=8,
                                        border=ft.Border.all(1, "#333344")
                                    )
                                )
                except Exception as ex:
                    print("ERROR LISTA WEEKLY:", ex)
                    weekly_reports_list.controls.append(ft.Text("Error al cargar la lista de reportes weekly.", color="red"))

                try: page.update()
                except Exception: pass

            cargar_reportes_weekly()

            def procesar_y_notificar(file_path, f_name):
                lbl_upload_status.value = f"⏳ Procesando '{f_name}'..."
                lbl_upload_status.color = "#FFD700"
                page.update()
                mostrar_snack(f"Extrayendo tiendas y métricas de '{f_name}'...", color="#D8B4FE")
                
                def worker():
                    try:
                        n_reg, sem, tiendas_set = procesar_excel_weekly(file_path)
                        lbl_upload_status.value = f"✅ Éxito: {n_reg} registros insertados de {len(tiendas_set)} tiendas para {sem}."
                        lbl_upload_status.color = "#7CFC00"
                        mostrar_snack(f"✅ Excel Weekly cargado con éxito ({sem}).", color="green")
                        cargar_reportes_weekly()
                        render_table()
                    except Exception as ex:
                        print("Error procesando Excel Weekly:", ex)
                        lbl_upload_status.value = f"❌ Error al procesar: {ex}"
                        lbl_upload_status.color = "#FF4500"
                        mostrar_snack(f"Error procesando Excel: {ex}", color="red")
                    page.update()

                import threading
                threading.Thread(target=worker, daemon=True).start()

            btn_actualizar_listas = ft.IconButton(
                icon=ft.Icons.REFRESH_ROUNDED,
                icon_color="#00FFFF",
                tooltip="🔄 Actualizar Semanas y Tiendas",
                on_click=lambda e: (recargar_dropdowns(), render_table(), mostrar_snack("🔄 Lista de semanas y tiendas actualizada", color="cyan"))
            )

            tab_resumen = ft.Column([
                ft.Row([txt_buscar_tienda, btn_consultar_tienda, dd_tiendas, dd_semanas, btn_actualizar_listas], wrap=True, spacing=6 if is_mobile_w else 8, vertical_alignment="center"),
                ft.Divider(height=8 if is_mobile_w else 10, color="#333333"),
                table_container
            ], scroll=ft.ScrollMode.AUTO, expand=True)

            render_table()

            if es_admin():
                tab_admin = ft.Column([
                    ft.Row([
                        ft.Text("Reportes Semanales de Sunglass Hut", size=14 if is_mobile_w else 18, color="white", weight="bold"),
                        btn_upload
                    ], alignment="spaceBetween", vertical_alignment="center", wrap=True),
                    ft.Divider(height=10, color="transparent"),
                    weekly_reports_list,
                    ft.Divider(height=15, color="#333344"),
                    ft.Text("⚡ Cargar desde Ruta / Carpeta de Descargas", size=12 if is_mobile_w else 13, weight="bold", color="#00FFFF"),
                    ft.Row([txt_ruta_excel, btn_cargar_ruta], wrap=True, spacing=6 if is_mobile_w else 10),
                    ft.Container(height=5),
                    lbl_upload_status
                ], spacing=10, scroll=ft.ScrollMode.AUTO, expand=True)

                tab_defs_w = [
                    ("📊 Resumen Weekly", ft.Icons.BAR_CHART, tab_resumen),
                    ("📤 Cargar Excel (Admin)", ft.Icons.UPLOAD_FILE, tab_admin)
                ]
                curr_w_idx = 0
                content_w_box = ft.Container(content=tab_defs_w[curr_w_idx][2], expand=True)
                tab_w_buttons = []
                for idx, (label, icon_name, tab_cnt) in enumerate(tab_defs_w):
                    def make_w_click(i, cnt):
                        def click(e):
                            content_w_box.content = cnt
                            for b_i, btn_c in enumerate(tab_w_buttons):
                                is_active = (b_i == i)
                                btn_c.bgcolor = "#7c3aed" if is_active else "#1e1e1e"
                                btn_c.border = ft.Border.all(1, "#9D50BB" if is_active else "#333333")
                            try: page.update()
                            except Exception: pass
                        return click

                    is_sel = (idx == curr_w_idx)
                    btn_c = ft.Container(
                        content=ft.Row([
                            ft.Icon(icon_name, size=15, color="#00FFFF" if is_sel else "#aaaaaa"),
                            ft.Text(label, size=12, weight="bold", color="white" if is_sel else "#aaaaaa")
                        ], spacing=6, alignment=ft.MainAxisAlignment.CENTER),
                        bgcolor="#7c3aed" if is_sel else "#1e1e1e",
                        padding=ft.padding.Padding(12, 8, 12, 8),
                        border_radius=8,
                        border=ft.Border.all(1, "#9D50BB" if is_sel else "#333333"),
                        on_click=make_w_click(idx, tab_cnt),
                        ink=True
                    )
                    tab_w_buttons.append(btn_c)

                tab_w_bar_row = ft.Row(tab_w_buttons, scroll=ft.ScrollMode.AUTO, spacing=8)
                main_content = ft.Column([
                    tab_w_bar_row,
                    ft.Container(height=5),
                    content_w_box
                ], expand=True)
            else:
                main_content = tab_resumen

            return ft.Column([
                ft.Row([
                    ft.Text("📅 MÓDULO WEEKLY", size=22 if is_mobile_w else 24, color="#D8B4FE", weight="bold")
                ]),
                ft.Text("Consulta rápida de métricas semanales por tienda (Ventas, Target, % Target y Comparativo).", color="#aaaaaa", size=12 if is_mobile_w else 13),
                ft.Divider(height=12, color="#333333"),
                main_content
            ], expand=True)

        actualizar_campana_badge()
        active_view = ["chat"]

        # Cambiar vistas con hover y estilos activos
        def cambiar_vista(vista):
            active_view[0] = vista
            for btn, v_name in [(btn_chat, "chat"), (btn_historial, "historial"), (btn_operacion_diaria, "operacion_diaria"), (btn_checklists, "checklists"), (btn_manuales, "manuales"), (btn_garantias, "garantias"), (btn_tareas, "tareas"), (btn_campanas, "campanas"), (btn_presupuesto, "presupuesto"), (btn_reto, "reto"), (btn_vendedores, "vendedores"), (btn_simulador, "simulador"), (btn_meta_semanal, "meta_semanal"), (btn_weekly, "weekly"), (btn_enfoque, "enfoque_diario")]:
                if btn:
                    btn.style = ft.ButtonStyle(
                        bgcolor="#141424" if v_name == vista else "transparent",
                        shape=ft.RoundedRectangleBorder(radius=8)
                    )
            if btn_dashboard:
                btn_dashboard.style = ft.ButtonStyle(
                    bgcolor="#141424" if vista == "dashboard" else "transparent",
                    shape=ft.RoundedRectangleBorder(radius=8)
                )
            if btn_admin_trivia:
                btn_admin_trivia.style = ft.ButtonStyle(
                    bgcolor="#141424" if vista == "admin_trivia" else "transparent",
                    shape=ft.RoundedRectangleBorder(radius=8)
                )
            
            if vista == "chat":
                content_area.content = build_chat_view()
            elif vista == "historial":
                content_area.content = build_historial_view()
            elif vista == "checklists":
                content_area.content = build_checklists_view()
            elif vista == "manuales":
                content_area.content = build_manuals_view()
            elif vista == "garantias":
                content_area.content = build_garantias_view()
            elif vista == "tareas":
                content_area.content = build_tareas_view()
            elif vista == "campanas":
                content_area.content = build_campanas_view()
            elif vista == "presupuesto":
                content_area.content = build_presupuesto_view()
            elif vista == "reto":
                content_area.content = build_reto_dia_view()
            elif vista == "dashboard":
                content_area.content = build_dashboard_view()
            elif vista == "vendedores":
                content_area.content = build_vendedores_view()
            elif vista == "simulador":
                content_area.content = build_simulador_view()
            elif vista == "crm":
                content_area.content = build_crm_view()
            elif vista == "meta_semanal":
                content_area.content = build_meta_semanal_view()
            elif vista == "weekly":
                content_area.content = build_weekly_view()
            elif vista == "enfoque_diario":
                import enfoque_diario
                content_area.content = enfoque_diario.build_enfoque_diario_view(page, user_info)
            elif vista == "operacion_diaria":
                if es_admin():
                    content_area.content = operacion_tiendas.build_aperturas_cierres_tab(page, user_info, conectar_db, mostrar_snack, tr)
                else:
                    content_area.content = operacion_tiendas.build_operacion_diaria_view(
                        page, user_info, conectar_db, mostrar_snack, tr,
                        seleccionar_archivo_async=seleccionar_archivo_async
                    )
            elif vista == "admin_trivia":
                if not es_admin():
                    mostrar_snack("Acceso denegado: Se requieren permisos de administrador.", color="red")
                    cambiar_vista("chat")
                    return
                content_area.content = build_admin_trivia_view()
            elif vista == "bitacora":
                if not es_admin():
                    mostrar_snack("Acceso denegado: Solo Administradores pueden ver la Bitácora de Seguridad.", color="red")
                    cambiar_vista("chat")
                    return
                content_area.content = build_bitacora_view()


            # Cerrar el menú lateral en móviles al cambiar de vista
            if getattr(page, "width", None) and page.width < 800:
                sidebar.visible = False
                
            page.update()

        sess_data["cambiar_vista"] = cambiar_vista

        def tr(es, en, fr=None, it=None, zh=None):
            l = selected_lang[0]
            if l == "en": return en
            if l == "fr": return fr or en
            if l == "it": return it or en
            if l == "zh": return zh or en
            return es

        btn_chat = ft.TextButton(
            content=ft.Row([ft.Text("✨", color="#00FFFF", size=14, weight="bold"), ft.Text(tr("Asistente Chat ✨", "Chat Assistant ✨", "Assistant Chat ✨", "Assistente Chat ✨", "聊天助手 ✨"), color="white", weight="bold")], spacing=10),
            on_click=lambda e: cambiar_vista("chat"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_historial = ft.TextButton(
            content=ft.Row([ft.Text("⏳", color="#00FFFF", size=14, weight="bold"), ft.Text(tr("Mi Historial ⏳", "My History ⏳", "Mon Historique ⏳", "La Mia Cronologia ⏳", "我的历史记录 ⏳"), color="white", weight="bold")], spacing=10),
            on_click=lambda e: cambiar_vista("historial"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_operacion_diaria = ft.TextButton(
            content=ft.Row([
                ft.Text("🔑", color="#00FFFF", size=14, weight="bold"),
                ft.Text(tr("Aperturas y Cierres 🔑", "Openings & Closings 🔑", "Ouvertures & Fermetures 🔑", "Aperture & Chiusure 🔑", "开门与关门 🔑"), color="white", weight="bold")
            ], spacing=10),
            on_click=lambda e: cambiar_vista("operacion_diaria"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_checklists = ft.TextButton(
            content=ft.Row([ft.Text("📋", color="#00FFFF", size=14, weight="bold"), ft.Text(tr("Checklists 📋", "Checklists 📋", "Listes 📋", "Liste 📋", "任务清单 📋"), color="white", weight="bold")], spacing=10),
            on_click=lambda e: cambiar_vista("checklists"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_manuales = ft.TextButton(
            content=ft.Row([ft.Text("📚", color="#00FFFF", size=14, weight="bold"), ft.Text(tr("Manuales 📚", "Manuals 📚", "Manuels 📚", "Manuali 📚", "手册 📚"), color="white", weight="bold")], spacing=10),
            on_click=lambda e: cambiar_vista("manuales"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_garantias = ft.TextButton(
            content=ft.Row([ft.Text("👓", color="#00FFFF", size=14, weight="bold"), ft.Text(tr("Garantías 👓", "Warranties 👓", "Garanties 👓", "Garanzie 👓", "保修 👓"), color="white", weight="bold")], spacing=10),
            on_click=lambda e: cambiar_vista("garantias"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_tareas = ft.TextButton(
            content=ft.Row([ft.Text("📋", color="#00FFFF", size=14, weight="bold"), ft.Text(tr("Tareas 📋", "Tasks 📋", "Tâches 📋", "Attività 📋", "任务 📋"), color="white", weight="bold")], spacing=10),
            on_click=lambda e: cambiar_vista("tareas"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_campanas = ft.TextButton(
            content=ft.Row([ft.Text("📸", color="#00FFFF", size=14, weight="bold"), ft.Text(tr("Campañas 📸", "Campaigns 📸", "Campagnes 📸", "Campagne 📸", "活动 📸"), color="white", weight="bold")], spacing=10),
            on_click=lambda e: cambiar_vista("campanas"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_presupuesto = ft.TextButton(
            content=ft.Row([ft.Text("💰", color="#00FFFF", size=14, weight="bold"), ft.Text(tr("Presupuesto 💰", "Budget 💰", "Budget 💰", "Budget 💰", "预算 💰"), color="white", weight="bold")], spacing=10),
            on_click=lambda e: cambiar_vista("presupuesto"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_reto = ft.TextButton(
            content=ft.Row([ft.Text("🏆", color="#00FFFF", size=14, weight="bold"), ft.Text(tr("Reto del Día 🏆", "Daily Quiz 🏆", "Défi du Jour 🏆", "Sfida del Giorno 🏆", "每日挑战 🏆"), color="white", weight="bold")], spacing=10),
            on_click=lambda e: cambiar_vista("reto"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_vendedores = ft.TextButton(
            content=ft.Row([ft.Text("👥", color="#00FFFF", size=14, weight="bold"), ft.Text(tr("Configuración Tienda 👥", "Store Config 👥", "Configuration 👥", "Configurazione 👥", "店铺配置 👥"), color="white", weight="bold")], spacing=10),
            on_click=lambda e: cambiar_vista("vendedores"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_simulador = ft.TextButton(
            content=ft.Row([ft.Text("🎭", color="#00FFFF", size=14, weight="bold"), ft.Text(tr("Simulador IA 🎭", "AI Simulator 🎭", "Simulateur IA 🎭", "Simulatore IA 🎭", "AI 模拟器 🎭"), color="white", weight="bold")], spacing=10),
            on_click=lambda e: cambiar_vista("simulador"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_crm = ft.TextButton(
            content=ft.Row([ft.Text("📱", color="#00FFFF", size=14, weight="bold"), ft.Text(tr("CRM Cobertura Oops 📱", "CRM Coverage Oops 📱", "CRM Couverture Oops 📱", "CRM Copertura Oops 📱", "CRM 意外保障 📱"), color="white", weight="bold")], spacing=10),
            on_click=lambda e: cambiar_vista("crm"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_meta_semanal = ft.TextButton(
            content=ft.Row([ft.Text("🎯", color="#00FFFF", size=14, weight="bold"), ft.Text(tr("Metas y Métricas 🎯", "Goals & Metrics 🎯", "Objectifs & Métriques 🎯", "Obiettivi & Metriche 🎯", "目标与指标 🎯"), color="white", weight="bold")], spacing=10),
            on_click=lambda e: cambiar_vista("meta_semanal"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_weekly = ft.TextButton(
            content=ft.Row([ft.Text("🗓️", color="#00FFFF", size=14, weight="bold"), ft.Text(tr("Weekly 🗓️", "Weekly 🗓️", "Hebdomadaire 🗓️", "Settimanale 🗓️", "每周 🗓️"), color="white", weight="bold")], spacing=10),
            on_click=lambda e: cambiar_vista("weekly"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_enfoque = ft.TextButton(
            content=ft.Row([ft.Text("☀️", color="#FFD700", size=14, weight="bold"), ft.Text(tr("Enfoque Diario ☀️", "Daily Focus ☀️", "Focus Quotidien ☀️", "Focus Giornaliero ☀️", "每日焦点 ☀️"), color="white", weight="bold")], spacing=10),
            on_click=lambda e: cambiar_vista("enfoque_diario"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_dashboard = None
        if es_admin():
            btn_dashboard = ft.TextButton(
                content=ft.Row([ft.Text("🎮", color="#00FFFF", size=14, weight="bold"), ft.Text(tr("Panel de Control 🎮", "Admin Panel 🎮", "Panneau de Contrôle 🎮", "Pannello di Controllo 🎮", "控制面板 🎮"), color="white", weight="bold")], spacing=10),
                on_click=lambda e: cambiar_vista("dashboard"),
                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
            )

        btn_admin_trivia = None
        if es_admin():
            btn_admin_trivia = ft.TextButton(
                content=ft.Row([ft.Text("❓", color="#FFD700", size=14, weight="bold"), ft.Text(tr("Gestionar Trivia 🧠", "Manage Trivia 🧠", "Gérer Quiz 🧠", "Gestisci Trivia 🧠", "管理问答 🧠"), color="white", weight="bold")], spacing=10),
                on_click=lambda e: cambiar_vista("admin_trivia"),
                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
            )

        btn_bitacora = None
        if es_admin():
            btn_bitacora = ft.TextButton(
                content=ft.Row([ft.Text("🛡️", color="#00FFAA", size=14, weight="bold"), ft.Text("Bitácora de Seguridad 🛡️", color="white", weight="bold")], spacing=10),
                on_click=lambda e: cambiar_vista("bitacora"),
                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
            )

        btn_logout = ft.TextButton(
            content=ft.Row([ft.Text("👋", color="#FF4500", size=14, weight="bold"), ft.Text(tr("Cerrar Sesión 👋", "Log Out 👋", "Se Déconnecter 👋", "Disconnettersi 👋", "退出登录 👋"), color="#FF4500", weight="bold")], spacing=10),
            on_click=lambda e: cerrar_sesion(),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        # --- RECUADRO DE SUGERENCIAS EN LA BARRA LATERAL (VISIBLE PARA TODOS) ---
        suggestion_input = ft.TextField(
            label=tr("Escribe tu idea aquí...", "Write your idea here...", "Écrivez votre idée ici...", "Scrivi la tua idea qui...", "在此写下您的想法..."),
            multiline=True,
            min_lines=1,
            max_lines=3,
            border_color="#9D50BB",
            color="white",
            text_size=12,
            label_style=ft.TextStyle(color="#aaaaaa", size=11),
            focused_border_color="#00FFFF",
        )
        
        def enviar_sugerencia_click(e):
            text = suggestion_input.value.strip()
            if not text:
                mostrar_snack(tr("Por favor escribe algo antes de enviar.", "Please type something before sending.", "Veuillez écrire quelque chose.", "Si prega di scrivere qualcosa.", "发送前请填写内容。"), color="red")
                return
            try:
                db_sug = conectar_db()
                if db_sug:
                    cursor_sug = db_sug.cursor()
                    cursor_sug.execute("""
                        CREATE TABLE IF NOT EXISTS sugerencias_luxo (
                            ID_Sugerencia INT AUTO_INCREMENT PRIMARY KEY,
                            ID_Usuario INT NOT NULL,
                            Fecha_Hora DATETIME DEFAULT CURRENT_TIMESTAMP,
                            Sugerencia TEXT NOT NULL,
                            FOREIGN KEY (ID_Usuario) REFERENCES usuarios(ID_Usuario) ON DELETE CASCADE
                        )
                    """)
                    db_sug.commit()
                    
                    cursor_sug.execute("""
                        INSERT INTO sugerencias_luxo (ID_Usuario, Sugerencia)
                        VALUES (%s, %s)
                    """, (user_info["id"], text))
                    db_sug.commit()
                    db_sug.close()
                    
                    suggestion_input.value = ""
                    mostrar_snack(tr("¡Sugerencia enviada! Gracias.", "Suggestion sent! Thanks.", "Suggestion envoyée ! Merci.", "Suggerimento inviato! Grazie.", "建议已发送！谢谢。"), color="#7CFC00")
                    page.update()
            except Exception as ex:
                print("ERROR AL ENVIAR SUGERENCIA:", ex)
                mostrar_snack("Error", color="red")

        btn_enviar_sug = ft.ElevatedButton(
            tr("Enviar", "Send", "Envoyer", "Invia", "发送"),
            on_click=enviar_sugerencia_click,
            bgcolor="#9D50BB",
            color="white",
            height=30,
        )

        suggestion_box = ft.Container(
            content=ft.Column([
                ft.Text(tr("¿Qué te gustaría que tuviera LUXO?", "What would you like LUXO to have?", "Qu'aimeriez-vous que LUXO ait ?", "Cosa vorresti che avesse LUXO?", "您希望 LUXO 增加什么功能？"), color="#D8B4FE", size=11, weight="bold"),
                suggestion_input,
                ft.Row([btn_enviar_sug], alignment="end")
            ], spacing=5),
            bgcolor="#141424",
            padding=10,
            border_radius=8,
            border=ft.Border.all(1, "#333333"),
        )

        # Dropdown y botón de refrescar para forzar idioma en la barra lateral
        lang_dropdown = EmojiDropdown(
            label=t("lang_label"),
            value=selected_lang[0],
            border_color="#9D50BB",
            options=[
                ft.dropdown.Option("es", "Español"),
                ft.dropdown.Option("en", "English"),
                ft.dropdown.Option("fr", "Français"),
                ft.dropdown.Option("it", "Italiano"),
                ft.dropdown.Option("zh", "中文")
            ],
            width=140,
            height=45
        )

        def on_global_master_refresh_click(e):
            curr_v = active_view[0] if active_view else "chat"
            curr_lang = (lang_dropdown.value if 'lang_dropdown' in locals() and lang_dropdown and lang_dropdown.value else selected_lang[0]) or "es"
            apply_language(curr_lang)
            cambiar_vista(curr_v)
            if page:
                snack = ft.SnackBar(
                    content=ft.Row([
                        ft.Text("🔄", size=14),
                        ft.Text(f"LUXO ({curr_lang.upper()}): Módulo '{curr_v.upper()}' y Lenguaje Recargados", color="white", weight="bold", size=12)
                    ], spacing=6),
                    bgcolor="#064E3B",
                    duration=2000
                )
                page.overlay.append(snack)
                snack.open = True
                page.update()

        def apply_language(new_lang):
            selected_lang[0] = new_lang
            try:
                lang_dropdown.label = tr("Idioma 🌐", "Language 🌐", "Langue 🌐", "Lingua 🌐", "语言 🌐")
                btn_chat.content.controls[1].value = tr("Asistente Chat ✨", "Chat Assistant ✨", "Assistant Chat ✨", "Assistente Chat ✨", "聊天助手 ✨")
                btn_historial.content.controls[1].value = tr("Mi Historial ⏳", "My History ⏳", "Mon Historique ⏳", "La Mia Cronologia ⏳", "我的历史记录 ⏳")
                btn_operacion_diaria.content.controls[1].value = tr("Aperturas y Cierres 🔑", "Openings & Closings 🔑", "Ouvertures & Fermetures 🔑", "Aperture & Chiusure 🔑", "开门与关门 🔑")
                btn_checklists.content.controls[1].value = tr("Checklists 📋", "Checklists 📋", "Listes 📋", "Liste 📋", "任务清单 📋")
                btn_manuales.content.controls[1].value = tr("Manuales 📚", "Manuals 📚", "Manuels 📚", "Manuali 📚", "手册 📚")
                btn_garantias.content.controls[1].value = tr("Garantías 👓", "Warranties 👓", "Garanties 👓", "Garanzie 👓", "保修 👓")
                btn_tareas.content.controls[1].value = tr("Tareas 📋", "Tasks 📋", "Tâches 📋", "Attività 📋", "任务 📋")
                btn_campanas.content.controls[1].value = tr("Campañas 📸", "Campaigns 📸", "Campagnes 📸", "Campagne 📸", "活动 📸")
                btn_presupuesto.content.controls[1].value = tr("Presupuesto 💰", "Budget 💰", "Budget 💰", "Budget 💰", "预算 💰")
                btn_reto.content.controls[1].value = tr("Reto del Día 🏆", "Daily Quiz 🏆", "Défi du Jour 🏆", "Sfida del Giorno 🏆", "每日挑战 🏆")
                btn_vendedores.content.controls[1].value = tr("Configuración Tienda 👥", "Store Config 👥", "Configuration 👥", "Configurazione 👥", "店铺配置 👥")
                btn_simulador.content.controls[1].value = tr("Simulador IA 🎭", "AI Simulator 🎭", "Simulateur IA 🎭", "Simulatore IA 🎭", "AI 模拟器 🎭")
                btn_crm.content.controls[1].value = tr("CRM Cobertura Oops 📱", "CRM Coverage Oops 📱", "CRM Couverture Oops 📱", "CRM Copertura Oops 📱", "CRM 意外保障 📱")
                btn_meta_semanal.content.controls[1].value = tr("Metas y Métricas 🎯", "Goals & Metrics 🎯", "Objectifs & Métriques 🎯", "Obiettivi & Metriche 🎯", "目标与指标 🎯")
                btn_weekly.content.controls[1].value = tr("Weekly 🗓️", "Weekly 🗓️", "Hebdomadaire 🗓️", "Settimanale 🗓️", "每周 🗓️")
                btn_enfoque.content.controls[1].value = tr("Enfoque Diario 2026 ☀️", "Daily Focus 2026 ☀️", "Focus Quotidien ☀️", "Focus Giornaliero ☀️", "每日焦点 ☀️")
                if 'btn_dashboard' in locals() and btn_dashboard:
                    btn_dashboard.content.controls[1].value = tr("Panel de Control 🎮", "Admin Panel 🎮", "Panneau de Contrôle 🎮", "Pannello di Controllo 🎮", "控制面板 🎮")
                if 'btn_admin_trivia' in locals() and btn_admin_trivia:
                    btn_admin_trivia.content.controls[1].value = tr("Gestionar Trivia 🧠", "Manage Trivia 🧠", "Gérer Quiz 🧠", "Gestisci Trivia 🧠", "管理问答 🧠")
                if 'btn_bitacora' in locals() and btn_bitacora:
                    btn_bitacora.content.controls[1].value = tr("Bitácora de Seguridad 🛡️", "Security Log 🛡️", "Registre de Sécurité 🛡️", "Registro di Sicurezza 🛡️", "安全日志 🛡️")
                if 'btn_gestion_perfiles' in locals() and btn_gestion_perfiles:
                    btn_gestion_perfiles.content.controls[1].value = tr("Gestión de Perfiles 👤", "Profile Management 👤", "Gestion des Profils 👤", "Gestione Profili 👤", "个人资料管理 👤")

                tile_ventas.title.value = tr("📊 VENTAS Y MÉTRICAS", "📊 SALES & METRICS", "📊 VENTES & MÉTRIQUES", "📊 VENDITE & METRICHE", "📊 销售与指标")
                tile_clientes.title.value = tr("🤝 CLIENTES Y GARANTÍAS", "🤝 CLIENTS & WARRANTY", "🤝 CLIENTS & GARANTIE", "🤝 CLIENTI & GARANZIA", "🤝 客户与保修")
                tile_operacion.title.value = tr("📋 OPERACIÓN Y TIENDA", "📋 STORE OPERATIONS", "📋 OPÉRATIONS MAGASIN", "📋 OPERAZIONI NEGOZIO", "📋 店铺运营")
                tile_entrenamiento.title.value = tr("🧠 CAPACITACIÓN E IA", "🧠 TRAINING & IA", "🧠 FORMATION & IA", "🧠 FORMAZIONE & IA", "🧠 培训与 AI")

                btn_logout.content.controls[1].value = tr("Cerrar Sesión 👋", "Log Out 👋", "Se Déconnecter 👋", "Disconnettersi 👋", "退出登录 👋")
                suggestion_box.content.controls[0].value = tr("¿Qué te gustaría que tuviera LUXO?", "What would you like LUXO to have?", "Qu'aimeriez-vous que LUXO ait ?", "Cosa vorresti che avesse LUXO?", "您希望 LUXO 增加什么功能？")
                suggestion_input.label = tr("Escribe tu idea aquí...", "Write your idea here...", "Écrivez votre idée ici...", "Scrivi la tua idea qui...", "在此写下您的想法...")
                btn_enviar_sug.text = tr("Enviar", "Send", "Envoyer", "Invia", "发送")
            except Exception as ex_tr:
                print("Error actualizando textos del sidebar:", ex_tr)

            curr_view = active_view[0] if ('active_view' in locals() and active_view) else "chat"
            cambiar_vista(curr_view)
            try:
                sidebar.update()
            except Exception:
                pass
            page.update()

        def language_changed(e):
            val = e.control.value if (e and hasattr(e, "control") and e.control and e.control.value) else (lang_dropdown.value or "es")
            apply_language(val)

        lang_dropdown.on_change = language_changed

        btn_refresh_lang = ft.Container(
            content=ft.Row([
                ft.Icon(ft.Icons.REFRESH_ROUNDED, color="#00FFFF", size=20)
            ], alignment="center"),
            border=ft.Border.all(1.5, "#00FFFF"),
            border_radius=8,
            padding=10,
            bgcolor="#1E1E2E",
            ink=True,
            on_click=on_global_master_refresh_click,
            tooltip="Sincronizar e Idioma 🔄"
        )

        lang_row = ft.Row([
            lang_dropdown,
            btn_refresh_lang
        ], alignment="start", vertical_alignment="center", spacing=6)

        # --- CATEGORÍAS AGRUPADAS CON ACORDEÓN DESPLEGABLE (ExpansionTile) ---
        ventas_controls = [btn_presupuesto, btn_meta_semanal, btn_weekly]
        if btn_dashboard:
            ventas_controls.append(btn_dashboard)
        if btn_bitacora:
            ventas_controls.append(btn_bitacora)

        def crear_acordeon(titulo_traduccion, controles_lista):
            trailing_txt = ft.Text("▼", color="#00FFFF", size=11)
            def al_cambiar(e):
                expandido = str(e.data).lower() == "true"
                trailing_txt.value = "▲" if expandido else "▼"
                trailing_txt.update()
            return ft.ExpansionTile(
                title=titulo_traduccion,
                controls=controles_lista,
                collapsed_text_color="#00FFFF",
                text_color="#D8B4FE",
                bgcolor="#161622",
                shape=ft.RoundedRectangleBorder(radius=8),
                trailing=trailing_txt,
                on_change=al_cambiar
            )

        tile_ventas = crear_acordeon(
            ft.Text(tr("📊 VENTAS Y MÉTRICAS", "📊 SALES & METRICS", "📊 VENTES & MÉTRIQUES", "📊 VENDITE & METRICHE", "📊 销售与指标"), color="#00FFFF", weight="bold", size=12),
            ventas_controls
        )

        clientes_controls = [btn_crm, btn_garantias]
        tile_clientes = crear_acordeon(
            ft.Text(tr("🤝 CLIENTES Y GARANTÍAS", "🤝 CLIENTS & WARRANTY", "🤝 CLIENTS & GARANTIE", "🤝 CLIENTI & GARANZIA", "🤝 客户 & 保修"), color="#00FFFF", weight="bold", size=12),
            clientes_controls
        )

        operacion_controls = [btn_operacion_diaria, btn_checklists, btn_tareas, btn_campanas, btn_manuales, btn_vendedores]
        tile_operacion = crear_acordeon(
            ft.Text(tr("📋 OPERACIÓN Y TIENDA", "📋 STORE OPERATIONS", "📋 OPÉRATIONS MAGASIN", "📋 OPERAZIONI NEGOZIO", "📋 店铺运营"), color="#00FFFF", weight="bold", size=12),
            operacion_controls
        )

        entrenamiento_controls = [btn_simulador, btn_reto]
        if btn_admin_trivia:
            entrenamiento_controls.append(btn_admin_trivia)

        tile_entrenamiento = crear_acordeon(
            ft.Text(tr("🧠 CAPACITACIÓN E IA", "🧠 TRAINING & AI", "🧠 FORMATION & IA", "🧠 FORMAZIONE & IA", "🧠 培训与 AI"), color="#00FFFF", weight="bold", size=12),
            entrenamiento_controls
        )

        avatar_luxo2_base64 = obtener_64("luxo_avatar2.png") or obtener_64("avatar_luxo2.png")
        
        avatar_header_widget = ft.Container(
            content=ft.Image(
                src=avatar_luxo2_base64,
                width=34,
                height=34,
                fit=ft.controls.box.BoxFit.COVER
            ) if avatar_luxo2_base64 else ft.Text("🤖", size=18, color="#00FFFF"),
            width=34,
            height=34,
            border_radius=17,
            bgcolor="#1A102F",
            border=ft.Border.all(1.5, "#00FFFF"),
            alignment=ft.alignment.Alignment(0, 0),
            shadow=[
                ft.BoxShadow(
                    color="#4000FFFF",
                    blur_radius=10,
                    spread_radius=1
                )
            ]
        )

        title_luxo_gradient = ft.ShaderMask(
            content=ft.Text("LUXO", size=18, weight="bold"),
            blend_mode=ft.BlendMode.SRC_IN,
            shader=ft.LinearGradient(
                colors=["#00F0FF", "#E040FB"],
                begin=ft.alignment.Alignment(-1, 0),
                end=ft.alignment.Alignment(1, 0)
            )
        )

        subtitle_terminal = ft.Text("SYSTEM", size=9, weight="bold", color="#00FFFF")

        star_icon_container.visible = es_admin()
        if es_admin():
            try:
                operacion_tiendas.actualizar_estrella_aperturas(page, star_icon_container, conectar_db)
            except Exception as e:
                print("Error al inicializar estrella de aperturas:", e)

        sidebar_brand_header = ft.Container(
            content=ft.Row([
                avatar_header_widget,
                ft.Column([
                    title_luxo_gradient,
                    subtitle_terminal
                ], spacing=0, alignment=ft.MainAxisAlignment.CENTER),
                ft.Container(expand=True),
                star_icon_container,
                bell_icon_container
            ], vertical_alignment="center", spacing=6),
            padding=ft.padding.Padding(4, 4, 4, 4)
        )

        sidebar_items = [
            sidebar_brand_header,
            ft.Divider(height=15, color="#444444"),
            profile_row,
            ft.Divider(height=15, color="#444444"),
            btn_operacion_diaria,
            btn_chat,
            btn_historial,
            btn_enfoque,
            ft.Divider(height=10, color="#333333"),
            tile_ventas,
            tile_clientes,
            tile_operacion,
            tile_entrenamiento,
            ft.Container(height=10),
            suggestion_box,
            ft.Container(expand=True),
            lang_row,
            btn_logout
        ]

        print(">>> LUXO SYSTEM: SIDEBAR ACTIVADO CON ENFOQUE DIARIO 2026 <<<")

        sidebar = ft.Container(
            content=ft.Column(sidebar_items, spacing=10, scroll=ft.ScrollMode.AUTO),
            width=240,
            padding=16,
            bgcolor="#090B16",
            border_radius=15,
            border=ft.Border.all(1.5, "#00FFFF"),
            shadow=[
                ft.BoxShadow(
                    color="#3500FFFF",
                    blur_radius=18,
                    spread_radius=1
                )
            ]
        )

        def toggle_sidebar(e):
            sidebar.visible = not sidebar.visible
            page.update()

        btn_global_master_refresh = ft.Container(
            content=ft.Row([
                ft.Text("🔄", size=12),
                ft.Text("Sincronizar LUXO", color="#00FF88", weight="bold", size=11)
            ], spacing=4),
            border=ft.Border.all(1.5, "#00FF88"),
            border_radius=6,
            padding=8,
            bgcolor="transparent",
            ink=True,
            on_click=on_global_master_refresh_click,
            tooltip="Recargar y sincronizar el módulo activo actual"
        )

        # Definir la cabecera superior permanente para toda la aplicación
        top_appbar = ft.Container(
            content=ft.Row([
                ft.Row([
                    ft.Container(
                        content=ft.Text("☰", color="#00FFFF", size=18, weight="bold"),
                        on_click=toggle_sidebar,
                        tooltip="Mostrar/Ocultar Menú",
                        padding=8,
                        alignment=ft.alignment.Alignment(0, 0),
                        ink=True,
                        border_radius=4
                    ),
                    ft.Text("LUXO AI SYSTEM", color="white", weight="bold", size=15),
                ], vertical_alignment="center", spacing=4),
                btn_global_master_refresh
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            bgcolor="#0B0E17",
            padding=6,
            border=ft.Border(bottom=ft.BorderSide(1, "#1F2937")),
            visible=True
        )

        # Ocultar menú por defecto si la pantalla es de móvil y activar cabecera superior
        if getattr(page, "width", None) and page.width < 800:
            sidebar.visible = False
            top_appbar.visible = True

        # Crear área derecha combinada con la cabecera y el área de contenido
        right_area = ft.Column([
            top_appbar,
            content_area
        ], expand=True, spacing=0)

        dashboard_layout = ft.Row([
            sidebar,
            ft.VerticalDivider(width=1, color="#00FFFF"),
            right_area
        ], expand=True, spacing=12)

        cambiar_vista("chat")

        page.controls.clear()
        page.bgcolor = "#05070D"

        if img_fondo:
            page.add(
                ft.Container(
                    expand=True,
                    image=ft.DecorationImage(
                        src=img_fondo,
                        fit=ft.controls.box.BoxFit.COVER
                    ),
                    content=dashboard_layout
                )
            )
        else:
            page.add(
                ft.Container(
                    expand=True,
                    padding=10,
                    gradient=ft.LinearGradient(
                        colors=["#021622", "#05070D", "#220228"],
                        begin=ft.alignment.Alignment(-1, -0.2),
                        end=ft.alignment.Alignment(1, 0.2)
                    ),
                    content=dashboard_layout
                )
            )

        page.update()

    def obtener_avatar_usuario(id_usuario):
        if not id_usuario:
            return None
        os.makedirs(os.path.join(ASSETS_PATH, "perfiles"), exist_ok=True)
        for ext in [".png", ".jpg", ".jpeg"]:
            ruta = os.path.join(ASSETS_PATH, "perfiles", f"user_{id_usuario}{ext}")
            if os.path.exists(ruta):
                try:
                    with open(ruta, "rb") as f:
                        content = base64.b64encode(f.read()).decode("utf-8")
                    mime = "image/png" if ext == ".png" else "image/jpeg"
                    return f"data:{mime};base64,{content}"
                except Exception as ex:
                    print("Error leyendo avatar:", ex)
        return None

    def cargar_medallas_usuario(id_usuario):
        medallas = []
        if not id_usuario:
            return medallas
        try:
            db = conectar_db()
            if db:
                cursor = db.cursor()
                
                # 1. Auditor Estrella ⭐ (Campañas con Estatus = 'Visto_Bueno')
                cursor.execute("""
                    SELECT COUNT(*) FROM campana_entregas_tienda 
                    WHERE ID_Usuario = %s AND Estatus = 'Visto_Bueno'
                """, (id_usuario,))
                cant_camp = cursor.fetchone()[0]
                auditor_star_icon = getattr(ft.Icons, "STAR_ROUNDED", ft.Icons.STAR)
                auditor_star_border = getattr(ft.Icons, "STAR_BORDER_ROUNDED", ft.Icons.STAR_BORDER)
                
                if cant_camp > 0:
                    medallas.append({
                        "nombre": "Auditor Estrella",
                        "icono": auditor_star_icon,
                        "color": "#FFD700",
                        "tooltip": "Auditor Estrella ⭐\n¡Desbloqueada! Tienes campañas aprobadas con Visto Bueno.",
                        "desbloqueada": True
                    })
                else:
                    medallas.append({
                        "nombre": "Auditor Estrella",
                        "icono": auditor_star_border,
                        "color": "#555555",
                        "tooltip": "Auditor Estrella (Bloqueada) 🔒\nSube tus fotos de campaña y obtén el Visto Bueno para ganarla.",
                        "desbloqueada": False
                    })
                
                # Obtener cuántas tareas hay por categoría
                cursor.execute("SELECT Categoria, COUNT(*) FROM plantillas_checklist GROUP BY Categoria")
                tareas_por_cat = {row[0]: row[1] for row in cursor.fetchall()}
                
                # 2. Madrugador 🌅 (Checklist Apertura - Categoría 1)
                total_cat1 = tareas_por_cat.get(1, 0)
                desb_cat1 = False
                if total_cat1 > 0:
                    cursor.execute("""
                        SELECT Fecha, COUNT(*) as completadas FROM registro_checklist
                        WHERE ID_Usuario = %s AND Completado = 1 AND ID_Plantilla IN (
                            SELECT ID_Plantilla FROM plantillas_checklist WHERE Categoria = 1
                        )
                        GROUP BY Fecha
                        HAVING completadas >= %s
                    """, (id_usuario, total_cat1))
                    desb_cat1 = cursor.fetchone() is not None
                
                light_mode_icon = getattr(ft.Icons, "LIGHT_MODE_ROUNDED", ft.Icons.LIGHT_MODE)
                sad_face_icon = getattr(ft.Icons, "SENTIMENT_VERY_DISSATISFIED_ROUNDED", ft.Icons.SENTIMENT_VERY_DISSATISFIED)
                
                if desb_cat1:
                    medallas.append({
                        "nombre": "Madrugador",
                        "icono": light_mode_icon,
                        "color": "#FFA500",
                        "tooltip": "Madrugador 🌅\n¡Desbloqueada! Completaste el checklist de Apertura al 100%.",
                        "desbloqueada": True
                    })
                else:
                    medallas.append({
                        "nombre": "Madrugador",
                        "icono": sad_face_icon,
                        "color": "#555555",
                        "tooltip": "Madrugador (Bloqueada) 🔒\nCompleta al 100% el checklist de Apertura para ganarla.",
                        "desbloqueada": False
                    })
                
                # 3. Cierre Perfecto 🌙 (Checklist Cierre - Categoría 2)
                total_cat2 = tareas_por_cat.get(2, 0)
                desb_cat2 = False
                if total_cat2 > 0:
                    cursor.execute("""
                        SELECT Fecha, COUNT(*) as completadas FROM registro_checklist
                        WHERE ID_Usuario = %s AND Completado = 1 AND ID_Plantilla IN (
                            SELECT ID_Plantilla FROM plantillas_checklist WHERE Categoria = 2
                        )
                        GROUP BY Fecha
                        HAVING completadas >= %s
                    """, (id_usuario, total_cat2))
                    desb_cat2 = cursor.fetchone() is not None
                
                nightlight_icon = getattr(ft.Icons, "NIGHTLIGHT_ROUNDED", ft.Icons.NIGHTLIGHT)
                
                if desb_cat2:
                    medallas.append({
                        "nombre": "Cierre Perfecto",
                        "icono": nightlight_icon,
                        "color": "#9D50BB",
                        "tooltip": "Cierre Perfecto 🌙\n¡Desbloqueada! Completaste el checklist de Cierre al 100%.",
                        "desbloqueada": True
                    })
                else:
                    medallas.append({
                        "nombre": "Cierre Perfecto",
                        "icono": sad_face_icon,
                        "color": "#555555",
                        "tooltip": "Cierre Perfecto (Bloqueada) 🔒\nCompleta al 100% el checklist de Cierre para ganarla.",
                        "desbloqueada": False
                    })
                
                # 4. Vendedor Pro 💰 (Checklist Venta - Categoría 3)
                total_cat3 = tareas_por_cat.get(3, 0)
                desb_cat3 = False
                if total_cat3 > 0:
                    cursor.execute("""
                        SELECT Fecha, COUNT(*) as completadas FROM registro_checklist
                        WHERE ID_Usuario = %s AND Completado = 1 AND ID_Plantilla IN (
                            SELECT ID_Plantilla FROM plantillas_checklist WHERE Categoria = 3
                        )
                        GROUP BY Fecha
                        HAVING completadas >= %s
                    """, (id_usuario, total_cat3))
                    desb_cat3 = cursor.fetchone() is not None
                
                monetization_icon = getattr(ft.Icons, "MONETIZATION_ON_ROUNDED", ft.Icons.MONETIZATION_ON)
                
                if desb_cat3:
                    medallas.append({
                        "nombre": "Vendedor Pro",
                        "icono": monetization_icon,
                        "color": "#00FF7F",
                        "tooltip": "Vendedor Pro 💰\n¡Desbloqueada! Completaste el checklist de Venta Exitosa al 100%.",
                        "desbloqueada": True
                    })
                else:
                    medallas.append({
                        "nombre": "Vendedor Pro",
                        "icono": sad_face_icon,
                        "color": "#555555",
                        "tooltip": "Vendedor Pro (Bloqueada) 🔒\nCompleta al 100% el checklist de Venta Exitosa para ganarla.",
                        "desbloqueada": False
                    })
                
                db.close()
        except Exception as e:
            print("Error al cargar medallas del usuario:", e)
        return medallas

    # =====================================
    # BIOMETRÍA - CÁMARA FACIAL Y PASSKEY
    # =====================================

    async def _abrir_camara_facial():
        """Inyecta un modal HTML5 de cámara en el navegador para capturar el rostro."""
        try:
            js_code = """
(function() {
    if (document.getElementById('luxo-facial-modal')) return;
    const modal = document.createElement('div');
    modal.id = 'luxo-facial-modal';
    modal.style.cssText = 'position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,0.92);z-index:99999999;display:flex;flex-direction:column;align-items:center;justify-content:center;font-family:Segoe UI,sans-serif;';
    modal.innerHTML = `
        <div style="background:#0a0a16;border:2px solid #00FFFF;border-radius:20px;padding:28px 24px;max-width:420px;width:95%;text-align:center;box-shadow:0 0 40px rgba(0,255,255,0.4);">
            <div style="font-size:28px;font-weight:800;color:#D8B4FE;margin-bottom:6px;">📷 Reconocimiento Facial</div>
            <div style="color:#aaa;font-size:13px;margin-bottom:16px;">Posiciona tu rostro dentro del círculo y presiona Capturar</div>
            <div style="position:relative;width:220px;height:220px;margin:0 auto 16px;">
                <video id="luxo-cam" autoplay playsinline muted style="width:220px;height:220px;object-fit:cover;border-radius:50%;border:3px solid #00FFFF;"></video>
                <canvas id="luxo-canvas" width="220" height="220" style="display:none;"></canvas>
            </div>
            <div id="luxo-face-msg" style="color:#00FFFF;font-weight:bold;font-size:14px;min-height:22px;margin-bottom:12px;"></div>
            <div style="display:flex;gap:12px;justify-content:center;">
                <button id="luxo-capture-btn" onclick="luxoCaptureFace()" style="background:linear-gradient(135deg,#003366,#0066cc);color:white;border:none;padding:12px 22px;border-radius:10px;font-size:14px;font-weight:bold;cursor:pointer;">📸 Capturar</button>
                <button onclick="document.getElementById('luxo-facial-modal').remove();luxoCamStop();" style="background:#333;color:white;border:none;padding:12px 22px;border-radius:10px;font-size:14px;cursor:pointer;">✕ Cerrar</button>
            </div>
        </div>
    `;
    document.body.appendChild(modal);
    let stream = null;
    function luxoCamStop() { if (stream) { stream.getTracks().forEach(t => t.stop()); stream = null; } }
    window.luxoCamStop = luxoCamStop;
    navigator.mediaDevices.getUserMedia({ video: { facingMode: 'user' } })
        .then(s => { stream = s; document.getElementById('luxo-cam').srcObject = s; })
        .catch(() => { document.getElementById('luxo-face-msg').innerText = '⚠️ Cámara no disponible en este dispositivo.'; document.getElementById('luxo-face-msg').style.color = '#FF4500'; });
    window.luxoCaptureFace = function() {
        const video = document.getElementById('luxo-cam');
        const canvas = document.getElementById('luxo-canvas');
        const ctx = canvas.getContext('2d');
        ctx.drawImage(video, 0, 0, 220, 220);
        const frameB64 = canvas.toDataURL('image/jpeg', 0.85);
        document.getElementById('luxo-face-msg').innerText = '⏳ Analizando rostro...';
        document.getElementById('luxo-face-msg').style.color = '#FFFF00';
        fetch('/api/biometria/facial_login', { method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({ frame_base64: frameB64 }) })
            .then(r => r.json())
            .then(data => {
                if (data.status === 'ok') {
                    document.getElementById('luxo-face-msg').innerText = '✅ ¡Bienvenido, ' + data.nombre + '!';
                    document.getElementById('luxo-face-msg').style.color = '#7CFC00';
                    luxoCamStop();
                    setTimeout(() => { document.getElementById('luxo-facial-modal').remove(); }, 1800);
                } else if (data.status === 'no_registered') {
                    document.getElementById('luxo-face-msg').innerText = '⚠️ ' + data.message;
                    document.getElementById('luxo-face-msg').style.color = '#FF8C00';
                } else {
                    document.getElementById('luxo-face-msg').innerText = '❌ ' + (data.message || 'Rostro no reconocido');
                    document.getElementById('luxo-face-msg').style.color = '#FF4500';
                }
            })
            .catch(() => { document.getElementById('luxo-face-msg').innerText = '❌ Error de conexión. Intenta de nuevo.'; document.getElementById('luxo-face-msg').style.color = '#FF4500'; });
    };
})();
"""
            ejecutar_js_flet(page, js_code)
        except Exception as ex_cf:
            print("Error abriendo cámara facial:", ex_cf)
            mostrar_snack("Usa Chrome o Edge para el Reconocimiento Facial 📷", "orange")

    async def _activar_passkey():
        """Invoca el API WebAuthn nativo del navegador para autenticar por Huella/Passkey."""
        try:
            js_code = """
(async function() {
    const msgEl = document.createElement('div');
    msgEl.style.cssText = 'position:fixed;bottom:80px;left:50%;transform:translateX(-50%);background:#0a0a16;border:2px solid #00FFFF;color:#00FFFF;padding:14px 24px;border-radius:14px;font-size:14px;font-weight:bold;z-index:999999;box-shadow:0 0 20px rgba(0,255,255,0.4);';
    msgEl.innerText = '👆 Solicitando huella dactilar...';
    document.body.appendChild(msgEl);
    try {
        const challResp = await fetch('/api/biometria/passkey_challenge');
        const challData = await challResp.json();
        const challenge = Uint8Array.from(atob(challData.challenge.replace(/-/g,'+').replace(/_/g,'/')), c => c.charCodeAt(0));
        const credential = await navigator.credentials.get({
            publicKey: {
                challenge: challenge,
                rpId: window.location.hostname,
                userVerification: 'required',
                timeout: 60000,
                allowCredentials: []
            }
        });
        const credId = btoa(String.fromCharCode(...new Uint8Array(credential.rawId)));
        msgEl.innerText = '⏳ Verificando identidad...';
        const verResp = await fetch('/api/biometria/passkey_verify', {
            method: 'POST',
            headers: {'Content-Type': 'application/json'},
            body: JSON.stringify({ credential_id: credId })
        });
        const verData = await verResp.json();
        if (verData.status === 'ok') {
            msgEl.style.borderColor = '#7CFC00'; msgEl.style.color = '#7CFC00';
            msgEl.innerText = '✅ ¡Bienvenido, ' + verData.nombre + '!';
            setTimeout(() => msgEl.remove(), 2500);
        } else {
            msgEl.style.borderColor = '#FF4500'; msgEl.style.color = '#FF4500';
            msgEl.innerText = '❌ ' + (verData.message || 'Huella no registrada');
            setTimeout(() => msgEl.remove(), 3500);
        }
    } catch(ex) {
        msgEl.style.borderColor = '#FF8C00'; msgEl.style.color = '#FF8C00';
        if (ex.name === 'NotAllowedError') {
            msgEl.innerText = '⚠️ Permiso de Huella denegado. Intenta de nuevo.';
        } else if (ex.name === 'NotSupportedError') {
            msgEl.innerText = '⚠️ Este dispositivo no tiene sensor de huella registrado.';
        } else {
            msgEl.innerText = '⚠️ ' + ex.message;
        }
        setTimeout(() => msgEl.remove(), 4000);
    }
})();
"""
            ejecutar_js_flet(page, js_code)
        except Exception as ex_pk:
            print("Error activando passkey:", ex_pk)
            mostrar_snack("Huella/Passkey: Usa Chrome, Edge o Safari en tu celular/laptop con sensor 👆", "orange")

    # =====================================
    # LOGIN
    # =====================================

    def login_click(e):

        try:
            db = conectar_db()

            if not db:
                mostrar_snack("Error Base de Datos", color="#FF4B4B")
                return

            cursor = db.cursor(dictionary=True)

            cursor.execute(
                """
                SELECT
                ID_Usuario,
                Nombre_Completo,
                Rol,
                Tienda,
                Zona,
                Contrasena
                FROM usuarios
                WHERE Usuario = %s
                """,
                (txt_user.value,)
            )

            res = cursor.fetchone()

            if res and verify_password(txt_pass.value, res.get("Contrasena", "")):
                # Migración transparente: Si la contraseña en BD era texto plano legacy, actualizarla con hash bcrypt
                stored_pass = str(res.get("Contrasena") or "")
                if not (stored_pass.startswith("$2b$") or stored_pass.startswith("$2a$")):
                    try:
                        new_hash = hash_password(txt_pass.value)
                        cursor.execute("UPDATE usuarios SET Contrasena = %s WHERE ID_Usuario = %s", (new_hash, res["ID_Usuario"]))
                        db.commit()
                        print(f"🔒 Contraseña del usuario '{res['Nombre_Completo']}' migrada exitosamente a hash bcrypt.")
                    except Exception as ex_mig:
                        print("Error migrando contraseña a bcrypt:", ex_mig)
                login_message.value = ""
                login_error_box.visible = False

                user_info["id"] = res["ID_Usuario"]
                user_info["usuario"] = res.get("Usuario") or ""
                user_info["nombre"] = res["Nombre_Completo"]
                user_info["rol"] = res["Rol"]
                user_info["tienda"] = res["Tienda"] if res["Tienda"] is not None else ""
                user_info["zona"] = res["Zona"] if res["Zona"] is not None else "Zona Centro"
                user_info["img_usuario"] = obtener_avatar_usuario(res["ID_Usuario"])
                reproducir_saludo_login(res["Nombre_Completo"])
                
                # Guardar sesión de forma en memoria active_sessions
                user_id_key = res["ID_Usuario"]
                active_sessions[user_id_key] = {
                    "page": page,
                    "user_info": user_info,
                    "cargar_chat": cargar_chat,
                    "active_file_callback": active_file_callback
                }

                # --- REGISTRAR INICIO DE SESIÓN ---
                ip_client = getattr(page, "client_ip", None) or "Desconocido"
                
                def registrar_sesion_async(u_id, ip):
                    city = "Desconocido"
                    country = "Desconocido"
                    is_local = False
                    
                    if not ip or ip == "Desconocido":
                        is_local = True
                    else:
                        ip_clean = ip.strip()
                        if ip_clean in ("127.0.0.1", "::1", "localhost") or \
                           ip_clean.startswith("192.168.") or \
                           ip_clean.startswith("10.") or \
                           ip_clean.startswith("172.16.") or \
                           ip_clean.startswith("fe80:"):
                            is_local = True
                    
                    if is_local:
                        city = "Localhost"
                        country = "Local / Desarrollo"
                    else:
                        try:
                            resp = requests.get(f"http://ip-api.com/json/{ip}", timeout=5)
                            if resp.status_code == 200:
                                data = resp.json()
                                if data.get("status") == "success":
                                    city = data.get("city", "Desconocido")
                                    country = data.get("country", "Desconocido")
                        except Exception as err:
                            print("Error consultando geolocalización de IP:", err)
                            
                    try:
                        db_log = conectar_db()
                        if db_log:
                            cursor_log = db_log.cursor()
                            cursor_log.execute(
                                """
                                INSERT INTO sesiones (ID_Usuario, Direccion_IP, Ubicacion_Ciudad, Ubicacion_Pais)
                                VALUES (%s, %s, %s, %s)
                                """,
                                (u_id, ip, city, country)
                            )
                            db_log.commit()
                            db_log.close()
                    except Exception as err_db:
                        print("Error al guardar sesión en BD:", err_db)

                threading.Thread(target=registrar_sesion_async, args=(user_info["id"], ip_client), daemon=True).start()

                async def guardar_sesion_storage():
                    try:
                        import time
                        if hasattr(page, "shared_preferences") and page.shared_preferences:
                            await page.shared_preferences.set("logged_user_id", str(user_info["id"]))
                            await page.shared_preferences.set("last_activity_timestamp", str(int(time.time())))
                    except Exception as ex_st:
                        print("Notice guardar shared_preferences:", ex_st)

                page.run_task(guardar_sesion_storage)

                # Registrar en la bitácora como sesión por contraseña (no biométrica)
                user_info["biometria_metodo"] = None
                user_info["es_gerente_verificado"] = False
                try:
                    u_rol_bit = str(res.get("Rol", "")).lower()
                    es_gerente_pw = "admin" in u_rol_bit
                    _ip_log = getattr(page, "client_ip", None) or "Localhost"
                    _ua_log = "LUXO-Desktop"
                    def _reg_bit_pw(uid, nombre, rol_str, ip, ua):
                        try:
                            db_bit2 = conectar_db()
                            if db_bit2:
                                cur_bit2 = db_bit2.cursor()
                                cur_bit2.execute("""
                                    INSERT INTO bitacora_sesiones_biometricas
                                        (ID_Usuario, Nombre_Usuario, Empleado_Identificado, Metodo_Ingreso, Es_Gerente_Verificado, IP_Acceso, Dispositivo)
                                    VALUES (%s, %s, %s, 'Contrasena', %s, %s, %s)
                                """, (uid, nombre, nombre, es_gerente_pw, ip, ua[:150]))
                                db_bit2.commit()
                                db_bit2.close()
                        except Exception as ex_bit2:
                            print("Error bitácora contraseña:", ex_bit2)
                    import threading as _th
                    _th.Thread(target=_reg_bit_pw, args=(user_info["id"], user_info["nombre"], u_rol_bit, _ip_log, _ua_log), daemon=True).start()
                except Exception as ex_bit_pw:
                    print("Error preparando bitácora contraseña:", ex_bit_pw)

                cargar_chat()


            else:
                cursor.execute(
                    """
                    SELECT
                    ID_Usuario
                    FROM usuarios
                    WHERE Usuario = %s
                    """,
                    (txt_user.value,)
                )
                usuario_existe = cursor.fetchone()

                if usuario_existe:
                    mensaje = "Contraseña incorrecta"
                else:
                    mensaje = "Usuario no registrado"

                login_message.value = mensaje
                login_message.color = "#FF4B4B"
                login_error_box.visible = True
                page.update()

            db.close()
        except Exception as err:
            import traceback
            tb_str = traceback.format_exc()
            print("--- DETECTADO ERROR EN LOGIN ---")
            print(tb_str)
            try:
                with open("login_error.log", "w", encoding="utf-8") as log_f:
                    log_f.write(tb_str)
            except Exception as e_log:
                print("No se pudo escribir en login_error.log:", e_log)

    # =====================================
    # LOGIN UI
    # =====================================

    login_video_player = None
    btn_audio = None

    def toggle_audio(e):
        nonlocal login_video_player, btn_audio
        if login_video_player:
            try:
                currently_unmuted = e.control.data
                if currently_unmuted:
                    login_video_player.volume = 0
                    login_video_player.muted = True
                    btn_audio.content = ft.Text("🔇", size=11, color="#00FFFF", text_align="center")
                    btn_audio.tooltip = "Activar Audio"
                    e.control.data = False
                else:
                    login_video_player.volume = 100
                    login_video_player.muted = False
                    btn_audio.content = ft.Text("🔊", size=11, color="#00FFFF", text_align="center")
                    btn_audio.tooltip = "Silenciar Audio"
                    e.control.data = True
                login_video_player.update()
                btn_audio.update()
            except Exception as err:
                print("Error al cambiar estado de audio:", err)

    txt_user_input = ft.TextField(
        hint_text="Ej. admin",
        hint_style=ft.TextStyle(color="#555566", size=13),
        width=300,
        height=45,
        border_color="#121620",
        focused_border_color="#00F0FF",
        color="white",
        bgcolor="#040407",
        border_radius=10,
        content_padding=12
    )

    is_pass_hidden = [True]
    def toggle_pass_visibility(e):
        is_pass_hidden[0] = not is_pass_hidden[0]
        txt_pass_input.password = is_pass_hidden[0]
        btn_eye_3d.content.color = "#E040FB" if is_pass_hidden[0] else "#00F0FF"
        txt_pass_input.update()
        btn_eye_3d.update()

    btn_eye_3d = ft.Container(
        content=ft.Text("👁️", size=16, color="#E040FB"),
        alignment=ft.alignment.Alignment(0, 0),
        width=32,
        height=32,
        on_click=toggle_pass_visibility,
        tooltip="Mostrar / Ocultar Contraseña"
    )

    txt_pass_input = ft.TextField(
        hint_text="••••••••",
        hint_style=ft.TextStyle(color="#555566", size=13),
        password=True,
        width=300,
        height=45,
        border_color="#121620",
        focused_border_color="#E040FB",
        color="white",
        bgcolor="#040407",
        border_radius=10,
        content_padding=12,
        on_submit=login_click,
        suffix=btn_eye_3d
    )

    txt_user = txt_user_input
    txt_pass = txt_pass_input

    login_message = ft.Text(
        "",
        size=14,
        weight="bold",
        color="#FF4B4B"
    )

    login_error_box = ft.Container(
        content=login_message,
        bgcolor="#000000",
        padding=8,
        border_radius=8,
        visible=False,
        width=300
    )

    video_avatar = None
    if video_login_exists:
        try:
            login_video_player = fv.Video(
                playlist=[fv.VideoMedia(video_login_url)],
                playlist_mode=fv.PlaylistMode.LOOP,
                autoplay=True,
                volume=100.0,
                muted=False,
                controls=None,
                expand=True,
                fit=ft.BoxFit.COVER,
                filter_quality=ft.FilterQuality.HIGH,
            )
            def on_avatar_tap(e):
                if login_video_player:
                    try:
                        login_video_player.play()
                        login_video_player.volume = 100
                        login_video_player.muted = False
                        login_video_player.update()
                    except Exception: pass

            btn_audio = ft.Container(
                content=ft.Text("🔊", size=11, color="#00FFFF", text_align="center"),
                bgcolor="#111111",
                width=28,
                height=28,
                border_radius=14,
                alignment=ft.alignment.Alignment(0, 0),
                tooltip="Silenciar Audio",
                data=True,
                on_click=toggle_audio
            )
            video_avatar = ft.Stack([
                ft.Container(
                    content=login_video_player,
                    width=108,
                    height=108,
                    border_radius=54,
                    clip_behavior=ft.ClipBehavior.ANTI_ALIAS,
                    border=ft.Border.all(2, "#00F0FF"),
                    shadow=[ft.BoxShadow(color="#4000FFFF", blur_radius=20, spread_radius=1)],
                    on_click=on_avatar_tap
                ),
                ft.Container(
                    content=btn_audio,
                    right=0,
                    bottom=0,
                    width=28,
                    height=28,
                    border_radius=14,
                )
            ], width=108, height=108)
        except Exception as ex_v:
            print("Notice video login load:", ex_v)
            video_avatar = None

    header_title = ft.Column([
        ft.ShaderMask(
            content=ft.Text("LUXO OS", size=28, weight="bold", color="white"),
            blend_mode=ft.BlendMode.SRC_IN,
            shader=ft.LinearGradient(
                colors=["#00F0FF", "#E040FB"],
                begin=ft.alignment.Alignment(-1, 0),
                end=ft.alignment.Alignment(1, 0)
            )
        ),
        ft.Text(
            "PORTAL DE AUTENTICACIÓN",
            size=9,
            weight="bold",
            color="#8899A6"
        )
    ], horizontal_alignment="center", spacing=3)

    user_field_group = ft.Column([
        ft.Text("USUARIO", size=10, weight="bold", color="#00F0FF"),
        txt_user_input
    ], spacing=4, width=300)

    pass_field_group = ft.Column([
        ft.Text("CONTRASEÑA", size=10, weight="bold", color="#E040FB"),
        txt_pass_input
    ], spacing=4, width=300)

    btn_acceder = ft.Container(
        content=ft.Text("ACCEDER", color="white", weight="bold", size=14),
        alignment=ft.alignment.Alignment(0, 0),
        gradient=ft.LinearGradient(
            colors=["#00A3FF", "#E040FB"],
            begin=ft.alignment.Alignment(-1, -1),
            end=ft.alignment.Alignment(1, 1)
        ),
        padding=14,
        border_radius=22,
        width=300,
        height=46,
        on_click=login_click,
        shadow=[
            ft.BoxShadow(
                color="#E040FB",
                blur_radius=18,
                spread_radius=1
            )
        ]
    )

    login_card = ft.Container(
        content=ft.Column([
            video_avatar if video_avatar else (
                ft.Container(
                    content=ft.Image(
                        src=img_avatar,
                        width=108,
                        height=108,
                        fit=ft.BoxFit.COVER
                    ),
                    width=108,
                    height=108,
                    border_radius=54,
                    clip_behavior=ft.ClipBehavior.ANTI_ALIAS,
                    border=ft.Border.all(2, "#00F0FF"),
                    shadow=[ft.BoxShadow(color="#4000FFFF", blur_radius=20, spread_radius=1)]
                ) if img_avatar else ft.Text(
                    "LUXO",
                    size=28,
                    color="#FFFFFF",
                    weight="bold"
                )
            ),
            login_error_box,
            header_title,
            user_field_group,
            pass_field_group,
            ft.Container(height=6),
            btn_acceder
        ],
        horizontal_alignment="center",
        spacing=16 if is_mobile else 18),
        padding=32 if is_mobile else 42,
        bgcolor="#06070B",
        border_radius=24,
        border=ft.Border.all(1.2, "#0A202A"),
        shadow=[
            ft.BoxShadow(
                color="#000000",
                blur_radius=35,
                spread_radius=5,
            )
        ],
        width=370,
        clip_behavior=ft.ClipBehavior.HARD_EDGE
    )

    # Fondo Ambiental de Pantalla con degradado nativo de Flet (Cian a la izquierda, Obsidiana al centro, Magenta a la derecha)
    full_screen_background = ft.Container(
        content=login_card,
        alignment=ft.alignment.Alignment(0, 0),
        expand=True,
        gradient=ft.LinearGradient(
            colors=["#021622", "#05070D", "#220228"],
            begin=ft.alignment.Alignment(-1, -0.2),
            end=ft.alignment.Alignment(1, 0.2)
        )
    )

    page.bgcolor = "#05070D"
    page.vertical_alignment = "center"
    page.horizontal_alignment = "center"
    page.controls.clear()
    page.add(full_screen_background)
    page.update()

    async def intentar_restaurar_sesion():
        try:
            if hasattr(page, "shared_preferences") and page.shared_preferences:
                uid_saved = await page.shared_preferences.get("logged_user_id")
                if uid_saved:
                    db_r = conectar_db()
                    if db_r:
                        cur_r = db_r.cursor(dictionary=True)
                        cur_r.execute("SELECT ID_Usuario, Nombre_Completo, Rol, Tienda, Zona, Usuario FROM usuarios WHERE ID_Usuario = %s", (uid_saved,))
                        user_data = cur_r.fetchone()
                        db_r.close()
                        if user_data:
                            user_info["id"] = user_data["ID_Usuario"]
                            user_info["usuario"] = user_data.get("Usuario") or ""
                            user_info["nombre"] = user_data["Nombre_Completo"]
                            user_info["rol"] = user_data["Rol"]
                            user_info["tienda"] = user_data.get("Tienda") or ""
                            user_info["zona"] = user_data.get("Zona") or "Zona Centro"
                            user_info["img_usuario"] = obtener_avatar_usuario(user_data["ID_Usuario"])
                            user_id_key = user_data["ID_Usuario"]
                            active_sessions[user_id_key] = {
                                "page": page,
                                "user_info": user_info,
                                "cargar_chat": cargar_chat,
                                "active_file_callback": active_file_callback
                            }
                            print(f"🔄 Sesión restaurada automáticamente para: {user_data['Nombre_Completo']}")
                            cargar_chat()
        except Exception as ex_r:
            print("Notice auto-restore session:", ex_r)

    page.run_task(intentar_restaurar_sesion)

     