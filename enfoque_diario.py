# -*- coding: utf-8 -*-
"""
MÓDULO ENFOQUE DIARIO 2026 - SUNGLASS HUT (SGH)
Sistema LUXO - Replicación fiel del Excel Oficial 2026 SGH ENFOQUE DIARIO.

Este módulo es 100% independiente y modular.
No modifica ninguna función existente del sistema.
"""

import os
import math
import datetime
import flet as ft

# Intentar importar reportlab para generación de PDF
try:
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

BASE_PATH = os.path.dirname(__file__)

import mysql.connector
from dotenv import load_dotenv
load_dotenv()

def conectar_db():
    try:
        return mysql.connector.connect(
            host=os.getenv('DB_HOST', 'localhost'),
            user=os.getenv('DB_USER', 'root'),
            password=os.getenv('DB_PASSWORD', ''),
            database=os.getenv('DB_NAME', 'sgh_portal'),
            port=int(os.getenv('DB_PORT', 3306))
        )
    except Exception as ex:
        print("Notice conectar_db in enfoque_diario:", ex)
        return None


# --- ESTADO GLOBAL Y MATRIZ DE DATOS ---
DIAS = ["DOMINGO", "LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES", "SÁBADO"]
COLOR_TABS = {
    "SEMANAL": "#A100F2",
    "DOMINGO": "#10B981",
    "PLAN.ACCIÓN_D": "#F59E0B",
    "LUNES": "#EF4444",
    "PLAN.ACCIÓN_L": "#F59E0B",
    "MARTES": "#EC4899",
    "PLAN.ACCIÓN_MA": "#F59E0B",
    "MIÉRCOLES": "#8B5CF6",
    "PLAN.ACCIÓN_MI": "#F59E0B",
    "JUEVES": "#3B82F6",
    "PLAN.ACCIÓN_J": "#F59E0B",
    "VIERNES": "#06B6D4",
    "PLAN.ACCIÓN_V": "#F59E0B",
    "SÁBADO": "#10B981",
    "PLAN.ACCIÓN_S": "#F59E0B"
}

# --- MATRIZ OFICIAL DE TIENDAS Y NÚMEROS REALES SGH ---
MAPEO_TIENDAS_SGH = {
    "3502": "Interlomas",
    "9277": "Ampliación Interlomas",
    "3019": "Toreo",
    "3645": "Vallejo",
    "3488": "Atizapán",
    "3586": "Punta Norte 1",
    "5256": "Punta Norte 2",
    "3583": "Satélite",
    "3507": "Lindavista",
    "8839": "Fortuna",
    "c964": "Explanada Pachuca",
    "q382": "Parque Tepeyac",
    "3542": "Galerías Pachuca",
    "3519": "Santa Fe",
    "9539": "Town Square Metepec",
    "108024": "Plaza Satélite"
}
MAPEO_NOMBRE_A_NUMERO_SGH = {v.upper(): k for k, v in MAPEO_TIENDAS_SGH.items()}

def cargar_mapeo_tiendas_db():
    db = conectar_db()
    if db:
        try:
            cur = db.cursor(dictionary=True)
            cur.execute("SELECT Usuario, Tienda FROM usuarios WHERE Usuario LIKE 'sgh%' AND Tienda IS NOT NULL AND Tienda != ''")
            for row in cur.fetchall():
                u_str = str(row["Usuario"]).lower().replace("sgh", "").strip()
                t_str = str(row["Tienda"]).strip()
                if u_str and t_str:
                    MAPEO_TIENDAS_SGH[u_str] = t_str
                    MAPEO_NOMBRE_A_NUMERO_SGH[t_str.upper()] = u_str
            db.close()
        except Exception as ex:
            print("Notice cargar_mapeo_tiendas_db:", ex)

try:
    cargar_mapeo_tiendas_db()
except Exception: pass

def default_global_meta():
    return {
        "semana": "30",
        "tienda": "Vallejo",
        "num_tienda": "3645"
    }

def default_store_state():
    state = {}
    for d in DIAS:
        state[d] = {
            "meta_diaria": 4758.0,
            "trafico_esperado": 8,
            "conversion_target": 0.13,
            "vta_ly": 4758.0,
            "wearables_pct": 0.15,
            "kids_pct": 0.05,
            "carekits_pct": 0.30,
            "atv_dia": 3620.0,
            "aur_dia": 3620.0,
            "atv_mtd": 7597.0,
            "aur_mtd": 3362.0,
            "estrellas_logro": 5,
            "trafico_bloques": [4, 2, 2, 0, 0],
            "colaboradores": [
                {"nombre": "", "horas": 0.0, "interacciones": 0, "convertidos": 0, "vta_cierre": 0.0, "ana_cierre": 0, "wea_demos": 0, "wea_cierre": 0, "kid_cierre": 0}
                for _ in range(8)
            ],
            "venta_neta_dia": 0.0,
            "venta_unidades_dia": 0,
            "slp_dia": "",
            "onesight_dia": "",
            "enfoque_hoy": "Enfocar el 100% del equipo en ofrecer la solución limpiadora y bandeja de opciones para maximizar venta múltiple.",
            "logros_hoy": "Excelente retención de clientes y venta cruzada.",
            "plan_accion": [
                {"colaborador": "", "compromiso": ""} for _ in range(3)
            ]
        }
    return state

user_states = {}

def get_state_file(user_id):
    return os.path.join(BASE_PATH, f"enfoque_diario_state_{user_id}.json")

def init_user_state(user_id):
    if user_id not in user_states:
        user_states[user_id] = {
            "global_meta": default_global_meta(),
            "store_state": default_store_state(),
            "historico_semanal_state": {},
            "active_tab": ["DOMINGO"]
        }
        cargar_estado_persistente(user_id)

import json
import threading

def _db_save_worker(user_id, json_str, sem_str, tienda_id):
    db = conectar_db()
    if db:
        try:
            cursor = db.cursor()
            cursor.execute("""
            INSERT INTO enfoque_diario_guardado (user_id, tienda_id, semana, estado_json)
            VALUES (%s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE tienda_id=VALUES(tienda_id), estado_json=VALUES(estado_json), fecha_actualizacion=NOW()
            """, (str(user_id), tienda_id, sem_str, json_str))
            db.commit()
        except Exception as ex_db:
            print(f"Notice DB save enfoque_diario for user {user_id}:", ex_db)
        finally:
            try: db.close()
            except: pass

def guardar_estado_persistente(user_id):
    try:
        if user_id not in user_states: return
        payload = {
            "global_meta": user_states[user_id]["global_meta"],
            "store_state": user_states[user_id]["store_state"],
            "historico_semanal_state": user_states[user_id]["historico_semanal_state"],
            "active_tab": user_states[user_id].get("active_tab", ["DOMINGO"])
        }
        json_str = json.dumps(payload, ensure_ascii=False)

        # 1. Guardar en archivo JSON local de respaldo (rápido en disco)
        with open(get_state_file(user_id), "w", encoding="utf-8") as f:
            f.write(json_str)

        # 2. Guardar en MySQL en segundo plano (Thread para NO congelar ni alentar la interfaz)
        sem_str = str(user_states[user_id]["global_meta"].get("semana", "34"))
        tienda_id = int(user_states[user_id]["global_meta"].get("num_tienda", 0))
        t = threading.Thread(target=_db_save_worker, args=(user_id, json_str, sem_str, tienda_id), daemon=True)
        t.start()
    except Exception as ex:
        print(f"Error al guardar estado de enfoque diario para {user_id}:", ex)

def cargar_estado_persistente(user_id):
    loaded_from_db = False
    try:
        # 1. Intentar cargar prioritariamente desde MySQL
        db = conectar_db()
        if db:
            try:
                cursor = db.cursor(dictionary=True)
                sem_str = str(user_states[user_id]["global_meta"].get("semana", "34"))
                cursor.execute(
                    "SELECT estado_json FROM enfoque_diario_guardado WHERE user_id=%s AND semana=%s ORDER BY id DESC LIMIT 1",
                    (str(user_id), sem_str)
                )
                row = cursor.fetchone()
                if not row:
                    cursor.execute(
                        "SELECT estado_json FROM enfoque_diario_guardado WHERE user_id=%s ORDER BY id DESC LIMIT 1",
                        (str(user_id),)
                    )
                    row = cursor.fetchone()
                if row and row.get("estado_json"):
                    payload = json.loads(row["estado_json"])
                    if "store_state" in payload:
                        for d in DIAS:
                            if d in payload["store_state"]:
                                user_states[user_id]["store_state"][d].update(payload["store_state"][d])
                    if "global_meta" in payload:
                        user_states[user_id]["global_meta"].update(payload["global_meta"])
                    if "historico_semanal_state" in payload:
                        user_states[user_id]["historico_semanal_state"].update(payload["historico_semanal_state"])
                    if "active_tab" in payload:
                        user_states[user_id]["active_tab"] = payload["active_tab"]
                    loaded_from_db = True
                    print(f"✅ Estado de Enfoque Diario para usuario {user_id} cargado desde MySQL con éxito.")
            except Exception as ex_db:
                print(f"Notice DB load enfoque_diario for user {user_id}:", ex_db)
            finally:
                try: db.close()
                except: pass

        # 2. Respaldo: Cargar desde archivo JSON local si no estaba en MySQL
        if not loaded_from_db:
            sf = get_state_file(user_id)
            if os.path.exists(sf):
                with open(sf, "r", encoding="utf-8") as f:
                    payload = json.load(f)
                    if "store_state" in payload:
                        for d in DIAS:
                            if d in payload["store_state"]:
                                user_states[user_id]["store_state"][d].update(payload["store_state"][d])
                    if "global_meta" in payload:
                        user_states[user_id]["global_meta"].update(payload["global_meta"])
                    if "historico_semanal_state" in payload:
                        user_states[user_id]["historico_semanal_state"].update(payload["historico_semanal_state"])
                    if "active_tab" in payload:
                        user_states[user_id]["active_tab"] = payload["active_tab"]
            else:
                guardar_estado_persistente(user_id)
    except Exception as ex:
        print(f"Error al cargar estado de enfoque diario para {user_id}:", ex)

def guardar_semana_historico(user_id):
    g_meta = user_states[user_id]["global_meta"]
    s_state = user_states[user_id]["store_state"]
    h_state = user_states[user_id]["historico_semanal_state"]
    key = f"S{g_meta['semana']}_{g_meta['num_tienda']}_{g_meta['tienda']}"
    import copy
    h_state[key] = copy.deepcopy(s_state)
    guardar_estado_persistente(user_id)

def cargar_semana_historico(user_id, num_semana):
    g_meta = user_states[user_id]["global_meta"]
    s_state = user_states[user_id]["store_state"]
    h_state = user_states[user_id]["historico_semanal_state"]
    g_meta["semana"] = str(num_semana)
    key = f"S{num_semana}_{g_meta['num_tienda']}_{g_meta['tienda']}"
    if key in h_state:
        import copy
        s_state.clear()
        s_state.update(copy.deepcopy(h_state[key]))
    guardar_estado_persistente(user_id)

def sincronizar_colaboradores_db(user_info=None, tienda_name=None, user_id=None):
    """Consulta los colaboradores registrados en la base de datos de Configuración de Tienda y los auto-llena en Enfoque Diario 2026."""
    if not user_id: return
    g_meta = user_states[user_id]["global_meta"]
    s_state = user_states[user_id]["store_state"]
    db_names = []
    try:
        target_t = tienda_name or g_meta.get("tienda", "Vallejo")
        db = conectar_db()
        if db:
            cursor = db.cursor(dictionary=True)
            cursor.execute("""
                SELECT v.Nombre_Completo 
                FROM vendedores v
                JOIN usuarios u ON v.ID_Usuario_Tienda = u.ID_Usuario
                WHERE LOWER(u.Tienda) = LOWER(%s) AND v.Activo = 1
                ORDER BY v.ID_Vendedor ASC
            """, (target_t,))
            rows = cursor.fetchall()
            if not rows and isinstance(user_info, dict):
                user_id = user_info.get("id", 1)
                cursor.execute("""
                    SELECT Nombre_Completo FROM vendedores
                    WHERE (ID_Usuario_Tienda = %s OR ID_Usuario_Tienda = 1) AND Activo = 1
                    ORDER BY ID_Vendedor ASC
                """, (user_id,))
                rows = cursor.fetchall()
            db.close()
            db_names = [r["Nombre_Completo"] for r in rows if r.get("Nombre_Completo")]
    except Exception as ex:
        print("Error sincronizando colaboradores DB:", ex)

    if db_names:
        for d in DIAS:
            for i in range(8):
                if i < len(db_names):
                    s_state[d]["colaboradores"][i]["nombre"] = db_names[i]
                    if s_state[d]["colaboradores"][i]["horas"] <= 0:
                        s_state[d]["colaboradores"][i]["horas"] = 10.0 if i == 0 else 8.0
                else:
                    s_state[d]["colaboradores"][i]["nombre"] = ""
                    s_state[d]["colaboradores"][i]["horas"] = 0.0
        guardar_estado_persistente(user_id)

# --- FUNCIONES MATEMÁTICAS EXPORTADAS AL MÓDULO ---
def calcular_dia(d_name, user_id):
    if user_id not in user_states: init_user_state(user_id)
    s_state = user_states[user_id]["store_state"]
    data = s_state[d_name]
    m_diaria = data["meta_diaria"]
    analogos = m_diaria * 0.85
    wearables = m_diaria * 0.15
    
    trafico = data["trafico_esperado"]
    conv = data["conversion_target"]
    transacciones = math.ceil(trafico * conv) if trafico > 0 else 0
    meta_ideal = m_diaria * 1.10
    total_unidades = max(transacciones, 1)

    vta_neta_prod = (m_diaria / total_unidades) if total_unidades > 0 else 0.0
    vta_ly = data["vta_ly"]

    b_trafico = data["trafico_bloques"]
    tot_trafico_b = sum(b_trafico)
    b_pesos = [(t / tot_trafico_b) if tot_trafico_b > 0 else 0.0 for t in b_trafico]
    b_metas = [p * m_diaria for p in b_pesos]

    colabs = data["colaboradores"]
    tot_horas = sum(c["horas"] for c in colabs if c["nombre"].strip() and c["horas"] > 0)
    u_prod = round(total_unidades / tot_horas, 2) if tot_horas > 0 else 0.0

    colab_rows = []
    for c in colabs:
        nom = c["nombre"].strip()
        hrs = c["horas"]
        if hrs > 0 and tot_horas > 0:
            m_vta = (m_diaria / tot_horas) * hrs
            
            tot_wea_unid = total_unidades * data.get("wearables_pct", 0.15)
            tot_kids_unid = total_unidades * data.get("kids_pct", 0.05)
            tot_ck_unid = total_unidades * data.get("carekits_pct", 0.30)
            
            calc_kid = math.ceil(max((tot_kids_unid / tot_horas) * hrs, 1))
            calc_ck = math.ceil(max((tot_ck_unid / tot_horas) * hrs, 1))
            calc_ana = math.ceil(max(((total_unidades - tot_wea_unid) / tot_horas) * hrs, 1))
            calc_wea = math.ceil(max((tot_wea_unid / tot_horas) * hrs, 1))
            
            def get_manual_or_calc(key, default_calc):
                val = c.get(key, "")
                if val != "":
                    try: return int(val)
                    except ValueError: return default_calc
                return default_calc

            m_ana = get_manual_or_calc("meta_ana", calc_ana)
            m_wea = get_manual_or_calc("meta_wea", calc_wea)
            m_kid = get_manual_or_calc("meta_kid", calc_kid)
            m_ck = get_manual_or_calc("meta_ck", calc_ck)
        else:
            m_vta = 0.0
            m_ana = 0
            m_wea = 0
            m_kid = 0
            m_ck = 0
        
        colab_rows.append({
            "nombre": nom,
            "horas": hrs,
            "meta_vta": m_vta,
            "meta_ana": m_ana,
            "meta_wea": m_wea,
            "meta_kid": m_kid,
            "meta_ck": m_ck,
            # CÓMO VAMOS (Bottom Half)
            "interacciones": c.get("interacciones", 0),
            "convertidos": c.get("convertidos", 0),
            "vta_cierre": c.get("vta_cierre", 0.0),
            "ana_cierre": c.get("ana_cierre", 0),
            "wea_demos": c.get("wea_demos", 0),
            "wea_cierre": c.get("wea_cierre", 0),
            "kid_cierre": c.get("kid_cierre", 0),
            "ck_cierre": c.get("ck_cierre", 0),
            "conversion_cierre": (c.get("convertidos", 0) / c.get("interacciones", 1)) if c.get("interacciones", 0) > 0 else 0.0,
            "conversion_wea": (c.get("wea_cierre", 0) / c.get("wea_demos", 1)) if c.get("wea_demos", 0) > 0 else 0.0
        })

    # Cálculos globales de CÓMO VAMOS
    tot_interacciones = sum(r["interacciones"] for r in colab_rows)
    tot_convertidos = sum(r["convertidos"] for r in colab_rows)
    tot_vta_cierre = sum(r["vta_cierre"] for r in colab_rows)
    tot_ana_cierre = sum(r["ana_cierre"] for r in colab_rows)
    tot_wea_demos = sum(r["wea_demos"] for r in colab_rows)
    tot_wea_cierre = sum(r["wea_cierre"] for r in colab_rows)
    tot_kid_cierre = sum(r["kid_cierre"] for r in colab_rows)
    tot_ck_cierre = sum(r["ck_cierre"] for r in colab_rows)
    
    venta_neta_dia = data.get("venta_neta_dia", 0.0)
    venta_unidades_dia = data.get("venta_unidades_dia", 0)
    
    conversion_dia = (venta_unidades_dia / tot_interacciones) if tot_interacciones > 0 else 0.0
    crecimiento_conversion = conversion_dia - conv
    
    # User overrideable Wearables/Kids %
    def get_pct_override(key, default_calc):
        val = data.get(key, "")
        if val != "":
            try: return float(val) / 100.0 if "%" not in str(val) else float(str(val).replace("%","")) / 100.0
            except ValueError: return default_calc
        return default_calc
        
    calc_wearables_pct = (tot_wea_cierre / venta_unidades_dia) if venta_unidades_dia > 0 else 0.0
    calc_kids_pct = (tot_kid_cierre / venta_unidades_dia) if venta_unidades_dia > 0 else 0.0
    
    cv_wearables_pct = get_pct_override("cv_wearables_pct", calc_wearables_pct)
    cv_kids_pct = get_pct_override("cv_kids_pct", calc_kids_pct)
    
    cv_carekits_pct = (tot_ck_cierre / venta_unidades_dia) if venta_unidades_dia > 0 else 0.0
    
    return {
        "meta_diaria": m_diaria,
        "analogos": analogos,
        "wearables": wearables,
        "total_unidades": total_unidades,
        "transacciones": transacciones,
        "meta_ideal": meta_ideal,
        "vta_neta_prod": vta_neta_prod,
        "u_prod": u_prod,
        "vta_ly": vta_ly,
        "tot_trafico_b": tot_trafico_b,
        "b_pesos": b_pesos,
        "b_metas": b_metas,
        "tot_horas": tot_horas,
        "colab_rows": colab_rows,
        # CÓMO VAMOS
        "tot_interacciones": tot_interacciones,
        "tot_convertidos": tot_convertidos,
        "tot_vta_cierre": tot_vta_cierre,
        "tot_ana_cierre": tot_ana_cierre,
        "tot_wea_demos": tot_wea_demos,
        "tot_wea_cierre": tot_wea_cierre,
        "tot_kid_cierre": tot_kid_cierre,
        "tot_ck_cierre": tot_ck_cierre,
        "venta_neta_dia": venta_neta_dia,
        "venta_unidades_dia": venta_unidades_dia,
        "conversion_dia": conversion_dia,
        "crecimiento_conversion": crecimiento_conversion,
        "wearables_pct": cv_wearables_pct,
        "kids_pct": cv_kids_pct,
        "carekits_pct": cv_carekits_pct
    }

# --- GENERADOR DE EXCEL OFICIAL SGH (.xlsx) ---
def generar_excel_enfoque(d_name, user_id, page=None):
    try:
        d_real = "DOMINGO" if d_name == "SEMANAL" else d_name
        calc = calcular_dia(d_real, user_id)
        g_meta = user_states[user_id]["global_meta"]
        s_state = user_states[user_id]["store_state"]

        template_path = os.path.abspath(os.path.join(BASE_PATH, "plantilla_sgh_2026.xlsx"))
        if not os.path.exists(template_path):
            template_path = os.path.abspath(os.path.join(BASE_PATH, "2026 SGH ENFOQUE DIARIO- Nuestra meta y plan de accion FINAL.xlsx"))

        excel_filename = f"Enfoque_Diario_{d_name}_SGH_2026.xlsx"
        uploads_dir = os.path.abspath(os.path.join(BASE_PATH, "uploads"))
        os.makedirs(uploads_dir, exist_ok=True)
        web_excel_path = os.path.abspath(os.path.join(uploads_dir, excel_filename))

        if os.path.exists(template_path):
            import shutil
            shutil.copy(template_path, web_excel_path)

            try:
                import win32com.client
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                excel.ScreenUpdating = False

                wb = excel.Workbooks.Open(web_excel_path)
                ws_names = [ws.Name for ws in wb.Worksheets]

                for d in DIAS:
                    if d in ws_names:
                        ws = wb.Worksheets(d)
                        d_data = s_state[d]

                        ws.Range('I1').Value = int(g_meta['semana']) if str(g_meta['semana']).isdigit() else g_meta['semana']
                        ws.Range('M1').Value = g_meta['tienda']

                        ws.Range('C5').Value = d_data['meta_diaria']
                        ws.Range('F5').Value = d_data['trafico_esperado']
                        ws.Range('F6').Value = d_data['conversion_target']
                        ws.Range('E9').Value = d_data['vta_ly']
                        ws.Range('I5').Value = d_data['wearables_pct']
                        ws.Range('I6').Value = d_data['kids_pct']
                        ws.Range('I7').Value = d_data['carekits_pct']

                        for i, val in enumerate(d_data['trafico_bloques']):
                            col_letter = ['C', 'D', 'E', 'F', 'G'][i]
                            ws.Range(f'{col_letter}12').Value = val

                        ws.Range('P7').Value = d_data.get('atv_dia', 7500.0)
                        ws.Range('P9').Value = d_data.get('aur_dia', 4617.0)
                        ws.Range('P13').Value = d_data.get('atv_mtd', 6578.0)
                        ws.Range('P15').Value = d_data.get('aur_mtd', 4312.0)

                        for i, c in enumerate(d_data['colaboradores']):
                            r_idx = 17 + i
                            if r_idx <= 24:
                                ws.Range(f'B{r_idx}').Value = c['nombre']
                                ws.Range(f'D{r_idx}').Value = c['horas']
                                if c.get("meta_ana", "") != "": ws.Range(f'F{r_idx}').Value = int(c["meta_ana"])
                                if c.get("meta_wea", "") != "": ws.Range(f'G{r_idx}').Value = int(c["meta_wea"])
                                if c.get("meta_kid", "") != "": ws.Range(f'H{r_idx}').Value = int(c["meta_kid"])
                                if c.get("meta_ck", "") != "": ws.Range(f'I{r_idx}').Value = int(c["meta_ck"])
                                
                            r_cv_idx = 33 + i
                            if r_cv_idx <= 40:
                                ws.Range(f'E{r_cv_idx}').Value = c.get('interacciones', 0)
                                ws.Range(f'G{r_cv_idx}').Value = c.get('convertidos', 0)
                                ws.Range(f'J{r_cv_idx}').Value = c.get('vta_cierre', 0.0)
                                ws.Range(f'K{r_cv_idx}').Value = c.get('ana_cierre', 0)
                                ws.Range(f'L{r_cv_idx}').Value = c.get('wea_demos', 0)
                                ws.Range(f'M{r_cv_idx}').Value = c.get('wea_cierre', 0)
                                ws.Range(f'O{r_cv_idx}').Value = c.get('kid_cierre', 0)
                                ws.Range(f'P{r_cv_idx}').Value = c.get('ck_cierre', 0)

                        ws.Range('E30').Value = d_data.get('venta_neta_dia', 0.0)
                        ws.Range('G30').Value = d_data.get('venta_unidades_dia', 0)

                wb.Save()
                wb.Close(False)
                excel.Quit()
                print(f"✅ Excel generado exitosamente vía win32com ({os.path.getsize(web_excel_path)/(1024*1024):.2f} MB)")
                return web_excel_path
            except Exception as ex_com:
                print("Notice win32com excel save fallback:", ex_com)
                import openpyxl
                wb = openpyxl.load_workbook(template_path)
                ws = wb[d_real if d_real in wb.sheetnames else "DOMINGO"]
                ws['C5'] = s_state.get(d_real, {}).get('meta_diaria', 0.0)
                wb.save(web_excel_path)
                wb.close()
                return web_excel_path
        return web_excel_path
    except Exception as ex:
        print("Error en generar_excel_enfoque:", ex)
        return None

        home_dir = os.path.expanduser("~")
        possible_desktops = [
            os.path.join(home_dir, "Desktop"),
            os.path.join(home_dir, "OneDrive", "Desktop")
        ]
        primary_desk = None
        for d_path in possible_desktops:
            if os.path.exists(d_path):
                target_file = os.path.join(d_path, excel_filename)
                try:
                    import shutil
                    shutil.copy2(web_excel_path, target_file)
                    if not primary_desk:
                        primary_desk = target_file
                except Exception:
                    pass

        if primary_desk and os.path.exists(primary_desk):
            try:
                os.startfile(primary_desk)
            except Exception:
                pass

        if page:
            snack = ft.SnackBar(ft.Text(f"📊 Excel oficial generado y guardado en tu Escritorio: {excel_filename}", color="white"), bgcolor="#10B981")
            page.overlay.append(snack)
            snack.open = True
            page.update()

        return web_excel_path

    except Exception as ex_excel:
        print("Error generando Excel Enfoque Diario:", ex_excel)
        if page:
            snack = ft.SnackBar(ft.Text(f"❌ Error generando Excel: {ex_excel}", color="white"), bgcolor="red")
            page.overlay.append(snack)
            snack.open = True
            page.update()
        return None

# --- GENERADOR DE REPORTES EN PDF ---
def generar_pdf_enfoque_file(d_name, user_id):
    try:
        d_real = "DOMINGO" if d_name == "SEMANAL" else d_name
        excel_path = generar_excel_enfoque(d_real, user_id)

        pdf_filename = f"Enfoque_Diario_{d_name}_SGH_2026.pdf"
        uploads_dir = os.path.join(BASE_PATH, "uploads")
        os.makedirs(uploads_dir, exist_ok=True)
        web_pdf_path = os.path.join(uploads_dir, pdf_filename)

        abs_excel = os.path.abspath(excel_path)
        abs_pdf = os.path.abspath(web_pdf_path)

        # Conversión directa del Excel Oficial a PDF via Windows Excel COM
        try:
            import win32com.client
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(abs_excel)
            sheet_name = d_real if d_real in [ws.Name for ws in wb.Worksheets] else "DOMINGO"
            ws = wb.Worksheets(sheet_name)
            ws.ExportAsFixedFormat(0, abs_pdf)
            wb.Close(False)
            excel.Quit()
            if os.path.exists(abs_pdf):
                print(f"✅ PDF convertido idéntico desde Excel oficial a través de win32com: {abs_pdf}")
                return abs_pdf
        except Exception as ex_com:
            print("Notice win32com PDF conversion:", ex_com)

        return web_pdf_path if os.path.exists(web_pdf_path) else None
    except Exception as ex_gen:
        print("Error en generar_pdf_enfoque_file:", ex_gen)
        return None

        story = []
        story.append(Paragraph(f"<b>SUNGLASS HUT (SGH) - ENFOQUE DIARIO 2026</b>", title_style))
        story.append(Paragraph(f"<b>DÍA:</b> {d_name} | <b>SEMANA:</b> {g_meta['semana']} | <b>TIENDA:</b> {g_meta['tienda']} | <b>FECHA EMISIÓN:</b> {datetime.date.today().strftime('%d/%m/%Y')}", sub_title_style))
        story.append(Spacer(1, 6))

        story.append(Paragraph("<b>1. META DEL DÍA Y NO NEGOCIABLES (RESUMEN OPERATIVO)</b>", h2_style))
        meta_data_table = [
            [
                Paragraph("<b>META DEL DÍA</b>", black_hdr_style),
                Paragraph("<b>CONVERSIÓN</b>", black_hdr_style),
                Paragraph("<b>NO NEGOCIABLES</b>", black_hdr_style),
                Paragraph("<b>PRODUCTIVIDAD</b>", black_hdr_style)
            ],
            [
                f"Meta Diaria: ${calc['meta_diaria']:,.2f}\nAnálogos (85%): ${calc['analogos']:,.2f}\nWearables (15%): ${calc['wearables']:,.2f}\nTotal Unidades: {calc['total_unidades']}\nEvaluación: {'⭐' * data['estrellas_logro']}",
                f"Tráfico Esperado: {data['trafico_esperado']}\nConversión LY+1: {int(data['conversion_target']*100)}%\nTransacciones Target: {calc['transacciones']}\nMeta Ideal (110%): ${calc['meta_ideal']:,.2f}",
                f"Wearables: 15% (Min 1)\nKids: 5% (Min 1)\nCarekits: 30% (Min 1)",
                f"Vta Neta: ${calc['vta_neta_prod']:,.2f}\nU.Prod: {calc['u_prod']}\nVta LY: ${calc['vta_ly']:,.2f}"
            ]
        ]
        t_meta = Table(meta_data_table, colWidths=[135, 135, 135, 135])
        t_meta.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#000000')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#999999')),
            ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#F3F4F6')),
            ('PADDING', (0,0), (-1,-1), 6),
            ('VALIGN', (0,0), (-1,-1), 'TOP')
        ]))
        story.append(t_meta)
        story.append(Spacer(1, 8))

        story.append(Paragraph("<b>2. DESGLOSE HORARIO DE VENTA</b>", h2_style))
        bloques_names = ["Apertura-1pm", "1pm - 3pm", "3pm - 5pm", "5pm - 7pm", "7pm - Cierre", "TOTAL"]
        hdr_b = [Paragraph(f"<b>{b}</b>", black_hdr_style) for b in bloques_names]
        hdr_b.insert(0, Paragraph("<b>BLOQUE / INDICADOR</b>", black_hdr_style))

        row_trafico = ["Tráfico (⚪)"] + [str(b) for b in data["trafico_bloques"]] + [str(calc["tot_trafico_b"])]
        row_peso = ["Peso % (🟩)"] + [f"{p*100:.1f}%" for p in calc["b_pesos"]] + ["100%"]
        row_meta = ["Meta $ (🟩)"] + [f"${m:,.0f}" for m in calc["b_metas"]] + [f"${calc['meta_diaria']:,.0f}"]

        t_bloques = Table([hdr_b, row_trafico, row_peso, row_meta], colWidths=[115, 70, 70, 70, 70, 75, 70])
        t_bloques.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#000000')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#999999')),
            ('BACKGROUND', (1,2), (-1,3), colors.HexColor('#E6F4EA')),
            ('PADDING', (0,0), (-1,-1), 5)
        ]))
        story.append(t_bloques)
        story.append(Spacer(1, 8))

        story.append(Paragraph("<b>3. ASIGNACIÓN POR COLABORADOR</b>", h2_style))
        colab_hdr = [Paragraph(h, black_hdr_style) for h in ["COLABORADOR", "HORAS", "META VENTA", "ANÁLOGOS", "WEARABLES", "KIDS", "CAREKITS"]]
        colab_table_data = [colab_hdr]

        for r in calc["colab_rows"]:
            if r["nombre"]:
                colab_table_data.append([
                    r["nombre"], f"{r['horas']:.1f}", f"${r['meta_vta']:,.2f}",
                    str(r["meta_ana"]), str(r["meta_wea"]), str(r["meta_kid"]), str(r["meta_ck"])
                ])

        colab_table_data.append([
            "TOTAL TIENDA", f"{calc['tot_horas']:.1f}", f"${calc['meta_diaria']:,.2f}",
            str(sum(r["meta_ana"] for r in calc["colab_rows"])),
            str(sum(r["meta_wea"] for r in calc["colab_rows"])),
            str(sum(r["meta_kid"] for r in calc["colab_rows"])),
            str(sum(r["meta_ck"] for r in calc["colab_rows"]))
        ])

        t_colab = Table(colab_table_data, colWidths=[120, 55, 95, 65, 70, 65, 70])
        t_colab.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#000000')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#999999')),
            ('BACKGROUND', (2,1), (-1,-2), colors.HexColor('#E6F4EA')),
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#D1D5DB')),
            ('TEXTCOLOR', (0,-1), (-1,-1), colors.black),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('PADDING', (0,0), (-1,-1), 5)
        ]))
        story.append(t_colab)
        story.append(Spacer(1, 8))

        story.append(Paragraph("<b>4. ¿CÓMO VAMOS? (RESULTADOS AL CIERRE)</b>", h2_style))
        
        cv_hdr1 = [Paragraph(h, black_hdr_style) for h in ["META", "VTA NETA", "META UNID", "VTA UNID", "CONVERSIÓN", "CRECIM.", "WEARABLES %", "KIDS %", "CAREKITS %"]]
        cv_row1 = [
            f"${calc['meta_diaria']:,.2f}", f"${calc['venta_neta_dia']:,.2f}", str(calc['total_unidades']),
            str(calc['venta_unidades_dia']), f"{calc['conversion_dia']*100:.1f}%", f"{calc['crecimiento_conversion']*100:.1f}%",
            f"{calc['wearables_pct']*100:.1f}%", f"{calc['kids_pct']*100:.1f}%", f"{calc['carekits_pct']*100:.1f}%"
        ]
        t_cv1 = Table([cv_hdr1, cv_row1], colWidths=[60, 60, 60, 60, 60, 50, 70, 50, 70])
        t_cv1.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F59E0B')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#999999')),
            ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#FEF3C7')),
            ('PADDING', (0,0), (-1,-1), 5),
            ('ALIGN', (0,0), (-1,-1), 'CENTER')
        ]))
        story.append(t_cv1)
        story.append(Spacer(1, 6))

        cv_colab_hdr = [Paragraph(h, black_hdr_style) for h in ["COLABORADOR", "INT.", "CONV.", "CONV.%", "VTA CIERRE", "WEA.C.", "KIDS C.", "CK C."]]
        cv_colab_data = [cv_colab_hdr]
        for r in calc["colab_rows"]:
            if r["nombre"]:
                cv_colab_data.append([
                    r["nombre"], str(r["interacciones"]), str(r["convertidos"]), f"{r['conversion_cierre']*100:.1f}%",
                    f"${r['vta_cierre']:,.2f}", str(r["wea_cierre"]), str(r["kid_cierre"]), str(r["ck_cierre"])
                ])
        cv_colab_data.append([
            "TOTAL TIENDA", str(calc["tot_interacciones"]), str(calc["tot_convertidos"]), "",
            f"${calc['tot_vta_cierre']:,.2f}", str(calc["tot_wea_cierre"]), str(calc["tot_kid_cierre"]), str(calc["tot_ck_cierre"])
        ])
        t_cv_colab = Table(cv_colab_data, colWidths=[100, 40, 40, 50, 70, 50, 50, 50])
        t_cv_colab.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F59E0B')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#999999')),
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#D1D5DB')),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('PADDING', (0,0), (-1,-1), 5),
            ('ALIGN', (1,0), (-1,-1), 'CENTER')
        ]))
        story.append(t_cv_colab)
        story.append(Spacer(1, 8))

        story.append(Paragraph("<b>5. PLAN DE ACCIÓN, LOS 5 SECRETOS Y CUSTOMER JOURNEY</b>", h2_style))
        story.append(Paragraph(f"<b>Los 5 Secretos:</b> 1. Pulir es poder | 2. Póntelos | 3. Diviértete más | 4. Cuídalos | 5. Ajuste perfecto", normal_style))
        story.append(Paragraph(f"<b>Customer Journey:</b> 1. Empieza una relación | 2. Gánate su confianza | 3. Interactúa y relaciona | 4. Descubre y aprende | 5. Ve más allá", normal_style))
        story.append(Spacer(1, 4))
        story.append(Paragraph(f"<b>Tu Enfoque Para Hoy:</b> {data['enfoque_hoy'] or 'Seguimiento continuo al 100% de la Meta Diaria.'}", normal_style))
        story.append(Spacer(1, 4))
        story.append(Paragraph(f"<b>Logros y Oportunidades:</b> {data['logros_hoy'] or 'Mantener la disciplina en el Customer Journey y Los 5 Secretos.'}", normal_style))

        doc = SimpleDocTemplate(
            web_pdf_path, pagesize=letter,
            rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36
        )
        doc.build(story)

        home_dir = os.path.expanduser("~")
        possible_desktops = [
            os.path.join(home_dir, "Desktop"),
            os.path.join(home_dir, "OneDrive", "Desktop")
        ]
        for d_path in possible_desktops:
            if os.path.exists(d_path):
                try:
                    import shutil
                    shutil.copy2(web_pdf_path, os.path.join(d_path, pdf_filename))
                except Exception:
                    pass

        return web_pdf_path
    except Exception as ex_gen:
        print("Error en generar_pdf_enfoque_file:", ex_gen)
        return None


def build_enfoque_diario_view(page: ft.Page, session_user: dict = None):
    """
    Construye la vista principal del Módulo Enfoque Diario 2026 SGH
    con navegación por pestañas (Resumen Semanal, Días y Planes de Acción).
    """
    if session_user is None:
        session_user = {"user": "invitado"}
    user_id = session_user.get("user", "invitado")
    init_user_state(user_id)
    g_meta = user_states[user_id]["global_meta"]

    # Sincronizar dinámicamente tienda y número de tienda según el usuario en sesión
    if isinstance(session_user, dict):
        s_tienda = session_user.get("tienda")
        s_usuario = str(session_user.get("usuario", "")).lower()
        if s_tienda and s_tienda != "Tienda Luxo":
            g_meta["tienda"] = s_tienda
            if s_usuario.startswith("sgh") and len(s_usuario) > 3:
                g_meta["num_tienda"] = s_usuario.replace("sgh", "")
            elif s_tienda.upper() in MAPEO_NOMBRE_A_NUMERO_SGH:
                g_meta["num_tienda"] = MAPEO_NOMBRE_A_NUMERO_SGH[s_tienda.upper()]
    s_state = user_states[user_id]["store_state"]
    h_state = user_states[user_id]["historico_semanal_state"]
    
    # Sincronización asíncrona de colaboradores para NO bloquear el cambio de pestaña
    threading.Thread(target=sincronizar_colaboradores_db, args=(session_user, g_meta["tienda"], user_id), daemon=True).start()

    def generar_pdf_enfoque(target_day=None):
        if not REPORTLAB_AVAILABLE:
            snack = ft.SnackBar(ft.Text("❌ La librería ReportLab no está instalada para generar PDF.", color="white"), bgcolor="red")
            page.overlay.append(snack)
            snack.open = True
            page.update()
            return

        try:
            if target_day:
                d_name = target_day
            else:
                curr_tab = user_states[user_id]['active_tab'][0]
                d_name = curr_tab.replace("PLAN.ACCIÓN_", "")
                code_map = {"D": "DOMINGO", "L": "LUNES", "MA": "MARTES", "MI": "MIÉRCOLES", "J": "JUEVES", "V": "VIERNES", "S": "SÁBADO"}
                d_name = code_map.get(d_name, d_name if d_name in DIAS else "DOMINGO")

            pdf_path = generar_pdf_enfoque_file(d_name, user_id)
            pdf_filename = f"Enfoque_Diario_{d_name}_SGH_2026.pdf"
            pdf_url = f"/uploads/{pdf_filename}"

            def cerrar_dialogo(e):
                dlg.open = False
                page.update()

            dlg = ft.AlertDialog(
                title=ft.Row([
                    ft.Icon(ft.Icons.PICTURE_AS_PDF_ROUNDED, color="#10B981", size=24),
                    ft.Text(f"Reporte PDF ({d_name}) Listo 📄", color="white", weight="bold", size=16)
                ], spacing=8),
                content=ft.Column([
                    ft.Text("Tu reporte ha sido generado y guardado en tu Escritorio de Windows:", color="#CCCCCC", size=12),
                    ft.Container(
                        content=ft.Row([
                            ft.Icon(ft.Icons.FOLDER_ROUNDED, color="#FFD700", size=18),
                            ft.Text(pdf_filename, color="#00FFFF", weight="bold", size=12)
                        ], spacing=6),
                        bgcolor="#111827", padding=12, border_radius=8, border=ft.Border.all(1, "#374151")
                    ),
                    ft.Text("Haz clic en el botón de abajo para abrir o descargar el PDF directamente en tu navegador:", color="#AAAAAA", size=11),
                ], spacing=12, main_axis_alignment="center", tight=True),
                actions=[
                    ft.ElevatedButton(
                        content=ft.Row([
                            ft.Icon(ft.Icons.OPEN_IN_NEW_ROUNDED, color="white", size=16),
                            ft.Text("⬇️ Abrir / Descargar PDF en Navegador", color="white", weight="bold", size=12)
                        ], spacing=6),
                        style=ft.ButtonStyle(bgcolor="#10B981", shape=ft.RoundedRectangleBorder(radius=8)),
                        url=pdf_url,
                        on_click=cerrar_dialogo
                    ),
                    ft.TextButton("Cerrar", on_click=cerrar_dialogo)
                ],
                actions_alignment="end",
                bgcolor="#06070B",
            )
            page.dialog = dlg
            dlg.open = True
            page.update()

            if pdf_path and os.path.exists(pdf_path):
                try:
                    os.startfile(pdf_path)
                except Exception:
                    pass

        except Exception as ex:
            print("Error en generar_pdf_enfoque:", ex)

    # --- COMPONENTES VISUALES ---
    tab_view_cache = {}

    # Re-renderizador del contenedor activo
    def update_active_view():
        curr_tab = user_states[user_id]['active_tab'][0]
        if curr_tab not in tab_view_cache:
            if curr_tab == "SEMANAL":
                tab_view_cache[curr_tab] = build_semanal_ui()
            elif curr_tab.startswith("PLAN.ACCIÓN_"):
                d_code = curr_tab.replace("PLAN.ACCIÓN_", "")
                code_map = {"D": "DOMINGO", "L": "LUNES", "MA": "MARTES", "MI": "MIÉRCOLES", "J": "JUEVES", "V": "VIERNES", "S": "SÁBADO"}
                d_real = code_map.get(d_code, "DOMINGO")
                tab_view_cache[curr_tab] = build_plan_accion_ui(d_real)
            else:
                tab_view_cache[curr_tab] = build_sheet_ui(curr_tab)

        tab_content_container.content = tab_view_cache[curr_tab]
            
        try:
            if curr_tab == "SEMANAL":
                btn_download_excel.url = f"/api/download_excel/SEMANAL?user_id={user_id}"
                btn_download_pdf.url = f"/print_enfoque/SEMANAL?user_id={user_id}"
            else:
                d_name = curr_tab
                if d_name.startswith("PLAN.ACCIÓN_"):
                    d_code = d_name.replace("PLAN.ACCIÓN_", "")
                    code_map = {"D": "DOMINGO", "L": "LUNES", "MA": "MARTES", "MI": "MIÉRCOLES", "J": "JUEVES", "V": "VIERNES", "S": "SÁBADO"}
                    d_name = code_map.get(d_code, "DOMINGO")
                btn_download_excel.url = f"/api/download_excel/{d_name}?user_id={user_id}"
                btn_download_pdf.url = f"/print_enfoque/{d_name}?user_id={user_id}"

            btn_download_excel.update()
            btn_download_pdf.update()
        except Exception:
            pass
            
        page.update()

    # Callback al modificar celdas globales de Semana/Tienda
    def on_global_header_change(e):
        if e.control.data == "semana":
            g_meta["semana"] = e.control.value
        elif e.control.data == "tienda":
            g_meta["tienda"] = e.control.value
        guardar_estado_persistente(user_id)

    # --- CONSTRUCCIÓN DE INTERFAZ GRÁFICA DE HOJA DIARIA ---
    def build_sheet_ui(d_name):
        calc = calcular_dia(d_name, user_id)
        data = s_state[d_name]
        green_txts = {}

        def make_green_calc(key, val_str, width=120):
            t_obj = ft.Text(val_str, size=12, weight="bold", color="#00FF88")
            green_txts[key] = t_obj
            return ft.Container(
                content=t_obj,
                bgcolor="#052C1E",
                border=ft.Border.all(1, "#10B981"),
                border_radius=6,
                padding=8,
                alignment=ft.alignment.Alignment(0, 0),
                width=width
            )

        def sync_green_cells():
            c = calcular_dia(d_name, user_id)
            if "analogos" in green_txts: green_txts["analogos"].value = f"${c['analogos']:,.2f}"
            if "wearables" in green_txts: green_txts["wearables"].value = f"${c['wearables']:,.2f}"
            if "total_unidades" in green_txts: green_txts["total_unidades"].value = f"{c['total_unidades']} Pza"
            if "transacciones" in green_txts: green_txts["transacciones"].value = f"{c['transacciones']} Transac."
            if "meta_ideal" in green_txts: green_txts["meta_ideal"].value = f"${c['meta_ideal']:,.2f}"
            if "vta_neta_prod" in green_txts: green_txts["vta_neta_prod"].value = f"${c['vta_neta_prod']:,.2f}"
            if "u_prod" in green_txts: green_txts["u_prod"].value = f"{c['u_prod']}"
            if "tot_trafico_b" in green_txts: green_txts["tot_trafico_b"].value = str(c["tot_trafico_b"])
            
            for idx, p in enumerate(c["b_pesos"]):
                if f"b_peso_{idx}" in green_txts: green_txts[f"b_peso_{idx}"].value = f"{p*100:.1f}%"
            for idx, m in enumerate(c["b_metas"]):
                if f"b_meta_{idx}" in green_txts: green_txts[f"b_meta_{idx}"].value = f"${m:,.0f}"
            if "b_meta_tot" in green_txts: green_txts["b_meta_tot"].value = f"${c['meta_diaria']:,.0f}"

            for idx, r in enumerate(c["colab_rows"]):
                if f"colab_vta_{idx}" in green_txts: green_txts[f"colab_vta_{idx}"].value = f"${r['meta_vta']:,.2f}"
                if f"colab_ana_{idx}" in green_txts: green_txts[f"colab_ana_{idx}"].value = str(r['meta_ana'])
                if f"colab_wea_{idx}" in green_txts: green_txts[f"colab_wea_{idx}"].value = str(r['meta_wea'])
                if f"colab_kid_{idx}" in green_txts: green_txts[f"colab_kid_{idx}"].value = str(r['meta_kid'])
                if f"colab_ck_{idx}" in green_txts: green_txts[f"colab_ck_{idx}"].value = str(r['meta_ck'])
                
                # CÓMO VAMOS green cells
                if f"colab_cv_hrs_{idx}" in green_txts: green_txts[f"colab_cv_hrs_{idx}"].value = f"{r['horas']:.1f}"
                if f"colab_conv_cierre_{idx}" in green_txts: green_txts[f"colab_conv_cierre_{idx}"].value = f"{r['conversion_cierre']*100:.1f}%"
                if f"colab_conv_wea_{idx}" in green_txts: green_txts[f"colab_conv_wea_{idx}"].value = f"{r['conversion_wea']*100:.1f}%"

            if "tot_colab_hrs" in green_txts: green_txts["tot_colab_hrs"].value = f"{c['tot_horas']:.1f} hrs"
            if "tot_colab_vta" in green_txts: green_txts["tot_colab_vta"].value = f"${c['meta_diaria']:,.2f}"
            if "tot_colab_ana" in green_txts: green_txts["tot_colab_ana"].value = str(sum(r["meta_ana"] for r in c["colab_rows"]))
            if "tot_colab_wea" in green_txts: green_txts["tot_colab_wea"].value = str(sum(r["meta_wea"] for r in c["colab_rows"]))
            if "tot_colab_kid" in green_txts: green_txts["tot_colab_kid"].value = str(sum(r["meta_kid"] for r in c["colab_rows"]))
            if "tot_colab_ck" in green_txts: green_txts["tot_colab_ck"].value = str(sum(r["meta_ck"] for r in c["colab_rows"]))

            # Global CÓMO VAMOS
            if "tot_cv_meta" in green_txts: green_txts["tot_cv_meta"].value = f"${c['meta_diaria']:,.2f}"
            if "tot_cv_meta_unidades" in green_txts: green_txts["tot_cv_meta_unidades"].value = str(c["total_unidades"])
            if "tot_cv_conversion" in green_txts: green_txts["tot_cv_conversion"].value = f"{c['conversion_dia']*100:.1f}%"
            if "tot_cv_crecimiento" in green_txts: green_txts["tot_cv_crecimiento"].value = f"{c['crecimiento_conversion']*100:.1f}%"
            if "tot_cv_wearables_pct" in green_txts: green_txts["tot_cv_wearables_pct"].value = f"{c['wearables_pct']*100:.1f}%"
            if "tot_cv_kids_pct" in green_txts: green_txts["tot_cv_kids_pct"].value = f"{c['kids_pct']*100:.1f}%"

            # Sumas columna CÓMO VAMOS
            if "sum_cv_interacciones" in green_txts: green_txts["sum_cv_interacciones"].value = str(c["tot_interacciones"])
            if "sum_cv_convertidos" in green_txts: green_txts["sum_cv_convertidos"].value = str(c["tot_convertidos"])
            if "sum_cv_vta_cierre" in green_txts: green_txts["sum_cv_vta_cierre"].value = f"${c['tot_vta_cierre']:,.2f}"
            if "sum_cv_ana_cierre" in green_txts: green_txts["sum_cv_ana_cierre"].value = str(c["tot_ana_cierre"])
            if "sum_cv_wea_demos" in green_txts: green_txts["sum_cv_wea_demos"].value = str(c["tot_wea_demos"])
            if "sum_cv_wea_cierre" in green_txts: green_txts["sum_cv_wea_cierre"].value = str(c["tot_wea_cierre"])
            if "sum_cv_kid_cierre" in green_txts: green_txts["sum_cv_kid_cierre"].value = str(c["tot_kid_cierre"])

        def on_white_cell_change(e):
            try:
                v_raw = e.control.value or ""
                v = v_raw.replace(",", "").strip()
                if e.control.data == "meta_diaria":
                    data["meta_diaria"] = float(v) if v else 0.0
                elif e.control.data == "trafico_esperado":
                    data["trafico_esperado"] = int(v) if v else 0
                elif e.control.data == "conversion_target":
                    data["conversion_target"] = (float(v) / 100.0) if v else 0.0
                elif e.control.data == "vta_ly":
                    val_num = float(v) if v else 0.0
                    data["vta_ly"] = val_num
                    if d_name == "DOMINGO":
                        for day_other in DIAS:
                            if day_other in s_state:
                                s_state[day_other]["vta_ly"] = val_num
                elif e.control.data == "atv_dia":
                    val_num = float(v) if v else 0.0
                    data["atv_dia"] = val_num
                    if d_name == "DOMINGO":
                        for day_other in DIAS:
                            if day_other in s_state:
                                s_state[day_other]["atv_dia"] = val_num
                elif e.control.data == "aur_dia":
                    val_num = float(v) if v else 0.0
                    data["aur_dia"] = val_num
                    if d_name == "DOMINGO":
                        for day_other in DIAS:
                            if day_other in s_state:
                                s_state[day_other]["aur_dia"] = val_num
                elif e.control.data == "atv_mtd":
                    val_num = float(v) if v else 0.0
                    data["atv_mtd"] = val_num
                    if d_name == "DOMINGO":
                        for day_other in DIAS:
                            if day_other in s_state:
                                s_state[day_other]["atv_mtd"] = val_num
                elif e.control.data == "aur_mtd":
                    val_num = float(v) if v else 0.0
                    data["aur_mtd"] = val_num
                    if d_name == "DOMINGO":
                        for day_other in DIAS:
                            if day_other in s_state:
                                s_state[day_other]["aur_mtd"] = val_num
                elif e.control.data == "venta_neta_dia":
                    data["venta_neta_dia"] = float(v) if v else 0.0
                elif e.control.data == "venta_unidades_dia":
                    data["venta_unidades_dia"] = int(v) if v else 0
                elif e.control.data == "slp_dia":
                    data["slp_dia"] = v_raw
                elif e.control.data == "onesight_dia":
                    data["onesight_dia"] = v_raw
                elif e.control.data.startswith("trafico_b_"):
                    idx = int(e.control.data.split("_")[-1])
                    data["trafico_bloques"][idx] = int(v) if v else 0
                elif e.control.data.startswith("colab_nom_"):
                    idx = int(e.control.data.split("_")[-1])
                    data["colaboradores"][idx]["nombre"] = v_raw
                elif e.control.data.startswith("colab_hrs_"):
                    idx = int(e.control.data.split("_")[-1])
                    data["colaboradores"][idx]["horas"] = float(v) if v else 0.0
                elif e.control.data.startswith("colab_int_"):
                    idx = int(e.control.data.split("_")[-1])
                    data["colaboradores"][idx]["interacciones"] = int(v) if v else 0
                elif e.control.data.startswith("colab_conv_"):
                    idx = int(e.control.data.split("_")[-1])
                    data["colaboradores"][idx]["convertidos"] = int(v) if v else 0
                elif e.control.data.startswith("colab_vtac_"):
                    idx = int(e.control.data.split("_")[-1])
                    data["colaboradores"][idx]["vta_cierre"] = float(v) if v else 0.0
                elif e.control.data.startswith("colab_anac_"):
                    idx = int(e.control.data.split("_")[-1])
                    data["colaboradores"][idx]["ana_cierre"] = int(v) if v else 0
                elif e.control.data.startswith("colab_wead_"):
                    idx = int(e.control.data.split("_")[-1])
                    data["colaboradores"][idx]["wea_demos"] = int(v) if v else 0
                elif e.control.data.startswith("colab_weac_"):
                    idx = int(e.control.data.split("_")[-1])
                    data["colaboradores"][idx]["wea_cierre"] = int(v) if v else 0
                elif e.control.data.startswith("colab_kidc_"):
                    idx = int(e.control.data.split("_")[-1])
                    data["colaboradores"][idx]["kid_cierre"] = int(v) if v else 0
                elif e.control.data.startswith("colab_ckc_"):
                    idx = int(e.control.data.split("_")[-1])
                    data["colaboradores"][idx]["ck_cierre"] = int(v) if v else 0
                elif e.control.data.startswith("colab_wea_"):
                    idx = int(e.control.data.split("_")[-1])
                    data["colaboradores"][idx]["meta_wea"] = str(v_raw)
                elif e.control.data.startswith("colab_kid_"):
                    idx = int(e.control.data.split("_")[-1])
                    data["colaboradores"][idx]["meta_kid"] = str(v_raw)
                elif e.control.data == "cv_wearables_pct":
                    data["cv_wearables_pct"] = v_raw
                elif e.control.data == "cv_kids_pct":
                    data["cv_kids_pct"] = v_raw
            except Exception:
                pass
            
            sync_green_cells()
            guardar_estado_persistente(user_id)
            try: page.update()
            except Exception: pass

        # Componente Celda Blanca (Entrada editable ⚪)
        def make_white_input(val, data_id, width=110, suffix=""):
            val_formatted = str(val)
            try:
                if isinstance(val, (int, float)) and val != 0:
                    f_val = float(val)
                    if f_val.is_integer():
                        val_formatted = f"{int(f_val):,}"
                    else:
                        val_formatted = f"{f_val:,.2f}"
            except Exception: pass

            def on_blur_format_commas(e):
                try:
                    raw_txt = (e.control.value or "").replace(",", "").strip()
                    if raw_txt:
                        num = float(raw_txt)
                        if num.is_integer():
                            fmt = f"{int(num):,}"
                        else:
                            fmt = f"{num:,.2f}"
                        e.control.value = fmt
                        e.control.update()
                except Exception: pass

            return ft.Container(
                content=ft.TextField(
                    value=val_formatted,
                    data=data_id,
                    on_change=on_white_cell_change,
                    on_blur=on_blur_format_commas,
                    text_size=12,
                    text_style=ft.TextStyle(weight="bold", color="#FFFFFF"),
                    bgcolor="#1F2937",
                    border_color="#374151",
                    focused_border_color="#00FFFF",
                    content_padding=8,
                    suffix=ft.Text(suffix, color="#AAAAAA", size=11) if suffix else None
                ),
                width=width
            )

        # Render Estrellas ⭐
        def set_stars(num):
            data["estrellas_logro"] = num
            update_active_view()

        star_row = ft.Row([
            ft.IconButton(
                icon=ft.Icons.STAR_ROUNDED if i <= data["estrellas_logro"] else ft.Icons.STAR_OUTLINE_ROUNDED,
                icon_color="#FFD700" if i <= data["estrellas_logro"] else "#555555",
                icon_size=20,
                on_click=lambda e, idx=i: set_stars(idx),
                padding=0
            ) for i in range(1, 6)
        ], spacing=0)

        # 1. TARJETA CABECERA DE METAS GENERATION (OFICIAL SGH)
        card_metas = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Icon(ft.Icons.FLAG_ROUNDED, color="#00FFFF", size=18),
                    ft.Text(f"META DEL DÍA ({d_name}) Y NO NEGOCIABLES", color="white", weight="bold", size=13),
                    ft.Container(expand=True),
                    ft.Text("Evaluación:", color="#AAAAAA", size=11),
                    star_row
                ], spacing=6, vertical_alignment="center"),
                ft.Divider(height=8, color="#374151"),
                ft.Row([
                    # Columna 1: Meta del Día (Negro / Cian)
                    ft.Container(
                        content=ft.Column([
                            ft.Text("META DEL DÍA (NEGRO)", color="#00FFFF", weight="bold", size=11),
                            ft.Text("Meta Diaria (Manual)", color="#AAAAAA", size=10),
                            make_white_input(data["meta_diaria"], "meta_diaria", width=130, suffix="$"),
                            ft.Text("Análogos 85% (Auto)", color="#AAAAAA", size=10),
                            make_green_calc("analogos", f"${calc['analogos']:,.2f}", width=130),
                            ft.Text("Wearables 15% (Auto)", color="#AAAAAA", size=10),
                            make_green_calc("wearables", f"${calc['wearables']:,.2f}", width=130),
                            ft.Text("Total Unidades (Auto)", color="#AAAAAA", size=10),
                            make_green_calc("total_unidades", f"{calc['total_unidades']} Pza", width=130),
                        ], spacing=4),
                        bgcolor="#111827", padding=10, border_radius=8, border=ft.Border.all(1, "#374151"), width=220
                    ),
                    # Columna 2: Conversión (No Negociable)
                    ft.Container(
                        content=ft.Column([
                            ft.Text("CONVERSIÓN (NO NEGOCIABLE)", color="#F59E0B", weight="bold", size=11),
                            ft.Text("Tráfico Esperado (Manual)", color="#AAAAAA", size=10),
                            make_white_input(data["trafico_esperado"], "trafico_esperado", width=130),
                            ft.Text("Conversión LY+1 (Manual)", color="#AAAAAA", size=10),
                            make_white_input(int(data["conversion_target"]*100), "conversion_target", width=130, suffix="%"),
                            ft.Text("Transacciones (Auto)", color="#AAAAAA", size=10),
                            make_green_calc("transacciones", f"{calc['transacciones']} Transac.", width=130),
                            ft.Text("Meta Ideal 110% (Auto)", color="#AAAAAA", size=10),
                            make_green_calc("meta_ideal", f"${calc['meta_ideal']:,.2f}", width=130),
                        ], spacing=4),
                        bgcolor="#111827", padding=10, border_radius=8, border=ft.Border.all(1, "#374151"), width=220
                    ),
                    # Columna 3: Otros No Negociables
                    ft.Container(
                        content=ft.Column([
                            ft.Text("OTROS NO NEGOCIABLES", color="#E040FB", weight="bold", size=11),
                            ft.Text("Wearables: 15% (Auto)", color="#AAAAAA", size=10),
                            make_green_calc("fix_wea", f"15% (Min 1)", width=130),
                            ft.Text("Kids: 5% (Auto)", color="#AAAAAA", size=10),
                            make_green_calc("fix_kid", f"5% (Min 1)", width=130),
                            ft.Text("Carekits: 30% (Auto)", color="#AAAAAA", size=10),
                            make_green_calc("fix_ck", f"30% (Min 1)", width=130),
                        ], spacing=4),
                        bgcolor="#111827", padding=10, border_radius=8, border=ft.Border.all(1, "#374151"), width=220
                    ),
                    # Columna 4: Productividad & COMP LY
                    ft.Container(
                        content=ft.Column([
                            ft.Text("PRODUCTIVIDAD & COMP LY", color="#10B981", weight="bold", size=11),
                            ft.Text("Vta Neta / Prod (Auto)", color="#AAAAAA", size=10),
                            make_green_calc("vta_neta_prod", f"${calc['vta_neta_prod']:,.2f}", width=130),
                            ft.Text("U.Prod (U/Hr) (Auto)", color="#AAAAAA", size=10),
                            make_green_calc("u_prod", f"{calc['u_prod']}", width=130),
                            ft.Text("COMP LY - Vta LY (Manual)", color="#AAAAAA", size=10),
                            make_white_input(data["vta_ly"], "vta_ly", width=130, suffix="$"),
                        ], spacing=4),
                        bgcolor="#111827", padding=10, border_radius=8, border=ft.Border.all(1, "#374151"), width=220
                    ),
                    # Columna 5: CAPTURA DÍA COMP & MTD (ATV / AUR)
                    ft.Container(
                        content=ft.Column([
                            ft.Text("VALORES DÍA COMP & MTD", color="#FFD700", weight="bold", size=11),
                            ft.Text("ATV Día Comp (Manual)", color="#AAAAAA", size=10),
                            make_white_input(data.get("atv_dia", 3620.0), "atv_dia", width=130, suffix="$"),
                            ft.Text("AUR Día Comp (Manual)", color="#AAAAAA", size=10),
                            make_white_input(data.get("aur_dia", 3620.0), "aur_dia", width=130, suffix="$"),
                            ft.Text("ATV MTD (Manual)", color="#AAAAAA", size=10),
                            make_white_input(data.get("atv_mtd", 7597.0), "atv_mtd", width=130, suffix="$"),
                            ft.Text("AUR MTD (Manual)", color="#AAAAAA", size=10),
                            make_white_input(data.get("aur_mtd", 3362.0), "aur_mtd", width=130, suffix="$"),
                        ], spacing=4),
                        bgcolor="#111827", padding=10, border_radius=8, border=ft.Border.all(1, "#374151"), width=220
                    ),
                ], spacing=10, wrap=True)
            ]),
            bgcolor="#0B0E17",
            padding=14,
            border_radius=12,
            border=ft.Border.all(1.5, "#00FFFF"),
            shadow=[ft.BoxShadow(color="#2000FFFF", blur_radius=10, spread_radius=1)]
        )

        # 2. TARJETA DESGLOSE HORARIO POR BLOQUE
        is_mobile_w = (page.width < 800) if (page and page.width) else False
        w_lbl = 90 if is_mobile_w else 120
        w_cell = 75 if is_mobile_w else 95

        b_names = ["Apertura-1pm", "1pm - 3pm", "3pm - 5pm", "5pm - 7pm", "7pm - Cierre"]
        row_b_headers = [ft.Container(ft.Text("INDICADOR", weight="bold", color="#00FFFF", size=11), width=w_lbl)]
        for bn in b_names:
            row_b_headers.append(ft.Container(ft.Text(bn, weight="bold", color="white", size=11), width=w_cell, alignment=ft.alignment.Alignment(0, 0)))
        row_b_headers.append(ft.Container(ft.Text("TOTAL", weight="bold", color="#00FF88", size=11), width=w_cell, alignment=ft.alignment.Alignment(0, 0)))

        # Fila Tráfico (⚪)
        row_b_trafico = [ft.Container(ft.Text("Tráfico (⚪)", color="#AAAAAA", size=11), width=w_lbl)]
        for idx, tv in enumerate(data["trafico_bloques"]):
            row_b_trafico.append(make_white_input(tv, f"trafico_b_{idx}", width=w_cell))
        row_b_trafico.append(make_green_calc("tot_trafico_b", str(calc["tot_trafico_b"]), width=w_cell))

        # Fila Peso % (🟩)
        row_b_peso = [ft.Container(ft.Text("Peso % (🟩)", color="#AAAAAA", size=11), width=w_lbl)]
        for idx, p in enumerate(calc["b_pesos"]):
            row_b_peso.append(make_green_calc(f"b_peso_{idx}", f"{p*100:.1f}%", width=w_cell))
        row_b_peso.append(make_green_calc("b_peso_tot", "100%", width=w_cell))

        # Fila Meta $ (🟩)
        row_b_meta = [ft.Container(ft.Text("Meta $ (🟩)", color="#AAAAAA", size=11), width=w_lbl)]
        for idx, m in enumerate(calc["b_metas"]):
            row_b_meta.append(make_green_calc(f"b_meta_{idx}", f"${m:,.0f}", width=w_cell))
        row_b_meta.append(make_green_calc("b_meta_tot", f"${calc['meta_diaria']:,.0f}", width=w_cell))

        scrollable_horarios_table = ft.Row([
            ft.Column([
                ft.Row(row_b_headers, spacing=6),
                ft.Row(row_b_trafico, spacing=6),
                ft.Row(row_b_peso, spacing=6),
                ft.Row(row_b_meta, spacing=6)
            ], spacing=6)
        ], scroll=ft.ScrollMode.AUTO)

        card_horarios = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Icon(ft.Icons.ACCESS_TIME_ROUNDED, color="#E040FB", size=18),
                    ft.Text("DESGLOSE HORARIO DE VENTA DE TIENDA", color="white", weight="bold", size=13)
                ], spacing=6),
                ft.Divider(height=8, color="#374151"),
                scrollable_horarios_table
            ]),
            bgcolor="#0B0E17",
            padding=14,
            border_radius=12,
            border=ft.Border.all(1.5, "#E040FB"),
            shadow=[ft.BoxShadow(color="#20E040FB", blur_radius=10, spread_radius=1)]
        )

        # Componente Celda Bloqueada / Solo Lectura (🔒)
        def make_readonly_input(val, width=140):
            return ft.Container(
                content=ft.TextField(
                    value=str(val),
                    read_only=True,
                    text_size=12,
                    text_style=ft.TextStyle(weight="bold", color="#00FFFF"),
                    bgcolor="#111827",
                    border_color="#374151",
                    content_padding=8
                ),
                width=width
            )

        # 3. TABLA ASIGNACIÓN POR COLABORADOR
        w_colab_name = 100 if is_mobile_w else 140
        w_colab_hrs = 65 if is_mobile_w else 80
        w_colab_vta = 90 if is_mobile_w else 110
        w_colab_ana = 75 if is_mobile_w else 90
        w_colab_wea = 75 if is_mobile_w else 90
        w_colab_kid = 65 if is_mobile_w else 80
        w_colab_ck = 65 if is_mobile_w else 80

        colab_hdr_ui = ft.Row([
            ft.Container(ft.Text("COLABORADOR (🔒)", weight="bold", color="#00FFFF", size=11), width=w_colab_name),
            ft.Container(ft.Text("HORAS (⚪)", weight="bold", color="#00FFFF", size=11), width=w_colab_hrs),
            ft.Container(ft.Text("META VTA (🟩)", weight="bold", color="#00FF88", size=11), width=w_colab_vta),
            ft.Container(ft.Text("ANÁLOGOS (🟩)", weight="bold", color="#00FF88", size=11), width=w_colab_ana),
            ft.Container(ft.Text("WEARABLES (⚪)", weight="bold", color="#00FFFF", size=11), width=w_colab_wea),
            ft.Container(ft.Text("KIDS (⚪)", weight="bold", color="#00FFFF", size=11), width=w_colab_kid),
            ft.Container(ft.Text("CAREKITS (🟩)", weight="bold", color="#00FF88", size=11), width=w_colab_ck),
        ], spacing=6)

        colab_rows_ui = [colab_hdr_ui]
        for idx, c in enumerate(data["colaboradores"]):
            r_calc = calc["colab_rows"][idx]
            colab_rows_ui.append(
                ft.Row([
                    make_readonly_input(c["nombre"], width=w_colab_name),
                    make_white_input(c["horas"], f"colab_hrs_{idx}", width=w_colab_hrs),
                    make_green_calc(f"colab_vta_{idx}", f"${r_calc['meta_vta']:,.2f}", width=w_colab_vta),
                    make_green_calc(f"colab_ana_{idx}", str(r_calc['meta_ana']), width=w_colab_ana),
                    make_white_input(c.get("meta_wea", r_calc['meta_wea']), f"colab_wea_{idx}", width=w_colab_wea),
                    make_white_input(c.get("meta_kid", r_calc['meta_kid']), f"colab_kid_{idx}", width=w_colab_kid),
                    make_green_calc(f"colab_ck_{idx}", str(r_calc['meta_ck']), width=w_colab_ck),
                ], spacing=6)
            )

        # Fila Totales Colaboradores
        colab_rows_ui.append(
            ft.Row([
                ft.Container(ft.Text("TOTALES TIENDA", weight="bold", color="white", size=11), width=w_colab_name),
                make_green_calc("tot_colab_hrs", f"{calc['tot_horas']:.1f} hrs", width=w_colab_hrs),
                make_green_calc("tot_colab_vta", f"${calc['meta_diaria']:,.2f}", width=w_colab_vta),
                make_green_calc("tot_colab_ana", str(sum(r["meta_ana"] for r in calc["colab_rows"])), width=w_colab_ana),
                make_green_calc("tot_colab_wea", str(sum(r["meta_wea"] for r in calc["colab_rows"])), width=w_colab_wea),
                make_green_calc("tot_colab_kid", str(sum(r["meta_kid"] for r in calc["colab_rows"])), width=w_colab_kid),
                make_green_calc("tot_colab_ck", str(sum(r["meta_ck"] for r in calc["colab_rows"])), width=w_colab_ck),
            ], spacing=6)
        )

        scrollable_colabs_table = ft.Row([
            ft.Column(colab_rows_ui, spacing=5)
        ], scroll=ft.ScrollMode.AUTO)

        card_colabs = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Icon(ft.Icons.PEOPLE_ROUNDED, color="#00FFFF", size=18),
                    ft.Text("ASIGNACIÓN Y DISTRIBUCIÓN POR PERSONAL", color="white", weight="bold", size=13)
                ], spacing=6),
                ft.Text("💡 Celdas blancas (⚪) son editables. Las celdas verdes (🟩) calculan metas en tiempo real.", color="#AAAAAA", size=10),
                ft.Divider(height=8, color="#374151"),
                scrollable_colabs_table
            ]),
            bgcolor="#0B0E17",
            padding=14,
            border_radius=12,
            border=ft.Border.all(1.5, "#00FFFF"),
            shadow=[ft.BoxShadow(color="#2000FFFF", blur_radius=10, spread_radius=1)]
        )

        # 4. TABLA CÓMO VAMOS (TOTAL DEL DÍA)
        cv_hdr_ui = ft.Row([
            ft.Container(ft.Text("META", weight="bold", color="#00FFFF", size=11), width=80),
            ft.Container(ft.Text("VENTA NETA (⚪)", weight="bold", color="#00FFFF", size=11), width=100),
            ft.Container(ft.Text("META UNID.", weight="bold", color="#00FFFF", size=11), width=80),
            ft.Container(ft.Text("VENTA UNID. (⚪)", weight="bold", color="#00FFFF", size=11), width=110),
            ft.Container(ft.Text("CONVERSIÓN (🟩)", weight="bold", color="#00FF88", size=11), width=110),
            ft.Container(ft.Text("CRECIMIENTO (🟩)", weight="bold", color="#00FF88", size=11), width=110),
            ft.Container(ft.Text("WEARABLES % (⚪)", weight="bold", color="#00FFFF", size=11), width=110),
            ft.Container(ft.Text("KIDS % (⚪)", weight="bold", color="#00FFFF", size=11), width=90),
            ft.Container(ft.Text("CAREKITS % (🟩)", weight="bold", color="#00FF88", size=11), width=110),
        ], spacing=6)
        
        cv_row_ui = ft.Row([
            make_green_calc("tot_cv_meta", f"${calc['meta_diaria']:,.2f}", width=80),
            make_white_input(data.get("venta_neta_dia", 0.0), "venta_neta_dia", width=100),
            make_green_calc("tot_cv_meta_unidades", str(calc["total_unidades"]), width=80),
            make_white_input(data.get("venta_unidades_dia", 0), "venta_unidades_dia", width=110),
            make_green_calc("tot_cv_conversion", f"{calc['conversion_dia']*100:.1f}%", width=110),
            make_green_calc("tot_cv_crecimiento", f"{calc['crecimiento_conversion']*100:.1f}%", width=110),
            make_white_input(f"{calc['wearables_pct']*100:.1f}%", "cv_wearables_pct", width=110),
            make_white_input(f"{calc['kids_pct']*100:.1f}%", "cv_kids_pct", width=90),
            make_green_calc("tot_cv_carekits_pct", f"{calc['carekits_pct']*100:.1f}%", width=110),
        ], spacing=6)

        # 5. TABLA CÓMO VAMOS (POR COLABORADOR AL CIERRE)
        cv_colab_hdr = ft.Row([
            ft.Container(ft.Text("COLABORADOR", weight="bold", color="#00FFFF", size=10), width=90),
            ft.Container(ft.Text("HORAS", weight="bold", color="#00FFFF", size=10), width=50),
            ft.Container(ft.Text("INTERACC. (⚪)", weight="bold", color="#00FFFF", size=10), width=90),
            ft.Container(ft.Text("CONVERTIDOS (⚪)", weight="bold", color="#00FFFF", size=10), width=110),
            ft.Container(ft.Text("CONVERSIÓN (🟩)", weight="bold", color="#00FF88", size=10), width=100),
            ft.Container(ft.Text("VTA CIERRE (⚪)", weight="bold", color="#00FFFF", size=10), width=100),
            ft.Container(ft.Text("ANÁLOGAS (⚪)", weight="bold", color="#00FFFF", size=10), width=90),
            ft.Container(ft.Text("DEMOS WEA. (⚪)", weight="bold", color="#00FFFF", size=10), width=100),
            ft.Container(ft.Text("UNID. WEA. (⚪)", weight="bold", color="#00FFFF", size=10), width=100),
            ft.Container(ft.Text("CONV. WEA (🟩)", weight="bold", color="#00FF88", size=10), width=90),
            ft.Container(ft.Text("KIDS CIERRE (⚪)", weight="bold", color="#00FFFF", size=10), width=90),
            ft.Container(ft.Text("CK CIERRE (⚪)", weight="bold", color="#00FFFF", size=10), width=100),
        ], spacing=6)

        cv_colab_rows = [cv_colab_hdr]
        for idx, c in enumerate(data["colaboradores"]):
            r_calc = calc["colab_rows"][idx]
            cv_colab_rows.append(
                ft.Row([
                    make_readonly_input(c["nombre"], width=90),
                    make_green_calc(f"colab_cv_hrs_{idx}", f"{r_calc['horas']:.1f}", width=50),
                    make_white_input(c.get("interacciones", 0), f"colab_int_{idx}", width=90),
                    make_white_input(c.get("convertidos", 0), f"colab_conv_{idx}", width=110),
                    make_green_calc(f"colab_conv_cierre_{idx}", f"{r_calc['conversion_cierre']*100:.1f}%", width=100),
                    make_white_input(c.get("vta_cierre", 0.0), f"colab_vtac_{idx}", width=100),
                    make_white_input(c.get("ana_cierre", 0), f"colab_anac_{idx}", width=90),
                    make_white_input(c.get("wea_demos", 0), f"colab_wead_{idx}", width=100),
                    make_white_input(c.get("wea_cierre", 0), f"colab_weac_{idx}", width=100),
                    make_green_calc(f"colab_conv_wea_{idx}", f"{r_calc['conversion_wea']*100:.1f}%", width=90),
                    make_white_input(c.get("kid_cierre", 0), f"colab_kidc_{idx}", width=90),
                    make_white_input(c.get("ck_cierre", 0), f"colab_ckc_{idx}", width=100),
                ], spacing=6)
            )

        cv_colab_rows.append(
            ft.Row([
                ft.Container(ft.Text("TOTALES TIENDA", weight="bold", color="white", size=11), width=90),
                make_green_calc("sum_cv_horas", f"{calc['tot_horas']:.1f}", width=50),
                make_green_calc("sum_cv_interacciones", str(calc["tot_interacciones"]), width=90),
                make_green_calc("sum_cv_convertidos", str(calc["tot_convertidos"]), width=110),
                ft.Container(width=100), # Espacio para conversion global
                make_green_calc("sum_cv_vta_cierre", f"${calc['tot_vta_cierre']:,.2f}", width=100),
                make_green_calc("sum_cv_ana_cierre", str(calc["tot_ana_cierre"]), width=90),
                make_green_calc("sum_cv_wea_demos", str(calc["tot_wea_demos"]), width=100),
                make_green_calc("sum_cv_wea_cierre", str(calc["tot_wea_cierre"]), width=100),
                ft.Container(width=90), # Espacio conv wea global
                make_green_calc("sum_cv_kid_cierre", str(calc["tot_kid_cierre"]), width=90),
                make_green_calc("sum_cv_ck_cierre", str(calc["tot_ck_cierre"]), width=100),
            ], spacing=6)
        )

        card_como_vamos = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Icon(ft.Icons.TRENDING_UP_ROUNDED, color="#F59E0B", size=18),
                    ft.Text("¿CÓMO VAMOS? (RESULTADOS DE CIERRE)", color="white", weight="bold", size=13)
                ], spacing=6),
                ft.Divider(height=8, color="#374151"),
                ft.Text("TOTAL DEL DÍA", weight="bold", color="#F59E0B", size=11),
                ft.Row([ft.Column([cv_hdr_ui, cv_row_ui], spacing=5)], scroll=ft.ScrollMode.AUTO),
                ft.Divider(height=8, color="#374151"),
                ft.Text("RESULTADOS POR COLABORADOR AL CIERRE", weight="bold", color="#F59E0B", size=11),
                ft.Row([ft.Column(cv_colab_rows, spacing=5)], scroll=ft.ScrollMode.AUTO),
            ]),
            bgcolor="#0B0E17",
            padding=14,
            border_radius=12,
            border=ft.Border.all(1.5, "#F59E0B"),
            shadow=[ft.BoxShadow(color="#20F59E0B", blur_radius=10, spread_radius=1)]
        )

        return ft.Column([
            card_metas,
            ft.Container(height=8),
            card_horarios,
            ft.Container(height=8),
            card_colabs,
            ft.Container(height=8),
            card_como_vamos
        ], scroll=ft.ScrollMode.AUTO)

    # --- INTERFAZ PLAN DE ACCIÓN ---
    def build_plan_accion_ui(d_name):
        data = s_state[d_name]

        def on_enfoque_change(e):
            data["enfoque_hoy"] = e.control.value
            guardar_estado_persistente(user_id)

        def on_logros_change(e):
            data["logros_hoy"] = e.control.value
            guardar_estado_persistente(user_id)

        # Tarjetas de Los 5 Secretos (Brandies SGH)
        card_secretos = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Icon(ft.Icons.AUTO_AWESOME_ROUNDED, color="#FFD700", size=18),
                    ft.Text(f"LOS 5 SECRETOS Y PLAN DE ACCIÓN ({d_name}) - OFICIAL SGH", color="white", weight="bold", size=13),
                    ft.Container(expand=True),
                    ft.ElevatedButton(
                        content=ft.Row([
                            ft.Icon(ft.Icons.TABLE_CHART_ROUNDED, color="white", size=15),
                            ft.Text(f"📊 Exportar Excel Plan ({d_name})", color="white", weight="bold", size=11)
                        ], spacing=4),
                        style=ft.ButtonStyle(bgcolor="#059669", shape=ft.RoundedRectangleBorder(radius=6)),
                        url=f"/api/download_excel/{d_name}?user_id={user_id}"
                    ),
                    ft.Container(width=6),
                    ft.ElevatedButton(
                        content=ft.Row([
                            ft.Icon(ft.Icons.PRINT_ROUNDED, color="white", size=15),
                            ft.Text(f"📄 Descargar PDF Plan ({d_name})", color="white", weight="bold", size=11)
                        ], spacing=4),
                        style=ft.ButtonStyle(bgcolor="#10B981", shape=ft.RoundedRectangleBorder(radius=6)),
                        on_click=lambda e, day=d_name: generar_pdf_enfoque(day)
                    )
                ], spacing=6, vertical_alignment="center"),
                ft.Divider(height=8, color="#374151"),
                ft.Row([
                    ft.Container(ft.Column([ft.Text("✨ 1. PULIR ES PODER", weight="bold", color="#00FFFF", size=11), ft.Text("Ofrece limpiar sus lentes al iniciar la conversación.", color="#CCCCCC", size=10)]), bgcolor="#111827", padding=8, border_radius=8, width=190),
                    ft.Container(ft.Column([ft.Text("🕶️ 2. PÓNTELOS", weight="bold", color="#E040FB", size=11), ft.Text("Invítalo a probar diferentes modelos en la bandeja.", color="#CCCCCC", size=10)]), bgcolor="#111827", padding=8, border_radius=8, width=190),
                    ft.Container(ft.Column([ft.Text("🎉 3. DIVIÉRTETE MÁS", weight="bold", color="#FFD700", size=11), ft.Text("Muestra 3 o 4 opciones adicionales para venta múltiple.", color="#CCCCCC", size=10)]), bgcolor="#111827", padding=8, border_radius=8, width=190),
                    ft.Container(ft.Column([ft.Text("🧼 4. CUÍDALOS", weight="bold", color="#10B981", size=11), ft.Text("Ofrece estuche, solución limpiadora y carekits.", color="#CCCCCC", size=10)]), bgcolor="#111827", padding=8, border_radius=8, width=190),
                    ft.Container(ft.Column([ft.Text("📐 5. AJUSTE PERFECTO", weight="bold", color="#3B82F6", size=11), ft.Text("Ajusta los armazones a la medida exacta del cliente.", color="#CCCCCC", size=10)]), bgcolor="#111827", padding=8, border_radius=8, width=190),
                ], wrap=True, spacing=8)
            ]),
            bgcolor="#0B0E17",
            padding=14,
            border_radius=12,
            border=ft.Border.all(1.5, "#FFD700")
        )

        # Tarjetas del Customer Journey
        card_journey = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Icon(ft.Icons.MAP_ROUNDED, color="#00FFFF", size=18),
                    ft.Text("CUSTOMER JOURNEY (EXPERIENCIA DEL CLIENTE)", color="white", weight="bold", size=13)
                ], spacing=6),
                ft.Divider(height=8, color="#374151"),
                ft.Row([
                    ft.Container(ft.Column([ft.Text("🤝 1. Empieza una relación", weight="bold", color="#00FFFF", size=11), ft.Text("Bienvenida cálida e idéntica a los estándares SGH.", color="#CCCCCC", size=10)]), bgcolor="#111827", padding=8, border_radius=8, width=190),
                    ft.Container(ft.Column([ft.Text("🛡️ 2. Gánate su confianza", weight="bold", color="#E040FB", size=11), ft.Text("Escucha activa y conocimiento de producto.", color="#CCCCCC", size=10)]), bgcolor="#111827", padding=8, border_radius=8, width=190),
                    ft.Container(ft.Column([ft.Text("💬 3. Interactúa y relaciona", weight="bold", color="#FFD700", size=11), ft.Text("Conecta necesidades con beneficios clave.", color="#CCCCCC", size=10)]), bgcolor="#111827", padding=8, border_radius=8, width=190),
                    ft.Container(ft.Column([ft.Text("🔍 4. Descubre y aprende", weight="bold", color="#10B981", size=11), ft.Text("Preguntas abiertas sobre estilo de vida.", color="#CCCCCC", size=10)]), bgcolor="#111827", padding=8, border_radius=8, width=190),
                    ft.Container(ft.Column([ft.Text("🚀 5. Ve más allá", weight="bold", color="#3B82F6", size=11), ft.Text("Cierre perfecto, garantía y despedida memorable.", color="#CCCCCC", size=10)]), bgcolor="#111827", padding=8, border_radius=8, width=190),
                ], wrap=True, spacing=8)
            ]),
            bgcolor="#0B0E17",
            padding=14,
            border_radius=12,
            border=ft.Border.all(1.5, "#00FFFF")
        )

        card_enfoque_texto = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Icon(ft.Icons.EDIT_NOTE_ROUNDED, color="#00FFFF", size=18),
                    ft.Text("📝 TU ENFOQUE PARA HOY", color="#00FFFF", weight="bold", size=13)
                ], spacing=6),
                ft.TextField(
                    value=data["enfoque_hoy"],
                    on_change=on_enfoque_change,
                    hint_text="Escribe aquí las acciones clave y estrategia del día...",
                    multiline=True,
                    min_lines=3,
                    max_lines=5,
                    bgcolor="#111827",
                    border_color="#374151",
                    color="white",
                    text_size=12
                )
            ]),
            bgcolor="#0B0E17",
            padding=14,
            border_radius=12,
            border=ft.Border.all(1.5, "#00FFFF")
        )

        # Evaluador de 5 estrellas para logros
        def set_stars(num):
            data["estrellas_logro"] = num
            update_active_view()

        star_row_logros = ft.Row([
            ft.Text("Calificación del día: ", color="white", weight="bold", size=12),
            *[
                ft.IconButton(
                    icon=ft.Icons.STAR_ROUNDED if i <= data["estrellas_logro"] else ft.Icons.STAR_OUTLINE_ROUNDED,
                    icon_color="#FFD700" if i <= data["estrellas_logro"] else "#555555",
                    icon_size=24,
                    on_click=lambda e, idx=i: set_stars(idx),
                    padding=0
                ) for i in range(1, 6)
            ]
        ], spacing=4, vertical_alignment="center")

        card_logros_texto = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Icon(ft.Icons.EMOJI_EVENTS_ROUNDED, color="#E040FB", size=18),
                    ft.Text("🏆 LOGROS DE HOY Y OPORTUNIDADES PARA MAÑANA", color="#E040FB", weight="bold", size=13),
                    ft.Container(expand=True),
                    star_row_logros
                ], spacing=6),
                ft.TextField(
                    value=data["logros_hoy"],
                    on_change=on_logros_change,
                    hint_text="Resumen del cierre del día, compromisos y áreas de oportunidad...",
                    multiline=True,
                    min_lines=3,
                    max_lines=5,
                    bgcolor="#111827",
                    border_color="#374151",
                    color="white",
                    text_size=12
                )
            ]),
            bgcolor="#0B0E17",
            padding=14,
            border_radius=12,
            border=ft.Border.all(1.5, "#E040FB")
        )

        return ft.Column([
            card_secretos,
            ft.Container(height=8),
            card_journey,
            ft.Container(height=8),
            card_enfoque_texto,
            ft.Container(height=8),
            card_logros_texto
        ], scroll=ft.ScrollMode.AUTO, expand=True)

    # --- INTERFAZ RESUMEN SEMANAL (RÉPLICA EXACTA DE HOJA EXCEL SGH DE 3 SECCIONES) ---
    def build_semanal_ui():
        is_mobile_w = (page.width < 800) if (page and hasattr(page, 'width') and isinstance(page.width, (int, float))) else False

        # Valores por defecto en h_state (semanal)
        h_state.setdefault("meta_conversion", 0.16)
        h_state.setdefault("wea_pct", 0.15)
        h_state.setdefault("kids_pct", 0.05)
        h_state.setdefault("ck_pct", 0.30)
        h_state.setdefault("comply_sem", 22519.0)
        h_state.setdefault("atv_sem", 7500.0)
        h_state.setdefault("aur_sem", 4617.0)
        h_state.setdefault("horas_colab", {})
        h_state.setdefault("cierre_semanal", {})
        h_state.setdefault("vta_neta_sem", 0.0)
        h_state.setdefault("vta_unid_sem", 0)

        # Totales agregados de las pestañas diarias
        tot_meta_sem = sum(s_state[d]["meta_diaria"] for d in DIAS)
        tot_ana_sem = sum(s_state[d]["meta_diaria"] * 0.85 for d in DIAS)
        tot_wea_sem = sum(s_state[d]["meta_diaria"] * 0.15 for d in DIAS)
        
        tot_trafico_sem = 0
        tot_transac_sem = 0
        tot_horas_sem = 0.0
        
        dias_calc = {}
        for d in DIAS:
            c = calcular_dia(d, user_id)
            dias_calc[d] = c
            tot_trafico_sem += c.get("trafico", 0)
            tot_transac_sem += c.get("transacciones", 0)
            tot_horas_sem += c.get("tot_horas", 0.0)

        meta_conversion = float(h_state.get("meta_conversion", 0.16))
        meta_transacciones_sem = int(tot_trafico_sem * meta_conversion)
        meta_ideal_sem = tot_meta_sem * 1.10
        total_unidades_sem = int(tot_meta_sem / h_state["aur_sem"]) if h_state.get("aur_sem", 4617.0) > 0 else (int(tot_meta_sem / 4617) if tot_meta_sem > 0 else 0)

        wea_pct = float(h_state.get("wea_pct", 0.15))
        unidades_wea_sem = max(1, int(tot_wea_sem / 8100)) if tot_wea_sem > 0 else 1

        kids_pct = float(h_state.get("kids_pct", 0.05))
        unidades_kids_sem = max(1, int(total_unidades_sem * kids_pct)) if total_unidades_sem > 0 else 1

        ck_pct = float(h_state.get("ck_pct", 0.30))
        unidades_ck_sem = max(1, int(total_unidades_sem * ck_pct)) if total_unidades_sem > 0 else 1

        comply_sem = float(h_state.get("comply_sem", 22519.0))
        atv_sem = float(h_state.get("atv_sem", 7500.0))
        aur_sem = float(h_state.get("aur_sem", 4617.0))

        # Callbacks para guardar cambios al editar cuadros blancos de entrada
        def on_param_change(field_key, val_str, is_pct=False):
            try:
                val = float(val_str.replace("$", "").replace(",", "").replace("%", "").strip() or 0)
                if is_pct:
                    val = val / 100.0 if val > 1.0 else val
                h_state[field_key] = val
                guardar_estado_persistente(user_id)
                update_active_view()
            except Exception as ex:
                print(f"Error on_param_change {field_key}:", ex)

        # Helper para crear celda editable blanca/oscura con borde cyan
        def make_input_field(value_str, on_change_fn, width=80, suffix=""):
            return ft.Container(
                content=ft.TextField(
                    value=value_str,
                    width=width,
                    on_change=on_change_fn,
                    text_size=11,
                    text_style=ft.TextStyle(weight="bold", color="#FFFFFF"),
                    bgcolor="#1F2937",
                    border_color="#00FFFF",
                    focused_border_color="#00FF88",
                    content_padding=5,
                    suffix=ft.Text(suffix, color="#AAAAAA", size=10) if suffix else None
                ),
                width=width
            )

        # --- SECCIÓN 1: TARJETAS DE ENCABEZADO Y METAS ---
        card_w = 240 if is_mobile_w else 270

        card_metas = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Icon(ft.Icons.STAR_ROUNDED, color="#A100F2", size=16),
                    ft.Text("SEMANAL", color="#A100F2", weight="bold", size=12)
                ]),
                ft.Row([ft.Text("Meta Semanal:", color="#aaaaaa", size=11), ft.Text(f"${tot_meta_sem:,.2f}", color="#00FF88", weight="bold", size=11)], alignment="spaceBetween"),
                ft.Row([ft.Text("Análogos:", color="#aaaaaa", size=11), ft.Text(f"${tot_ana_sem:,.2f}", color="#00FF88", size=11)], alignment="spaceBetween"),
                ft.Row([ft.Text("Wearables:", color="#aaaaaa", size=11), ft.Text(f"${tot_wea_sem:,.2f}", color="#00FF88", size=11)], alignment="spaceBetween"),
                ft.Row([ft.Text("Total de Unidades:", color="#aaaaaa", size=11), ft.Text(str(total_unidades_sem), color="white", weight="bold", size=11)], alignment="spaceBetween"),
            ], spacing=6),
            bgcolor="#111827", padding=12, border_radius=10, border=ft.Border.all(1.5, "#A100F2"), width=card_w
        )

        card_conversion = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Icon(ft.Icons.SHOW_CHART_ROUNDED, color="#00FF88", size=16),
                    ft.Text("CONVERSIÓN SEMANAL", color="#00FF88", weight="bold", size=12)
                ]),
                ft.Row([ft.Text("Tráfico Esperado:", color="#aaaaaa", size=11), ft.Text(str(tot_trafico_sem), color="white", weight="bold", size=11)], alignment="spaceBetween"),
                ft.Row([
                    ft.Text("Meta Conversión:", color="#aaaaaa", size=11),
                    make_input_field(f"{meta_conversion*100:.1f}", lambda e: on_param_change("meta_conversion", e.control.value, is_pct=True), width=70, suffix="%")
                ], alignment="spaceBetween"),
                ft.Row([ft.Text("Meta Transacciones:", color="#aaaaaa", size=11), ft.Text(str(meta_transacciones_sem), color="white", size=11)], alignment="spaceBetween"),
                ft.Row([ft.Text("Meta Ideal ($ NS):", color="#aaaaaa", size=11), ft.Text(f"${meta_ideal_sem:,.2f}", color="#00FFFF", weight="bold", size=11)], alignment="spaceBetween"),
            ], spacing=6),
            bgcolor="#111827", padding=12, border_radius=10, border=ft.Border.all(1.5, "#00FF88"), width=card_w
        )

        # OTROS NO NEGOCIABLES: SOLO LECTURA (NO EDITABLE)
        card_otros = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Icon(ft.Icons.CHECK_CIRCLE_OUTLINE, color="#00FFFF", size=16),
                    ft.Text("OTROS NO NEGOCIABLES", color="#00FFFF", weight="bold", size=12)
                ]),
                ft.Row([ft.Text("Wearables (15%):", color="#aaaaaa", size=11), ft.Text(f"{unidades_wea_sem} unid.", color="#00FFFF", weight="bold", size=11)], alignment="spaceBetween"),
                ft.Row([ft.Text("Kids (5%):", color="#aaaaaa", size=11), ft.Text(f"{unidades_kids_sem} unid.", color="#00FFFF", weight="bold", size=11)], alignment="spaceBetween"),
                ft.Row([ft.Text("Carekits (30%):", color="#aaaaaa", size=11), ft.Text(f"{unidades_ck_sem} unid.", color="#00FFFF", weight="bold", size=11)], alignment="spaceBetween"),
            ], spacing=6),
            bgcolor="#111827", padding=12, border_radius=10, border=ft.Border.all(1.5, "#00FFFF"), width=card_w
        )

        card_comply = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Icon(ft.Icons.VERIFIED_ROUNDED, color="#FF8C00", size=16),
                    ft.Text("COMPLY E INDICADORES", color="#FF8C00", weight="bold", size=12)
                ]),
                ft.Row([
                    ft.Text("Comply Semanal:", color="#aaaaaa", size=11),
                    make_input_field(f"{comply_sem:.2f}", lambda e: on_param_change("comply_sem", e.control.value), width=95)
                ], alignment="spaceBetween"),
                ft.Row([
                    ft.Text("ATV Semanal:", color="#aaaaaa", size=11),
                    make_input_field(f"{atv_sem:.2f}", lambda e: on_param_change("atv_sem", e.control.value), width=90)
                ], alignment="spaceBetween"),
                ft.Row([
                    ft.Text("AUR Semanal:", color="#aaaaaa", size=11),
                    make_input_field(f"{aur_sem:.2f}", lambda e: on_param_change("aur_sem", e.control.value), width=90)
                ], alignment="spaceBetween"),
            ], spacing=6),
            bgcolor="#111827", padding=12, border_radius=10, border=ft.Border.all(1.5, "#FF8C00"), width=card_w
        )

        grid_cards = ft.Row([card_metas, card_conversion, card_otros, card_comply], spacing=10, wrap=True)

        # --- SECCIÓN 2: TABLA DE COLABORADORES Y METAS INDIVIDUALES (DESDE COLABORADORES DB) ---
        h_state.setdefault("wea_colab_manual", {})

        w_colab_name = 110 if is_mobile_w else 130
        w_colab_hrs = 80 if is_mobile_w else 100
        w_colab_vta = 95 if is_mobile_w else 110
        w_colab_ana = 75 if is_mobile_w else 90
        w_colab_wea = 75 if is_mobile_w else 90
        w_colab_kid = 65 if is_mobile_w else 80
        w_colab_ck = 65 if is_mobile_w else 80

        rows_colabs = []
        colabs_list = s_state.get("DOMINGO", {}).get("colaboradores", [])
        
        tot_horas_plantilla = 0.0
        horas_colab_map = {}
        for colab in colabs_list:
            if not isinstance(colab, dict):
                continue
            c_name = colab.get("nombre", "")
            if not c_name.strip():
                continue
            h_acc = 0.0
            for d in DIAS:
                c_day_list = s_state.get(d, {}).get("colaboradores", [])
                for p_item in c_day_list:
                    if isinstance(p_item, dict) and p_item.get("nombre") == c_name:
                        h_acc += float(p_item.get("horas", 0.0) or 0.0)
            
            if c_name in h_state["horas_colab"]:
                h_acc = float(h_state["horas_colab"][c_name])

            horas_colab_map[c_name] = h_acc
            tot_horas_plantilla += h_acc

        if tot_horas_plantilla <= 0:
            tot_horas_plantilla = 1.0

        tot_colab_meta_venta = 0.0
        tot_colab_analogos = 0
        tot_colab_wearables = 0
        tot_colab_kids = 0
        tot_colab_carekits = 0

        def on_colab_hrs_change(c_name, val_str):
            try:
                h_val = float(val_str.strip() or 0)
                h_state["horas_colab"][c_name] = h_val
                guardar_estado_persistente(user_id)
                update_active_view()
            except Exception:
                pass

        def on_colab_wea_change(c_name, val_str):
            try:
                w_val = int(val_str.strip() or 0)
                h_state["wea_colab_manual"][c_name] = w_val
                guardar_estado_persistente(user_id)
                update_active_view()
            except Exception:
                pass

        for colab in colabs_list:
            if not isinstance(colab, dict):
                continue
            c_name = colab.get("nombre", "")
            if not c_name.strip():
                continue
            h_p = horas_colab_map.get(c_name, 0.0)
            
            m_venta_colab = (tot_meta_sem / tot_horas_plantilla) * h_p
            ana_colab = max(1, int((total_unidades_sem / tot_horas_plantilla) * h_p)) if (tot_ana_sem > 0 and h_p > 0) else 0
            
            if c_name in h_state["wea_colab_manual"]:
                wea_colab = int(h_state["wea_colab_manual"][c_name])
            else:
                wea_colab = max(1, int((unidades_wea_sem / tot_horas_plantilla) * h_p)) if h_p > 0 else 0

            kids_colab = max(1, int((unidades_kids_sem / tot_horas_plantilla) * h_p)) if h_p > 0 else 0
            ck_colab = max(1, int((unidades_ck_sem / tot_horas_plantilla) * h_p)) if h_p > 0 else 0

            tot_colab_meta_venta += m_venta_colab
            tot_colab_analogos += ana_colab
            tot_colab_wearables += wea_colab
            tot_colab_kids += kids_colab
            tot_colab_carekits += ck_colab

            rows_colabs.append(
                ft.Row([
                    ft.Container(ft.Text(c_name, weight="bold", color="white", size=11), width=w_colab_name),
                    make_input_field(f"{h_p:.1f}", lambda e, cn=c_name: on_colab_hrs_change(cn, e.control.value), width=w_colab_hrs, suffix="h"),
                    ft.Container(ft.Text(f"${m_venta_colab:,.2f}", color="#00FF88", weight="bold", size=11), width=w_colab_vta),
                    ft.Container(ft.Text(str(ana_colab), color="#00FF88", size=11), width=w_colab_ana),
                    make_input_field(str(wea_colab), lambda e, cn=c_name: on_colab_wea_change(cn, e.control.value), width=w_colab_wea),
                    ft.Container(ft.Text(str(kids_colab), color="#00FFFF", size=11), width=w_colab_kid),
                    ft.Container(ft.Text(str(ck_colab), color="#00FFFF", size=11), width=w_colab_ck),
                ], spacing=6)
            )

        tabla_colabs_content = ft.Column([
            ft.Row([
                ft.Container(ft.Text("COLABORADOR", weight="bold", color="#A100F2", size=11), width=w_colab_name),
                ft.Container(ft.Text("HORAS PROG.", weight="bold", color="white", size=11), width=w_colab_hrs),
                ft.Container(ft.Text("META VENTA", weight="bold", color="#00FF88", size=11), width=w_colab_vta),
                ft.Container(ft.Text("ANÁLOGOS", weight="bold", color="#00FF88", size=11), width=w_colab_ana),
                ft.Container(ft.Text("WEARABLES", weight="bold", color="#00FFFF", size=11), width=w_colab_wea),
                ft.Container(ft.Text("KIDS", weight="bold", color="#00FFFF", size=11), width=w_colab_kid),
                ft.Container(ft.Text("CAREKITS", weight="bold", color="#00FFFF", size=11), width=w_colab_ck),
            ], spacing=6),
            ft.Column(rows_colabs, spacing=6),
            ft.Divider(height=8, color="#374151"),
            ft.Row([
                ft.Container(ft.Text("TOTAL", weight="bold", color="white", size=12), width=w_colab_name),
                ft.Container(ft.Text(f"{tot_horas_plantilla:.1f} hrs", weight="bold", color="white", size=12), width=w_colab_hrs),
                ft.Container(ft.Text(f"${tot_colab_meta_venta:,.2f}", weight="bold", color="#00FF88", size=12), width=w_colab_vta),
                ft.Container(ft.Text(str(tot_colab_analogos), weight="bold", color="#00FF88", size=12), width=w_colab_ana),
                ft.Container(ft.Text(str(tot_colab_wearables), weight="bold", color="#00FFFF", size=12), width=w_colab_wea),
                ft.Container(ft.Text(str(tot_colab_kids), weight="bold", color="#00FFFF", size=12), width=w_colab_kid),
                ft.Container(ft.Text(str(tot_colab_carekits), weight="bold", color="#00FFFF", size=12), width=w_colab_ck),
            ], spacing=6)
        ])

        card_tabla_colabs = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Icon(ft.Icons.PEOPLE_ALT_ROUNDED, color="#A100F2", size=18),
                    ft.Text("DISTRIBUCIÓN DE METAS POR COLABORADOR (SEMANAL)", color="white", weight="bold", size=13)
                ], spacing=6),
                ft.Divider(height=8, color="#374151"),
                ft.Row([tabla_colabs_content], scroll=ft.ScrollMode.AUTO)
            ]),
            bgcolor="#0B0E17", padding=14, border_radius=12, border=ft.Border.all(1.5, "#A100F2")
        )

        # --- SECCIÓN 3: SEGUIMIENTO Y RESULTADOS AL CIERRE SEMANAL (EXCEL RÉPLICA FILAS 25-37) ---
        w3_colab = 110 if is_mobile_w else 125
        w3_hrs = 65 if is_mobile_w else 75
        w3_num = 75 if is_mobile_w else 85
        w3_vta = 90 if is_mobile_w else 105
        w3_pct = 75 if is_mobile_w else 85
        w3_demos = 80 if is_mobile_w else 95
        w3_wea = 75 if is_mobile_w else 85
        w3_kids = 65 if is_mobile_w else 75
        w3_ck = 65 if is_mobile_w else 75

        def on_cierre_val_change(c_name, field_name, val_str, is_float=False):
            try:
                val = float(val_str.replace("$", "").replace(",", "").strip() or 0) if is_float else int(val_str.strip() or 0)
                h_state["cierre_semanal"].setdefault(c_name, {})[field_name] = val
                guardar_estado_persistente(user_id)
                update_active_view()
            except Exception:
                pass

        rows_cierre_colabs = []
        tot_cierre_hrs = 0.0
        tot_cierre_inter = 0
        tot_cierre_conv = 0
        tot_cierre_vta = 0.0
        tot_cierre_ana = 0
        tot_cierre_demos = 0
        tot_cierre_wea = 0
        tot_cierre_kids = 0
        tot_cierre_ck = 0

        for colab in colabs_list:
            if not isinstance(colab, dict):
                continue
            c_name = colab.get("nombre", "")
            if not c_name.strip():
                continue
            
            h_p = horas_colab_map.get(c_name, 0.0)
            c_data = h_state["cierre_semanal"].get(c_name, {})

            inter = int(c_data.get("interacciones", 0))
            conv = int(c_data.get("convertidos", 0))
            conv_pct = (conv / inter * 100.0) if inter > 0 else 0.0
            vta_c = float(c_data.get("vta_cierre", 0.0))
            ana_c = int(c_data.get("ana_cierre", 0))
            wea_demos = int(c_data.get("wea_demos", 0))
            wea_c = int(c_data.get("wea_cierre", 0))
            wea_conv_pct = (wea_c / wea_demos * 100.0) if wea_demos > 0 else 0.0
            kid_c = int(c_data.get("kid_cierre", 0))
            ck_c = int(c_data.get("ck_cierre", 0))

            tot_cierre_hrs += h_p
            tot_cierre_inter += inter
            tot_cierre_conv += conv
            tot_cierre_vta += vta_c
            tot_cierre_ana += ana_c
            tot_cierre_demos += wea_demos
            tot_cierre_wea += wea_c
            tot_cierre_kids += kid_c
            tot_cierre_ck += ck_c

            rows_cierre_colabs.append(
                ft.Row([
                    ft.Container(ft.Text(c_name, weight="bold", color="white", size=11), width=w3_colab),
                    ft.Container(ft.Text(f"{h_p:.1f}", color="#AAAAAA", size=11), width=w3_hrs),
                    make_input_field(str(inter), lambda e, cn=c_name: on_cierre_val_change(cn, "interacciones", e.control.value), width=w3_num),
                    make_input_field(str(conv), lambda e, cn=c_name: on_cierre_val_change(cn, "convertidos", e.control.value), width=w3_num),
                    ft.Container(ft.Text(f"{conv_pct:.1f}%", color="#00FF88", size=11), width=w3_pct),
                    make_input_field(f"{vta_c:.2f}", lambda e, cn=c_name: on_cierre_val_change(cn, "vta_cierre", e.control.value, is_float=True), width=w3_vta),
                    make_input_field(str(ana_c), lambda e, cn=c_name: on_cierre_val_change(cn, "ana_cierre", e.control.value), width=w3_num),
                    make_input_field(str(wea_demos), lambda e, cn=c_name: on_cierre_val_change(cn, "wea_demos", e.control.value), width=w3_demos),
                    make_input_field(str(wea_c), lambda e, cn=c_name: on_cierre_val_change(cn, "wea_cierre", e.control.value), width=w3_wea),
                    ft.Container(ft.Text(f"{wea_conv_pct:.1f}%", color="#00FFFF", size=11), width=w3_pct),
                    make_input_field(str(kid_c), lambda e, cn=c_name: on_cierre_val_change(cn, "kid_cierre", e.control.value), width=w3_kids),
                    make_input_field(str(ck_c), lambda e, cn=c_name: on_cierre_val_change(cn, "ck_cierre", e.control.value), width=w3_ck),
                ], spacing=6)
            )

        tot_conv_pct_gen = (tot_cierre_conv / tot_cierre_inter * 100.0) if tot_cierre_inter > 0 else 0.0
        tot_wea_conv_gen = (tot_cierre_wea / tot_cierre_demos * 100.0) if tot_cierre_demos > 0 else 0.0
        crec_conv_gen = tot_conv_pct_gen - (meta_conversion * 100.0)
        wea_pct_real = (tot_cierre_wea / (tot_cierre_ana + tot_cierre_wea) * 100.0) if (tot_cierre_ana + tot_cierre_wea) > 0 else 0.0
        kids_pct_real = (tot_cierre_kids / (tot_cierre_ana + tot_cierre_wea) * 100.0) if (tot_cierre_ana + tot_cierre_wea) > 0 else 0.0
        ck_pct_str = f"{(tot_cierre_ck / (tot_cierre_ana + tot_cierre_wea) * 100.0):.1f}%" if (tot_cierre_ana + tot_cierre_wea) > 0 else "#¡DIV/0!"

        # TABLA EXCEL CABECERA SUPERIOR AL CIERRE (FILAS 25-26 EXCEL OFICIAL)
        ex_header = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Container(ft.Text("META", weight="bold", color="white", size=10), width=95, alignment=ft.alignment.Alignment(0, 0)),
                    ft.Container(ft.Text("VENTA NETA", weight="bold", color="white", size=10), width=105, alignment=ft.alignment.Alignment(0, 0)),
                    ft.Container(ft.Text("META UNIDADES", weight="bold", color="white", size=10), width=110, alignment=ft.alignment.Alignment(0, 0)),
                    ft.Container(ft.Text("VENTA UNIDADES", weight="bold", color="white", size=10), width=105, alignment=ft.alignment.Alignment(0, 0)),
                    ft.Container(ft.Text("CONVERSION (%)", weight="bold", color="white", size=10), width=105, alignment=ft.alignment.Alignment(0, 0)),
                    ft.Container(ft.Text("CRECIMIENTO CONVERSION (%)", weight="bold", color="white", size=9), width=140, alignment=ft.alignment.Alignment(0, 0)),
                    ft.Container(ft.Text("WEARABLES %", weight="bold", color="white", size=10), width=100, alignment=ft.alignment.Alignment(0, 0)),
                    ft.Container(ft.Text("KIDS%", weight="bold", color="white", size=10), width=85, alignment=ft.alignment.Alignment(0, 0)),
                    ft.Container(ft.Text("CAREKITS%", weight="bold", color="white", size=10), width=95, alignment=ft.alignment.Alignment(0, 0)),
                ], spacing=4),
                ft.Row([
                    ft.Container(ft.Text(f"${tot_meta_sem:,.0f}", color="#00FF88", weight="bold", size=11), width=95, bgcolor="#065F46", padding=4, border_radius=4, alignment=ft.alignment.Alignment(0, 0)),
                    make_input_field(f"{h_state['vta_neta_sem']:.2f}", lambda e: on_param_change("vta_neta_sem", e.control.value), width=105),
                    ft.Container(ft.Text("SEMANAL", color="#00FF88", weight="bold", size=11), width=110, bgcolor="#065F46", padding=4, border_radius=4, alignment=ft.alignment.Alignment(0, 0)),
                    make_input_field(str(int(h_state['vta_unid_sem'])), lambda e: on_param_change("vta_unid_sem", e.control.value), width=105),
                    ft.Container(ft.Text(f"{tot_conv_pct_gen:.1f}%", color="#00FF88", weight="bold", size=11), width=105, bgcolor="#065F46", padding=4, border_radius=4, alignment=ft.alignment.Alignment(0, 0)),
                    ft.Container(ft.Text(f"{crec_conv_gen:.1f}%", color="#00FFFF", weight="bold", size=11), width=140, bgcolor="#065F46", padding=4, border_radius=4, alignment=ft.alignment.Alignment(0, 0)),
                    ft.Container(ft.Text(f"{wea_pct_real:.1f}%", color="white", size=11), width=100, bgcolor="#1F2937", padding=4, border_radius=4, alignment=ft.alignment.Alignment(0, 0)),
                    ft.Container(ft.Text(f"{kids_pct_real:.1f}%", color="white", size=11), width=85, bgcolor="#1F2937", padding=4, border_radius=4, alignment=ft.alignment.Alignment(0, 0)),
                    ft.Container(ft.Text(ck_pct_str, color="#FF8C00" if "DIV" in ck_pct_str else "#00FF88", weight="bold", size=11), width=95, bgcolor="#065F46", padding=4, border_radius=4, alignment=ft.alignment.Alignment(0, 0)),
                ], spacing=4)
            ]),
            padding=8, bgcolor="#111827", border_radius=8, border=ft.Border.all(1, "#374151")
        )

        tabla_cierre_header = ft.Row([
            ft.Container(ft.Text("COLABORADOR", weight="bold", color="#00FFFF", size=11), width=w3_colab),
            ft.Container(ft.Text("HORAS", weight="bold", color="white", size=11), width=w3_hrs),
            ft.Container(ft.Text("INTERACCIONES", weight="bold", color="#00FFFF", size=11), width=w3_num),
            ft.Container(ft.Text("CONVERTIDOS", weight="bold", color="#00FF88", size=11), width=w3_num),
            ft.Container(ft.Text("CONVERSIÓN", weight="bold", color="#00FF88", size=11), width=w3_pct),
            ft.Container(ft.Text("VENTA NETA", weight="bold", color="#00FF88", size=11), width=w3_vta),
            ft.Container(ft.Text("ANÁLOGOS", weight="bold", color="#00FF88", size=11), width=w3_num),
            ft.Container(ft.Text("DEMOS WEAR.", weight="bold", color="#00FFFF", size=11), width=w3_demos),
            ft.Container(ft.Text("WEARABLES", weight="bold", color="#00FFFF", size=11), width=w3_wea),
            ft.Container(ft.Text("CONV. WEAR.", weight="bold", color="#00FFFF", size=11), width=w3_pct),
            ft.Container(ft.Text("KIDS", weight="bold", color="#00FFFF", size=11), width=w3_kids),
            ft.Container(ft.Text("CAREKITS", weight="bold", color="#00FFFF", size=11), width=w3_ck),
        ], spacing=6)

        tabla_cierre_totales = ft.Row([
            ft.Container(ft.Text("TOTAL", weight="bold", color="white", size=12), width=w3_colab),
            ft.Container(ft.Text(f"{tot_cierre_hrs:.1f}", weight="bold", color="white", size=12), width=w3_hrs),
            ft.Container(ft.Text(str(tot_cierre_inter), weight="bold", color="#00FFFF", size=12), width=w3_num),
            ft.Container(ft.Text(str(tot_cierre_conv), weight="bold", color="#00FF88", size=12), width=w3_num),
            ft.Container(ft.Text(f"{tot_conv_pct_gen:.1f}%", weight="bold", color="#00FF88", size=12), width=w3_pct),
            ft.Container(ft.Text(f"${tot_cierre_vta:,.2f}", weight="bold", color="#00FF88", size=12), width=w3_vta),
            ft.Container(ft.Text(str(tot_cierre_ana), weight="bold", color="#00FF88", size=12), width=w3_num),
            ft.Container(ft.Text(str(tot_cierre_demos), weight="bold", color="#00FFFF", size=12), width=w3_demos),
            ft.Container(ft.Text(str(tot_cierre_wea), weight="bold", color="#00FFFF", size=12), width=w3_wea),
            ft.Container(ft.Text(f"{tot_wea_conv_gen:.1f}%", weight="bold", color="#00FFFF", size=12), width=w3_pct),
            ft.Container(ft.Text(str(tot_cierre_kids), weight="bold", color="#00FFFF", size=12), width=w3_kids),
            ft.Container(ft.Text(str(tot_cierre_ck), weight="bold", color="#00FFFF", size=12), width=w3_ck),
        ], spacing=6)

        tabla_cierre_content = ft.Column([
            ft.Row([ex_header], scroll=ft.ScrollMode.AUTO),
            ft.Divider(height=10, color="#374151"),
            tabla_cierre_header,
            ft.Column(rows_cierre_colabs, spacing=6),
            ft.Divider(height=8, color="#374151"),
            tabla_cierre_totales
        ])

        card_cierre_semanal = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Icon(ft.Icons.ANALYTICS_ROUNDED, color="#00FFFF", size=18),
                    ft.Text(f"SEGUIMIENTO DE RESULTADOS AL CIERRE SEMANAL - SEMANA {g_meta['semana']} ({g_meta['tienda']})", color="white", weight="bold", size=13)
                ], spacing=6),
                ft.Divider(height=8, color="#374151"),
                ft.Row([tabla_cierre_content], scroll=ft.ScrollMode.AUTO)
            ]),
            bgcolor="#0B0E17", padding=14, border_radius=12, border=ft.Border.all(1.5, "#00FFFF")
        )

        return ft.Container(
            content=ft.Column([
                grid_cards,
                ft.Container(height=10),
                card_tabla_colabs,
                ft.Container(height=10),
                card_cierre_semanal
            ], spacing=10),
            padding=5
        )

    # --- NAVEGACIÓN Y CAMBIO DE PESTAÑAS ---
    tab_content_container = ft.Container()

    def select_tab(tab_name):
        user_states[user_id]['active_tab'][0] = tab_name
        for btn, t_id in tab_buttons:
            is_sel = (t_id == tab_name)
            col_base = COLOR_TABS.get(t_id, "#00FFFF")
            btn.content.color = "white" if is_sel else "#CCCCCC"
            btn.bgcolor = col_base if is_sel else "#111827"
            btn.border = ft.Border.all(1.5, col_base if is_sel else "#374151")

        update_active_view()
        try:
            page.update()
        except Exception:
            try:
                tab_navigation_bar.update()
                tab_content_container.update()
            except Exception:
                pass

    # Construcción de botones de pestañas
    TABS_LIST = [
        ("📊 SEMANAL", "SEMANAL"),
        ("☀️ DOMINGO", "DOMINGO"),
        ("📋 PLAN DOMINGO", "PLAN.ACCIÓN_D"),
        ("🌙 LUNES", "LUNES"),
        ("📋 PLAN LUNES", "PLAN.ACCIÓN_L"),
        ("🔥 MARTES", "MARTES"),
        ("📋 PLAN MARTES", "PLAN.ACCIÓN_MA"),
        ("⚡ MIÉRCOLES", "MIÉRCOLES"),
        ("📋 PLAN MIÉRCOLES", "PLAN.ACCIÓN_MI"),
        ("🚀 JUEVES", "JUEVES"),
        ("📋 PLAN JUEVES", "PLAN.ACCIÓN_J"),
        ("💎 VIERNES", "VIERNES"),
        ("📋 PLAN VIERNES", "PLAN.ACCIÓN_V"),
        ("👑 SÁBADO", "SÁBADO"),
        ("📋 PLAN SÁBADO", "PLAN.ACCIÓN_S")
    ]

    tab_buttons = []
    tab_row_controls = []
    for label, t_id in TABS_LIST:
        is_sel = (t_id == user_states[user_id]['active_tab'][0])
        col_base = COLOR_TABS.get(t_id, "#00FFFF")
        btn = ft.Container(
            content=ft.Text(label, color="white" if is_sel else "#CCCCCC", size=11, weight="bold"),
            bgcolor=col_base if is_sel else "#111827",
            padding=ft.Padding(12, 8, 12, 8),
            border_radius=8,
            border=ft.Border.all(1.5, col_base if is_sel else "#374151"),
            on_click=lambda e, tid=t_id: select_tab(tid),
            ink=True
        )
        tab_buttons.append((btn, t_id))
        tab_row_controls.append(btn)

    tab_navigation_bar = ft.Row(tab_row_controls, scroll=ft.ScrollMode.AUTO, spacing=6)

    # --- BARRA SUPERIOR CON BOTÓN PDF Y CAMPOS DE ENCABEZADO ---
    # --- VERIFICACIÓN DE ROL Y USUARIO ---
    es_admin = False
    if session_user and isinstance(session_user, dict):
        u_rol = str(session_user.get("rol", "")).lower()
        u_name = str(session_user.get("user", "")).lower()
        if u_rol in ["admin", "administrador"] or u_name in ["mx204562", "clorio", "ricardo", "gerry", "cesar", "manuel"]:
            es_admin = True

    if not es_admin and session_user and isinstance(session_user, dict):
        u_tienda = session_user.get("tienda") or session_user.get("Tienda")
        u_user = session_user.get("usuario") or session_user.get("Usuario") or ""
        num_code = u_user.lower().replace("sgh", "").strip()
        if u_tienda and u_tienda != "Tienda Luxo":
            g_meta["tienda"] = u_tienda
            if num_code in MAPEO_TIENDAS_SGH:
                g_meta["num_tienda"] = num_code
            elif u_tienda.upper() in MAPEO_NOMBRE_A_NUMERO_SGH:
                g_meta["num_tienda"] = MAPEO_NOMBRE_A_NUMERO_SGH[u_tienda.upper()]

    # Sincronizar colaboradores iniciales
    sincronizar_colaboradores_db(session_user, g_meta["tienda"])

    # 0. Selector de Año (2025, 2026, 2027, 2028)
    g_meta.setdefault("anio", "2026")

    def on_anio_change(e):
        g_meta["anio"] = e.control.value
        guardar_estado_persistente(user_id)
        update_active_view()

    dd_anio = ft.Dropdown(
        value=str(g_meta["anio"]),
        options=[ft.dropdown.Option(str(y)) for y in [2026, 2025, 2027, 2028]],
        width=80,
        content_padding=6,
        text_size=11,
        text_style=ft.TextStyle(weight="bold", color="#FFFFFF"),
        bgcolor="#1F2937",
        border_color="#374151",
        border_radius=6
    )
    dd_anio.on_change = on_anio_change

    # 1. Selector de Semana (1 a 52)
    def on_semana_change(e):
        guardar_semana_historico(user_id)
        n_sem = e.control.value
        cargar_semana_historico(user_id, n_sem)
        guardar_estado_persistente(user_id)
        update_active_view()
        if page:
            snack = ft.SnackBar(ft.Text(f"📅 Semana {n_sem} cargada.", color="white"), bgcolor="#059669")
            page.overlay.append(snack)
            snack.open = True
            page.update()

    dd_semana = ft.Dropdown(
        value=str(g_meta["semana"]),
        options=[ft.dropdown.Option(str(w)) for w in range(1, 53)],
        width=75,
        content_padding=6,
        text_size=11,
        text_style=ft.TextStyle(weight="bold", color="#FFFFFF"),
        bgcolor="#1F2937",
        border_color="#374151",
        border_radius=6
    )
    dd_semana.on_change = on_semana_change

    # 2. Campo # TIENDA (Número de Tienda SGH)
    def on_num_tienda_change(e):
        val = e.control.value.strip()
        g_meta["num_tienda"] = val
        if val in MAPEO_TIENDAS_SGH:
            t_matched = MAPEO_TIENDAS_SGH[val]
            g_meta["tienda"] = t_matched
            dd_tienda.value = t_matched
            sincronizar_colaboradores_db(session_user, t_matched)
            guardar_estado_persistente(user_id)
            update_active_view()

    txt_num_tienda = ft.TextField(
        value=g_meta["num_tienda"],
        read_only=not es_admin,
        width=85,
        text_size=11,
        text_style=ft.TextStyle(weight="bold", color="#00FFFF" if es_admin else "#00FF88"),
        bgcolor="#111827" if es_admin else "#1F2937",
        border_color="#00FFFF" if es_admin else "#374151",
        content_padding=6,
        on_change=on_num_tienda_change if es_admin else None,
        tooltip="Número de Tienda SGH (🔒 Fijo para Gerentes)" if not es_admin else "Ingrese número de tienda"
    )

    # 3. Desplegable NOMBRE DE TIENDA (16 Tiendas Oficiales)
    options_tiendas = [ft.dropdown.Option(v) for v in MAPEO_TIENDAS_SGH.values()]

    def on_store_select_change(e):
        selected_tienda = e.control.value
        g_meta["tienda"] = selected_tienda
        if selected_tienda.upper() in MAPEO_NOMBRE_A_NUMERO_SGH:
            num_code = MAPEO_NOMBRE_A_NUMERO_SGH[selected_tienda.upper()]
            g_meta["num_tienda"] = num_code
            txt_num_tienda.value = num_code

        sincronizar_colaboradores_db(session_user, selected_tienda)
        guardar_estado_persistente(user_id)
        update_active_view()
        if page:
            snack = ft.SnackBar(ft.Text(f"🏢 Tienda cambiada a {selected_tienda} (# {g_meta['num_tienda']}).", color="white"), bgcolor="#059669")
            page.overlay.append(snack)
            snack.open = True
            page.update()

    dd_tienda = ft.Dropdown(
        value=g_meta["tienda"],
        options=options_tiendas,
        disabled=not es_admin,
        width=150,
        content_padding=6,
        text_size=11,
        text_style=ft.TextStyle(weight="bold", color="#00FFFF" if es_admin else "#00FF88"),
        bgcolor="#111827" if es_admin else "#1F2937",
        border_color="#00FFFF" if es_admin else "#374151",
        border_radius=6
    )
    if es_admin:
        dd_tienda.on_change = on_store_select_change

    btn_download_excel = ft.ElevatedButton(
        content=ft.Text("📊 Exportar Excel SGH", color="white", weight="bold", size=11),
        style=ft.ButtonStyle(
            bgcolor="#059669",
            shape=ft.RoundedRectangleBorder(radius=8)
        ),
        url=f"/api/download_excel/DOMINGO?user_id={user_id}"
    )

    btn_download_pdf = ft.ElevatedButton(
        content=ft.Text("🖨️ PDF / Imprimir Día", color="white", weight="bold", size=11),
        style=ft.ButtonStyle(
            bgcolor="#10B981",
            shape=ft.RoundedRectangleBorder(radius=8)
        ),
        url=f"/print_enfoque/DOMINGO?user_id={user_id}"
    )

    # Master Refresh Button 🔄 (Flat Neon Green Outline Style)
    def on_master_refresh_click(e):
        curr_tab = user_states[user_id]['active_tab'][0]
        d_name = curr_tab.replace("PLAN.ACCIÓN_", "")
        code_map = {"D": "DOMINGO", "L": "LUNES", "MA": "MARTES", "MI": "MIÉRCOLES", "J": "JUEVES", "V": "VIERNES", "S": "SÁBADO"}
        d_name = code_map.get(d_name, d_name if d_name in DIAS else "DOMINGO")
        
        calcular_dia(d_name, user_id)
        update_active_view()

        if page:
            snack = ft.SnackBar(
                content=ft.Text(f"🔄 LUXO: Datos de {d_name} ({g_meta['tienda']} #{g_meta['num_tienda']}) Sincronizados", color="white", weight="bold", size=12),
                bgcolor="#064E3B",
                duration=2000
            )
            page.overlay.append(snack)
            snack.open = True
            page.update()

    btn_refresh_master = ft.Container(
        content=ft.Text("🔄 Recalcular Todo", color="#00FF88", weight="bold", size=11),
        border=ft.Border.all(1.5, "#00FF88"),
        border_radius=6,
        padding=8,
        bgcolor="transparent",
        ink=True,
        on_click=on_master_refresh_click,
        tooltip="Sincronizar y recalcular todas las celdas y fórmulas"
    )

    header_module = ft.Container(
        content=ft.Column([
            # Fila 1: Título Principal + Botones de Acción de Exportación y Recálculo
            ft.Row([
                ft.Text("🎯 ENFOQUE DIARIO - SUNGLASS HUT", size=14, weight="bold", color="#00FFFF"),
                ft.Row([
                    btn_refresh_master,
                    btn_download_excel,
                    btn_download_pdf
                ], vertical_alignment="center", spacing=6, wrap=True)
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, vertical_alignment=ft.CrossAxisAlignment.CENTER, wrap=True),
            ft.Divider(height=4, color="#1F2937"),
            # Fila 2: Subtítulo + Filtros de Año, Semana, # Tienda y Nombre Tienda
            ft.Row([
                ft.Text("Replicación fiel del Excel Oficial SGH (Celdas blancas ⚪ = Entrada | Celdas verdes 🟩 = Cálculo automático).", size=10, color="#AAAAAA"),
                ft.Row([
                    ft.Text("AÑO:", color="white", weight="bold", size=11),
                    dd_anio,
                    ft.Container(width=2),
                    ft.Text("SEMANA:", color="white", weight="bold", size=11),
                    dd_semana,
                    ft.Container(width=2),
                    ft.Text("# TIENDA:", color="white", weight="bold", size=11),
                    txt_num_tienda,
                    ft.Container(width=2),
                    ft.Text("TIENDA:", color="white", weight="bold", size=11),
                    dd_tienda
                ], vertical_alignment="center", spacing=4, wrap=True)
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, vertical_alignment=ft.CrossAxisAlignment.CENTER, wrap=True)
        ], spacing=6),
        bgcolor="#0B0E17",
        padding=10,
        border_radius=10,
        border=ft.Border.all(1.5, "#1F2937")
    )

    update_active_view()

    return ft.Column([
        header_module,
        ft.Divider(height=8, color="#374151"),
        tab_navigation_bar,
        ft.Container(height=6),
        tab_content_container
    ], scroll=ft.ScrollMode.AUTO)
